from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse,JsonResponse,HttpResponseRedirect
from django.contrib.auth.models import User
from applications.academic_information.models import Student
import django. utils. timezone as timezone
from collections import defaultdict
import openpyxl


from .models import (
    CounsellingFAQ,
    CounsellingIssue,
    CounsellingIssueCategory,
    FacultyCounsellingTeam,
    StudentCounsellingTeam,
    StudentCounsellingInfo,
    CounsellingMeeting

)
from .handlers import (
    add_counselling_faq,
    add_student_counsellors
)
from applications.academic_information.models import Student,ExtraInfo
from applications.globals.models import Faculty, HoldsDesignation,Designation
# Create your views here.


def counselling_cell(request):
    user = request.user
    extra_info = ExtraInfo.objects.get(user=user)
    user_role = extra_info.user_type
    meetings = CounsellingMeeting.objects.all().select_related()
    year = timezone.now().year
    third_year_students = Student.objects.filter(batch=year-3).select_related()
    second_year_students = Student.objects.filter(batch=year-2).select_related()
    faqs = CounsellingFAQ.objects.all().select_related()
    categories = CounsellingIssueCategory.objects.all().select_related
    student_coordinators = StudentCounsellingTeam.objects.filter(student_position="student_coordinator").select_related()
    student_guide = StudentCounsellingTeam.objects.filter(student_position="student_guide").select_related()
    statuses =[]

    issues = CounsellingIssue.objects.filter(issue_status="status_unresolved").select_related()
    
    mapped_data = StudentCounsellingInfo.objects.all().select_related()
    student_and_student_guide = {}
    for x in mapped_data :
        if x.student_guide not in student_and_student_guide :
            student_and_student_guide[x.student_guide] = [x.student]
        else:
            student_and_student_guide[x.student_guide].append(x.student)
    if extra_info.user_type == 'student':
        statuses = CounsellingIssue.objects.filter(student=Student.objects.get(id=extra_info)).select_related()
        student = Student.objects.get(id=extra_info)
        user_role = "student"
        student = StudentCounsellingTeam.objects.filter(student = student).first()
        if student :
            student_des = student.student_position
            user_role = student_des
            # print(student_des)
            if student.student_position == "student_guide" :
             
           
                newlist = list()
                for i in student_and_student_guide.keys():
                    newlist.append(i)
                

                if student in newlist:


                    issues = CounsellingIssue.objects.filter(issue_status="status_unresolved",student__in=student_and_student_guide[student]).select_related()
                else :
                     
                    issues=[]    
    elif extra_info.user_type == 'faculty':
        designation = Designation.objects.filter(name= "counselling_head").first()
        faculty=Faculty.objects.get(id=extra_info)
        user_role="faculty"
        faculty=FacultyCounsellingTeam.objects.filter(faculty=faculty).first()
        user_role=faculty
        if faculty:
            print("kgf2")
            faculty_des=faculty.faculty_position
            user_role=faculty_des
    context = {
        "faqs":faqs,
        "meetings":meetings,
        "categories":categories,
        "third_year_students":third_year_students,
        "second_year_students":second_year_students,
        "student_counsellors":student_coordinators,
        "student_guide":student_guide,
        "statuses":statuses,
        "issues":issues,
        "user_role":user_role,
        "student_and_student_guide":student_and_student_guide
    }
    return render(request, "counselling_cell/counselling.html",context)
    

@csrf_exempt
def raise_issue(request):
    if request.method == 'POST':
        category_id = request.POST.get("category")
        print(category_id)
        category = CounsellingIssueCategory.objects.get(id = category_id)
        description = request.POST.get("description")
        user = request.user
        extra_info = ExtraInfo.objects.get(user=user)
        student = Student.objects.get(id=extra_info)
        issue = CounsellingIssue(
            issue_category = category,
            issue = description,
            student = student,
        )
        issue.save()
    
    return HttpResponseRedirect("/counselling/")

@csrf_exempt
def schedule_meeting(request):
    if request.method == 'POST':
        student_invities = request.POST.getlist("student")
        print(student_invities)
        temp=[]
        for i in student_invities:
            temp.append(i)
        venue = request.POST.get("venue")
        agenda = request.POST.get("agenda")
        date = request.POST.get("meeting_date")
        time = request.POST.get("meeting_time")
        user = request.user
        meeting_host = user
        extra_info = ExtraInfo.objects.get(user=user)
        meeting = CounsellingMeeting.objects.create(meeting_host=extra_info,meeting_date = date,meeting_time=time,agenda=agenda,venue=venue,student_invities=' '.join(temp))
        meeting.save()
        
    return HttpResponseRedirect("/counselling/") 

@csrf_exempt
def respond_issue(request):
    if request.method == 'POST':
        
        remark = request.POST.get("remark")
        idd = request.POST.get("id")
        print(remark,idd)
        user = request.user
        extra_info = ExtraInfo.objects.get(user=user)
        CounsellingIssue.objects.filter(id=idd).update(
            response_remark = remark,
            issue_status = "status_resolved",
            resolved_by=extra_info
        )
    
    return HttpResponseRedirect("/counselling/")


@csrf_exempt
# @login_required
# @transaction.atomic
def submit_counselling_faq(request):
    """
    This function is to record new faq submitted
    :param request:
        user: Current logged in user
    :variable:
         extra_info: Extra information of the user
    :return:
        data: to record success or any errors
    """
    student = request.user
    data = add_counselling_faq(request, student)
    return JsonResponse(data)

@csrf_exempt
def appoint_student_counsellors(request):
    data = add_student_counsellors(request)
    return JsonResponse(data)

@csrf_exempt
def dismiss_student_coordinator(request):
    data = remove_student_coordinator(request)
    return JsonResponse(data)

@csrf_exempt
def assign_student_to_sg(request):

    studentToStudentGuide=defaultdict(lambda:[])
    year = timezone.now().year
    # third_year_students = Student.objects.filter(batch=year-3)
    
    if request.method == 'POST' and request.FILES:
        profiles=request.FILES['mappedStudent']
        # excel = xlrd.open_workbook(file_contents=profiles.read())
        wb_obj = openpyxl.load_workbook(profiles)
        sheet = wb_obj.active
        for i in range(2,sheet.max_row+1):
            if sheet.cell(i,2).value : 
                student_roll_no=str(int(sheet.cell(i,2).value))
            if sheet.cell(i,1).value : 
                sg_roll_no=str(int(sheet.cell(i,1).value))  

                checkForSG = StudentCounsellingTeam.objects.filter(student_id=sg_roll_no,student_position="student_guide")
                if  len(checkForSG) == 0:
                    return JsonResponse({
                            'status':1,
                            'message':"Student Guide Not Found"
                        })
            studentToStudentGuide[sg_roll_no].append(student_roll_no)
        for sg,students in studentToStudentGuide.items() :
            sg = StudentCounsellingTeam.objects.filter(student_id=sg,student_position="student_guide").first()
            for student in students:    
                mappedStudent = StudentCounsellingInfo(student_guide=sg,student_id=Student(id=ExtraInfo(user=User(username=student))))
                mappedStudent.save()    
    return HttpResponseRedirect("/counselling/")
    
