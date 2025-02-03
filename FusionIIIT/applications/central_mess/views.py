from datetime import date, datetime, timedelta
from django.contrib.auth.decorators import login_required
from django.db import transaction
from threading import Thread
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import redirect, render
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.views.generic import View
from django.db.models import Count
from django.forms.models import model_to_dict
from django.db.models import Q
from django.contrib.auth.models import User
from .utils import render_to_pdf
from applications.academic_information.models import Student
from applications.globals.models import ExtraInfo, HoldsDesignation, Designation
from .forms import MinuteForm, MessInfoForm,RegistrationRequest,UpdatePaymentRequest
from .tasks import *
from .models import (Feedback, Menu, Menu_change_request, Mess_meeting,
                     Mess_minutes, Mess_reg, Messinfo, Monthly_bill,
                    Payments, Rebate, 
                     Special_request, Vacation_food, MessBillBase,Registration_Request, Reg_records ,Reg_main,Deregistration_Request,Semdates,Update_Payment)
from .handlers import (add_mess_feedback, add_sem_dates, add_vacation_food_request,
                       add_menu_change_request, handle_menu_change_response, handle_vacation_food_request,
                       add_mess_registration_time, add_leave_request, add_mess_meeting_invitation,
                       handle_rebate_response, add_special_food_request, handle_update_payment_response,
                       handle_special_request, add_bill_base_amount, add_mess_committee,  handle_reg_response, handle_dreg_response, update_month_bill,handle_add_reg)

from notification.views import central_mess_notif

import csv
import openpyxl
 

today_g = datetime.datetime.now()
month_g = today_g.month
month_g_l = today_g.strftime('%B')
year_g = today_g.year
tomorrow_g = today_g + timedelta(days=1)
first_day_of_this_month = date.today().replace(day=1)           
first_day_of_next_month = (date.today().replace(day=28) + timedelta(days=4)).replace(day=1)
last_day_of_this_month = first_day_of_next_month - timedelta(days=1)
next_month = first_day_of_next_month.month
last_day_prev_month = first_day_of_this_month - timedelta(days=1)
month_last_g = last_day_prev_month.month
year_last_g = last_day_prev_month.year
previous_month = last_day_prev_month.strftime('%B')

@login_required
def mess(request):
    """
    This view get the access to the central mess dashboard. View all details and apply for any changes.
    It also shows the previous feedback submitted by the user.
    """
    user = request.user
    notifs=request.user.notifications.all()
    extrainfo = ExtraInfo.objects.select_related().get(user=user)
    current_date = date.today()
    holds_designations = HoldsDesignation.objects.select_related().filter(user=user)
    desig = holds_designations
    form = MinuteForm()
    # mess_reg = Mess_reg.objects.select_related().last()
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    count6 = 0
    count7 = 0
    count8 = 0
    reg_form = RegistrationRequest()
    if extrainfo.user_type == 'student':
        # def deleteEntries():
        #     Registration_Request.objects.all().delete()
        # deleteEntries()
        student = Student.objects.select_related('id','id__user','id__department').get(id=extrainfo)
        vaca_obj = Vacation_food.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(student_id=student)
        feedback_obj = Feedback.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(student_id=student).order_by('-fdate')
        monthly_bill = Monthly_bill.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(student_id=student)
        payments = Payments.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(student_id=student)
        rebates = Rebate.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(student_id=student).order_by('-app_date')
        splrequest = Special_request.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(student_id=student).order_by('-app_date') 
        
        reg_form = RegistrationRequest()
        update_form=UpdatePaymentRequest()

        reg_request = Registration_Request.objects.filter(student_id=student)
        update_payment_request = Update_Payment.objects.filter(student_id=student)
        de_reg_request = Deregistration_Request.objects.filter(student_id=student)

        menu_data = Menu.objects.all()


        try:
            mess_optn = Reg_main.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').get(student_id=student)
            y = Menu.objects.filter(mess_option=mess_optn.mess_option)
            current_rem_balance = mess_optn.balance
            current_mess_status = mess_optn.current_mess_status
        except:
            mess_optn={}
            mess_optn={'mess_option':'no-mess'}
            y = Menu.objects.filter(mess_option="mess1")
            current_rem_balance = 0
            current_mess_status = 'Deregistered'
      
        

        reg_record = Reg_records.objects.filter(student_id=student)
        monthly_bill=monthly_bill[::-1]

        # tot_am=0
        # if len(payments)>0:
        #     tot_am=payments[0].amount_paid
        # else:
        #     tot_am=0
        #     for x in monthly_bill:
        #         tot_am=tot_am+x.total_bill
        #     Payments.objects.create(student_id=student,amount_paid=(-tot_am))
        #     payments = Payments.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(student_id=student)

            
            

        # for i in range(0,len(monthly_bill)):
        #     if(monthly_bill[i].paid):
        #         monthly_bill[i].due_amount=0;
        #     elif monthly_bill[i].total_bill+tot_am<0:
        #         monthly_bill[i].due_amount=(monthly_bill[i].total_bill)
        #     else:
        #         monthly_bill[i].due_amount=(-tot_am)
        #     tot_am+=monthly_bill[i].total_bill
        # amount_due=-payments[0].amount_paid
        # amount_due = 0
        ## adding the batch of student if btech or bdes then value of programme is 1 or else 0, holds value of phd and mtech.
        
        # if student.programme == 'B.Tech' or student.programme == 'B.Des':
        #     programme = 1
        # else:
        #     programme = 0
        # meeting = Mess_meeting.objects.all()
        # minutes = Mess_minutes.objects.all()
        # count = 0

        # try:
        #     mess_optn = Messinfo.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').get(student_id=student)
        #     y = Menu.objects.filter(mess_option=mess_optn.mess_option)


        
            # bill = Monthly_bill.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(Q(student_id=student) & Q(month=month_g_l) & Q(year=year_g))
            # amount_c = MessBillBase.objects.latest('timestamp')
            # rebate_count = 0
            
            # for r in rebates:
            #     if r.status == '2':
            #         if r.start_date.month == month_g:
            #             if r.end_date.month == next_month:
            #                 rebate_count = rebate_count + abs((last_day_of_this_month - r.start_date).days) + 1
            #             else:
            #                 rebate_count = rebate_count + abs((r.end_date - r.start_date).days) + 1
            #         elif r.end_date.month == month_g:
            #             rebate_count = rebate_count + abs((r.end_date - first_day_of_this_month).days) + 1
            #         else:
            #             rebate_count = 0
            # rebate_amount = rebate_count * amount_c.bill_amount / 30
            # total_bill = amount_c.bill_amount - rebate_amount 
            # if bill:
            #     bill.update(student_id = student,
            #                 month = month_g_l,
            #                 year = year_g,
            #                 amount = amount_c.bill_amount,
            #                 rebate_count = rebate_count,
            #                 rebate_amount = rebate_amount,
            #                 total_bill = total_bill)

            # else:
            #     bill_object = Monthly_bill(student_id=student,
            #                             amount=amount_c.bill_amount,
            #                             rebate_count=rebate_count,
            #                             rebate_amount=rebate_amount,
            #                             total_bill=total_bill,
            #                             month=month_g_l,
            #                             year=year_g)
            #     bill_object.save()
        # except:
        #     mess_optn={'mess_option':'no-mess'}
        #     y = Menu.objects.filter(mess_option="mess1")

        # for d in desig:
        #     if d.designation.name == 'mess_committee' or d.designation.name == 'mess_convener':
        #         newmenu = Menu_change_request.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department','dish').filter(dish__mess_option='mess1').order_by('-app_date')
        #         meeting = Mess_meeting.objects.all()
        #         minutes = Mess_minutes.objects.select_related().all()
        #         feed = Feedback.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(mess='mess1').order_by('-fdate')
        #         feed2 = Feedback.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(mess='mess2').order_by('-fdate')
        #         sprequest = Special_request.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(status='1').order_by('-app_date')
        #         sprequest_past = Special_request.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(status='2').order_by('-app_date')
        #         menuchangerequest= Menu_change_request.objects.select_related('student_id').filter().order_by('-app_date')

        #         # menu_data = Menu.objects.all()

        #         for f in feed:
        #             if f.feedback_type == 'Maintenance' :
        #                 count1 += 1

        #             elif f.feedback_type == 'Food' :
        #                 count2 += 1

        #             elif f.feedback_type == 'Cleanliness' :
        #                 count3 += 1

        #             elif f.feedback_type == 'Others' :
        #                 count4 += 1
        #         count5=0
        #         count6=0
        #         count7=0
        #         count8=0

        #         context = {
        #             'menu': menu_data,
        #             'reg_menu': y,
        #             'messinfo': mess_optn,
        #             'newmenu': newmenu,
        #             'monthly_bill': monthly_bill,
        #             'total_due': amount_due,
                    
        #             'vaca': vaca_obj,
        #             'info': extrainfo,
        #             'feedback': feedback_obj,
        #             'feed1': feed,
        #             'feed2':'',
        #             'student': student,
        #             'mess_reg': mess_reg,
        #             'current_date': current_date,
        #             'count': count,
        #             'rebates': rebates,
        #             'meeting': meeting,
        #             'minutes': minutes,
        #             'sprequest': sprequest,
        #             'splrequest': splrequest,
        #             'sprequest_past': sprequest_past,
        #             'menuchangerequest':menuchangerequest,
        #             'programme':programme,
        #             'count1': count1,
        #             'count2': count2,
        #             'count3': count3,
        #             'count4': count4,
        #             'count5': count5,
        #             'count6': count6,
        #             'count7': count7,
        #             'count8': count8,
        #             'form': form,
        #             'desig': desig,
        #             'reg_form':reg_form,
        #             'reg_request':reg_request,
        #             'reg_main':mess_optn,
        #             'reg_record':reg_record,
        #             'de_reg_request':de_reg_request,

        #         }
        #         return render(request, "messModule/mess.html", context)

        #     if d.designation.name == 'mess_committee_mess2' or d.designation.name == 'mess_convener_mess2':
        #         newmenu = Menu_change_request.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department','dish').filter(dish__mess_option='mess2').order_by('-app_date')
        #         meeting = Mess_meeting.objects.all()
        #         minutes = Mess_minutes.objects.select_related().all()
        #         feed = Feedback.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(mess='mess2').order_by('-fdate')
        #         feed2 = Feedback.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(mess='mess1').order_by('-fdate')
        #         sprequest = Special_request.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(status='1').order_by('-app_date')
        #         sprequest_past = Special_request.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(status='2').order_by('-app_date')
        #         menuchangerequest= Menu_change_request.objects.select_related('student_id').filter().order_by('-app_date')

        #         # menu_data = Menu.objects.all().order_by()

        #         count5=0
        #         count6=0
        #         count7=0
        #         count8=0
        #         for f in feed:
        #             if f.feedback_type == 'Maintenance' :
        #                 count1 += 1

        #             elif f.feedback_type == 'Food' :
        #                 count2 += 1

        #             elif f.feedback_type == 'Cleanliness' :
        #                 count3 += 1

        #             elif f.feedback_type == 'Others' :
        #                 count4 += 1

        #         context = {
        #             'menu': menu_data,
        #             'reg_menu': y,
        #             'messinfo': mess_optn,
        #             'newmenu': newmenu,
        #             'monthly_bill': monthly_bill,
        #             'total_due': amount_due,
        #             'vaca': vaca_obj,
        #             'info': extrainfo,
        #             'feedback': feedback_obj,
        #             'feed2': feed,
        #             'feed1':'',
        #             'student': student,
        #             # 'data': data,
        #             'mess_reg': mess_reg,
        #             'current_date': current_date,
        #             'count': count,
        #             'rebates': rebates,
        #             'programme': programme,
        #             'meeting': meeting,
        #             'minutes': minutes,
        #             'splrequest': splrequest,
        #             'sprequest': sprequest,
        #             'sprequest_past': sprequest_past,
        #             'menuchangerequest':menuchangerequest,  
        #             'count1': count1,
        #             'count2': count2,
        #             'count3': count3,
        #             'count4': count4,
        #             'count5': count5,
        #             'count6': count6,
        #             'count7': count7,
        #             'count8': count8,
        #             'form': form,
        #             'desig': desig,
        #             'reg_form':reg_form,
        #             'reg_request':reg_request,
        #             'reg_main':mess_optn,
        #             'reg_record':reg_record,
        #             'de_reg_request':de_reg_request,
        #         }
        #         return render(request, "messModule/mess.html", context)

        context = {
                    'menu': menu_data,
                #    'reg_menu': y,
                   'messinfo': mess_optn,
                   'monthly_bill': monthly_bill,
                #    'total_due': amount_due,
                   'vaca': vaca_obj,
                   'info': extrainfo,
                   'feedback': feedback_obj,
                   'student': student,
                #    'mess_reg': mess_reg,
                   'current_date': current_date,
                #    'count': count,
                   'rebates': rebates,
                   'splrequest': splrequest,
                   'form': form,
                #    'programme': programme,
                   'desig': desig,
                #    'minutes': minutes,
                #    'meeting': meeting,
                   'reg_form':reg_form,
                   'update_form':update_form,
                   'update_payment_request':update_payment_request,
                   'reg_main_stud':mess_optn,
                   'reg_request':reg_request,
                   'reg_record':reg_record,
                   'de_reg_request':de_reg_request,
                   'payments': payments,
                   'curr_balance': current_rem_balance,
                   'curr_status':current_mess_status,
                   'notifications':notifs
                  }

        return render(request, "messModule/mess.html", context)


        

    elif extrainfo.user_type == 'staff':
        for d in desig:
            if(d.designation.name == 'mess_manager'):
                try:
                    current_bill = MessBillBase.objects.latest('timestamp')
                except:
                    new_entry = MessBillBase(bill_amount=150)
                    new_entry.save()
                    current_bill = MessBillBase.objects.latest('timestamp')
                newmenu = Menu_change_request.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department','dish').all().order_by('-app_date')
                vaca_all = Vacation_food.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').all().order_by('-app_date')
                members_mess = HoldsDesignation.objects.select_related().filter(Q(designation__name__contains='mess_convener')
                                                            | Q(designation__name__contains='mess_committee'))
                y = Menu.objects.all()
                leave = Rebate.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(status='1').order_by('-app_date')
                leave_past = Rebate.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(status='2').order_by('-app_date')
                # meeting = Mess_meeting.objects.all()
                # minutes = Mess_minutes.objects.all()
                feed1 = Feedback.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(mess='mess1').order_by('-fdate')
                feed2 = Feedback.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(mess='mess2').order_by('-fdate')
                        
                for f in feed1:
                    if f.feedback_type == 'Maintenance' :
                        count1 += 1

                    elif f.feedback_type == 'Food' :
                        count2 += 1

                    elif f.feedback_type == 'Cleanliness' :
                        count3 += 1

                    elif f.feedback_type == 'Others' :
                        count4 += 1

                for f in feed2:
                    if f.feedback_type == 'Maintenance':
                        count5 += 1

                    elif f.feedback_type == 'Food':
                        count6 += 1

                    elif f.feedback_type == 'Cleanliness':
                        count7 += 1

                    elif f.feedback_type == 'Others':
                        count8 += 1
            
                sprequest = Special_request.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(status='1').order_by('-app_date')
                sprequest_past = Special_request.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(status='2').order_by('-app_date')

                reg_request = Registration_Request.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').all().filter(status='pending')
                update_pay_request=Update_Payment.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').all().filter(status='pending')
                de_reg_request = Deregistration_Request.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').all().filter(status='pending')
                reg_main = Reg_main.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(current_mess_status='Registered')
                reg_record = Reg_records.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').all()
                bills = Monthly_bill.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').all()
                # bills = Monthly_bill.objects.all()
                context = {
                    'bill_base': current_bill,
                    'today': today_g.date(),
                    'tomorrow': tomorrow_g.date(),
                    'members': members_mess,
                    'menu': y,
                    'newmenu': newmenu,
                    'vaca_all': vaca_all,
                    'info': extrainfo,
                    'leave': leave,
                    'leave_past': leave_past,
                    'current_date': current_date,
                    # 'mess_reg': mess_reg,
                    'desig': desig,
                    # 'meeting': meeting,
                    # 'minutes': minutes,
                    'sprequest': sprequest,
                    'sprequest_past': sprequest_past,
                    'count1': count1,
                    'count2': count2, 'count3': count3, 'feed1': feed1,'feed2':feed2,
                    'count4': count4, 'form': form, 'count5': count5,
                    'count6': count6, 'count7': count7, 'count8': count8, 'desig': desig,
                    'reg_request':reg_request,'reg_record':reg_record,'reg_main':reg_main,
                    'de_reg_request':de_reg_request,
                    'bill': bills,
                    'reg_form':reg_form,
                    'update_pay_request':update_pay_request
                }
                return render(request, "messModule/mess.html", context)

    elif extrainfo.user_type == 'faculty':
        for d in desig:
            if(d.designation.name == 'mess_warden'):
                
                feed1 = Feedback.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(mess='mess1').order_by('-fdate')
                feed2 = Feedback.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(mess='mess2').order_by('-fdate')
                y = Menu.objects.all()

                reg_main = Reg_main.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(current_mess_status='Registered')
                reg_record = Reg_records.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').all()
                bills = Monthly_bill.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').all()

                for f in feed1:
                    if f.feedback_type == 'Maintenance' :
                        count1 += 1

                    elif f.feedback_type == 'Food' :
                        count2 += 1

                    elif f.feedback_type == 'Cleanliness' :
                        count3 += 1

                    elif f.feedback_type == 'Others' :
                        count4 += 1

                for f in feed2:
                    if f.feedback_type == 'Maintenance':
                        count5 += 1

                    elif f.feedback_type == 'Food':
                        count6 += 1

                    elif f.feedback_type == 'Cleanliness':
                        count7 += 1

                    elif f.feedback_type == 'Others':
                        count8 += 1
                context = {
                    'info': extrainfo,
                    'desig': desig,
                    'menu': y,
                    'count1': count1,
                    'count2': count2, 'count3': count3, 'feed1': feed1,'feed2':feed2,
                    'count4': count4, 'form': form, 'count5': count5,
                    'count6': count6, 'count7': count7, 'count8': count8, 'desig': desig,
                    'reg_record':reg_record,'reg_main':reg_main,'bill': bills,
                }
                return render(request, 'messModule/mess.html', context)


@login_required
@transaction.atomic
def mess_info(request):
    """
    Register someone in mess 1 or mess 2

    @params:
        request: contains metadata about the requested page

    @variables:
        user_id: user id of the current user
        student_id: Student Object of given user_id
        mess_option: requested mess : {"mess1" || "mess2"}
    """
    if (request.method == "POST"):
        user_id = request.user
        student_id = Student.objects.select_related('id').only('id__id').get(id__id=user_id)
        form = MessInfoForm(request.POST)
        if form.is_valid():
            mess_option =  form.cleaned_data['mess_option']
            Messinfo.objects.create(student_id=student_id, mess_option=mess_option)
        return HttpResponseRedirect("/mess")
    user_id = request.user
    student_id = Student.objects.select_related(
        'id').only('id__id').get(id__id=user_id)
    if Messinfo.objects.filter(student_id=student_id).exists():
        return HttpResponseRedirect("/mess")
    form = MessInfoForm()
    context = {
        "form": form
    }
    return render(request, "messModule/messInfoForm.html", context)





@csrf_exempt
@login_required
@transaction.atomic
def submit_mess_feedback(request):
    """
    This function is to record the feedback submitted

    @param:
        request: contains metadata about the requested page

    @variable:
        user: Current logged in user
        extra_info: Extra information of the user

    @return:
        data: to record success or any errors
    """
    user = request.user
    extra_info = ExtraInfo.objects.select_related().get(user=user)
    student = Student.objects.select_related('id','id__user','id__department').get(id=extra_info)
    if extra_info.user_type == 'student':
        data = add_mess_feedback(request, student)
        central_mess_notif(request.user, request.user, 'feedback_submitted')
        return JsonResponse(data)


@csrf_exempt
@login_required
@transaction.atomic
def mess_vacation_submit(request):
    """
    This function is to record vacation food requests

    @param
        request: contains metadata about the requested page

    @variables:
        user: Current user information
        extra_info: Extra information of the user
        student: Student information about the current user

    @return:
        data: JsonResponse
    """
    user = request.user
    extra_info = ExtraInfo.objects.select_related().get(user=user)
    student = Student.objects.select_related('id','id__user','id__department').get(id=extra_info)

    if extra_info.user_type == 'student':
        data = add_vacation_food_request(request, student)
        return JsonResponse(data)


@login_required
@transaction.atomic
def submit_mess_menu(request):
    """
    This function is to record mess menu change requests by the  mess_committee

    @param
        request: contains metadata about the requested page

    @variables:
        user: Current user information
        holds_designations: designation of current user
        extrainfo: Extra information of the user
        student: Student information about the current user
    """
    # TODO add ajax for this
    user = request.user
    holds_designations = HoldsDesignation.objects.select_related().filter(user=user)
    extrainfo = ExtraInfo.objects.select_related().get(user=user)
    designation = holds_designations
    student = Student.objects.select_related('id','id__user','id__department').get(id=extrainfo)

    context = {}


    data = add_menu_change_request(request,student)
    if data['status'] == 1:
        return HttpResponseRedirect("/mess")

    return render(request, 'messModule/mess.html', context)


@login_required
def menu_change_response(request):
    """
    This function is to respond to mess menu requests

    @param
        request: contains metadata about the requested page

    @variables:
        user: Current user information
        holds_designations: designation of current user
    """
    user = request.user
    holds_designations = HoldsDesignation.objects.select_related().filter(user=user)

    designation = holds_designations
    data = handle_menu_change_response(request)
    return JsonResponse(data)


@login_required
def response_vacation_food(request, ap_id):
    """
    This function records the response to vacation food requests

    @param
        request: contains metadata about the requested page

    @variables:
        user: Current user information
        holds_designations: designation of current user
    """
    user = request.user

    holds_designations = HoldsDesignation.objects.select_related().filter(user=user)
    designation = holds_designations

    for d in designation:
        if d.designation.name == 'mess_manager':
            data = handle_vacation_food_request(request, ap_id)
    return HttpResponseRedirect("/mess")


@login_required
@transaction.atomic
def regsubmit(request):
    """
    This function ise used to change mess option

    @param
        request: contains metadata about the requested page

    @variables:
        user: Current user information
        extrainfo: Extra information of the user
        student: Student information about the current user
        mess_info_inst: MessInfo Object of the current user
        monthly_bill_obj: Monthly_bill Object of the current user
        mess_monthly_bill: monthly bill of each month of a semester
    """
    i = 0
    j = 0
    month_1 = ['January', 'February', 'March', 'April', 'May', 'June']
    month_2 = ['July', 'August', 'September', 'October', 'November', 'December']
    user = request.user
    extrainfo = ExtraInfo.objects.select_related().get(user=user)

    if extrainfo.user_type == 'student':
        student = Student.objects.select_related('id','id__user','id__department').get(id=extrainfo)
        mess = request.POST.get('mess_type')
        mess_info_inst = Messinfo.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').get(student_id=student)
        mess_info_inst.mess_option = mess
        mess_info_inst.save()
        mess_reg = Mess_reg.objects.last()

        if Monthly_bill.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(student_id=student):
            return HttpResponseRedirect("/mess")

        else:

            if mess_reg.end_reg.strftime("%B") in month_1:
                mess_monthly_bill = []
                while i<=5:
                    monthly_bill_obj = Monthly_bill(student_id=student, month=month_1[i], year=year_last_g)
                    mess_monthly_bill.append(monthly_bill_obj)
                    i = i+1
                Monthly_bill.objects.bulk_create(mess_monthly_bill)

            else:
                mess_monthly_bill = []
                while j<=5:
                    monthly_bill_obj = Monthly_bill(student_id=student, month=month_2[j], year=year_last_g)
                    mess_monthly_bill.append(monthly_bill_obj)
                    j = j+1
                Monthly_bill.objects.bulk_create(mess_monthly_bill)

        return HttpResponseRedirect("/mess")

    else:
        return redirect('mess')


@login_required
@transaction.atomic
def start_mess_registration(request):
    """
    This function is to start mess registration

    @param
        request: contains metadata about the requested page

    @variables:
        user: Current user information
        designation: designation of current user to validate proper platform
    """
    #   TODO ajax convert add a section to see previous sessions as well as close a session
    user = request.user
    designation = HoldsDesignation.objects.select_related().filter(user=user)
    for d in designation:
        if d.designation.name == 'mess_manager':
            data = add_mess_registration_time(request)
            return JsonResponse(data)

@csrf_exempt
def closeRegistration(request):
    mess_reg = Mess_reg.objects.last()
    yesterday = date.today() - timedelta(days=1)
    Mess_reg.objects.filter(id=mess_reg.id).update(end_reg=yesterday)
    return HttpResponseRedirect('/mess')

@transaction.atomic
@csrf_exempt
def mess_leave_request(request):
    """
    This function is to record and validate leave requests

    @param
        request: contains metadata about the requested page

    @variables:
        user: Current user information
        extra_info: Extra information of the user
        student: Student information about the current user
    """
    user = request.user
    extra_info = ExtraInfo.objects.select_related().get(user=user)
    student = Student.objects.select_related('id','id__user','id__department').get(id=extra_info)
    data = add_leave_request(request, student)
    return JsonResponse(data)


@login_required
@transaction.atomic
def minutes(request):
    """
    This function is used to upload the minutes of the meeting

    @param
        request: contains metadata about the requested page

    @variables:
        form: MinuteForm Object to create minutes of the meeting
    """
    if request.method == 'POST' and request.FILES:
        form = MinuteForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect('/mess')
        else:
            return HttpResponseRedirect('/mess')


@csrf_exempt
@transaction.atomic
def invitation(request):
    """
    This function is to schedule a mess committee meeting

    @param
        request: contains metadata about the requested page
    """
    # todo add ajax to this page as well
    data = add_mess_meeting_invitation(request)
    # return HttpResponseRedirect("/mess")
    return JsonResponse(data)


@login_required
@transaction.atomic
@csrf_exempt
def rebate_response(request):
    """
    This function is to respond to rebate requests

    @param request: 
        request - contains metadata about the requested page 

    @variables: 
        user: Current user details
        designation : designation of the user

    @return:
        data: returns the status of the application
    """
    data = {
        'status': 1
    }
    user = request.user
    designation = HoldsDesignation.objects.select_related().filter(user=user)

    for d in designation:
        if d.designation.name == 'mess_manager':
            data = handle_rebate_response(request)
            print(data)
            print(request)
    return JsonResponse(data)


@login_required
@transaction.atomic
@csrf_exempt
def place_request(request):
    """
        This function is to place special food requests ( used by students )
        
        @params:
            request - contains metadata about the requested page 

        @variables:
            user: Current user details
        
        @return:
            data['status']: returns status of the application
    """
    user = request.user
    extra_info = ExtraInfo.objects.select_related().get(user=user)
    if extra_info.user_type == 'student':
        extra_info = ExtraInfo.objects.select_related().get(user=user)
        student = Student.objects.select_related('id','id__user','id__department').get(id=extra_info)
        data = add_special_food_request(request, student)
        return JsonResponse(data)


@login_required
@transaction.atomic
@csrf_exempt
def special_request_response(request):
    """
       This function is to respond to special request for food submitted by students
       data: message regarding the request
    """

    data = handle_special_request(request)
    return JsonResponse(data)


@login_required
@transaction.atomic
@csrf_exempt
def update_cost(request):
    """
    This function is to update the base cost of the monthly central mess bill
    
    @param:
        request - contains metadata about the requested page 

    @variables:  
        user - contains user details
    """
    user = request.user
    data = add_bill_base_amount(request)
    return JsonResponse(data)

@login_required
def update_semdates(request):
    """
    This function is to update the semester start and end date
    
    @param:
        request - contains metadata about the requested page 

    @variables:  
        user - contains user details
    """
    user = request.user
    data = add_sem_dates(request)
    return HttpResponseRedirect('/mess')
@csrf_exempt
@login_required
def update_bill(request):
    # user = request.user
    update_month_bill(request)
    return HttpResponseRedirect('/mess')

def generate_mess_bill(request):
    """
    This function is to generate the bill of the students

    @param:
        request - contains metadata about the requested page 

    @variables:
        user: stores current user information
        nonveg_data : stores records of non-veg ordered by a student
        year_now: current year
        month_now: current month
        amount_m: monhly base amount
        students: information of all students
        mess_info: Mess Information, mainly choice of mess
        rebates: Rebate records of students
    """
    # todo generate proper logic for generate_mess_bill
    user = request.user
    # t1 = Thread(target=generate_bill, args=())
    # t1.setDaemon(True)
    # t1.start()
    generate_bill()
    data ={
        'status': 1
    }
    return JsonResponse(data)

    
class MenuPDF(View):

    def post(self, request, *args, **kwargs):
        user = request.user
        extra_info = ExtraInfo.objects.select_related().get(user=user)
        y = Menu.objects.all()

        context = {
            'menu': y,
            'mess_option': 'mess2',
            'date':str(today_g.date())
        }
        return render_to_pdf('messModule/menudownloadable2.html', context)
    # return HttpResponse(pdf, content_type='application/pdf')


class MenuPDF1(View):
    """
    This function is to generate the menu in pdf format (downloadable) for mess 1

    @param:
        request - contains metadata about the requested page 

    @variables:
        current_user - get user from request
        user_details - extract details of the user from the database
    """
    def post(self, request, *args, **kwargs):
        user = request.user
        # extrainfo = ExtraInfo.objects.get(user=user)
        y = Menu.objects.all()
        context = {
            'menu': y,
            'mess_option': 'mess1',
            'date':str(today_g.date())
        }
        return render_to_pdf('messModule/menudownloadable1.html', context)
    
class BillPDFStudent(View):
    def post(self, request, *args, **kwargs):
        user = request.user
        extra_info = ExtraInfo.objects.select_related().get(user=user)
        student = Student.objects.select_related('id','id__user','id__department').get(id=extra_info)
        # reg_student = Reg_records.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').get(student_id_id=student)
        try:
            monthly_bill = Monthly_bill.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(student_id=student)
            if monthly_bill.exists():
                context = {
                    'student_bill': monthly_bill
                }
                return render_to_pdf('messModule/billpdfexport.html', context)
            else :
                return HttpResponseRedirect('/mess')
        except:
            return HttpResponseRedirect('/mess')



def menu_change_request(request):
    """
    This function is to request a change in menu

    @param:
        request - contains metadata about the requested page 

    @variables:
        current_user - get user from request
        user_details - extract details and designation of the user from the database
        new_menu - the new menu to replace the previous menu
    """
    newmenu = Menu_change_request.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department','dish').filter(status=2)
    data = model_to_dict(newmenu)
    return JsonResponse(data)


@csrf_exempt
def submit_mess_committee(request):
    """
    This function is to add the new mess committee

    @param:
        request - contains metadata about the requested page 

    @variables:
        current_user - get user from request
        user_details - extract details and designation of the user from the database
    """
    roll_number = str(request.POST.get('roll_number')).upper()
    data = add_mess_committee(request, roll_number)
    return HttpResponseRedirect("/mess")


def remove_mess_committee(request):
    """
    This function is to remove the current mess committee

    @param:
        request - contains metadata about the requested page 

    @variables:
        current_user - get user from request
        user_details - extract details and designation of the user from the database

    """
    member_id = request.POST['member_id']
    data_m = member_id.split("-")
    roll_number = data_m[1]
    if data_m[0] == 'mess_committee':
        designation = Designation.objects.get(name='mess_committee')
    elif data_m[0] == 'mess_convener':
        designation = Designation.objects.get(name='mess_convener')
    elif data_m[0] == 'mess_committee_mess2':
        designation = Designation.objects.get(name='mess_committee_mess2')
    else:
        designation = Designation.objects.get(name='mess_convener_mess2')
    remove_object = HoldsDesignation.objects.select_related().get(Q(user__username=roll_number) & Q(designation=designation))
    remove_object.delete()
    data = {
        'status': 1,
        'message': 'Successfully removed '
    }
    return JsonResponse(data)


def get_leave_data(request):
    leave_data = Rebate.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(Q(start_date__lte=today_g)&Q(end_date__gte=today_g)).count()
    leave_data_t = Rebate.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(Q(start_date__lte=tomorrow_g)&Q(end_date__gte=tomorrow_g)).count()
    data = {
        'status': 1,
        'message': 'HI I AM WORKING',
        'today': today_g.date(),
        'tomorrow': tomorrow_g.date(),
        'counttoday': leave_data,
        'counttomorrow':leave_data_t
    }
    return JsonResponse(data)


def accept_vacation_leaves(request):
    """
    This function is to accept vacation leave request

    @param:
        request - contains metadata about the requested page 
        request - details about leave

    @variables:
        current_user - get user from request
        user_details - extract details and designation of the user from the database
        start_date_leave - Starting date of leave
        end_date_leave - Ending date of leave
    """
    start_date_leave = request.GET['start_date']
    end_date_leave = request.GET['end_date']
    leave_data = Rebate.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(Q(start_date__gte=start_date_leave)
                                       &Q(end_date__lte=end_date_leave)
                                       &Q(leave_type="vacation")
                                       &Q(status='1'))

    if leave_data:
        for item in leave_data:
            item.status = '2'
            item.save()

    data = {
        'status': 1,
        'display': 'Vacation Leaves Successfully Accepted'
    }
    return JsonResponse(data)


def select_mess_convener(request):
    """
    This function is to select a new convenor for mess

    @param:
        request - contains metadata about the requested page 
        
    @variables:
        current_user - get user from request
        user_details - extract details and designation of the user from the database
        designation - to get the designation of the user
        first_day_of_the_month - first day of the month
        last_day_of_the_month - last day of the month
        previous_month - last month
        bill_object - accessing the bill from Monthly bills
    """
    member_id = request.POST['member_id_add']
    data_m = member_id.split("-")
    roll_number = data_m[1]

    if data_m[0] == 'mess_committee':
        designation = Designation.objects.get(name='mess_committee')
        new_designation = Designation.objects.get(name='mess_convener')
        # One mess can have only one mess convener
        existing_check = HoldsDesignation.objects.select_related().filter(designation=new_designation)
        if existing_check.count():
            data = {
                'status': 1,
                'message': 'Mess Convener already exists for Mess 1 ! \nRemove the existing convener to add new one'
            }
            return JsonResponse(data)
        else:
            modify_object = HoldsDesignation.objects.select_related().get(Q(user__username=roll_number) & Q(designation=designation))
            modify_object.designation = new_designation
            modify_object.save()
    else:
        designation = Designation.objects.get(name='mess_committee_mess2')
        new_designation = Designation.objects.get(name='mess_convener_mess2')
        existing_check = HoldsDesignation.objects.select_related().filter(designation=new_designation)
        if existing_check.count():
            data = {
                'status': 1,
                'message': 'Mess Convener already exists for Mess 2 ! \n Remove the existing convener to add new one'
            }
            return JsonResponse(data)
        else:
            modify_object = HoldsDesignation.objects.select_related().get(Q(user__username=roll_number) & Q(designation=designation))
            modify_object.designation = new_designation
            modify_object.save()

    data = {
        'status': 1,
        'message': 'Successfully added as mess convener ! '
    }
    return JsonResponse(data)


def download_bill_mess(request):
    """
    This function is to get the mess bill for current month

    @param:
        request - contains metadata about the requested page 
        request - first day of the month and last day of the previous month

    @variables:
        current_user - get user from request
        user_details - extract details of the user from the database
        first_day_of_the_month - first day of the month
        last_day_of_the_month - last day of the month
        previous_month - last month
        bill_object - accessing the bill from Monthly bills
    """
    user = request.user
    extra_info = ExtraInfo.objects.select_related().get(user=user)
    first_day_of_this_month = date.today().replace(day=1)
    last_day_prev_month = first_day_of_this_month - timedelta(days=1)
    previous_month = last_day_prev_month.strftime('%B')
    bill_object = Monthly_bill.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(Q(month=previous_month)&Q(year=year_last_g))
    # bill_object = Monthly_bill.objects.all()

    context = {
        'bill': bill_object,
    }
    return render_to_pdf('messModule/billpdfexport.html', context)

def add_leave_manager(request):
    """
    This function is to apply for leave

    @param:
        request - contains metadata about the requested page 
        request - start date, end date and type of leave

    @variables:
        current_user - get user from request
        user_details - extract details of the user from the database
        start_date - starting date of the leave
        end_data - ending date of leave
        type - type of leave
    """
    flag = 1
    start_date = request.POST.get('l_startd')
    end_date = request.POST.get('l_endd')
    roll_number = request.POST.get('l_rollno')
    type = request.POST.get('l_type')
    purpose = request.POST.get('l_purpose')
    student = Student.objects.select_related('id','id__user','id__department').get(id__id=roll_number)
    add_obj = Rebate(student_id = student,
                     start_date = start_date,
                     end_date = end_date,
                     purpose = purpose,
                     status='2',
                     leave_type=type)

    if (end_date < start_date):
        data = {
            'status': 3,
            'message': "Please check the dates"
        }
        flag = 0
        return HttpResponse('Check the dates')

    date_format = "%Y-%m-%d"
    b = datetime.strptime(str(start_date), date_format)
    d = datetime.strptime(str(end_date), date_format)

    rebates = Rebate.objects.select_related('student_id','student_id__id','student_id__id__user','student_id__id__department').filter(student_id=student)
    rebate_check = rebates.filter(status='2')

    for r in rebate_check:
        a = datetime.strptime(str(r.start_date), date_format)
        c = datetime.strptime(str(r.end_date), date_format)
        if ((b <= a and (d >= a and d <= c)) or (b >= a and (d >= a and d <= c))
                or (b <= a and (d >= c)) or ((b >= a and b <= c) and (d >= c))):
            flag = 0
            data = {
                'status': 3,
                'message': "Already applied for these dates",
            }
            return HttpResponse('You are seeing this page : As the leave has been applied for these days already')
    if flag == 1:
        message = 'Your leave request has been accepted between dates ' + str(b.date()) + ' and ' + str(d.date())
        central_mess_notif(request.user, student.id.user, 'leave_request', message)
        add_obj.save()
    return HttpResponseRedirect('/mess')

def update_menu2(request):
    if (request.method == "POST"):
        mb = request.POST['MB2']
        ml = request.POST['ML2']
        md = request.POST['MD2']
        sud = request.POST['SUD2']
        sul = request.POST['SUL2']
        sub = request.POST['SUB2']
        sd = request.POST['SD2']
        sl = request.POST['SL2']
        sb = request.POST['SB2']
        fd = request.POST['FD2']
        fl = request.POST['FL2']
        fb = request.POST['FB2']
        thd = request.POST['THD2']
        thl = request.POST['THL2']
        thb = request.POST['THB2']
        wd = request.POST['WD2']
        wl = request.POST['WL2']
        wb = request.POST['WB2']
        td = request.POST['TD2']
        tl = request.POST['TL2']
        tb = request.POST['TB2']
        
        print("mb", mb)
        Menu.objects.filter(mess_option = 'mess2',meal_time='MB').update(dish = mb)
        Menu.objects.filter(mess_option = 'mess2',meal_time='ML').update(dish = ml)
        Menu.objects.filter(mess_option = 'mess2',meal_time='MD').update(dish = md)
        Menu.objects.filter(mess_option = 'mess2',meal_time='TB').update(dish = tb)
        Menu.objects.filter(mess_option = 'mess2',meal_time='TL').update(dish = tl)
        Menu.objects.filter(mess_option = 'mess2',meal_time='TD').update(dish = td)
        Menu.objects.filter(mess_option = 'mess2',meal_time='WB').update(dish = wb)
        Menu.objects.filter(mess_option = 'mess2',meal_time='WL').update(dish = wl)
        Menu.objects.filter(mess_option = 'mess2',meal_time='WD').update(dish = wd)
        Menu.objects.filter(mess_option = 'mess2',meal_time='THB').update(dish = thb)
        Menu.objects.filter(mess_option = 'mess2',meal_time='THL').update(dish = thl)
        Menu.objects.filter(mess_option = 'mess2',meal_time='THD').update(dish = thd)
        Menu.objects.filter(mess_option = 'mess2',meal_time='FB').update(dish = fb)
        Menu.objects.filter(mess_option = 'mess2',meal_time='FL').update(dish = fl)
        Menu.objects.filter(mess_option = 'mess2',meal_time='FD').update(dish = fd)
        Menu.objects.filter(mess_option = 'mess2',meal_time='SB').update(dish = sb)
        Menu.objects.filter(mess_option = 'mess2',meal_time='SL').update(dish = sl)
        Menu.objects.filter(mess_option = 'mess2',meal_time='SD').update(dish = sd)
        Menu.objects.filter(mess_option = 'mess2',meal_time='SUB').update(dish = sub)
        Menu.objects.filter(mess_option = 'mess2',meal_time='SUL').update(dish = sul)
        Menu.objects.filter(mess_option = 'mess2',meal_time='SUD').update(dish = sud)


    return redirect('/mess')
def update_menu1(request):
    if (request.method == "POST"):
        mb1 = request.POST['MB1']
        ml1 = request.POST['ML1']
        md1 = request.POST['MD1']
        sud1 = request.POST['SUD1']
        sul1 = request.POST['SUL1']
        sub1 = request.POST['SUB1']
        sd1 = request.POST['SD1']
        sl1 = request.POST['SL1']
        sb1 = request.POST['SB1']
        fd1 = request.POST['FD1']
        fl1 = request.POST['FL1']
        fb1 = request.POST['FB1']
        thd1 = request.POST['THD1']
        thl1 = request.POST['THL1']
        thb1 = request.POST['THB1']
        wd1 = request.POST['WD1']
        wl1 = request.POST['WL1']
        wb1 = request.POST['WB1']
        td1 = request.POST['TD1']
        tl1 = request.POST['TL1']
        tb1 = request.POST['TB1']
        
        print("mb", mb1)
        Menu.objects.filter(mess_option = 'mess1',meal_time='MB').update(dish = mb1)
        Menu.objects.filter(mess_option = 'mess1',meal_time='ML').update(dish = ml1)
        Menu.objects.filter(mess_option = 'mess1',meal_time='MD').update(dish = md1)
        Menu.objects.filter(mess_option = 'mess1',meal_time='TB').update(dish = tb1)
        Menu.objects.filter(mess_option = 'mess1',meal_time='TL').update(dish = tl1)
        Menu.objects.filter(mess_option = 'mess1',meal_time='TD').update(dish = td1)
        Menu.objects.filter(mess_option = 'mess1',meal_time='WB').update(dish = wb1)
        Menu.objects.filter(mess_option = 'mess1',meal_time='WL').update(dish = wl1)
        Menu.objects.filter(mess_option = 'mess1',meal_time='WD').update(dish = wd1)
        Menu.objects.filter(mess_option = 'mess1',meal_time='THB').update(dish = thb1)
        Menu.objects.filter(mess_option = 'mess1',meal_time='THL').update(dish = thl1)
        Menu.objects.filter(mess_option = 'mess1',meal_time='THD').update(dish = thd1)
        Menu.objects.filter(mess_option = 'mess1',meal_time='FB').update(dish = fb1)
        Menu.objects.filter(mess_option = 'mess1',meal_time='FL').update(dish = fl1)
        Menu.objects.filter(mess_option = 'mess1',meal_time='FD').update(dish = fd1)
        Menu.objects.filter(mess_option = 'mess1',meal_time='SB').update(dish = sb1)
        Menu.objects.filter(mess_option = 'mess1',meal_time='SL').update(dish = sl1)
        Menu.objects.filter(mess_option = 'mess1',meal_time='SD').update(dish = sd1)
        Menu.objects.filter(mess_option = 'mess1',meal_time='SUB').update(dish = sub1)
        Menu.objects.filter(mess_option = 'mess1',meal_time='SUL').update(dish = sul1)
        Menu.objects.filter(mess_option = 'mess1',meal_time='SUD').update(dish = sud1)

    return redirect('/mess')

@csrf_exempt
def searchAddOrRemoveStudent(request):
    if request.method=='GET':
        submitType=request.GET.get('type')
        msg=""
        if submitType=='searchStudent':
            studentId=str((request.GET.get('roll_number'))).upper()
            try:
                reg_main = Reg_main.objects.values('current_mess_status','mess_option').get(student_id=studentId)
                if(reg_main['current_mess_status']=="Registered"):
                    msg= str(studentId)+" is registered for "+str(reg_main['mess_option'])
                else:
                    msg=str(studentId)+" is not registered for Mess" 
            except:
                msg="unable to find this student in database."                    

            # try:    
            #     mess_optn = Messinfo.objects.select_related().values('mess_option').get(student_id=studentId)
            #     msg= str(studentId)+" is registered for "+str(mess_optn['mess_option'])
            # except:
            #     msg=str(studentId)+" is not registered for Mess" 
                
        elif submitType=='addStudent':
            messNo=request.GET.get('messNo')  
            studentId = str((request.GET.get('roll_number'))).upper()
            try:
                reg_main = Reg_main.objects.get(student_id=studentId)

                if(reg_main.current_mess_status=="Registered"):
                    msg=str(studentId)+" is already registered for "+str(reg_main.mess_option) 
                else:
                    reg_main.current_mess_status="Registered"
                    reg_main.mess_option=str(messNo)
                    reg_main.save()
                    msg="success"
            except:
                msg="unable to find this student in database."        

            # try:
            #     mess_optn = Messinfo.objects.select_related().values('mess_option').get(student_id=studentId)
            #     msg=str(studentId)+" is already registered for "+str(mess_optn['mess_option']) 
            # except:
            #     try:
            #         studentHere = Student.objects.select_related('id','id__user','id__department').get(id=studentId)
            #         newData=Messinfo(student_id=studentHere,mess_option=str(messNo))
            #         newData.save()
            #         msg=str(studentId)+" is successfully registered for Mess."
            #     except:
            #         msg="unable to find this student in database."
                
        elif submitType=='removeStudent':
            studentId = str((request.GET.get('roll_number'))).upper()
            try:
                reg_main = Reg_main.objects.get(student_id=studentId,current_mess_status="Registered")
                reg_main.current_mess_status="Deregistered"
                reg_main.save()
                msg=str(studentId)+" is successfully removed from mess." 
            except:
                msg=str(studentId)+" is not registered for mess." 

            # try:
            #     studentHere = Student.objects.select_related('id','id__user','id__department').get(id=studentId)
            #     data=Messinfo.objects.get(student_id=studentId)
            #     data.delete()
            #     Messinfo.objects.all()
            #     msg=str(studentId)+" is successfully removed from mess." 
            # except:
            #     msg=str(studentId)+" is not registered for mess." 
        elif (submitType=='removeAllStudent1' or submitType=='removeAllStudent2'):
            messNo=request.GET.get('mess')
    
            try:
                reg_main = Reg_main.objects.filter(mess_option=str(messNo),current_mess_status="Registered")
                for reg in reg_main:
                    reg.current_mess_status="Deregistered"
                    reg.save()
                msg="All students removed successfully from "+str(messNo)
            except:
                msg="can't remove students." 

        return JsonResponse({'message':msg})
    else:
        if(request.FILES):
            # if 'excelUpload1' in request.POST:
            #     messNo='mess1'
            #     excel_file = request.FILES['excel_file1']
            # else: 
            #     messNo='mess2'
            #     excel_file = request.FILES['excel_file2']
            try:
                latest = Semdates.objects.latest('end_date')
                latest_end_date=latest.end_date
                print(latest_end_date)
            except:
                latest_end_date=None
            excel_file = request.FILES['excel_file1']
            wb = openpyxl.load_workbook(excel_file)
            flag = False
            for row in wb.active:
                if(flag==False):
                    flag=True
                    continue
                studentId=(str(row[0].value)).upper()
                studentHere = Student.objects.select_related('id','id__user','id__department').get(id=studentId)
                balance=row[1].value
                messNo = row[2].value
                try:                    
                    reg_main = Reg_main.objects.get(student_id=studentId)
                    reg_main.current_mess_status="Registered"
                    reg_main.mess_option=str(messNo)
                    reg_main.balance=reg_main.balance+balance
                    reg_main.save()
                    # if Messinfo.objects.filter(student_id=studentId).exists():
                    #     Messinfo.objects.filter(student_id=studentId).update(mess_option=str(messNo))
                    # else:
                    #     newData=Messinfo(student_id=studentHere,mess_option=str(messNo))
                    #     newData.save()
                except:
                    reg_main = Reg_main(student_id=reg_main.student_id,program=studentHere.programme,current_mess_status="Registered",mess_option=str(messNo),balance=balance)
                    reg_main.save()

                new_reg_record = Reg_records(student_id=reg_main.student_id,start_date=today_g,end_date=latest_end_date)
                new_reg_record.save()
    # messages.success(request,"Done.")
    return HttpResponseRedirect("/mess")
        
@csrf_exempt
def uploadPaymentDue(request):
    if(request.FILES):
        
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file)

        for row in wb.active:
            studentId=(str(row[0].value)).upper()
            amount=(row[1].value)
            try:
                studentHere = Student.objects.get(id=studentId)
                monthly_bill = Monthly_bill.objects.select_related('student_id').filter(student_id=studentHere)
                try:
                    Payments.objects.filter(student_id=studentHere).delete()
                except:
                    1
                Payments.objects.create(student_id=studentHere,amount_paid=(-1*(amount)))
                if amount<=0:
                    for x in monthly_bill:
                        Monthly_bill.objects.filter(student_id=studentHere).filter(month=x.month).filter(year=x.year).update(paid=True)
                else:
                    monthly_bill=monthly_bill[::-1]
                    curr_amount=amount
                    for x in monthly_bill:
                        if(curr_amount<=0):
                            Monthly_bill.objects.filter(student_id=studentHere).filter(month=x.month).filter(year=x.year).update(paid=True)
                        else:
                            Monthly_bill.objects.filter(student_id=studentHere).filter(month=x.month).filter(year=x.year).update(paid=False)
                        curr_amount-=x.total_bill;    
                        print(x)
            except:
                1
    messages.success(request,"Done.")
    return HttpResponseRedirect("/mess")
        
        
@csrf_exempt
@login_required
def respond_to_reg(request):
    """
    This function is used to respond to registeration requests

    @param request: 
        request - contains metadata about the requested page 

    @variables: 
        user: Current user details
        designation : designation of the user

    @return:
        data: returns the status of the application
    """
    data = {
        'status': 1
    }
    user = request.user
    designation = HoldsDesignation.objects.select_related().filter(user=user)
    type = request.POST['type']
    for d in designation:
        if d.designation.name == 'mess_manager':
            if(type=='reg'):
                data = handle_reg_response(request)
            elif(type=='dreg'):
                data = handle_dreg_response(request)    
    return JsonResponse(data)
    
@csrf_exempt
@login_required
def respond_to_update_payment(request):
    data = {
        'status': 1
    }
    user = request.user
    designation = HoldsDesignation.objects.select_related().filter(user=user)
    for d in designation:
        if d.designation.name == 'mess_manager':
            data = handle_update_payment_response(request)
            
    return JsonResponse(data)



def reg_request(request):

    user = request.user
    extra_info = ExtraInfo.objects.select_related().get(user=user)
    try:
        if request.POST['input_roll']:
            # print(request.POST)
            studentID = str(request.POST['input_roll']).upper()
            handle_add_reg(request)
            form = RegistrationRequest(request.POST, request.FILES)
            student = Student.objects.select_related('id','id__user','id__department').get(id=studentID)
            if form.is_valid():
                temp=form.save(commit=False)
                temp.student_id=student
                temp.status='accept'
                temp.save()
            return HttpResponseRedirect("/mess")  
    except:
        student = Student.objects.select_related('id','id__user','id__department').get(id=extra_info)
        if request.method == 'POST':
            form = RegistrationRequest(request.POST, request.FILES)

            if form.is_valid():
                temp=form.save(commit=False)
                temp.student_id=student
                temp.save()
                return HttpResponseRedirect("/mess")  

            
def update_payment(request):
    user = request.user
    extra_info = ExtraInfo.objects.select_related().get(user=user)    
    student = Student.objects.select_related('id','id__user','id__department').get(id=extra_info)
    if request.method == 'POST':
        form = UpdatePaymentRequest(request.POST, request.FILES)

        if form.is_valid():
            temp=form.save(commit=False)
            temp.student_id=student
            temp.save()
            return HttpResponseRedirect("/mess")

@csrf_exempt
def update_bill_excel(request):
    if(request.FILES):   
            excel_file = request.FILES['excel_file_bill']
            wb = openpyxl.load_workbook(excel_file)
            flag = False
            for row in wb.active:
                if(flag==False):
                    flag=True
                    continue
                studentId=(str(row[0].value)).upper()
                studentHere = Student.objects.select_related('id','id__user','id__department').get(id=studentId)
                month=str(row[1].value)
                year = row[2].value
                amt = row[3].value
                rebate_cnt = row[4].value
                rebate_amt = row[5].value
                total_amt = row[6].value
                try:                    
                    bill = Monthly_bill.objects.get(student_id=studentId,month=month,year=year)
                    reg_main = Reg_main.objects.get(student_id=studentId)
                    reg_main.balance=reg_main.balance+bill.total_bill
                    bill.amount=amt
                    bill.rebate_count=rebate_cnt
                    bill.rebate_amount=rebate_amt
                    bill.total_bill=total_amt
                    reg_main.balance=reg_main.balance-total_amt
                    bill.save()
                    reg_main.save()
                except:
                    bill = Monthly_bill(student_id=studentHere,month=month,year=year,amount=amt,rebate_count=rebate_cnt,rebate_amount=rebate_amt,total_bill=total_amt)
                    bill.save()
    # messages.success(request,"Done.")
    return HttpResponseRedirect("/mess")
  
def de_reg_request(request):
    try:
        if request.POST['input_roll']:
            # print(request.POST)
            studentID = str(request.POST['input_roll']).upper()
            end_date = request.POST.get("end_date")
            try:
                reg_main = Reg_main.objects.get(student_id=studentID)
                
                if(end_date == str(date.today())):
                    reg_main.current_mess_status = 'Deregistered'
                    reg_main.save()
                reg_record = Reg_records.objects.filter(student_id=studentID).latest('start_date')
                reg_record.end_date=end_date
                reg_record.save()
            except:
                pass
            return  HttpResponseRedirect('/mess')
    except:
        data={
            'message':'request submitted successfully'
        }
        user = request.user
        end_date = request.POST.get("end_date")
        print(end_date)
        extra_info = ExtraInfo.objects.select_related().get(user=user)
        student = Student.objects.select_related('id','id__user','id__department').get(id=extra_info)
        new_req=Deregistration_Request(student_id=student, end_date=end_date)
        new_req.save()
        return  HttpResponseRedirect('/mess')

