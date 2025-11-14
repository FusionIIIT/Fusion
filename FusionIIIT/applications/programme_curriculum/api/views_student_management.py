import json
import pandas as pd
import openpyxl
import random
import secrets
import string
import sys
import os
from io import BytesIO
from datetime import datetime, date
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db import transaction, connection
from django.db.models import Q, Count
from django.conf import settings
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from applications.academic_information.models import Student as AcademicStudent
from applications.globals.models import ExtraInfo, Designation, HoldsDesignation
from django.contrib.auth.models import User
from applications.programme_curriculum.models import (
    Programme, Curriculum, Batch, Discipline
)

from applications.programme_curriculum.views_password_email import send_password_email_smtp
from applications.programme_curriculum.models_password_email import PasswordEmailLog

try:
    from applications.programme_curriculum.models_student_management import (
        StudentBatchUpload, BatchConfiguration, StudentStatusLog
    )
except ImportError:
    from django.db import models
    from django.contrib.auth.models import User
    pass

def parse_request_data(request, field_mappings=None):
    """
    Helper function to parse request data from various formats with field mapping
    
    Args:
        request: Django request object
        field_mappings: Dict mapping backend_field -> frontend_field(s)
                       e.g. {'year': ['year', 'batchYear'], 'total_seats': 'totalSeats'}
    
    Returns:
        Parsed and mapped data dictionary
    """
    try:
        data = json.loads(request.body)

        if not isinstance(data, dict):
            return {}
    except json.JSONDecodeError:
        # Fallback: accept form-encoded or multipart form data from frontend
        try:
            data = request.POST.dict() if hasattr(request, 'POST') else {}
        except Exception:
            data = {}
    
    # Apply field mappings if provided
    if field_mappings:
        mapped_data = {}
        for backend_field, frontend_fields in field_mappings.items():
            if isinstance(frontend_fields, list):
                for field in frontend_fields:
                    if data.get(field):
                        mapped_data[backend_field] = data.get(field)
                        break
            else:
                mapped_data[backend_field] = data.get(frontend_fields) or data.get(backend_field)

        mapped_frontend_fields = set()
        for frontend_fields in field_mappings.values():
            if isinstance(frontend_fields, list):
                mapped_frontend_fields.update(frontend_fields)
            else:
                mapped_frontend_fields.add(frontend_fields)
        
        for key, value in data.items():
            if key not in mapped_frontend_fields and key not in mapped_data:
                mapped_data[key] = value
        
        return mapped_data
    
    return data

def sanitize_phone_number(phone_value):
    if phone_value is None:
        return phone_value
    phone_str = str(phone_value)
    if phone_str.endswith('.0'):
        phone_str = phone_str[:-2]
    return phone_str

def validate_phone_numbers(student_data, field_map=None):
    """
    Returns (student_phone, father_phone, errors) tuple
    """
    if field_map is None:
        field_map = {
            'student': ['phone_number', 'Mobile No', 'phoneNumber'],
            'father': ['father_mobile', 'Father Mobile Number', 'fatherMobile']
        }
    
    errors = []
    student_phone = None
    father_phone = None
    
    # Get student phone
    student_phone_raw = None
    for field in field_map['student']:
        if student_data.get(field):
            student_phone_raw = student_data.get(field)
            break
    
    # Get father phone  
    father_phone_raw = None
    for field in field_map['father']:
        if student_data.get(field):
            father_phone_raw = student_data.get(field)
            break
    
    # Validate and sanitize phones
    if student_phone_raw:
        sp = sanitize_phone_number(student_phone_raw)
        student_phone = sp.replace(' ', '').replace('-', '') if sp else None
        if student_phone and (not student_phone.isdigit() or len(student_phone) != 10):
            errors.append(f'Invalid student mobile: {student_phone_raw}')
    
    if father_phone_raw:
        fp = sanitize_phone_number(father_phone_raw)  
        father_phone = fp.replace(' ', '').replace('-', '') if fp else None
        if father_phone and (not father_phone.isdigit() or len(father_phone) != 10):
            errors.append(f'Invalid father mobile: {father_phone_raw}')

    if student_phone and father_phone and student_phone == father_phone:
        errors.append("Father's mobile number should not be same as student's mobile number")
    
    return student_phone, father_phone, errors

def sanitize_rank_value(rank_value):
    if rank_value is None or rank_value == '' or rank_value == 'null':
        return None
    rank_str = str(rank_value).strip()
    if not rank_str or rank_str.lower() == 'nan':
        return None
    if rank_str.endswith('.0'):
        rank_str = rank_str[:-2]
    return rank_str

def _safe_int_conversion(value):
    """Convert a value to int, returning None for empty/invalid values"""
    if value is None or value == '' or value == 'null':
        return None
    try:
        value_str = str(value).replace(',', '').strip()
        if value_str and value_str.isdigit():
            return int(value_str)
    except (ValueError, TypeError):
        pass
    return None

def parse_date_flexible(date_value):
    if date_value is None or date_value == '':
        return None
    
    try:
        if hasattr(date_value, 'date'):
            return date_value.date()

        date_str = str(date_value).strip()
        if not date_str or date_str.lower() == 'nan':
            return None

        date_formats = [
            '%d-%m-%Y',    # dd-mm-yyyy (Excel common format)
            '%d/%m/%Y',    # dd/mm/yyyy
            '%Y-%m-%d',    # yyyy-mm-dd (ISO format)
            '%m/%d/%Y',    # mm/dd/yyyy (US format)
            '%d-%m-%y',    # dd-mm-yy
            '%d/%m/%y',    # dd/mm/yy
            '%Y/%m/%d',    # yyyy/mm/dd
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        
        return None
        
    except Exception:
        return None

def get_batch_year_from_academic_year(academic_year):
    if isinstance(academic_year, str):
        if '-' in academic_year:
            return int(academic_year.split('-')[0])
        else:
            return int(academic_year)
    return int(academic_year)

def get_academic_year_from_batch_year(batch_year):
    batch_year = int(batch_year)
    next_year = batch_year + 1
    return f"{batch_year}-{str(next_year)[-2:]}"

def calculate_batch_filled_seats(batch):
    """
    Centralized function to calculate filled seats for a batch using priority-based algorithm.
    """
    try:
        from applications.academic_information.models import Student
        direct_count = Student.objects.filter(batch_id=batch).count()
        if direct_count > 0:
            return direct_count
        else:
            discipline_name = batch.discipline.name if batch.discipline else ''
            discipline_acronym = batch.discipline.acronym if batch.discipline else ''
            discipline_count = Student.objects.filter(
                batch=batch.year,
                id__department__name__in=[discipline_acronym, discipline_name.split()[0] if discipline_name else '']
            ).count() if discipline_name else 0
            if discipline_count > 0:
                return discipline_count
            else:
                if batch.name and batch.year:
                    programme_count = Student.objects.filter(
                        programme__icontains=batch.name.split()[0] if batch.name else '',
                        batch=batch.year,
                        id__department__name__in=[discipline_acronym] if discipline_acronym else []
                    ).count()
                    return programme_count
                else:
                    return 0
                    
    except Exception as e:
        return 0

def normalize_year_input(year_input):
    try:
        if isinstance(year_input, str):
            if '-' in year_input:
                parts = year_input.split('-')
                if len(parts) != 2:
                    raise ValueError(f"Invalid academic year format: {year_input}")
                batch_year = int(parts[0])
                second_year = int(parts[1])
                if second_year != (batch_year + 1) % 100:
                    raise ValueError(f"Invalid academic year format: {year_input}")
                academic_year = year_input
            else:
                batch_year = int(year_input)
                academic_year = get_academic_year_from_batch_year(batch_year)
        else:
            batch_year = int(year_input)
            academic_year = get_academic_year_from_batch_year(batch_year)
        
        return batch_year, academic_year
    except (ValueError, TypeError) as e:
        raise ValueError(f"Invalid year format: {year_input}. Expected formats: 2025, '2025', or '2025-26'")

def get_current_academic_year():
    current_year = datetime.now().year
    if datetime.now().month >= 7:
        batch_year = current_year
    else:
        batch_year = current_year - 1
    
    academic_year = get_academic_year_from_batch_year(batch_year)
    return batch_year, academic_year

def validate_and_normalize_year(selected_year):
    """
    Returns (batch_year, academic_year) or raises JsonResponse error
    """
    if selected_year:
        try:
            return normalize_year_input(selected_year)
        except ValueError as e:
            return JsonResponse({
                'success': False,
                'message': f'Invalid year format: {str(e)}'
            }, status=400)
    else:
        return get_current_academic_year()

def validate_curriculum_exists(action_context="performing action"):
    """
    Returns JsonResponse error if no curriculums found, None if valid
    """
    from applications.programme_curriculum.models import Curriculum
    
    available_curriculums = Curriculum.objects.filter(working_curriculum=True)
    if not available_curriculums.exists():
        return JsonResponse({
            'success': False,
            'message': f'CURRICULUM REQUIRED: No working curriculums found. Please create a curriculum first before {action_context}.',
            'validation_error': 'missing_curriculum'
        }, status=400)
    return None

def validate_batch_curriculum_requirements(batch_year, academic_year, action_context="performing action"):
    """
    Returns JsonResponse error if validation fails, None if valid
    """
    from applications.programme_curriculum.models import Batch

    existing_batches = Batch.objects.filter(year=batch_year, running_batch=True)
    if not existing_batches.exists():
        return JsonResponse({
            'success': False,
            'message': f'BATCH REQUIRED: No active batches found for academic year {academic_year} (batch year {batch_year}). Please create batches with assigned curriculums first before {action_context}.',
            'validation_error': 'missing_batch',
            'academic_year': academic_year,
            'batch_year': batch_year
        }, status=400)

    batches_without_curriculum = existing_batches.filter(curriculum__isnull=True)
    
    if batches_without_curriculum.exists():
        batch_names = [f"{batch.name} {batch.discipline.acronym}" for batch in batches_without_curriculum]
        return JsonResponse({
            'success': False,
            'message': f'The following batches for {academic_year} exist but have no curriculum assigned: {", ".join(batch_names)}. Please assign a curriculum to all batches first.',
            'validation_error': 'batch_missing_curriculum',
            'batches_without_curriculum': batch_names,
            'academic_year': academic_year
        }, status=400)
    
    return None

def calculate_current_semester(academic_year, current_date=None):
    if current_date is None:
        current_date = timezone.now().date()
    
    current_year = current_date.year
    current_month = current_date.month
    
    years_completed = 0
    
    if current_month >= 8:
        years_completed = current_year - academic_year
        semester_in_year = 1
    else:
        years_completed = current_year - academic_year - 1
        if current_month <= 5:
            semester_in_year = 2
        else:
            semester_in_year = 2
    
    total_semester = (years_completed * 2) + semester_in_year
    total_semester = max(1, min(total_semester, 8))
    
    return total_semester

@csrf_exempt
@require_http_methods(["GET"])
def admin_batches_overview(request):
    request.GET = request.GET.copy()
    request.GET['format'] = 'nested'
    return admin_batches_unified(request)

def get_display_branch_name(discipline):
    branch_mappings = {
        'Computer Science and Engineering': 'CSE',
        'Electronics and Communication Engineering': 'ECE',
        'Mechanical Engineering': 'ME',
        'Smart Manufacturing': 'SM',
        'Design': 'DES'
    }
    return branch_mappings.get(discipline, discipline)

@csrf_exempt
@require_http_methods(["POST"])
def process_excel_upload(request):
    try:
        if 'file' not in request.FILES:
            return JsonResponse({
                'success': False,
                'message': 'No file uploaded'
            }, status=400)
        
        file = request.FILES['file']
        programme_type = request.POST.get('programme_type', 'ug')
        
        year_result = validate_and_normalize_year(request.POST.get('academic_year'))
        if isinstance(year_result, JsonResponse):
            return year_result
        batch_year, academic_year = year_result

        curriculum_error = validate_curriculum_exists("uploading student data")
        if curriculum_error:
            return curriculum_error

        batch_error = validate_batch_curriculum_requirements(batch_year, academic_year, "uploading student data")
        if batch_error:
            return batch_error
        
        if not file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({
                'success': False,
                'message': 'Invalid file format. Please upload an Excel file (.xlsx or .xls)'
            }, status=400)
        
        try:
            if file.name.endswith('.xlsx'):
                df = pd.read_excel(file, engine='openpyxl')
            else:
                df = pd.read_excel(file, engine='xlrd')
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error reading Excel file: {str(e)}'
            }, status=400)
        
        valid_students = []
        invalid_students = []
        
        column_mapping = {
            'sno': ['sno', 's.no', 'serial number', 's no'],
            'jee_app_no': ['jee main application number', 'jee app. no./ccmt roll. no.', 'jee app. no./ccmt roll no.', 'jee application no./ccmt roll no.', 'jee application no./ccmt roll. no.', 'jee app. no.', 'jee application number', 'jee main app number', 'ccmt roll. no.', 'ccmt roll no', 'jee app no', 'rollno', 'jee app. no / ccmt roll no.', 'jee app. no / ccmt roll no', 'jee app no / ccmt roll no', 'jee app no / ccmt roll no.', 'jee app. no / ccmt roll no', 'jee app no / ccmt roll no', 'jee app. no./ccmt roll no', 'jee app no./ccmt roll no', 'jee app. no. / ccmt roll no.'],
            'roll_number': ['institute roll number', 'roll number', 'rollno', 'inst roll number'],
            'name': ['name', 'student name', 'full name'],
            'discipline': ['discipline', 'branch', 'department'],
            'specialization': ['specialization', 'spec', 'specialisation'],
            'gender': ['gender', 'sex'],
            'category': ['category', 'caste'],
            'pwd': ['pwd', 'disability', 'pwb'],
            'minority': ['minority', 'minority status'],
            'phone_number': ['mobileno', 'mobile', 'phone', 'mobile no'],
            'institute_email': ['institute email id', 'institute email', 'inst email id'],
            'personal_email': ['alternate email id', 'email', 'personal email'],
            'father_name': ['father\'s name', 'father name', 'fathers name'],
            'father_occupation': ['father\'s occupation', 'father occupation', 'fathers occupation'],
            'father_mobile': ['father mobile number', 'father mobile', 'fathers mobile'],
            'mother_name': ['mother\'s name', 'mother name', 'mothers name'],
            'mother_occupation': ['mother\'s occupation', 'mother occupation', 'mothers occupation'],
            'mother_mobile': ['mother mobile number', 'mother mobile', 'mothers mobile'],
            'date_of_birth': ['date of birth', 'dob', 'birth date'],
            'ai_rank': ['ai rank', 'jee rank', 'all india rank'],
            'category_rank': ['category rank', 'cat rank'],
            'allotted_category': ['allottedcat', 'allotted category', 'alloted category'],
            'allotted_gender': ['allotted gender', 'alloted gender'],
            'state': ['state'],
            'address': ['full address', 'address', 'complete address'],
            'parent_email': ['parent\'s email', 'parent email', 'parents email'],
            'country': ['country'],
            'nationality': ['nationality'],
            'blood_group': ['blood group', 'bloodgroup'],
            'blood_group_remarks': ['blood group remarks', 'bloodgroup remarks'],
            'pwd_category': ['pwd category', 'pwb category', 'disability category'],
            'pwd_category_remarks': ['pwd category remarks', 'pwb category remarks', 'disability remarks'],
            'admission_mode': ['admission mode', 'admission type'],
            'admission_mode_remarks': ['admission mode remarks', 'admission type remarks'],
            'income_group': ['income group', 'family income group'],
            'income': ['income', 'family income', 'annual income']
        }

        df.columns = df.columns.str.lower().str.strip()

        mapped_data = {}
        for field, possible_columns in column_mapping.items():
            for col in possible_columns:
                if col in df.columns:
                    mapped_data[field] = col
                    break
        
        for index, row in df.iterrows():
            try:
                student_data = {}

                for field, excel_col in mapped_data.items():
                    if excel_col in df.columns:
                        value = row[excel_col]
                        if pd.notna(value):
                            if field == 'minority':
                                student_data[field] = str(value).strip()
                            # Fix serial number to start from 1
                            elif field == 'sno':
                                student_data[field] = str(index + 1)  # Force serial number to be 1-based
                            elif field in ['phone_number', 'father_mobile', 'mother_mobile']:
                                student_data[field] = sanitize_phone_number(value)
                            elif field in ['ai_rank', 'category_rank']:
                                student_data[field] = sanitize_rank_value(value)
                            elif field == 'date_of_birth':
                                parsed_date = parse_date_flexible(value)
                                student_data[field] = parsed_date.isoformat() if parsed_date else str(value).strip()
                            else:
                                student_data[field] = str(value).strip()

                required_fields = ['name', 'discipline', 'roll_number', 'institute_email']
                missing_fields = [field for field in required_fields if not student_data.get(field)]

                errors = []
                
                if student_data.get('gender'):
                    gender = student_data['gender'].lower()
                    if gender not in ['male', 'female', 'other']:
                        errors.append(f'Invalid gender: {student_data["gender"]}')

                if student_data.get('pwd'):
                    pwd = student_data['pwd'].upper()
                    if pwd not in ['YES', 'NO']:
                        errors.append(f'Invalid PWD value: {student_data["pwd"]}')


                if student_data.get('institute_email'):
                    email = student_data['institute_email']
                    if '@' not in email:
                        errors.append(f'Invalid institute email format: {email}')
                
                student_phone, father_phone, phone_errors = validate_phone_numbers(student_data)
                errors.extend(phone_errors)
                
                if missing_fields or errors:
                    error_msg = []
                    if missing_fields:
                        error_msg.append(f'Missing required fields: {", ".join(missing_fields)}')
                    if errors:
                        error_msg.extend(errors)
                    
                    invalid_students.append({
                        'row': index + 2,  # Excel row number (1-indexed + header)
                        'error': '; '.join(error_msg),
                        'data': student_data
                    })
                else:

                    cleaned_data = {
                        'Sno': student_data.get('sno', ''),
                        'jee_app_no': student_data.get('jee_app_no', ''),  # Add this for frontend compatibility
                        'JEE App. No./CCMT Roll. No.': student_data.get('jee_app_no', ''),
                        'JEE App. No / CCMT Roll No': student_data.get('jee_app_no', ''),  # Alternative format
                        'Jee Main Application Number': student_data.get('jee_app_no', ''),  # Alternative format
                        'Institute Roll Number': student_data.get('roll_number', ''),
                        'Name': student_data.get('name', ''),
                        'Discipline': student_data.get('discipline', ''),
                        'Specialization': student_data.get('specialization', ''),
                        'Gender': student_data.get('gender', ''),
                        'Category': student_data.get('category', ''),
                        'PWD': student_data.get('pwd', ''),
                        'Minority': student_data.get('minority', ''),
                        'Mobile No': student_data.get('phone_number', ''),
                        'Institute Email ID': student_data.get('institute_email', ''),
                        'Alternate Email ID': student_data.get('personal_email', ''),
                        'Father\'s Name': student_data.get('father_name', ''),
                        'Father\'s Occupation': student_data.get('father_occupation', ''),
                        'Father Mobile Number': student_data.get('father_mobile', ''),
                        'Mother\'s Name': student_data.get('mother_name', ''),
                        'Mother\'s Occupation': student_data.get('mother_occupation', ''),
                        'Mother Mobile Number': student_data.get('mother_mobile', ''),
                        'Date of Birth': student_data.get('date_of_birth', ''),
                        'AI rank': student_data.get('ai_rank', ''),
                        'Category Rank': student_data.get('category_rank', ''),
                        'allottedcat': student_data.get('allotted_category', ''),
                        'Allotted Gender': student_data.get('allotted_gender', ''),
                        'State': student_data.get('state', ''),
                        'Full Address': student_data.get('address', ''),
                        'Parent\'s Email': student_data.get('parent_email', ''),
                        'Country': student_data.get('country', ''),
                        'Nationality': student_data.get('nationality', ''),
                        'Blood Group': student_data.get('blood_group', ''),
                        'Blood Group Remarks': student_data.get('blood_group_remarks', ''),
                        'PwD Category': student_data.get('pwd_category', ''),
                        'PwD Category Remarks': student_data.get('pwd_category_remarks', ''),
                        'Admission Mode': student_data.get('admission_mode', ''),
                        'Admission Mode Remarks': student_data.get('admission_mode_remarks', ''),
                        'Income Group': student_data.get('income_group', ''),
                        'Income': student_data.get('income', '')
                    }         
                    valid_students.append(cleaned_data)
                    
            except Exception as e:
                invalid_students.append({
                    'row': index + 2,
                    'error': f'Error processing row: {str(e)}',
                    'data': {}
                })
        
        return JsonResponse({
            'success': True,
            'valid_students': valid_students,
            'invalid_students': invalid_students,
            'valid_records': len(valid_students),
            'invalid_records': len(invalid_students),
            'total_records': len(df),
            'message': f'Excel file processed successfully. {len(valid_students)} valid records, {len(invalid_students)} invalid records.'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to process Excel file: {str(e)}'
        }, status=500)

# =============================================================================
# STUDENT BATCH OPERATIONS
# =============================================================================

@csrf_exempt
@require_http_methods(["POST"])
def check_student_duplicate(student, duplicate_check_fields):

    field_mapping = {
        'jeeAppNo': 'jee_app_no',
        'rollNumber': 'roll_number',
        'instituteEmail': 'institute_email',
        'fname': 'father_name',
        'personalEmail': 'personal_email',
        'mobile': 'mobile_number'
    }
    
    try:
        for field in duplicate_check_fields:
            backend_field = field_mapping.get(field, field.lower())
            student_value = student.get(field)
            
            if not student_value:
                continue

            if backend_field == 'jee_app_no':
                existing = StudentBatchUpload.objects.filter(jee_app_no=student_value).first()
                if existing:
                    return True, f"JEE Application Number {student_value} already exists for {existing.name}"
                    
            elif backend_field == 'roll_number':
                existing = StudentBatchUpload.objects.filter(roll_number=student_value).first()
                if existing:
                    return True, f"Roll Number {student_value} already exists for {existing.name}"
                    
            elif backend_field == 'institute_email':
                existing = StudentBatchUpload.objects.filter(institute_email=student_value).first()
                if existing:
                    return True, f"Institute Email {student_value} already exists for {existing.name}"
                    
            elif backend_field == 'personal_email':
                existing = StudentBatchUpload.objects.filter(personal_email=student_value).first()
                if existing:
                    return True, f"Personal Email {student_value} already exists for {existing.name}"
                    
            elif backend_field == 'mobile_number':
                existing = StudentBatchUpload.objects.filter(mobile_number=student_value).first()
                if existing:
                    return True, f"Mobile Number {student_value} already exists for {existing.name}"
        
        return False, ""
        
    except Exception as e:
        return False, ""


@csrf_exempt
@require_http_methods(["POST"])
def save_students_batch(request):
    try:
        data = json.loads(request.body)
        students = data.get('students', [])
        programme_type = data.get('programme_type', 'ug')
        year_result = validate_and_normalize_year(data.get('academic_year'))
        if isinstance(year_result, JsonResponse):
            return year_result
        batch_year, academic_year = year_result
        
        curriculum_error = validate_curriculum_exists("saving student data")
        if curriculum_error:
            return curriculum_error

        batch_error = validate_batch_curriculum_requirements(batch_year, academic_year, "saving student data")
        if batch_error:
            return batch_error
        
        skip_duplicates = data.get('skip_duplicates', False)
        duplicate_check_fields = data.get('duplicate_check_fields', ['jeeAppNo', 'rollNumber', 'instituteEmail'])
        
        if not students:
            return JsonResponse({
                'success': False,
                'message': 'No student data provided'
            }, status=400)

        # Initialize counters
        successful_uploads = 0
        failed_uploads = 0
        skipped_duplicates = 0
        validation_errors = 0
        errors = []
        
        # Filter out duplicates if requested
        if skip_duplicates:
            filtered_students = []
            duplicate_students = []
            
            for student in students:
                is_duplicate, duplicate_info = check_student_duplicate(student, duplicate_check_fields)
                
                if is_duplicate:
                    skipped_duplicates += 1
                    duplicate_students.append({
                        'student': student,
                        'duplicate_reason': duplicate_info
                    })
                else:
                    filtered_students.append(student)
            
            students = filtered_students

        # Filter out students with validation errors before processing
        valid_students_only = []
        skipped_invalid = 0
        
        for i, student in enumerate(students):
            student_phone, father_phone, phone_errors = validate_phone_numbers(student)
            if phone_errors:
                skipped_invalid += 1
                errors.append(f"Skipped student {student.get('Name', 'Unknown')}: {'; '.join(phone_errors)}")
                continue
            
            valid_students_only.append(student)

        processed_students = process_batch_allocation(valid_students_only, programme_type, batch_year)
        
        for i, student_data in enumerate(processed_students):
            try:
                with transaction.atomic():
                    student_name = student_data.get('Name') or student_data.get('name', 'Unknown')
                    name = (student_data.get('Name') or student_data.get('name', '')).strip()
                    
                    if not name:
                        validation_errors += 1
                        errors.append(f"Student at row has no name - skipping")
                        continue
                    
                    dob_value = student_data.get('Date of Birth') or student_data.get('dob') or student_data.get('date_of_birth')
                    dob = parse_date_flexible(dob_value)
                    
                    discipline_name = student_data.get('Discipline') or student_data.get('branch', '')
                    specialization = student_data.get('Specialization') or student_data.get('specialization', '')
                    
                    # For M.Tech students, use specialization-specific batch names
                    if programme_type == 'pg' and specialization:
                        if 'design' in discipline_name.lower():
                            batch_name = 'M.Des'
                        elif specialization == 'Mechatronics':
                            batch_name = 'M.Tech'  # Mechatronics batch is just named "M.Tech"
                        else:
                            batch_name = f'M.Tech {specialization}'
                    else:
                        batch_name = get_batch_name_from_discipline(discipline_name, programme_type)

                    if programme_type == 'pg' and specialization:
                        if 'design' in discipline_name.lower():
                            discipline_obj = get_or_create_discipline('Design')
                        elif specialization in ['Data Science', 'AI & ML']:
                            discipline_obj = get_or_create_discipline('Computer Science and Engineering')
                        elif specialization in ['Communication and Signal Processing', 'Nanoelectronics and VLSI Design', 'Power & Control']:
                            discipline_obj = get_or_create_discipline('Electronics and Communication Engineering')
                        elif specialization in ['Design', 'CAD/CAM', 'Manufacturing and Automation']:
                            discipline_obj = get_or_create_discipline('Mechanical Engineering')
                        elif specialization == 'Mechatronics':
                            discipline_obj = get_or_create_discipline('Mechatronics')
                        else:
                            discipline_obj = get_or_create_discipline(discipline_name)
                    else:
                        discipline_obj = get_or_create_discipline(discipline_name)
                    
                    try:
                        batch_obj = Batch.objects.get(
                            name=batch_name, 
                            discipline=discipline_obj, 
                            year=batch_year,
                            running_batch=True
                        )
                    except Batch.DoesNotExist:
                        name_year_matches = Batch.objects.filter(name=batch_name, year=batch_year, running_batch=True)
                        discipline_year_matches = Batch.objects.filter(discipline=discipline_obj, year=batch_year, running_batch=True)
                        existing_batches = Batch.objects.filter(year=batch_year, running_batch=True)
                        
                        error_msg = f"No active batch exists for {batch_name} with discipline '{discipline_obj.name}' in Year-{batch_year}. Please create the batch via Admin Batch Management first."
                        failed_uploads += 1
                        errors.append({
                            'student': student_data.get('Name', 'Unknown'),
                            'roll_number': student_data.get('Institute Roll Number', ''),
                            'error': error_msg,
                            'validation_error': 'missing_batch',
                            'required_batch': f"{batch_name} {discipline_obj.name} {batch_year}",
                            'required_action': f'Create batch for {batch_name} {discipline_obj.name} Year-{batch_year} via Admin Batch Management first'
                        })
                        continue

                    student_name = student_data.get('Name') or student_data.get('name', 'Unknown')
                    
                    student_upload = StudentBatchUpload.objects.create(
                        # Core identification - handle both field name formats
                        jee_app_no=student_data.get('JEE App. No./CCMT Roll. No.') or student_data.get('JEE App. No / CCMT Roll No') or student_data.get('Jee Main Application Number') or student_data.get('jee_app_no') or student_data.get('jeeAppNo') or None,
                        roll_number=student_data.get('Institute Roll Number') or student_data.get('rollNumber', ''),
                        institute_email=student_data.get('Institute Email ID') or student_data.get('instituteEmail', ''),
                        
                        name=student_data.get('Name') or student_data.get('name', ''),
                        father_name=student_data.get("Father's Name") or student_data.get('fname', ''),
                        mother_name=student_data.get("Mother's Name") or student_data.get('mname', ''),
                        gender=student_data.get('Gender') or student_data.get('gender', ''),
                        category=student_data.get('Category') or student_data.get('category', ''),
                        pwd=student_data.get('PWD') or student_data.get('pwd', 'NO'),
                        minority=student_data.get('Minority') or student_data.get('minority', ''),

                        phone_number=sanitize_phone_number(student_data.get('Mobile No') or student_data.get('phoneNumber', '')),
                        personal_email=student_data.get('Alternate Email ID') or student_data.get('email', '') or student_data.get('alternateEmail', '') or student_data.get('personalEmail', '') or student_data.get('personal_email', ''),
                        parent_email=student_data.get('Parent Email') or student_data.get('parentEmail', '') or student_data.get('parent_email', ''),
                        address=student_data.get('Full Address') or student_data.get('address', ''),
                        state=student_data.get('State') or student_data.get('state', ''),
                        country=student_data.get('Country') or student_data.get('country', 'India'),
                        nationality=student_data.get('Nationality') or student_data.get('nationality', 'Indian'),
                        blood_group=student_data.get('Blood Group') or student_data.get('bloodGroup', ''),
                        blood_group_remarks=student_data.get('Blood Group Remarks') or student_data.get('bloodGroupRemarks', ''),
                        pwd_category=student_data.get('PWD Category') or student_data.get('pwdCategory', ''),
                        pwd_category_remarks=student_data.get('PWD Category Remarks') or student_data.get('pwdCategoryRemarks', ''),
                        admission_mode=student_data.get('Admission Mode') or student_data.get('admissionMode', ''),
                        admission_mode_remarks=student_data.get('Admission Mode Remarks') or student_data.get('admissionModeRemarks', ''),
                        income_group=student_data.get('Income Group') or student_data.get('incomeGroup', ''),
                        income=student_data.get('Income') or student_data.get('income', None),

                        branch=student_data.get('Discipline') or student_data.get('branch', ''),
                        specialization=student_data.get('Specialization') or student_data.get('specialization', ''),
                        date_of_birth=dob,
                        ai_rank=_safe_int_conversion(sanitize_rank_value(student_data.get('AI rank') or student_data.get('jeeRank'))),
                        category_rank=_safe_int_conversion(sanitize_rank_value(student_data.get('Category Rank') or student_data.get('categoryRank'))),

                        father_occupation=student_data.get("Father's Occupation") or student_data.get('fatherOccupation', ''),
                        father_mobile=sanitize_phone_number(student_data.get('Father Mobile Number') or student_data.get('fatherMobile', '')),
                        mother_occupation=student_data.get("Mother's Occupation") or student_data.get('motherOccupation', ''),
                        mother_mobile=sanitize_phone_number(student_data.get('Mother Mobile Number') or student_data.get('motherMobile', '')),

                        allotted_category=student_data.get('allottedcat') or student_data.get('allottedCategory', ''),
                        allotted_gender=student_data.get('Allotted Gender') or student_data.get('allottedGender', ''),

                        year=batch_year,
                        academic_year=academic_year,  # Use the normalized academic year
                        programme_type=programme_type,
                        reported_status='NOT_REPORTED',
                        source='excel_upload',  # Track that this came from Excel upload
                        uploaded_by=request.user if request.user.is_authenticated else None
                    )
                    
                    # AUTOMATIC USER ACCOUNT CREATION with HASHED PASSWORD
                    if student_upload.roll_number:
                        try:
                            user_account, plain_password = student_upload.create_user_account()
                            if user_account and plain_password:
                                # Store password for potential emailing (temporarily)
                                student_data['auto_generated_password'] = plain_password
                                student_data['user_created'] = True
                            else:
                                student_data['user_created'] = False
                        except Exception as user_error:
                            student_data['user_created'] = False
                    else:
                        student_data['user_created'] = False

                    if student_data.get('Institute Roll Number'):
                        create_or_update_main_student_record(student_data, batch_obj, batch_year)
                    
                    successful_uploads += 1
                    
            except Exception as e:
                failed_uploads += 1
                error_msg = f"Failed to save student {student_name}: {str(e)}"
                errors.append(error_msg)

        total_processed = successful_uploads + failed_uploads + validation_errors + skipped_invalid
        response_data = {
            'success': True,
            'data': {
                'successful_uploads': successful_uploads,
                'failed_uploads': failed_uploads,
                'skipped_duplicates': skipped_duplicates,
                'validation_errors': validation_errors,
                'skipped_invalid': skipped_invalid,
                'total_processed': total_processed,
                'original_count': len(data.get('students', []))
            },
            'students': processed_students,
            'summary': get_allocation_summary(processed_students, programme_type),
            'errors': errors
        }
        
        messages = []
        if successful_uploads > 0:
            messages.append(f'{successful_uploads} students uploaded successfully')
        if skipped_duplicates > 0:
            messages.append(f'{skipped_duplicates} duplicates skipped')
        if skipped_invalid > 0:
            messages.append(f'{skipped_invalid} students with validation errors skipped')
        if failed_uploads > 0:
            messages.append(f'{failed_uploads} uploads failed')
        if validation_errors > 0:
            messages.append(f'{validation_errors} validation errors')
            
        response_data['message'] = '. '.join(messages) + '.'
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to save students batch: {str(e)}'
        }, status=500)

def process_batch_allocation(students, programme_type, batch_year=None):
    """Process batch allocation algorithm"""
    
    if batch_year is None:
        current_year = datetime.now().year
        batch_year = current_year if datetime.now().month >= 7 else current_year - 1
    
    display_year = batch_year 

    branch_groups = {}
    
    for i, student in enumerate(students):
        branch_field = student.get('Discipline', '') or student.get('branch', '')
        branch_code = get_branch_code(branch_field, programme_type)
        
        if branch_code not in branch_groups:
            branch_groups[branch_code] = []
        
        branch_groups[branch_code].append({
            **student,
            'branch_code': branch_code,
            'display_branch': get_display_branch_name(branch_field)
        })
  
    for branch_code in branch_groups:
        branch_groups[branch_code].sort(key=lambda x: x.get('Name', '') or x.get('name', ''))
    
    # Allocate roll numbers
    processed_students = []
    branch_counters = {}
    
    for branch_code, students_in_branch in branch_groups.items():
        for student in students_in_branch:
            processed_students.append({
                **student,
                # Use provided roll number and email from Excel/manual entry with correct field names
                'roll_number': student.get('Institute Roll Number', '') or student.get('rollNumber', ''),
                'institute_email': student.get('Institute Email ID', '') or student.get('instituteEmail', ''),
                'year': display_year,
                'programme': programme_type.upper(),
                'allocation_date': timezone.now().isoformat(),
                'status': 'ALLOCATED',
                'reported_status': 'NOT_REPORTED'  # Password will be generated when this changes to REPORTED
            })
            
    
    processed_students.sort(key=lambda x: x['roll_number'])
    
    return processed_students

def get_branch_code(branch_name, programme_type):
    branch_lower = branch_name.lower()
    
    if programme_type == 'ug':
        if 'computer' in branch_lower or 'cse' in branch_lower:
            return 'BCS'
        elif 'electronics' in branch_lower or 'ece' in branch_lower:
            return 'BEC'
        elif 'mechanical' in branch_lower or 'me' in branch_lower:
            return 'BME'
        elif 'smart' in branch_lower or 'manufacturing' in branch_lower:
            return 'BSM'
        elif 'design' in branch_lower:
            return 'BDS'
    elif programme_type == 'pg':
        if 'computer' in branch_lower or 'cse' in branch_lower:
            return 'MCS'
        elif 'electronics' in branch_lower or 'ece' in branch_lower:
            return 'MEC'
        elif 'mechanical' in branch_lower or 'me' in branch_lower:
            return 'MME'
        elif 'smart' in branch_lower or 'manufacturing' in branch_lower:
            return 'MSM'
        elif 'design' in branch_lower:
            return 'MDS'
    elif programme_type == 'phd':
        if 'computer' in branch_lower or 'cse' in branch_lower:
            return 'PCS'
        elif 'electronics' in branch_lower or 'ece' in branch_lower:
            return 'PEC'
        elif 'mechanical' in branch_lower or 'me' in branch_lower:
            return 'PME'
        elif 'smart' in branch_lower or 'manufacturing' in branch_lower:
            return 'PSM'
        elif 'design' in branch_lower:
            return 'PDS'
    
    return 'UNK'  # Unknown

def generate_roll_number(branch_code, year, sequence):
    year_suffix = str(year)[-2:]
    sequence_str = str(sequence).zfill(3)
    return f"{year_suffix}{branch_code}{sequence_str}"

def generate_institute_email(roll_number):
    return f"{roll_number.lower()}@iiitdmj.ac.in"

def generate_password():
    """Generate a cryptographically secure password"""
    import secrets
    import string
    length = 12
    charset = string.ascii_letters + string.digits + "!@#$%"
    password = ''.join(secrets.choice(charset) for _ in range(length))
    
    return password

def get_allocation_summary(students, programme_type):
    branch_counts = {}
    for student in students:
        branch = student.get('branch_code', 'Unknown')
        branch_counts[branch] = branch_counts.get(branch, 0) + 1
    
    return {
        'total_students': len(students),
        'branch_counts': branch_counts,
        'programme': programme_type.upper(),
        'allocation_date': timezone.now().isoformat()
    }

# =============================================================================
# SINGLE STUDENT OPERATIONS
# =============================================================================

@csrf_exempt
@require_http_methods(["POST"])
def add_single_student(request):
    try:
        data = json.loads(request.body)
        programme_type = data.get('programme_type', 'ug')
        
        year_result = validate_and_normalize_year(data.get('academic_year'))
        if isinstance(year_result, JsonResponse):
            return year_result
        batch_year, academic_year = year_result
        
        curriculum_error = validate_curriculum_exists("adding student")
        if curriculum_error:
            return curriculum_error

        batch_error = validate_batch_curriculum_requirements(batch_year, academic_year, "adding student")
        if batch_error:
            return batch_error

        field_mapping = {
            'fname': 'father_name',
            'mname': 'mother_name',
            'jeeAppNo': 'jee_app_no',
            'phoneNumber': 'phone_number',
            'email': 'personal_email',
            'dob': 'date_of_birth',
            'aadharNumber': 'aadhar_number',
            'tenthMarks': 'tenth_marks',
            'twelfthMarks': 'twelfth_marks',
            'jeeRank': 'ai_rank',
            'categoryRank': 'category_rank',
            'allottedGender': 'allotted_gender',
            'allottedCategory': 'allotted_category',
            'fatherOccupation': 'father_occupation',
            'fatherMobile': 'father_mobile',
            'motherOccupation': 'mother_occupation',
            'motherMobile': 'mother_mobile',
            'parentEmail': 'parent_email',
            'bloodGroup': 'blood_group',
            'bloodGroupRemarks': 'blood_group_remarks',
            'pwdCategory': 'pwd_category',
            'pwdCategoryRemarks': 'pwd_category_remarks',
            'admissionMode': 'admission_mode',
            'admissionModeRemarks': 'admission_mode_remarks',
            'incomeGroup': 'income_group',
        }

        mapped_data = {}
        for key, value in data.items():
            mapped_key = field_mapping.get(key, key)
            mapped_data[mapped_key] = value

            if key in ['rollNumber', 'instituteEmail']:
                mapped_data[key] = value  # Keep original name too

        data = mapped_data

        required_fields = ['name', 'father_name', 'mother_name', 'branch', 'gender', 'category', 'pwd', 'address']
        missing_fields = [field for field in required_fields if not data.get(field)]
        
        if missing_fields:
            return JsonResponse({
                'success': False,
                'message': f'Missing required fields: {", ".join(missing_fields)}'
            }, status=400)

        processed_students = process_batch_allocation([data], programme_type, batch_year)
        
        if not processed_students:
            return JsonResponse({
                'success': False,
                'message': 'Failed to process student data'
            }, status=400)
        
        student_data = processed_students[0]

        jee_app_no = data.get('jee_app_no')
        if jee_app_no:
            existing_student = StudentBatchUpload.objects.filter(jee_app_no=jee_app_no).first()
            if existing_student:
                return JsonResponse({
                    'success': False,
                    'message': f'Student with JEE Application Number {jee_app_no} already exists (Roll Number: {existing_student.roll_number})'
                }, status=400)

        dob = parse_date_flexible(data.get('date_of_birth'))
        
        # Save to database with ALL Excel-equivalent fields for complete synchronization
        with transaction.atomic():
                student = StudentBatchUpload.objects.create(
                    name=student_data.get('name'),
                    jee_app_no=student_data.get('jee_app_no'),
                    roll_number=student_data.get('roll_number'),
                    institute_email=student_data.get('institute_email'),

                    father_name=student_data.get('father_name'),
                    mother_name=student_data.get('mother_name'),
                    gender=student_data.get('gender'),
                    category=student_data.get('category'),
                    pwd=student_data.get('pwd'),
                    minority=data.get('minority', ''),
                    date_of_birth=dob or data.get('date_of_birth'),

                    phone_number=sanitize_phone_number(data.get('phone_number', '') or data.get('MobileNo', '')),
                    personal_email=data.get('personal_email', '') or data.get('email', '') or data.get('alternateEmail', '') or data.get('Alternate Email ID', ''),
                    parent_email=data.get('parent_email', '') or data.get('parentEmail', ''),
                    address=student_data.get('address'),
                    state=data.get('state', '') or data.get('State', ''),
                    country=data.get('country', 'India'),
                    nationality=data.get('nationality', 'Indian'),
                    blood_group=data.get('blood_group', '') or data.get('bloodGroup', ''),
                    blood_group_remarks=data.get('blood_group_remarks', '') or data.get('bloodGroupRemarks', ''),
                    pwd_category=data.get('pwd_category', '') or data.get('pwdCategory', ''),
                    pwd_category_remarks=data.get('pwd_category_remarks', '') or data.get('pwdCategoryRemarks', ''),
                    admission_mode=data.get('admission_mode', '') or data.get('admissionMode', ''),
                    admission_mode_remarks=data.get('admission_mode_remarks', '') or data.get('admissionModeRemarks', ''),
                    income_group=data.get('income_group', '') or data.get('incomeGroup', ''),
                    income=data.get('income', None),

                    father_occupation=data.get('father_occupation', '') or data.get("Father's Occupation", ''),
                    father_mobile=sanitize_phone_number(data.get('father_mobile', '') or data.get('Father Mobile Number', '')),
                    mother_occupation=data.get('mother_occupation', '') or data.get("Mother's Occupation", ''),
                    mother_mobile=sanitize_phone_number(data.get('mother_mobile', '') or data.get('Mother Mobile Number', '')),

                    branch=student_data.get('branch'),
                    ai_rank=sanitize_rank_value(data.get('ai_rank') or data.get('AI rank')),
                    category_rank=sanitize_rank_value(data.get('category_rank') or data.get('Category Rank')),

                    allotted_category=data.get('allotted_category', '') or data.get('allottedcat', ''),
                    allotted_gender=data.get('allotted_gender', '') or data.get('allottedGender', ''),

                    year=batch_year, 
                    programme_type=programme_type,
                    reported_status='NOT_REPORTED',
                    academic_year=academic_year,
                    allocation_status='ALLOCATED',
                    source='manual_entry'
                )
        
        return JsonResponse({
            'success': True,
            'data': {
                'student_id': student.id,
                'roll_number': student.roll_number or student_data.get('roll_number'),
                'institute_email': student.institute_email or student_data.get('institute_email'),
                'name': student.name,
                'personal_email': student.personal_email,
                'parent_email': student.parent_email,
                'blood_group': student.blood_group,
                'country': student.country,
                'nationality': student.nationality
            },
            'message': f'Student {student.name} added successfully. Use "Update Status" to transfer to main academic system when ready.'
        })
        
    except json.JSONDecodeError as e:
        return JsonResponse({
            'success': False,
            'message': f'Invalid JSON data: {str(e)}'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to add student: {str(e)}'
        }, status=500)

# =============================================================================
# SEAT MANAGEMENT
# =============================================================================

@csrf_exempt
@require_http_methods(["PUT"])
def set_total_seats(request):
    try:
        data = json.loads(request.body)
        
        programme = data.get('programme')
        discipline = data.get('discipline')
        year = data.get('year')
        total_seats = data.get('total_seats')
        
        if not all([programme, discipline, year, total_seats]):
            return JsonResponse({
                'success': False,
                'message': 'Missing required fields: programme, discipline, year, total_seats'
            }, status=400)

        batch_config, created = BatchConfiguration.objects.update_or_create(
            programme=programme,
            discipline=discipline,
            year=year,
            defaults={'total_seats': total_seats}
        )

        batch_config.calculate_seats()
        
        return JsonResponse({
            'success': True,
            'message': f'Total seats updated to {total_seats} for {programme} - {discipline}',
            'data': {
                'programme': programme,
                'discipline': discipline,
                'year': year,
                'total_seats': total_seats,
                'filled_seats': batch_config.filled_seats,
                'available_seats': batch_config.available_seats
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to update total seats: {str(e)}'
        }, status=500)

# =============================================================================
# STUDENT STATUS MANAGEMENT
# =============================================================================

@csrf_exempt
@require_http_methods(["PUT", "POST", "OPTIONS"])
def update_student_status(request):
    """
    Update student reported status
    """
    
    if request.method == 'OPTIONS':
        response = JsonResponse({})
        response['Access-Control-Allow-Origin'] = '*'
        response['Access-Control-Allow-Methods'] = 'PUT, POST, OPTIONS'
        response['Access-Control-Allow-Headers'] = 'Content-Type, X-CSRFToken, Authorization'
        return response
    
    try:
        data = json.loads(request.body)
        
        student_id = data.get('studentId')
        reported_status = data.get('reportedStatus')
        
        if not student_id or not reported_status:
            return JsonResponse({
                'success': False,
                'message': 'Missing required fields: studentId, reportedStatus'
            }, status=400)
        
        # Validate reported_status
        if reported_status not in ['REPORTED', 'NOT_REPORTED', 'WITHDRAWAL', 'PENDING']:
            return JsonResponse({
                'success': False,
                'message': 'Invalid reportedStatus. Must be REPORTED, NOT_REPORTED, or PENDING'
            }, status=400)
        
        try:
            student = StudentBatchUpload.objects.get(id=student_id)
        except StudentBatchUpload.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Student not found'
            }, status=404)
        
        old_status = student.reported_status
        student.reported_status = reported_status
        student.save()
        
        # Auto-transfer to main tables when status becomes REPORTED
        # OR remove from main tables when status becomes NOT_REPORTED
        transfer_success = False
        transfer_message = ""
        user_created = False
        auto_generated_password = None
        
        if reported_status == 'REPORTED' and old_status != 'REPORTED':
            try:
                # Simple transfer to main tables
                from applications.globals.models import ExtraInfo, DepartmentInfo
                from applications.academic_information.models import Student as AcademicStudent
                from applications.programme_curriculum.models import Batch, Curriculum
                from django.db import transaction
                
                with transaction.atomic():
                    
                    if not student.user:
                        try:
                            user, password = student.create_user_account()
                            auto_generated_password = password
                            user_created = True
                        except Exception as user_error:
                            user_created = False
                    else:
                        if not student.email_password:
                            auto_generated_password = student.generate_secure_password()
                            student.email_password = auto_generated_password
                            student.password_generated_at = timezone.now()
                            student.password_email_sent = False
                            student.save()
                        else:
                            auto_generated_password = student.email_password
                    
                    # Create ExtraInfo if not exists
                    if student.roll_number:
                        from applications.globals.models import DepartmentInfo
                        branch_field = student.branch or ''
                        branch_upper = branch_field.upper()

                        if 'COMPUTER SCIENCE' in branch_upper or 'CSE' in branch_upper:
                            dept_name = 'CSE'
                            discipline_name = 'Computer Science and Engineering'
                        elif 'ELECTRONICS' in branch_upper or 'ECE' in branch_upper:
                            dept_name = 'ECE'
                            discipline_name = 'Electronics and Communication Engineering'
                        elif 'MECHANICAL' in branch_upper or 'ME' in branch_upper:
                            dept_name = 'ME'
                            discipline_name = 'Mechanical Engineering'
                        elif 'SMART MANUFACTURING' in branch_upper or 'SMART' in branch_upper or 'SM' in branch_upper:
                            dept_name = 'SM'
                            discipline_name = 'Smart Manufacturing'
                        elif 'DESIGN' in branch_upper or 'DES' in branch_upper or 'B.DES' in branch_upper:
                            dept_name = 'Design'  # Exact database name
                            discipline_name = 'Design'
                            discipline_acronym = 'Des.'  # Use existing acronym
                        else:
                            # Default fallback
                            dept_name = 'CSE'
                            discipline_name = 'Computer Science and Engineering'

                        try:
                            department = DepartmentInfo.objects.get(name=dept_name)
                        except DepartmentInfo.DoesNotExist:
                            department = DepartmentInfo.objects.create(name=dept_name)
                        
                        extra_info, created = ExtraInfo.objects.get_or_create(
                            id=student.roll_number,
                            defaults={
                                'user': student.user,
                                'title': 'Mr.' if student.gender == 'Male' else ('Ms.' if student.gender == 'Female' else 'Mr.'),
                                'sex': 'M' if student.gender == 'Male' else ('F' if student.gender == 'Female' else 'O'),
                                'date_of_birth': student.date_of_birth or timezone.now().date(),
                                'address': student.address or '',
                                'phone_no': int(sanitize_phone_number(student.phone_number)) if student.phone_number and sanitize_phone_number(student.phone_number).isdigit() else 9999999999,
                                'user_type': 'student',  # CRITICAL: Must be 'student' for proper access
                                'department': department
                            }
                        )
                        
                        # IMPORTANT: Update ALL fields even if ExtraInfo already existed
                        if not created:  # If record already existed, update ALL relevant fields
                            extra_info.department = department
                            extra_info.user = student.user  # Also ensure user is set
                            extra_info.title = 'Mr.' if student.gender == 'Male' else ('Ms.' if student.gender == 'Female' else 'Mr.')
                            extra_info.sex = 'M' if student.gender == 'Male' else ('F' if student.gender == 'Female' else 'O')
                            extra_info.date_of_birth = student.date_of_birth or extra_info.date_of_birth
                            extra_info.address = student.address or extra_info.address
                            sanitized_phone = sanitize_phone_number(student.phone_number)
                            if sanitized_phone and sanitized_phone.isdigit():
                                extra_info.phone_no = int(sanitized_phone)
                            extra_info.user_type = 'student'  # CRITICAL: Ensure user_type is always 'student'
                            extra_info.save()

                        # FETCH from existing main tables instead of creating new ones
                        from applications.programme_curriculum.models import Batch, Discipline, Curriculum, Programme
                        from applications.globals.models import DepartmentInfo
                        
                        # FIND existing discipline based on department mapping
                        discipline = None
                        try:
                            # For PG students with specialization, use the same mapping as upload logic
                            if student.programme_type == 'pg' and student.specialization:
                                if student.specialization in ['Data Science', 'AI & ML']:
                                    discipline_name = 'Computer Science and Engineering'
                                elif student.specialization in ['Communication and Signal Processing', 'Nanoelectronics and VLSI Design', 'Power & Control']:
                                    discipline_name = 'Electronics and Communication Engineering'
                                elif student.specialization == 'Design':
                                    discipline_name = 'Design'
                                elif student.specialization in ['CAD/CAM', 'Manufacturing and Automation']:
                                    discipline_name = 'Mechanical Engineering'
                                elif student.specialization == 'Mechatronics':
                                    discipline_name = 'Mechatronics'
                                else:
                                    branch_field = student.branch or ''
                                    branch_upper = branch_field.upper()
                                    if 'COMPUTER SCIENCE' in branch_upper or 'CSE' in branch_upper:
                                        discipline_name = 'Computer Science and Engineering'
                                    elif 'ELECTRONICS' in branch_upper or 'ECE' in branch_upper:
                                        discipline_name = 'Electronics and Communication Engineering'
                                    elif 'MECHANICAL' in branch_upper or 'ME' in branch_upper:
                                        discipline_name = 'Mechanical Engineering'
                                    elif 'DESIGN' in branch_upper or 'DES' in branch_upper:
                                        discipline_name = 'Design'
                                    else:
                                        discipline_name = 'Computer Science and Engineering'

                                discipline = Discipline.objects.filter(name__exact=discipline_name).first()
                                if not discipline:
                                    discipline = Discipline.objects.filter(name__icontains=discipline_name.split()[0]).first()
                            else:
                                branch_field = student.branch or ''
                                branch_upper = branch_field.upper()

                                if 'COMPUTER SCIENCE' in branch_upper or 'CSE' in branch_upper:
                                    dept_name = 'CSE'
                                    discipline_name = 'Computer Science and Engineering'
                                elif 'ELECTRONICS' in branch_upper or 'ECE' in branch_upper:
                                    dept_name = 'ECE'
                                    discipline_name = 'Electronics and Communication Engineering'
                                elif 'MECHANICAL' in branch_upper or 'ME' in branch_upper:
                                    dept_name = 'ME'
                                    discipline_name = 'Mechanical Engineering'
                                elif 'SMART MANUFACTURING' in branch_upper or 'SMART' in branch_upper or 'SM' in branch_upper:
                                    dept_name = 'SM'
                                    discipline_name = 'Smart Manufacturing'
                                elif 'DESIGN' in branch_upper or 'DES' in branch_upper or 'B.DES' in branch_upper:
                                    dept_name = 'Design'  # Exact database name
                                    discipline_name = 'Design'
                                    discipline_acronym = 'Des.'  # Use existing acronym
                                else:
                                    dept_name = 'CSE'
                                    discipline_name = 'Computer Science and Engineering'
                                
                                if dept_name == 'Design':
                                    discipline = Discipline.objects.filter(
                                        acronym='Des.',
                                        name='Design'
                                    ).first()
                                    
                                    if not discipline:
                                        discipline = Discipline.objects.filter(name__icontains='design').first()
                                else:
                                    discipline = Discipline.objects.filter(
                                        acronym=dept_name,
                                        name__exact=discipline_name
                                    ).first()
                                
                                if not discipline:
                                    disciplines = Discipline.objects.filter(acronym=dept_name).order_by('name')
                                    discipline = disciplines.first()
                                    
                        except Exception as e:
                            pass

                        if not discipline:
                            if student.programme_type == 'pg' and student.specialization:
                                discipline = Discipline.objects.filter(name__icontains='Computer Science').first()  # Default fallback
                            else:
                                discipline = Discipline.objects.filter(acronym='CSE').first()  # Default fallback
                        
                        # FIND existing programme based on correct naming pattern
                        if student.programme_type == 'ug':
                            if 'design' in student.branch.lower():
                                programme_name = 'B.Des'  # Match the actual batch name in database
                            else:
                                programme_name = 'B.Tech'
                            programme_category = 'UG'
                        elif student.programme_type == 'pg':
                            if 'design' in student.branch.lower():
                                programme_name = 'M.Des'  # Match the actual batch name for Design PG
                            else:
                                programme_name = 'M.Tech'
                            programme_category = 'PG'
                        else:
                            programme_name = 'Ph.D'
                            programme_category = 'PhD'
                            
                        # FIND existing programme (look for existing ones first)
                        programme = None
                        try:
                            if dept_name == 'Design':
                                programme = Programme.objects.filter(
                                    name=programme_name,  # Use the correct programme_name (B.Des for UG, M.Des for PG)
                                    category=programme_category,
                                    discipline=discipline
                                ).first()
                            else:
                                programme_search_name = f"{programme_name} {dept_name}"
                                programme = Programme.objects.filter(
                                    name__icontains=programme_search_name,
                                    category=programme_category
                                ).first()
                                
                                if not programme:
                                    programme = Programme.objects.filter(
                                        name__icontains=dept_name,
                                        category=programme_category
                                    ).first()
                            
                            if not programme:
                                programme = Programme.objects.get_or_create(
                                    name=programme_name,
                                    category=programme_category,
                                    defaults={'discipline': discipline}
                                )[0]
                        except Exception as e:
                            programme = Programme.objects.get_or_create(
                                name=programme_name,
                                category=programme_category,
                                defaults={'discipline': discipline}
                            )[0]

                        from applications.programme_curriculum.models import Curriculum

                        discipline_from_excel = student.branch or ''
                        
                        # For PG students, use the actual specialization field; for UG students, map from discipline
                        if student.programme_type == 'pg' and student.specialization:
                            specialization = student.specialization[:35]  # Limit to 40 chars as per model field limit
                        else:
                            if 'Computer Science' in discipline_from_excel or 'CSE' in discipline_from_excel:
                                specialization = 'Computer Science and Engineering'
                            elif 'Electronics' in discipline_from_excel or 'ECE' in discipline_from_excel:
                                specialization = 'Electronics and Communication'
                            elif 'Mechanical' in discipline_from_excel or 'ME' in discipline_from_excel:
                                specialization = 'Mechanical Engineering'
                            elif 'Smart Manufacturing' in discipline_from_excel or 'SM' in discipline_from_excel:
                                specialization = 'Smart Manufacturing'
                            elif 'Design' in discipline_from_excel:
                                specialization = 'Design'
                            else:
                                specialization = discipline_from_excel[:35]  # Fallback

                        transfer_message_addition = ""

                        curriculum_assigned = False
                        if student.programme_type == 'pg':
                            curriculum_id = get_curriculum_by_specialization(student.specialization, discipline)
                            
                            if curriculum_id:
                                try:
                                    from applications.programme_curriculum.models import Curriculum
                                    student_specific_curriculum = Curriculum.objects.get(id=curriculum_id, working_curriculum=True)
                                    curriculum_assigned = True
                                    specialization_display = student.specialization if student.specialization else 'No specialization'
                                    transfer_message_addition += f" | Curriculum: {student_specific_curriculum.name} (Specialization: {specialization_display})"
                                except Curriculum.DoesNotExist:
                                    transfer_message_addition += f" | WARNING: Curriculum with ID {curriculum_id} not found"
                            else:
                                specialization_display = student.specialization if student.specialization else 'empty/blank'
                                transfer_message_addition += f" | No curriculum mapping found for specialization '{specialization_display}'"
                        else:
                            student_specific_curriculum = None
                        

                        try:
                            batch_obj = None
                            batch_created = False
                            
                            # For PG students with specialization, only use existing batches
                            if programme_category == 'PG' and student.specialization and student_specific_curriculum:
                                if student.specialization == 'Design':
                                    batch_obj = Batch.objects.filter(
                                        name=programme_name,  # M.Des
                                        year=student.year,
                                        discipline=discipline,
                                        running_batch=True
                                    ).first()
                                elif student.specialization == 'Mechatronics':
                                    # Look for existing M.Tech batch for Mechatronics
                                    batch_obj = Batch.objects.filter(
                                        name='M.Tech',
                                        year=student.year,
                                        discipline=discipline,
                                        running_batch=True
                                    ).first()
                                else:
                                    specialization_batch_name = f"{programme_name} {student.specialization}"
                                    batch_obj = Batch.objects.filter(
                                        name=specialization_batch_name,
                                        year=student.year,
                                        discipline=discipline,
                                        running_batch=True
                                    ).first()

                                    if not batch_obj:
                                        batch_obj = Batch.objects.filter(
                                            name=programme_name,
                                            year=student.year,
                                            discipline=discipline,
                                            running_batch=True
                                        ).first()

                                if not batch_obj:
                                    return JsonResponse({
                                        'success': False,
                                        'message': f'No batch found for {programme_name} {discipline.name} Year-{student.year} with specialization {student.specialization}. Please create the required batch manually first.',
                                        'error_code': 'BATCH_NOT_FOUND',
                                        'required_batch': f'{programme_name} {student.specialization}' if student.specialization != 'Design' else programme_name,
                                        'discipline': discipline.name,
                                        'year': student.year
                                    }, status=400)

                            if not batch_obj:
                                batch_obj = Batch.objects.filter(
                                    name=programme_name,    
                                    year=student.year,   
                                    discipline=discipline,  
                                    running_batch=True
                                ).first()
                            
                            if not batch_obj and programme_category == 'PG':
                                batch_obj = Batch.objects.filter(
                                    year=student.year,
                                    discipline=discipline,
                                    running_batch=True
                                ).first()
                                
                            if not batch_obj and programme_category == 'PG':
                                batch_obj = Batch.objects.filter(
                                    year=student.year,
                                    running_batch=True
                                ).first()

                            if not batch_obj and programme_category == 'PG':
                                batch_obj = Batch.objects.filter(
                                    discipline=discipline,
                                    running_batch=True
                                ).first()

                            if not batch_obj and programme_category == 'PG':
                                batch_obj = Batch.objects.filter(
                                    running_batch=True
                                ).first()
                            elif not batch_obj and programme_category == 'UG':
                                batch_obj = Batch.objects.filter(
                                    year=student.year,
                                    discipline=discipline,
                                    running_batch=True
                                ).first()
                                
                                if not batch_obj:
                                    batch_obj = Batch.objects.filter(
                                        year=student.year,
                                        running_batch=True
                                    ).first()
                                
                        except Exception as e:
                            if programme_category == 'PG':
                                batch_obj = Batch.objects.filter(
                                    running_batch=True
                                ).first()
                            elif programme_category == 'UG':
                                batch_obj = Batch.objects.filter(
                                    running_batch=True
                                ).first()
                            else:
                                batch_obj = Batch.objects.filter(running_batch=True).first()
                            transfer_message_addition = f" | Warning: Batch matching error: {str(e)}, using fallback batch"

                        transfer_message_addition = transfer_message_addition if 'transfer_message_addition' in locals() else ""

                        current_semester = calculate_current_semester(int(student.year))

                        # USE EXISTING BATCH (NO AUTOMATIC BATCH CREATION)
                        final_batch = batch_obj
                        
                        final_batch = batch_obj

                        if batch_created:
                            transfer_message_addition += f" | Created new batch: {final_batch.name} with curriculum: {final_batch.curriculum.name}"
                        elif final_batch.curriculum:
                            transfer_message_addition += f" | Using batch: {final_batch.name} with curriculum: {final_batch.curriculum.name}"
                        else:
                            transfer_message_addition += f" | Using batch: {final_batch.name} (no curriculum assigned to batch)"

                        academic_student, created = AcademicStudent.objects.get_or_create(
                            id=extra_info, 
                            defaults={
                                'batch_id': final_batch,
                                'specialization': specialization,
                                'programme': programme_name,
                                'batch': student.year,  
                                'father_name': student.father_name or '',  
                                'mother_name': student.mother_name or '', 
                                'category': student.category or '',  
                                'cpi': 0.0,  
                                'curr_semester_no': current_semester,  
                                'hall_no': 0,  
                                'room_no': '',
                            }
                        )
                        
                        # IMPORTANT: Update ALL fields even if Student already existed
                        if not created:                         
                            academic_student.specialization = specialization
                            academic_student.batch_id = final_batch
                            academic_student.programme = programme_name 
                            academic_student.batch = student.year 
                            academic_student.father_name = student.father_name or ''  
                            academic_student.mother_name = student.mother_name or '' 
                            academic_student.category = student.category or ''  
                            academic_student.curr_semester_no = current_semester 
                            from applications.academic_information.models import Constants
                            valid_choices = [choice[0] for choice in Constants.MTechSpecialization]
                            
                            try:
                                academic_student.save()
                            except Exception as save_error:
                                pass
                        else:
                            pass

                        academic_student.refresh_from_db()
                            
                        try:
                            pass
                        except Exception as e:
                            pass
                        
                        # CREATE STUDENT DESIGNATION
                        try:
                            if student.user:
                                student_designation, created = Designation.objects.get_or_create(
                                    name='student',
                                    defaults={
                                        'full_name': 'Student',
                                        'type': 'Academic'
                                    }
                                )

                                existing_designation = HoldsDesignation.objects.filter(
                                    user=student.user,
                                    designation=student_designation
                                ).first()
                                
                                if not existing_designation:
                                    holds_designation = HoldsDesignation.objects.create(
                                        user=student.user,
                                        working=student.user,
                                        designation=student_designation,
                                        held_at=timezone.now()
                                    )
                                    transfer_message_addition = f" | Student designation assigned"
                                else:
                                    transfer_message_addition = f" | Student designation already exists"
                                    
                        except Exception as designation_error:
                            transfer_message_addition = f" | Designation error: {str(designation_error)}"
                        
                        # COURSE REGISTRATION - MANUAL PROCESS

                        course_registration_message = " | Course registration: Manual process (to be done by academic admin)"

                        transfer_message_addition += course_registration_message
                        
                        # COMPREHENSIVE VALIDATION - Check for missing critical components
                        validation_warnings = []
                        try:
                            if hasattr(extra_info, 'student') and extra_info.student:
                                pass
                            else:
                                validation_warnings.append("Missing ExtraInfo.student relationship")

                            if hasattr(student.user, 'extrainfo') and student.user.extrainfo:
                                pass
                            else:
                                validation_warnings.append("Missing User.extrainfo relationship")

                            if academic_student.batch_id and academic_student.batch_id.curriculum:
                                pass
                            else:
                                validation_warnings.append("Missing batch curriculum assignment")

                            if academic_student.curr_semester_no >= 1:
                                pass
                            else:
                                validation_warnings.append("Invalid semester number")

                            designations = student.user.holds_designations.all()
                            if designations.filter(designation__name='student').exists():
                                pass
                            else:
                                validation_warnings.append("Missing student designation access")

                            if validation_warnings:
                                transfer_message_addition += f" | Validation warnings: {'; '.join(validation_warnings)}"
                                
                        except Exception as validation_error:
                            transfer_message_addition += f" | Validation check failed: {str(validation_error)}"
                        
                        transfer_success = True

                        curriculum_status = ""
                        if batch_obj and batch_obj.curriculum:
                            curriculum_status = f" | Batch: {batch_obj.name} {batch_obj.discipline.acronym} {batch_obj.year} | Curriculum: {batch_obj.curriculum.name}"
                        elif batch_obj:
                            curriculum_status = f" | Batch: {batch_obj.name} {batch_obj.discipline.acronym} {batch_obj.year} | NO CURRICULUM ASSIGNED - Please assign via Batch Form!"
                        else:
                            curriculum_status = f" | NO BATCH FOUND - Critical error!"
                        
                        transfer_message = f"Student successfully transferred to main academic tables. Roll: {student.roll_number}{curriculum_status}{transfer_message_addition}"
                        
                        # Send welcome email with login credentials using existing email system
                        email_password = auto_generated_password or student.email_password
                        if email_password and not student.password_email_sent:
                            try:

                                email_sent, email_message = send_password_email_smtp(
                                    student_email=(student.institute_email or student.personal_email or '').lower(), 
                                    student_name=student.name,
                                    password=email_password, 
                                    roll_number=student.roll_number,
                                    student=student
                                )
                                
                                if email_sent:
                                    student.password_email_sent = True  
                                    student.save()
                                    transfer_message += f" | Email sent to {student.institute_email or student.personal_email}"
                                else:
                                    transfer_message += f" | Email failed: {email_message}"
                                    
                            except Exception as email_error:
                                transfer_message += f" | Email error: {str(email_error)}"             
                    else:
                        transfer_message = "Transfer skipped: No roll number assigned"
                        
            except Exception as e:
                transfer_message = f"Status updated but transfer failed: {str(e)}"
        
        # REVERT: Remove from main tables when status becomes NOT_REPORTED
        elif reported_status in ['NOT_REPORTED', 'PENDING'] and old_status == 'REPORTED':
            try:
                from applications.globals.models import ExtraInfo
                from applications.academic_information.models import Student as AcademicStudent
                from django.db import transaction
                
                with transaction.atomic():
                    if student.roll_number:
                        try:
                            extra_info = ExtraInfo.objects.get(id=student.roll_number)
                            academic_student = AcademicStudent.objects.get(id=extra_info)

                            batch_to_check = academic_student.batch_id
                            
                            academic_student.delete()
                            transfer_message += f"Removed AcademicStudent record for {student.roll_number}. "

                            if batch_to_check:
                                remaining_students = AcademicStudent.objects.filter(batch_id=batch_to_check).count()
                                if remaining_students == 0:
                                    transfer_message += f"Batch {batch_to_check.name} now empty. "
                                    
                        except AcademicStudent.DoesNotExist:
                            transfer_message += f"No AcademicStudent record found for {student.roll_number}. "
                        except ExtraInfo.DoesNotExist:
                            transfer_message += f"No ExtraInfo record found for {student.roll_number}. "

                        try:
                            extra_info = ExtraInfo.objects.get(id=student.roll_number)
                            extra_info.delete()
                            transfer_message += f"Removed ExtraInfo record for {student.roll_number}. "
                        except ExtraInfo.DoesNotExist:
                            transfer_message += f"No ExtraInfo record found for {student.roll_number}. "

                        if student.user:
                            try:
                                username = student.roll_number if student.roll_number else f"temp_{student.jee_app_no}"
                                if student.user.username == username:
                                    user_to_delete = student.user
                                    student.user = None  # Clear the reference first
                                    student.save()
                                    user_to_delete.delete()
                                    transfer_message += f"Removed User account for {username}. "
                                else:
                                    student.user = None
                                    student.save()
                                    transfer_message += f"Cleared User reference for {student.roll_number}. "
                            except Exception as e:
                                transfer_message += f"Error removing User: {str(e)}. "
                        
                        try:
                            if student.user:
                                student_designation = Designation.objects.get(name='student')
                                holds_designation = HoldsDesignation.objects.filter(
                                    user=student.user,
                                    designation=student_designation
                                ).first()
                                
                                if holds_designation:
                                    holds_designation.delete()
                                    transfer_message += f"Removed student designation for {student.roll_number}. "
                                else:
                                    transfer_message += f"No student designation found for {student.roll_number}. "
                                    
                        except Designation.DoesNotExist:
                            transfer_message += f"Student designation type does not exist. "
                        except Exception as designation_error:
                            transfer_message += f"Error removing designation: {str(designation_error)}. "

                        try:  
                            pass 

                        except Exception as e:
                            transfer_message += f"Additional cleanup warning: {str(e)}. "
                        
                        transfer_success = True
                        if not transfer_message:
                            transfer_message = f"Student {student.roll_number} successfully removed from main academic tables"
                    else:
                        transfer_message = "Revert skipped: No roll number assigned"
                        
            except Exception as e:
                transfer_message = f"Status updated but revert failed: {str(e)}"
        
        # Log the status change
        try:
            from django.contrib.auth.models import User
            user = None
            if hasattr(request, 'user') and request.user.is_authenticated:
                user = request.user
            
        except Exception:
            pass
        
        # Prepare descriptive response message
        if reported_status == 'REPORTED' and old_status != 'REPORTED':
            main_message = f'Student status updated to {reported_status} and transferred to main academic system'
        elif reported_status in ['NOT_REPORTED', 'PENDING'] and old_status == 'REPORTED':
            main_message = f'Student status reverted to {reported_status} and removed from main academic system'
        else:
            main_message = f'Student status updated to {reported_status}'
        
        response_data = {
            'success': True,
            'message': main_message,
            'data': {
                'student_id': student.id,
                'roll_number': student.roll_number,
                'old_status': old_status,
                'new_status': reported_status,
                'transfer_success': transfer_success,
                'transfer_message': transfer_message,
                'user_created': user_created,
                'password_generated': auto_generated_password if auto_generated_password else None,
                'email_address': student.institute_email or student.personal_email,
                'username': student.roll_number if user_created else None
            }
        }
        
        # Update batch seat calculations after status change
        try:
            from applications.programme_curriculum.models_student_management import BatchConfiguration
            from django.db.models import Q

            discipline_q = Q()
            if student.branch:
                branch_upper = student.branch.upper()
                if 'COMPUTER SCIENCE' in branch_upper or 'CSE' in branch_upper:
                    discipline_q = Q(discipline='Computer Science and Engineering')
                elif 'ELECTRONICS' in branch_upper or 'ECE' in branch_upper:
                    discipline_q = Q(discipline='Electronics and Communication Engineering')
                elif 'DESIGN' in branch_upper:
                    discipline_q = Q(discipline='Design')
                elif 'MECHANICAL' in branch_upper:
                    discipline_q = Q(discipline='Mechanical Engineering')
                elif 'SMART' in branch_upper:
                    discipline_q = Q(discipline='Smart Manufacturing')
            
            # Update seat calculations for matching batches
            if discipline_q:
                matching_batches = BatchConfiguration.objects.filter(
                    discipline_q,
                    year=student.year
                )
                for batch in matching_batches:
                    batch.calculate_seats()
        except Exception as e:
            pass
        
        response = JsonResponse(response_data, status=200)

        response['Access-Control-Allow-Origin'] = '*'
        return response
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to update student status: {str(e)}'
        }, status=500)

# =============================================================================
# EXPORT FUNCTIONALITY
# =============================================================================

@csrf_exempt
@require_http_methods(["GET"])
def export_students(request, programme_type):
    """
    Export student data to Excel
    """
    try:
        students = StudentBatchUpload.objects.filter(programme_type=programme_type).order_by('roll_number')
        
        if not students.exists():
            return JsonResponse({
                'success': False,
                'message': f'No students found for programme type: {programme_type}'
            }, status=404)

        output = BytesIO()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = f'{programme_type.upper()} Students'

        headers = [
            'S.No', 'JEE App. No./CCMT Roll. No.', 'Institute Roll Number', 'Name',
            'Discipline', 'Specialization', 'Gender', 'Category', 'PWD', 'Minority', 'Mobile No',
            'Institute Email ID', 'Alternate Email ID', 'Father\'s Name',
            'Father\'s Occupation', 'Father Mobile Number', 'Mother\'s Name',
            'Mother\'s Occupation', 'Mother Mobile Number', 'Date of Birth',
            'AI Rank', 'Category Rank', 'Allotted Category', 'Allotted Gender',
            'State', 'Full Address', 'Reported Status'
        ]
        
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

        for row, student in enumerate(students, 2):
            data = [
                row - 1,  # S.No
                student.jee_app_no,
                student.roll_number,
                student.name,
                student.branch,
                getattr(student, 'specialization', ''),
                student.gender,
                student.category,
                student.pwd,
                getattr(student, 'minority', ''),
                student.phone_number,
                student.institute_email,
                student.personal_email,
                student.father_name,
                getattr(student, 'father_occupation', ''),
                getattr(student, 'father_mobile', ''),
                student.mother_name,
                getattr(student, 'mother_occupation', ''),
                getattr(student, 'mother_mobile', ''),
                student.date_of_birth.strftime('%d/%m/%Y') if student.date_of_birth else '',
                student.ai_rank,
                student.category_rank,
                getattr(student, 'allotted_category', ''),
                getattr(student, 'allotted_gender', ''),
                getattr(student, 'state', ''),
                student.address,
                student.reported_status
            ]
            
            for col, value in enumerate(data, 1):
                worksheet.cell(row=row, column=col, value=value)
        
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{programme_type}_students_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to export student data: {str(e)}'
        }, status=500)

# =============================================================================
# UPLOAD HISTORY
# =============================================================================

@csrf_exempt
@require_http_methods(["GET"])
def upload_history(request):
    """
    Get upload history
    """
    try:
        from django.db.models import Count
        
        history = StudentBatchUpload.objects.values('created_at__date', 'programme_type').annotate(
            count=Count('id')
        ).order_by('-created_at__date')
        
        return JsonResponse({
            'success': True,
            'data': list(history),
            'message': 'Upload history fetched successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to fetch upload history: {str(e)}'
        }, status=500)

# =============================================================================
# STUDENT LISTING
# =============================================================================

@csrf_exempt
@require_http_methods(["GET"])
def list_students(request):
    """
    List students with filters
    """
    try:
        programme_type = request.GET.get('programme_type')
        batch_id = request.GET.get('batch_id')
        year = request.GET.get('year')
        discipline = request.GET.get('discipline')
        specialization = request.GET.get('specialization')
        
        students = StudentBatchUpload.objects.all()
        
        if programme_type:
            students = students.filter(programme_type=programme_type)
        
        if year:
            students = students.filter(year=year)
        
        # For PG students, filter by specialization first, then discipline
        if programme_type == 'pg':
            if specialization:
                students = students.filter(specialization__icontains=specialization)
            elif discipline:
                # Handle discipline filtering with typo tolerance
                discipline_filters = Q()
                discipline_filters |= Q(branch__icontains=discipline)

                if 'Engineering' in discipline:
                    typo_name = discipline.replace('Engineering', 'Enginnering')
                    discipline_filters |= Q(branch__icontains=typo_name)
                elif 'Enginnering' in discipline:
                    correct_name = discipline.replace('Enginnering', 'Engineering')
                    discipline_filters |= Q(branch__icontains=correct_name)
                
                students = students.filter(discipline_filters)
        elif discipline:  # For UG or general discipline filter
            discipline_filters = Q()
            discipline_filters |= Q(branch__icontains=discipline)

            if 'Engineering' in discipline:
                typo_name = discipline.replace('Engineering', 'Enginnering')
                discipline_filters |= Q(branch__icontains=typo_name)
            elif 'Enginnering' in discipline:
                correct_name = discipline.replace('Enginnering', 'Engineering')
                discipline_filters |= Q(branch__icontains=correct_name)
            
            students = students.filter(discipline_filters)
        
        if batch_id:
            students = students.filter(id=batch_id)
        
        students = students.order_by('roll_number')
        
        student_list = []
        for student in students:
            student_list.append({
                'id': student.id,
                'sno': getattr(student, 'id', ''),
                'jee_app_no': getattr(student, 'jee_app_no', ''),
                'JEE App. No / CCMT Roll No': getattr(student, 'jee_app_no', ''),
                'Jee Main Application Number': getattr(student, 'jee_app_no', ''),
                'name': student.name,
                'roll_number': student.roll_number,
                'institute_email': student.institute_email,
                'father_name': student.father_name,
                'category': student.category,
                'pwd': student.pwd,
                'minority': getattr(student, 'minority', ''),
                'reported_status': student.reported_status,
                'branch': student.branch,
                'specialization': getattr(student, 'specialization', ''),
                'year': student.year,
                'source': getattr(student, 'source', 'unknown'),  # Include source field
                'parent_email': getattr(student, 'parent_email', ''),
                'parentEmail': getattr(student, 'parent_email', ''),  # Camel case for frontend
                'alternateEmail': getattr(student, 'personal_email', ''),
                'country': getattr(student, 'country', ''),
                'nationality': getattr(student, 'nationality', ''),
                'blood_group': getattr(student, 'blood_group', ''),
                'bloodGroup': getattr(student, 'blood_group', ''),  # Camel case for frontend
                'blood_group_remarks': getattr(student, 'blood_group_remarks', ''),
                'bloodGroupRemarks': getattr(student, 'blood_group_remarks', ''),  # Camel case for frontend
                'pwd_category': getattr(student, 'pwd_category', ''),
                'pwdCategory': getattr(student, 'pwd_category', ''),  # Camel case for frontend
                'pwd_category_remarks': getattr(student, 'pwd_category_remarks', ''),
                'pwdCategoryRemarks': getattr(student, 'pwd_category_remarks', ''),  # Camel case for frontend
                'admission_mode': getattr(student, 'admission_mode', ''),
                'admissionMode': getattr(student, 'admission_mode', ''),  # Camel case for frontend
                'admission_mode_remarks': getattr(student, 'admission_mode_remarks', ''),
                'admissionModeRemarks': getattr(student, 'admission_mode_remarks', ''),  # Camel case for frontend
                'income_group': getattr(student, 'income_group', ''),
                'incomeGroup': getattr(student, 'income_group', ''),  # Camel case for frontend
                'income': getattr(student, 'income', '')
            })
        
        return JsonResponse({
            'success': True,
            'data': student_list,
            'total': len(student_list),
            'message': f'Students listed successfully - Found {len(student_list)} students'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to list students: {str(e)}'
        }, status=500)

# =============================================================================
# BATCH CRUD OPERATIONS
# =============================================================================

@csrf_exempt
@require_http_methods(["POST"])
def create_batch(request):
    """
    Create new batch
    """
    try:
        field_mappings = {
            'programme': ['programme', 'batch_name'],
            'discipline': 'discipline',
            'year': ['year', 'batchYear'],
            'total_seats': ['total_seats', 'totalSeats'],
            'curriculum_data': ['curriculum', 'curriculum_id', 'disciplineBatch'],
            'specialization': 'specialization'
        }
        
        data = parse_request_data(request, field_mappings)
        
        # Extract mapped fields
        programme = data.get('programme')
        discipline = data.get('discipline')
        year = data.get('year')
        total_seats = data.get('total_seats')
        curriculum_data = data.get('curriculum_data')
        specialization = data.get('specialization', '')
        
        if not all([programme, discipline, year, total_seats]):
            missing_fields = []
            if not programme: missing_fields.append('programme')
            if not discipline: missing_fields.append('discipline') 
            if not year: missing_fields.append('year')
            if not total_seats: missing_fields.append('total_seats/totalSeats')
            
            return JsonResponse({
                'success': False,
                'message': f'Missing required fields: {", ".join(missing_fields)}'
            }, status=400)

        from applications.programme_curriculum.models import Curriculum

        curriculum_obj = None
        if curriculum_data:
            curriculum_id = None
            
            if isinstance(curriculum_data, list) and curriculum_data:
                curriculum_id = str(curriculum_data[0])  # Take first one for single selection
            elif isinstance(curriculum_data, str):
                curriculum_id = curriculum_data.strip()
            else:
                curriculum_id = str(curriculum_data)

            if curriculum_id and curriculum_id not in ['null', 'undefined', '']:
                try:
                    curriculum_obj = Curriculum.objects.get(id=int(curriculum_id), working_curriculum=True)
                except (Curriculum.DoesNotExist, ValueError) as e:
                    return JsonResponse({
                        'success': False,
                        'message': f'Invalid curriculum ID: {curriculum_id}. Curriculum not found or not working.',
                        'validation_error': 'invalid_curriculum'
                    }, status=400)
        
        from applications.programme_curriculum.models import Discipline
        
        try:
            discipline_obj = Discipline.objects.get(id=discipline)
        except Discipline.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': f'Invalid discipline ID: {discipline}'
            }, status=400)

        batch_name = programme
        existing_batch = Batch.objects.filter(
            name=batch_name,
            discipline=discipline_obj,
            year=year,
            running_batch=True
        ).first()
        
        if existing_batch:
            return JsonResponse({
                'success': False,
                'message': f'Batch "{existing_batch.name}" already exists for {existing_batch.discipline.acronym} {existing_batch.year}',
                'validation_error': 'duplicate_batch'
            }, status=400)
        
        # Validate data types
        try:
            year = int(year)
            total_seats = int(total_seats)
        except (ValueError, TypeError):
            return JsonResponse({
                'success': False,
                'message': 'Year and total_seats must be integers'
            }, status=400)
        
        try:
            batch = Batch.objects.create(
                name=batch_name,
                discipline=discipline_obj,
                year=year,
                curriculum=curriculum_obj,
                total_seats=total_seats,
                running_batch=True
            )
                
        except Exception as batch_error:
            return JsonResponse({
                'success': False,
                'message': f'Failed to create batch: {str(batch_error)}'
            }, status=500)

        if curriculum_obj:
            success_message = f'Batch created successfully with curriculum: {curriculum_obj.name}'
            curriculum_info = {
                'curriculum': curriculum_obj.name,
                'curriculum_id': curriculum_obj.id
            }
        else:
            success_message = 'Batch created successfully (no curriculum assigned)'
            curriculum_info = {
                'curriculum': None,
                'curriculum_id': None
            }
        
        return JsonResponse({
            'success': True,
            'data': {
                'id': batch.id,
                'programme': batch.name,
                'discipline': batch.discipline.name,
                'disciplineAcronym': batch.discipline.acronym,
                'year': batch.year,
                'total_seats': batch.total_seats,
                'totalSeats': batch.total_seats,  # Return both formats
                'specialization': specialization,
                'running_batch': batch.running_batch,
                **curriculum_info  # Merge curriculum info into response
            },
            'message': success_message
        })
        
    except json.JSONDecodeError as e:
        return JsonResponse({
            'success': False,
            'message': f'Invalid JSON data: {str(e)}'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to create batch: {str(e)}'
        }, status=500)

@csrf_exempt
@require_http_methods(["PUT"])
def update_batch(request, batch_id):
    """
    Update existing batch
    """
    try:
        data = json.loads(request.body)
        
        try:
            batch = BatchConfiguration.objects.get(id=batch_id)
        except BatchConfiguration.DoesNotExist:
            # Try to find it in regular Batch model
            try:
                from applications.programme_curriculum.models import Batch
                regular_batch = Batch.objects.get(id=batch_id)
                
                # Create a corresponding BatchConfiguration if it doesn't exist
                batch_config = BatchConfiguration.objects.create(
                    programme=regular_batch.name,
                    discipline=regular_batch.discipline.name if regular_batch.discipline else "Unknown",
                    year=regular_batch.year,
                    total_seats=regular_batch.total_seats
                )
                batch_config.id = batch_id  # Force the same ID
                batch_config.save()
                batch = batch_config
                
            except:
                return JsonResponse({
                    'success': False,
                    'message': f'Batch with ID {batch_id} not found in any model'
                }, status=404)

        # Update fields
        if 'programme' in data:
            batch.programme = data['programme']
        if 'discipline' in data:
            batch.discipline = data['discipline']
        if 'year' in data:
            batch.year = data['year']
        if 'total_seats' in data or 'totalSeats' in data:
            new_total_seats = data.get('total_seats') or data.get('totalSeats')
            batch.total_seats = new_total_seats

        if 'displayBranch' in data:
            batch.discipline = data['displayBranch']
        
        batch.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Batch updated successfully',
            'batch': {
                'id': batch.id,
                'programme': batch.programme,
                'discipline': batch.discipline,
                'year': batch.year,
                'total_seats': batch.total_seats
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to update batch: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def list_batches_with_status(request):
    """
    List all batches with status
    """
    try:
        batches = BatchConfiguration.objects.all().order_by('programme', 'discipline', 'year')
        
        batch_list = []
        for batch in batches:
            batch_list.append({
                'id': batch.id,
                'programme': batch.programme,
                'discipline': batch.discipline,
                'year': batch.year,
                'total_seats': batch.total_seats,
                'filled_seats': batch.filled_seats,
                'available_seats': batch.available_seats
            })
        
        return JsonResponse({
            'success': True,
            'data': batch_list,
            'message': 'Batches listed successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to list batches: {str(e)}'
        }, status=500)

# =============================================================================
# STUDENT STATUS CRUD
# =============================================================================

@csrf_exempt
@require_http_methods(["PUT"])
def update_student_status_crud(request, student_id):
    """
    Update student status
    """
    try:
        data = json.loads(request.body)
        reported_status = data.get('reported_status')
        
        if not reported_status:
            return JsonResponse({
                'success': False,
                'message': 'Missing reported_status field'
            }, status=400)
        
        try:
            student = StudentBatchUpload.objects.get(id=student_id)
        except StudentBatchUpload.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Student not found'
            }, status=404)
        
        old_status = student.reported_status
        student.reported_status = reported_status
        student.save()

        try:
            from applications.programme_curriculum.models_student_management import BatchConfiguration
            from django.db.models import Q

            discipline_q = Q()
            if student.branch:
                branch_upper = student.branch.upper()
                if 'COMPUTER SCIENCE' in branch_upper or 'CSE' in branch_upper:
                    discipline_q = Q(discipline='Computer Science and Engineering')
                elif 'ELECTRONICS' in branch_upper or 'ECE' in branch_upper:
                    discipline_q = Q(discipline='Electronics and Communication Engineering')
                elif 'DESIGN' in branch_upper:
                    discipline_q = Q(discipline='Design')
                elif 'MECHANICAL' in branch_upper:
                    discipline_q = Q(discipline='Mechanical Engineering')
                elif 'SMART' in branch_upper:
                    discipline_q = Q(discipline='Smart Manufacturing')

            if discipline_q:
                matching_batches = BatchConfiguration.objects.filter(
                    discipline_q,
                    year=student.year
                )
                for batch in matching_batches:
                    batch.calculate_seats()
        except Exception as e:
            pass
        
        return JsonResponse({
            'success': True,
            'message': f'Student status updated from {old_status} to {reported_status}'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to update student status: {str(e)}'
        }, status=500)

# =============================================================================
# PASSWORD MANAGEMENT
# =============================================================================

@csrf_exempt
@require_http_methods(["POST"])
def auto_generate_passwords_for_batch(request):
    """
    Auto-generate passwords for a batch
    """
    try:
        data = json.loads(request.body)
        batch_id = data.get('batch_id')
        
        if not batch_id:
            return JsonResponse({
                'success': False,
                'message': 'Missing batch_id'
            }, status=400)

        try:
            batch = BatchConfiguration.objects.get(id=batch_id)
        except BatchConfiguration.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Batch not found'
            }, status=404)
        
        students = StudentBatchUpload.objects.filter(
            branch__icontains=batch.discipline,
            year=batch.year
        )
        
        updated_count = 0
        for student in students:
            if not hasattr(student, 'password') or not student.password:
                student.password = generate_password()
                student.save()
                updated_count += 1
        
        return JsonResponse({
            'success': True,
            'message': f'Generated passwords for {updated_count} students',
            'data': {
                'updated_count': updated_count,
                'total_students': students.count()
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to generate passwords: {str(e)}'
        }, status=500)

# =============================================================================
# HELPER FUNCTIONS FOR STUDENT MANAGEMENT
# =============================================================================

def get_available_curriculums_for_batch(batch_obj):
    """
    Get available curriculums for a batch based on programme and discipline
    For multi-curriculum PG batches (M.Tech, M.Des), return all relevant curriculums
    """
    if not batch_obj:
        return []
    
    from applications.programme_curriculum.models import Curriculum
    
    if hasattr(batch_obj, 'curriculum_options') and batch_obj.curriculum_options:
        return batch_obj.curriculum_options

    # Check if this is a multi-curriculum PG batch (M.Tech, M.Des, PhD, etc.)
    if batch_obj.name in ['M.Tech', 'M.Des', 'Phd'] and not batch_obj.curriculum:
        # Get all working curriculums for this discipline and programme
        available_curriculums = Curriculum.objects.filter(
            working_curriculum=True,
            programme__name__icontains=batch_obj.name
        )
        
        # Filter by discipline if possible (based on common patterns)
        discipline_name = batch_obj.discipline.name.lower()
        discipline_filtered = available_curriculums
        
        # Apply discipline-specific filtering for various disciplines
        if 'computer science' in discipline_name or 'cse' in discipline_name:
            discipline_filtered = available_curriculums.filter(
                Q(name__icontains='AI') |
                Q(name__icontains='Data Science') |
                Q(name__icontains='Computer Science')
            )
        elif 'electronics' in discipline_name or 'ece' in discipline_name:
            discipline_filtered = available_curriculums.filter(
                Q(name__icontains='Communication and Signal Processing') |
                Q(name__icontains='Nanoelectronics and VLSI Design') |
                Q(name__icontains='Power & Control')
            )
        elif 'mechanical engineering' in discipline_name or 'me' in discipline_name:
            discipline_filtered = available_curriculums.filter(
                Q(name__icontains='Design') |
                Q(name__icontains='CAD/CAM') |
                Q(name__icontains='Manufacturing and Automation')
            )
        elif 'mechatronics' in discipline_name or 'mt' in discipline_name:
            discipline_filtered = available_curriculums.filter(
                Q(name__icontains='Mechatronics')
            )
        elif 'design' in discipline_name or 'des' in discipline_name:
            discipline_filtered = available_curriculums.filter(
                Q(name__icontains='Design')
            )
        final_curriculums = discipline_filtered if discipline_filtered.exists() else available_curriculums
        return [{'id': c.id, 'name': c.name, 'version': c.version} for c in final_curriculums]
    
    elif batch_obj.curriculum:
        # Single curriculum batch
        return [{'id': batch_obj.curriculum.id, 'name': batch_obj.curriculum.name, 'version': batch_obj.curriculum.version}]
    
    return []

def get_batch_curriculum_display(batch_obj):
    """
    Get curriculum display information for a batch
    Returns appropriate curriculum display text based on batch type
    """
    if not batch_obj:
        return "No curriculum assigned"
    
    # Get available curriculums dynamically
    available_curriculums = get_available_curriculums_for_batch(batch_obj)
    if len(available_curriculums) > 1:
        curriculum_names = [curr['name'] for curr in available_curriculums]
        return f"{len(curriculum_names)} curriculums: {', '.join(curriculum_names)}"
    elif len(available_curriculums) == 1:
        return available_curriculums[0]['name']
    elif batch_obj.curriculum:
        return batch_obj.curriculum.name
    
    return "Multi-curriculum batch (curriculums assigned based on specialization)"

def get_curriculum_by_specialization(specialization, discipline_obj):
    """
    Map student specialization to appropriate curriculum based on IIITDMJ official specializations
    Returns curriculum ID based on specialization and discipline
    """
    if not specialization or not discipline_obj:
        return None
    
    from applications.programme_curriculum.models import Curriculum
    from django.db.models import Q
    
    specialization_lower = specialization.lower().strip()
    discipline_name = discipline_obj.name.lower()
    
    # M.Tech Computer Science & Engineering (CSE) specializations
    if 'computer science' in discipline_name or 'cse' in discipline_name:
        # AI & ML specialization patterns
        ai_ml_patterns = ['ai & ml', 'ai and ml', 'artificial intelligence', 'machine learning', 'ai/ml']
        if any(pattern in specialization_lower for pattern in ai_ml_patterns):
            try:
                curriculum = Curriculum.objects.filter(
                    Q(name__icontains='AI & ML') |
                    Q(name__icontains='AI and ML') |
                    (Q(name__icontains='AI') & Q(name__icontains='ML')),
                    working_curriculum=True,
                    programme__name__icontains='M.Tech'
                ).first()
                if curriculum:
                    return curriculum.id
            except:
                pass
        
        # Data Science specialization patterns
        data_science_patterns = ['data science', 'data analytics', 'big data']
        if any(pattern in specialization_lower for pattern in data_science_patterns):
            try:
                curriculum = Curriculum.objects.filter(
                    Q(name__icontains='Data Science') |
                    Q(name__icontains='Data Analytics'),
                    working_curriculum=True,
                    programme__name__icontains='M.Tech'
                ).first()
                if curriculum:
                    return curriculum.id
            except:
                pass
        
        # General CSE patterns
        general_cse_patterns = ['total (cse)', 'total cse', 'computer science', 'general cse']
        if any(pattern in specialization_lower for pattern in general_cse_patterns):
            try:
                curriculum = Curriculum.objects.filter(
                    working_curriculum=True,
                    programme__name__icontains='M.Tech',
                    programme__discipline=discipline_obj
                ).exclude(
                    Q(name__icontains='AI') | Q(name__icontains='Data Science')
                ).first()
                if curriculum:
                    return curriculum.id
            except:
                pass
    
    # M.Tech Electronics and Communication Engineering (ECE) specializations  
    elif 'electronics' in discipline_name or 'ece' in discipline_name:
        # Communication and Signal Processing patterns
        comm_patterns = ['communication and signal processing', 'signal processing', 'communication', 'dsp']
        if any(pattern in specialization_lower for pattern in comm_patterns):
            try:
                curriculum = Curriculum.objects.filter(
                    Q(name__icontains='Communication') |
                    Q(name__icontains='Signal Processing'),
                    working_curriculum=True,
                    programme__name__icontains='M.Tech'
                ).first()
                if curriculum:
                    return curriculum.id
            except:
                pass
        
        # VLSI and Nanoelectronics patterns
        vlsi_patterns = ['nanoelectronics and vlsi design', 'vlsi', 'nanoelectronics', 'chip design']
        if any(pattern in specialization_lower for pattern in vlsi_patterns):
            try:
                curriculum = Curriculum.objects.filter(
                    Q(name__icontains='VLSI') |
                    Q(name__icontains='Nanoelectronics'),
                    working_curriculum=True,
                    programme__name__icontains='M.Tech'
                ).first()
                if curriculum:
                    return curriculum.id
            except:
                pass
        
        # Power & Control patterns
        power_patterns = ['power & control', 'power and control', 'power systems', 'control systems']
        if any(pattern in specialization_lower for pattern in power_patterns):
            try:
                curriculum = Curriculum.objects.filter(
                    Q(name__icontains='Power') |
                    Q(name__icontains='Control'),
                    working_curriculum=True,
                    programme__name__icontains='M.Tech'
                ).first()
                if curriculum:
                    return curriculum.id
            except:
                pass
        
        # General ECE patterns
        general_ece_patterns = ['total (ece)', 'total ece', 'electronics', 'general ece']
        if any(pattern in specialization_lower for pattern in general_ece_patterns):
            try:
                curriculum = Curriculum.objects.filter(
                    working_curriculum=True,
                    programme__name__icontains='M.Tech',
                    programme__discipline=discipline_obj
                ).first()
                if curriculum:
                    return curriculum.id
            except:
                pass

    elif 'mechatronics' in discipline_name or 'mt' in discipline_name:
        mechatronics_patterns = ['mechatronics', 'robotics', 'automation']
        if any(pattern in specialization_lower for pattern in mechatronics_patterns):
            try:
                curriculum = Curriculum.objects.filter(
                    Q(name__icontains='MT PG') | Q(name__icontains='Mechatronics'),
                    working_curriculum=True,
                    programme__name__icontains='M.Tech'
                ).first()
                if curriculum:
                    return curriculum.id
            except Exception as e:
                pass
    
    # M.Tech Mechanical Engineering (ME) specializations
    elif 'mechanical' in discipline_name or 'me' in discipline_name:
        # Only match mechanical-specific design patterns, not generic "design"
        design_patterns = ['product design', 'engineering design', 'mechanical design']
        if any(pattern in specialization_lower for pattern in design_patterns):
            try:
                curriculum = Curriculum.objects.filter(
                    name__icontains='Design',
                    working_curriculum=True,
                    programme__name__icontains='M.Tech'
                ).first()
                if curriculum:
                    return curriculum.id
            except:
                pass
        
        # CAD/CAM patterns
        cad_patterns = ['cad/cam', 'cad', 'cam', 'computer aided']
        if any(pattern in specialization_lower for pattern in cad_patterns):
            try:
                curriculum = Curriculum.objects.filter(
                    Q(name__icontains='CAD') | Q(name__icontains='CAM'),
                    working_curriculum=True,
                    programme__name__icontains='M.Tech'
                ).first()
                if curriculum:
                    return curriculum.id
            except:
                pass
        
        # Manufacturing patterns
        manufacturing_patterns = ['manufacturing and automation', 'manufacturing', 'automation', 'production']
        if any(pattern in specialization_lower for pattern in manufacturing_patterns):
            try:
                curriculum = Curriculum.objects.filter(
                    Q(name__icontains='Manufacturing') | Q(name__icontains='Automation'),
                    working_curriculum=True,
                    programme__name__icontains='M.Tech'
                ).first()
                if curriculum:
                    return curriculum.id
            except:
                pass
        
        # General ME patterns
        general_me_patterns = ['total (me)', 'total me', 'mechanical', 'general me']
        if any(pattern in specialization_lower for pattern in general_me_patterns):
            try:
                curriculum = Curriculum.objects.filter(
                    working_curriculum=True,
                    programme__name__icontains='M.Tech',
                    programme__discipline=discipline_obj
                ).first()
                if curriculum:
                    return curriculum.id
            except:
                pass
    # Design specializations - Handle Design discipline with Design specialization
    elif 'design' in discipline_name:
        try:
            curriculum = Curriculum.objects.filter(
                name__icontains='Design',
                working_curriculum=True,
                programme__name__icontains='M.Des',
                programme__discipline=discipline_obj
            ).first()
            if curriculum:
                return curriculum.id
            # Fallback to M.Tech Design if M.Des not found
            curriculum = Curriculum.objects.filter(
                name__icontains='Design',
                working_curriculum=True,
                programme__name__icontains='M.Tech',
                programme__discipline=discipline_obj
            ).first()
            if curriculum:
                return curriculum.id
        except:
            pass
    
    # No fallback - only assign curriculum for exact specialization matches
    return None


def get_batch_name_from_discipline(discipline_name, programme_type):
    if programme_type == 'ug':
        if 'design' in discipline_name.lower():
            batch_name = 'B.Des'
        else:
            batch_name = 'B.Tech'
    elif programme_type == 'pg':
        if 'design' in discipline_name.lower():
            batch_name = 'M.Des'
        else:
            batch_name = 'M.Tech'
    else:
        batch_name = programme_type.upper()
    
    return batch_name

def get_or_create_discipline(discipline_name):
    from applications.programme_curriculum.models import Discipline

    normalized_name = discipline_name.strip()
    discipline_lower = normalized_name.lower()

    discipline_mapping = {
        'computer science and engineering': 'Computer Science and Engineering',
        'electronics and communication engineering': 'Electronics and Communication Engineering', 
        'mechanical engineering': 'Mechanical Engineering',
        'smart manufacturing': 'Smart Manufacturing',
        'design': 'Design'
    }

    database_discipline_name = None
    for key, value in discipline_mapping.items():
        if key in discipline_lower:
            database_discipline_name = value
            break
    
    if not database_discipline_name:
        database_discipline_name = normalized_name
    
    try:
        discipline = Discipline.objects.get(name=database_discipline_name)
        return discipline
    except Discipline.DoesNotExist:
        # Create new discipline as fallback
        discipline = Discipline.objects.create(
            name=database_discipline_name,
            acronym=get_discipline_acronym(database_discipline_name)
        )
        return discipline

def get_discipline_acronym(discipline_name):

    discipline_mappings = {
        'Computer Science and Engineering': 'CSE',
        'Electronics and Communication Engineering': 'ECE',
        'Mechanical Engineering': 'ME',
        'Smart Manufacturing': 'SM',
        'Design': 'DES'
    }
    
    for full_name, acronym in discipline_mappings.items():
        if full_name.lower() in discipline_name.lower():
            return acronym

    words = discipline_name.split()
    return ''.join([word[0].upper() for word in words if word])[:5]

def create_or_update_main_student_record(student_data, batch_obj, batch_year):

    try:
        roll_number = student_data.get('Institute Roll Number', '').strip()
        if not roll_number:
            return None

        try:
            user = User.objects.get(username=roll_number)
        except User.DoesNotExist:
            full_name = student_data.get('Name', '')
            name_parts = full_name.split() if full_name else ['']
            user = User.objects.create_user(
                username=roll_number,
                first_name=name_parts[0] if name_parts else '',
                last_name=' '.join(name_parts[1:]) if len(name_parts) > 1 else '',
                email=student_data.get('Institute Email ID', ''),
                password='student123'  # Default password
            )

        extra_info, created = ExtraInfo.objects.get_or_create(
            user=user,
            defaults={
                'id': roll_number,
                'user_type': 'student',
                'phone_no': student_data.get('Mobile No', ''),
                'address': student_data.get('Full Address', ''),
                'sex': 'M' if student_data.get('Gender', '').lower() == 'male' else 'F' if student_data.get('Gender', '').lower() == 'female' else 'O',
                'date_of_birth': None  # Will be set later if date parsing works
            }
        )

        dob_value = student_data.get('Date of Birth') or student_data.get('date_of_birth') or student_data.get('dob')
        if dob_value:
            dob = parse_date_flexible(dob_value)
            if dob:
                extra_info.date_of_birth = dob
                extra_info.save()

        category_mapping = {
            'General': 'GEN',
            'GEN': 'GEN',
            'OBC': 'OBC', 
            'SC': 'SC',
            'ST': 'ST'
        }
        category = category_mapping.get(student_data.get('Category', 'GEN'), 'GEN')

        current_semester = calculate_current_semester(batch_year)

        academic_student, created = AcademicStudent.objects.get_or_create(
            id=extra_info,
            defaults={
                'programme': batch_obj.name,
                'batch': batch_year,
                'batch_id': batch_obj,
                'category': category,
                'father_name': student_data.get("Father's Name", ''),
                'mother_name': student_data.get("Mother's Name", ''),
                'cpi': 0.0,
                'curr_semester_no': current_semester  # Dynamic semester calculation
            }
        )
        
        return academic_student
        
    except Exception as e:
        return None


# =============================================================================
# INDIVIDUAL STUDENT CRUD OPERATIONS
# =============================================================================

@csrf_exempt
@require_http_methods(["GET"])
def get_student(request, student_id):
    """
    Get a single student by ID
    """
    try:
        student = StudentBatchUpload.objects.get(id=student_id)

        student_data = {
            # Core identification
            'id': student.id,
            'jee_app_no': student.jee_app_no,
            'jeeAppNo': student.jee_app_no,
            'roll_number': student.roll_number,
            'rollNumber': student.roll_number,
            'institute_email': student.institute_email,
            'instituteEmail': student.institute_email,

            'name': student.name,
            'father_name': student.father_name,
            'fatherName': student.father_name,
            'fname': student.father_name,  # Additional alias
            'mother_name': student.mother_name,
            'motherName': student.mother_name,
            'gender': student.gender,
            'category': student.category,
            'pwd': student.pwd,
            'minority': getattr(student, 'minority', ''),
            'date_of_birth': student.date_of_birth.isoformat() if student.date_of_birth else '',
            'dateOfBirth': student.date_of_birth.isoformat() if student.date_of_birth else '',

            'phone_number': student.phone_number,
            'phoneNumber': student.phone_number,
            'mobile': student.phone_number,  # Additional alias
            'personalEmail': student.personal_email,
            'address': student.address,
            'state': student.state,

            'branch': student.branch,
            'specialization': getattr(student, 'specialization', ''),
            'ai_rank': student.ai_rank,
            'aiRank': student.ai_rank,
            'category_rank': student.category_rank,
            'categoryRank': student.category_rank,
            'tenth_marks': student.tenth_marks,
            'tenthMarks': student.tenth_marks,
            'twelfth_marks': student.twelfth_marks,
            'twelfthMarks': student.twelfth_marks,

            'father_occupation': student.father_occupation,
            'fatherOccupation': student.father_occupation,
            'father_mobile': student.father_mobile,
            'fatherMobile': student.father_mobile,
            'mother_occupation': student.mother_occupation,
            'motherOccupation': student.mother_occupation,
            'mother_mobile': student.mother_mobile,
            'motherMobile': student.mother_mobile,

            'allotted_category': student.allotted_category,
            'allottedCategory': student.allotted_category,
            'allotted_gender': student.allotted_gender,
            'allottedGender': student.allotted_gender,
            'aadhar_number': student.aadhar_number,
            'aadharNumber': student.aadhar_number,
            'parent_email': getattr(student, 'parent_email', ''),
            'parentEmail': getattr(student, 'parent_email', ''),  # Camel case for frontend
            'alternateEmail': getattr(student, 'personal_email', ''),
            'nationality': getattr(student, 'nationality', ''),
            'blood_group': getattr(student, 'blood_group', ''),
            'bloodGroup': getattr(student, 'blood_group', ''),  # Camel case for frontend
            'blood_group_remarks': getattr(student, 'blood_group_remarks', ''),
            'bloodGroupRemarks': getattr(student, 'blood_group_remarks', ''),  # Camel case for frontend
            'pwd_category': getattr(student, 'pwd_category', ''),
            'pwdCategory': getattr(student, 'pwd_category', ''),  # Camel case for frontend
            'pwd_category_remarks': getattr(student, 'pwd_category_remarks', ''),
            'pwdCategoryRemarks': getattr(student, 'pwd_category_remarks', ''),  # Camel case for frontend
            'admission_mode': getattr(student, 'admission_mode', ''),
            'admissionMode': getattr(student, 'admission_mode', ''),  # Camel case for frontend
            'admission_mode_remarks': getattr(student, 'admission_mode_remarks', ''),
            'admissionModeRemarks': getattr(student, 'admission_mode_remarks', ''),  # Camel case for frontend
            'income_group': getattr(student, 'income_group', ''),
            'incomeGroup': getattr(student, 'income_group', ''),  # Camel case for frontend
            'income': getattr(student, 'income', ''),

            'reported_status': student.reported_status,
            'reportedStatus': student.reported_status,
            'status': student.reported_status,  # Additional alias
            'year': student.year,
            'programme_type': student.programme_type,
            'programmeType': student.programme_type,
            'created_at': student.created_at.isoformat() if hasattr(student, 'created_at') and student.created_at else '',
            'createdAt': student.created_at.isoformat() if hasattr(student, 'created_at') and student.created_at else '',
            'updated_at': student.updated_at.isoformat() if hasattr(student, 'updated_at') and student.updated_at else '',
            'updatedAt': student.updated_at.isoformat() if hasattr(student, 'updated_at') and student.updated_at else '',
        }
        
        return JsonResponse({
            'success': True,
            'student': student_data,
            'message': f'Student {student.name} retrieved successfully'
        })
        
    except StudentBatchUpload.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': f'Student with ID {student_id} not found'
        }, status=404)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to retrieve student: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["PUT", "POST"])
def update_student(request, student_id):
    """
    Update a student by ID
    """
    try:
        student = StudentBatchUpload.objects.get(id=student_id)
        data = json.loads(request.body)
        old_discipline = student.branch
        discipline_changed = False

        field_mapping = {
            # Identification fields
            'jeeAppNo': 'jee_app_no',
            'rollNumber': 'roll_number',
            'instituteEmail': 'institute_email',

            # Names
            'fatherName': 'father_name',
            'fname': 'father_name',
            'mname': 'mother_name',
            'motherName': 'mother_name',

            # Dates & phones
            'dateOfBirth': 'date_of_birth',
            'dob': 'date_of_birth',
            'phoneNumber': 'phone_number',
            'mobile': 'phone_number',

            # Email / alternate
            'email': 'personal_email',
            'alternateEmail': 'personal_email',

            # Academic ranks
            'jeeRank': 'ai_rank',
            'categoryRank': 'category_rank',
            'tenthMarks': 'tenth_marks',
            'twelfthMarks': 'twelfth_marks',

            # Parents / contacts
            'fatherOccupation': 'father_occupation',
            'fatherMobile': 'father_mobile',
            'motherOccupation': 'mother_occupation',
            'motherMobile': 'mother_mobile',
            'parentEmail': 'parent_email',
            'parent_email': 'parent_email',

            # Allotment
            'allottedCategory': 'allotted_category',
            'allottedGender': 'allotted_gender',

            # New fields added
            'country': 'country',
            'nationality': 'nationality',
            'bloodGroup': 'blood_group',
            'blood_group': 'blood_group',
            'bloodGroupRemarks': 'blood_group_remarks',
            'blood_group_remarks': 'blood_group_remarks',
            'pwdCategory': 'pwd_category',
            'pwd_category': 'pwd_category',
            'pwdCategoryRemarks': 'pwd_category_remarks',
            'admissionMode': 'admission_mode',
            'admissionModeRemarks': 'admission_mode_remarks',
            'incomeGroup': 'income_group',
            'income': 'income',

            'aadharNumber': 'aadhar_number',
            'reportedStatus': 'reported_status',
            'programmeType': 'programme_type'
        }

        
        jee_app_no = data.get('jeeAppNo') or data.get('jee_app_no')
        roll_number = data.get('rollNumber') or data.get('roll_number')
        institute_email = data.get('instituteEmail') or data.get('institute_email')

        if jee_app_no and jee_app_no != student.jee_app_no:
            if StudentBatchUpload.objects.filter(jee_app_no=jee_app_no).exclude(id=student_id).exists():
                return JsonResponse({
                    'success': False,
                    'message': f'JEE Application Number {jee_app_no} already exists for another student'
                }, status=400)

        if roll_number and roll_number != student.roll_number:
            if StudentBatchUpload.objects.filter(roll_number=roll_number).exclude(id=student_id).exists():
                return JsonResponse({
                    'success': False,
                    'message': f'Roll Number {roll_number} already exists for another student'
                }, status=400)

        if institute_email and institute_email != student.institute_email:
            if StudentBatchUpload.objects.filter(institute_email=institute_email).exclude(id=student_id).exists():
                return JsonResponse({
                    'success': False,
                    'message': f'Institute Email {institute_email} already exists for another student'
                }, status=400)

        for frontend_field, backend_field in field_mapping.items():
            if frontend_field in data:
                value = data[frontend_field]
                if hasattr(student, backend_field):
                    setattr(student, backend_field, value)

        direct_fields = [
            'name', 'gender', 'category', 'pwd', 'minority', 'address', 'state', 'branch', 'specialization',
            'personal_email', 'parent_email', 'country', 'nationality',
            'blood_group', 'blood_group_remarks', 'pwd_category', 'pwd_category_remarks',
            'admission_mode', 'admission_mode_remarks', 'income_group', 'income'
        ]
        
        for field in direct_fields:
            if field in data:
                if field == 'branch' and data[field] != old_discipline:
                    discipline_changed = True
                setattr(student, field, data[field])

        dob_value = data.get('dob') or data.get('dateOfBirth') or data.get('date_of_birth')
        if dob_value:
            try:
                if isinstance(dob_value, str) and dob_value.strip():
                    if '/' in dob_value:
                        student.date_of_birth = datetime.strptime(dob_value, '%m/%d/%Y').date()
                    elif '-' in dob_value:
                        student.date_of_birth = datetime.strptime(dob_value, '%Y-%m-%d').date()
                else:
                    student.date_of_birth = dob_value
            except Exception as e:
                pass

        jee_rank_value = data.get('jeeRank') or data.get('aiRank') or data.get('ai_rank')
        if jee_rank_value:
            try:
                sanitized_rank = sanitize_rank_value(jee_rank_value)
                student.ai_rank = int(sanitized_rank.replace(',', '')) if sanitized_rank.replace(',', '').isdigit() else None
            except (ValueError, TypeError) as e:
                student.ai_rank = None

        rank_fields = ['category_rank', 'tenth_marks', 'twelfth_marks']
        for field in rank_fields:
            frontend_field = field
            if field == 'category_rank':
                frontend_field = 'categoryRank'
            elif field == 'tenth_marks':
                frontend_field = 'tenthMarks'
            elif field == 'twelfth_marks':
                frontend_field = 'twelfthMarks'
                
            # Get value from either frontend field name or backend field name
            value = data.get(frontend_field, data.get(field))
            if value is None or value == '' or value == 'null':
                setattr(student, field, None)
            else:
                try:
                    if field == 'category_rank':
                        sanitized_value = sanitize_rank_value(value)
                        if sanitized_value and sanitized_value.replace(',', '').isdigit():
                            setattr(student, field, int(sanitized_value.replace(',', '')))
                        else:
                            setattr(student, field, None)
                    else:
                        value_str = str(value).replace(',', '')
                        if value_str.isdigit():
                            setattr(student, field, int(value_str))
                        else:
                            setattr(student, field, None)
                except (ValueError, TypeError):
                    setattr(student, field, None)
        
        # Update timestamp if field exists
        if hasattr(student, 'updated_at'):
            student.updated_at = timezone.now()
        
        student.save()
        
        # Handle discipline change using existing batch change API logic if discipline was changed
        if discipline_changed and student.reported_status == 'REPORTED':
            try:
                from applications.academic_information.models import Student as AcademicStudent
                from applications.programme_curriculum.models import Batch, Discipline
                from applications.academic_procedures.models import BatchChangeHistory
                
                academic_student = AcademicStudent.objects.filter(id__id=student.roll_number).first()
                
                if academic_student:
                    new_discipline_obj = Discipline.objects.filter(name__iexact=student.branch).first()
                    if not new_discipline_obj:
                        discipline_upper = student.branch.upper()
                        if 'COMPUTER SCIENCE' in discipline_upper or 'CSE' in discipline_upper:
                            new_discipline_obj = Discipline.objects.filter(name__icontains='Computer Science').first()
                        elif 'ELECTRONICS' in discipline_upper or 'ECE' in discipline_upper:
                            new_discipline_obj = Discipline.objects.filter(name__icontains='Electronics').first()
                        elif 'MECHANICAL' in discipline_upper or 'ME' in discipline_upper:
                            new_discipline_obj = Discipline.objects.filter(name__icontains='Mechanical').first()
                        elif 'SMART' in discipline_upper or 'MANUFACTURING' in discipline_upper:
                            new_discipline_obj = Discipline.objects.filter(name__icontains='Smart Manufacturing').first()
                        elif 'DESIGN' in discipline_upper:
                            new_discipline_obj = Discipline.objects.filter(name__icontains='Design').first()
                    
                    if new_discipline_obj:
                        new_batch = Batch.objects.filter(
                            discipline=new_discipline_obj,
                            year=student.year,
                            running_batch=True
                        ).first()
                        
                        if new_batch and academic_student.batch_id != new_batch:
                            old_batch = academic_student.batch_id
                            BatchChangeHistory.objects.create(
                                student=academic_student,
                                old_batch=old_batch,
                                new_batch=new_batch,
                            )

                            academic_student.batch_id = new_batch
                            academic_student.save()
                            
                            # Update department and specialization
                            try:
                                from applications.globals.models import DepartmentInfo
                                dept_name = new_batch.discipline.acronym
                                department = DepartmentInfo.objects.filter(name=dept_name).first()
                                if department:
                                    academic_student.id.department = department
                                    academic_student.id.save()
                                academic_student.specialization = dept_name
                                academic_student.save()
                            except:
                                pass
                            
            except Exception as batch_change_error:
                pass
        
        return JsonResponse({
            'success': True,
            'message': f'Student {student.name} updated successfully',
            'student_id': student.id,
            'discipline_changed': discipline_changed,
            'old_discipline': old_discipline if discipline_changed else None,
            'new_discipline': student.branch if discipline_changed else None
        })
        
    except StudentBatchUpload.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': f'Student with ID {student_id} not found'
        }, status=404)
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid JSON data provided'
        }, status=400)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to update student: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["DELETE", "POST"])
def delete_student(request, student_id):
    """
    Delete a student from the batch upload table AND perform cascade deletion from main academic tables
    URL: /programme_curriculum/api/student/{id}/delete/
    
    This performs COMPLETE CASCADE DELETION:
    - Removes from StudentBatchUpload (source table)
    - Removes from AcademicStudent (main academic table) 
    - Removes from ExtraInfo (personal info table)
    - Removes from User (authentication table)
    """
    try:
        from applications.academic_information.models import Student as AcademicStudent
        from applications.globals.models import ExtraInfo
        from django.contrib.auth.models import User
        from django.db import transaction

        student = StudentBatchUpload.objects.get(id=student_id)
        student_name = student.name
        student_roll = student.roll_number

        deletion_info = {
            'id': student.id,
            'name': student_name,
            'roll_number': student_roll,
            'jee_app_no': student.jee_app_no,
            'deleted_by': getattr(request, 'user', None) and hasattr(request.user, 'username') and request.user.username or 'Unknown',
            'deleted_at': timezone.now().isoformat()
        }

        deleted_records = {
            'student_batch_upload': False,
            'academic_student': False,
            'extra_info': False,
            'user_account': False
        }
        
        # Perform CASCADE DELETION in transaction with proper error handling
        try:
            with transaction.atomic():
                if student_roll:
                    try:
                        academic_student = AcademicStudent.objects.get(id__id=student_roll)
                        academic_student.delete()
                        deleted_records['academic_student'] = True
                    except AcademicStudent.DoesNotExist:
                        pass

                if student_roll:
                    try:
                        extra_info = ExtraInfo.objects.get(id=student_roll)
                        extra_info.delete()
                        deleted_records['extra_info'] = True
                    except ExtraInfo.DoesNotExist:
                        pass

                user_to_delete = None
                if student.user:
                    user_to_delete = student.user
                    student.user = None 
                    student.save()
                elif student_roll: 
                    try:
                        user_to_delete = User.objects.get(username=student_roll)  
                    except User.DoesNotExist:
                        pass
                student.delete()
                deleted_records['student_batch_upload'] = True
        
        except Exception as transaction_error:
            return JsonResponse({
                'success': False,
                'message': f'Failed to delete student records: {str(transaction_error)}'
            }, status=500)
        # Handle User deletion separately to avoid transaction conflicts
        if user_to_delete:
            try:
                try:
                    remaining_students = StudentBatchUpload.objects.filter(user=user_to_delete)
                    if remaining_students.exists():
                        remaining_students.update(user=None)
                except Exception as cleanup_error:
                    pass

                with transaction.atomic():
                    user_to_delete.delete()
                    deleted_records['user_account'] = True
                    
            except Exception as user_deletion_error:
                pass
        
        # Build deletion summary
        deletion_summary = []
        if deleted_records['student_batch_upload']:
            deletion_summary.append("StudentBatchUpload")
        if deleted_records['academic_student']:
            deletion_summary.append("AcademicStudent")
        if deleted_records['extra_info']:
            deletion_summary.append("ExtraInfo")
        if deleted_records['user_account']:
            deletion_summary.append("User Account")
        
        return JsonResponse({
            'success': True,
            'message': f'Student {student_name} completely deleted from all systems',
            'cascade_deletion': True,
            'deleted_from': deletion_summary,
            'deleted_student': {
                'id': deletion_info['id'],
                'name': deletion_info['name'],
                'roll': deletion_info['roll_number']
            },
            'deletion_details': deleted_records
        }, status=200)
    
    except StudentBatchUpload.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': f'Student with ID {student_id} not found'
        }, status=404)
    
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        
        return JsonResponse({
            'success': False,
            'message': f'Failed to delete student: {str(e)}',
            'error_details': str(e)
        }, status=500)


# =============================================================================
# BULK STATUS UPDATE FUNCTIONALITY  
# =============================================================================

@csrf_exempt
@require_http_methods(["POST"])
def bulk_update_student_status(request):
    """
    Bulk update status for multiple students with automatic curriculum assignment
    URL: /programme_curriculum/api/bulk_update_student_status/
    
    Payload: {
        "student_ids": [1, 2, 3],
        "reported_status": "REPORTED"
    }
    """
    try:
        data = json.loads(request.body)
        student_ids = data.get('student_ids', [])
        new_status = data.get('reported_status', '')
        
        if not student_ids:
            return JsonResponse({
                'success': False,
                'message': 'No student IDs provided'
            }, status=400)
        
        if new_status not in ['REPORTED', 'NOT_REPORTED', 'WITHDRAWAL', 'PENDING']:
            return JsonResponse({
                'success': False,
                'message': 'Invalid status. Must be REPORTED, NOT_REPORTED, WITHDRAWAL, or PENDING'
            }, status=400)
        
        results = []
        success_count = 0
        error_count = 0
        curriculum_assignments = {}
        
        for student_id in student_ids:
            try:
                # USE DATABASE TRANSACTION ISOLATION FOR EACH STUDENT
                from django.db import transaction
                with transaction.atomic():
                    # Call individual update_student_status function for each student in isolation
                    # This ensures curriculum assignment logic is applied consistently without interference
                    from django.test import RequestFactory
                    factory = RequestFactory()
                    individual_data = {'studentId': student_id, 'reportedStatus': new_status}
                    individual_request = factory.post('/api/update/', 
                                                    data=json.dumps(individual_data), 
                                                    content_type='application/json')
                    individual_request.user = request.user
                    
                    response = update_student_status(individual_request)
                    response_data = json.loads(response.content.decode('utf-8'))
                    
                    if response_data.get('success'):
                        success_count += 1
                        
                        # Track curriculum assignments for summary
                        if new_status == 'REPORTED':
                            transfer_message = response_data.get('data', {}).get('transfer_message', '')
                            if 'Curriculum:' in transfer_message:
                                # Extract curriculum info from transfer message
                                curriculum_part = transfer_message.split('Curriculum:')[1].split('|')[0].strip()
                                if curriculum_part not in curriculum_assignments:
                                    curriculum_assignments[curriculum_part] = 0
                                curriculum_assignments[curriculum_part] += 1
                        
                        results.append({
                            'student_id': student_id,
                            'status': 'success',
                            'message': response_data.get('message', 'Status updated successfully')
                        })
                    else:
                        error_count += 1
                        results.append({
                            'student_id': student_id,
                            'status': 'error',
                            'message': response_data.get('message', 'Unknown error')
                        })
                        
            except Exception as individual_error:
                error_count += 1
                results.append({
                    'student_id': student_id,
                    'status': 'error', 
                    'message': f'Processing error: {str(individual_error)}'
                })
        
        # Build summary message
        summary_message = f"Bulk status update completed: {success_count} successful, {error_count} failed"
        if curriculum_assignments:
            summary_message += f". Curriculum assignments: {dict(curriculum_assignments)}"
        
        return JsonResponse({
            'success': success_count > 0,
            'message': summary_message,
            'data': {
                'total_processed': len(student_ids),
                'success_count': success_count,
                'error_count': error_count,
                'new_status': new_status,
                'curriculum_assignments': curriculum_assignments,
                'individual_results': results
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid JSON data provided'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Bulk update failed: {str(e)}'
        }, status=500)


# =============================================================================
# STUDENT TRANSFER TO MAIN TABLES
# =============================================================================

@csrf_exempt
@require_http_methods(["POST"])
def bulk_transfer_students(request):
    """
    Bulk transfer all reported students to main academic tables with curriculum assignment
    ENHANCED: Now uses the same curriculum assignment logic as individual transfers
    """
    try:
        data = json.loads(request.body)
        year_filter = data.get('year')

        queryset = StudentBatchUpload.objects.filter(
            reported_status='REPORTED',
            user__isnull=True  # Not yet transferred
        )
        
        if year_filter:
            queryset = queryset.filter(year=year_filter)
        
        total_count = queryset.count()
        
        if total_count == 0:
            return JsonResponse({
                'success': True,
                'message': 'No students found to transfer',
                'data': {
                    'total_processed': 0,
                    'success_count': 0,
                    'error_count': 0,
                    'errors': []
                }
            })

        # Use the new bulk update function to transfer all at once with curriculum assignment
        student_ids = list(queryset.values_list('id', flat=True))
        
        from django.test import RequestFactory
        factory = RequestFactory()
        bulk_data = {'student_ids': student_ids, 'reported_status': 'REPORTED'}
        bulk_request = factory.post('/api/bulk_update/', 
                                  data=json.dumps(bulk_data), 
                                  content_type='application/json')
        bulk_request.user = request.user
        
        response = bulk_update_student_status(bulk_request)
        response_data = json.loads(response.content.decode('utf-8'))
        
        return JsonResponse({
            'success': response_data.get('success', False),
            'message': f"Bulk transfer completed with curriculum assignment: {response_data.get('data', {}).get('success_count', 0)} students transferred successfully",
            'data': response_data.get('data', {})
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Bulk transfer failed: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def manual_transfer_student(request):
    """
    Manually transfer a specific student to main academic tables
    """
    try:
        data = json.loads(request.body)
        student_id = data.get('studentId')
        
        if not student_id:
            return JsonResponse({
                'success': False,
                'message': 'Student ID is required'
            }, status=400)
        
        try:
            student = StudentBatchUpload.objects.get(id=student_id)
        except StudentBatchUpload.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Student not found'
            }, status=404)
        
        if student.reported_status != 'REPORTED':
            return JsonResponse({
                'success': False,
                'message': 'Student must have REPORTED status to be transferred'
            }, status=400)
        
        if student.user:
            return JsonResponse({
                'success': False,
                'message': 'Student already transferred to main tables'
            }, status=400)
        
        # Manual transfer not yet implemented - use status update API instead
        return JsonResponse({
            'success': False,
            'message': 'Manual transfer not yet implemented. Use the status update API to change student status to REPORTED for automatic transfer.',
            'data': {
                'student_id': student.id,
                'roll_number': student.roll_number,
                'suggestion': 'Use /programme_curriculum/api/update_student_status/ with status REPORTED'
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Transfer failed: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def check_transfer_status(request):
    """
    Check transfer status of students - how many are transferred vs pending
    """
    try:
        year = request.GET.get('year')
        
        queryset = StudentBatchUpload.objects.all()
        if year:
            queryset = queryset.filter(year=year)
        
        stats = {
            'total_students': queryset.count(),
            'reported_students': queryset.filter(reported_status='REPORTED').count(),
            'not_reported_students': queryset.filter(reported_status='NOT_REPORTED').count(),
            'transferred_students': queryset.filter(reported_status='REPORTED', user__isnull=False).count(),
            'pending_transfer': queryset.filter(reported_status='REPORTED', user__isnull=True).count(),
        }
        
        return JsonResponse({
            'success': True,
            'data': stats
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to get transfer status: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def get_batch_students(request, batch_id):
    """
    Get students for a specific batch - ONLY from StudentBatchUpload table
    This is the primary source for upcoming batches admin management.
    When students are marked as REPORTED, they get transferred to academic_information tables.
    URL: /programme_curriculum/api/batches/{batch_id}/students/
    """
    try:
        from applications.programme_curriculum.models import Batch

        try:
            batch = Batch.objects.get(id=batch_id)
        except Batch.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': f'Batch with ID {batch_id} not found'
            }, status=404)

        programme_type = 'ug' if batch.name.startswith('B.') else 'pg' if batch.name.startswith('M.') else 'phd'
        specialization = request.GET.get('specialization')
        discipline = request.GET.get('discipline')

        students = StudentBatchUpload.objects.filter(
            year=batch.year,
            programme_type=programme_type
        )
        
        # Apply disciplinary filtering based on programme type
        if programme_type == 'pg':
            if specialization:
                students = students.filter(specialization__icontains=specialization)
            else:

                discipline_name = batch.discipline.name
                discipline_filters = Q()

                discipline_filters |= Q(branch__icontains=discipline_name)

                if 'Engineering' in discipline_name:
                    typo_name = discipline_name.replace('Engineering', 'Enginnering')
                    discipline_filters |= Q(branch__icontains=typo_name)

                if 'Computer Science' in discipline_name:
                    discipline_filters |= Q(branch__icontains='CSE')
                    discipline_filters |= Q(branch__icontains='Computer Science')
                elif 'Electronics and Communication' in discipline_name:
                    discipline_filters |= Q(branch__icontains='ECE')
                    discipline_filters |= Q(branch__icontains='Electronics')
                elif 'Mechanical' in discipline_name:
                    discipline_filters |= Q(branch__icontains='ME')
                    discipline_filters |= Q(branch__icontains='Mechanical')
                
                students = students.filter(discipline_filters)
        else:
            discipline_name = batch.discipline.name
            discipline_filters = Q()
            
            discipline_filters |= Q(branch__icontains=discipline_name)

            if 'Engineering' in discipline_name:
                typo_name = discipline_name.replace('Engineering', 'Enginnering')
                discipline_filters |= Q(branch__icontains=typo_name)
            
            students = students.filter(discipline_filters)
        
        students = students.order_by('roll_number')

        upload_students = []
        for student in students:
            upload_students.append({
                'id': student.id,
                'name': student.name,
                'roll_number': student.roll_number,
                'institute_email': student.institute_email,
                'jee_app_no': student.jee_app_no,

                'father_name': student.father_name,
                'mother_name': getattr(student, 'mother_name', ''),
                'gender': getattr(student, 'gender', ''),
                'category': student.category,
                'pwd': getattr(student, 'pwd', ''),
                'minority': getattr(student, 'minority', ''),
                'date_of_birth': student.date_of_birth.isoformat() if student.date_of_birth else '',

                'phone_number': getattr(student, 'phone_number', ''),
                'address': getattr(student, 'address', ''),
                'state': getattr(student, 'state', ''),

                'branch': student.branch,
                'specialization': getattr(student, 'specialization', ''),
                'ai_rank': getattr(student, 'ai_rank', None),
                'category_rank': getattr(student, 'category_rank', None),
                'tenth_marks': getattr(student, 'tenth_marks', None),
                'twelfth_marks': getattr(student, 'twelfth_marks', None),

                'father_occupation': getattr(student, 'father_occupation', ''),
                'father_mobile': getattr(student, 'father_mobile', ''),
                'mother_occupation': getattr(student, 'mother_occupation', ''),
                'mother_mobile': getattr(student, 'mother_mobile', ''),
                'aadhar_number': getattr(student, 'aadhar_number', ''),

                'allotted_category': getattr(student, 'allotted_category', ''),
                'allotted_gender': getattr(student, 'allotted_gender', ''),
                'parent_email': getattr(student, 'parent_email', ''),
                'parentEmail': getattr(student, 'parent_email', ''),  # Camel case for frontend
                'alternateEmail': getattr(student, 'personal_email', ''),
                'country': getattr(student, 'country', ''),
                'nationality': getattr(student, 'nationality', ''),
                'blood_group': getattr(student, 'blood_group', ''),
                'bloodGroup': getattr(student, 'blood_group', ''),  # Camel case for frontend
                'blood_group_remarks': getattr(student, 'blood_group_remarks', ''),
                'bloodGroupRemarks': getattr(student, 'blood_group_remarks', ''),  # Camel case for frontend
                'pwd_category': getattr(student, 'pwd_category', ''),
                'pwdCategory': getattr(student, 'pwd_category', ''),  # Camel case for frontend
                'pwd_category_remarks': getattr(student, 'pwd_category_remarks', ''),
                'pwdCategoryRemarks': getattr(student, 'pwd_category_remarks', ''),  # Camel case for frontend
                'admission_mode': getattr(student, 'admission_mode', ''),
                'admissionMode': getattr(student, 'admission_mode', ''),  # Camel case for frontend
                'admission_mode_remarks': getattr(student, 'admission_mode_remarks', ''),
                'admissionModeRemarks': getattr(student, 'admission_mode_remarks', ''),  # Camel case for frontend
                'income_group': getattr(student, 'income_group', ''),
                'incomeGroup': getattr(student, 'income_group', ''),  # Camel case for frontend
                'income': getattr(student, 'income', ''),

                'year': student.year,
                'academic_year': getattr(student, 'academic_year', ''),
                'programme_type': getattr(student, 'programme_type', ''),
                'allocation_status': getattr(student, 'allocation_status', ''),
                'reported_status': student.reported_status,
                'status_display': 'Reported' if student.reported_status == 'REPORTED' else 'Not Reported',
                'created_at': student.created_at.isoformat() if hasattr(student, 'created_at') else '',
                'updated_at': student.updated_at.isoformat() if hasattr(student, 'updated_at') else '',
                'source': getattr(student, 'source', 'unknown')  # Track actual source from model
            })

        academic_students = []
        try:
            from applications.academic_information.models import Student as AcademicStudent
            from applications.globals.models import ExtraInfo

            academic_student_records = AcademicStudent.objects.filter(
                batch=batch.year,
                programme__icontains=programme_type.upper()
            )

            if batch.discipline:
                academic_student_records = academic_student_records.filter(
                    id__department__name__icontains=batch.discipline.name
                )
            
            for academic_student in academic_student_records:
                try:
                    assigned_batch = academic_student.batch_id
                    batch_info = {
                        'name': assigned_batch.name if assigned_batch else 'No batch assigned',
                        'curriculum': assigned_batch.curriculum.name if assigned_batch and assigned_batch.curriculum else 'No curriculum',
                        'id': assigned_batch.id if assigned_batch else None
                    }
                    academic_students.append({
                        'id': f"academic_{academic_student.id.id}",  # ExtraInfo ID
                        'name': f"{academic_student.id.user.first_name} {academic_student.id.user.last_name}".strip(),
                        'roll_number': academic_student.id.id,
                        'institute_email': academic_student.id.user.email,
                        'specialization': academic_student.specialization or '',
                        'batch_assigned': batch_info,
                        'programme': academic_student.programme,
                        'batch_year': academic_student.batch,
                        'cpi': academic_student.cpi,
                        'curr_semester_no': academic_student.curr_semester_no,
                        'reported_status': 'REPORTED',
                        'status_display': 'Reported (Academic System)',
                        'source': 'academic_system'
                    })
                except Exception as e:
                    pass
                    continue
                    
        except Exception as e:
            pass

        all_students = upload_students + academic_students
        
        response_data = {
            'success': True,
            'batch': {
                'id': batch.id,
                'name': batch.name,
                'discipline': batch.discipline.acronym,
                'year': batch.year,
                'curriculum': batch.curriculum.name if batch.curriculum else None
            },
            'students': all_students, 
            'total_students': len(all_students),
            'count': len(all_students),

            'uploadStudents': upload_students, 
            'upload_students': upload_students,  
            'academicStudents': academic_students,  
            'academic_students': academic_students,  
            'combinedBeforeDedup': all_students,
            'combined_before_dedup': all_students,  
            'totalStudentsAfterDedup': len(all_students), 
            'total_students_after_dedup': len(all_students), 
            'totalCount': len(all_students),  
            'total_count': len(all_students), 

            'data': all_students,
            'results': all_students,
            'items': all_students,
            'student_list': all_students,
            'studentList': all_students
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False,
            'message': f'Failed to fetch students: {str(e)}'
        }, status=500)


# =============================================================================
# BATCH SYNCHRONIZATION AND WORKFLOW ENFORCEMENT
# =============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_batches_unified(request):
    """
    UNIFIED API for both 'Batches' tab and 'Upcoming Batches' tab
    Returns the same data structure for consistent frontend handling
    URL: /programme_curriculum/api/admin_batches/ 
    """
    try:
        from applications.programme_curriculum.models import Batch

        from applications.globals.models import ExtraInfo, HoldsDesignation
        user_details = ExtraInfo.objects.get(user=request.user)
        des = HoldsDesignation.objects.all().filter(user=request.user).first()

        if des and des.designation.name not in ['acadadmin', 'studentacadadmin']:
            return JsonResponse({
                'success': False,
                'message': 'Access denied. Only academic admins can view batch data.',
                'batches': []
            }, status=403)

        all_batches = Batch.objects.filter(running_batch=True).select_related(
            'discipline', 'curriculum'
        ).order_by('-year', 'discipline__name')
        
        unified_batches = []
        
        for batch in all_batches:
            # Use centralized filled seats calculation function
            actual_filled = calculate_batch_filled_seats(batch)

            if hasattr(batch, 'total_seats') and batch.total_seats:
                total_seats = batch.total_seats
            else:
                default_seats = {
                    'CSE': 300,
                    'ECE': 120, 
                    'ME': 120,
                    'SM': 60,
                    'Des.': 80,
                    'Design': 80
                }
                total_seats = default_seats.get(batch.discipline.acronym, 100)
                if hasattr(batch, 'total_seats'):
                    batch.total_seats = total_seats
                    batch.save()
            
            available_seats = max(0, total_seats - actual_filled)

            students_data = []
            if request.GET.get('include_students') == 'true':
                try:
                    students = StudentBatchUpload.objects.filter(
                        year=batch.year,
                        branch__icontains=batch.discipline.name
                    ).order_by('roll_number')
                    
                    for student in students:
                        students_data.append({
                            'id': student.id,
                            'name': student.name,
                            'roll_number': student.roll_number,
                            'institute_email': student.institute_email,
                            'reported_status': student.reported_status,
                            'status_display': 'Reported' if student.reported_status == 'REPORTED' else 'Not Reported'
                        })
                except Exception as e:
                    students_data = []

            unified_batch = {
                'id': batch.id,
                'name': batch.name,  
                'programme': batch.name, 

                'discipline': batch.discipline.acronym,  
                'disciplineName': batch.discipline.name, 
                'disciplineId': batch.discipline.id,

                'year': batch.year,
                'batchYear': batch.year,
                'academicYear': f"{batch.year}-{str(batch.year + 1)[-2:]}",

                'curriculum': batch.curriculum.name if batch.curriculum else None,
                'curriculumId': batch.curriculum.id if batch.curriculum else None,
                'curriculum_id': batch.curriculum.id if batch.curriculum else None,
                'curriculumVersion': batch.curriculum.version if batch.curriculum else None,
                'curriculum_display': get_batch_curriculum_display(batch),
                'available_curriculums': [
                    {'id': c.id, 'name': c.name} for c in get_available_curriculums_for_batch(batch)
                ],
                'hasCurriculum': True if (batch.curriculum or len(get_available_curriculums_for_batch(batch)) > 0) else False,

                'totalSeats': total_seats,
                'total_seats': total_seats,
                'filledSeats': actual_filled // 2,
                'filled_seats': actual_filled,  
                'availableSeats': available_seats,
                'available_seats': available_seats,  
                'student_count': actual_filled // 2, 

                'running_batch': batch.running_batch,
                'status': 'READY' if (batch.curriculum or len(get_available_curriculums_for_batch(batch)) > 0) else 'NEEDS_CURRICULUM',
                'canAcceptStudents': True if (batch.curriculum or len(get_available_curriculums_for_batch(batch)) > 0) else False,

                'students': students_data,
                'studentCount': len(students_data),

                'displayBranch': batch.discipline.acronym,
                'fullDisciplineName': batch.discipline.name,
            }
            
            unified_batches.append(unified_batch)

        if request.GET.get('format') == 'nested':
            ug_batches = [b for b in unified_batches if b['name'].startswith('B.')]
            pg_batches = [b for b in unified_batches if b['name'].startswith('M.')]
            phd_batches = [b for b in unified_batches if b['name'] == 'PhD']
            
            return JsonResponse({
                'success': True,
                'data': {
                    'ug': ug_batches,
                    'pg': pg_batches,
                    'phd': phd_batches
                },
                'batches': unified_batches,
                'message': 'Batch data fetched successfully'
            })

        response_data = {
            'success': True,
            'batches': unified_batches,
            'total_batches': len(unified_batches),
            'message': 'Unified batch data fetched successfully'
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        import traceback
        
        return JsonResponse({
            'success': False,
            'message': f'Failed to fetch batch data: {str(e)}',
            'batches': [],
            'error_details': str(e)
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def sync_batch_data(request):
    """
    Sync total seats and curriculum between 'Batches' tab and 'Upcoming Batches' tab
    URL: /programme_curriculum/api/batches/sync/
    """
    try:
        from applications.programme_curriculum.models import Batch
        from applications.academic_information.models import Student

        all_batches = Batch.objects.filter(running_batch=True).select_related(
            'discipline', 'curriculum'
        ).order_by('year', 'discipline__name')
        
        sync_results = []
        
        for batch in all_batches:
            # Use centralized filled seats calculation function
            actual_filled = calculate_batch_filled_seats(batch)

            if hasattr(batch, 'total_seats') and batch.total_seats:
                available_seats = max(0, batch.total_seats - actual_filled)
            else:

                default_seats = {
                    'CSE': 300,
                    'ECE': 120, 
                    'ME': 80,
                    'SM': 80,
                    'Des.': 80,
                }
                batch.total_seats = default_seats.get(batch.discipline.acronym, 100)
                batch.save()
                available_seats = max(0, batch.total_seats - actual_filled)

            curriculum_display = get_batch_curriculum_display(batch)
            
            sync_results.append({
                'batch_id': batch.id,
                'name': batch.name,
                'discipline': batch.discipline.acronym,
                'discipline_name': batch.discipline.name,
                'year': batch.year,
                'total_seats': batch.total_seats,
                'filled_seats': actual_filled,
                'available_seats': available_seats,
                'curriculum': curriculum_display,
                'curriculum_display': curriculum_display,
                'curriculum_id': batch.curriculum.id if batch.curriculum else None,
                'status': 'READY'
            })
        
        return JsonResponse({
            'success': True,
            'message': 'Batch data synchronized successfully',
            'batches': sync_results,
            'total_batches': len(sync_results)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to sync batch data: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def validate_batch_prerequisites(request):
    """
    Validate that required batches exist before allowing student upload
    Enforces workflow: Batches must be created first, then students added
    URL: /programme_curriculum/api/batches/validate_prerequisites/
    """
    try:
        data = json.loads(request.body)
        academic_year = data.get('academic_year')  # e.g., 2025
        
        if not academic_year:
            return JsonResponse({
                'success': False,
                'message': 'Academic year is required'
            }, status=400)
        
        from applications.programme_curriculum.models import Batch, Discipline

        required_disciplines = ['Computer Science and Engineering', 'Electronics and Communication Engineering', 
                               'Mechanical Engineering', 'Smart Manufacturing', 'Design']
        
        missing_batches = []
        existing_batches = []
        
        for discipline_name in required_disciplines:
            try:
                discipline = Discipline.objects.filter(name=discipline_name).first()
                if discipline:
                    batch = Batch.objects.filter(
                        year=academic_year,
                        discipline=discipline,
                        running_batch=True
                    ).first()
                    
                    if batch:
                        existing_batches.append({
                            'discipline': discipline.name,
                            'acronym': discipline.acronym,
                            'batch_id': batch.id,
                            'curriculum_assigned': batch.curriculum is not None,
                            'curriculum_name': batch.curriculum.name if batch.curriculum else None,
                            'total_seats': getattr(batch, 'total_seats', 0)
                        })
                    else:
                        missing_batches.append({
                            'discipline': discipline.name,
                            'acronym': discipline.acronym,
                            'action_required': f'Create {discipline.acronym} batch for year {academic_year}'
                        })
            except Exception as e:
                pass

        can_upload_students = len(missing_batches) == 0
        
        return JsonResponse({
            'success': True,
            'can_upload_students': can_upload_students,
            'academic_year': academic_year,
            'existing_batches': existing_batches,
            'missing_batches': missing_batches,
            'message': 'All required batches exist' if can_upload_students else 'Some batches are missing',
            'workflow_status': 'READY' if can_upload_students else 'BATCHES_REQUIRED'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to validate prerequisites: {str(e)}'
        }, status=500)


# =============================================================================
# BATCH CURRICULUM STATUS CHECK
# =============================================================================

@csrf_exempt
@require_http_methods(["GET"])
def check_batches_curriculum_status(request):
    """
    Check curriculum assignment status for all batches
    Helps admins identify which batches need curriculum assignment
    URL: /programme_curriculum/api/batches/curriculum_status/
    """
    try:
        from applications.programme_curriculum.models import Batch

        batches = Batch.objects.filter(running_batch=True).select_related(
            'discipline', 'curriculum'
        ).order_by('-year', 'discipline__name')
        
        batch_status = []
        
        for batch in batches:
            student_count = StudentBatchUpload.objects.filter(
                year=batch.year,
                branch__icontains=batch.discipline.name
            ).count()

            reported_students = StudentBatchUpload.objects.filter(
                year=batch.year,
                branch__icontains=batch.discipline.name,
                reported_status='REPORTED'
            ).count()
            
            batch_info = {
                'batch_id': batch.id,
                'batch_name': batch.name,
                'discipline': batch.discipline.name,
                'discipline_acronym': batch.discipline.acronym,
                'year': batch.year,
                'total_seats': batch.total_seats,
                'student_count': student_count,
                'reported_students': reported_students,
                'has_curriculum': batch.curriculum is not None,
                'curriculum_name': batch.curriculum.name if batch.curriculum else None,
                'curriculum_id': batch.curriculum.id if batch.curriculum else None,
                'status': 'READY' if batch.curriculum else 'NEEDS_CURRICULUM',
                'warning': None if batch.curriculum else 'Curriculum assignment required before student upload/entry'
            }
            
            batch_status.append(batch_info)
        
        # Summary statistics
        summary = {
            'total_batches': len(batch_status),
            'batches_with_curriculum': len([b for b in batch_status if b['has_curriculum']]),
            'batches_without_curriculum': len([b for b in batch_status if not b['has_curriculum']]),
            'total_students': sum(b['student_count'] for b in batch_status),
            'reported_students': sum(b['reported_students'] for b in batch_status)
        }
        
        return JsonResponse({
            'success': True,
            'data': {
                'summary': summary,
                'batches': batch_status
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to check batch curriculum status: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def validate_student_upload_prerequisites(request):
    """
    Validate if all prerequisites are met for student data upload
    URL: /programme_curriculum/api/validate_student_upload_prerequisites/
    """
    try:
        academic_year_param = request.GET.get('academic_year')
        if academic_year_param:
            try:
                batch_year, academic_year = normalize_year_input(academic_year_param)
            except ValueError as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Invalid year format: {str(e)}'
                }, status=400)
        else:
            batch_year, academic_year = get_current_academic_year()
        
        from applications.programme_curriculum.models import Curriculum, Batch

        available_curriculums = Curriculum.objects.filter(working_curriculum=True)
        curriculum_check = {
            'exists': available_curriculums.exists(),
            'count': available_curriculums.count(),
            'status': 'PASS' if available_curriculums.exists() else 'FAIL',
            'message': f'Found {available_curriculums.count()} working curriculums' if available_curriculums.exists() 
                      else 'No working curriculums found - please create curriculums first'
        }

        existing_batches = Batch.objects.filter(year=batch_year, running_batch=True)
        batch_check = {
            'exists': existing_batches.exists(),
            'count': existing_batches.count(),
            'year': batch_year,
            'academic_year': academic_year,
            'status': 'PASS' if existing_batches.exists() else 'FAIL',
            'message': f'Found {existing_batches.count()} active batches for {academic_year}' if existing_batches.exists()
                      else f'No active batches found for {academic_year} - please create batches first'
        }

        batches_with_curriculum = existing_batches.filter(curriculum__isnull=False)
        batches_without_curriculum = existing_batches.filter(curriculum__isnull=True)
        
        curriculum_assignment_check = {
            'all_assigned': not batches_without_curriculum.exists(),
            'assigned_count': batches_with_curriculum.count(),
            'unassigned_count': batches_without_curriculum.count(),
            'status': 'PASS' if not batches_without_curriculum.exists() else 'FAIL',
            'message': f'All {batches_with_curriculum.count()} batches have curriculums assigned' if not batches_without_curriculum.exists()
                      else f'{batches_without_curriculum.count()} batches need curriculum assignment',
            'unassigned_batches': [f"{batch.name} {batch.discipline.acronym}" for batch in batches_without_curriculum]
        }
  
        overall_status = 'PASS' if (curriculum_check['status'] == 'PASS' and 
                                   batch_check['status'] == 'PASS' and 
                                   curriculum_assignment_check['status'] == 'PASS') else 'FAIL'
        
        validation_result = {
            'overall_status': overall_status,
            'can_upload_students': overall_status == 'PASS',
            'academic_year': academic_year,
            'batch_year': batch_year,
            'checks': {
                'curriculum_exists': curriculum_check,
                'batches_exist': batch_check,
                'curriculums_assigned': curriculum_assignment_check
            },
            'requirements': {
                'step_1': 'Create working curriculums',
                'step_2': 'Create batches for the academic year',
                'step_3': 'Assign curriculums to all batches',
                'step_4': 'Upload/enter student data'
            }
        }
        
        return JsonResponse({
            'success': True,
            'data': validation_result
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to validate prerequisites: {str(e)}'
        }, status=500)


# =============================================================================
# CURRICULUM REDUNDANCY CLEANUP UTILITIES
# =============================================================================

@csrf_exempt
@require_http_methods(["GET"])
def find_duplicate_curriculums(request):
    """
    Find duplicate curriculums that have the same name and programme
    URL: /programme_curriculum/api/find_duplicate_curriculums/
    """
    try:
        from applications.programme_curriculum.models import Curriculum
        from django.db.models import Count

        duplicates = (Curriculum.objects
                     .values('name', 'programme__name', 'programme__id')
                     .annotate(count=Count('id'))
                     .filter(count__gt=1, working_curriculum=True)
                     .order_by('name'))
        
        duplicate_details = []
        for dup in duplicates:
            curriculum_name = dup['name']
            programme_name = dup['programme__name']
            programme_id = dup['programme__id']
            count = dup['count']

            curricula = Curriculum.objects.filter(
                name=curriculum_name,
                programme__id=programme_id,
                working_curriculum=True
            ).order_by('version')
            
            curriculum_list = []
            for curr in curricula:
                batch_count = curr.batches.count()
                curriculum_list.append({
                    'id': curr.id,
                    'version': str(curr.version),
                    'latest_version': curr.latest_version,
                    'batch_count': batch_count,
                    'created_date': curr.id 
                })
            
            duplicate_details.append({
                'curriculum_name': curriculum_name,
                'programme_name': programme_name,
                'programme_id': programme_id,
                'duplicate_count': count,
                'curricula': curriculum_list
            })
        
        return JsonResponse({
            'success': True,
            'total_duplicate_groups': len(duplicate_details),
            'duplicates': duplicate_details
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to find duplicate curriculums: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def consolidate_duplicate_curriculums(request):
    """
    Consolidate duplicate curriculums by keeping the latest version and reassigning batches
    URL: /programme_curriculum/api/consolidate_duplicate_curriculums/
    """
    try:
        data = json.loads(request.body)
        curriculum_name = data.get('curriculum_name')
        programme_id = data.get('programme_id')
        keep_curriculum_id = data.get('keep_curriculum_id')
        
        if not all([curriculum_name, programme_id, keep_curriculum_id]):
            return JsonResponse({
                'success': False,
                'message': 'Missing required fields: curriculum_name, programme_id, keep_curriculum_id'
            }, status=400)
        
        from applications.programme_curriculum.models import Curriculum, Batch

        try:
            keep_curriculum = Curriculum.objects.get(id=keep_curriculum_id)
        except Curriculum.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': f'Curriculum with ID {keep_curriculum_id} not found'
            }, status=400)

        duplicate_curriculums = Curriculum.objects.filter(
            name=curriculum_name,
            programme__id=programme_id,
            working_curriculum=True
        ).exclude(id=keep_curriculum_id)

        reassigned_batches = 0
        for dup_curriculum in duplicate_curriculums:
            batches = dup_curriculum.batches.all()
            for batch in batches:
                batch.curriculum = keep_curriculum
                batch.save()
                reassigned_batches += 1

        duplicate_count = duplicate_curriculums.count()
        duplicate_curriculums.update(working_curriculum=False, latest_version=False)

        keep_curriculum.latest_version = True
        keep_curriculum.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Consolidated {duplicate_count} duplicate curriculums. Reassigned {reassigned_batches} batches to "{keep_curriculum.name} v{keep_curriculum.version}".',
            'kept_curriculum': {
                'id': keep_curriculum.id,
                'name': keep_curriculum.name,
                'version': str(keep_curriculum.version)
            },
            'removed_duplicates': duplicate_count,
            'reassigned_batches': reassigned_batches
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to consolidate curriculums: {str(e)}'
        }, status=500)


# =============================================================================
# BATCH REDUNDANCY CLEANUP UTILITIES
# =============================================================================

@csrf_exempt
@require_http_methods(["GET"])
def find_duplicate_batches(request):
    """
    Find duplicate batches that have the same programme+discipline+year combination
    URL: /programme_curriculum/api/find_duplicate_batches/
    """
    try:
        from applications.programme_curriculum.models import Batch
        from django.db.models import Count

        duplicates = (Batch.objects
                     .values('name', 'discipline__name', 'discipline__acronym', 'discipline__id', 'year')
                     .annotate(count=Count('id'))
                     .filter(count__gt=1, running_batch=True)
                     .order_by('year', 'discipline__name'))
        
        duplicate_details = []
        for dup in duplicates:
            programme_name = dup['name']
            discipline_name = dup['discipline__name']
            discipline_acronym = dup['discipline__acronym']
            discipline_id = dup['discipline__id']
            year = dup['year']
            count = dup['count']

            batches = Batch.objects.filter(
                name=programme_name,
                discipline__id=discipline_id,
                year=year,
                running_batch=True
            ).order_by('id')
            
            batch_list = []
            for batch in batches:
                try:
                    student_count = StudentBatchUpload.objects.filter(
                        year=batch.year,
                        branch__icontains=batch.discipline.name
                    ).count()
                except:
                    student_count = 0
                
                batch_list.append({
                    'id': batch.id,
                    'total_seats': batch.total_seats,
                    'student_count': student_count,
                    'curriculum': batch.curriculum.name if batch.curriculum else None,
                    'curriculum_id': batch.curriculum.id if batch.curriculum else None,
                    'created_order': batch.id 
                })
            
            duplicate_details.append({
                'programme_name': programme_name,
                'discipline_name': discipline_name,
                'discipline_acronym': discipline_acronym,
                'discipline_id': discipline_id,
                'year': year,
                'duplicate_count': count,
                'batches': batch_list
            })
        
        return JsonResponse({
            'success': True,
            'total_duplicate_groups': len(duplicate_details),
            'duplicates': duplicate_details
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to find duplicate batches: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def consolidate_duplicate_batches(request):
    """
    Consolidate duplicate batches by keeping one and reassigning students/data
    URL: /programme_curriculum/api/consolidate_duplicate_batches/
    """
    try:
        data = json.loads(request.body)
        programme_name = data.get('programme_name')
        discipline_id = data.get('discipline_id')
        year = data.get('year')
        keep_batch_id = data.get('keep_batch_id')
        
        if not all([programme_name, discipline_id, year, keep_batch_id]):
            return JsonResponse({
                'success': False,
                'message': 'Missing required fields: programme_name, discipline_id, year, keep_batch_id'
            }, status=400)
        
        from applications.programme_curriculum.models import Batch

        try:
            keep_batch = Batch.objects.get(id=keep_batch_id)
        except Batch.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': f'Batch with ID {keep_batch_id} not found'
            }, status=400)

        duplicate_batches = Batch.objects.filter(
            name=programme_name,
            discipline__id=discipline_id,
            year=year,
            running_batch=True
        ).exclude(id=keep_batch_id)

        reassigned_students = 0
        total_seats_added = 0
        
        for dup_batch in duplicate_batches:
            if dup_batch.total_seats > 0:
                keep_batch.total_seats += dup_batch.total_seats
                total_seats_added += dup_batch.total_seats

            if dup_batch.curriculum and not keep_batch.curriculum:
                keep_batch.curriculum = dup_batch.curriculum

            try:
                from applications.academic_information.models import Student
                students = Student.objects.filter(batch_id=dup_batch)
                for student in students:
                    student.batch_id = keep_batch
                    student.save()
                    reassigned_students += 1
            except:
                pass 
        
        keep_batch.save()

        duplicate_count = duplicate_batches.count()
        duplicate_batches.update(running_batch=False)
        
        return JsonResponse({
            'success': True,
            'message': f'Consolidated {duplicate_count} duplicate batches into "{keep_batch.name} {keep_batch.discipline.acronym} {keep_batch.year}". Added {total_seats_added} seats, reassigned {reassigned_students} students.',
            'kept_batch': {
                'id': keep_batch.id,
                'name': keep_batch.name,
                'discipline': keep_batch.discipline.acronym,
                'year': keep_batch.year,
                'total_seats': keep_batch.total_seats,
                'curriculum': keep_batch.curriculum.name if keep_batch.curriculum else None
            },
            'removed_duplicates': duplicate_count,
            'total_seats_added': total_seats_added,
            'reassigned_students': reassigned_students
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to consolidate batches: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def fix_stuck_reported_students(request):
    """
    Fix students that are stuck in REPORTED status and should be transferred to academic system
    """
    try:
        stuck_students = StudentBatchUpload.objects.filter(
            reported_status='REPORTED'
        ).exclude(
            Q(first_name__isnull=True) | Q(first_name='') |
            Q(last_name__isnull=True) | Q(last_name='') |
            Q(roll_number__isnull=True) | Q(roll_number='')
        )
        
        if not stuck_students.exists():
            return JsonResponse({
                'success': True,
                'message': 'No stuck students found',
                'data': {
                    'total_stuck': 0,
                    'fixed_count': 0,
                    'error_count': 0,
                    'errors': []
                }
            })
        
        fixed_count = 0
        errors = []
        
        with transaction.atomic():
            for student in stuck_students:
                try:
                    # Transfer student to academic system
                    transfer_result = transfer_student_to_academic_system(student)
                    
                    if transfer_result.get('success'):
                        # Update status to transferred
                        student.reported_status = 'TRANSFERRED'
                        student.save()
                        fixed_count += 1
                    else:
                        errors.append(f"Student {student.first_name} {student.last_name} ({student.roll_number}): {transfer_result.get('message', 'Unknown error')}")
                        
                except Exception as e:
                    errors.append(f"Student {student.first_name} {student.last_name} ({student.roll_number}): {str(e)}")
        
        return JsonResponse({
            'success': True,
            'message': f'Fixed {fixed_count} stuck students',
            'data': {
                'total_stuck': stuck_students.count(),
                'fixed_count': fixed_count,
                'error_count': len(errors),
                'errors': errors[:10]
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to fix stuck students: {str(e)}'
        }, status=500)


def transfer_student_to_academic_system(student):
    """
    Transfer a student from StudentBatchUpload to the academic_information.Student table
    This function handles the complete transfer process including User and ExtraInfo creation
    """
    try:
        from applications.academic_information.models import Student as AcademicStudent
        from applications.globals.models import ExtraInfo, Designation, HoldsDesignation
        from django.contrib.auth.models import User

        existing_academic_student = AcademicStudent.objects.filter(
            id__id=student.roll_number
        ).first()
        
        if existing_academic_student:
            return {
                'success': True,
                'message': 'Student already exists in academic system',
                'user_id': existing_academic_student.id.user.id,
                'extra_info_id': existing_academic_student.id.id,
                'academic_student_id': existing_academic_student.id.id
            }

        existing_user = User.objects.filter(username=student.roll_number).first()
        
        if existing_user:
            user = existing_user
        else:
            user = User.objects.create_user(
                username=student.roll_number,
                email=student.institute_email,
                first_name=student.first_name,
                last_name=student.last_name,
                password=student.password if student.password else 'changeme123'
            )

        existing_extra_info = ExtraInfo.objects.filter(id=student.roll_number).first()
        
        if existing_extra_info:
            extra_info = existing_extra_info
        else:
            extra_info = ExtraInfo.objects.create(
                id=student.roll_number,
                user=user,
                title='Mr' if student.gender == 'M' else 'Ms',
                sex='M' if student.gender == 'M' else 'F',
                date_of_birth=student.date_of_birth,
                user_type='student',
                address=student.permanent_address or '',
                phone_no=student.phone_number or '',
                department=student.batch.discipline if student.batch else None,
                profile_picture=None
            )

        academic_student = AcademicStudent.objects.create(
            id=extra_info,
            programme=student.programme,
            discipline=student.batch.discipline if student.batch else None,
            batch=student.batch,
            current_semester=calculate_current_semester(student.academic_year),
            current_year=student.academic_year,
            cpi=0.0,
            category=student.category or 'General',
            father_name=student.father_name or '',
            mother_name=student.mother_name or '',
            hall_no=0,
            room_no='',
            specialization=student.specialization or ''
        )

        student_designation = Designation.objects.filter(name='student').first()
        if student_designation:
            holds_designation, created = HoldsDesignation.objects.get_or_create(
                user=user,
                designation=student_designation,
                defaults={'working': user}
            )
        
        return {
            'success': True,
            'message': 'Student transferred successfully',
            'user_id': user.id,
            'extra_info_id': extra_info.id,
            'academic_student_id': academic_student.id.id
        }
        
    except Exception as e:
        import traceback
        return {
            'success': False,
            'message': f'Transfer failed: {str(e)}'
        }


@csrf_exempt
@require_http_methods(["POST"])
def sync_batches_to_configuration(request):
    """
    Sync batches from Batch model to BatchConfiguration model
    This ensures that all batches are available in the BatchConfiguration model
    """
    try:
        from applications.programme_curriculum.models import Batch
        
        synced_count = 0
        created_count = 0
        updated_count = 0
        
        for batch in Batch.objects.all():
            # Check if BatchConfiguration already exists with this ID
            try:
                batch_config = BatchConfiguration.objects.get(id=batch.id)
                # Update existing
                batch_config.programme = batch.name
                batch_config.discipline = batch.discipline.name if batch.discipline else "Unknown"
                batch_config.year = batch.year
                batch_config.total_seats = batch.total_seats
                batch_config.save()
                updated_count += 1
                
            except BatchConfiguration.DoesNotExist:
                # Create new
                batch_config = BatchConfiguration.objects.create(
                    programme=batch.name,
                    discipline=batch.discipline.name if batch.discipline else "Unknown",
                    year=batch.year,
                    total_seats=batch.total_seats
                )
                # Force the same ID to maintain consistency
                batch_config.id = batch.id
                batch_config.save()
                created_count += 1
            
            synced_count += 1
        
        return JsonResponse({
            'success': True,
            'message': f'Synced {synced_count} batches. Created: {created_count}, Updated: {updated_count}',
            'stats': {
                'total_synced': synced_count,
                'created': created_count,
                'updated': updated_count
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Sync failed: {str(e)}'
        }, status=500)


def ensure_default_curriculum_exists(discipline_obj, programme_name):
    """
    Ensure that a default curriculum exists for the given discipline and programme.
    This function creates a basic curriculum if none exists.
    
    Args:
        discipline_obj: Discipline instance
        programme_name: Programme name (e.g., 'M.Tech', 'B.Tech')
    
    Returns:
        curriculum_id: ID of the curriculum (existing or newly created), or None if failed
    """
    try:
        from applications.programme_curriculum.models import Curriculum, Programme
        
        # First, try to find an existing curriculum
        existing_curriculum = Curriculum.objects.filter(
            programme__name__icontains=programme_name,
            programme__discipline=discipline_obj,
            working_curriculum=True
        ).first()
        
        if existing_curriculum:
            return existing_curriculum.id
        
        # If no curriculum exists, create a default one
        # First, ensure the programme exists
        programme_category = 'PG' if programme_name in ['M.Tech', 'M.Des'] else ('UG' if programme_name in ['B.Tech', 'B.Des'] else 'PHD')
        
        programme, created = Programme.objects.get_or_create(
            name=programme_name,
            category=programme_category,
            defaults={'programme_begin_year': 2024}
        )
        
        # Add discipline to programme if not already associated
        if not discipline_obj.programmes.filter(id=programme.id).exists():
            discipline_obj.programmes.add(programme)
        
        # Create a default curriculum
        default_curriculum = Curriculum.objects.create(
            programme=programme,
            name=f"{programme_name} {discipline_obj.acronym} Default Curriculum",
            version=1.0,
            working_curriculum=True,
            no_of_semester=8 if programme_category in ['UG', 'PG'] else 12,
            min_credit=120 if programme_category == 'UG' else (60 if programme_category == 'PG' else 180)
        )
        
        # Log the creation for audit purposes
        import logging
        pass
        
        return default_curriculum.id
        
    except Exception as e:
        # Log the error but don't raise it to avoid breaking the student reporting process
        pass
        return None
