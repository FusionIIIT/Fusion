from rest_framework.views import APIView

import json

from django.http import JsonResponse

from rest_framework.decorators import api_view, permission_classes,authentication_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.authentication import TokenAuthentication
from rest_framework import status



from django.http import (Http404, HttpResponse, HttpResponseNotFound,
                         HttpResponseRedirect)
from django.shortcuts import render, reverse

from django.contrib import messages

from applications.globals.models import HoldsDesignation, Designation,ExtraInfo

from .models import Bank, Company, Payments, Paymentscheme, Receipts

from django.views.generic import View

from .render import Render

from datetime import datetime

import calendar
import openpyxl
from openpyxl import load_workbook
from django.contrib.auth.decorators import login_required
from Fusion.settings.common import LOGIN_URL, LOGIN_REDIRECT_URL


# making login mandatory to access this view
@login_required(login_url=LOGIN_URL)
def financeModule(request):
    """
    This function verifies the designation of the employee whether he/she is in the accounts department or not and if yes on the basis of designation he/she
    is redirected to the corresponding page

    @param:
    request - designation of the logged in user

    @variables:
    context - is used for pasing the information from the login page to the corresponding designated html page
    b - object of the context
    """

    context = {
    }
    k = HoldsDesignation.objects.select_related().filter(working=request.user)
    flag = 0
    for z in k:
        if(str(z.designation) == 'dealing assistant'):
            flag = 1
            b = Paymentscheme.objects.filter(view=True, senior_verify=False)
            context = {
                'b': b
            }
            return render(request, "financeAndAccountsModule/financeAndAccountsModuleds.html", context)

        if (str(z.designation) == 'adminstrator'):
            flag = 1
            return render(request, "financeAndAccountsModule/financeAndAccountsModulead.html", context)

        if (str(z.designation) == 'sr dealing assitant'):
            flag = 1
            b = Paymentscheme.objects.filter(
                senior_verify=True, view=True, ass_registrar_verify=False)
            context = {
                'b': b
            }
            return render(request, "financeAndAccountsModule/financeAndAccountsModulesrda.html", context)

        if (str(z.designation) == 'asst. registrar fa'):
            flag = 1
            b = Paymentscheme.objects.filter(
                ass_registrar_verify=True, view=True, ass_registrar_aud_verify=False)
            context = {
                'b': b
            }
            return render(request, "financeAndAccountsModule/financeAndAccountsModulearfa.html", context)

        if (str(z.designation) == 'asst. registrar aud'):
            flag = 1
            b = Paymentscheme.objects.filter(
                ass_registrar_aud_verify=True, view=True, registrar_director_verify=False)
            context = {
                'b': b
            }
            return render(request, "financeAndAccountsModule/finanaceAndAccountsModulearaud.html", context)

        if (str(z.designation) == 'Registrar'):
            flag = 1
            b = Paymentscheme.objects.filter(
                registrar_director_verify=True, view=True, )
            context = {
                'b': b
            }
            return render(request, "financeAndAccountsModule/financeAndAccountsModule.html", context)

        if (str(z.designation) == 'Director'):
            flag = 1
            b = Paymentscheme.objects.filter(
                registrar_director_verify=True, view=True)
            context = {
                'b': b
            }
            return render(request, "financeAndAccountsModule/financeAndAccountsModule.html", context) 
        if (str(z.designation) == 'student'):
            flag = 1
           
            return HttpResponse("You are not authorised to visit this page!!")
    if(flag == 0):
        return render(request, "financeAndAccountsModule/employee.html", context)


# making login mandatory to access this view
@login_required(login_url=LOGIN_URL)
def previewing(request):
    """
    This function allows the dealing assistant to access the Payment Scheme model to view the payroll of the individual employees

    @param
    Payment scheme model form the model.py is previewed for the creation of the payroll

    @variables
    Basic details of the employee with all the salary components are previewed in this function
    """
    k = HoldsDesignation.objects.select_related().filter(working=request.user)

    flag = 0
    for z in k:
        if(str(z.designation) == 'dealing assistant'):
            flag = 1
            month = request.POST.get("month")
            year = request.POST.get("number1")
            pf = request.POST.get("number2")
            name = request.POST.get("name")
            designation = request.POST.get("designation")
            pay = request.POST.get("number3")
            gr_pay = request.POST.get("number4")
            da = request.POST.get("number5")
            ta = request.POST.get("number6")
            hra = request.POST.get("number7")
            fpa = request.POST.get("number8")
            special_allow = request.POST.get("number9")
            nps = request.POST.get("number10")
            gpf = request.POST.get("number11")
            income_tax = request.POST.get("number12")
            p_tax = request.POST.get("number13")
            gslis = request.POST.get("number14")
            gis = request.POST.get("number15")
            license_fee = request.POST.get("number16")
            electricity_charges = request.POST.get("number17")
            others = request.POST.get("number18")
            gr_reduction = int(nps) + int(gpf) + int(income_tax) + int(p_tax) + int(
                gslis) + int(gis) + int(license_fee) + int(electricity_charges) + int(others)
            income = int(pay) + int(gr_pay) + int(da) + int(ta) + \
                int(hra) + int(fpa) + int(special_allow)
            net_payment = (income - gr_reduction)
            
            
            a = Paymentscheme(month=month, year=year, pf=pf, name=name, designation=designation, pay=pay, gr_pay=gr_pay, da=da, ta=ta, hra=hra, fpa=fpa, special_allow=special_allow, nps=nps, gpf=gpf,
                              income_tax=income_tax, p_tax=p_tax, gslis=gslis, gis=gis, license_fee=license_fee, electricity_charges=electricity_charges, others=others, gr_reduction=gr_reduction, net_payment=net_payment)
            a.save()
            context = {
            }
            return render(request, "financeAndAccountsModule/financeAndAccountsModuleds.html", context)
        # if (str(z.designation) == 'adminstrator'):
        #     flag = 1
        #     return render(request, "financeAndAccountsModule/financeAndAccountsModulead.html", context)

        if (str(z.designation) == 'sr dealing assitant'):
            flag = 1
            b = Paymentscheme.objects.filter(
                senior_verify=True, view=True, ass_registrar_verify=False)
            context = {
                'b': b
            }
            return render(request, "financeAndAccountsModule/financeAndAccountsModulesrda.html", context)

        if (str(z.designation) == 'asst. registrar fa'):
            flag = 1
            b = Paymentscheme.objects.filter(
                ass_registrar_verify=True, view=True, ass_registrar_aud_verify=False)
            context = {
                'b': b
            }
            return render(request, "financeAndAccountsModule/financeAndAccountsModulearfa.html", context)

        if (str(z.designation) == 'asst. registrar aud'):
            flag = 1
            b = Paymentscheme.objects.filter(
                ass_registrar_aud_verify=True, view=True, registrar_director_verify=False)
            context = {
                'b': b
            }
            return render(request, "financeAndAccountsModule/finanaceAndAccountsModulearaud.html", context)

        if (str(z.designation) == 'Registrar'):
            flag = 1
            b = Paymentscheme.objects.filter(
                registrar_director_verify=True, view=True)
            context = {
                'b': b
            }
            return render(request, "financeAndAccountsModule/financeAndAccountsModule.html", context)

        if (str(z.designation) == 'director'):
            flag = 1
            b = Paymentscheme.objects.filter(
                registrar_director_verify=True, view=True)
            context = {
                'b': b
            }
            return render(request, "financeAndAccountsModule/financeAndAccountsModule.html", context)
    # if(flag == 0):
    #     return render(request, "financeAndAccountsModule/employee.html", context)


# making login mandatory to access this view
@login_required(login_url=LOGIN_URL)
def verifying(request):
    """
    This function verification of the salary slips starting from the range of the dealing assistant to sr. dealing assistant to asst. registrar FA to
    asst. registrar aud to registrar/director

    @param
    request - trivial
officeOfRegistrar
    @variables
    a - object of the verifying method
    id - primary key of the payment scheme model
    """
    k = HoldsDesignation.objects.select_related().filter(working=request.user)

    for z in k:
        if request.method == "POST":
            if(str(z.designation) == 'dealing assistant'):
                a = request.POST.getlist('box')
                pay_scheme = []
                for i in range(len(a)):
                    if "verify" in request.POST:
                        p = Paymentscheme.objects.get(id=a[i])
                        p.senior_verify = True
                        pay_scheme.append(p)

                    if "delete" in request.POST:
                        p = Paymentscheme.objects.get(id=a[i])
                        p.senior_verify = False
                        pay_scheme.append(p)
                Paymentscheme.objects.bulk_update(
                    pay_scheme, ['senior_verify'])

            if(str(z.designation) == 'sr dealing assitant'):
                a = request.POST.getlist('box')
                sr_pay_scheme = []
                for i in range(len(a)):
                    if "verify" in request.POST:
                        p = Paymentscheme.objects.get(id=a[i])
                        p.ass_registrar_verify = True
                        sr_pay_scheme.append(p)

                    if "delete" in request.POST:
                        p = Paymentscheme.objects.get(id=a[i])
                        p.senior_verify = False
                        sr_pay_scheme.append(p)
                Paymentscheme.objects.bulk_update(
                    sr_pay_scheme, ['senior_verify', 'ass_registrar_verify'])

            if(str(z.designation) == 'asst. registrar fa'):
                a = request.POST.getlist('box')
                asst_pay_scheme = []
                for i in range(len(a)):
                    if "verify" in request.POST:
                        p = Paymentscheme.objects.get(id=a[i])
                        p.ass_registrar_aud_verify = True
                        asst_pay_scheme.append(p)

                    if "delete" in request.POST:
                        p = Paymentscheme.objects.get(id=a[i])
                        p.ass_registrar_verify = False
                        asst_pay_scheme.append(p)
                Paymentscheme.objects.bulk_update(
                    asst_pay_scheme, ['ass_registrar_verify', 'ass_registrar_aud_verify'])

            if(str(z.designation) == 'asst. registrar aud'):
                a = request.POST.getlist('box')
                aud_pay_scheme = []
                for i in range(len(a)):
                    if "verify" in request.POST:
                        p = Paymentscheme.objects.get(id=a[i])
                        p.registrar_director_verify = True
                        aud_pay_scheme.append(p)

                    if "delete" in request.POST:
                        p = Paymentscheme.objects.get(id=a[i])
                        p.ass_registrar_aud_verify = False
                        aud_pay_scheme.append(p)
                Paymentscheme.objects.bulk_update(
                    aud_pay_scheme, ['registrar_director_verify', 'ass_registrar_aud_verify'])

            if(str(z.designation) == 'Registrar'):
                a = request.POST.getlist('box')
                reg_pay_scheme = []
                for i in range(len(a)):
                    if "verify" in request.POST:
                        p = Paymentscheme.objects.get(id=a[i])
                        p.runpayroll = True
                        p.view = False
                        reg_pay_scheme.append(p)

                    if "delete" in request.POST:
                        p = Paymentscheme.objects.get(id=a[i])
                        p.registrar_director_verify = False
                        p.view = True
                        reg_pay_scheme.append(p)
                Paymentscheme.objects.bulk_update(
                    reg_pay_scheme, ['runpayroll', 'view', 'registrar_director_verify'])

            if(str(z.designation) == 'director'):
                a = request.POST.getlist('box')
                dir_pay_scheme = []
                for i in range(len(a)):
                    if "verify" in request.POST:
                        p = Paymentscheme.objects.get(id=a[i])
                        p.runpayroll = True
                        p.view = False
                        dir_pay_scheme.append(p)

                    if "delete" in request.POST:
                        p = Paymentscheme.objects.get(id=a[i])
                        p.registrar_director_verify = False
                        p.view = True
                        dir_pay_scheme.append(p)
                Paymentscheme.objects.bulk_update(
                    dir_pay_scheme, ['view', 'runpayroll', 'registrar_director_verify'])

    return HttpResponseRedirect("/finance/finance/")


# making login mandatory to access this view
@login_required(login_url=LOGIN_URL)
def previous(request):
    """
        This method is used to view the already executed payroll run the registrar, and this method will take the input of the month and year which generates
        the entire payments of the employee of particular month and year

        @param
        request - trivial

        @variables
        a,b - these are taken as inout variables of month and year respectively for the display of the corresponding month's payments
    """
    k = HoldsDesignation.objects.select_related().filter(working=request.user)

    for z in k:
        if request.method == "POST":
            if(str(z.designation) == 'dealing assistant'):
                a = request.POST.get('selectmonth')
                b = request.POST.get('selectyear')

                c = Paymentscheme.objects.filter(
                    month=a, year=b, runpayroll=True)
                context = {
                    'c': c,

                }
                return render(request, "financeAndAccountsModule/financeAndAccountsModuleds.html", context)

            if(str(z.designation) == 'sr dealing assitant'):
                a = request.POST.get('selectmonth')
                b = request.POST.get('selectyear')  

                c = Paymentscheme.objects.filter(
                    month=a, year=b, runpayroll=True)
                context = {
                    'c': c
                }
                return render(request, "financeAndAccountsModule/financeAndAccountsModulesrda.html", context)

            if(str(z.designation) == 'asst. registrar fa'):
                a = request.POST.get('selectmonth')
                b = request.POST.get('selectyear')

                c = Paymentscheme.objects.filter(
                    month=a, year=b, runpayroll=True)
                context = {
                    'c': c
                }
                return render(request, "financeAndAccountsModule/financeAndAccountsModulearfa.html", context)

            if(str(z.designation) == 'asst. registrar aud'):
                a = request.POST.get('selectmonth')
                b = request.POST.get('selectyear')

                c = Paymentscheme.objects.filter(
                    month=a, year=b, runpayroll=True)
                context = {
                    'c': c
                }
                return render(request, "financeAndAccountsModule/finanaceAndAccountsModulearaud.html", context)

            if(str(z.designation) == 'Registrar'):
                a = request.POST.get('selectmonth')
                b = request.POST.get('selectyear')

                c = Paymentscheme.objects.filter(
                    month=a, year=b, runpayroll=True)
                context = {
                    'c': c
                }
                return render(request, "financeAndAccountsModule/financeAndAccountsModule.html", context)

            if(str(z.designation) == 'director'):
                a = request.POST.get('selectmonth')
                b = request.POST.get('selectyear')

                c = Paymentscheme.objects.filter(
                    month=a, year=b, runpayroll=True)
                context = {
                    'c': c
                }
                return render(request, "financeAndAccountsModule/financeAndAccountsModule.html", context)
    return HttpResponseRedirect("/finance/finance/")

# making login mandatory to access this view


@login_required(login_url=LOGIN_URL)
def createPayments(request):
    """
        This method is used to create a new payement transaction by the Assistant Registrar(fa), Registrar, Assistant Registrar(aud).

        @param
        request - trivial

        @variables
        Basic details of a new payement transaction.
    """
    k = HoldsDesignation.objects.select_related().filter(working=request.user)

    for z in k:
        if(str(z.designation) == "asst. registrar fa"):
            t_id = request.POST.get("t_id")
            toWhom = request.POST.get("toWhom")
            fromWhom = request.POST.get("fromWhom")
            purpose = request.POST.get("purpose")
            date = request.POST.get("date")

            p = Payments(TransactionId=t_id, ToWhom=toWhom,
                         FromWhom=fromWhom, Purpose=purpose, Date=date)
            p.save()
            context = {
            }
            return render(request, "financeAndAccountsModule/financeAndAccountsModulearfa.html", context)

        if(str(z.designation) == "asst. registrar aud"):
            t_id = request.POST.get("t_id")
            toWhom = request.POST.get("toWhom")
            fromWhom = request.POST.get("fromWhom")
            purpose = request.POST.get("purpose")
            date = request.POST.get("date")

            p = Payments(TransactionId=t_id, ToWhom=toWhom,
                         FromWhom=fromWhom, Purpose=purpose, Date=date)
            p.save()
            context = {
            }
            return render(request, "financeAndAccountsModule/financeAndAccountsModulearaud.html", context)

        if(str(z.designation) == "Registrar"):
            t_id = request.POST.get("t_id")
            toWhom = request.POST.get("toWhom")
            fromWhom = request.POST.get("fromWhom")
            purpose = request.POST.get("purpose")
            date = request.POST.get("date")

            p = Payments(TransactionId=t_id, ToWhom=toWhom,
                         FromWhom=fromWhom, Purpose=purpose, Date=date)
            p.save()
            context = {
            }
            return render(request, "financeAndAccountsModule/financeAndAccountsModule.html", context)

        if(str(z.designation) == "director"):
            t_id = request.POST.get("t_id")
            toWhom = request.POST.get("toWhom")
            fromWhom = request.POST.get("fromWhom")
            purpose = request.POST.get("purpose")
            date = request.POST.get("date")

            p = Payments(TransactionId=t_id, ToWhom=toWhom,
                         FromWhom=fromWhom, Purpose=purpose, Date=date)
            p.save()
            context = {
            }
            return render(request, "financeAndAccountsModule/financeAndAccountsModule.html", context)

# Function to get month number
# making login mandatory to access this view


@login_required(login_url=LOGIN_URL)
def getMonth(i):
    month = {
        "january": 1,
        "february": 2,
        "march": 3,
        "april": 4,
        "may": 5,
        "june": 6,
        "july": 7,
        "august": 8,
        "september": 9,
        "october": 10,
        "november": 11,
        "december": 12
    }

    return month.get(i, -1)


# making login mandatory to access this view
@login_required(login_url=LOGIN_URL)
def previousPayments(request):
    """
        This method is used to view the already created payement transaction, and this method will take the input of the month and year which generates
        the entire payments of particular month and year, made by the concerned authorities.

        @param
        request - trivial

        @variables
        a,b - these are taken as inout variables of month and year respectively for the display of the corresponding month's payments.
    """
    k = HoldsDesignation.objects.select_related().filter(working=request.user)

    for z in k:
        if request.method == "POST":
            if(str(z.designation) == 'asst. registrar fa'):
                a = request.POST.get('selectmonth')
                b = request.POST.get('selectyear')
                month_no = getMonth(a)

                c = Payments.objects.filter(Date__month=month_no, Date__year=b)
                context = {
                    'c': c
                }
                return render(request, "financeAndAccountsModule/financeAndAccountsModulearfa.html", context)

            if(str(z.designation) == 'asst. registrar aud'):
                a = request.POST.get('selectmonth')
                b = request.POST.get('selectyear')
                month_no = getMonth(a)

                c = Payments.objects.filter(Date__month=month_no, Date__year=b)
                context = {
                    'c': c
                }
                return render(request, "financeAndAccountsModule/finanaceAndAccountsModulearaud.html", context)

            if(str(z.designation) == 'Registrar'):
                a = request.POST.get('selectmonth')
                b = request.POST.get('selectyear')
                month_no = getMonth(a)

                c = Payments.objects.filter(Date__month=month_no, Date__year=b)

                context = {
                    'c': c
                }
                return render(request, "financeAndAccountsModule/financeAndAccountsModule.html", context)

            if(str(z.designation) == 'director'):
                a = request.POST.get('selectmonth')
                b = request.POST.get('selectyear')
                month_no = getMonth(a)

                c = Payments.objects.filter(Date__month=month_no, Date__year=b)
                context = {
                    'c': c
                }
                return render(request, "financeAndAccountsModule/financeAndAccountsModule.html", context)
    return HttpResponseRedirect("/finance/finance/")

# making login mandatory to access this view


@login_required(login_url=LOGIN_URL)
def createReceipts(request):
    """
        This method is used to create a receipt of a new transaction by the Assistant Registrar(fa), Registrar, Assistant Registrar(aud).

        @param
        request - trivial

        @variables
        Basic details of a new payement transaction for generating receipts.
    """
    k = HoldsDesignation.objects.select_related().filter(working=request.user)

    for z in k:
        if(str(z.designation) == "asst. registrar fa"):
            t_id = request.POST.get("t_id")
            toWhom = request.POST.get("toWhom")
            fromWhom = request.POST.get("fromWhom")
            purpose = request.POST.get("purpose")
            date = request.POST.get("date")

            p = Receipts(TransactionId=t_id, ToWhom=toWhom,
                         FromWhom=fromWhom, Purpose=purpose, Date=date)
            p.save()
            context = {
            }
            return render(request, "financeAndAccountsModule/financeAndAccountsModulearfa.html", context)

        if(str(z.designation) == "asst. registrar aud"):
            t_id = request.POST.get("t_id")
            toWhom = request.POST.get("toWhom")
            fromWhom = request.POST.get("fromWhom")
            purpose = request.POST.get("purpose")
            date = request.POST.get("date")

            p = Receipts(TransactionId=t_id, ToWhom=toWhom,
                         FromWhom=fromWhom, Purpose=purpose, Date=date)
            p.save()
            context = {
            }
            return render(request, "financeAndAccountsModule/financeAndAccountsModulearaud.html", context)

        if(str(z.designation) == "Registrar"):
            t_id = request.POST.get("t_id")
            toWhom = request.POST.get("toWhom")
            fromWhom = request.POST.get("fromWhom")
            purpose = request.POST.get("purpose")
            date = request.POST.get("date")

            p = Receipts(TransactionId=t_id, ToWhom=toWhom,
                         FromWhom=fromWhom, Purpose=purpose, Date=date)
            p.save()
            context = {
            }
            return render(request, "financeAndAccountsModule/financeAndAccountsModule.html", context)

        if(str(z.designation) == "director"):
            t_id = request.POST.get("t_id")
            toWhom = request.POST.get("toWhom")
            fromWhom = request.POST.get("fromWhom")
            purpose = request.POST.get("purpose")
            date = request.POST.get("date")

            p = Receipts(TransactionId=t_id, ToWhom=toWhom,
                         FromWhom=fromWhom, Purpose=purpose, Date=date)
            p.save()
            context = {
            }
            return render(request, "financeAndAccountsModule/financeAndAccountsModule.html", context)

# making login mandatory to access this view


@login_required(login_url=LOGIN_URL)
def previousReceipts(request):
    """
        This method is used to view Receipts of already created payement transaction, and this method will take the input of the month and year which generates
        the entire payments of particular month and year, made by the concerned authorities.

        @param
        request - trivial

        @variables
        a,b - these are taken as inout variables of month and year respectively for the display of the corresponding month's payment recei.
    """
    k = HoldsDesignation.objects.select_related().filter(working=request.user)

    for z in k:
        if request.method == "POST":
            if(str(z.designation) == 'asst. registrar fa'):
                a = request.POST.get('selectmonth')
                b = request.POST.get('selectyear')
                month_no = getMonth(a)

                c = Receipts.objects.filter(Date__month=month_no, Date__year=b)
                context = {
                    'x': c
                }
                return render(request, "financeAndAccountsModule/financeAndAccountsModulearfa.html", context)

            if(str(z.designation) == 'asst. registrar aud'):
                a = request.POST.get('selectmonth')
                b = request.POST.get('selectyear')
                month_no = getMonth(a)

                c = Receipts.objects.filter(Date__month=month_no, Date__year=b)
                context = {
                    'x': c
                }
                return render(request, "financeAndAccountsModule/finanaceAndAccountsModulearaud.html", context)

            if(str(z.designation) == 'Registrar'):
                a = request.POST.get('selectmonth')
                b = request.POST.get('selectyear')
                month_no = getMonth(a)

                c = Receipts.objects.filter(Date__month=month_no, Date__year=b)
                context = {
                    'x': c
                }
                return render(request, "financeAndAccountsModule/financeAndAccountsModule.html", context)

            if(str(z.designation) == 'director'):
                a = request.POST.get('selectmonth')
                b = request.POST.get('selectyear')
                month_no = getMonth(a)

                c = Receipts.objects.filter(Date__month=month_no, Date__year=b)
                context = {
                    'x': c
                }
                return render(request, "financeAndAccountsModule/financeAndAccountsModule.html", context)
    return HttpResponseRedirect("/finance/finance/")

# making login mandatory to access this view


@login_required(login_url=LOGIN_URL)
def createBank(request):
    """
        This method is used to create a new Bank Branch by the Administrator.

        @param
        request - trivial

        @variables
        Basic details of a new bank Branch.
    """
    k = HoldsDesignation.objects.select_related().filter(
        working=request.user, designation=Designation.objects.get(name='adminstrator'))

    acc_no = request.POST.get("acc_no")
    bank_Name = request.POST.get("bank_name")
    IFSC_code = request.POST.get("IFSC_code")
    branch = request.POST.get("branch")

    p = Bank(Account_no=acc_no, Bank_Name=bank_Name,
             IFSC_Code=IFSC_code, Branch_Name=branch)
    p.save()
    context = {
    }
    return render(request, "financeAndAccountsModule/financeAndAccountsModulead.html", context)

# making login mandatory to access this view


@login_required(login_url=LOGIN_URL)
def createCompany(request):
    """
        This method is used to create a new Company by the Administrator.

        @param
        request - trivial

        @variables
        Basic details of a new Company.
    """
    k = HoldsDesignation.objects.select_related().filter(
        working=request.user, designation=Designation.objects.get(name='adminstrator'))

    c_name = request.POST.get("c_name")
    start_date = request.POST.get("start_date")
    end_date = request.POST.get("end_date") or None
    description = request.POST.get("description")
    status = request.POST.get("status")

    p = Company(Company_Name=c_name, Start_Date=start_date,
                End_Date=end_date, Description=description, Status=status)
    p.save()
    context = {
    }
    return render(request, "financeAndAccountsModule/financeAndAccountsModulead.html", context)

# def alterBank(request):
#     """
#         This method is used to update the details of Bank Branch by the Administrator.

#         @param
#         request - trivial

#         @variables
#         Basic details of a new bank Branch.
#     """
#     k = HoldsDesignation.objects.filter(working = request.user, designation = Designation.objects.get(name = 'adminstrator'))

#     acc_no = request.POST.get("acc_no")
#     bank_Name = request.POST.get("bank_name")
#     IFSC_code = request.POST.get("IFSC_code")
#     branch = request.POST.get("branch")

#     if request.method == "POST" :
#         a = request.POST.getlist('box')
#         print(a)
#         for i in range(len(a)) :
#             if "update" in request.POST :
#                 p = Bank.objects.get(bank_id = a[i])
#                 if(p.Account_no == acc_no and p.Bank_Name == bank_Name):
#                     p.Account_no = acc_no
#                     p.Bank_Name = bank_Name
#                     p.IFSC_Code = IFSC_code
#                     p.Branch = branch
#                     p.save()
#                 else:
#                     messeges.success(request, "Please Add the curresponding bank first!")
#     context = {
#     }
#     return render(request, "financeAndAccountsModule/financeAndAccountsModulead.html")

# def alterCompany(request):
#     """
#         This method is used to update the details of Bank Branch by the Administrator.

#         @param
#         request - trivial

#         @variables
#         Basic details of a new bank Branch.
#     """
#     k = HoldsDesignation.objects.filter(working = request.user, designation = Designation.objects.get(name = 'adminstrator'))

#     c_name = request.POST.get("c_name")
#     start_date = request.POST.get("start_date")
#     end_date = request.POST.get("end_date")
#     description = request.POST.get("description")
#     status = request.POST.get("status")

#     a = request.POST.getlist('box')
#     for i in range(len(a)) :
#         if "update" in request.POST :
#             p = Company.objects.get(id = a[i])
#             p.Company_Name = c_name
#             p.Start_Date = start_date
#             p.End_Date = end_date
#             p.Description = description
#             p.Status = status

#             p.save()

#     context = {
#     }
#     return render(request, "financeAndAccountsModule/financeAndAccountsModulead.html")

# making login mandatory to access this view


@login_required(login_url=LOGIN_URL)
def printSalary(request):
    """
      This method is used to print the salary details of an employee by the Administrator.

      @param
      request - trivial

      @variables
      Basic details of an employee's salary including basic pay, ta, da, hra, nps etc.

    """
    user=request.user
    k = HoldsDesignation.objects.select_related().filter(
        working=request.user, designation=Designation.objects.get(name='adminstrator'))

    month = request.POST.get("month")
    year = request.POST.get("year")
    runpayroll=True
    c = Paymentscheme.objects.filter(month=month, year=year)
    context = {
        'c': c,
    }

    return Render.render('financeAndAccountsModule/payroll_content4.html', context)


@login_required(login_url=LOGIN_URL)
def previewing_file(request):
    if request.method == 'POST':
       
            # Load the Excel file
      workbook = load_workbook(request.FILES['file'])
      worksheet = workbook.active
        # Loop through each row in the worksheet and create a new model instance
      for row in worksheet.iter_rows(min_row=2, values_only=True):
            month=row[0]
            year=row[1]
            pf=row[2]
            name=row[3]
            designation=row[4]
            pay=row[5]
            gr_pay=row[6] 
            da=row[7]
            ta =row[8]
            hra=row[9]
            fpa=row[10]
            special_allow=row[11]
            nps=row[12]
            gpf=row[13]
            income_tax=row[14]
            p_tax=row[15]
            gslis=row[16]
            gis=row[17]
            license_fee=row[18]
            electricity_charges =row[19]
            others=row[20]
            gr_reduction = int(row[12])+int(row[13])+int(row[14])+int(row[15])+int(row[16])+ int(row[17])+int(row[18])+int(row[19])+int(row[20])
            income=int(row[5])+int(row[6])+int(row[7])+int(row[8]) + int(row[9])+int(row[10])+int(row[11])
            net_payment =int(income) -int(gr_reduction) 
          
            
            a = Paymentscheme(month=month, year=year, pf=pf, name=name, designation=designation, pay=pay, gr_pay=gr_pay, da=da, ta=ta, hra=hra, fpa=fpa, special_allow=special_allow, nps=nps, gpf=gpf,
                              income_tax=income_tax, p_tax=p_tax, gslis=gslis, gis=gis, license_fee=license_fee, electricity_charges=electricity_charges, others=others, gr_reduction=gr_reduction, net_payment=net_payment)
        
            a.save()
         
            
            
         
    return render(request, "financeAndAccountsModule/financeAndAccountsModuleds.html")
  
  


#API 




@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
class PaymentschemeApi(APIView):
    def get(self, request):
        Paymentscheme_obj = Paymentscheme.objects.all();
        serialized_obj = PaymentschemeSerializer(serialized_obj, many=True)
        return Response({'status':200, 'payload':serialized_obj.data})

    def post(self, request):
        request_body = json.loads(request.body)
        month = request_body.get('month')
        year = request_body.get('year')
        pf = request_body.get('pf')
        name = request_body.get('name')
        designation =request_body.get('designation')
        pay = request_body.get('pay')
        gr_pay = request_body.get('gr_pay')
        da = request_body.get('da')
        ta = request_body.get('ta')
        hra = request_body.get('hra')
        fpa = request_body.get('fpa')
        special_allow = request_body.get('special_allow')
        nps = request_body.get('nps')
        gpf = request_body.get('gpf')
        income_tax = request_body.get('income_tax')
        p_tax = request_body.get('p_tax')
        gslis = request_body.get('gslis')
        gis = request_body.get('gis')
        license_fee = request_body.get('license_fee')
        electricity_charges = request_body.get('electricity_charges')
        others = request_body.get('others')
        gr_reduction = request_body.get('gr_reduction')
        net_payment = request_body.get('net_payment')
        senior_verify = request_body.get('senior_verify')
        ass_registrar_verify = request_body.get('ass_registrar_verify')
        ass_registrar_aud_verify = request_body.get('ass_registrar_aud_verify')
        registrar_director_verify = request_body.get('registrar_director_verify')
        runpayroll = request_body.get('runpayroll')
        view = request_body.get('view')

        query = Paymentscheme.objects.all()
        if month:
            query = query.filter(month=month)
        if year:
            query = query.filter(year=year)
        if pf:
            query = query.filter(pf=pf)
        if name:
            query = query.filter(name=name)
        if designation:
            query = query.filter(designation=designation)
        if pay:
            query = query.filter(pay=pay)
        if gr_pay:
            query = query.filter(gr_pay=gr_pay)
        if da:
            query = query.filter(da=da)
        if ta:
            query = query.filter(ta=ta)
        if hra:
            query = query.filter(hra=hra)
        if fpa:
            query = query.filter(fpa=fpa)
        if special_allow:
            query = query.filter(special_allow=special_allow)
        if nps:
            query = query.filter(nps=nps)
        if gpf:
            query = query.filter(gpf=gpf)
        if income_tax:
            query = query.filter(income_tax=income_tax)
        if p_tax:
            query = query.filter(p_tax=p_tax)
        if gslis:
            query = query.filter(gslis=gslis)
        if gis:
            query = query.filter(gis=gis)
        if  license_fee:
            query = query.filter( license_fee= license_fee)
        if electricity_charges:
            query = query.filter(electricity_charges=electricity_charges)
        if others:
            query = query.filter(others=others)
        if gr_reduction:
            query = query.filter(gr_reduction=gr_reduction)
        if net_payment:
            query = query.filter(net_payment=net_payment)
        if senior_verify:
            query = query.filter(senior_verify=senior_verify)
        if ass_registrar_verify:
            query = query.filter(ass_registrar_verify=ass_registrar_verify)
        if ass_registrar_aud_verify:
            query = query.filter(ass_registrar_aud_verify=ass_registrar_aud_verify)
        if registrar_director_verify:
            query = query.filter(registrar_director_verify=registrar_director_verify)
        if runpayroll:
            query = query.filter( runpayroll= runpayroll)
        if view:
            query = query.filter(view=view)
       
        results = list(query.values())
        return JsonResponse({'results': results})

@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
class ReceiptsApi(APIView):
    def get(self, request):
        receipts_obj = Receipts.objects.all();
        serialized_obj = ReceiptsSerializer(receipts_obj, many=True)
        return Response({'status':200, 'payload':serialized_obj.data})
    def post(self, request):
        request_body = json.loads(request.body)
        receipt_id = request_body.get('receipt_id')
        transaction_id = request_body.get('transaction_id')
        to_whom = request_body.get('to_whom')
        from_whom = request_body.get('from_whom')
        purpose = request_body.get('purpose')
        date = request_body.get('date')

        query = Receipts.objects.all()
        if receipt_id:
            query = query.filter(receipt_id=receipt_id)
        if transaction_id:
            query = query.filter(TransactionId=transaction_id)
        if to_whom:
            query = query.filter(ToWhom__icontains=to_whom)
        if from_whom:
            query = query.filter(FromWhom__icontains=from_whom)
        if purpose:
            query = query.filter(Purpose__icontains=purpose)
        if date:
            query = query.filter(Date=date)

        results = list(query.values())
        return JsonResponse({'results': results})


@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
class PaymentsApi(APIView):
    def get(self, request):
        receipts_obj = Receipts.objects.all();
        serialized_obj = ReceiptsSerializer(receipts_obj, many=True)
        return Response({'status':200, 'payload':serialized_obj.data})

    def post(self, request):
        request_body = json.loads(request.body)
        paymentID= request_body.get('payment_id')
        transaction_id = request_body.get('TransactionId')
        to_whom = request_body.get('ToWhom')
        from_whom = request_body.get('FromWhom')
        purpose = request_body.get('Purpose')
        date = request_body.get('Date')

        query = Payments.objects.all()
        if paymentID:
            query = query.filter(payment_id=paymentID)
        if transaction_id:
            query = query.filter(TransactionId=transaction_id)
        if to_whom:
            query = query.filter(ToWhom__icontains=to_whom)
        if from_whom:
            query = query.filter(FromWhom__icontains=from_whom)
        if purpose:
            query = query.filter(Purpose__icontains=purpose)
        if date:
            query = query.filter(Date=date)
        
        results = list(query.values())
        return JsonResponse({'results': results})

@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
class BankApi(APIView):
    def get(self, request):
        receipts_obj = Receipts.objects.all();
        serialized_obj = ReceiptsSerializer(receipts_obj, many=True)
        return Response({'status':200, 'payload':serialized_obj.data})

    def post(self, request):
        request_body = json.loads(request.body)
        bank_id= request_body.get('bank_id')
        Account_no = request_body.get('Account_no')
        Bank_Name = request_body.get('Bank_Name')
        IFSC_Code = request_body.get('IFSC_Code')
        Branch_Name = request_body.get('Branch_Name')

        query = Bank.objects.all()
        if bank_id:
            query = query.filter(bank_id=bank_id)
        if Account_no:
            query = query.filter(Account_no=Account_no)
        if Bank_Name:
            query = query.filter(Bank_Name=Bank_Name)
        if IFSC_Code:
            query = query.filter(IFSC_Code=IFSC_Code)
        if Branch_Name:
            query = query.filter(Branch_Name=Branch_Name)

        bank_data = list(query.values())
        return JsonResponse({'banks': bank_data})

@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
class CompanyApi(APIView):
    def get(self, request):
        Company_obj = Company.objects.all();
        Company_obj = CompanySerializer(Company_obj, many=True)
        return Response({'status':200, 'payload':serialized_obj.data})

    def post(self, request):
        request_body = json.loads(request.body)
        company_id= request_body.get('company_id')
        Company_Name = request_body.get('Company_Name')
        Start_Date = request_body.get('Start_Date')
        End_Date = request_body.get('End_Date')
        Description = request_body.get('Description')
        Status = request_body.get('Status')

        companies = Company.objects.all()
        if company_id:
            companies = companies.filter(company_id=company_id)
        if Company_Name:
            companies = companies.filter(company_name=Company_Name)
        if Start_Date:
            companies = companies.filter(Start_Date=Start_Date)
        if End_Date:
            companies = companies.filter(End_date=End_Date)
        if Description:
            companies = companies.filter(Description=Description)
        if Status:
            companies = companies.filter(Status=Status)
        
        results = list(companies.values())
        return JsonResponse({'results': results})


