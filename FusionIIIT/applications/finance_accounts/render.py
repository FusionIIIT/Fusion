from io import BytesIO
from django.http import HttpResponse
from django.template.loader import get_template
import xhtml2pdf.pisa as pisa 
import openpyxl
from bs4 import BeautifulSoup


class Render:

    # @staticmethod
    # def render(path: str, params: dict):
    #     template = get_template(path)
    #     html = template.render(params)
    #     # hello this is my new commit

    #     response = HttpResponse(content_type='application/pdf')
    #     response['Content-Disposition'] = 'attachment; filename="salary.pdf"'
    #     # create a pdf
    #     pisa_status = pisa.CreatePDF(
    #         html, dest=response)
    #     # if error then show some funy view
    #     if pisa_status.err:
    #         return HttpResponse('We had some errors <pre>' + html + '</pre>')
    #     return response


    @staticmethod
    def render(path: str, params: dict):
    # Load the HTML file and extract the table data using BeautifulSoup
        template = get_template(path)
        f = template.render(params)
        # print(f)
        soup = BeautifulSoup(f, 'html.parser')
        table = soup.find('table')
        headers = [th.text.strip() for th in table.find_all('th')]
        rows = table.find_all('tr')
        data = []
        for row in rows:
            cols = row.find_all('td')
            cols = [col.text.strip() for col in cols]
            data.append(cols)
    
    # Create the Excel workbook and worksheet objects
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        # Write the headers to the worksheet
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
        
        # Write the table data to the worksheet
        for row_num, row_data in enumerate(data, 1):
            for col_num, col_data in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = col_data
 
        
            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename="salary.xlsx"'
            workbook.save(response)
        return response
