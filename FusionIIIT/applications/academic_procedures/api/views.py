import datetime
import random
import logging
import traceback
import re
from collections import defaultdict, deque, OrderedDict
from functools import wraps
from datetime import date
from django.utils import timezone
from django.conf import settings
logger = logging.getLogger(__name__)
from django.contrib.auth import get_user_model
from django.shortcuts import get_object_or_404, redirect
from django.db import transaction, IntegrityError
from django.db.models import Prefetch
from django.db.models.functions import Concat,ExtractYear,ExtractMonth,ExtractDay,Cast
from django.db.models import Max,Value,IntegerField,CharField,F,Sum, Case, When, Count
from io import BytesIO
import json
import xlrd
from xlsxwriter.workbook import Workbook
from django.db.models import Prefetch
import json
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponse
from rest_framework.permissions import IsAuthenticated
from rest_framework.authentication import TokenAuthentication
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes,authentication_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
import pandas as pd
from rest_framework.decorators import (
    api_view, parser_classes, permission_classes
)
from rest_framework.parsers    import MultiPartParser, FormParser

from applications.globals.models import Faculty, HoldsDesignation, Designation, ExtraInfo
from applications.globals.decorators import role_required
from applications.globals.programme_scope import (
    ALL_ACAD_ROLES,
    batch_in_scope,
    scope_batches,
    scope_students,
    scope_via_student,
    scoped_ids,
    scopes_for,
    student_in_scope,
)
from notifications.signals import notify
from applications.programme_curriculum.models import ( CourseInstructor, CourseSlot, Course as Courses, Batch, Discipline, Semester)
# from applications.programme_curriculum.models import Course

from applications.academic_procedures.models import ( MTechGraduateSeminarReport, PhDProgressExamination, Student, Curriculum , ThesisTopicProcess, InitialRegistrations,
                                                     FinalRegistration, SemesterMarks,backlog_course,
                                                     BranchChange , StudentRegistrationChecks, Semester , FeePayments , course_registration, course_replacement, AssistantshipClaim, Assignment, StipendRequest, CourseReplacementRequest, SwayamReplacementRequest, CourseDropRequest, CourseAddRequest, BatchChangeHistory, FeedbackQuestion, FeedbackResponse, FeedbackFilled, FeedbackOption, PhDCourseRegistrationRequest)

from applications.academic_information.models import (Curriculum_Instructor , Calendar)
from applications.online_cms.models import Student_grades

from applications.academic_procedures.views import (get_user_semester, get_acad_year,
                                                    get_currently_registered_courses,
                                                    get_current_credits, get_branch_courses,
                                                    Constants, get_faculty_list,
                                                    get_registration_courses, get_add_course_options,
                                                    get_final_registration_eligibility,
                                                    get_final_registration_window,
                                                    get_add_or_drop_course_date_eligibility,
                                                    get_detailed_sem_courses,
                                                    InitialRegistration)

from applications.academic_procedures.views import get_sem_courses, get_student_registrtion_check, get_cpi, academics_module_notif, get_final_registration_choices, get_currently_registered_course, get_add_course_options, get_drop_course_options, get_replace_course_options
from applications.examination.api.views import parse_academic_year, calculate_cpi_for_student

from . import serializers

User = get_user_model()

date_time = datetime.datetime.now()

def make_label(no: int, sem_type: str) -> str:
    """
    - odd → "Semester <no>"
    - even & Even Semester → "Semester <no>"
    - even & Summer Semester → "Summer <no//2>"
    """
    if no % 2 == 1:
        return f"Semester {no}"
    if sem_type == "Summer Semester":
        return f"Summer {no // 2}"
    return f"Semester {no}"


def get_semester_type(semester):
    """
    Returns the semester type string.
    """
    if semester % 2 == 1:
        return "Odd Semester"
    elif semester % 2 == 0:
        return "Even Semester"
    else:
        return "Summer Semester"


def generate_current_session(current_year, semester) :
    """
    Returns a tuple of (session, semester_type).
    """
    semester_type = get_semester_type(semester)

    if semester_type == "Odd Semester":
        session = f"{current_year}-{str(current_year + 1)[-2:]}"
    else:  # Even or Summer
        session = f"{current_year - 1}-{str(current_year)[-2:]}"
    
    return session, semester_type


def generate_next_session(current_year, next_semester) :
    """
    Returns a tuple of (session, semester_type) for student's next semester.
    """
    semester_type = get_semester_type(next_semester)

    if semester_type == "Odd Semester" or semester_type == "Even Semester":
        session = f"{current_year}-{str(current_year + 1)[-2:]}"
    else:
        session = f"{current_year - 1}-{str(current_year)[-2:]}"
    
    return session, semester_type





#--------------------------------------- APIs of student----------------------------------------------------------

demo_date = timezone.now()

@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def get_all_courses(request):
    try:
        obj = Courses.objects.all()
        serializer = serializers.CourseSerializer(obj, many=True).data
        
        return Response(serializer, status=status.HTTP_200_OK)
    except Exception as e:
        return Response(data = str(e) , status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# @api_view(['POST'])
# @role_required(['acadadmin'])
# def gen_roll_list(request):
#     try:
#         batch = request.data['batch']
#         course_id = request.data['course']
#         course = Courses.objects.get(id = course_id)
#         #obj = course_registration.objects.all().filter(course_id = course)
#         obj=course_registration.objects.filter(course_id__id=course_id, student_id__batch=batch).select_related(
#         'student_id__id__user','student_id__id__department').only('student_id__batch', 
#         'student_id__id__user__first_name', 'student_id__id__user__last_name',
#         'student_id__id__department__name','student_id__id__user__username')
#     except Exception as e:
#         batch=""
#         course=""
#         obj=""
#     students = []
#     for i in obj:
#         students.append({"rollno":i.student_id.id.user.username, 
#         "name":i.student_id.id.user.first_name+" "+i.student_id.id.user.last_name, 
#         "department":i.student_id.id.department.name})
#     # {'students': students, 'batch':batch, 'course':course_id}
#     return JsonResponse({'students': students, 'batch':batch, 'course':course_id}, status=200)


# api for student for adding courses  
# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def add_course(request):
#     try:
#         current_user = request.user
#         current_user = ExtraInfo.objects.all().filter(user=current_user).first()
#         current_user = Student.objects.all().filter(id=current_user.id).first()

#         sem_id_instance = Semester.objects.get(id = request.data['semester'])
        
#         count = request.data['ct']
#         count = int(count)
#         reg_curr = []

#         for i in range(1, count+1):
#             choice = "choice["+str(i)+"]"
#             slot = "slot["+str(i)+"]"
#             try:
#                 course_id_instance = Courses.objects.get(id = request.data[choice])
#                 courseslot_id_instance = CourseSlot.objects.get(id = request.data[slot])
                
#                 print(courseslot_id_instance.max_registration_limit)
#                 if course_registration.objects.filter(working_year = current_user.batch_id.year, course_id = course_id_instance).count() < courseslot_id_instance.max_registration_limit and (course_registration.objects.filter(course_id=course_id_instance, student_id=current_user).count() == 0):
#                     print("space left = True")
#                     p = course_registration(
#                         course_id=course_id_instance,
#                         student_id=current_user,
#                         course_slot_id=courseslot_id_instance,
#                         semester_id=sem_id_instance
#                     )
#                     print(serializers.course_registration(p))
#                     if p not in reg_curr:
#                         reg_curr.append(p)
#                     else:
#                         print("already exist")
#             except Exception as e:
#                 error_message = str(e) 
#                 resp = {'message': 'Course addition failed', 'error': error_message}
#                 return Response(resp, status=status.HTTP_400_BAD_REQUEST)
#         print(reg_curr)
#         course_registration_data = course_registration.objects.bulk_create(reg_curr)
#         course_registration_data = serializers.CourseRegistrationSerializer(course_registration_data , many = True).data
#         res = {'message' : 'Courses successfully added' , "courses_added" : course_registration_data }
#         return Response(data = res , status = status.HTTP_200_OK)
#     except Exception as e:
#         print(e)
#         return Response(data = str(e) , status= status.HTTP_500_INTERNAL_SERVER_ERROR)

def block_pg_phd(view_func):
    """403 for PG/PhD students; these UG-only flows are hidden from them in the UI."""
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        try:
            student = Student.objects.select_related('batch_id__curriculum').get(
                id__user=request.user
            )
        except Student.DoesNotExist:
            return JsonResponse({'error': 'Student record not found'}, status=404)
        if _is_phd_student(student):
            return JsonResponse(
                {'error': 'This section is not available for PG/PhD students'},
                status=403,
            )
        return view_func(request, *args, **kwargs)
    return _wrapped


# API for student to add BL courses
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
@role_required(['student'])
@block_pg_phd
def add_course(request):
    try:
        student = Student.objects.select_related('batch_id__curriculum').get(
            id__user=request.user
        )

        eligibility_resp = get_add_registration_eligibility(
            timezone.now().date(), 
            student.curr_semester_no, 
            datetime.datetime.now().year
        )
        if isinstance(eligibility_resp, JsonResponse):
            return eligibility_resp

        course = Courses.objects.get(id=request.data.get('course_id'))
        slot = CourseSlot.objects.get(id=request.data.get('slot_id'))

        if not slot.name.startswith('BL'):
            return Response({
                'error': f'Only BL slots allowed. "{slot.name}" is not a BL slot.'
            }, status=status.HTTP_400_BAD_REQUEST)

        backlog_grades = ['F', 'X', 'CD']
        improvement_grades = ['C', 'D+', 'D']
        allowed_grades = backlog_grades + improvement_grades
        
        student_grade = Student_grades.objects.filter(
            roll_no=request.user.username,
            course_id=course
        ).order_by('-year', '-semester').first()
        
        if not student_grade:
            return Response({
                'error': 'You can only register for BL courses if you have a grade below C+ in this course. No grade record found.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if student_grade.grade not in allowed_grades:
            return Response({
                'error': f'You can only register for BL courses with grades below C+. Your grade: {student_grade.grade}'
            }, status=status.HTTP_400_BAD_REQUEST)

        current_year = datetime.datetime.now().year
        session, semester_type = generate_current_session(current_year, student.curr_semester_no)
        working_year = parse_academic_year(academic_year=session, semester_type=semester_type)[0]

        # Resolve the offering (section) to register into. Prefer the student's own
        # section; if that section doesn't run the course, they must pick one of the
        # running sections (course_instructor id) — a cross-section backlog/improvement.
        from applications.academic_information.models import resolve_offering
        offering = resolve_offering(student, course, working_year, semester_type)
        ci_id = request.data.get('course_instructor_id')
        if ci_id:
            offering = CourseInstructor.objects.filter(
                id=ci_id, course_id=course, year=working_year, semester_type=semester_type,
            ).first()
            if not offering:
                return Response({
                    'error': 'Selected section is not running this course this semester.'
                }, status=status.HTTP_400_BAD_REQUEST)
        elif offering is None:
            running = CourseInstructor.objects.filter(
                course_id=course, year=working_year, semester_type=semester_type,
            ).exists()
            if running:
                return Response({
                    'error': 'This course is not running in your section. Please select a section to register in.'
                }, status=status.HTTP_400_BAD_REQUEST)

        old_course_reg = course_registration.objects.filter(
            student_id=student,
            course_id=course
        ).order_by('-working_year', '-semester_id__semester_no').first()

        existing_request = CourseAddRequest.objects.filter(
            student=student,
            course=course,
            course_slot=slot,
            academic_year=session,
            semester_type=semester_type,
            status__in=['Pending', 'Approved']
        ).first()
        
        if existing_request:
            return Response({
                'error': f'You already have a {existing_request.status.lower()} request for this slot in the current semester.'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            add_request = CourseAddRequest.objects.create(
                student=student,
                course=course,
                course_slot=slot,
                academic_year=session,
                semester_type=semester_type,
                old_course_registration=old_course_reg,
                course_instructor=offering,
                status='Pending'
            )
        except Exception as create_error:
            import logging
            logger = logging.getLogger(__name__)
            error_msg = str(create_error)
            if 'duplicate' in error_msg.lower() or 'unique constraint' in error_msg.lower():
                return Response({
                    'error': 'You have already submitted a request for this slot in the current semester.'
                }, status=status.HTTP_400_BAD_REQUEST)
            else:
                logger.error(f"Error creating add course request: {error_msg}", exc_info=True)
                return Response({
                    'error': 'Failed to submit course add request. Please try again or contact support.'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        return Response({
            'message': 'Course add request submitted successfully. Awaiting Academic approval.',
            'data': {
                'id': add_request.id,
                'course': course.code,
                'course_name': course.name,
                'slot': slot.name,
                'status': add_request.status,
                'created_at': add_request.created_at.isoformat()
            }
        }, status=status.HTTP_201_CREATED)
        
    except Student.DoesNotExist:
        return Response({'error': 'Student not found'}, status=status.HTTP_404_NOT_FOUND)
    except Courses.DoesNotExist:
        return Response({'error': 'Course not found'}, status=status.HTTP_404_NOT_FOUND)
    except CourseSlot.DoesNotExist:
        return Response({'error': 'Course slot not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# API to get available course slots for student (only BL slots)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
@role_required(['student'])
@block_pg_phd
def get_student_add_course_slots(request):
    try:
        current_user = request.user
        extra_info = ExtraInfo.objects.filter(user=current_user).first()
        
        if not extra_info:
            return Response({
                'error': 'User information not found'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        student = Student.objects.filter(id=extra_info.id).first()
        
        if not student:
            return Response({
                'error': 'Student information not found'
            }, status=status.HTTP_400_BAD_REQUEST)

        eligibility_resp = get_add_registration_eligibility(
            timezone.now().date(), 
            student.curr_semester_no, 
            datetime.datetime.now().year
        )
        if isinstance(eligibility_resp, JsonResponse):
            return eligibility_resp
        
        batch = student.batch_id
        if not batch or not batch.curriculum:
            return Response({
                'error': 'Student batch or curriculum not found'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        current_semester = Semester.objects.filter(
            curriculum=batch.curriculum,
            semester_no=student.curr_semester_no
        ).first()
        
        if not current_semester:
            return Response({
                'error': 'Current semester not found'
            }, status=status.HTTP_404_NOT_FOUND)

        bl_slots = CourseSlot.objects.filter(
            semester=current_semester,
            name__startswith='BL'
        )
        
        registered_slots = course_registration.objects.filter(
            student_id=student,
            semester_id=current_semester
        ).values_list('course_slot_id', flat=True)

        current_year = datetime.datetime.now().year
        academic_year, semester_type = generate_current_session(current_year, student.curr_semester_no)

        pending_request_slots = CourseAddRequest.objects.filter(
            student=student,
            academic_year=academic_year,
            semester_type=semester_type,
            status='Pending'
        ).values_list('course_slot_id', flat=True)
        
        available_slots = []
        
        for slot in bl_slots:
            if slot.id not in registered_slots and slot.id not in pending_request_slots:
                course_count = slot.courses.count()
                
                available_slots.append({
                    'id': slot.id,
                    'name': slot.name,
                    'type': slot.type,
                    'max_registration_limit': slot.max_registration_limit,
                    'academic_year': academic_year,
                    'semester_type': semester_type,
                    'semester_no': student.curr_semester_no,
                    'course_count': course_count
                })
        
        return Response(available_slots, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# API to get available courses for a specific slot (only BL slots)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
@role_required(['student'])
@block_pg_phd
def get_student_add_courses(request):
    try:
        current_user = request.user
        extra_info = ExtraInfo.objects.filter(user=current_user).first()
        
        if not extra_info:
            return Response({
                'error': 'User information not found'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        student = Student.objects.filter(id=extra_info.id).first()
        
        if not student:
            return Response({
                'error': 'Student information not found'
            }, status=status.HTTP_400_BAD_REQUEST)

        batch = student.batch_id
        if not batch or not batch.curriculum:
            return Response({
                'error': 'Student batch or curriculum not found'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        curriculum = batch.curriculum
        current_sem_no = student.curr_semester_no
        
        current_semester = Semester.objects.filter(
            curriculum=curriculum,
            semester_no=current_sem_no
        ).first()
        
        if not current_semester:
            return Response({
                'error': 'Current semester not found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        slot_id = request.query_params.get('slot_id')

        if slot_id:
            bl_slots = CourseSlot.objects.filter(
                id=slot_id,
                semester=current_semester,
                name__startswith='BL'
            )
        else:
            bl_slots = CourseSlot.objects.filter(
                semester=current_semester,
                name__startswith='BL'
            )
        
        # Current term, to find where each course is actually running this sem.
        cur_year = datetime.datetime.now().year
        session, sem_type = generate_current_session(cur_year, current_sem_no)
        working_year = parse_academic_year(academic_year=session, semester_type=sem_type)[0]
        student_section = student.section or ''
        from applications.academic_information.models import offerings_by_course, pick_offering

        courses_list = []

        # One query each for slots+courses, registrations and offerings, instead of
        # three per course: the BL slots of a big semester hold hundreds of electives.
        slot_courses = [
            (slot, course)
            for slot in bl_slots.prefetch_related('courses')
            for course in slot.courses.all()
        ]
        course_ids = [course.id for _, course in slot_courses]
        registered_ids = set(course_registration.objects.filter(
            student_id=student, course_id__in=course_ids,
        ).values_list('course_id', flat=True))
        offerings_of = offerings_by_course(course_ids, working_year, sem_type)

        for slot, course in slot_courses:
            # Sections where this course is running this term. If the student's
            # own section isn't among them, they pick one of these (backlog/improvement).
            offerings = offerings_of.get(course.id, [])
            sections = []
            for o in offerings:
                u = o.instructor_id.id.user
                sections.append({
                    'course_instructor_id': o.id,
                    'section': o.section_label or '',
                    'instructor': f"{u.first_name} {u.last_name}".strip(),
                })
            # Auto-resolvable (own section, or a single/unsectioned offering) -> no pick needed; only genuinely sectioned courses missing the student's section prompt one.
            own_section_running = pick_offering(student, offerings) is not None

            courses_list.append({
                'id': course.id,
                'code': course.code,
                'name': course.name,
                'credit': course.credit,
                'slot': slot.name,
                'slot_id': slot.id,
                'already_registered': course.id in registered_ids,
                'sections': sections,
                'own_section_running': own_section_running,
                'student_section': student_section,
            })

        return Response(courses_list, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def drop_course(request):
#     if not request.user.is_authenticated:
#         return Response({'message': 'Login required '}, status=status.HTTP_400_BAD_REQUEST)
#     data = request.GET.get('id')
#     reg_id = int(data)
#     current_user = request.user
#     current_user = ExtraInfo.objects.all().filter(user=current_user).first()
#     current_user = Student.objects.all().filter(id = current_user.id).first()
#     try:
#         course_registration.objects.filter(id = reg_id, student_id = current_user).delete()
#     except Exception as e:
#         resp = {"message" : "Course drop failed", "error" : str(e)}
#         return Response(data = resp, status = status.HTTP_400_BAD_REQUEST)
    
#     resp = {"message" : "Course successfully dropped"}
#     return Response(data = resp , status = status.HTTP_200_OK)


# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def student_swayam_add_course(request):
#     if not request.user.is_authenticated:
#         return JsonResponse({'message': 'Login required '}, status=401)
#     try:
#         current_user = request.user
#         current_user = ExtraInfo.objects.all().filter(user=current_user).first()
#         current_user = Student.objects.all().filter(id = current_user.id).first()
#         course_id = request.POST["course_id"]
#         courseslot_id = request.POST["courseslot_id"]
#         registration_type = request.POST["registration_type"]
#         if (not course_id) or (not courseslot_id) or (not registration_type):
#             return JsonResponse({'message': 'Enter Complete Form Details '}, status=400)
#         course = Courses.objects.get(id=course_id)
#         courseslot = CourseSlot.objects.get(id=courseslot_id)
#         semester_no = current_user.curr_semester_no
#         curr_id = current_user.batch_id.curriculum
#         semester = Semester.objects.get(curriculum = curr_id, semester_no = semester_no)
#         try:
#             course_registration.objects.get(course_slot_id = courseslot, student_id = current_user)
#             return JsonResponse({'message': 'already registered a course in course slot'}, status=400)
#         except:
#             pass
#         try:
#             course_registration.objects.get(course_id = course, student_id = current_user)
#             return JsonResponse({'message': 'already registered a particular course'}, status=400)
#         except:
#             pass
#         cr = course_registration(
#             course_slot_id=courseslot, course_id=course, student_id=current_user, semester_id=semester , working_year = datetime.datetime.now().year, registration_type=registration_type)
#         cr.save()
#         return JsonResponse({'message': 'Successfully added swayam course' }, status=200)
#     except Exception as e:
#         print(str(e))
#         return JsonResponse({'message': 'Error adding course '}, status=500)
        

# simple api for getting to know the details of user who have logined in the system
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_user_info(request):
    current_user = request.user
    details1 = serializers.UserSerializer(current_user).data
    details2 = serializers.ExtraInfoSerializer(current_user.extrainfo).data
    details = {
        "user_serializer_Data" : details1,
        "ExtraInfoSerializer_Data" : details2
    }
    return Response(data = details  , status= status.HTTP_200_OK)


# with this api student can see the list of courses offered to him in upcoming semester
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['student'])
def view_offered_courses(request):
    try : 
        obj = Curriculum.objects.filter(
            programme = request.data['programme'],
            branch = request.data['branch'],
            batch = request.data["batch"],
            sem = request.data["semester"]
        )
        serializer = serializers.CurriculumSerializer(obj, many=True).data
        return Response(serializer, status=status.HTTP_200_OK)
    except Exception as e:
            return Response(data = str(e) , status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    # try:
    #     ug_flag = True
    #     masters_flag = False
    #     phd_flag = False
    #     current_semester =  get_user_semester(request.user, ug_flag, masters_flag, phd_flag)
    #     current_year = date_time.date().year
        
    #     return Response(data= { } , status=status.HTTP_200_OK)
    # except Exception as e:
    #     return Response(data = {"error" : str(e)} , status= status.HTTP_500_INTERNAL_SERVER_ERROR)


#  with this student can know status of pre registration and final registration
@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(['student'])
def student_view_registration(request):
    try:
        # getting the registration status of current user for the given semester
        current_user = request.user
        student_id = current_user.extrainfo.id
        
        sem_id = Semester.objects.get(id = request.data.get('semester'))
        sem_id = serializers.SemesterSerializer(sem_id).data["id"]

        # filter based on the semester id and student id
        obj = StudentRegistrationChecks.objects.filter(semester_id_id = sem_id,  student_id = student_id)

        # serialize the data for displaying
        serializer = serializers.StudentRegistrationChecksSerializer(obj, many=True).data

        return Response(serializer, status=status.HTTP_200_OK)
    except Exception as e:
        return Response(data = str(e) , status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# with this student can do his pre registration for the upcoming semester
# @api_view(['POST'])
# @transaction.atomic
# def student_pre_registration(request):
#     try:
#         current_user = request.user
#         current_user_id = serializers.UserSerializer(current_user).data["id"]
#         s_id = current_user.extrainfo.id

#         current_user = ExtraInfo.objects.all().select_related('user','department').filter(user=current_user_id).first()
#         current_user = serializers.ExtraInfoSerializer(current_user).data

#         current_user_instance = Student.objects.all().filter(id=current_user["id"]).first()
#         current_user = serializers.StudentSerializers(current_user_instance).data

#         sem_id_instance = Semester.objects.get(id = request.data.get('semester'))
#         sem_id = serializers.SemesterSerializer(sem_id_instance).data["id"]

#         # filter based on the semester id and student id
#         obj = StudentRegistrationChecks.objects.filter(semester_id_id = sem_id,  student_id = s_id)
#         # serialize the data for displaying
#         student_registration_check = serializers.StudentRegistrationChecksSerializer(obj, many=True).data

#         try:
#             # check if user have already done pre registration
#             if(student_registration_check and student_registration_check[0]["pre_registration_flag"] ):
#                 return Response(data = {"message" : "You have already registered for this semester" }, status=status.HTTP_400_BAD_REQUEST)
#         except Exception as e:
#             return Response(data = str(e) , status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#         course_slots=request.data.get("course_slot")
#         reg_curr = []
        

#         for course_slot in course_slots :
#             course_priorities = request.data.get("course_priority-"+course_slot)
#             if(course_priorities[0] == 'NULL'):
#                 print("NULL FOUND")
#                 continue
#             course_slot_id_for_model = CourseSlot.objects.get(id = int(course_slot))

#             # return Response(data = course_slots , status=status.HTTP_200_OK)
#             for course_priority in course_priorities:
#                 priority_of_current_course,course_id = map(int,course_priority.split("-"))
#                 # get course id for the model
#                 course_id_for_model = Courses.objects.get(id = course_id)
#                 print("check")
#                 p = InitialRegistration(
#                     course_id = course_id_for_model,
#                     semester_id = sem_id_instance,
#                     student_id = current_user_instance,
#                     course_slot_id = course_slot_id_for_model,
#                     priority = priority_of_current_course
#                 )
#                 p.save()
#                 reg_curr.append(p)

        
#         try:
#             serialized_reg_curr = serializers.InitialRegistrationSerializer(reg_curr, many=True).data
            
#             registration_check = StudentRegistrationChecks(
#                         student_id = current_user_instance,
#                         pre_registration_flag = True,
#                         final_registration_flag = False,
#                         semester_id = sem_id_instance
#                     )
#             registration_check.save()
#             return Response(data={"message": "Successfully Registered for the courses.", "registrations": serialized_reg_curr}, status=status.HTTP_200_OK)
#         except Exception as e:
#             return Response(data = {"message" : "Error in Registration." , "error" : str(e)} , status=status.HTTP_500_INTERNAL_SERVER_ERROR)
#     except Exception as e:
#         return Response(data = {"message" : "Error in Registration." , "error" : str(e)} , status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['student'])
@block_pg_phd
def final_registration(request):
    try:
        with transaction.atomic():
            current_user = request.user
            extra_info = current_user.extrainfo
            student = Student.objects.filter(id=extra_info).first()

            sem_id = Semester.objects.get(id=request.data.get('semester'))

            mode = str(request.data.get('mode'))
            transaction_id = str(request.data.get('transaction_id'))
            deposit_date = request.data.get('deposit_date')
            utr_number = str(request.data.get('utr_number'))
            fee_paid = request.data.get('fee_paid')
            actual_fee = request.data.get('actual_fee')
            reason = str(request.data.get('reason')) or None  # Handle empty string
            fee_receipt = request.FILES['fee_receipt']
            # Save FeePayments object
            obj = FeePayments(
                student_id=student,
                semester_id=sem_id,
                mode=mode,
                transaction_id=transaction_id,
                deposit_date=deposit_date,
                utr_number=utr_number,
                fee_paid=fee_paid,
                actual_fee=actual_fee,
                reason=reason,
                fee_receipt=fee_receipt
            )
            obj.save()

            # Update StudentRegistrationChecks
            StudentRegistrationChecks.objects.filter(
                student_id=student,
                semester_id=sem_id
            ).update(final_registration_flag=True)

            return JsonResponse({'message': 'Final Registration Successful'})
        
    except Exception as e:
        return JsonResponse({'message': f'Final Registration Failed: {str(e)}'}, status=500)
        
        
# with this student can do his final registration for the upcoming semester
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@transaction.atomic
def student_final_registration(request):
    try:
        current_user = request.user
        current_user_id = serializers.UserSerializer(current_user).data["id"]
        s_id = current_user.extrainfo.id

        current_user = ExtraInfo.objects.all().select_related('user','department').filter(user=current_user).first()
        current_user = serializers.ExtraInfoSerializer(current_user).data

        current_user_instance = Student.objects.all().filter(id=current_user["id"]).first()
        current_user = serializers.StudentSerializers(current_user_instance).data

        # these details we need from the body of the request fot doing final registration
        sem_id_instance = Semester.objects.get(id = request.data.get('semester'))
        sem_id = serializers.SemesterSerializer(sem_id_instance).data["id"]
        registration_status = StudentRegistrationChecks.objects.filter(student_id = current_user["id"], semester_id = sem_id)
        registration_status = serializers.StudentRegistrationChecksSerializer(registration_status , many = True ).data
        
        if(len(registration_status)>0 and registration_status[0]["pre_registration_flag"] == False):
            return Response(data = {"message" : "Student haven't done pre registration yet."} , status= status.HTTP_400_BAD_REQUEST )
        mode = str(request.data.get('mode'))
        transaction_id = str(request.data.get('transaction_id'))
        deposit_date = request.data.get('deposit_date')
        utr_number = str(request.data.get('utr_number'))
        fee_paid = request.data.get('fee_paid')
        actual_fee = request.data.get('actual_fee')
        reason = str(request.data.get('reason'))
        if reason=="":
            reason=None
        # fee_receipt = request.FILES['fee_receipt']

        # print(fee_receipt)
        obj = FeePayments(
            student_id = current_user_instance,
            semester_id = sem_id_instance,
            mode = mode,
            transaction_id = transaction_id,
            # fee_receipt = fee_receipt,
            deposit_date = deposit_date,
            utr_number = utr_number,
            fee_paid = fee_paid,
            actual_fee = actual_fee,
            reason = reason
            )
        obj.save()
        try:
            registration_status = StudentRegistrationChecks.objects.filter(student_id = current_user_instance, semester_id = sem_id).update(final_registration_flag = True)
            return Response(data = {"message" : "Final Registration Successfull" } , status= status.HTTP_200_OK)
        except Exception as e:
            return Response(data = {"message" : "Final Registration Failed " , "error" : str(e)} , status = status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        return Response(data = {"message" : "Final Registration Failed " , "error" : str(e)} , status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# with this api student can get his backlog courses list
# @api_view(['GET'])
# def student_backlog_courses(request):
#     try : 
#         stu_id = Student.objects.select_related('id','id__user','id__department').get(id=request.user.username)
#         backlogCourseList = []
#         backlogCourses = backlog_course.objects.select_related('course_id' , 'student_id' , 'semester_id' ).filter(student_id=stu_id)
#         for i in backlogCourses:
#             obj = {
#                 "course_id" : i.course_id.id,
#                 "course_name" : i.course_id.course_name,
#                 "faculty" : i.course_id.course_details,
#                 "semester" : i.semester_id.semester_no,
#                 "is_summer_course" : i.is_summer_course
#             }
#             backlogCourseList.append(obj)

#         return Response(backlogCourseList, status=status.HTTP_200_OK)
#     except Exception as e:
#             return Response(data = str(e) , status=status.HTTP_500_INTERNAL_SERVER_ERROR)



#--------------------------------------- APIs of acad person----------------------------------------------------------


# with this acad admin can fetch the list of courses for any batch , semester and brach
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def get_course_list(request):
    
    programme = request.data['programme']
    branch = request.data['branch']
    batch = request.data['batch']

    try : 
        obj = Curriculum.objects.filter(
            programme = request.data['programme'],
            branch = request.data['branch'],
            batch = request.data["batch"]
        )
        serializer = serializers.CurriculumSerializer(obj, many=True).data
        return Response(serializer, status=status.HTTP_200_OK)
    except Exception as e:
            return Response(data = str(e) , status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    # obj = Curriculum.objects.filter(curriculum_id_=curriculum_id, course_type_ = course_type, programme_ = programme, batch_ = batch, branch_ = branch, sem_ = sem, optional_ = optional)


#  with this api acad person can see the list of students who have completed their pre and final registrations for any semester
@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def acad_view_reigstrations(request):
    try:
        semester = request.data["semester"]
        sem_id_instance = Semester.objects.get(id = request.data.get('semester'))
        sem_id = serializers.SemesterSerializer(sem_id_instance).data["id"]
        obj = StudentRegistrationChecks.objects.filter(semester_id_id = sem_id,   final_registration_flag =True)
        student_registration_check = serializers.StudentRegistrationChecksSerializer(obj, many=True).data

        return Response(data= student_registration_check  , status=status.HTTP_200_OK)
    except Exception as e:
        return Response(data = {"error" : str(e)} , status= status.HTTP_500_INTERNAL_SERVER_ERROR)


# with this api acad person set the date of pre registration date for any semester
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['acadadmin'])
def configure_pre_registration_date(request):
    try:
        try:
            from_date = request.data.get('from_date')
            to_date = request.data.get('to_date')
            semester = request.data.get('semester')
            current_year = date_time.date().year
            desc = "Pre Registration " + str(semester) +" " + str(current_year)
            from_date = from_date.split('-')
            from_date = [int(i) for i in from_date]
            from_date = datetime.datetime(*from_date).date()
            to_date = to_date.split('-')
            to_date = [int(i) for i in to_date]
            to_date = datetime.datetime(*to_date).date()
        except Exception as e:
            from_date=""
            to_date=""
            desc=""
            pass
        c = Calendar(
            from_date=from_date,
            to_date=to_date,
            description=desc)
        c.save()
        return Response(data = {"message" : "Pre registration for semester " + str(semester) + " will be opened from " + str(from_date) + " to " + str(to_date) + ". "  ,  } , status= status.HTTP_200_OK)
    except Exception as e:
        return Response(data = {"error " : str(e)} , status= status.HTTP_500_INTERNAL_SERVER_ERROR)


# with this api request acad person can set the date of final registration
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['acadadmin'])
def configure_final_registration_date(request):
    try:
        try:
            from_date = request.data.get('from_date')
            to_date = request.data.get('to_date')
            semester = request.data.get('semester')
            current_year = date_time.date().year
            desc = "Physical Reporting at the Institute"
            from_date = from_date.split('-')
            from_date = [int(i) for i in from_date]
            from_date = datetime.datetime(*from_date).date()
            to_date = to_date.split('-')
            to_date = [int(i) for i in to_date]
            to_date = datetime.datetime(*to_date).date()
        except Exception as e:
            from_date=""
            to_date=""
            desc=""
            pass
        c = Calendar(
            from_date=from_date,
            to_date=to_date,
            description=desc)
        c.save()
        return Response(data = {"message" : "Physical Reporting at the Institute will be opened from " + str(from_date) + " to " + str(to_date) + ". "  ,  } , status= status.HTTP_200_OK)
    except Exception as e:
        return Response(data = {"error " : str(e)} , status= status.HTTP_500_INTERNAL_SERVER_ERROR)

# with this api request acad person can add any courses in a specific slot   
# @api_view(['POST'])
# def add_course_to_slot(request):
#     course_code = request.data.get('course_code')
#     course_slot_name = request.data.get('course_slot_name')
#     try:
#         course_slot = CourseSlot.objects.get(name=course_slot_name)
#         course = Courses.objects.get(code=course_code)
#         course_slot.courses.add(course)
        
#         return JsonResponse({'message': f'Course {course_code} added to slot {course_slot_name} successfully.'}, status=200)
#     except CourseSlot.DoesNotExist:
#         return JsonResponse({'error': 'Course slot does not exist.'}, status=400)
#     except Courses.DoesNotExist:
#         return JsonResponse({'error': 'Course does not exist.'}, status=400)

# # with this api request acad person can remove any course from a specific slot   
# @api_view(['POST'])
# def remove_course_from_slot(request):
#     course_code = request.data.get('course_code')
#     course_slot_name = request.data.get('course_slot_name')
#     try:
#         course_slot = CourseSlot.objects.get(name=course_slot_name)
#         course = Courses.objects.get(code=course_code)
#         course_slot.courses.remove(course)
#         return JsonResponse({'message': f'Course {course_code} removed from slot {course_slot_name} successfully.'}, status=200)
#     except CourseSlot.DoesNotExist:
#         return JsonResponse({'error': 'Course slot does not exist.'}, status=400)
#     except Course.DoesNotExist:
#         return JsonResponse({'error': 'Course does not exist.'}, status=400)
  


#--------------------------------------- APIs of faculty----------------------------------------------------------

# with this api faculty can know what are the courses assigned to him 
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def faculty_assigned_courses(request):
    
    
    try:
        current_user = request.user
        curriculum_ids = Curriculum_Instructor.objects.filter(instructor_id=current_user.id).values_list('curriculum_id', flat=True)
        course_infos = []
        for curriculum_id in curriculum_ids:
            course_info = Curriculum.objects.filter(curriculum_id=curriculum_id).values_list('course_code','course_type','programme','branch','sem','course_id_id').first()
            # course_infos.append(course_info)
            context = {
                "course_code": course_info[0],
                "course_type": course_info[1],
                "programme": course_info[2],
                "branch": course_info[3],
                "sem": course_info[4],
                "course_id": course_info[5]
            }
            course_infos.append(context)
    
        return Response(data= course_infos , status=status.HTTP_200_OK)
    except Exception as e:
        return Response(data = {"error" : str(e)} , status= status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['student'])
def get_next_sem_courses(request):
    try:
        next_sem = request.data.get('next_sem')
        branch = request.data.get('branch')
        programme = request.data.get('programme')
        batch = request.data.get('batch')

        #  we go to student table and apply filters and get batch_id of the students with these filter
        batch_id = Student.objects.filter(programme = programme , batch = batch , specialization = branch)[0].batch_id

        curr_id = batch_id.curriculum
        next_sem_id = Semester.objects.get(curriculum = curr_id, semester_no = next_sem)
        
        if next_sem_id:
            next_sem_registration_courses = get_detailed_sem_courses(next_sem_id )
            return JsonResponse(next_sem_registration_courses, safe=False)
    except Exception as e:
        return Response(data = {"error" : str(e)} , status= status.HTTP_500_INTERNAL_SERVER_ERROR)

# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def add_one_course(request):
#     try:    
#         print(request.data)
#         current_user = get_object_or_404(User, username=request.data.get('user'))
#         current_user = ExtraInfo.objects.all().filter(user=current_user).first()
#         current_user = Student.objects.all().filter(id=current_user.id).first()

#         sem_id = Semester.objects.get(id=request.data.get('semester'))
#         choice = request.data.get('choice')
#         slot = request.data.get('slot')

#         try:
#             course_id = Courses.objects.get(id=choice)
#             courseslot_id = CourseSlot.objects.get(id=slot)
#             print(courseslot_id.id)
#             print(courseslot_id.type)
#             print(courseslot_id.max_registration_limit)
#             if course_registration.objects.filter(course_slot_id_id=courseslot_id, student_id=current_user).count() == 1 and courseslot_id.type != "Swayam":
#                 already_registered_course_id = course_registration.objects.filter(course_slot_id_id=courseslot_id, student_id=current_user)[0].course_id
#                 print(already_registered_course_id)
#                 msg = 'Already Registered in the course : ' +already_registered_course_id.code + '-'+ already_registered_course_id.name
#                 return JsonResponse({'message' : msg})
#             if((course_registration.objects.filter(course_id=course_id, student_id=current_user).count() >= 1)):
#                 return JsonResponse({'message': 'Already registered in this course!'}, status=200)
#             # Check if maximum course registration limit has not been reached
#             if course_registration.objects.filter(student_id__batch_id__year=current_user.batch_id.year, course_id=course_id).count() < courseslot_id.max_registration_limit and \
#                     (course_registration.objects.filter(course_id=course_id, student_id=current_user).count() == 0):
#                 p = course_registration(
                    
#                     course_id=course_id,
#                     student_id=current_user,
#                     course_slot_id=courseslot_id,
#                     semester_id=sem_id
#                 )
#                 p.save()
#                 return JsonResponse({'message': 'Course added successfully'}, status=200)
#             else:
#                 return JsonResponse({'message': 'Course not added because seats are full!'}, status=404)
#         except Exception as e:
#             print(e)
#             return JsonResponse({'message': 'Error adding course'}, status=500)
#     except Exception as e:
#         return JsonResponse({'message': 'Error adding course'}, status=500)
    

@transaction.atomic
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
@role_required(list(ALL_ACAD_ROLES))
def verify_registration(request):
    data = json.loads(request.body)
    if data.get('status_req') == "accept" :
        student_id = data.get('student_id')
        student = Student.objects.get(id = student_id)
        batch = student.batch_id
        curr_id = batch.curriculum
        
        # Verify whichever semester the student actually registered for: a fresh
        # intake registers into the semester it is already in, so assuming the
        # next one would find nothing to verify and promote it a semester early.
        sem_no = pre_registration_target_semester(student)
        if sem_no >= 9:
            sem_no = 8
        sem_id = Semester.objects.get(curriculum = curr_id, semester_no = sem_no)
        # print('----------------------------------------------------------------' , student.curr_semester_no)
        
        final_register_list = FinalRegistration.objects.all().filter(student_id = student, verified = False, semester_id = sem_id)
        
        # final_register_list = FinalRegistration.objects.all().filter(student_id = student, verified = False)
        
        with transaction.atomic():
            from applications.academic_information.models import resolve_offering
            ver_reg = []
            for obj in final_register_list:
                _work_year = datetime.datetime.now().year
                _sem_no    = obj.semester_id.semester_no
                _sem_type  = "Odd Semester" if _sem_no % 2 == 1 else "Even Semester"
                if _sem_type == "Odd Semester":
                    _session = f"{_work_year}-{str(_work_year + 1)[-2:]}"
                else:
                    _session = f"{_work_year - 1}-{str(_work_year)[-2:]}"
                # Bind this registration to the offering (course+section+faculty)
                # the student belongs to, so grades are attributable per instructor.
                offering = resolve_offering(student, obj.course_id, _work_year, _sem_type)
                p = course_registration(
                    course_id=obj.course_id,
                    student_id=student,
                    semester_id=obj.semester_id,
                    course_slot_id=obj.course_slot_id,
                    working_year=_work_year,
                    registration_type=obj.registration_type,
                    session=_session,
                    semester_type=_sem_type,
                    course_instructor=offering,
                    )
                # ver_reg.append(p)
                p.save()
                if (obj.old_course_registration):
                    course_replacement.objects.create(new_course_registration=p, old_course_registration=obj.old_course_registration)
                o = FinalRegistration.objects.filter(id= obj.id).update(verified = True, course_instructor = offering)
            # course_registration.objects.bulk_create(ver_reg)
            academics_module_notif(request.user, student.id.user, 'registration_approved')

            # Only a registration for a later semester moves the student on.
            if sem_no > student.curr_semester_no:
                Student.objects.filter(id = student_id).update(curr_semester_no = sem_no)
            return JsonResponse({'status': 'success', 'message': 'Successfully Accepted'})
         
    elif data.get('status_req') == "reject" :
        reject_reason = data.get('reason', '')
        student_id = data.get('student_id')
        student_id = Student.objects.get(id = student_id)
        batch = student_id.batch_id
        curr_id = batch.curriculum
        if(student_id.curr_semester_no+1 >= 9):
            sem_no = 8
        else:
            sem_no = student_id.curr_semester_no+1
        sem_id = Semester.objects.get(curriculum = curr_id, semester_no = sem_no)
        with transaction.atomic():
            academicadmin = get_object_or_404(User, username = request.user.username)
            # print(sem_id)
            # FinalRegistration.objects.filter(student_id = student_id, verified = False, semester_id = sem_id).delete()
            # StudentRegistrationChecks.objects.filter(student_id = student_id, semester_id = sem_id).update(final_registration_flag = False)
            FeePayments.objects.filter(student_id = student_id, semester_id = sem_id).delete()
            academics_module_notif(academicadmin, student_id.id.user, 'Registration Declined - '+reject_reason)
            return JsonResponse({'status': 'success', 'message': 'Successfully Rejected'})
        
    return JsonResponse({'status': 'error', 'message': 'Error in processing'})

@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def verify_course(request):
    roll_no = request.data.get("rollno")
    if not roll_no:
        return Response({"error": "rollno is required"}, status=status.HTTP_400_BAD_REQUEST)
    
    # Convert to uppercase after null check
    roll_no = roll_no.strip().upper()

    # First check main academic tables
    student = Student.objects.filter(id_id=roll_no).first()
    if not student:
        # If not found in main tables, check StudentBatchUpload
        try:
            from applications.programme_curriculum.models_student_management import StudentBatchUpload
            batch_student = StudentBatchUpload.objects.filter(roll_number=roll_no).first()
            if batch_student:
                if batch_student.reported_status == 'REPORTED':
                    return Response({
                        "error": f"Student {roll_no} has reported but not yet transferred to main academic system. Please contact academic office."
                    }, status=status.HTTP_400_BAD_REQUEST)
                else:
                    return Response({
                        "error": f"Student {roll_no} found in upcoming batches but status is '{batch_student.reported_status}'. Student must report first."
                    }, status=status.HTTP_400_BAD_REQUEST)
        except ImportError:
            pass
        
        return Response({"error": "Student record not found"}, status=status.HTTP_400_BAD_REQUEST)


    extra = student.id
    user_obj = extra.user
    
    # name & roll for frontend
    dict2 = {
        "roll_no": roll_no,
        "firstname": user_obj.first_name or "",
        "lastname": user_obj.last_name or "",
        "programme_category": _student_programme_category(student),
    }

    # current curriculum & semester - handle None batch_id case
    if not student.batch_id:
        return Response({
            "error": f"Student {roll_no} does not have a valid batch assignment. Please contact academic office to complete student setup."
        }, status=status.HTTP_400_BAD_REQUEST)
    
    if not student.batch_id.curriculum:
        return Response({
            "error": f"Student {roll_no}'s batch does not have a curriculum assigned. Please contact academic office."
        }, status=status.HTTP_400_BAD_REQUEST)
    
    curr = student.batch_id.curriculum
    curr_sem = Semester.objects.filter(curriculum=curr, semester_no=student.curr_semester_no).first()
    if not curr_sem:
        return Response({"error": "Current semester not found"}, status=status.HTTP_404_NOT_FOUND)

    # gather registered courses for this semester
    regs = course_registration.objects.filter(student_id=student).order_by('-semester_id__semester_no')
    details = []
    for reg in regs:
        slot_course = Courses.objects.filter(id=reg.course_id.id).first()
        repl_qs = course_replacement.objects.filter(old_course_registration=reg)
        replaced_list = []
        for repl in repl_qs:
            nr = repl.new_course_registration
            replaced_list.append({
                "course_id": {
                    "code": nr.course_id.code,
                    "name": nr.course_id.name,
                },
                "semester_id": {
                    "semester_no": nr.semester_id.semester_no,
                },
            })

        details.append({
            "id": reg.id,
            "reg_id": reg.id,
            "rid": f"{roll_no} - {reg.course_id.code}",
            "course_id": reg.course_id.code,
            "course_name": reg.course_id.name,
            "sem": reg.semester_id.semester_no,
            "semester_type" : reg.semester_type,
            "credits": slot_course.credit if slot_course else 0,
            "registration_type": reg.registration_type,
            "replaced_by": replaced_list,
        })

    # lists for selects (no serializers)
    course_list = list(Courses.objects.values("id", "code", "name", "credit"))
    
    # For PhD students, show only current semester + next semester (no summer terms)
    batch_name = student.batch_id.name if student.batch_id else ""
    is_phd = batch_name.upper().startswith('PHD')
    
    if is_phd:
        current_sem = student.curr_semester_no
        all_semesters = Semester.objects.filter(
            curriculum=curr
        ).exclude(
            semester_no__in=[4, 6, 8, 10, 12]
        ).order_by('semester_no')

        phd_semester_list = []
        for sem in all_semesters:
            if sem.semester_no == current_sem or sem.semester_no == current_sem + 1:
                phd_semester_list.append({
                    "id": sem.id, 
                    "semester_no": sem.semester_no
                })
        
        semester_list = phd_semester_list if phd_semester_list else list(
            Semester.objects.filter(curriculum=curr).values("id", "semester_no")
        )
    else:
        semester_list = list(
            Semester.objects.filter(curriculum=curr).values("id", "semester_no")
        )
    
    courseslot_list = list(
        CourseSlot.objects.filter(semester__in=[s["id"] for s in semester_list]).values("id", "name")
    )

    # academic year & semflag
    today = date.today()
    year = today.year
    semflag = 1 if today.month >= 7 else 2
    yearr = f"{year}-{year+1}"
    if today.month >= 7:
        current_semester_type = "Odd Semester"
    else:
        current_semester_type = "Even Semester"

    return Response({
        "details": details,
        "dict2": dict2,
        "course_list": course_list,
        "semester_list": semester_list,
        "courseslot_list": courseslot_list,
        "date": {"year": yearr, "semflag": semflag},
        "current_semester": {
            "semester_no": student.curr_semester_no,
            "semester_type": current_semester_type
        }
    })


#  These apis were implemented before but now don't use them they have some errors


# @api_view(['GET'])
# def academic_procedures_faculty(request):
#     current_user = request.user
#     user_details = current_user.extrainfo
#     des = current_user.holds_designations.all().first()

#     if str(des.designation) == 'student':
#         return Response({'error':'Not a faculty'}, status=status.HTTP_400_BAD_REQUEST)
#     elif str(current_user) == 'acadadmin':
#         return Response({'error':'User is acadadmin'}, status=status.HTTP_400_BAD_REQUEST)

#     elif str(des.designation) == "Associate Professor" or str(des.designation) == "Professor" or str(des.designation) == "Assistant Professor":
#         faculty_object = user_details.faculty
#         month = int(date_time.month)
#         sem = []
#         if month>=7 and month<=12:
#             sem = [1,3,5,7]
#         else:
#             sem = [2,4,6,8]
#         student_flag = False
#         fac_flag = True

#         thesis_supervision_request_list = faculty_object.thesistopicprocess_supervisor.all()
#         thesis_supervision_request_list_data = serializers.ThesisTopicProcessSerializer(thesis_supervision_request_list, many=True).data
#         approved_thesis_request_list = serializers.ThesisTopicProcessSerializer(thesis_supervision_request_list.filter(approval_supervisor = True), many=True).data
#         pending_thesis_request_list = serializers.ThesisTopicProcessSerializer(thesis_supervision_request_list.filter(pending_supervisor = True), many=True).data
#         courses_list = serializers.CurriculumInstructorSerializer(user_details.curriculum_instructor_set.all(), many=True).data
#         fac_details = serializers.UserSerializer(current_user).data

#         resp = {
#             'student_flag' : student_flag,
#             'fac_flag' : fac_flag,
#             'thesis_supervision_request_list' : thesis_supervision_request_list_data,
#             'pending_thesis_request_list' : pending_thesis_request_list,
#             'approved_thesis_request_list' : approved_thesis_request_list,
#             'courses_list': courses_list,
#             'faculty': fac_details
#         }
#         return Response(data=resp, status=status.HTTP_200_OK)






# @api_view(['GET'])
# def academic_procedures_student(request):
#     current_user = request.user
#     current_user_data = {
#         'first_name': current_user.first_name,
#         'last_name': current_user.last_name,
#         'username': current_user.username,
#         'email': current_user.email
#     }
#     user_details = current_user.extrainfo
#     des = current_user.holds_designations.all().first()
#     if str(des.designation) == 'student':
#         obj = user_details.student

#         if obj.programme.upper() == "PH.D":
#             student_flag = True
#             ug_flag = False
#             masters_flag = False
#             phd_flag = True
#             fac_flag = False
#             des_flag = False

#         elif obj.programme.upper() == "M.DES":
#             student_flag = True
#             ug_flag = False
#             masters_flag = True
#             phd_flag = False
#             fac_flag = False
#             des_flag = True

#         elif obj.programme.upper() == "B.DES":
#             student_flag = True
#             ug_flag = True
#             masters_flag = False
#             phd_flag = False
#             fac_flag = False
#             des_flag = True

#         elif obj.programme.upper() == "M.TECH":
#             student_flag = True
#             ug_flag = False
#             masters_flag = True
#             phd_flag = False
#             fac_flag = False
#             des_flag = False

#         elif obj.programme.upper() == "B.TECH":
#             student_flag = True
#             ug_flag = True
#             masters_flag = False
#             phd_flag = False
#             fac_flag = False
#             des_flag = False

#         else:
#             return Response({'message':'Student has no record'}, status=status.HTTP_400_BAD_REQUEST)

#         current_date = date_time.date()
#         current_year = date_time.year
#         batch = obj.batch_id
#         user_sem = get_user_semester(request.user, ug_flag, masters_flag, phd_flag)
#         acad_year = get_acad_year(user_sem, current_year)
#         user_branch = user_details.department.name
#         cpi = obj.cpi
#         cur_spi='Sem results not available' # To be fetched from db if result uploaded

#         details = {
#             'current_user': current_user_data,
#             'year': acad_year,
#             'user_sem': user_sem,
#             'user_branch' : str(user_branch),
#             'cpi' : cpi,
#             'spi' : cur_spi
#         }
        
#         currently_registered_courses = get_currently_registered_courses(user_details.id, user_sem)
#         currently_registered_courses_data = serializers.CurriculumSerializer(currently_registered_courses, many=True).data
#         try:
#             pre_registered_courses = obj.initialregistrations_set.all().filter(semester = user_sem)
#             pre_registered_courses_show = obj.initialregistrations_set.all().filter(semester = user_sem+1)
#         except:
#             pre_registered_courses =  None
#             pre_registered_courses_show=None
#         try:
#             final_registered_courses = obj.finalregistrations_set.all().filter(semester = user_sem)
#         except:
#             final_registered_courses = None

#         pre_registered_courses_data = serializers.InitialRegistrationsSerializer(pre_registered_courses, many=True).data
#         pre_registered_courses_show_data = serializers.InitialRegistrationsSerializer(pre_registered_courses_show, many=True).data
#         final_registered_courses_data = serializers.FinalRegistrationsSerializer(final_registered_courses, many=True).data

#         current_credits = get_current_credits(currently_registered_courses)
#         print(current_user, user_sem+1, user_branch)
#         try:
#             next_sem_branch_courses = get_branch_courses(current_user, user_sem+1, user_branch)
#         except Exception as e:
#             return Response(data = str(e))
#         next_sem_branch_courses_data = serializers.CurriculumSerializer(next_sem_branch_courses, many=True).data

#         fee_payment_mode_list = dict(Constants.PaymentMode)

#         next_sem_branch_registration_courses = get_registration_courses(next_sem_branch_courses)
#         next_sem_branch_registration_courses_data = []
#         for choices in next_sem_branch_registration_courses:
#             next_sem_branch_registration_courses_data.append(serializers.CurriculumSerializer(choices, many=True).data)
#         # next_sem_branch_registration_courses_data = serializers.CurriculumSerializer(next_sem_branch_registration_courses, many=True).data

#         final_registration_choices = get_registration_courses(get_branch_courses(request.user, user_sem, user_branch))
#         final_registration_choices_data = []
#         for choices in final_registration_choices:
#             final_registration_choices_data.append(serializers.CurriculumSerializer(choices, many=True).data)

#         performance_list = []
#         result_announced = False
#         for i in currently_registered_courses:
#             try:
#                 performance_obj = obj.semestermarks_set.all().filter(curr_id = i).first()
#             except:
#                 performance_obj = None
#             performance_list.append(performance_obj)
#         performance_list_data = serializers.SemesterMarksSerializer(performance_list, many=True).data

#         thesis_request_list = serializers.ThesisTopicProcessSerializer(obj.thesistopicprocess_set.all(), many=True).data

#         pre_existing_thesis_flag = True if obj.thesistopicprocess_set.all() else False

#         current_sem_branch_courses = get_branch_courses(current_user, user_sem, user_branch)

#     #     pre_registration_date_flag = get_pre_registration_eligibility(current_date)
#         final_registration_date_flag = get_final_registration_eligibility(current_date)

#         add_or_drop_course_date_flag = get_add_or_drop_course_date_eligibility(current_date)

#         student_registration_check_pre = obj.studentregistrationcheck_set.all().filter(semester=user_sem+1)
#         student_registration_check_final = obj.studentregistrationcheck_set.all().filter(semester=user_sem)
#         pre_registration_flag = False
#         final_registration_flag = False
#         if(student_registration_check_pre):
#             pre_registration_flag = student_registration_check_pre.pre_registration_flag
#         if(student_registration_check_final):
#             final_registration_flag = student_registration_check_final.final_registration_flag

#         teaching_credit_registration_course = None
#         if phd_flag:
#             teaching_credit_registration_course = Curriculum.objects.all().filter(batch = 2016, sem =6)
#         teaching_credit_registration_course_data = serializers.CurriculumSerializer(teaching_credit_registration_course, many=True).data

#         if student_flag:
#             try:
#                 due = obj.dues_set.get()
#                 lib_d = due.library_due
#                 pc_d = due.placement_cell_due
#                 hos_d = due.hostel_due
#                 mess_d = due.mess_due
#                 acad_d = due.academic_due
#             except:
#                 lib_d, pc_d, hos_d, mess_d, acad_d = 0, 0, 0, 0, 0

#         tot_d = lib_d + acad_d + pc_d + hos_d + mess_d

#         registers = obj.register_set.all()
#         course_list = []
#         for i in registers:
#             course_list.append(i.curr_id)
#         attendence = []
#         for i in course_list:
#             instructors = i.curriculum_instructor_set.all()
#             pr,ab=0,0
#             for j in list(instructors):

#                 presents = obj.student_attendance_set.all().filter(instructor_id=j, present=True)
#                 absents = obj.student_attendance_set.all().filter(instructor_id=j, present=False)
#                 pr += len(presents)
#                 ab += len(absents)
#             attendence.append((i,pr,pr+ab))
#         attendance_data = {}
#         for course in attendence:
#             attendance_data[course[0].course_id.course_name] = {
#                 'present' : course[1],
#                 'total' : course[2]
#             }

#         branchchange_flag = False
#         if user_sem == 2:
#             branchchange_flag=True

#         faculty_list = serializers.HoldsDesignationSerializer(get_faculty_list(), many=True).data

#         resp = {
#             'details': details,
#             'currently_registered': currently_registered_courses_data,
#             # 'pre_registered_courses' : pre_registered_courses_data,
#             # 'pre_registered_courses_show' : pre_registered_courses_show_data,
#             'final_registered_courses' : final_registered_courses_data,
#             'current_credits' : current_credits,
#             'courses_list': next_sem_branch_courses_data,
#             'fee_payment_mode_list' : fee_payment_mode_list,
#             'next_sem_branch_registration_courses' : next_sem_branch_registration_courses_data,
#             'final_registration_choices' : final_registration_choices_data,
#             'performance_list' : performance_list_data,
#             'thesis_request_list' : thesis_request_list,
#             'student_flag' : student_flag,
#             'ug_flag' : ug_flag,
#             'masters_flag' : masters_flag,
#             'phd_flag' : phd_flag,
#             'fac_flag' : fac_flag,
#             'des_flag' : des_flag,
#             'thesis_flag' : pre_existing_thesis_flag,
#             'drop_courses_options' : currently_registered_courses_data,
#             # 'pre_registration_date_flag': pre_registration_date_flag,
#             'final_registration_date_flag': final_registration_date_flag,
#             'add_or_drop_course_date_flag': add_or_drop_course_date_flag,
#             # 'pre_registration_flag' : pre_registration_flag,
#             'final_registration_flag': final_registration_flag,
#             'teaching_credit_registration_course' : teaching_credit_registration_course_data,
#             'lib_d':lib_d,
#             'acad_d':acad_d,
#             'mess_d':mess_d,
#             'pc_d':pc_d,
#             'hos_d':hos_d,
#             'tot_d':tot_d,
#             'attendance': attendance_data,
#             'Branch_Change_Flag':branchchange_flag
#             # 'faculty_list' : faculty_list
#         }
#     return Response(data=resp, status=status.HTTP_200_OK)






















#               These apis are not needed in this module



# @api_view(['POST'])
# def add_thesis(request):
#     current_user = request.user
#     profile = current_user.extrainfo
#     if profile.user_type == 'student':
#         if not 'thesis_topic' in request.data:
#             return Response({'error':'Thesis topic is required'}, status=status.HTTP_400_BAD_REQUEST)
#         if not 'research_area' in request.data:
#             return Response({'error':'Research area is required'}, status=status.HTTP_400_BAD_REQUEST)
#         if 'supervisor_id' in request.data:
#             try:
#                 supervisor_faculty = User.objects.get(username=request.data['supervisor_id'])
#                 supervisor_faculty = supervisor_faculty.extrainfo
#                 request.data['supervisor_id'] = supervisor_faculty
#             except:
#                 return Response({'error':'Wrong supervisor id. User does not exist.'}, status=status.HTTP_400_BAD_REQUEST)
#         else:
#             return Response({'error':'supervisor id is required'}, status=status.HTTP_400_BAD_REQUEST)
#         if 'co_supervisor_id' in request.data:
#             try:
#                 co_supervisor_faculty = User.objects.get(username=request.data['co_supervisor_id'])
#                 co_supervisor_faculty = co_supervisor_faculty.extrainfo
#                 request.data['co_supervisor_id'] = co_supervisor_faculty
#             except:
#                 return Response({'error':'Wrong co_supervisor id. User does not exist.'}, status=status.HTTP_400_BAD_REQUEST)
#         else:
#             co_supervisor_faculty = None
#         if 'curr_id' in request.data:
#             curr_id = None
#         student = profile.student
#         request.data['student_id'] = profile
#         request.data['submission_by_student'] = True
#         serializer = serializers.ThesisTopicProcessSerializer(data=request.data)
#         if serializer.is_valid():
#             serializer.save()
#             return Response(serializer.data, status=status.HTTP_200_OK)
#         else:
#             return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
#     else:
#         return Response({'error':'Cannot add thesis'}, status=status.HTTP_400_BAD_REQUEST)

















# @api_view(['PUT'])
# def approve_thesis(request, id):
#     current_user = request.user
#     profile = current_user.extrainfo
#     if profile.user_type == 'faculty':
#         try:
#             thesis = ThesisTopicProcess.objects.get(id=id)
#         except:
#             return Response({'error':'This thesis does not exist'}, status=status.HTTP_400_BAD_REQUEST)
#         if 'member1' in request.data:
#             try:
#                 user1 = User.objects.get(username=request.data['member1'])
#                 member1 = user1.extrainfo
#                 request.data['member1'] = member1
#             except:
#                 return Response({'error':'Wrong username of member 1. User does not exist.'}, status=status.HTTP_400_BAD_REQUEST)
#         else:
#             return Response({'error':'Member 1 is required'}, status=status.HTTP_400_BAD_REQUEST)
#         if 'member2' in request.data:
#             try:
#                 user2 = User.objects.get(username=request.data['member2'])
#                 member2 = user2.extrainfo
#                 request.data['member2'] = member2
#             except:
#                 return Response({'error':'Wrong username of member 2. User does not exist.'}, status=status.HTTP_400_BAD_REQUEST)
#         else:
#             return Response({'error':'Member 2 is required'}, status=status.HTTP_400_BAD_REQUEST)
#         if 'member3' in request.data:
#             try:
#                 user3 = User.objects.get(username=request.data['member3'])
#                 member3 = user3.extrainfo
#                 request.data['member3'] = member3
#             except:
#                 return Response({'error':'Wrong username of member 3. User does not exist.'}, status=status.HTTP_400_BAD_REQUEST)
#         else:
#             member3 = None
#         if not 'approval' in request.data:
#             return Response({'error':'Approval value is required.'}, status=status.HTTP_400_BAD_REQUEST)
#         elif request.data['approval'] != 'yes' and request.data['approval'] != 'no':
#             return Response({'error':'Wrong approval value provided. Approval value should be yes or no'}, status=status.HTTP_400_BAD_REQUEST)
#         if request.data['approval'] == 'yes':
#             request.data.pop('approval', None)
#             request.data['pending_supervisor'] = False
#             request.data['approval_supervisor'] = True
#             request.data['forwarded_to_hod'] = True
#             request.data['pending_hod'] = True
#         else:
#             request.data.pop('approval', None)
#             request.data['pending_supervisor'] = False
#             request.data['approval_supervisor'] = False
#             request.data['forwarded_to_hod'] = False
#             request.data['pending_hod'] = False
#         serializer = serializers.ThesisTopicProcessSerializer(thesis, data=request.data, partial=True)
#         if serializer.is_valid():
#             serializer.save()
#             return Response(serializer.data, status=status.HTTP_200_OK)
#         else:
#             return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
#     else:
#         return Response({'error':'Cannot approve thesis'}, status=status.HTTP_400_BAD_REQUEST)

#--------------------------------------- New APIs Made for React ----------------------------------------------------------


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(['student'])
@block_pg_phd
def final_registration_page(request):
    try:
        current_user = request.user
        user_details = current_user.extrainfo
        student = Student.objects.get(id=user_details)
        curr_id = student.batch_id.curriculum
        next_sem_id = Semester.objects.get(curriculum=curr_id, semester_no=student.curr_semester_no+1)
        current_date = datetime.date.today()
        final_registration_date_flag = get_final_registration_eligibility(current_date)
        student_registration_check = get_student_registrtion_check(student, next_sem_id)
        final_registration_flag = False
        if student_registration_check:
            final_registration_flag = student_registration_check.final_registration_flag

        final_registration = FinalRegistration.objects.filter(
            student_id=user_details.id, semester_id=next_sem_id
        )
        if final_registration.exists():
            final_registration = serializers.FinalRegistrationSerializer(final_registration, many=True).data
        else:
            final_registration = None
        frd_window = get_final_registration_window()
        resp = {
            'frd': final_registration_date_flag,
            'frd_configured': frd_window is not None,
            'frd_from': frd_window.from_date if frd_window else None,
            'frd_to': frd_window.to_date if frd_window else None,
            'final_registration_flag': final_registration_flag,
            'final_registration': final_registration,
        }
        return Response(data=resp, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

@api_view(['GET', 'POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def student_list(request):
    if request.method == 'POST':
        excel_export = request.GET.get("excel_export", "false")
        data = json.loads(request.body)
        batch = data.get('batch')
        
        year = demo_date.year
        month = demo_date.month
        yearr = f'{year}-{year+1}'
        semflag = 1 if month >= 7 else 2
        queryflag = 1

        batch_id = Batch.objects.get(id=batch)
        if not batch_in_scope(batch_id, scopes_for(request.user)):
            return Response({'detail': 'Batch not found.'},
                            status=status.HTTP_404_NOT_FOUND)
        student_obj = FeePayments.objects.all().select_related('student_id').filter(student_id__batch_id=batch_id)

        if excel_export == "false":
            if student_obj:
                reg_table = list(student_obj.prefetch_related('student_id__studentregistrationchecks')
                    .filter(semester_id=student_obj[0].semester_id, student_id__studentregistrationchecks__final_registration_flag=True, 
                            student_id__finalregistration__verified=False, student_id__finalregistration__semester_id=student_obj[0].semester_id)
                    .select_related('student_id', 'student_id__id', 'student_id__id__user', 'student_id__id__department')
                    .values(
                        'student_id__id', 'student_id__id__user__first_name', 'student_id__id__user__last_name',
                        'student_id__batch', 'student_id__id__department__name', 'student_id__programme',
                        'student_id__curr_semester_no', 'student_id__id__sex', 'student_id__id__phone_no',
                        'student_id__category', 'student_id__specialization', 'mode', 'transaction_id', 'deposit_date',
                        'fee_paid', 'utr_number', 'reason', 'fee_receipt', 'actual_fee',
                        'student_id__id__user__username'
                    ).distinct())

            else:
                reg_table = []

            response_data = {
                'date': {'year': yearr, 'month': month, 'semflag': semflag, 'queryflag': queryflag},
                'students': reg_table
            }

            return JsonResponse(response_data, safe=False)

        elif excel_export == "true":
            if student_obj:
                table = [("Admission Year", "Semester", "Roll Number", "Full Name", "Program", "Discipline", "Specialization", "Gender", "Category", "Mobile Number", "Actual Fee", "Fee Paid By Student", "Reason", "Date", "Mode", "UTR Number", "Fee Receipt")]
                
                table += student_obj.prefetch_related('student_id__studentregistrationchecks').filter(semester_id=student_obj[0].semester_id, student_id__studentregistrationchecks__final_registration_flag=True).select_related('student_id', 'student_id__id', 'student_id__id__user', 'student_id__id__department').annotate(
                    admission_year = F('student_id__batch'),
                    semester=F('student_id__curr_semester_no') + 1,
                    roll_no=F('student_id__id__user__username'),
                    full_name=Concat('student_id__id__user__first_name', Value(' '), 'student_id__id__user__last_name'),
                    program=F('student_id__programme'),
                    discipline=F('student_id__id__department__name'),
                    specialization=F('student_id__specialization'),
                    gender=F('student_id__id__sex'),
                    category=F('student_id__category'),
                    phone_no=F('student_id__id__phone_no'),
                    date_deposited=Concat(Cast(ExtractDay('deposit_date'), CharField()), Value('/'),
                                      Cast(ExtractMonth('deposit_date'), CharField()), Value('/'),
                                      Cast(ExtractYear('deposit_date'), CharField()), output_field=CharField())
                    ).values_list('admission_year', 'semester', 'roll_no', 'full_name', 'program', 'discipline', 'specialization', 'gender', 'category', 'phone_no', 'actual_fee', 'fee_paid', 'reason', 'date_deposited', 'mode', 'utr_number', 'fee_receipt').distinct()

                excel_response = BytesIO()
                final_register_workbook = Workbook(excel_response)
                final_register_worksheet = final_register_workbook.add_worksheet()

                for i, row in enumerate(table):
                    for j, cell in enumerate(row):
                        final_register_worksheet.write(i, j, cell)

                final_register_workbook.close()
                excel_response.seek(0)

                response = HttpResponse(excel_response.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="{batch_id.name}_{batch_id.discipline.acronym}_{batch_id.year}_final_registered.xlsx"'
                return response

            else:
                return JsonResponse({'error': 'No registered students found'}, status=404)

    return JsonResponse({'error': 'Invalid request'}, status=400)


@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def course_list(request):
    request_body = json.loads(request.body)
    student_id = request_body['student_id']
    semester_no = request_body['semester_no']

    # final_registration_table = FinalRegistration.objects.all().filter(semester_id = semester_id, verified = False)
    # final = final_registration_table.filter(student_id = student_id, semester_id = semester_id)
    final = FinalRegistration.objects.all().filter(semester_id__semester_no = semester_no, student_id__id=student_id, verified = False)
    if final.exists():
        final_registration = serializers.FinalRegistrationSerializer(final, many=True).data
    else:
        final_registration = None
    resp = {
        'final_registration': final_registration,
    }
    return Response(data=resp, status=status.HTTP_200_OK)

@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def dropcourseadmin(request):
    try:
        reg_id = request.data.get('id')
        roll_no = request.data.get('roll_no')

        if not reg_id or not roll_no:
            return JsonResponse({'error': 'Missing registration ID or roll number'}, status=400)

        reg_id = int(reg_id)
        course_registration.objects.filter(id=reg_id).delete()

        return JsonResponse({'message': 'Success!'}, status=200)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def acad_add_course(request):
    data = request.data
    for fld in ("roll_no", "semester_id", "courseslot_id", "course_id", "academic_year", "registration_type", "semester_type"):
        if not data.get(fld):
            return Response({ "error": f"{fld} is required" }, status=status.HTTP_400_BAD_REQUEST)
    student = get_object_or_404(Student, id=data["roll_no"].upper())
    semester = get_object_or_404(Semester,   id=data["semester_id"])
    slot     = get_object_or_404(CourseSlot, id=data["courseslot_id"])
    course   = get_object_or_404(Courses,     id=data["course_id"])
    session  = data["academic_year"]
    reg_type = data["registration_type"]
    old_id   = data.get("old_course")
    sem_type = data["semester_type"]
    working_year = parse_academic_year(academic_year=session, semester_type=sem_type)[0]

    # Resolve the offering (section). Prefer the chosen section, else the student's
    # own. Registering into a section other than the student's own is only allowed
    # for Backlog/Improvement.
    from applications.academic_information.models import resolve_offering
    ci_id = data.get("course_instructor_id")
    if ci_id:
        offering = CourseInstructor.objects.filter(
            id=ci_id, course_id=course, year=working_year, semester_type=sem_type,
        ).first()
        if not offering:
            return Response({"error": "Selected section is not running this course this semester."},
                            status=status.HTTP_400_BAD_REQUEST)
    else:
        offering = resolve_offering(student, course, working_year, sem_type)

    # A sectioned course can't be registered without a chosen section.
    if offering is None:
        sectioned = CourseInstructor.objects.filter(
            course_id=course, year=working_year, semester_type=sem_type,
        ).exclude(section_label__isnull=True).exclude(section_label="").exists()
        if sectioned:
            return Response(
                {"error": "This course runs in sections — select a section to register in."},
                status=status.HTTP_400_BAD_REQUEST,
            )

    student_section = getattr(student, "section", None)
    if (offering is not None and offering.section_label
            and offering.section_label != student_section
            and reg_type not in ("Backlog", "Improvement")):
        return Response(
            {"error": "Registering into another section is only allowed as Backlog/Improvement."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    with transaction.atomic():
        cr = course_registration.objects.create(
            student_id       = student,
            semester_id      = semester,
            course_slot_id   = slot,
            course_id        = course,
            session          = session,
            registration_type= reg_type,
            semester_type = sem_type,
            working_year = working_year,
            course_instructor = offering,
        )
        if old_id:
            old = course_registration.objects.filter(id=old_id, student_id=student).first()
            if old:
                course_replacement.objects.create(
                    old_course_registration = old,
                    new_course_registration = cr,
                )

    return Response({ "message": "Course added successfully" }, status=status.HTTP_200_OK)


# ===========================================================================
# Admin manual "Add Thesis / Progress Seminar / Teaching Credit" for one
# student -- the same manual-override spirit as acad_add_course above, and
# following the exact same Slot -> Course two-step shape (get_add_course_slots
# / get_add_course_courses / acad_add_course), just against the three
# slot-based PhD/PG registration models instead of CourseSlot/Courses. Each
# of these skips the eligibility gate the student self-registration flow
# enforces (dean-approved topic for Thesis/Seminar, comprehensive exam passed
# for Teaching Credit) since admin is trusted to only use this when
# appropriate, and lands the registration as already 'verified' -- same as
# acad_add_course creating a course_registration directly rather than a
# pending request.
# ===========================================================================

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def get_thesis_slots(request):
    sem_id = request.query_params.get('semester_id')
    if not sem_id:
        return Response({'error': 'semester_id query parameter is required'}, status=status.HTTP_400_BAD_REQUEST)
    get_object_or_404(Semester, id=sem_id)
    slots = ThesisSlot.objects.filter(semester_id=sem_id).values('id', 'name', 'evaluation_type')
    return Response(list(slots), status=status.HTTP_200_OK)


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def get_thesis_courses(request):
    slot_id = request.query_params.get('slot_id')
    if not slot_id:
        return Response({'error': 'slot_id query parameter is required'}, status=status.HTTP_400_BAD_REQUEST)
    slot = get_object_or_404(ThesisSlot, id=slot_id)
    return Response(list(slot.theses.all().values('id', 'code', 'name', 'credit')), status=status.HTTP_200_OK)


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def get_progress_seminar_slots(request):
    sem_id = request.query_params.get('semester_id')
    if not sem_id:
        return Response({'error': 'semester_id query parameter is required'}, status=status.HTTP_400_BAD_REQUEST)
    get_object_or_404(Semester, id=sem_id)
    slots = ProgressSeminarSlot.objects.filter(semester_id=sem_id).values('id', 'name')
    return Response(list(slots), status=status.HTTP_200_OK)


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def get_progress_seminar_courses(request):
    slot_id = request.query_params.get('slot_id')
    if not slot_id:
        return Response({'error': 'slot_id query parameter is required'}, status=status.HTTP_400_BAD_REQUEST)
    slot = get_object_or_404(ProgressSeminarSlot, id=slot_id)
    return Response(list(slot.seminars.all().values('id', 'code', 'name', 'credit')), status=status.HTTP_200_OK)


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def get_teaching_credit_slots(request):
    sem_id = request.query_params.get('semester_id')
    if not sem_id:
        return Response({'error': 'semester_id query parameter is required'}, status=status.HTTP_400_BAD_REQUEST)
    get_object_or_404(Semester, id=sem_id)
    slots = TeachingCreditSlot.objects.filter(semester_id=sem_id).values('id', 'name')
    return Response(list(slots), status=status.HTTP_200_OK)


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def get_teaching_credit_courses(request):
    slot_id = request.query_params.get('slot_id')
    if not slot_id:
        return Response({'error': 'slot_id query parameter is required'}, status=status.HTTP_400_BAD_REQUEST)
    slot = get_object_or_404(TeachingCreditSlot, id=slot_id)
    return Response(list(slot.teaching_credits.all().values('id', 'code', 'name', 'credit')), status=status.HTTP_200_OK)


@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def admin_add_thesis(request):
    roll_no = request.data.get('roll_no')
    semester_id = request.data.get('semester_id')
    thesis_slot_id = request.data.get('thesis_slot_id')
    thesis_id = request.data.get('thesis_id')
    if not roll_no or not semester_id or not thesis_slot_id or not thesis_id:
        return Response({'error': 'roll_no, semester_id, thesis_slot_id and thesis_id are required'}, status=status.HTTP_400_BAD_REQUEST)

    student = get_object_or_404(Student, id=roll_no.upper())
    semester = get_object_or_404(Semester, id=semester_id)
    thesis_slot = get_object_or_404(ThesisSlot, id=thesis_slot_id)
    thesis = get_object_or_404(thesis_slot.theses, id=thesis_id)

    if ThesisRegistration.objects.filter(student=student, semester=semester).exists():
        return Response({'error': 'Student is already registered for this semester'}, status=status.HTTP_400_BAD_REQUEST)

    ALLOWED_THESIS_CREDITS = [3, 6, 9, 12]
    try:
        credits = int(request.data.get('credits', 6))
    except (TypeError, ValueError):
        credits = 6
    if credits not in ALLOWED_THESIS_CREDITS:
        return Response({'error': f'Invalid credit value. Choose from {ALLOWED_THESIS_CREDITS}'}, status=status.HTTP_400_BAD_REQUEST)

    # PG students have a fixed credit value per evaluation_type -- no free
    # choice: 3 for a block-graded (S/X) semester, 12 for the decimal-graded
    # semester. Enforced here too (not just in the frontend's locked Select)
    # since this endpoint is reachable directly.
    if _student_programme_category(student) == 'PG':
        expected = 3 if thesis_slot.evaluation_type == 'blocks_sx' else 12
        if credits != expected:
            return Response({'error': f'PG students in a {thesis_slot.evaluation_type} slot must register for {expected} credits'}, status=status.HTTP_400_BAD_REQUEST)

    now = _dt.datetime.now()
    session = f"{now.year}-{str(now.year + 1)[2:]}" if now.month >= 7 else f"{now.year - 1}-{str(now.year)[2:]}"
    topic = ThesisTopic.objects.filter(student=student).order_by('-created_at').first()

    with transaction.atomic():
        reg = ThesisRegistration.objects.create(
            student=student, thesis_slot=thesis_slot, thesis=thesis, thesis_topic=topic,
            semester=semester, credits=credits, working_year=now.year,
            academic_session=session, status='verified', verified_on=now,
        )
        # Mirrors admin_verify_enrollments' side-effect -- see that function.
        if thesis_slot.evaluation_type == 'decimal':
            evaluation, _created = ThesisEvaluation.objects.get_or_create(registration=reg, block_number=1)
            ThesisEvaluationScore.objects.get_or_create(evaluation=evaluation)
        else:
            for blk in range(1, credits // 3 + 1):
                ThesisEvaluation.objects.get_or_create(registration=reg, block_number=blk)

    return Response({
        'id': reg.id, 'student': roll_no.upper(), 'semester_no': semester.semester_no,
        'thesis': thesis.code, 'credits': credits, 'status': reg.status,
    }, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def admin_add_progress_seminar(request):
    roll_no = request.data.get('roll_no')
    semester_id = request.data.get('semester_id')
    seminar_slot_id = request.data.get('seminar_slot_id')
    seminar_id = request.data.get('seminar_id')
    if not roll_no or not semester_id or not seminar_slot_id or not seminar_id:
        return Response({'error': 'roll_no, semester_id, seminar_slot_id and seminar_id are required'}, status=status.HTTP_400_BAD_REQUEST)

    student = get_object_or_404(Student, id=roll_no.upper())
    semester = get_object_or_404(Semester, id=semester_id)
    slot = get_object_or_404(ProgressSeminarSlot, id=seminar_slot_id)
    seminar = get_object_or_404(slot.seminars, id=seminar_id)

    if ProgressSeminarRegistration.objects.filter(student=student, semester=semester).exists():
        return Response({'error': 'Student is already registered for this semester'}, status=status.HTTP_400_BAD_REQUEST)

    reg = ProgressSeminarRegistration.objects.create(
        student=student, progress_seminar_slot=slot, seminar=seminar, semester=semester,
        working_year=_dt.datetime.now().year, status='verified',
    )
    return Response({
        'id': reg.id, 'student': roll_no.upper(), 'semester_no': semester.semester_no,
        'seminar': seminar.code, 'status': reg.status,
    }, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def admin_add_teaching_credit(request):
    roll_no = request.data.get('roll_no')
    semester_id = request.data.get('semester_id')
    teaching_credit_slot_id = request.data.get('teaching_credit_slot_id')
    teaching_credit_id = request.data.get('teaching_credit_id')
    if not roll_no or not semester_id or not teaching_credit_slot_id or not teaching_credit_id:
        return Response({'error': 'roll_no, semester_id, teaching_credit_slot_id and teaching_credit_id are required'}, status=status.HTTP_400_BAD_REQUEST)

    student = get_object_or_404(Student, id=roll_no.upper())
    semester = get_object_or_404(Semester, id=semester_id)
    slot = get_object_or_404(TeachingCreditSlot, id=teaching_credit_slot_id)
    teaching_credit = get_object_or_404(slot.teaching_credits, id=teaching_credit_id)

    if TeachingCreditRegistration.objects.filter(student=student, semester=semester).exists():
        return Response({'error': 'Student is already registered for this semester'}, status=status.HTTP_400_BAD_REQUEST)

    now = _dt.datetime.now()
    session = f"{now.year}-{str(now.year + 1)[2:]}" if now.month >= 7 else f"{now.year - 1}-{str(now.year)[2:]}"

    reg = TeachingCreditRegistration.objects.create(
        student=student, teaching_credit_slot=slot, teaching_credit=teaching_credit, semester=semester,
        working_year=now.year, academic_session=session, status='verified',
    )
    return Response({
        'id': reg.id, 'student': roll_no.upper(), 'semester_no': semester.semester_no,
        'teaching_credit': teaching_credit.code, 'status': reg.status,
    }, status=status.HTTP_201_CREATED)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def academic_procedures_faculty_api(request):
    try:
        # Ensure the user is a faculty member
        if request.user.extrainfo.user_type != 'faculty':
            return Response({"error": "Unauthorized access. Faculty only."}, status=403)

        user_details = ExtraInfo.objects.select_related("department").get(user=request.user)

        # Get courses taught by the faculty
        courses = CourseInstructor.objects.filter(instructor_id=user_details.id).select_related("course_id")

        current_year = timezone.now().year
        response_data = []

        for course in courses:
            # Calculate academic year from calendar year + semester
            
            response_data.append({
                "course_id": course.course_id.id,
                "course_code": course.course_id.code,
                "course_name": course.course_id.name,
                "version": course.course_id.version,
                "semester_type": course.semester_type,
                "academic_year": course.academic_year,
                # This offering's section (A-F), or "" for old/single-offering courses.
                "section": course.section_label or "",
            })

        return Response({"assigned_courses": response_data})

    except Exception as e:
        return Response({"error": str(e)}, status=500)

@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def search_preregistration(request):
    try:
        roll_no=request.data.get("roll_no")
        sem_no=request.data.get("sem_no")
        looked_up = Student.objects.filter(id_id=roll_no).first()
        if looked_up and not student_in_scope(looked_up, scopes_for(request.user)):
            return JsonResponse({'error': 'Student record not found'}, status=400)
        initial_registrations = InitialRegistration.objects.filter(
            student_id_id=roll_no, semester_id__semester_no=sem_no
        )
        student_registration_check = StudentRegistrationChecks.objects.filter(
            student_id_id=roll_no, semester_id__semester_no=sem_no
        ).first()
        initial = serializers.InitialRegistrationSerializer(initial_registrations, many=True)
        student_registration_check_data = serializers.StudentRegistrationChecksSerializer(student_registration_check)
        return Response({
            "initial_registration": initial.data,  # Send serialized initial registrations
            "student_registration_check": student_registration_check_data.data if student_registration_check else None
        })
    except Exception as e:
        return Response(
            {"error": f"An error occurred: {str(e)}"}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    

@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def delete_preregistration(request):
    try:
        # Extract roll_no and sem_no from the request
        roll_no = request.data.get("roll_no")
        sem_no = request.data.get("sem_no")
        target = Student.objects.filter(id_id=roll_no).first()
        if target and not student_in_scope(target, scopes_for(request.user)):
            return JsonResponse({'error': 'Student record not found'}, status=400)

        # Validate input data
        if not roll_no or not sem_no:
            return Response(
                {"error": "Both roll_no and sem_no are required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Delete initial registration entries
        initial_registrations = InitialRegistration.objects.filter(
            student_id_id=roll_no, semester_id__semester_no=sem_no
        )
        initial_count = initial_registrations.delete()

        # Delete student registration check entries
        student_registration_check = StudentRegistrationChecks.objects.filter(
            student_id_id=roll_no, semester_id__semester_no=sem_no
        )
        student_registration_check_count = student_registration_check.delete()

        # Return a success response with counts of deleted entries
        return Response({
            "message": "Successfully Deleted."
        })

    except Exception as e:
        return Response(
            {"error": f"An error occurred: {str(e)}"}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def allot_courses(request):
    if 'allotedCourses' not in request.FILES:
        return Response({'error': 'Excel file not provided.'},
                        status=status.HTTP_400_BAD_REQUEST)

    batch_id = request.data.get('batch')
    if batch_id and not batch_in_scope(
            Batch.objects.filter(id=batch_id).first(), scopes_for(request.user)):
        return Response({'detail': 'Batch not found.'},
                        status=status.HTTP_404_NOT_FOUND)
    sem_no = request.data.get('semester')
    sem_type = request.data.get('semester_type')
    academic_year = request.data.get('academic_year')
    working_year, _ = parse_academic_year(academic_year=academic_year, semester_type=sem_type)
    from applications.academic_information.models import resolve_offering

    if not all([batch_id, sem_no, sem_type, academic_year]):
        return Response({'error': 'Missing required fields.'},
                        status=status.HTTP_400_BAD_REQUEST)

    try:
        sem_no = int(sem_no)
    except ValueError:
        return Response({'error': 'Semester must be integer.'},
                        status=status.HTTP_400_BAD_REQUEST)

    try:
        with transaction.atomic():
            batch = Batch.objects.get(id=batch_id)
            sem = Semester.objects.get(
                curriculum=batch.curriculum,
                semester_no=sem_no
            )
            book = xlrd.open_workbook(file_contents=request.FILES['allotedCourses'].read())
            sheet = book.sheet_by_index(0)

            checks, pre_regs, final_regs, course_regs = [], [], [], []
            row_errors = []
            seen = set()

            for i in range(1, sheet.nrows):

                try:
                    roll_no = str(sheet.cell_value(i,0)).split('.')[0].strip()
                    slot_name = sheet.cell_value(i,1).strip()
                    code = sheet.cell_value(i,2).strip()

                    # user = User.objects.get(username=roll_no)
                    student = Student.objects.get(id__user__username=roll_no)
                    slot = CourseSlot.objects.get(name=slot_name, semester=sem)
                    course = slot.courses.get(code=code)
                    if roll_no not in seen:
                        checks.append(StudentRegistrationChecks(
                            student_id=student,
                            semester_id=sem,
                            pre_registration_flag=True,
                            final_registration_flag=True
                        ))
                        seen.add(roll_no)

                    pre_regs.append(InitialRegistration(
                        student_id=student,
                        course_slot_id=slot,
                        course_id=course,
                        semester_id=sem,
                        priority=1
                    ))
                    final_regs.append(FinalRegistration(
                        student_id=student,
                        course_slot_id=slot,
                        course_id=course,
                        semester_id=sem,
                        verified=True
                    ))
                    offering = resolve_offering(student, course, working_year, sem_type)
                    course_regs.append(course_registration(
                        session=academic_year,
                        working_year = working_year,
                        course_id=course,
                        semester_id=sem,
                        student_id=student,
                        course_slot_id=slot,
                        semester_type = sem_type,
                        course_instructor = offering
                    ))
                except Exception as e:
                    row_errors.append({
                        "row": i + 1,
                        "roll_no": str(sheet.cell_value(i, 0)).strip() if sheet.ncols > 0 else "",
                        "slot": str(sheet.cell_value(i, 1)).strip() if sheet.ncols > 1 else "",
                        "course_code": str(sheet.cell_value(i, 2)).strip() if sheet.ncols > 2 else "",
                        "error": str(e)
                    })

            inserted = len(course_regs)
            if inserted == 0:
                return Response(
                    {
                        'error': 'No valid rows were found in the uploaded file.',
                        'failed_rows': row_errors[:25],
                        'failed_rows_count': len(row_errors)
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )

            StudentRegistrationChecks.objects.bulk_create(checks, ignore_conflicts=True)
            InitialRegistration.objects.bulk_create(pre_regs, ignore_conflicts=True)
            FinalRegistration.objects.bulk_create(final_regs, ignore_conflicts=True)
            course_registration.objects.bulk_create(course_regs, ignore_conflicts=True)

        if row_errors:
            return Response({
                'message': 'Upload completed with partial success.',
                'inserted_rows': inserted,
                'failed_rows_count': len(row_errors),
                'failed_rows': row_errors[:25]
            }, status=status.HTTP_207_MULTI_STATUS)

        return Response({'message': 'Successfully uploaded!', 'inserted_rows': inserted})
    except Batch.DoesNotExist:
        return Response({'error': 'Invalid batch id.'}, status=status.HTTP_400_BAD_REQUEST)
    except Semester.DoesNotExist:
        return Response({'error': 'Invalid semester or type.'}, status=status.HTTP_400_BAD_REQUEST)
    except xlrd.XLRDError:
        return Response({'error': 'Invalid Excel format.'}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({'error': f'Processing error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ===========================================================================
# Bulk "Allot Thesis / Progress Seminar / Teaching Credit" -- same manual-
# override spirit as allot_courses above (one Excel row per student, lands
# directly as 'verified'), for the three slot-based PhD/PG registration
# models. Same 4-column row shape as allot_courses (RollNo | Slot | Code |
# Name), just against ThesisSlot/SeminarSlot/TeachingCreditSlot + their
# catalog entries instead of CourseSlot/Courses.
# ===========================================================================

@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def allot_thesis(request):
    if 'allotedThesis' not in request.FILES:
        return Response({'error': 'Excel file not provided.'}, status=status.HTTP_400_BAD_REQUEST)

    batch_id = request.data.get('batch')
    if batch_id and not batch_in_scope(
            Batch.objects.filter(id=batch_id).first(), scopes_for(request.user)):
        return Response({'detail': 'Batch not found.'},
                        status=status.HTTP_404_NOT_FOUND)
    sem_no = request.data.get('semester')
    if not batch_id or not sem_no:
        return Response({'error': 'Missing required fields.'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        sem_no = int(sem_no)
    except ValueError:
        return Response({'error': 'Semester must be integer.'}, status=status.HTTP_400_BAD_REQUEST)

    now = _dt.datetime.now()
    session = f"{now.year}-{str(now.year + 1)[2:]}" if now.month >= 7 else f"{now.year - 1}-{str(now.year)[2:]}"
    ALLOWED_THESIS_CREDITS = [3, 6, 9, 12]

    try:
        with transaction.atomic():
            batch = Batch.objects.get(id=batch_id)
            semester = Semester.objects.get(curriculum=batch.curriculum, semester_no=sem_no)

            book = xlrd.open_workbook(file_contents=request.FILES['allotedThesis'].read())
            sheet = book.sheet_by_index(0)

            inserted = 0
            row_errors = []
            for i in range(1, sheet.nrows):
                roll_no = str(sheet.cell_value(i, 0)).split('.')[0].strip()
                if not roll_no:
                    continue
                try:
                    slot_name = sheet.cell_value(i, 1).strip()
                    code = sheet.cell_value(i, 2).strip()

                    student = Student.objects.get(id__user__username=roll_no)
                    thesis_slot = ThesisSlot.objects.get(name=slot_name, semester=semester)
                    thesis = thesis_slot.theses.get(code=code)

                    if ThesisRegistration.objects.filter(student=student, semester=semester).exists():
                        raise ValueError('Already registered for this semester')

                    raw = sheet.cell_value(i, 4) if sheet.ncols > 4 else ''
                    credits = int(raw) if raw != '' else 6
                    if credits not in ALLOWED_THESIS_CREDITS:
                        raise ValueError(f'Invalid credits {credits}; choose from {ALLOWED_THESIS_CREDITS}')

                    if _student_programme_category(student) == 'PG':
                        expected = 3 if thesis_slot.evaluation_type == 'blocks_sx' else 12
                        if credits != expected:
                            raise ValueError(f'PG students in a {thesis_slot.evaluation_type} slot must register for {expected} credits')

                    topic = ThesisTopic.objects.filter(student=student).order_by('-created_at').first()
                    reg = ThesisRegistration.objects.create(
                        student=student, thesis_slot=thesis_slot, thesis=thesis, thesis_topic=topic,
                        semester=semester, credits=credits, working_year=now.year,
                        academic_session=session, status='verified', verified_on=now,
                    )
                    if thesis_slot.evaluation_type == 'decimal':
                        evaluation, _created = ThesisEvaluation.objects.get_or_create(registration=reg, block_number=1)
                        ThesisEvaluationScore.objects.get_or_create(evaluation=evaluation)
                    else:
                        for blk in range(1, credits // 3 + 1):
                            ThesisEvaluation.objects.get_or_create(registration=reg, block_number=blk)
                    inserted += 1
                except Exception as e:
                    row_errors.append({'row': i + 1, 'roll_no': roll_no, 'error': str(e)})

            if inserted == 0:
                return Response({
                    'error': 'No valid rows were found in the uploaded file.',
                    'failed_rows': row_errors[:25], 'failed_rows_count': len(row_errors),
                }, status=status.HTTP_400_BAD_REQUEST)

        if row_errors:
            return Response({
                'message': 'Upload completed with partial success.',
                'inserted_rows': inserted, 'failed_rows_count': len(row_errors),
                'failed_rows': row_errors[:25],
            }, status=status.HTTP_207_MULTI_STATUS)
        return Response({'message': 'Successfully uploaded!', 'inserted_rows': inserted})
    except Batch.DoesNotExist:
        return Response({'error': 'Invalid batch id.'}, status=status.HTTP_400_BAD_REQUEST)
    except Semester.DoesNotExist:
        return Response({'error': 'Invalid semester.'}, status=status.HTTP_400_BAD_REQUEST)
    except xlrd.XLRDError:
        return Response({'error': 'Invalid Excel format.'}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({'error': f'Processing error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def allot_progress_seminar(request):
    if 'allotedSeminar' not in request.FILES:
        return Response({'error': 'Excel file not provided.'}, status=status.HTTP_400_BAD_REQUEST)

    batch_id = request.data.get('batch')
    if batch_id and not batch_in_scope(
            Batch.objects.filter(id=batch_id).first(), scopes_for(request.user)):
        return Response({'detail': 'Batch not found.'},
                        status=status.HTTP_404_NOT_FOUND)
    sem_no = request.data.get('semester')
    if not batch_id or not sem_no:
        return Response({'error': 'Missing required fields.'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        sem_no = int(sem_no)
    except ValueError:
        return Response({'error': 'Semester must be integer.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        with transaction.atomic():
            batch = Batch.objects.get(id=batch_id)
            semester = Semester.objects.get(curriculum=batch.curriculum, semester_no=sem_no)

            book = xlrd.open_workbook(file_contents=request.FILES['allotedSeminar'].read())
            sheet = book.sheet_by_index(0)

            inserted = 0
            row_errors = []
            for i in range(1, sheet.nrows):
                roll_no = str(sheet.cell_value(i, 0)).split('.')[0].strip()
                if not roll_no:
                    continue
                try:
                    slot_name = sheet.cell_value(i, 1).strip()
                    code = sheet.cell_value(i, 2).strip()

                    student = Student.objects.get(id__user__username=roll_no)
                    slot = ProgressSeminarSlot.objects.get(name=slot_name, semester=semester)
                    seminar = slot.seminars.get(code=code)

                    if ProgressSeminarRegistration.objects.filter(student=student, semester=semester).exists():
                        raise ValueError('Already registered for this semester')
                    ProgressSeminarRegistration.objects.create(
                        student=student, progress_seminar_slot=slot, seminar=seminar, semester=semester,
                        working_year=_dt.datetime.now().year, status='verified',
                    )
                    inserted += 1
                except Exception as e:
                    row_errors.append({'row': i + 1, 'roll_no': roll_no, 'error': str(e)})

            if inserted == 0:
                return Response({
                    'error': 'No valid rows were found in the uploaded file.',
                    'failed_rows': row_errors[:25], 'failed_rows_count': len(row_errors),
                }, status=status.HTTP_400_BAD_REQUEST)

        if row_errors:
            return Response({
                'message': 'Upload completed with partial success.',
                'inserted_rows': inserted, 'failed_rows_count': len(row_errors),
                'failed_rows': row_errors[:25],
            }, status=status.HTTP_207_MULTI_STATUS)
        return Response({'message': 'Successfully uploaded!', 'inserted_rows': inserted})
    except Batch.DoesNotExist:
        return Response({'error': 'Invalid batch id.'}, status=status.HTTP_400_BAD_REQUEST)
    except Semester.DoesNotExist:
        return Response({'error': 'Invalid semester.'}, status=status.HTTP_400_BAD_REQUEST)
    except xlrd.XLRDError:
        return Response({'error': 'Invalid Excel format.'}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({'error': f'Processing error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def allot_teaching_credit(request):
    if 'allotedTeachingCredit' not in request.FILES:
        return Response({'error': 'Excel file not provided.'}, status=status.HTTP_400_BAD_REQUEST)

    batch_id = request.data.get('batch')
    if batch_id and not batch_in_scope(
            Batch.objects.filter(id=batch_id).first(), scopes_for(request.user)):
        return Response({'detail': 'Batch not found.'},
                        status=status.HTTP_404_NOT_FOUND)
    sem_no = request.data.get('semester')
    if not batch_id or not sem_no:
        return Response({'error': 'Missing required fields.'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        sem_no = int(sem_no)
    except ValueError:
        return Response({'error': 'Semester must be integer.'}, status=status.HTTP_400_BAD_REQUEST)

    now = _dt.datetime.now()
    session = f"{now.year}-{str(now.year + 1)[2:]}" if now.month >= 7 else f"{now.year - 1}-{str(now.year)[2:]}"

    try:
        with transaction.atomic():
            batch = Batch.objects.get(id=batch_id)
            semester = Semester.objects.get(curriculum=batch.curriculum, semester_no=sem_no)

            book = xlrd.open_workbook(file_contents=request.FILES['allotedTeachingCredit'].read())
            sheet = book.sheet_by_index(0)

            inserted = 0
            row_errors = []
            for i in range(1, sheet.nrows):
                roll_no = str(sheet.cell_value(i, 0)).split('.')[0].strip()
                if not roll_no:
                    continue
                try:
                    slot_name = sheet.cell_value(i, 1).strip()
                    code = sheet.cell_value(i, 2).strip()

                    student = Student.objects.get(id__user__username=roll_no)
                    slot = TeachingCreditSlot.objects.get(name=slot_name, semester=semester)
                    teaching_credit = slot.teaching_credits.get(code=code)

                    if TeachingCreditRegistration.objects.filter(student=student, semester=semester).exists():
                        raise ValueError('Already registered for this semester')
                    TeachingCreditRegistration.objects.create(
                        student=student, teaching_credit_slot=slot, teaching_credit=teaching_credit, semester=semester,
                        working_year=now.year, academic_session=session, status='verified',
                    )
                    inserted += 1
                except Exception as e:
                    row_errors.append({'row': i + 1, 'roll_no': roll_no, 'error': str(e)})

            if inserted == 0:
                return Response({
                    'error': 'No valid rows were found in the uploaded file.',
                    'failed_rows': row_errors[:25], 'failed_rows_count': len(row_errors),
                }, status=status.HTTP_400_BAD_REQUEST)

        if row_errors:
            return Response({
                'message': 'Upload completed with partial success.',
                'inserted_rows': inserted, 'failed_rows_count': len(row_errors),
                'failed_rows': row_errors[:25],
            }, status=status.HTTP_207_MULTI_STATUS)
        return Response({'message': 'Successfully uploaded!', 'inserted_rows': inserted})
    except Batch.DoesNotExist:
        return Response({'error': 'Invalid batch id.'}, status=status.HTTP_400_BAD_REQUEST)
    except Semester.DoesNotExist:
        return Response({'error': 'Invalid semester.'}, status=status.HTTP_400_BAD_REQUEST)
    except xlrd.XLRDError:
        return Response({'error': 'Invalid Excel format.'}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({'error': f'Processing error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
@role_required(['student'])
def student_next_sem_courses(request):
    """
    REST API endpoint to return the courses_list as JSON.  Uses DRF authentication.
    """

    user_details = ExtraInfo.objects.select_related('user', 'department').get(user=request.user) # Changed to user=request.user
    des = HoldsDesignation.objects.all().select_related().filter(user=request.user).first()

    if str(des.designation) != "student":
        return Response({"error": "User is not a student"}, status=status.HTTP_403_FORBIDDEN)  # 403 Forbidden - DRF style

    obj = Student.objects.select_related('id', 'id__user', 'id__department').get(id=user_details.id)
    
    # Check if PhD student - they don't have semester-based courses like UG/PG
    is_phd_student = obj.programme and obj.programme.upper() == 'PHD'
    if is_phd_student:
        return Response({
            "courses_list": [],
            "message": "PhD students don't follow semester-based course structure"
        }, status=status.HTTP_200_OK)
    
    batch = obj.batch_id
    if not batch:
        return Response({"error": "Student batch not found"}, status=status.HTTP_404_NOT_FOUND)
    
    curr_id = batch.curriculum
    if not curr_id:
        return Response({"error": "Curriculum not found for student batch"}, status=status.HTTP_404_NOT_FOUND)

    try:
        semester_no = obj.curr_semester_no
        sem_no = semester_no + 1
        next_sem_id = Semester.objects.get(curriculum=curr_id, semester_no=sem_no)
    except Semester.DoesNotExist:  # Handle the case where next semester doesn't exist.
        return Response({"error": "Next semester not found"}, status=status.HTTP_404_NOT_FOUND)


    # Serialize the data (using DRF serializers is highly recommended)
    course_slot = CourseSlot.objects.all().filter(semester_id = next_sem_id).prefetch_related(Prefetch('courses', queryset=Courses.objects.all()))
    serializer = serializers.CourseSlotSerializer(course_slot, many=True) # Assuming you have a CourseSerializer
    courses_list_data = serializer.data

    return Response({"courses_list": courses_list_data}, status=status.HTTP_200_OK)

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(['student'])
def course_registration_view(request):
    try:
        current_user = request.user
        user_details = current_user.extrainfo
        student = Student.objects.get(id=user_details)

        # Check if PhD student - they don't have course registrations like UG/PG
        is_phd_student = student.programme and student.programme.upper() == 'PHD'
        if is_phd_student:
            return Response({
                "reg_data": [],
                "sem_no": 1,
                "semester_type": "Odd Semester",
                "message": "PhD students don't follow traditional course registration"
            }, status=status.HTTP_200_OK)

        semester_no = request.query_params.get('semester', student.curr_semester_no)
        semester_type = request.query_params.get('semester_type', 'Even Semester' if student.curr_semester_no%2==0 else 'Odd Semester')
        try:
            semester = Semester.objects.get(curriculum=student.batch_id.curriculum, semester_no=semester_no)
        except Semester.DoesNotExist:
            return JsonResponse({"error": "Semester not found."}, status=404)

        courses = course_registration.objects.filter(student_id=student, semester_id=semester, semester_type=semester_type)

        result = []
        for reg in courses:
            course_data = serializers.CourseRegistrationSerializer(reg).data

            replacements = course_replacement.objects.filter(old_course_registration=reg)
            replaced_by_list = []

            for replacement in replacements:
                new_reg = replacement.new_course_registration
                replaced_by_list.append({
                    "code": new_reg.course_id.code,
                    "name": new_reg.course_id.name,
                    "semester_no": new_reg.semester_id.semester_no,
                    "label" : make_label(new_reg.semester_id.semester_no, new_reg.semester_type)
                })

            course_data["replaced_by"] = replaced_by_list
            result.append(course_data)
        return Response({"reg_data": result, "sem_no": semester_no, "semester_type": semester_type}, status=status.HTTP_200_OK)

    except Student.DoesNotExist:
        return Response({"error": "Student profile not found"}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def _calendar_window_bounds(event):
    # Combine each date with its optional time (whole-day default) -> naive local datetimes.
    start = datetime.datetime.combine(event.from_date, event.from_time or datetime.time.min)
    end = datetime.datetime.combine(event.to_date, event.to_time or datetime.time.max)
    return start, end

def _fmt_window_dt(dt):
    # Human window bound with AM/PM; drops the time part for whole-day (midnight) starts.
    return dt.strftime('%Y-%m-%d') if dt.time() in (datetime.time.min, datetime.time.max) else dt.strftime('%Y-%m-%d %I:%M %p')

def _check_registration_window(description, action_label):
    # None if now is inside the event's window; else a JsonResponse(400). Honors from_time/to_time.
    try:
        event = Calendar.objects.get(description=description)
    except Calendar.DoesNotExist:
        # Name the event that is missing: the window is keyed by an exact
        # calendar description, and "not yet decided" gave the office no way to
        # tell a missing event from one entered under another name.
        return JsonResponse({
            "error": f"{action_label} date is not yet decided",
            "detail": f"No academic calendar event named \"{description}\".",
            "expected_event": description,
        }, status=400)
    start, end = _calendar_window_bounds(event)
    now = datetime.datetime.now()
    if now < start:
        return JsonResponse({"error": f"{action_label} will start from {_fmt_window_dt(start)} to {_fmt_window_dt(end)}"}, status=400)
    if now > end:
        return JsonResponse({"error": f"{action_label} period has ended"}, status=400)
    return None

def get_add_registration_eligibility(current_date, user_sem, year = datetime.datetime.now().year):
    return _check_registration_window(f"Add {user_sem} {year}", "Add course")

def get_drop_registration_eligibility(current_date, user_sem, year = datetime.datetime.now().year):
    return _check_registration_window(f"Drop {user_sem} {year}", "Drop course")

def get_replace_registration_eligibility(current_date, user_sem, year = datetime.datetime.now().year):
    return _check_registration_window(f"Replace {user_sem} {year}", "Replace course")

def pre_registration_target_semester(student):
    """The semester a student's pre-registration belongs to.

    Continuing students choose courses for the semester after the one they are
    in. A newly admitted batch has nothing registered in its first semester yet,
    so it registers into that semester rather than skipping past it.
    """
    current = student.curr_semester_no or 0
    if current <= 1 and not course_registration.objects.filter(
            student_id=student, semester_id__semester_no=current).exists():
        return current
    return current + 1


def get_pre_registration_eligibility(current_date, user_sem, year = datetime.datetime.now().year):
    # Keyed by the semester being registered into, so a newly admitted batch --
    # which registers into the semester it is already in -- opens on
    # "Pre Registration 1", and every later cohort on its own target semester.
    return _check_registration_window(f"Pre Registration {user_sem} {year}", "Pre Registration")

def get_swayam_registration_eligibility(current_date, user_sem, year = datetime.datetime.now().year):
    return _check_registration_window(f"Swayam {user_sem} {year}", "Swayam Registration")

def get_pg_phd_registration_eligibility(current_date, category, user_sem, year=datetime.datetime.now().year):
    # PhD runs Odd AND Even intakes, so the same sem number recurs across terms in one calendar year -> key by current term (from the date) + academic year. PG is single-intake, keyed by sem + year.
    if str(category).upper() == "PHD":
        if current_date.month >= 7:
            term, ay = "Odd", f"{current_date.year}-{str(current_date.year + 1)[-2:]}"
        else:
            term, ay = "Even", f"{current_date.year - 1}-{str(current_date.year)[-2:]}"
        return _check_registration_window(f"PhD Registration {term} {user_sem} {ay}", "PhD registration")
    return _check_registration_window(f"PG Registration {user_sem} {year}", "PG registration")

def get_student_registrtion_check(student, sem):
    return StudentRegistrationChecks.objects.filter(student_id=student, semester_id=sem).first()



def get_student_registrations(student, semester):
    """
    Returns a QuerySet of InitialRegistration entries for the given student and semester.
    
    Args:
        student (Student): The student instance.
        semester (Semester): The semester instance.
        
    Returns:
        QuerySet: Registrations for the student in the given semester.
    """
    return InitialRegistration.objects.filter(student_id=student, semester_id=semester)


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(['student'])
@block_pg_phd
def get_preregistration_data(request):
    """
    Returns the list of course slots available for the student's next semester,
    along with the list of courses available in each slot.
    If the student has already completed pre-registration, returns the registered
    courses with their priorities.
    """
    try:
        current_user = request.user
        user_details = current_user.extrainfo
        student = Student.objects.get(id=user_details)
        semester_no = student.curr_semester_no
        next_sem_no = pre_registration_target_semester(student)
        try:
            next_semester = Semester.objects.get(curriculum=student.batch_id.curriculum, semester_no=next_sem_no)
        except Semester.DoesNotExist:
            return JsonResponse({"error": "Not Eligible for Pre Registration"}, status=400)

        # Check if the student has already completed pre-registration.
        registration_check = get_student_registrtion_check(student, next_semester)
        course_slots = CourseSlot.objects.filter(semester=next_semester)\
            .exclude(name__icontains='SW')
        data = []
        if registration_check and registration_check.pre_registration_flag:
            registrations = get_student_registrations(student, next_semester)
            regular_regs =  registrations.exclude(course_slot_id__name__icontains='BL')
            backlog_regs = registrations.filter(
                course_slot_id__name__startswith='BL'
            )
            # Build a lookup dictionary: {(slot_id, course_id): priority}
            reg_lookup = {
                (reg.course_slot_id.id if reg.course_slot_id else None, reg.course_id.id): reg.priority 
                for reg in regular_regs
            }
            for slot in course_slots:
                courses = slot.courses.all()
                # print(slot.id)
                course_choices = [
                    {
                        "id": course.id,
                        "code": course.code,
                        "name": course.name,
                        "credits": course.credit,
                        "priority": reg_lookup.get((slot.id, course.id), "")
                    }
                    for course in courses
                ]
                data.append({
                    "sno": slot.id,
                    "slot_name": slot.name,
                    "slot_type": slot.type,
                    "semester": next_sem_no,
                    "course_choices": course_choices,
                })

                backlog_data = []
                for reg in backlog_regs:
                    backlog_data.append({
                        "sno": reg.course_slot_id.id if reg.course_slot_id else None,
                        "slot_name": reg.course_slot_id.name if reg.course_slot_id else "Unknown",
                        "course_choices": [{
                            "id": reg.course_id.id,
                            "code": reg.course_id.code,
                            "name": reg.course_id.name
                        }],
                        "prev_registration": {
                            "id": reg.old_course_registration.id if reg.old_course_registration else "",
                            "code": reg.old_course_registration.course_id.code if reg.old_course_registration and reg.old_course_registration.course_id else "",
                            "name": reg.old_course_registration.course_id.name if reg.old_course_registration and reg.old_course_registration.course_id else "",
                            "semester_no": reg.old_course_registration.semester_id.semester_no if reg.old_course_registration and reg.old_course_registration.semester_id else "",
                        }
                    })
            return JsonResponse({"message": "Already registered", "data": data, "backlog_data":backlog_data}, safe=False)
        else:
            # If not already registered, return slots without pre-set priorities.
            eligibility_resp = get_pre_registration_eligibility(timezone.now().date(), next_sem_no)
            if isinstance(eligibility_resp, JsonResponse):
                return eligibility_resp
            prev_registrations = serializers.CourseRegistrationSerializer(course_registration.objects.filter(student_id=student), many=True).data
            for slot in course_slots:
                courses = slot.courses.all()
                course_choices = [
                    {
                        "id": course.id,
                        "code": course.code,
                        "name": course.name,
                        "credits": course.credit
                    }
                    for course in courses
                ]
                data.append({
                    "sno": slot.id,
                    "slot_name": slot.name,
                    "slot_type": slot.type,
                    "semester": next_sem_no,
                    "course_choices": course_choices,
                    "prev_registrations": prev_registrations
                })
            return JsonResponse(data, safe=False)
    except Student.DoesNotExist:
        return Response({"error": "Student profile not found"}, status=400)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['student'])
@block_pg_phd
def submit_preregistration(request):
    """
    Expects a POST request with JSON data containing an array "registrations".
    Each registration entry should include:
      - slot_id: the ID of the CourseSlot
      - course_id: the chosen Course ID for that slot
      - priority: the priority assigned by the student
    If the student has not pre-registered, the registrations will be created.
    """
    if request.method != "POST":
        return JsonResponse({"error": "Invalid method"}, status=400)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return Response({"Invalid JSON"})

    registrations = data.get("registrations", [])
    backlog_registrations = data.get("backlog_registrations", [])

    # Marking pre-registration complete for an empty submission saved nothing and
    # then locked the student out, since the page returns an existing
    # registration before it offers the form again.
    if not registrations and not backlog_registrations:
        return JsonResponse(
            {"error": "Select at least one course before submitting your pre-registration."},
            status=400)

    try:
        current_user = request.user
        user_details = current_user.extrainfo
        student = Student.objects.get(id=user_details)
        semester_no = student.curr_semester_no
        next_sem_no = pre_registration_target_semester(student)
        try:
            next_semester = Semester.objects.get(curriculum=student.batch_id.curriculum, semester_no=next_sem_no)
        except Semester.DoesNotExist:
            return JsonResponse({"error": "Not Eligible for Pre Registration"}, status=400)
        eligibility_resp = get_pre_registration_eligibility(timezone.now().date(), next_sem_no)
        if isinstance(eligibility_resp, JsonResponse):
            return eligibility_resp
    except Student.DoesNotExist:
        return Response({"error": "Student profile not found"}, status=400)

    for reg in registrations:
        slot_id = reg.get("slot_id")
        course_id = reg.get("course_id")
        priority = reg.get("priority")

        InitialRegistration.objects.create(
            course_id_id=course_id,
            semester_id_id=next_semester.id,
            student_id=student,
            course_slot_id_id=slot_id,
            priority=priority,
            timestamp=timezone.now()
        )

    # The replaced registration is taken from the request body, so check it is
    # this student's before any row is written.
    backlog_entries = []
    for reg in backlog_registrations:
        prev_registration_id = reg.get("prev_registration_id")
        prev_reg = None
        if prev_registration_id:
            prev_reg = course_registration.objects.filter(
                id=prev_registration_id, student_id=student
            ).first()
            if not prev_reg:
                return JsonResponse(
                    {"error": "The course being repeated is not one of your registrations."},
                    status=400)
        backlog_entries.append((reg, prev_reg))

    for reg, prev_reg in backlog_entries:
        slot_id = reg.get("slot_id")
        course_id = reg.get("course_id")
        priority = reg.get("priority")

        # The repeated course carries the grade, so it decides backlog vs improvement.
        _sg = latest_grade(student.id_id, prev_reg.course_id) if prev_reg else None
        reg_type = registration_type_for_grade(_sg.grade if _sg else None)

        InitialRegistration.objects.create(
            course_id_id=course_id,
            semester_id_id=next_semester.id,
            student_id=student,
            course_slot_id_id=slot_id,
            priority=priority,
            registration_type=(reg_type if reg_type != 'Regular' else 'Backlog'),
            old_course_registration=prev_reg,
            timestamp=timezone.now()
        )

    # Optionally, update the StudentRegistrationChecks record to mark pre-registration as complete.
    reg_check, created = StudentRegistrationChecks.objects.get_or_create(
        student_id=student, semester_id_id=next_semester.id,
        defaults={'pre_registration_flag': True}
    )
    if not created:
        reg_check.pre_registration_flag = True
        reg_check.save()
        
    return JsonResponse({"status": "success"}, status=201)


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(['student'])
@block_pg_phd
def get_swayam_registration_data(request):
    """
    Returns the list of course slots available for Swayam registration for the student's next semester,
    along with the list of courses available in each slot.
    (Only course slots whose name starts with "SW" are returned.
    Excludes slots already used in pending/approved Extra Credits or Swayam_Replace requests.)
    """
    try:
        current_user = request.user
        user_details = current_user.extrainfo
        student = Student.objects.get(id=user_details)
        semester_no = student.curr_semester_no
        try:
            current_semester = Semester.objects.get(
                curriculum=student.batch_id.curriculum, 
                semester_no=semester_no
            )
        except Semester.DoesNotExist:
            return JsonResponse({"error": "Not Eligible for Swayam Registration"}, status=400)

        eligibility_resp = get_swayam_registration_eligibility(timezone.now().date(), semester_no)
        if isinstance(eligibility_resp, JsonResponse):
            return eligibility_resp
        
        # Exclude slots used in ANY request of either type (all statuses)
        used_slot_ids = list(SwayamReplacementRequest.objects.filter(
            student=student,
            request_type__in=['Extra_Credits', 'Swayam_Replace']
        ).values_list('new_course_slot__id', flat=True))

        registered_slot_ids = list(course_registration.objects.filter(
            student_id=student,
            semester_id=current_semester
        ).exclude(course_slot_id__isnull=True).values_list('course_slot_id__id', flat=True))

        all_excluded_slot_ids = used_slot_ids + registered_slot_ids

        course_slots = CourseSlot.objects.filter(
            semester=current_semester,
            name__startswith="SW"
        ).exclude(id__in=all_excluded_slot_ids)

        used_replace_course_ids = set(SwayamReplacementRequest.objects.filter(
            student=student,
            request_type='Swayam_Replace'
        ).values_list('new_course__id', flat=True))

        data = []
        for slot in course_slots:
            courses = slot.courses.exclude(id__in=used_replace_course_ids)
            course_choices = [
                {
                    "id": course.id,
                    "code": course.code,
                    "name": course.name,
                    "credits": course.credit
                }
                for course in courses
            ]
            data.append({
                "sno": slot.id,
                "slot_name": slot.name,
                "slot_type": slot.type,
                "semester": semester_no,
                "course_choices": course_choices,
            })
        return JsonResponse(data, safe=False)
    except Student.DoesNotExist:
        return Response({"error": "Student profile not found"}, status=404)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(['student'])
@block_pg_phd
def submit_swayam_registration(request):
    """
    Accepts a POST request with JSON data for Swayam Extra Credits registration.
    For each registration entry, a SwayamReplacementRequest record is created with
    request_type='Extra_Credits' and status='Pending' for admin approval.
    """
    if request.method != "POST":
        return JsonResponse({"error": "Invalid method"}, status=400)
    try:
        payload = json.loads(request.body)
    except Exception:
        return JsonResponse({"error": "Invalid JSON"}, status=400)
    try:
        current_user = request.user
        user_details = current_user.extrainfo
        student = Student.objects.get(id=user_details)
        semester_no = student.curr_semester_no
        try:
            semester = Semester.objects.get(
                curriculum=student.batch_id.curriculum, 
                semester_no=semester_no
            )
        except Semester.DoesNotExist:
            return JsonResponse({"error": "Not Eligible for Swayam Registration"}, status=400)
    except Student.DoesNotExist:
        return Response({"error": "Student not found"}, status=404)
    
    eligibility_resp = get_swayam_registration_eligibility(timezone.now().date(), semester_no)
    if isinstance(eligibility_resp, JsonResponse):
        return eligibility_resp

    current_year = timezone.now().year
    academic_year, semester_type = generate_current_session(current_year, semester_no)
    
    registrations = payload.get("registrations", [])
    errors = []
    success_count = 0
    
    for reg in registrations:
        slot_id = reg.get("slot_id")
        course_id = reg.get("course_id")
        
        try:
            course = Courses.objects.get(id=course_id)
            course_slot = CourseSlot.objects.get(id=slot_id)
        except Courses.DoesNotExist:
            errors.append(f"Course with ID {course_id} does not exist")
            continue
        except CourseSlot.DoesNotExist:
            errors.append(f"Course slot with ID {slot_id} does not exist")
            continue

        existing_registration = course_registration.objects.filter(
            course_id=course,
            semester_id=semester,
            student_id=student
        ).first()
        
        if existing_registration:
            errors.append(f"Already registered for course: {course.code} - {course.name}")
            continue
        
        existing_request = SwayamReplacementRequest.objects.filter(
            student=student,
            new_course_slot=course_slot,
            status__in=['Pending', 'Approved']
        ).first()
        
        if existing_request:
            errors.append(f"Already have a pending/approved request for slot {course_slot.name}")
            continue
        
        try:
            SwayamReplacementRequest.objects.create(
                student=student,
                semester=semester,
                academic_year=academic_year,
                semester_type=semester_type,
                request_type='Extra_Credits',
                old_course=None,
                new_course=course,
                course_slot=None,
                new_course_slot=course_slot,
                status='Pending'
            )
            success_count += 1
        except Exception as e:
            errors.append(f"Failed to register for {course.code}: {str(e)}")
    
    if errors and success_count == 0:
        return JsonResponse({"error": ", ".join(errors)}, status=400)
    elif errors:
        return JsonResponse({
            "status": "partial_success",
            "message": f"Submitted {success_count} Extra Credits request(s) successfully. Pending Academic Admin approval.",
            "errors": errors
        }, status=200)
    
    return JsonResponse({"message": f"Extra Credits request submitted successfully. Pending Academic Admin approval."}, status=201)


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(['student'])
@block_pg_phd
def swayam_replace_check(request):
    try:
        current_user = request.user
        user_details = current_user.extrainfo
        student = Student.objects.get(id=user_details)
        semester_no = student.curr_semester_no
        
        try:
            current_semester = Semester.objects.get(
                curriculum=student.batch_id.curriculum,
                semester_no=semester_no
            )
        except Semester.DoesNotExist:
            return JsonResponse({"error": "Invalid semester"}, status=400)

        eligibility_resp = get_swayam_registration_eligibility(timezone.now().date(), semester_no)
        if isinstance(eligibility_resp, JsonResponse):
            return eligibility_resp

        existing_request = SwayamReplacementRequest.objects.filter(
            student=student,
            semester=current_semester,
            request_type='Swayam_Replace',
            status__in=['Pending', 'Approved']
        ).first()

        if existing_request:
            return JsonResponse({
                "has_existing": False,
                "has_pending_request": True,
                "request_status": existing_request.status,
            })
        
        existing_sw = course_registration.objects.filter(
            student_id=student,
            semester_id=current_semester,
            course_id__code__startswith="SW",
            course_slot_id__name__startswith="OE"
        ).select_related('course_id', 'course_slot_id').first()
        
        if existing_sw:
            all_sw_slots = CourseSlot.objects.filter(
                semester=current_semester,
                name__startswith="SW"
            )
            all_sw_slots = all_sw_slots.exclude(id=existing_sw.course_slot_id.id)

            # Exclude slots used in ANY request of either type (all statuses)
            used_slot_ids = list(SwayamReplacementRequest.objects.filter(
                student=student,
                request_type__in=['Swayam_Replace', 'Extra_Credits']
            ).values_list('new_course_slot__id', flat=True))

            all_sw_slots = all_sw_slots.exclude(id__in=used_slot_ids)

            target_slots = [
                {
                    "id": slot.id,
                    "name": slot.name
                }
                for slot in all_sw_slots
            ]
            
            return JsonResponse({
                "has_existing": True,
                "has_pending_request": False,
                "is_current_semester": True,
                "single_slot_allowed": True,
                "existing_course": {
                    "id": existing_sw.course_id.id,
                    "code": existing_sw.course_id.code,
                    "name": existing_sw.course_id.name,
                    "semester": existing_sw.semester_id.semester_no,
                    "slot": existing_sw.course_slot_id.name,
                    "slot_id": existing_sw.course_slot_id.id,
                    "credits": existing_sw.course_id.credit
                },
                "target_slots": target_slots
            })

        available_semesters = []
        # Only show semesters up to current_semester - 1
        if semester_no >= 3:
            for sem_no in range(3, semester_no):
                try:
                    sem = Semester.objects.get(
                        curriculum=student.batch_id.curriculum,
                        semester_no=sem_no
                    )

                    has_oe_courses = course_registration.objects.filter(
                        student_id=student,
                        semester_id=sem,
                        course_slot_id__name__startswith="OE"
                    ).exists()
                    
                    if has_oe_courses:
                        available_semesters.append({
                            "semester_no": sem_no,
                            "label": f"Semester {sem_no}"
                        })
                except Semester.DoesNotExist:
                    continue

        all_sw_slots = CourseSlot.objects.filter(
            semester=current_semester,
            name__startswith="SW"
        )

        has_sw_slots = all_sw_slots.exists()
        
        if has_sw_slots:
            used_slot_ids = list(SwayamReplacementRequest.objects.filter(
                student=student,
                request_type__in=['Swayam_Replace', 'Extra_Credits']
            ).values_list('new_course_slot__id', flat=True))

            all_sw_slots = all_sw_slots.exclude(id__in=used_slot_ids)

            target_slots = [
                {
                    "id": slot.id,
                    "name": slot.name
                }
                for slot in all_sw_slots
            ]
        else:
            target_slots = []
        
        return JsonResponse({
            "has_existing": False,
            "has_pending_request": False,
            "is_current_semester": False,
            "single_slot_allowed": False,
            "available_semesters": available_semesters,
            "target_slots": target_slots,
            "has_sw_in_current": has_sw_slots,
            "message": "No auto-lockable course found. Please select semester, slot, and course manually."
        })
        
    except Student.DoesNotExist:
        return JsonResponse({"error": "Student profile not found. Please contact the administrator."}, status=404)
    except Exception as e:
        logger.error(f"Error in swayam_replace_check: {str(e)}", exc_info=True)
        return JsonResponse({
            "error": "Unable to check Swayam replacement eligibility. Please contact the administrator if this issue persists."
        }, status=500)


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(['student'])
@block_pg_phd
def swayam_replace_slots(request):
    try:
        semester_no = request.GET.get('semester_no')
        if not semester_no:
            return JsonResponse({"error": "semester_no parameter required"}, status=400)
        
        semester_no = int(semester_no)
        current_user = request.user
        user_details = current_user.extrainfo
        student = Student.objects.get(id=user_details)
        
        try:
            semester = Semester.objects.get(
                curriculum=student.batch_id.curriculum,
                semester_no=semester_no
            )
        except Semester.DoesNotExist:
            return JsonResponse({"error": "Invalid semester"}, status=400)

        registered_slots = course_registration.objects.filter(
            student_id=student,
            semester_id=semester,
            course_slot_id__name__startswith="OE"
        ).values_list('course_slot_id', flat=True).distinct()
        
        slots = CourseSlot.objects.filter(
            id__in=registered_slots
        ).values('id', 'name')
        
        return JsonResponse(list(slots), safe=False)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(['student'])
@block_pg_phd
def swayam_target_slots(request):
    try:
        semester_no = request.GET.get('semester_no')
        if not semester_no:
            return JsonResponse({"error": "semester_no parameter required"}, status=400)
        
        semester_no = int(semester_no)
        current_user = request.user
        user_details = current_user.extrainfo
        student = Student.objects.get(id=user_details)
        
        try:
            semester = Semester.objects.get(
                curriculum=student.batch_id.curriculum,
                semester_no=semester_no
            )
        except Semester.DoesNotExist:
            return JsonResponse({"error": "Invalid semester"}, status=400)
        
        registered_sw_slots = course_registration.objects.filter(
            student_id=student,
            semester_id=semester,
            course_id__code__startswith="SW"
        ).values_list('course_slot_id', flat=True).distinct()
        
        slots = CourseSlot.objects.filter(
            id__in=registered_sw_slots
        ).values('id', 'name')
        
        return JsonResponse(list(slots), safe=False)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(['student'])
@block_pg_phd
def swayam_replace_courses(request):
    try:
        slot_id = request.GET.get('slot_id')
        semester_no = request.GET.get('semester_no')
        
        if not slot_id:
            return JsonResponse({"error": "slot_id parameter required"}, status=400)
        if not semester_no:
            return JsonResponse({"error": "semester_no parameter required"}, status=400)
        
        semester_no = int(semester_no)
        current_user = request.user
        user_details = current_user.extrainfo
        student = Student.objects.get(id=user_details)
        
        try:
            semester = Semester.objects.get(
                curriculum=student.batch_id.curriculum,
                semester_no=semester_no
            )
        except Semester.DoesNotExist:
            return JsonResponse({"error": "Invalid semester"}, status=400)

        registrations = course_registration.objects.filter(
            student_id=student,
            semester_id=semester,
            course_slot_id=slot_id
        ).select_related('course_id')

        # Exclude courses where student already has a satisfactory grade (not eligible for replacement)
        # Grades are stored in online_cms_student_grades (Student_grades model) by roll_no (username)
        blocked_grades = ['O', 'A+', 'A', 'B+', 'B', 'C+']
        roll_no = current_user.username
        courses_list = []
        for reg in registrations:
            course = reg.course_id
            # Check if this course has a published good grade in Student_grades
            has_good_grade = Student_grades.objects.filter(
                roll_no=roll_no,
                course_id=course,
                grade__in=blocked_grades
            ).exists()
            if has_good_grade:
                continue  # Skip — this course has a satisfactory grade and cannot be replaced
            courses_list.append({
                'id': course.id,
                'code': course.code,
                'name': course.name,
                'credit': course.credit
            })
        
        return JsonResponse(courses_list, safe=False)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(['student'])
@block_pg_phd
def swayam_target_courses(request):
    try:
        slot_id = request.GET.get('slot_id')
        semester_no = request.GET.get('semester_no')
        
        if not slot_id:
            return JsonResponse({"error": "slot_id parameter required"}, status=400)
        if not semester_no:
            return JsonResponse({"error": "semester_no parameter required"}, status=400)
        
        semester_no = int(semester_no)
        current_user = request.user
        user_details = current_user.extrainfo
        student = Student.objects.get(id=user_details)
        
        try:
            semester = Semester.objects.get(
                curriculum=student.batch_id.curriculum,
                semester_no=semester_no
            )
        except Semester.DoesNotExist:
            return JsonResponse({"error": "Invalid semester"}, status=400)

        courses = course_registration.objects.filter(
            student_id=student,
            semester_id=semester,
            course_slot_id=slot_id,
            course_id__code__startswith="SW"
        ).select_related('course_id').values(
            'course_id__id', 'course_id__code', 'course_id__name', 'course_id__credit'
        )
        
        courses_list = []
        for c in courses:
            courses_list.append({
                'id': c['course_id__id'],
                'code': c['course_id__code'],
                'name': c['course_id__name'],
                'credit': c['course_id__credit']
            })
        
        return JsonResponse(courses_list, safe=False)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(['student'])
@block_pg_phd
def swayam_current_courses(request):
    try:
        slot_id = request.GET.get('slot_id')
        
        if not slot_id:
            return JsonResponse({"error": "slot_id parameter required"}, status=400)
        
        current_user = request.user
        user_details = current_user.extrainfo
        student = Student.objects.get(id=user_details)
        semester_no = student.curr_semester_no
        
        try:
            current_semester = Semester.objects.get(
                curriculum=student.batch_id.curriculum,
                semester_no=semester_no
            )
        except Semester.DoesNotExist:
            return JsonResponse({"error": "Invalid current semester"}, status=400)

        try:
            slot = CourseSlot.objects.get(id=slot_id, semester=current_semester)

            # Exclude courses already used in any Extra_Credits or Swayam_Replace request (any status)
            used_course_ids = set(SwayamReplacementRequest.objects.filter(
                student=student,
                request_type__in=['Extra_Credits', 'Swayam_Replace']
            ).values_list('new_course__id', flat=True))

            courses = slot.courses.filter(code__startswith="SW").exclude(id__in=used_course_ids).values(
                'id', 'code', 'name', 'credit'
            )

            courses_list = [
                {'id': c['id'], 'code': c['code'], 'name': c['name'], 'credit': c['credit']}
                for c in courses
            ]

            return JsonResponse(courses_list, safe=False)
        except CourseSlot.DoesNotExist:
            return JsonResponse({"error": "Invalid slot for current semester"}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(['student'])
@block_pg_phd
def swayam_replace_submit(request):
    try:
        current_user = request.user
        user_details = current_user.extrainfo
        student = Student.objects.get(id=user_details)
        semester_no = student.curr_semester_no

        current_year = datetime.datetime.now().year
        academic_year, semester_type = generate_current_session(current_year, semester_no)
        
        try:
            current_semester = Semester.objects.get(
                curriculum=student.batch_id.curriculum,
                semester_no=semester_no
            )
        except Semester.DoesNotExist:
            return JsonResponse({"error": "Invalid semester"}, status=400)
        
        data = request.data
        source_semester_no = data.get('source_semester')
        old_course_id = data.get('source_course')
        new_course_id_1 = data.get('target_course')
        new_course_slot_id_1 = data.get('target_slot')
        new_course_id_2 = data.get('target_course_2')
        new_course_slot_id_2 = data.get('target_slot_2')
        is_current_semester = data.get('is_current_semester', False)
        
        # Basic validation - at least first course is required
        if not all([source_semester_no, old_course_id, new_course_id_1, new_course_slot_id_1]):
            return JsonResponse({"error": "Source semester, old course, and at least one new course with slot are required"}, status=400)

        try:
            source_semester = Semester.objects.get(
                curriculum=student.batch_id.curriculum,
                semester_no=int(source_semester_no)
            )
        except Semester.DoesNotExist:
            return JsonResponse({"error": "Invalid source semester"}, status=400)

        try:
            old_course = Courses.objects.get(id=old_course_id)
            new_course_1 = Courses.objects.get(id=new_course_id_1)
            new_course_slot_1 = CourseSlot.objects.get(id=new_course_slot_id_1)

            new_course_2 = None
            new_course_slot_2 = None
            if new_course_id_2 and new_course_slot_id_2:
                new_course_2 = Courses.objects.get(id=new_course_id_2)
                new_course_slot_2 = CourseSlot.objects.get(id=new_course_slot_id_2)

            old_course_reg = course_registration.objects.get(
                student_id=student,
                semester_id=source_semester,
                course_id=old_course
            )
            old_course_slot = old_course_reg.course_slot_id
            
        except (Courses.DoesNotExist, CourseSlot.DoesNotExist) as e:
            return JsonResponse({"error": f"Invalid course or slot: {str(e)}"}, status=400)
        except course_registration.DoesNotExist:
            return JsonResponse({"error": "Old course registration not found in the selected semester"}, status=400)

        if old_course_id == new_course_id_1 or (new_course_id_2 and old_course_id == new_course_id_2):
            return JsonResponse({
                "error": "New course selection must be different from the course being replaced."
            }, status=400)

        if new_course_id_2 and new_course_id_1 == new_course_id_2:
            return JsonResponse({
                "error": "You must select two different new courses."
            }, status=400)

        if new_course_slot_id_2 and new_course_slot_id_1 == new_course_slot_id_2:
            return JsonResponse({
                "error": "You must select two different new slots."
            }, status=400)

        if not new_course_2 and not is_current_semester:
            return JsonResponse({
                "error": "Both slots are required. You must select two new Swayam courses for the replacement request."
            }, status=400)
        
        # Validate grade: source course must NOT have a satisfactory grade (O, A+, A, B+, B, C+)
        blocked_grades = ['O', 'A+', 'A', 'B+', 'B', 'C+', 'S']
        roll_no = current_user.username
        has_blocked_grade = Student_grades.objects.filter(
            roll_no=roll_no,
            course_id=old_course,
            grade__in=blocked_grades
        ).exists()
        
        if has_blocked_grade:
            return JsonResponse({
                "error": "You are not eligible for replacement of this course. Replacement is only allowed for courses with unsatisfactory grades."
            }, status=400)

        if is_current_semester:
            request_type = 'Swayam_Replace'
        else:
            request_type = 'Swayam_Replace'

        SwayamReplacementRequest.objects.create(
            student=student,
            semester=source_semester,
            academic_year=academic_year,
            semester_type=semester_type,
            request_type=request_type,
            old_course=old_course,
            new_course=new_course_1,
            course_slot=old_course_slot,
            new_course_slot=new_course_slot_1,
            status='Pending',
            is_current_semester=is_current_semester
        )

        if new_course_2 and new_course_slot_2:
            SwayamReplacementRequest.objects.create(
                student=student,
                semester=source_semester,
                academic_year=academic_year,
                semester_type=semester_type,
                request_type=request_type,
                old_course=old_course,
                new_course=new_course_2,
                course_slot=old_course_slot,
                new_course_slot=new_course_slot_2,
                status='Pending',
                is_current_semester=is_current_semester
            )
            message = "Replacement request submitted successfully (2 courses). Pending Academic Admin approval."
        else:
            message = "Replacement request submitted successfully (1 course). Pending Academic Admin approval."
        
        if is_current_semester:
            message += " Note: Current semester course will be DROPPED and new course(s) will be REGISTERED."
        
        return JsonResponse({"message": message}, status=201)
    
    except Student.DoesNotExist:
        return JsonResponse({"error": "Student not found"}, status=404)
    except Semester.DoesNotExist:
        return JsonResponse({"error": "Invalid semester"}, status=400)
    except Exception as e:
        logger.error(f"Error in swayam_replace_submit: {str(e)}", exc_info=True)
        logger.error(traceback.format_exc())
        return JsonResponse({"error": f"Failed to submit: {str(e)}"}, status=500)


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(['student'])
@block_pg_phd
def student_swayam_requests(request):
    try:
        current_user = request.user
        user_details = current_user.extrainfo
        student = Student.objects.get(id=user_details)

        request_type = request.GET.get('request_type')
        requests_query = SwayamReplacementRequest.objects.filter(
            student=student,
            semester__semester_no=student.curr_semester_no
        )

        if request_type:
            requests_query = requests_query.filter(request_type=request_type)
        
        requests = requests_query.select_related(
            'old_course',
            'new_course',
            'new_course_slot',
            'semester'
        ).order_by('-submitted_at')
        
        requests_data = []
        for req in requests:
            requests_data.append({
                'id': req.id,
                'request_type': req.request_type,
                'old_course': {
                    'id': req.old_course.id,
                    'code': req.old_course.code,
                    'name': req.old_course.name
                } if req.old_course else None,
                'new_course': {
                    'id': req.new_course.id,
                    'code': req.new_course.code,
                    'name': req.new_course.name
                },
                'slot': {
                    'id': req.new_course_slot.id,
                    'name': req.new_course_slot.name
                },
                'status': req.status,
                'submitted_at': req.submitted_at.isoformat() if req.submitted_at else None,
                'processed_at': req.processed_at.isoformat() if req.processed_at else None,
                'academic_year': req.academic_year,
                'semester_type': req.semester_type,
                'semester_no': req.semester.semester_no if req.semester else None
            })
        
        return JsonResponse({'requests': requests_data}, status=200)
    
    except Student.DoesNotExist:
        return JsonResponse({"error": "Student not found"}, status=404)
    except Exception as e:
        logger.error(f"Error in student_swayam_requests: {str(e)}", exc_info=True)
        return JsonResponse({"error": "Failed to fetch requests"}, status=500)


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def admin_swayam_list_requests(request):
    try:
        request_type = request.GET.get('request_type')
        status_filter = request.GET.get('status')
        academic_year = request.GET.get('academic_year')
        semester_type = request.GET.get('semester_type')
        
        queryset = SwayamReplacementRequest.objects.select_related(
            'student',
            'student__id__user',
            'student__batch_id',
            'old_course',
            'new_course',
            'new_course_slot',
            'semester'
        ).order_by('-submitted_at')

        scopes = scopes_for(request.user)
        queryset = scope_via_student(queryset, scopes, 'student')

        count_queryset = scope_via_student(
            SwayamReplacementRequest.objects.all(), scopes, 'student')
        
        if request_type:
            count_queryset = count_queryset.filter(request_type=request_type)
        if academic_year:
            count_queryset = count_queryset.filter(academic_year=academic_year)
        if semester_type:
            count_queryset = count_queryset.filter(semester_type=semester_type)

        counts = {
            'pending': count_queryset.filter(status='Pending').count(),
            'approved': count_queryset.filter(status='Approved').count(),
            'rejected': count_queryset.filter(status='Rejected').count(),
            'total': count_queryset.count()
        }

        if request_type:
            queryset = queryset.filter(request_type=request_type)
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        if academic_year:
            queryset = queryset.filter(academic_year=academic_year)
        
        if semester_type:
            queryset = queryset.filter(semester_type=semester_type)
        
        requests_data = []
        for req in queryset:
            requests_data.append({
                'id': req.id,
                'student': {
                    'id': req.student.id.id,
                    'name': f"{req.student.id.user.first_name} {req.student.id.user.last_name}",
                    'roll_no': req.student.id.id,
                    'batch': req.student.batch_id.name if req.student.batch_id else 'N/A'
                },
                'request_type': req.request_type,
                'old_course': {
                    'id': req.old_course.id,
                    'code': req.old_course.code,
                    'name': req.old_course.name
                } if req.old_course else None,
                'new_course': {
                    'id': req.new_course.id,
                    'code': req.new_course.code,
                    'name': req.new_course.name
                },
                'slot': {
                    'id': req.new_course_slot.id,
                    'name': req.new_course_slot.name
                },
                'status': req.status,
                'submitted_at': req.submitted_at.isoformat() if req.submitted_at else None,
                'processed_at': req.processed_at.isoformat() if req.processed_at else None,
                'academic_year': req.academic_year,
                'semester_type': req.semester_type,
                'semester': {
                    'id': req.semester.id,
                    'semester_no': req.semester.semester_no
                }
            })
        
        return JsonResponse({
            'requests': requests_data,
            'counts': counts
        }, status=200)
    
    except Exception as e:
        logger.error(f"Error in admin_swayam_list_requests: {str(e)}", exc_info=True)
        return JsonResponse({"error": "Failed to fetch requests"}, status=500)


@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def admin_swayam_approve(request):
    try:
        request_id = request.data.get('request_id')
        if not request_id:
            return JsonResponse({"error": "request_id is required"}, status=400)
        
        req_obj = SwayamReplacementRequest.objects.get(id=request_id)
        if not student_in_scope(req_obj.student, scopes_for(request.user)):
            return Response({'error': 'Request not found'}, status=status.HTTP_404_NOT_FOUND)
        
        if req_obj.status != 'Pending':
            return JsonResponse({"error": "Only pending requests can be approved"}, status=400)

        if req_obj.request_type == 'Extra_Credits':
            existing_reg = course_registration.objects.filter(
                student_id=req_obj.student,
                semester_id=req_obj.semester,
                course_id=req_obj.new_course
            ).first()
            
            if not existing_reg:
                course_registration.objects.create(
                    student_id=req_obj.student,
                    semester_id=req_obj.semester,
                    course_id=req_obj.new_course,
                    course_slot_id=req_obj.new_course_slot,
                    registration_type='Extra Credits',
                    session=req_obj.academic_year,
                    semester_type=req_obj.semester_type
                )
        
        elif req_obj.request_type == 'Swayam_Replace':
            paired_request = SwayamReplacementRequest.objects.filter(
                student=req_obj.student,
                old_course=req_obj.old_course,
                request_type='Swayam_Replace',
                status='Pending'
            ).exclude(id=req_obj.id).first()

            old_reg = course_registration.objects.filter(
                student_id=req_obj.student,
                semester_id=req_obj.semester,
                course_id=req_obj.old_course
            ).first()

            roll_no = req_obj.student.id.user.username
            if req_obj.is_current_semester:
                new_registration_type = 'Extra Credits'
            else:
                backlog_grades = ['F', 'CD', 'X']
                improvement_grades = ['D', 'D+', 'C']
                src_grade_obj = latest_grade(roll_no, req_obj.old_course)
                if src_grade_obj and src_grade_obj.grade in backlog_grades:
                    new_registration_type = 'Backlog'
                elif src_grade_obj and src_grade_obj.grade in improvement_grades:
                    new_registration_type = 'Improvement'
                else:
                    new_registration_type = 'Improvement'

            if req_obj.is_current_semester:
                # DROP + REGISTER mode 
                existing_new = course_registration.objects.filter(
                    student_id=req_obj.student,
                    semester_id=req_obj.semester,
                    course_id=req_obj.new_course
                ).first()
                if not existing_new:
                    course_registration.objects.create(
                        student_id=req_obj.student,
                        semester_id=req_obj.semester,
                        course_id=req_obj.new_course,
                        course_slot_id=req_obj.new_course_slot,
                        registration_type=new_registration_type,
                        session=req_obj.academic_year,
                        semester_type=req_obj.semester_type
                    )

                if paired_request:
                    existing_new_2 = course_registration.objects.filter(
                        student_id=paired_request.student,
                        semester_id=paired_request.semester,
                        course_id=paired_request.new_course
                    ).first()
                    if not existing_new_2:
                        course_registration.objects.create(
                            student_id=paired_request.student,
                            semester_id=paired_request.semester,
                            course_id=paired_request.new_course,
                            course_slot_id=paired_request.new_course_slot,
                            registration_type=new_registration_type,
                            session=paired_request.academic_year,
                            semester_type=paired_request.semester_type
                        )
                    paired_request.status = 'Approved'
                    paired_request.processed_at = timezone.now()
                    paired_request.save()

                # Drop the old course registration AFTER both new ones are registered
                if old_reg:
                    old_reg.delete()

            else:
                # REPLACE mode (previous semester)
                student_obj = req_obj.student
                current_semester_no = student_obj.curr_semester_no
                try:
                    current_semester = Semester.objects.get(
                        curriculum=student_obj.batch_id.curriculum,
                        semester_no=current_semester_no
                    )
                except Semester.DoesNotExist:
                    return JsonResponse({"error": "Could not determine student's current semester."}, status=400)

                new_reg_1 = course_registration.objects.filter(
                    student_id=student_obj,
                    semester_id=current_semester,
                    course_id=req_obj.new_course
                ).first()

                if not new_reg_1:
                    new_reg_1 = course_registration.objects.create(
                        student_id=student_obj,
                        semester_id=current_semester,
                        course_id=req_obj.new_course,
                        course_slot_id=req_obj.new_course_slot,
                        registration_type=new_registration_type,
                        session=req_obj.academic_year,
                        semester_type=req_obj.semester_type
                    )

                if old_reg and new_reg_1:
                    course_replacement.objects.get_or_create(
                        old_course_registration=old_reg,
                        new_course_registration=new_reg_1
                    )

                if paired_request:
                    paired_request.status = 'Approved'
                    paired_request.processed_at = timezone.now()
                    paired_request.save()

                    new_reg_2 = course_registration.objects.filter(
                        student_id=student_obj,
                        semester_id=current_semester,
                        course_id=paired_request.new_course
                    ).first()
                    if not new_reg_2:
                        new_reg_2 = course_registration.objects.create(
                            student_id=student_obj,
                            semester_id=current_semester,
                            course_id=paired_request.new_course,
                            course_slot_id=paired_request.new_course_slot,
                            registration_type=new_registration_type,
                            session=paired_request.academic_year,
                            semester_type=paired_request.semester_type
                        )

                    if old_reg and new_reg_2:
                        course_replacement.objects.get_or_create(
                            old_course_registration=old_reg,
                            new_course_registration=new_reg_2
                        )

        req_obj.status = 'Approved'
        req_obj.processed_at = timezone.now()
        req_obj.save()
        
        return JsonResponse({"message": "Request approved successfully"}, status=200)
    
    except SwayamReplacementRequest.DoesNotExist:
        return JsonResponse({"error": "Request not found"}, status=404)
    except Exception as e:
        logger.error(f"Error in admin_swayam_approve: {str(e)}", exc_info=True)
        logger.error(traceback.format_exc())
        return JsonResponse({"error": f"Failed to approve request: {str(e)}"}, status=500)


@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def admin_swayam_reject(request):
    try:
        request_id = request.data.get('request_id')
        if not request_id:
            return JsonResponse({"error": "request_id is required"}, status=400)
        
        req_obj = SwayamReplacementRequest.objects.get(id=request_id)
        if not student_in_scope(req_obj.student, scopes_for(request.user)):
            return Response({'error': 'Request not found'}, status=status.HTTP_404_NOT_FOUND)
        
        if req_obj.status != 'Pending':
            return JsonResponse({"error": "Only pending requests can be rejected"}, status=400)
        
        req_obj.status = 'Rejected'
        req_obj.processed_at = timezone.now()
        req_obj.save()
        
        return JsonResponse({"message": "Request rejected successfully"}, status=200)
    
    except SwayamReplacementRequest.DoesNotExist:
        return JsonResponse({"error": "Request not found"}, status=404)
    except Exception as e:
        logger.error(f"Error in admin_swayam_reject: {str(e)}", exc_info=True)
        return JsonResponse({"error": "Failed to reject request"}, status=500)


@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def admin_swayam_revert(request):
    try:
        request_id = request.data.get('request_id')
        if not request_id:
            return JsonResponse({"error": "request_id is required"}, status=400)
        
        req_obj = SwayamReplacementRequest.objects.get(id=request_id)
        if not student_in_scope(req_obj.student, scopes_for(request.user)):
            return Response({'error': 'Request not found'}, status=status.HTTP_404_NOT_FOUND)
        
        if req_obj.status != 'Rejected':
            return JsonResponse({"error": "Only rejected requests can be reverted"}, status=400)
        
        req_obj.status = 'Pending'
        req_obj.processed_at = None
        req_obj.save()
        
        return JsonResponse({"message": "Request reverted to pending successfully"}, status=200)
    
    except SwayamReplacementRequest.DoesNotExist:
        return JsonResponse({"error": "Request not found"}, status=404)
    except Exception as e:
        logger.error(f"Error in admin_swayam_revert: {str(e)}", exc_info=True)
        return JsonResponse({"error": "Failed to revert request"}, status=500)


@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def admin_swayam_delete(request):
    try:
        request_id = request.data.get('request_id')
        if not request_id:
            return JsonResponse({"error": "request_id is required"}, status=400)
        
        req_obj = SwayamReplacementRequest.objects.get(id=request_id)
        if not student_in_scope(req_obj.student, scopes_for(request.user)):
            return Response({'error': 'Request not found'}, status=status.HTTP_404_NOT_FOUND)
        req_obj.delete()
        
        return JsonResponse({"message": "Request deleted successfully"}, status=200)
    
    except SwayamReplacementRequest.DoesNotExist:
        return JsonResponse({"error": "Request not found"}, status=404)
    except Exception as e:
        logger.error(f"Error in admin_swayam_delete: {str(e)}", exc_info=True)
        return JsonResponse({"error": "Failed to delete request"}, status=500)


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def get_add_course_slots(request):
    """
    GET /api/course-slots/?semester_id=<id>
    Returns JSON list of { id, name } for all slots in that semester.
    """
    sem_id = request.query_params.get("semester_id")
    if not sem_id:
        return Response(
            {"error": "semester_id query parameter is required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # ensure the semester exists (404 if not)
    get_object_or_404(Semester, id=sem_id)

    # fetch slots and return only id & name
    slots = CourseSlot.objects.filter(semester_id=sem_id).values("id", "name")
    return Response(list(slots), status=status.HTTP_200_OK)


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def get_add_course_courses(request):
    """
    GET /api/courses/?courseslot_id=<id>
    Returns JSON list of { id, code, name, credit } for all courses in that slot.
    """
    slot_id = request.query_params.get("courseslot_id")
    if not slot_id:
        return Response(
            {"error": "courseslot_id query parameter is required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # ensure the slot exists (404 if not)
    slot = get_object_or_404(CourseSlot, id=slot_id)

    # When the term is supplied, include the sections each course is running in
    # (so the admin can pick a section for a cross-section backlog/improvement).
    academic_year = request.query_params.get("academic_year")
    semester_type = request.query_params.get("semester_type")
    working_year = None
    if academic_year and semester_type:
        try:
            working_year = parse_academic_year(academic_year=academic_year, semester_type=semester_type)[0]
        except Exception:
            working_year = None

    result = []
    for c in slot.courses.all():
        entry = {"id": c.id, "code": c.code, "name": c.name, "credit": c.credit}
        if working_year is not None:
            offerings = CourseInstructor.objects.filter(
                course_id=c, year=working_year, semester_type=semester_type,
            ).select_related("instructor_id__id__user")
            entry["sections"] = [{
                "course_instructor_id": o.id,
                "section": o.section_label or "",
                "instructor": f"{o.instructor_id.id.user.first_name} {o.instructor_id.id.user.last_name}".strip(),
            } for o in offerings]
        result.append(entry)
    return Response(result, status=status.HTTP_200_OK)


def roman_to_int(s):
    roman = {'I': 1, 'II': 2, 'III': 3, 'IV': 4, 'V': 5, 'VI': 6,
             'VII': 7, 'VIII': 8, 'IX': 9, 'X': 10}
    return roman.get(s.strip().upper())

@api_view(['POST'])
@parser_classes([MultiPartParser])
@role_required(['acadadmin'])
def upload_excel_replacement(request):
    file = request.FILES.get('file')
    if not file:
        return Response({'error': 'No file uploaded'}, status=400)

    try:
        df = pd.read_excel(file)
    except Exception as e:
        return Response({'error': f'Failed to read Excel: {e}'}, status=400)

    expected = {'student_id', 'old_course_code', 'new_course_code',
                'old_semester_roman', 'new_semester_number'}
    if not expected.issubset(df.columns):
        return Response(
            {'error': f'Missing columns: {expected - set(df.columns)}'},
            status=400
        )

    valid_entries = []
    failed_rows = []

    # Validate all rows first
    for idx, row in df.iterrows():
        sid = str(row['student_id']).strip()
        old_code = str(row['old_course_code']).strip()
        new_code = str(row['new_course_code']).strip()
        old_rom = str(row['old_semester_roman']).strip()
        new_sem = int(row['new_semester_number'])

        sem_old = roman_to_int(old_rom)
        if sem_old is None:
            msg = f'Row {idx+2} Invalid Roman numeral {old_rom}'
            failed_rows.append(msg)
            continue

        try:
            student = Student.objects.get(id_id=sid)
        except Student.DoesNotExist:
            msg = f'Row {idx+2} Student {sid} not found'
            failed_rows.append(msg)
            continue

        try:
            old_reg = course_registration.objects.get(
                student_id=student,
                course_id__code=old_code,
                semester_id__semester_no=sem_old
            )
        except course_registration.DoesNotExist:
            msg = f'Row {idx+2} Old registration not found: {sid}, {old_code}, {sem_old}'
            failed_rows.append(msg)
            continue

        try:
            new_reg = course_registration.objects.get(
                student_id=student,
                course_id__code=new_code,
                semester_id__semester_no=new_sem
            )
        except course_registration.DoesNotExist:
            msg = f'Row {idx+2} New registration not found: {sid}, {new_code}, {new_sem}'
            failed_rows.append(msg)
            continue

        valid_entries.append((old_reg, new_reg))

    # Abort if any error was found
    if failed_rows:
        return Response({
            'error': 'Some rows are invalid. No changes were made.',
            'failed_rows': failed_rows
        }, status=400)

    # All rows valid: Perform atomic save
    try:
        with transaction.atomic():
            for old_reg, new_reg in valid_entries:
                course_replacement.objects.create(
                    old_course_registration=old_reg,
                    new_course_registration=new_reg
                )
                # The replaced course carries the grade, so it decides whether
                # the new one counts as a backlog or an improvement.
                _sg = latest_grade(old_reg.student_id.id_id, old_reg.course_id)
                reg_type = registration_type_for_grade(_sg.grade if _sg else None)
                if reg_type != 'Regular' and new_reg.registration_type != reg_type:
                    new_reg.registration_type = reg_type
                    new_reg.save(update_fields=['registration_type'])
    except Exception as e:
        return Response({'error': str(e)}, status=400)

    return Response({
        'message': f'{len(valid_entries)} replacements added successfully',
        'failed_rows': []
    }, status=200)



# ─── Mapping HOD designation → allowed Student.specialization ─────────────
HOD_SPECIALIZATION_MAPPING = {
    # designation code without spaces/brackets → list of allowed specializations
    'ECE': ['PNC'],        # HOD(ECE) can only assign PNC students
    'CSE': ['AIML'],       # HOD(CSE) can only assign AIML students
    # add more as needed
}

def check_role(request, required_role):
    # Authorize from the user's real designation, never a client-supplied role
    from django.db.models import Q
    user = request.user
    if not getattr(user, "is_authenticated", False):
        return False
    held = HoldsDesignation.objects.filter(Q(working=user) | Q(user=user))
    if required_role == 'hod':
        return held.filter(designation__name__istartswith='HOD').exists()
    if required_role == 'faculty':
        return (
            Faculty.objects.filter(id__user=user).exists()
            or held.filter(
                designation__name__in=[
                    'Professor', 'Associate Professor', 'Assistant Professor',
                ]
            ).exists()
        )
    return held.filter(designation__name=required_role).exists()

def get_allowed_specs(user):
    """
    Find the HOD's designation name (e.g. "HOD (ECE)"), extract "ECE",
    and return the list of allowed student specializations.
    """
    des = HoldsDesignation.objects.filter(
        working=user, designation__name__startswith='HOD'
    ).first()
    if not des:
        return None
    # extract between parentheses
    raw = des.designation.name
    try:
        code = raw.split('(',1)[1].split(')',1)[0].strip()
    except:
        return None
    return HOD_SPECIALIZATION_MAPPING.get(code)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def tas_list(request):
    """GET /api/tas/ → return all TA usernames."""
    data = [{'username': s.id.user.username} for s in Student.objects.all()]
    return Response({'tas': data})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def faculties_list(request):
    """GET /api/faculties/ → return all Faculty usernames."""
    data = [{'username': f.id.user.username} for f in Faculty.objects.all()]
    return Response({'faculties': data})


# --- HOD endpoints ---

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def hod_students(request):
    if not check_role(request,'hod'):
        return Response({'error':'role=hod required'}, status=status.HTTP_403_FORBIDDEN)
    allowed = get_allowed_specs(request.user)
    if not allowed:
        return Response({'error':'Invalid HOD designation or no mapping'}, status=status.HTTP_403_FORBIDDEN)
    qs = Student.objects.filter(specialization__in=allowed)
    data = [{'username': s.id.user.username, 'batch': s.batch} for s in qs]
    return Response({'students': data})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def hod_assign_manual(request):
    if not check_role(request,'hod'):
        return Response({'error':'role=hod required'}, status=status.HTTP_403_FORBIDDEN)
    allowed = get_allowed_specs(request.user)
    if not allowed:
        return Response({'error':'Invalid HOD designation or no mapping'}, status=status.HTTP_403_FORBIDDEN)
    d = request.data
    try:
        stu = Student.objects.get(id__user__username=d['ta_username'])
    except Student.DoesNotExist:
        return Response({'error':'TA not found'}, status=status.HTTP_404_NOT_FOUND)
    if stu.specialization not in allowed:
        return Response({'error':'Specialization mismatch'}, status=status.HTTP_403_FORBIDDEN)
    try:
        fac = Faculty.objects.get(id__user__username=d['faculty_username'])
    except Faculty.DoesNotExist:
        return Response({'error':'Faculty not found'}, status=status.HTTP_404_NOT_FOUND)

    try:
        a = Assignment.objects.create(
            ta=stu, faculty=fac,
            start_year=int(d['start_year']),  start_month=int(d['start_month']),
            end_year=int(d['end_year']),      end_month=int(d['end_month'])
        )
        return Response({'assignment_id': a.id}, status=status.HTTP_201_CREATED)
    except KeyError:
        return Response({'error':'Missing field'}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({'error':str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def hod_upload_excel(request):
    if not check_role(request,'hod'):
        return Response({'error':'role=hod required'}, status=status.HTTP_403_FORBIDDEN)
    allowed = get_allowed_specs(request.user)
    if not allowed:
        return Response({'error':'Invalid HOD designation or no mapping'}, status=status.HTTP_403_FORBIDDEN)
    f = request.FILES.get('file')
    if not f:
        return Response({'error':'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        df = pd.read_excel(f)
    except:
        return Response({'error':'Invalid Excel file'}, status=status.HTTP_400_BAD_REQUEST)

    created, errors = [], []
    for idx, row in df.iterrows():
        try:
            stu = Student.objects.get(id__user__username=row['ta_username'])
            if stu.specialization not in allowed:
                raise PermissionError('Specialization mismatch')
            fac = Faculty.objects.get(id__user__username=row['faculty_username'])
            a = Assignment.objects.create(
                ta=stu, faculty=fac,
                start_year=int(row['start_year']), start_month=int(row['start_month']),
                end_year=int(row['end_year']),     end_month=int(row['end_month'])
            )
            created.append(a.id)
        except Exception as e:
            errors.append({'row': int(idx), 'error': str(e)})

    return Response({'created': created, 'errors': errors})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def hod_pending(request):
    if not check_role(request,'hod'):
        return Response({'error':'role=hod required'}, status=status.HTTP_403_FORBIDDEN)
    qs = StipendRequest.objects.filter(status=StipendRequest.FAC_APPROVED)
    data = [{
        'id': s.id,
        'ta': s.assignment.ta.id.user.username,
        'faculty': s.assignment.faculty.id.user.username,
        'month': s.month, 'year': s.year,
        'faculty_remark': s.faculty_remark
    } for s in qs]
    return Response({'stipends': data})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def hod_approved(request):
    if not check_role(request,'hod'):
        return Response({'error':'role=hod required'}, status=status.HTTP_403_FORBIDDEN)
    qs = StipendRequest.objects.filter(status=StipendRequest.HOD_APPROVED)
    data = [{
        'id': s.id,
        'ta': s.assignment.ta.id.user.username,
        'faculty': s.assignment.faculty.id.user.username,
        'month': s.month, 'year': s.year,
    } for s in qs]
    return Response({'stipends': data})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def hod_approve(request, sid):
    if not check_role(request,'hod'):
        return Response({'error':'role=hod required'}, status=status.HTTP_403_FORBIDDEN)
    try:
        s = StipendRequest.objects.get(id=sid)
    except StipendRequest.DoesNotExist:
        return Response({'error':'Stipend not found'}, status=status.HTTP_404_NOT_FOUND)

    if s.status != StipendRequest.FAC_APPROVED:
        return Response({'error':'Faculty must approve first'}, status=status.HTTP_400_BAD_REQUEST)

    s.status = StipendRequest.HOD_APPROVED
    s.hod_remark = request.data.get('remark','')
    s.save()
    return Response({'success': True})




@api_view(['GET'])
@permission_classes([IsAuthenticated])
def faculty_assignments(request):
    # if not check_role(request,'faculty'):
    #     return Response({'error':'role=faculty required'}, status=status.HTTP_403_FORBIDDEN)
    try:
        faculty = Faculty.objects.get(id=request.user.extrainfo)
        qs = Assignment.objects.filter(faculty=faculty)
    except Faculty.DoesNotExist:
        return Response({'error': 'Faculty profile not found'}, status=status.HTTP_404_NOT_FOUND)
    data = [{
        'id': a.id,
        'ta_username': a.ta.id.user.username,
        'start_month': a.start_month, 'start_year': a.start_year,
        'end_month': a.end_month,     'end_year': a.end_year
    } for a in qs]
    return Response({'assignments': data})

from django.db.models import Q

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def faculty_pending(request):
    if not check_role(request,'faculty'):
        return Response({'error':'role=faculty required'}, status=status.HTTP_403_FORBIDDEN)
    try:
        faculty = Faculty.objects.get(id=request.user.extrainfo)
    except Faculty.DoesNotExist:
        return Response({'error': 'Faculty profile not found'}, status=status.HTTP_404_NOT_FOUND)
    now = datetime.datetime.now()
    qs = StipendRequest.objects.filter(
        assignment__faculty=faculty,
        status=StipendRequest.PENDING
    ).filter(
        Q(year__lt=now.year) |
        Q(year=now.year, month__lte=now.month)
    )
    data = [{'id': s.id, 'ta': s.assignment.ta.id.user.username,
             'month': s.month, 'year': s.year} for s in qs]
    return Response({'stipends': data})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def faculty_approved(request):
    if not check_role(request,'faculty'):
        return Response({'error':'role=faculty required'}, status=status.HTTP_403_FORBIDDEN)
    try:
        faculty = Faculty.objects.get(id=request.user.extrainfo)
    except Faculty.DoesNotExist:
        return Response({'error': 'Faculty profile not found'}, status=status.HTTP_404_NOT_FOUND)
    qs = StipendRequest.objects.filter(
        assignment__faculty=faculty,
        status=StipendRequest.FAC_APPROVED
    )
    data = [{'id': s.id, 'ta': s.assignment.ta.id.user.username,
             'month': s.month, 'year': s.year} for s in qs]
    return Response({'stipends': data})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def faculty_approve(request, sid):
    if not check_role(request,'faculty'):
        return Response({'error':'role=faculty required'}, status=status.HTTP_403_FORBIDDEN)
    try:
        s = StipendRequest.objects.get(id=sid, assignment__faculty=request.user.faculty)
        now = datetime.now()
        if (s.year, s.month) > (now.year, now.month):
            return Response({'error':'Cannot approve future'}, status=status.HTTP_400_BAD_REQUEST)
        s.status = StipendRequest.FAC_APPROVED
        s.faculty_remark = request.data.get('remark','')
        s.save()
        return Response({'success': True})
    except StipendRequest.DoesNotExist:
        return Response({'error':'Stipend not found'}, status=status.HTTP_404_NOT_FOUND)

# --- TA endpoints ---

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def ta_stipends(request):
    if not check_role(request,'ta'):
        return Response({'error':'role=ta required'}, status=status.HTTP_403_FORBIDDEN)
    qs = StipendRequest.objects.filter(assignment__ta=request.user.student)
    data = [{'month': s.month, 'year': s.year,
             'status': s.status, 'faculty_remark': s.faculty_remark}
            for s in qs]
    return Response({'stipends': data})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(['student'])
@block_pg_phd
def registered_slots(request):
    try:
        current_user = request.user
        user_details = current_user.extrainfo
        student = Student.objects.get(id=user_details)
        session, semester_type = generate_current_session(datetime.datetime.now().year, student.curr_semester_no) 
        eligibility_resp = get_replace_registration_eligibility(timezone.now().date(), student.curr_semester_no, datetime.datetime.now().year)
        if isinstance(eligibility_resp, JsonResponse):
            return JsonResponse([], safe=False)
        
        # Exclude slots with pending drop requests
        pending_drop_slots = CourseDropRequest.objects.filter(
            student=student,
            academic_year=session,
            semester_type=semester_type,
            status='Pending'
        ).values_list('course_slot_id', flat=True)
        
        regs = course_registration.objects.filter(
            student_id=student, 
            semester_id__semester_no=student.curr_semester_no
        ).exclude(
            course_slot_id__name__startswith='SW'
        ).exclude(
            course_slot_id__name__startswith='BL'
        ).exclude(
            course_slot_id__in=pending_drop_slots
        )
        
        payload = []
        for reg in regs:
            slot = reg.course_slot_id
            others = slot.courses.all().exclude(id=reg.course_id.id)

            payload.append({
                "id": slot.id,
                "name": slot.name,
                "academic_year": reg.session,
                "semester_type": reg.semester_type,
                "old_course": {"id": reg.course_id.id, "code": reg.course_id.code, "name" : reg.course_id.name},
                "new_courses": [
                    {"id": c.id, "code": c.code, "name": c.name, "seats_available": max(c.max_seats - (course_registration.objects.filter(course_id=c, session = session, semester_type = semester_type).count()), 0)} for c in others
                ],
            })
        return JsonResponse(payload, safe=False)   
    except Student.DoesNotExist:
        return Response({"error": "Student profile not found"}, status=400)
    except Exception as e:
        return Response({"error": str(e)}, status=500)
    
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['student'])
@block_pg_phd
def batch_create_requests(request):
    try:
        current_user = request.user
        user_details = current_user.extrainfo
        student = Student.objects.get(id=user_details)
        eligibility_resp = get_replace_registration_eligibility(timezone.now().date(), student.curr_semester_no, datetime.datetime.now().year)
        if isinstance(eligibility_resp, JsonResponse):
            return eligibility_resp
        data = json.loads(request.body).get('requests', [])

        created = []
        errors = []

        for idx, item in enumerate(data):
            try:
                slot_id = item.get('course_slot')
                old_id = item.get('old_course')
                new_id = item.get('new_course')
                ay = item.get('academic_year')
                sem = item.get('semester_type')

                if not all([slot_id, old_id, new_id, ay, sem]):
                    raise ValueError("Missing required fields.")

                slot = CourseSlot.objects.get(pk=slot_id)
                old = Courses.objects.get(pk=old_id)
                new = Courses.objects.get(pk=new_id)

                reg = course_registration.objects.get(
                    student_id=student,
                    course_slot_id=slot,
                    course_id=old,
                    session=ay,
                    semester_type=sem
                )

                if not slot.courses.filter(id=new.id).exists():
                    raise ValueError("New course not in selected slot.")

                with transaction.atomic():
                    req, created_flag = CourseReplacementRequest.objects.get_or_create(
                        student=student,
                        course_slot=slot,
                        academic_year=ay,
                        semester_type=sem,
                        defaults={
                            "old_course": old,
                            "new_course": new,
                        }
                    )

                    action = "created"
                    if not created_flag:
                        if req.old_course != old or req.new_course != new:
                            req.old_course = old
                            req.new_course = new
                            req.status = "Pending"
                            req.processed_at = None
                            req.save()
                            action = "updated"
                        else:
                            action = "unchanged"

                    created.append({
                        "id": req.id,
                        "slot": slot.name,
                        "old": old.code,
                        "new": new.code,
                        "status": req.status,
                        "action": action,
                    })

            except Exception as e:
                errors.append({"index": idx, "error": str(e)})

        return JsonResponse({"created": created, "errors": errors}, status=201)

    except Exception as e:
        return JsonResponse({"detail": "Something went wrong", "error": str(e)}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def admin_list_requests(request):
    qs = CourseReplacementRequest.objects.select_related(
        'student', 'student__id__user', 'course_slot', 'course_slot__semester', 'old_course', 'new_course'
    ).all().order_by('-created_at')
    qs = scope_via_student(qs, scopes_for(request.user), 'student')
    
    year = request.GET.get('academic_year')
    sem  = request.GET.get('semester_type')
    status_filter = request.GET.get('status')
    if year:
        qs = qs.filter(academic_year=year)
    if sem:
        qs = qs.filter(semester_type=sem)
    if status_filter:
        qs = qs.filter(status__iexact=status_filter)

    out = []
    for r in qs:
        out.append({
            'id': r.id,
            'student': r.student.id.user.username,
            'student_name': r.student.id.user.get_full_name() or r.student.id.user.username,
            'slot': r.course_slot.name if r.course_slot else 'N/A',
            'semester': (r.course_slot.semester.semester_no if r.course_slot and r.course_slot.semester_id else None),
            'old_course': r.old_course.code if r.old_course else 'N/A',
            'old_course_name': r.old_course.name if r.old_course else 'N/A',
            'new_course': r.new_course.code if r.new_course else 'N/A',
            'new_course_name': r.new_course.name if r.new_course else 'N/A',
            'status': r.status,
            'academic_year': r.academic_year,
            'semester_type': r.semester_type,
            'created_at': r.created_at.isoformat(),
            'processed_at': r.processed_at.isoformat() if r.processed_at else None,
        })
    return JsonResponse(out, safe=False)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(['student'])
@block_pg_phd
def student_list_requests(request):
    current_user = request.user
    user_details = current_user.extrainfo
    student = Student.objects.get(id=user_details)
    current_reg = course_registration.objects.filter(student_id=student, semester_id__semester_no=student.curr_semester_no).first()
    
    if current_reg:
        academic_year = current_reg.session
        semester_type = current_reg.semester_type
    else:
        academic_year, semester_type = generate_current_session(datetime.datetime.now().year, student.curr_semester_no)
    
    qs = CourseReplacementRequest.objects.filter(student=student, academic_year=academic_year, semester_type = semester_type).select_related('old_course', 'new_course', 'course_slot').order_by('-created_at')
    out = []
    for r in qs:
        out.append({
            'id': r.id,
            'slot': r.course_slot.name if r.course_slot else 'N/A',
            'old_course': r.old_course.code if r.old_course else 'N/A',
            'old_course_name': r.old_course.name if r.old_course else 'N/A',
            'new_course': r.new_course.code if r.new_course else 'N/A',
            'new_course_name': r.new_course.name if r.new_course else 'N/A',
            'status': r.status,
            'academic_year': r.academic_year,
            'semester_type': r.semester_type,
            'created_at': r.created_at.isoformat(),
        })
    return JsonResponse(out, safe=False)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
@transaction.atomic
@role_required(list(ALL_ACAD_ROLES))
def allocate_all(request):
    import json
    try:
        body = json.loads(request.body)
        year = body.get('academic_year')
        sem  = body.get('semester_type')
        request_ids = body.get('request_ids') or []
        request_ids = scoped_ids(CourseReplacementRequest, request_ids, scopes_for(request.user))
    except:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    pending = CourseReplacementRequest.objects.select_for_update().filter(
        status="Pending", academic_year=year, semester_type=sem
    )
    # When specific requests are selected, allot only those.
    if request_ids:
        pending = pending.filter(id__in=request_ids)
    by_course = defaultdict(list)
    for cr in pending:
        by_course[cr.new_course].append(cr)

    queue = deque()
    in_q  = set()
    for course, reqs in by_course.items():
        course = Courses.objects.select_for_update().get(pk=course.pk)
        used   = course_registration.objects.filter(course_id=course, session = year, semester_type = sem).count()
        free   = max(course.max_seats - used, 0)
        if free > 0:
            queue.append(course)
            in_q.add(course)

    results = []
    while queue:
        course = queue.popleft()
        in_q.discard(course)

        while True:
            used = course_registration.objects.filter(course_id=course, session=year, semester_type=sem).count()
            free = max(course.max_seats - used, 0)
            reqs = by_course[course]
            if free <= 0 or not reqs:
                break

            # Always pick one request at a time (FIFO)
            cr = reqs[0]
            # Find the existing registration to replace; if missing, skip without approving so the batch still runs.
            old_reg = course_registration.objects.select_for_update().filter(
                student_id=cr.student,
                course_slot_id=cr.course_slot,
                session=cr.academic_year,
                semester_type=cr.semester_type,
            ).first()
            if old_reg is None:
                by_course[course].remove(cr)
                results.append({'id': cr.id, 'status': 'Skipped',
                                'reason': 'No existing registration in this slot to replace.'})
                continue

            cr.status = "Approved"
            cr.processed_at = timezone.now()
            cr.save(update_fields=['status', 'processed_at'])
            results.append({'id': cr.id, 'status': 'Approved'})

            old_course = old_reg.course_id
            semester_id = old_reg.semester_id
            old_reg.delete()

            working_year, _ = parse_academic_year(cr.academic_year, cr.semester_type)
            course_registration.objects.create(
                student_id=cr.student,
                course_slot_id=cr.course_slot,
                course_id=course,
                session=cr.academic_year,
                semester_type=cr.semester_type,
                semester_id=semester_id,
                working_year=working_year
            )
            by_course[course].remove(cr)

            # cascade enqueue old_course if it now has free seats and pending requests
            used_old = course_registration.objects.filter(course_id=old_course, session=year, semester_type=sem).count()
            free_old = max(old_course.max_seats - used_old, 0)
            if free_old > 0 and by_course.get(old_course) and by_course[old_course] and old_course not in in_q:
                queue.append(old_course)
                in_q.add(old_course)

    # reject leftovers
    for reqs in by_course.values():
        for cr in reqs:
            cr.status = "Rejected"
            cr.processed_at = timezone.now()
            cr.save(update_fields=['status', 'processed_at'])
            results.append({'id': cr.id, 'status': 'Rejected'})

    return JsonResponse(results, safe=False)


# Change replacement request status back to pending
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@transaction.atomic
@role_required(list(ALL_ACAD_ROLES))
def revert_replacement_to_pending(request):
    import json
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        body = json.loads(request.body)
        request_ids = body.get('request_ids', [])
        request_ids = scoped_ids(CourseReplacementRequest, request_ids, scopes_for(request.user))
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON format'}, status=400)
    except Exception as e:
        logger.error(f"Error parsing request body: {str(e)}")
        return JsonResponse({'error': 'Invalid request format'}, status=400)

    if not request_ids or not isinstance(request_ids, list):
        return JsonResponse({'error': 'request_ids must be a non-empty array'}, status=400)

    reverted_count = 0
    errors = []
    
    for req_id in request_ids:
        try:
            req_id = int(req_id)
            replacement_request = CourseReplacementRequest.objects.get(id=req_id)
            
            if replacement_request.status != 'Rejected':
                errors.append({'id': req_id, 'error': f'Cannot revert request with status: {replacement_request.status}'})
                continue
            
            replacement_request.status = 'Pending'
            replacement_request.processed_at = None
            replacement_request.save(update_fields=['status', 'processed_at'])
            reverted_count += 1
            
        except CourseReplacementRequest.DoesNotExist:
            errors.append({'id': req_id, 'error': 'Request not found'})
        except Exception as e:
            logger.error(f"Error reverting replacement request {req_id}: {str(e)}")
            errors.append({'id': req_id, 'error': str(e)})
    
    return JsonResponse({
        'reverted': reverted_count,
        'total': len(request_ids),
        'errors': errors
    })


# Delete replacement requests
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@transaction.atomic
@role_required(list(ALL_ACAD_ROLES))
def delete_replacement_requests(request):
    import json
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        body = json.loads(request.body)
        request_ids = body.get('request_ids', [])
        request_ids = scoped_ids(CourseReplacementRequest, request_ids, scopes_for(request.user))
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON format'}, status=400)
    except Exception as e:
        logger.error(f"Error parsing request body: {str(e)}")
        return JsonResponse({'error': 'Invalid request format'}, status=400)

    if not request_ids or not isinstance(request_ids, list):
        return JsonResponse({'error': 'request_ids must be a non-empty array'}, status=400)

    deleted_count = 0
    errors = []
    
    for req_id in request_ids:
        try:
            req_id = int(req_id)
            replacement_request = CourseReplacementRequest.objects.get(id=req_id)
            replacement_request.delete()
            deleted_count += 1
        except CourseReplacementRequest.DoesNotExist:
            errors.append({'id': req_id, 'error': 'Request not found'})
        except Exception as e:
            logger.error(f"Error deleting replacement request {req_id}: {str(e)}")
            errors.append({'id': req_id, 'error': str(e)})
    
    return JsonResponse({
        'deleted': deleted_count,
        'total': len(request_ids),
        'errors': errors
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(['student'])
@block_pg_phd
def student_registrations_for_drop(request):
    """
    GET /api/student/registrations/
    List all active registrations for the logged-in student.
    Excludes courses with pending replacement requests.
    """
    current_user = request.user
    user_details = current_user.extrainfo
    student = Student.objects.get(id=user_details)
    eligibility_resp = get_drop_registration_eligibility(timezone.now().date(), student.curr_semester_no, datetime.datetime.now().year)
    if isinstance(eligibility_resp, JsonResponse):
        return eligibility_resp
    
    current_year = datetime.datetime.now().year
    academic_year, semester_type = generate_current_session(current_year, student.curr_semester_no)
    
    # Exclude slots with pending replacement or drop requests
    pending_replacement_slots = CourseReplacementRequest.objects.filter(
        student=student,
        academic_year=academic_year,
        semester_type=semester_type,
        status='Pending'
    ).values_list('course_slot_id', flat=True)
    
    pending_drop_slots = CourseDropRequest.objects.filter(
        student=student,
        academic_year=academic_year,
        semester_type=semester_type,
        status='Pending'
    ).values_list('course_slot_id', flat=True)
    
    regs = course_registration.objects.filter(
        student_id=student, 
        semester_id__semester_no=student.curr_semester_no
    ).exclude(
        course_slot_id__in=pending_replacement_slots
    ).exclude(
        course_slot_id__in=pending_drop_slots
    ).select_related('course_id', 'course_slot_id').order_by('course_slot_id__name')
    
    out = []
    for reg in regs:
        out.append({
            'id': reg.id,
            'slot': reg.course_slot_id.name,
            'course': reg.course_id.code,
            'course_name': reg.course_id.name,
            'academic_year': reg.session,
            'semester_type': reg.semester_type,
        })
    return Response(out, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['student'])
@block_pg_phd
def drop_course(request):
    """
    POST /api/student/drop-course/
    Body: { "registration_id": <int> }
    Creates a drop request instead of directly dropping the course.
    Validates eligibility before creating the request.
    """
    try:
        current_user = request.user
        user_details = getattr(current_user, 'extrainfo', None)
        if not user_details:
            return Response(
                {'error': 'User profile not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        student = Student.objects.select_related('id__user', 'batch_id').get(id=user_details)

        eligibility_resp = get_drop_registration_eligibility(
            timezone.now().date(), 
            student.curr_semester_no, 
            datetime.datetime.now().year
        )
        if isinstance(eligibility_resp, JsonResponse):
            return eligibility_resp

        reg_id = request.data.get('registration_id')
        if not reg_id:
            return Response(
                {'error': 'registration_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            reg_id = int(reg_id)
        except (ValueError, TypeError):
            return Response(
                {'error': 'Invalid registration_id format'},
                status=status.HTTP_400_BAD_REQUEST
            )

        with transaction.atomic():
            reg = course_registration.objects.select_related(
                'course_id', 'course_slot_id'
            ).get(id=reg_id, student_id=student)

            drop_request, created = CourseDropRequest.objects.get_or_create(
                student=student,
                course_slot=reg.course_slot_id,
                academic_year=reg.session,
                semester_type=reg.semester_type,
                defaults={'course': reg.course_id}
            )
            
            if not created:
                if drop_request.status == "Pending":
                    return Response(
                        {'error': 'Drop request already pending for this slot'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                drop_request.course = reg.course_id
                drop_request.status = "Pending"
                drop_request.processed_at = None
                drop_request.save(update_fields=['course', 'status', 'processed_at'])
            
            return Response(
                {
                    'status': 'pending',
                    'request_id': drop_request.id,
                    'registration_id': reg_id,
                    'message': 'Drop request submitted successfully. Waiting for Academic approval.'
                },
                status=status.HTTP_201_CREATED if created else status.HTTP_200_OK
            )
            
    except Student.DoesNotExist:
        return Response(
            {'error': 'Student profile not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except course_registration.DoesNotExist:
        return Response(
            {'error': 'Course registration not found or does not belong to you'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Drop course error for user {request.user.username}: {str(e)}", exc_info=True)
        return Response(
            {'error': 'An error occurred while processing your request'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(['student'])
def student_calendar_view(request):
    calendar_entries = Calendar.objects.all().order_by('from_date')

    result = [
        {
            "id": entry.id,
            "from_date": entry.from_date,
            "to_date": entry.to_date,
            "from_time": entry.from_time,
            "to_time": entry.to_time,
            "description": entry.description,
        }
        for entry in calendar_entries
    ]

    return Response({"calendar_events": result})

# List all drop requests for the logged-in student for current semester
@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(['student'])
@block_pg_phd
def student_list_drop_requests(request):
    try:
        current_user = request.user
        user_details = getattr(current_user, 'extrainfo', None)
        if not user_details:
            return JsonResponse({'error': 'User profile not found'}, status=404)
        
        student = Student.objects.select_related('id__user').get(id=user_details)
        
        current_reg = course_registration.objects.filter(
            student_id=student,
            semester_id__semester_no=student.curr_semester_no
        ).only('session', 'semester_type').first()
        
        if current_reg:
            academic_year = current_reg.session
            semester_type = current_reg.semester_type
        else:
            academic_year, semester_type = generate_current_session(
                datetime.datetime.now().year,
                student.curr_semester_no
            )

        qs = CourseDropRequest.objects.filter(
            student=student,
            academic_year=academic_year,
            semester_type=semester_type
        ).select_related('course', 'course_slot').order_by('-created_at')

        out = [
            {
                'id': r.id,
                'slot': r.course_slot.name,
                'semester': r.course_slot.semester.semester_no if r.course_slot and r.course_slot.semester_id else None,
                'course': r.course.code,
                'course_name': r.course.name,
                'status': r.status,
                'academic_year': r.academic_year,
                'semester_type': r.semester_type,
                'created_at': r.created_at.isoformat(),
                'processed_at': r.processed_at.isoformat() if r.processed_at else None,
            }
            for r in qs
        ]
        
        return JsonResponse(out, safe=False)
        
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student profile not found'}, status=404)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error listing drop requests for {request.user.username}: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'An error occurred while fetching requests'}, status=500)

# Counts the queues the acadadmin home page reports, without serialising rows.
@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def admin_pending_counts(request):
    scopes = scopes_for(request.user)
    pending = lambda model: scope_via_student(
        model.objects.filter(status__iexact='pending'), scopes, 'student').count()
    return JsonResponse({'counts': {
        'swayam': pending(SwayamReplacementRequest),
        'add': pending(CourseAddRequest),
        'drop': pending(CourseDropRequest),
        'replacement': pending(CourseReplacementRequest),
        'phdCourses': pending(PhDCourseRegistrationRequest),
        'thesisEnrolments': pending(ThesisRegistration),
        'progressSeminars': pending(ProgressSeminarRegistration),
        'teachingCredits': pending(TeachingCreditRegistration),
    }}, status=200)


 # Lists all course drop requests with optional filtering
@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def admin_list_drop_requests(request):
    try:
        qs = CourseDropRequest.objects.select_related(
            'student__id__user',
            'course',
            'course_slot',
            'course_slot__semester'
        ).all().order_by('-created_at')
        qs = scope_via_student(qs, scopes_for(request.user), 'student')

        year = request.GET.get('academic_year', '').strip()
        sem = request.GET.get('semester_type', '').strip()
        status_filter = request.GET.get('status', '').strip()

        if year:
            qs = qs.filter(academic_year=year)
        if sem:
            qs = qs.filter(semester_type=sem)
        # Filter before the cap, so a status query cannot be truncated away.
        if status_filter:
            qs = qs.filter(status__iexact=status_filter)

        # Pending first, so the rows needing action survive the cap below.
        qs = qs.order_by(
            Case(When(status__iexact='pending', then=0), default=1,
                 output_field=IntegerField()),
            '-created_at',
        )

        qs = qs[:500]

        out = [
            {
                'id': r.id,
                'student': r.student.id.user.username,
                'student_name': f"{r.student.id.user.first_name} {r.student.id.user.last_name}".strip(),
                'slot': r.course_slot.name,
                'semester': r.course_slot.semester.semester_no if r.course_slot and r.course_slot.semester_id else None,
                'course': r.course.code,
                'course_name': r.course.name,
                'status': r.status,
                'academic_year': r.academic_year,
                'semester_type': r.semester_type,
                'created_at': r.created_at.isoformat(),
                'processed_at': r.processed_at.isoformat() if r.processed_at else None,
            }
            for r in qs
        ]
        
        return JsonResponse(out, safe=False)
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error listing drop requests for admin: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'An error occurred while fetching requests'}, status=500)

# Processes multiple drop requests in batch
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@transaction.atomic
@role_required(list(ALL_ACAD_ROLES))
def approve_drop_requests(request):
    import json
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        body = json.loads(request.body)
        request_ids = body.get('request_ids', [])
        request_ids = scoped_ids(CourseDropRequest, request_ids, scopes_for(request.user))
        action = body.get('action', 'approve').lower().strip()
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON format'}, status=400)
    except Exception as e:
        logger.error(f"Error parsing request body: {str(e)}")
        return JsonResponse({'error': 'Invalid request format'}, status=400)

    if not request_ids or not isinstance(request_ids, list):
        return JsonResponse({'error': 'request_ids must be a non-empty array'}, status=400)
    
    if action not in ['approve', 'reject']:
        return JsonResponse({'error': 'action must be either "approve" or "reject"'}, status=400)

    results = []
    success_count = 0
    error_count = 0
    
    for req_id in request_ids:
        try:
            req_id = int(req_id)

            drop_request = CourseDropRequest.objects.select_for_update().select_related(
                'student', 'course', 'course_slot'
            ).get(id=req_id)

            if drop_request.status != "Pending":
                results.append({
                    'id': req_id,
                    'status': 'already_processed',
                    'current_status': drop_request.status
                })
                continue
            
            if action == 'approve':
                try:
                    reg = course_registration.objects.get(
                        student_id=drop_request.student,
                        course_slot_id=drop_request.course_slot,
                        course_id=drop_request.course,
                        session=drop_request.academic_year,
                        semester_type=drop_request.semester_type
                    )
                    reg.delete()
                    drop_request.status = "Approved"
                    drop_request.processed_at = timezone.now()
                    drop_request.save(update_fields=['status', 'processed_at'])
                    results.append({'id': req_id, 'status': 'approved'})
                    success_count += 1
                    logger.info(f"Approved drop request {req_id} for student {drop_request.student.id}")
                except course_registration.DoesNotExist:
                    drop_request.status = "Approved"
                    drop_request.processed_at = timezone.now()
                    drop_request.save(update_fields=['status', 'processed_at'])
                    results.append({'id': req_id, 'status': 'approved', 'note': 'registration_not_found'})
                    success_count += 1
                    logger.warning(f"Approved drop request {req_id} but registration not found")
            else:  # reject
                drop_request.status = "Rejected"
                drop_request.processed_at = timezone.now()
                drop_request.save(update_fields=['status', 'processed_at'])
                results.append({'id': req_id, 'status': 'rejected'})
                success_count += 1
                logger.info(f"Rejected drop request {req_id}")
                
        except (ValueError, TypeError):
            results.append({'id': req_id, 'status': 'error', 'detail': 'Invalid ID format'})
            error_count += 1
        except CourseDropRequest.DoesNotExist:
            results.append({'id': req_id, 'status': 'not_found'})
            error_count += 1
        except Exception as e:
            logger.error(f"Error processing drop request {req_id}: {str(e)}", exc_info=True)
            results.append({'id': req_id, 'status': 'error', 'detail': 'Processing error'})
            error_count += 1
    
    return JsonResponse({
        'results': results,
        'summary': {
            'total': len(request_ids),
            'success': success_count,
            'errors': error_count
        }
    }, safe=False)


# Delete drop course requests
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@transaction.atomic
@role_required(list(ALL_ACAD_ROLES))
def delete_drop_requests(request):
    import json
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        body = json.loads(request.body)
        request_ids = body.get('request_ids', [])
        request_ids = scoped_ids(CourseDropRequest, request_ids, scopes_for(request.user))
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON format'}, status=400)
    except Exception as e:
        logger.error(f"Error parsing request body: {str(e)}")
        return JsonResponse({'error': 'Invalid request format'}, status=400)

    if not request_ids or not isinstance(request_ids, list):
        return JsonResponse({'error': 'request_ids must be a non-empty array'}, status=400)

    deleted_count = 0
    errors = []
    
    for req_id in request_ids:
        try:
            req_id = int(req_id)
            drop_request = CourseDropRequest.objects.get(id=req_id)
            drop_request.delete()
            deleted_count += 1
        except CourseDropRequest.DoesNotExist:
            errors.append({'id': req_id, 'error': 'Request not found'})
        except Exception as e:
            logger.error(f"Error deleting drop request {req_id}: {str(e)}")
            errors.append({'id': req_id, 'error': str(e)})
    
    return JsonResponse({
        'deleted': deleted_count,
        'total': len(request_ids),
        'errors': errors
    })


# ===================== COURSE ADD REQUEST APIs =====================

# Lists all add course requests for the current student's active semester
@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(['student'])
@block_pg_phd
def student_list_add_requests(request):
    try:
        current_user = request.user
        user_details = getattr(current_user, 'extrainfo', None)
        if not user_details:
            return JsonResponse({'error': 'User profile not found'}, status=404)
        
        student = Student.objects.select_related('id__user').get(id=user_details)
        
        current_reg = course_registration.objects.filter(
            student_id=student,
            semester_id__semester_no=student.curr_semester_no
        ).only('session', 'semester_type').first()
        
        if current_reg:
            academic_year = current_reg.session
            semester_type = current_reg.semester_type
        else:
            academic_year, semester_type = generate_current_session(
                datetime.datetime.now().year,
                student.curr_semester_no
            )

        qs = CourseAddRequest.objects.filter(
            student=student,
            academic_year=academic_year,
            semester_type=semester_type
        ).select_related('course', 'course_slot').order_by('-created_at')

        out = [
            {
                'id': r.id,
                'slot': r.course_slot.name,
                'semester': r.course_slot.semester.semester_no if r.course_slot and r.course_slot.semester_id else None,
                'course': r.course.code,
                'course_name': r.course.name,
                'status': r.status,
                'academic_year': r.academic_year,
                'semester_type': r.semester_type,
                'created_at': r.created_at.isoformat(),
                'processed_at': r.processed_at.isoformat() if r.processed_at else None,
            }
            for r in qs
        ]
        
        return JsonResponse(out, safe=False)
        
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student profile not found'}, status=404)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error listing add requests for {request.user.username}: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'An error occurred while fetching requests'}, status=500)

# Lists all course add requests with optional filtering
@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def admin_list_add_requests(request):
    try:
        qs = CourseAddRequest.objects.select_related(
            'student__id__user',
            'course',
            'course_slot',
            'course_slot__semester'
        ).all().order_by('-created_at')
        qs = scope_via_student(qs, scopes_for(request.user), 'student')

        year = request.GET.get('academic_year', '').strip()
        sem = request.GET.get('semester_type', '').strip()
        status_filter = request.GET.get('status', '').strip()

        if year:
            qs = qs.filter(academic_year=year)
        if sem:
            qs = qs.filter(semester_type=sem)
        # Filter before the cap, so a status query cannot be truncated away.
        if status_filter:
            qs = qs.filter(status__iexact=status_filter)

        # Pending first, so the rows needing action survive the cap below.
        qs = qs.order_by(
            Case(When(status__iexact='pending', then=0), default=1,
                 output_field=IntegerField()),
            '-created_at',
        )

        qs = qs[:500]

        out = [
            {
                'id': r.id,
                'student': r.student.id.user.username,
                'student_name': f"{r.student.id.user.first_name} {r.student.id.user.last_name}".strip(),
                'slot': r.course_slot.name,
                'semester': r.course_slot.semester.semester_no if r.course_slot and r.course_slot.semester_id else None,
                'course': r.course.code,
                'course_name': r.course.name,
                'status': r.status,
                'academic_year': r.academic_year,
                'semester_type': r.semester_type,
                'created_at': r.created_at.isoformat(),
                'processed_at': r.processed_at.isoformat() if r.processed_at else None,
            }
            for r in qs
        ]
        
        return JsonResponse(out, safe=False)
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error listing add requests for admin: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'An error occurred while fetching requests'}, status=500)

# Processes multiple add requests in batch
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@transaction.atomic
@role_required(list(ALL_ACAD_ROLES))
def approve_add_requests(request):
    import json
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        body = json.loads(request.body)
        request_ids = body.get('request_ids', [])
        request_ids = scoped_ids(CourseAddRequest, request_ids, scopes_for(request.user))
        action = body.get('action', 'approve').lower().strip()
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON format'}, status=400)
    except Exception as e:
        logger.error(f"Error parsing request body: {str(e)}")
        return JsonResponse({'error': 'Invalid request format'}, status=400)

    if not request_ids or not isinstance(request_ids, list):
        return JsonResponse({'error': 'request_ids must be a non-empty array'}, status=400)
    
    if action not in ['approve', 'reject']:
        return JsonResponse({'error': 'action must be either "approve" or "reject"'}, status=400)

    results = []
    success_count = 0
    error_count = 0
    
    for req_id in request_ids:
        try:
            req_id = int(req_id)

            add_request = CourseAddRequest.objects.select_related(
                'student', 'course', 'course_slot', 'student__batch_id'
            ).select_for_update(of=('self',)).get(id=req_id)

            if add_request.status != "Pending":
                results.append({
                    'id': req_id,
                    'status': 'already_processed',
                    'current_status': add_request.status
                })
                continue
            
            if action == 'approve':
                existing = course_registration.objects.filter(
                    student_id=add_request.student,
                    course_id=add_request.course,
                    session=add_request.academic_year,
                    semester_type=add_request.semester_type
                ).exists()
                
                if existing:
                    add_request.status = "Rejected"
                    add_request.processed_at = timezone.now()
                    add_request.save(update_fields=['status', 'processed_at'])
                    results.append({'id': req_id, 'status': 'error', 'detail': 'Already registered'})
                    error_count += 1
                    continue

                try:
                    semester = Semester.objects.get(
                        curriculum=add_request.student.batch_id.curriculum,
                        semester_no=add_request.student.curr_semester_no
                    )
                except Semester.DoesNotExist:
                    add_request.status = "Rejected"
                    add_request.processed_at = timezone.now()
                    add_request.save(update_fields=['status', 'processed_at'])
                    results.append({'id': req_id, 'status': 'error', 'detail': 'Semester not found'})
                    error_count += 1
                    continue

                registration_type = 'Regular'
                try:
                    student_grade = Student_grades.objects.filter(
                        roll_no=add_request.student.id.user.username,
                        course_id=add_request.course
                    ).order_by('-year', '-semester').first()
                    
                    if student_grade and student_grade.grade:
                        backlog_grades = ['F', 'X', 'CD']
                        improvement_grades = ['C', 'D+', 'D']
                        
                        if student_grade.grade in backlog_grades:
                            registration_type = 'Backlog'
                        elif student_grade.grade in improvement_grades:
                            registration_type = 'Improvement'
                except Exception as grade_error:
                    logger.warning(f"Could not determine registration type for request {req_id}: {str(grade_error)}")

                try:
                    reg = course_registration(
                        student_id=add_request.student,
                        course_id=add_request.course,
                        course_slot_id=add_request.course_slot,
                        semester_id=semester,
                        session=add_request.academic_year,
                        semester_type=add_request.semester_type,
                        working_year=datetime.datetime.now().year,
                        registration_type=registration_type,
                        # Section chosen at request time (cross-section backlog/improvement).
                        course_instructor=add_request.course_instructor,
                    )
                    reg.save()

                    if add_request.old_course_registration:
                        course_replacement.objects.create(
                            old_course_registration=add_request.old_course_registration,
                            new_course_registration=reg
                        )
                        logger.info(f"Created course_replacement for request {req_id}: {add_request.old_course_registration.id} -> {reg.id}")
                    
                    add_request.status = "Approved"
                    add_request.processed_at = timezone.now()
                    add_request.save(update_fields=['status', 'processed_at'])
                    results.append({'id': req_id, 'status': 'approved'})
                    success_count += 1
                    logger.info(f"Approved add request {req_id} for student {add_request.student.id}")
                except Exception as reg_error:
                    logger.error(f"Error creating registration for request {req_id}: {str(reg_error)}", exc_info=True)
                    results.append({'id': req_id, 'status': 'error', 'detail': f'Registration failed: {str(reg_error)}'})
                    error_count += 1
            else:  # reject
                add_request.status = "Rejected"
                add_request.processed_at = timezone.now()
                add_request.save(update_fields=['status', 'processed_at'])
                results.append({'id': req_id, 'status': 'rejected'})
                success_count += 1
                logger.info(f"Rejected add request {req_id}")
                
        except (ValueError, TypeError):
            results.append({'id': req_id, 'status': 'error', 'detail': 'Invalid ID format'})
            error_count += 1
        except CourseAddRequest.DoesNotExist:
            results.append({'id': req_id, 'status': 'not_found'})
            error_count += 1
        except Exception as e:
            logger.error(f"Error processing add request {req_id}: {str(e)}", exc_info=True)
            results.append({'id': req_id, 'status': 'error', 'detail': f'Processing error: {str(e)}'})
            error_count += 1
    
    return JsonResponse({
        'results': results,
        'summary': {
            'total': len(request_ids),
            'success': success_count,
            'errors': error_count
        }
    }, safe=False)


# Delete add course requests
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@transaction.atomic
@role_required(list(ALL_ACAD_ROLES))
def delete_add_requests(request):
    import json
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        body = json.loads(request.body)
        request_ids = body.get('request_ids', [])
        request_ids = scoped_ids(CourseAddRequest, request_ids, scopes_for(request.user))
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON format'}, status=400)
    except Exception as e:
        logger.error(f"Error parsing request body: {str(e)}")
        return JsonResponse({'error': 'Invalid request format'}, status=400)

    if not request_ids or not isinstance(request_ids, list):
        return JsonResponse({'error': 'request_ids must be a non-empty array'}, status=400)

    deleted_count = 0
    errors = []
    
    for req_id in request_ids:
        try:
            req_id = int(req_id)
            add_request = CourseAddRequest.objects.get(id=req_id)
            add_request.delete()
            deleted_count += 1
        except CourseAddRequest.DoesNotExist:
            errors.append({'id': req_id, 'error': 'Request not found'})
        except Exception as e:
            logger.error(f"Error deleting add request {req_id}: {str(e)}")
            errors.append({'id': req_id, 'error': str(e)})
    
    return JsonResponse({
        'deleted': deleted_count,
        'total': len(request_ids),
        'errors': errors
    })


@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def student_search(request):
    roll_no = request.data.get('rollno','').upper()
    if not roll_no:
        return JsonResponse({'error':'rollno is required'},status=400)
    student = Student.objects.filter(id_id=roll_no).first()
    if not student or not student_in_scope(student, scopes_for(request.user)):
        return JsonResponse({'error':'Student record not found'},status=400)
    extra = student.id
    user = extra.user
    data = {
        'roll_no':roll_no,
        'full_name':f"{user.first_name} {user.last_name}".strip(),
        'date_of_birth':str(extra.date_of_birth),
        'user_status':extra.user_status,
        'address':extra.address,
        'phone_no':extra.phone_no,
        'department':extra.department.name if extra.department else None,
        'programme':student.programme,
        'batch':student.batch,
        'batch_name':student.batch_id.name if student.batch_id else None,
        'discipline':student.batch_id.discipline.name if student.batch_id and student.batch_id.discipline else None,
        'curriculum':student.batch_id.curriculum.name if student.batch_id and student.batch_id.curriculum else None,
        'cpi':student.cpi,
        'category':student.category,
        'father_name':student.father_name,
        'mother_name':student.mother_name,
        'hall_no':student.hall_no,
        'room_no':student.room_no,
        'specialization':student.specialization,
        'curr_semester_no':student.curr_semester_no,
    }
    return JsonResponse(data,status=200)


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(['student'])
def student_registration_semesters_view(request):
    """
    Return a list of distinct semesters in which the student has registrations.
    For new students, also include their current semester even if no registrations exist.
    """
    try:
        roll_number = request.user.username
        student = Student.objects.get(id_id=roll_number)

        # Pull distinct (semester_no, semester_type) from the student's course registrations
        qs = (course_registration.objects
              .filter(student_id=student)
              .values_list('semester_id__semester_no', 'semester_type')
              .distinct()
              .order_by('semester_id__semester_no'))

        unique = OrderedDict()
        for sem_no, sem_type in qs:
            label = make_label(sem_no, sem_type or "")
            unique[(sem_no, sem_type)] = label

        # For new students who haven't registered for any courses yet,
        # include their current semester
        if not unique and student.curr_semester_no:
            # Determine semester type based on semester number (odd/even)
            current_sem_type = "Odd Semester" if student.curr_semester_no % 2 == 1 else "Even Semester"
            current_label = make_label(student.curr_semester_no, current_sem_type)
            unique[(student.curr_semester_no, current_sem_type)] = current_label

        semesters = [
            {"semester_no": no, "semester_type": typ, "label": lbl}
            for (no, typ), lbl in unique.items()
        ]

        return JsonResponse({"success": True, "semesters": semesters}, status=200)

    except Student.DoesNotExist:
        return JsonResponse({"success": False, "error": "Student not found."}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)
    

@api_view(["GET"])
@permission_classes([IsAuthenticated])
@role_required(['student'])
def student_filled(request):
    roll_number = request.user.username
    student = Student.objects.get(id_id=roll_number)
    semester_no = student.curr_semester_no
    done = FeedbackFilled.objects.filter(student=student, semester_no = semester_no).exists()
    return Response({"filled": done})

@api_view(["GET"])
@permission_classes([IsAuthenticated])
@role_required(['student'])
def student_questions(request):
    try:
        roll_number = request.user.username
        student = Student.objects.get(id_id=roll_number)
    except Student.DoesNotExist:
        return Response({"detail": "Student profile not found."}, status=status.HTTP_404_NOT_FOUND)

    semester_no = student.curr_semester_no

    filled = FeedbackFilled.objects.filter(
        student=student,
        semester_no=semester_no
    ).exists()

    registrations = course_registration.objects.filter(
        student_id=student,
        semester_id__semester_no=semester_no,
    ).select_related("course_id")

    from applications.academic_information.models import resolve_offering

    courses = []
    for reg in registrations:
        course = reg.course_id
        academic_year, _ = parse_academic_year(reg.session, reg.semester_type)
        # The student's own section instructor, else any offering (no-section courses).
        instructor_entry = resolve_offering(student, course, academic_year, reg.semester_type) or \
            CourseInstructor.objects.filter(
                course_id=course,
                semester_type=reg.semester_type,
                year=academic_year,
            ).first()

        # No instructor -> nothing to give feedback on; skip.
        if not instructor_entry:
            continue

        instructor_name = (
            f"{instructor_entry.instructor_id.id.user.first_name} "
            f"{instructor_entry.instructor_id.id.user.last_name}"
        ).strip()

        courses.append({
            "course_id": course.id,
            "code": course.code,
            "name": course.name,
            "instructor_id": instructor_entry.id,
            "instructor_name": instructor_name,
        })

    questions = [
        {
            "id": question.id,
            "section": question.section,
            "text": question.text,
            "options": [{"id": option.id, "text": option.text} for option in question.options.all()],
        }
        for question in FeedbackQuestion.objects.all()
    ]

    return Response({
        "filled": filled,
        "courses": courses,
        "questions": questions,
    })

@api_view(["POST"])
@permission_classes([IsAuthenticated])
@role_required(['student'])
def student_submit(request):
    try:
        roll_number = request.user.username
        student = Student.objects.get(id_id=roll_number)
    except Student.DoesNotExist:
        return Response({"detail": "Student profile not found."}, status=status.HTTP_404_NOT_FOUND)

    semester_no = student.curr_semester_no
    data = request.data
    if FeedbackFilled.objects.filter(student=student, semester_no = semester_no).exists():
        return Response({"detail":"Already filled."}, status=status.HTTP_409_CONFLICT)

    from applications.academic_information.models import resolve_offering
    with transaction.atomic():
        for r in data["responses"]:

            reg = course_registration.objects.get(student_id =student, course_id_id = r["course_id"], semester_id__semester_no = student.curr_semester_no)
            # Attribute to the offering the student was shown (their section's, else first).
            academic_year, _ = parse_academic_year(reg.session, reg.semester_type)
            offering = resolve_offering(student, reg.course_id, academic_year, reg.semester_type) or \
                CourseInstructor.objects.filter(
                    course_id=reg.course_id,
                    year=academic_year,
                    semester_type=reg.semester_type,
                ).first()
            # Trust the server for question/section; only accept an option that
            # actually belongs to the question.
            question = FeedbackQuestion.objects.filter(id=r.get("question_id")).first()
            if not question:
                continue
            option = None
            if r.get("option_id"):
                option = FeedbackOption.objects.filter(
                    id=r["option_id"], question=question).first()
            FeedbackResponse.objects.create(
                question      = question,
                option        = option,
                text_answer   = r.get("text_answer",""),
                course_id     = r["course_id"],
                course_instructor = offering,
                section       = question.section,
                session       = reg.session,
                semester_type = reg.semester_type,
            )
        FeedbackFilled.objects.create(student=student, semester_no = student.curr_semester_no)

    return Response({"detail":"Submitted"}, status=status.HTTP_201_CREATED)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def inst_courses(request):
    """
    GET /inst/courses/?session=<str>&semester_type=<str>
    Returns the list of courses the logged-in instructor is teaching.
    """
    fac = request.user.username
    sess = request.query_params.get("session")
    semt = request.query_params.get("semester_type")
    if not sess or not semt:
        return Response({"detail": "Provide 'session' and 'semester_type'."}, status=status.HTTP_400_BAD_REQUEST)

    academic_year, _ = parse_academic_year(sess, semt)
    regs = CourseInstructor.objects.filter(
        instructor_id_id=fac,
        year=academic_year,
        semester_type=semt,
    )

    return Response([{
        "course_id": cr.course_id.id,
        "code":      cr.course_id.code,
        "name":      cr.course_id.name,
    } for cr in regs])


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def inst_all_stats(request):
    """
    GET /inst/stats/all/?session=&semester_type=&course_id=
    Returns per-question counts + comments for the 'Course Instructor' section.
    If no responses yet, returns {"detail": "No responses found till now."}.
    """
    sess = request.query_params.get("session")
    semt = request.query_params.get("semester_type")
    cid = request.query_params.get("course_id")

    academic_year, _ = parse_academic_year(sess, semt)
    if not CourseInstructor.objects.filter(
        course_id_id=cid,
        instructor_id_id=request.user.username,
        year=academic_year,
        semester_type=semt,
    ).exists():
        return Response(
            {"error": "Access denied: you are not assigned as instructor for this course."},
            status=status.HTTP_403_FORBIDDEN
        )

    has_any = FeedbackResponse.objects.filter(
        course_id=cid,
        session=sess,
        semester_type=semt,
        question__section="instructor",
    ).exists()

    if not has_any:
        return Response(
            {"detail": "No responses found till now."},
            status=status.HTTP_200_OK
        )

    out = []
    questions = FeedbackQuestion.objects.filter(section="instructor").order_by("order")
    
    for q in questions:
        base = FeedbackResponse.objects.filter(
            question=q,
            course_id=cid,
            session=sess,
            semester_type=semt,
        )
        counts = {
            o.text: base.filter(option=o).count()
            for o in FeedbackOption.objects.filter(question=q)
        }
        comments = list(
            base.filter(option__isnull=True).values_list("text_answer", flat=True)
        )
        out.append({
            "question_id": q.id,
            "text": q.text,
            "counts": counts,
            "comments": comments,
        })

    return Response(out, status=status.HTTP_200_OK)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
@role_required(['acadadmin'])
def admin_course_list(request):
    """
    GET /admin/courses/?session=<str>&semester_type=<str>
    """
    sess = request.query_params.get("session")
    semt = request.query_params.get("semester_type")
    if not sess or not semt:
        return Response(
            {"detail":"Provide 'session' & 'semester_type'."},
            status=status.HTTP_400_BAD_REQUEST
        )

    regs = FeedbackResponse.objects.filter(
        session=sess,
        semester_type=semt,
    ).select_related("course").distinct()

    academic_year, _ = parse_academic_year(sess, semt)

    def _instr_name(offering):
        u = offering.instructor_id.id.user
        return f"{u.first_name} {u.last_name}".strip()

    seen = set()
    courses = []
    for reg in regs:
        c = reg.course
        if c.id in seen:
            continue
        seen.add(c.id)

        offerings = list(CourseInstructor.objects.filter(
            course_id=c, year=academic_year, semester_type=semt,
        ).select_related("instructor_id__id__user"))

        sectioned = len(offerings) > 1 and any(o.section_label for o in offerings)

        if sectioned:
            # One row per section offering.
            for o in offerings:
                courses.append({
                    "course_id": c.id,
                    "code":      c.code,
                    "name":      c.name,
                    "instructor": _instr_name(o),
                    "section":    o.section_label or "",
                    "course_instructor_id": o.id,
                })
            # Surface responses not tied to an offering so the split doesn't hide them.
            if FeedbackResponse.objects.filter(
                course=c, session=sess, semester_type=semt,
                course_instructor__isnull=True,
            ).exists():
                courses.append({
                    "course_id": c.id,
                    "code":      c.code,
                    "name":      c.name,
                    "instructor": "Section not recorded",
                    "section":    "",
                    "course_instructor_id": "none",
                })
        else:
            names = []
            for o in offerings:
                nm = _instr_name(o)
                if nm and nm not in names:
                    names.append(nm)
            courses.append({
                "course_id": c.id,
                "code":      c.code,
                "name":      c.name,
                "instructor": ", ".join(names),
                "section":    "",
                "course_instructor_id": None,
            })

    return Response(courses)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
@role_required(['acadadmin'])
def admin_all_stats(request):
    """
    GET /admin/stats/all/?session=<str>&semester_type=<str>&course_id=<int>
    Returns a JSON payload grouped by section:
      {
        sections: [
          {
            section: "<section_key>",
            questions: [
              {
                question_id, text,
                counts: { option_text: count, ... },
                comments: [ ... ]
              },
              ...
            ]
          },
          ...
        ]
      }
    """
    sess = request.query_params.get("session")
    semt = request.query_params.get("semester_type")
    cid  = request.query_params.get("course_id")
    course_instructor = request.query_params.get("course_instructor")
    if not sess or not semt or not cid:
        return Response(
            {"detail":"Provide 'session', 'semester_type', and 'course_id'."},
            status=status.HTTP_400_BAD_REQUEST
        )

    raw = []
    for q in FeedbackQuestion.objects.all().order_by("order"):
        base = FeedbackResponse.objects.filter(
            question=q,
            course_id=cid,
            session=sess,
            semester_type=semt,
        )
        # "none" = responses with no offering recorded.
        if course_instructor == "none":
            base = base.filter(course_instructor__isnull=True)
        elif course_instructor:
            base = base.filter(course_instructor_id=course_instructor)
        counts = {
            o.text: base.filter(option=o).count()
            for o in FeedbackOption.objects.filter(question=q)
        }
        comments = list(
            base.filter(option__isnull=True)
                .values_list("text_answer", flat=True)
        )
        raw.append({
            "section":     q.section,
            "question_id": q.id,
            "text":        q.text,
            "counts":      counts,
            "comments":    comments,
        })

    grouped = {}
    for item in raw:
        sec = item["section"]
        grouped.setdefault(sec, []).append({
            "question_id": item["question_id"],
            "text":        item["text"],
            "counts":      item["counts"],
            "comments":    item["comments"],
        })

    # Fixed display order: course-related sections first, lab-related last.
    section_order = ["contents", "instructor", "attendance", "tutorial", "lab"]
    def _sec_rank(s):
        return section_order.index(s) if s in section_order else len(section_order)

    response = {
        "sections": [
            {"section": sec, "questions": qs}
            for sec, qs in sorted(grouped.items(), key=lambda kv: _sec_rank(kv[0]))
        ]
    }

    if not raw or all(len(v["questions"]) == 0 for v in response["sections"]):
        return Response({"detail":"No responses found till now."})

    return Response(response)



@api_view(["GET"])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def list_batches(request):
    batches = scope_batches(
        Batch.objects.filter(running_batch=True), scopes_for(request.user)
    ).select_related("discipline").order_by("year", "name")
    result = []
    for b in batches:
        label = f"{b.name} {b.discipline.acronym} {b.year}"
        result.append({"id": b.id, "label": label, "year": b.year})
    return Response(result)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def list_students_in_batch(request):
    batch_id = request.query_params.get("batch_id")
    if not batch_id:
        return Response({"detail": "batch_id required."}, status=status.HTTP_400_BAD_REQUEST)
    students = scope_students(
        Student.objects.filter(batch_id__id=batch_id), scopes_for(request.user))
    result = []
    for st in students:
        cb = st.batch_id
        cb_label = f"{cb.name} {cb.discipline.acronym} {cb.year}"
        result.append({
            "id": st.id_id,
            "username": str(st.id_id),
            "current_batch": cb_label,
            "current_batch_id": cb.id,
            "current_batch_year": st.batch,
        })
    return Response(result)

@api_view(["POST"])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def apply_batch_changes(request):
    data = request.data
    user = request.user
    errors = []

    with transaction.atomic():
        for idx, pair in enumerate(data):
            sid = pair.get("student_id")
            nid = pair.get("new_batch_id")
            nyear = pair.get("new_batch_year")
            if not sid or not nid or nyear is None:
                errors.append({"index": idx, "detail": "student_id, new_batch_id, new_batch_year required."})
                continue
            try:
                student = Student.objects.get(id=sid)
                if not student_in_scope(student, scopes_for(request.user)):
                    errors.append({"index": idx, "detail": f"Student {sid} not found."})
                    continue
            except Student.DoesNotExist:
                errors.append({"index": idx, "detail": f"Student {sid} not found."})
                continue

            old_batch = student.batch_id
            if old_batch and old_batch.id == nid and student.batch == nyear:
                continue
            try:
                new_batch = Batch.objects.get(id=nid)
            except Batch.DoesNotExist:
                errors.append({"index": idx, "detail": f"Batch {nid} not found."})
                continue

            # Section follows the target batch: reuse sections already present there (single -> it, multi -> emptiest).
            in_batch = Student.objects.filter(batch_id=new_batch)
            counts = {}
            for s in (in_batch.exclude(section__isnull=True).exclude(section='')
                      .values_list('section', flat=True)):
                counts[s] = counts.get(s, 0) + 1
            if counts:
                new_section = next(iter(counts)) if len(counts) == 1 else min(counts, key=lambda s: (counts[s], s))
            elif in_batch.exists():
                new_section = None
            else:
                errors.append({
                    "index": idx,
                    "detail": f"{student.id_id}: target batch '{new_batch}' has no students yet, so its section layout is unknown. Set up its sections first, then move students in.",
                })
                continue

            BatchChangeHistory.objects.create(
                student=student,
                old_batch=old_batch,
                new_batch=new_batch,
            )
            student.batch_id = new_batch
            student.batch = nyear
            student.section = new_section
            student.save()

            # Sync branch, department, and specialization
            try:
                from applications.programme_curriculum.models_student_management import StudentBatchUpload
                from applications.globals.models import DepartmentInfo
                student_upload = StudentBatchUpload.objects.filter(roll_number=student.id_id).first()
                if student_upload:
                    student_upload.branch = new_batch.discipline.name
                    student_upload.save()
                dept_name = new_batch.discipline.acronym
                department = DepartmentInfo.objects.filter(name=dept_name).first()
                if department:
                    student.id.department = department
                    student.id.save()
                student.specialization = dept_name
                student.save()
            except:
                pass

    if errors:
        return Response({"errors": errors}, status=status.HTTP_207_MULTI_STATUS)
    return Response({"detail": "Batch changes applied."}, status=status.HTTP_200_OK)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def list_students_in_batch_semester_promotion(request):
    batch_id = request.query_params.get("batch_id")
    if not batch_id:
        return Response({"detail": "batch_id required."}, status=status.HTTP_400_BAD_REQUEST)
    students = scope_students(
        Student.objects.filter(batch_id__id=batch_id), scopes_for(request.user)
    ).select_related('id__user', 'batch_id__discipline').order_by('id_id')
    result = []
    for st in students:
        user = st.id.user
        result.append({
            "id": st.id_id,
            "username": str(st.id_id),
            "name": f"{user.first_name} {user.last_name}".strip(),
            "discipline": (st.batch_id.discipline.acronym
                           if st.batch_id and st.batch_id.discipline_id else ''),
            "current_semester_no": st.curr_semester_no,
        })
    return Response(result)

@api_view(["POST"])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def apply_promotion(request):
    data = request.data  # list of student IDs
    user = request.user
    errors = []
    with transaction.atomic():
        for idx, sid in enumerate(data):
            try:
                student = Student.objects.get(id=sid)
                if not student_in_scope(student, scopes_for(request.user)):
                    errors.append({"index": idx, "detail": f"Student {sid} not found."})
                    continue
            except Student.DoesNotExist:
                errors.append({"index": idx, "detail": f"Student {sid} not found."})
                continue
            old_sem = student.curr_semester_no
            new_sem = old_sem + 1
            
            # For PhD students, dynamically create next semester
            is_phd = hasattr(student, 'programme') and student.programme == 'PHD'
            
            try:
                semester_obj = Semester.objects.get(curriculum=student.batch_id.curriculum, semester_no=new_sem)
            except Semester.DoesNotExist:
                if is_phd:
                    # Create the semester for PhD students
                    semester_obj = Semester.objects.create(
                        curriculum=student.batch_id.curriculum,
                        semester_no=new_sem,
                        semester_name=f"Semester {new_sem}"
                    )
                    # Note: Admin needs to manually add thesis course to this semester via CourseSlot
                else:
                    errors.append({"index": idx, "detail": f"Semester {new_sem} not defined for student {sid}."})
                    continue
            student.curr_semester_no = new_sem
            student.save()
            frs = FinalRegistration.objects.filter(student_id=student, verified=False, semester_id = semester_obj)
            for fr in frs:
                course = fr.course_id
                exists = course_registration.objects.filter(
                    student_id=student,
                    course_id=course,
                    semester_id=semester_obj
                ).exists()
                session, semester_type = generate_next_session(date_time.year, new_sem)
                if not exists:
                    new_cr = course_registration.objects.create(
                        student_id=student,
                        working_year=None,
                        semester_id=semester_obj,
                        course_id=course,
                        course_slot_id=fr.course_slot_id,
                        registration_type=fr.registration_type,
                        session=session,
                        semester_type=semester_type
                    )
                    if fr.old_course_registration:
                        course_replacement.objects.create(
                            old_course_registration=fr.old_course_registration,
                            new_course_registration=new_cr
                        )
                fr.verified = True
                fr.save()
    if errors:
        return Response({"errors": errors}, status=status.HTTP_207_MULTI_STATUS)
    return Response({"detail": "Promotion applied."}, status=status.HTTP_200_OK)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def apply_demotion(request):
    """Move selected students one semester back (correction for over-promotion).

    Only decrements curr_semester_no (floored at 1); it does not delete any
    course registrations, so it is a safe inverse of an accidental promotion.
    """
    data = request.data  # list of student IDs
    errors = []
    with transaction.atomic():
        for idx, sid in enumerate(data):
            try:
                student = Student.objects.get(id=sid)
                if not student_in_scope(student, scopes_for(request.user)):
                    errors.append({"index": idx, "detail": f"Student {sid} not found."})
                    continue
            except Student.DoesNotExist:
                errors.append({"index": idx, "detail": f"Student {sid} not found."})
                continue
            old_sem = student.curr_semester_no
            if old_sem is None or old_sem <= 1:
                errors.append({"index": idx, "detail": f"Student {student.id_id} is already in semester 1; cannot demote."})
                continue
            student.curr_semester_no = old_sem - 1
            student.save()
    if errors:
        return Response({"errors": errors}, status=status.HTTP_207_MULTI_STATUS)
    return Response({"detail": "Demotion applied."}, status=status.HTTP_200_OK)


# @api_view(["GET"])
# @permission_classes([IsAuthenticated])
# def download_user_template(request):
#     columns = [
#         "username", "first_name", "last_name", "email", "gender", "date_of_birth",
#         "user_status", "address", "phone_no", "user_type", "department",
#         "title", "about_me",
#         "programme", "batch", "batch_id", "category",
#         "father_name", "mother_name", "hall_no", "room_no", "specialization", "curr_semester_no"
#     ]
#     df = pd.DataFrame(columns=columns)
#     buffer = io.BytesIO()
#     df.to_excel(buffer, index=False)
#     buffer.seek(0)
#     resp = HttpResponse(buffer, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#     resp["Content-Disposition"] = "attachment; filename=student_upload_template.xlsx"
#     return resp

# @api_view(["POST"])
# @permission_classes([IsAuthenticated])
# def upload_users(request):
#     f = request.FILES.get("file")
#     if not f:
#         return JsonResponse({"detail": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)
#     try:
#         df = pd.read_excel(f)
#     except Exception:
#         return JsonResponse({"detail": "Invalid Excel file."}, status=status.HTTP_400_BAD_REQUEST)
#     required = ["username", "first_name", "last_name", "email", "gender", "date_of_birth", "user_type", "programme", "batch", "category"]
#     errors = []
#     created = []
#     with transaction.atomic():
#         for idx, row in df.iterrows():
#             rownum = idx + 2
#             for field in required:
#                 if pd.isna(row.get(field)):
#                     errors.append({"row": rownum, "detail": f"{field} is required."})
#                     break
#             else:
#                 uname = str(row["username"]).strip()
#                 if User.objects.filter(username=uname).exists():
#                     errors.append({"row": rownum, "detail": "Username already exists."})
#                     continue
#                 email = str(row["email"]).strip()
#                 gender = str(row["gender"]).strip().upper()[0]
#                 dob = row["date_of_birth"]
#                 if isinstance(dob, datetime.date) is False:
#                     errors.append({"row": rownum, "detail": "Invalid date_of_birth."})
#                     continue
#                 user = User.objects.create_user(username=uname, email=email, password="user@123")
#                 user.first_name = str(row["first_name"]).strip()
#                 user.last_name = str(row["last_name"]).strip()
#                 user.save()
#                 eid = uname  # using username as ExtraInfo.id
#                 dept_name = str(row.get("department", "")).strip()
#                 dept = None
#                 if dept_name:
#                     dept, _ = DepartmentInfo.objects.get_or_create(name=dept_name)
#                 ei = ExtraInfo.objects.create(
#                     id=eid,
#                     user=user,
#                     title=str(row.get("title", "")).strip() or None,
#                     sex=gender,
#                     date_of_birth=dob,
#                     user_status=str(row.get("user_status", "")).strip() or None,
#                     address=str(row.get("address", "")).strip() or None,
#                     phone_no=int(row.get("phone_no")) if not pd.isna(row.get("phone_no")) else None,
#                     user_type=str(row["user_type"]).strip(),
#                     department=dept,
#                     about_me=str(row.get("about_me", "")).strip() or None,
#                 )
#                 batch_year = int(row["batch"])
#                 prog = str(row["programme"]).strip()
#                 cat = str(row["category"]).strip()
#                 batch_id_val = row.get("batch_id")
#                 batch_obj = None
#                 if not pd.isna(batch_id_val):
#                     try:
#                         batch_obj = Batch.objects.get(id=int(batch_id_val))
#                     except Batch.DoesNotExist:
#                         errors.append({"row": rownum, "detail": "Invalid batch_id."})
#                         continue
#                 student = Student.objects.create(
#                     id=ei,
#                     programme=prog,
#                     batch=batch_year,
#                     batch_id=batch_obj,
#                     cpi=float(row.get("cpi", 0)) if not pd.isna(row.get("cpi")) else 0,
#                     category=cat,
#                     father_name=str(row.get("father_name", "")).strip() or None,
#                     mother_name=str(row.get("mother_name", "")).strip() or None,
#                     hall_no=int(row.get("hall_no")) if not pd.isna(row.get("hall_no")) else 0,
#                     room_no=str(row.get("room_no", "")).strip() or None,
#                     specialization=str(row.get("specialization", "")).strip() or None,
#                     curr_semester_no=int(row.get("curr_semester_no")) if not pd.isna(row.get("curr_semester_no")) else 1
#                 )
#                 created.append(uname)
#     status_code = status.HTTP_207_MULTI_STATUS if errors else status.HTTP_201_CREATED
#     return JsonResponse({"created": created, "errors": errors}, status=status_code)


# ============================================================================
# PhD-SPECIFIC VIEW FUNCTIONS (Added for PhD student management)
# ============================================================================
# These functions handle PhD-specific workflows like thesis registration,
# seminar reports, RPC committees, and external review invitations.
# They are separate from UG/PG functions to maintain production stability.
# ============================================================================

def thesis_to_dict(t):
    """Serialize a ThesisTopic instance for JSON responses."""
    try:
        programme_category = t.student.batch_id.curriculum.programme.category
    except AttributeError:
        programme_category = None
    return {
        "id": t.id,
        "student_roll": t.student.id.id,
        "student_name": t.student.id.user.get_full_name(),
        "student_discipline": t.student.specialization,
        "programme_category": programme_category,
        "category": t.category,
        "broad_area": t.broad_area,
        "research_theme": t.research_theme,
        "supervisor": {"id": t.supervisor.id.id, "name": str(t.supervisor), "discipline": (t.supervisor.id.department.name if t.supervisor.id.department else "")},
        "co_supervisor": (
            {"id": t.co_supervisor.id.id, "name": str(t.co_supervisor), "discipline": (t.co_supervisor.id.department.name if t.co_supervisor.id.department else "")}
            if t.co_supervisor else None
        ),
        "supervisor_consented": t.supervisor_consented,
        "co_supervisor_consented": t.co_supervisor_consented,
        "external": {
            "ext_name": t.external_name,
            "ext_email": t.external_email,
            "ext_discipline": t.external_discipline,
            "ext_institution": t.external_institution,
        },
        "load": {
            "pg_single": t.pg_single,
            "pg_shared": t.pg_shared,
            "phd_single": t.phd_single,
            "phd_shared": t.phd_shared,
        },
        "committee": [
            {
                "id": cm.member.id.id,
                "name": str(cm.member),
                "discipline": (cm.member.id.department.name if cm.member.id.department else ""),
            }
            for cm in CommitteeMember.objects.filter(thesis = t).all()
        ],
        "status": t.status,
        "hod_remarks": t.hod_remarks,
        "dean_remarks" : t.dean_remarks
    }


# 1. Student APIs

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def student_thesis_api(request):
    """
    GET  /stu/thesis/             → fetch ({} if none)
    POST /stu/thesis/             → create/update when status == supervisor_pending or new
    """
    user = request.user
    try:
        user_details = user.extrainfo
        student = Student.objects.get(id=user_details)
    except Student.DoesNotExist:
        return JsonResponse({"error": "Student record not found"}, status=404)
    except Exception as e:
        return JsonResponse({"error": f"User setup error: {type(e).__name__}: {e}"}, status=400)

    try:
        thesis = ThesisTopic.objects.filter(student=student).order_by('-created_at').first()

        if request.method == 'GET':
            return JsonResponse(thesis_to_dict(thesis) if thesis else {}, status=200)
    except Exception as e:
        return JsonResponse({"error": f"Internal error: {type(e).__name__}: {e}"}, status=500)

    # POST: only if no thesis yet or status is supervisor_pending
    if thesis and thesis.status != 'supervisor_pending':
        return JsonResponse(
            {"error": "Cannot edit once under review past supervisor."},
            status=403
        )

    data = request.data
    required = ['supervisor_id', 'category', 'broad_area', 'research_theme']
    missing = [f for f in required if not data.get(f)]
    if missing:
        return JsonResponse({"error": f"Missing required field(s): {', '.join(missing)}"}, status=400)
    if not thesis:
        thesis = ThesisTopic(student=student)

    supervisor_id = data.get('supervisor_id')
    co_supervisor_id = data.get('co_supervisor_id')

    if co_supervisor_id and co_supervisor_id == supervisor_id:
        return JsonResponse({"error": "Co-supervisor must be different from supervisor"}, status=400)

    try:
        supervisor = Faculty.objects.get(pk=supervisor_id)
    except Faculty.DoesNotExist:
        return JsonResponse({"error": "Invalid supervisor"}, status=400)
    if not supervisor.id.user.is_active:
        return JsonResponse({"error": "Selected supervisor is not an active faculty member"}, status=400)

    if co_supervisor_id:
        try:
            co_supervisor = Faculty.objects.get(pk=co_supervisor_id)
        except Faculty.DoesNotExist:
            return JsonResponse({"error": "Invalid co-supervisor"}, status=400)
        if not co_supervisor.id.user.is_active:
            return JsonResponse({"error": "Selected co-supervisor is not an active faculty member"}, status=400)

    thesis.category            = data.get('category')
    thesis.broad_area          = data.get('broad_area')
    thesis.research_theme      = data.get('research_theme')
    thesis.supervisor_id       = supervisor_id
    thesis.co_supervisor_id    = co_supervisor_id
    thesis.external_name       = data.get('external_name', '')
    thesis.external_email      = data.get('external_email', '')
    thesis.external_discipline = data.get('external_discipline', '')
    thesis.external_institution= data.get('external_institution', '')
    thesis.status              = 'supervisor_pending'
    # Any edit invalidates consents already given (e.g. supervisor consented
    # while co-supervisor consent is still pending) -- otherwise a changed
    # supervisor/topic could ride through on a stale consent.
    thesis.supervisor_consented    = False
    thesis.co_supervisor_consented = False
    thesis.save()

    _thesis_notify(
        sender=user,
        recipient=supervisor.id.user,
        verb='New thesis topic proposal awaiting your review',
        description=f"{student.id.user.get_full_name()} has submitted a thesis topic proposal for your review.",
    )
    if co_supervisor_id:
        _thesis_notify(
            sender=user,
            recipient=co_supervisor.id.user,
            verb='New thesis topic proposal awaiting your review',
            description=f"{student.id.user.get_full_name()} has submitted a thesis topic proposal naming you as co-supervisor.",
        )

    return JsonResponse(thesis_to_dict(thesis), status=201)


from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer,
    Table, TableStyle, Image
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_download_pdf_api(request):
    thesis = get_object_or_404(ThesisTopic, student__id=request.user.extrainfo)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=20 * mm,
        bottomMargin=15 * mm
    )

    styles = getSampleStyleSheet()
    normal = styles['Normal']
    bold = ParagraphStyle('Bold', parent=normal, fontName='Helvetica-Bold')
    elements = []

    # Header
    logo = Image('./media/logo2.jpg', width=25 * mm, height=25 * mm)
    college_name = Paragraph(
        '<b>Indian Institute of Information Technology, Design and Manufacturing, Jabalpur</b><br/>',
        ParagraphStyle('Header', parent=styles['Title'], alignment=1)
    )
    header_tbl = Table([[logo, college_name]], colWidths=[30 * mm, 150 * mm])
    header_tbl.setStyle(TableStyle([
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN',       (1, 0), (1, 0), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.extend([header_tbl, Spacer(1, 12)])
    elements.extend([Paragraph('<u>Thesis Topic Submission Form</u>', styles['Heading2']), Spacer(1, 20)])

    # Form fields data
    data = [
        [Paragraph('<b>Roll Number:</b>', bold), thesis.student.id.id],
        [Paragraph('<b>Student Name:</b>', bold), thesis.student.id.user.get_full_name()],
        [Paragraph('<b>Discipline:</b>', bold), thesis.student.specialization],
        [Paragraph('<b>Category:</b>', bold), thesis.category],
        [Paragraph('<b>Broad Area:</b>', bold), thesis.broad_area],
        [Paragraph('<b>Research Theme:</b>', bold),
         Paragraph(thesis.research_theme.replace('\n', '<br/>'), normal)],
        [Paragraph('<b>Supervisor:</b>', bold), thesis.supervisor.id.user.get_full_name()],
    ]
    if thesis.co_supervisor:
        data.append([Paragraph('<b>Co-Supervisor:</b>', bold), thesis.co_supervisor.id.user.get_full_name()])
    if thesis.external_name:
        data.extend([
            [Paragraph('<b>External Supervisor:</b>', bold), thesis.external_name],
            [Paragraph('<b>Email:</b>', bold), thesis.external_email],
            [Paragraph('<b>Discipline:</b>', bold), thesis.external_discipline],
            [Paragraph('<b>Institution:</b>', bold), thesis.external_institution],
        ])

    # Create the form table with increased row heights
    form_tbl = Table(
        data,
        colWidths=[55 * mm, 125 * mm],
        rowHeights=[13 * mm] * len(data)  # each row is 15 mm tall
    )
    form_tbl.setStyle(TableStyle([
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.grey),
        ('VALIGN',        (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING',   (0, 0), (-1, -1), 8),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 8),
        ('TOPPADDING',    (0, 0), (-1, -1), 10),  # extra breathing room
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('BACKGROUND',    (0, 0), (0, -1), colors.whitesmoke),
    ]))
    elements.extend([form_tbl, Spacer(1, 40)])

    # Signatures: two per row
    sig_line = '__________    Date: _______'
    row1 = [
        Paragraph('<b>Supervisor Sig.:</b>', bold), sig_line,
        Paragraph('<b>Co-Supervisor Sig.:</b>', bold) if thesis.co_supervisor else '',
        sig_line if thesis.co_supervisor else ''
    ]
    sig_tbl = Table([row1], colWidths=[30 * mm, 60 * mm, 30 * mm, 60 * mm])
    sig_tbl.setStyle(TableStyle([
        ('VALIGN',        (0, 0), (-1, -1), 'BOTTOM'),
        ('LEFTPADDING',   (0, 0), (-1, -1), 4),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
        ('TOPPADDING',    (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(sig_tbl)

    # Build and return PDF
    doc.build(elements)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')
# 2. Faculty list for dropdowns

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def faculty_list_api(request):
    """
    GET /faculty/ → all faculty {id, name, discipline}
    """
    qs = Faculty.objects.select_related('id__user', 'id__department')
    data = []
    for f in qs:
        user = f.id.user
        dept = f.id.department
        data.append({
            'id': f.id.id,
            'name': f"{user.first_name} {user.last_name}",
            'discipline': dept.name if dept else '',
        })
    return JsonResponse(data, safe=False)


# 3. Supervisor endpoints
from django.db import models

from django.db.models import Q
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from django.http import JsonResponse

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def supervisor_thesis_topic_dashboard(request):
    """
    GET /supervisor/dashboard/
    → returns { pending, forwarded }
      for any thesis where request.user is either supervisor OR co_supervisor.
    """
    ex = request.user

    qs = ThesisTopic.objects.filter(
        Q(supervisor__id=ex.username) | Q(co_supervisor__id=ex.username)
    )

    pending_statuses = ['supervisor_pending', 'hod_rejected']
    pending_qs = qs.filter(status__in=pending_statuses)

    forwarded_qs = qs.exclude(status__in=pending_statuses)

    def serialize_for_viewer(t):
        d = thesis_to_dict(t)
        is_sup = t.supervisor_id == ex.username
        is_co = bool(t.co_supervisor_id) and t.co_supervisor_id == ex.username
        d['is_supervisor'] = is_sup
        d['is_co_supervisor'] = is_co
        d['my_consent_given'] = (
            t.supervisor_consented if is_sup
            else t.co_supervisor_consented if is_co
            else False
        )
        return d

    return JsonResponse({
        'pending':   [serialize_for_viewer(t) for t in pending_qs],
        'forwarded': [serialize_for_viewer(t) for t in forwarded_qs],
    })


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def supervisor_review_api(request, pk):
    thesis = get_object_or_404(ThesisTopic, pk=pk)
    user_ex   = request.user.username
    is_sup    = (thesis.supervisor_id == user_ex)
    is_co     = (thesis.co_supervisor and thesis.co_supervisor_id == user_ex)

    if request.method == 'GET':
        if not (is_sup or is_co):
            return JsonResponse({"error": "Access denied."}, status=403)
        payload = thesis_to_dict(thesis)
        payload.update({"is_supervisor": is_sup, "is_co_supervisor": is_co})
        return JsonResponse(payload, status=200)

    if thesis.status != 'supervisor_pending' and thesis.status != 'hod_rejected':
        return JsonResponse({"error": "Cannot review at this stage."}, status=403)

    data = request.data

    if 'research_theme' in data and data['research_theme'] != thesis.research_theme:
        thesis.research_theme = data['research_theme']
        # Content changed -- any consent already given (e.g. supervisor
        # consented while co-supervisor's is still pending) no longer
        # reflects what's being approved, so it must be re-given.
        thesis.supervisor_consented    = False
        thesis.co_supervisor_consented = False

    if is_co and not is_sup:
        if thesis.co_supervisor_consented:
            return JsonResponse({"error": "Already consented."}, status=400)
        if data.get('co_supervisor_consented'):
            thesis.co_supervisor_consented = True
            thesis.save()
            _thesis_notify(
                sender=request.user,
                recipient=thesis.supervisor.id.user,
                verb='Co-supervisor has consented to the thesis topic',
                description=f"{thesis.co_supervisor.id.user.get_full_name()} has consented to "
                            f"{thesis.student.id.user.get_full_name()}'s thesis topic.",
            )
            return JsonResponse({"message": "Co-Supervisor consent recorded."}, status=200)
        return JsonResponse({"error": "Invalid consent payload."}, status=400)

    if is_sup:

        if not (thesis.supervisor_consented and
                (not thesis.co_supervisor or thesis.co_supervisor_consented)):

            # Committee/PG-load edits change what's being approved -- a
            # co-supervisor consent already given against the old values
            # must not silently carry over.
            thesis.co_supervisor_consented = False

            thesis.pg_single  = data.get('pg_single', thesis.pg_single)
            thesis.pg_shared  = data.get('pg_shared', thesis.pg_shared)
            thesis.phd_single = data.get('phd_single', thesis.phd_single)
            thesis.phd_shared = data.get('phd_shared', thesis.phd_shared)

            # This committee doubles as the live RPC for Comprehensive Exam /
            # Open Seminar (_exam_rpc_committee). Editing membership while an
            # attempt is actively rpc_pending would let a member be dropped
            # (or dropped-then-re-added) mid-review, letting finalize succeed
            # without their consent or silently reusing a stale one -- so
            # membership is frozen until that review reaches a decision.
            if ComprehensiveExamAttempt.objects.filter(exam__student=thesis.student, status='rpc_pending').exists() or \
               OpenSeminarAttempt.objects.filter(open_seminar__student=thesis.student, status='rpc_pending').exists():
                return JsonResponse(
                    {"error": "Cannot edit committee while a Comprehensive Exam or Open Seminar is awaiting RPC consent."},
                    status=403
                )

            CommitteeMember.objects.filter(thesis=thesis).delete()
            for member_id in data.get('committee', []):
                CommitteeMember.objects.create(thesis=thesis, member_id=member_id)

            CommitteeMember.objects.get_or_create(thesis=thesis, member_id=thesis.supervisor_id)
            if thesis.co_supervisor_id:
                CommitteeMember.objects.get_or_create(thesis=thesis, member_id=thesis.co_supervisor_id)

        if not thesis.supervisor_consented and data.get('supervisor_consented'):
            thesis.supervisor_consented = True

        sup_ok = thesis.supervisor_consented
        co_ok  = (not thesis.co_supervisor) or thesis.co_supervisor_consented

        if sup_ok and co_ok:
            total_rpc = CommitteeMember.objects.filter(thesis=thesis).count()
            if total_rpc < 3:
                return JsonResponse(
                    {"error": "Need at least 3 RPC members (including supervisor/co-supervisor)."},
                    status=400
                )
            thesis.status = 'hod_pending'
            thesis.save()
            student_discipline = (
                thesis.student.batch_id.discipline.acronym
                if thesis.student.batch_id and thesis.student.batch_id.discipline else None
            )
            _thesis_notify(
                sender=request.user,
                recipient=_hod_users_for_discipline(student_discipline),
                verb='Thesis topic pending your review',
                description=f"{thesis.student.id.user.get_full_name()}'s thesis topic has been "
                            f"forwarded by the supervisor and co-supervisor for your review.",
            )
            return JsonResponse(
                {"message": "Forwarded to HOD successfully.", "status": thesis.status},
                status=200
            )

        thesis.save()
        return JsonResponse(
            {"message": "Supervisor changes saved; awaiting all consents and RPC ≥ 3."},
            status=200
        )

    return JsonResponse({"error": "Not authorized."}, status=403)


def get_hod_disciplines(user):
    """Discipline acronyms this user is HOD of, parsed from designation
    names like 'HOD (CSE)' -> 'CSE'. Used to scope a dashboard listing to
    every discipline the user is HOD for."""
    hod_designations = HoldsDesignation.objects.filter(
        working=user,
        designation__name__icontains='HOD'
    ).values_list('designation__name', flat=True)

    hod_disciplines = []
    for des_name in hod_designations:
        if '(' in des_name and ')' in des_name:
            discipline = des_name[des_name.index('(')+1:des_name.index(')')].strip()
            hod_disciplines.append(discipline)
    return hod_disciplines


def is_hod_of_discipline(user, discipline_acronym):
    """True if `user` holds the exact 'HOD (<discipline_acronym>)' designation.
    Used to authorize a single action against one specific discipline."""
    if not discipline_acronym:
        return False
    return HoldsDesignation.objects.filter(
        working=user,
        designation__name=f"HOD ({discipline_acronym})"
    ).exists()


def _users_holding_designation(designation_name):
    """Users currently acting in the given designation (via HoldsDesignation.working
    -- the field documented as the correct one for permissions/current-holder lookups,
    covering officiating/temporary holders too, not just the permanent one)."""
    return User.objects.filter(
        current_designation__designation__name=designation_name
    ).distinct()


def _hod_users_for_discipline(discipline_acronym):
    """Users currently acting as HOD for one specific discipline acronym (e.g. 'CSE')."""
    if not discipline_acronym:
        return User.objects.none()
    return _users_holding_designation(f"HOD ({discipline_acronym})")


def _dean_academic_users():
    return _users_holding_designation('Dean Academic')


def _notify(sender, recipient, verb, description='', module=''):
    """Thin wrapper around notify.send for PhD workflow notifications.
    `recipient` may be a single User, None, or a queryset/list of Users --
    falsy/empty recipients are silently skipped rather than erroring."""
    if recipient is None:
        return
    if hasattr(recipient, 'exists') and not recipient.exists():
        return
    notify.send(
        sender=sender,
        recipient=recipient,
        verb=verb,
        description=description,
        url='',
        module=module,
    )


def _thesis_notify(sender, recipient, verb, description=''):
    _notify(sender, recipient, verb, description, module='Thesis Topic')


def _comprehensive_exam_notify(sender, recipient, verb, description=''):
    _notify(sender, recipient, verb, description, module='Comprehensive Exam')


def _open_seminar_notify(sender, recipient, verb, description=''):
    _notify(sender, recipient, verb, description, module='Open Seminar')


def _teaching_credit_notify(sender, recipient, verb, description=''):
    _notify(sender, recipient, verb, description, module='Teaching Credit')


def _progress_seminar_notify(sender, recipient, verb, description=''):
    _notify(sender, recipient, verb, description, module='Progress Seminar')


def _phd_course_registration_notify(sender, recipient, verb, description=''):
    _notify(sender, recipient, verb, description, module='PhD Course Registration')


def _academic_office_users():
    return _users_holding_designation('acadadmin')


# 4. HOD endpoints

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def hod_dashboard(request):
    """
    GET /hod/dashboard/ → { pending, approved, rejected }
    filtered by HOD designation held by the user.
    
    - pending statuses: ['dean_rejected', 'hod_pending']
    - approved statuses: ['hod_approved', 'dean_approved']
    - rejected statuses: ['hod_rejected']
    """
    user = request.user
    data = {'pending': [], 'approved': [], 'rejected': []}

    STATUS_PENDING  = ['dean_rejected', 'hod_pending']
    STATUS_APPROVED = ['hod_approved', 'dean_approved']
    STATUS_REJECTED = ['hod_rejected']
    all_statuses = STATUS_PENDING + STATUS_APPROVED + STATUS_REJECTED

    # Get HOD designations for this user
    hod_disciplines = get_hod_disciplines(user)

    qs = ThesisTopic.objects.filter(status__in=all_statuses).select_related('student', 'student__batch_id', 'student__batch_id__discipline')

    for thesis in qs:
        # Check if thesis student's discipline matches HOD's discipline
        # Use discipline acronym (e.g., "CSE") to match with HOD designation (e.g., "HOD (CSE)")
        student_discipline_acronym = None
        if thesis.student.batch_id and thesis.student.batch_id.discipline:
            student_discipline_acronym = thesis.student.batch_id.discipline.acronym
        
        if not student_discipline_acronym or student_discipline_acronym not in hod_disciplines:
            continue

        dto = thesis_to_dict(thesis)

        if thesis.status in STATUS_PENDING:
            data['pending'].append(dto)
        elif thesis.status in STATUS_APPROVED:
            data['approved'].append(dto)
        else:  # thesis.status in STATUS_REJECTED
            data['rejected'].append(dto)

    return JsonResponse(data)


@api_view(['GET','POST'])
@permission_classes([IsAuthenticated])
def hod_review_api(request, pk):
    thesis = get_object_or_404(ThesisTopic, pk=pk)
    user = request.user
    
    # Check if user is HOD for the student's discipline
    student_discipline_acronym = None
    if thesis.student.batch_id and thesis.student.batch_id.discipline:
        student_discipline_acronym = thesis.student.batch_id.discipline.acronym

    is_hod = is_hod_of_discipline(user, student_discipline_acronym)

    if request.method == 'GET':
        if not is_hod:
            return JsonResponse({"error": "Access denied."}, status=403)
        data = thesis_to_dict(thesis)
        return JsonResponse(data, status=200)

    # POST
    if not is_hod or thesis.status not in ['hod_pending','hod_rejected','dean_pending','dean_rejected']:
        return JsonResponse({"error":"Forbidden or invalid stage"}, status=403)

    d = request.data
    if d.get('approve'):
        thesis.status      = 'hod_approved'
        thesis.hod_remarks = ''
        thesis.save()
        _thesis_notify(
            sender=user,
            recipient=_dean_academic_users(),
            verb='Thesis topic pending your final approval',
            description=f"{thesis.student.id.user.get_full_name()}'s thesis topic has been "
                        f"approved by the HOD and is awaiting your final approval.",
        )
        for recipient in filter(None, [thesis.student.id.user, thesis.supervisor.id.user]):
            _thesis_notify(
                sender=user,
                recipient=recipient,
                verb='Thesis topic approved by HOD',
                description=f"{thesis.student.id.user.get_full_name()}'s thesis topic has been "
                            f"approved by the HOD and forwarded to Dean Academic.",
            )
    else:
        thesis.status      = 'hod_rejected'
        thesis.hod_remarks = d.get('remarks','')
        thesis.supervisor_consented    = False
        thesis.co_supervisor_consented = False
        thesis.dean_remarks            = ''
        thesis.save()
        for recipient in filter(None, [thesis.supervisor.id.user, thesis.student.id.user]):
            _thesis_notify(
                sender=user,
                recipient=recipient,
                verb='Thesis topic rejected by HOD',
                description=f"The HOD rejected {thesis.student.id.user.get_full_name()}'s thesis "
                            f"topic. Remarks: {thesis.hod_remarks or '—'}",
            )

    return JsonResponse({"status":thesis.status}, status=200)


# 5. Dean endpoints

@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(['Dean Academic'])
def dean_dashboard(request):
    """
    GET /dean/dashboard/ → {pending, approved}
    for theses with status in dean_pending/dean_approved.
    """
    data = {'pending': [], 'approved': [], 'rejected':[]}
    qs = ThesisTopic.objects.filter(status__in=['dean_pending','dean_approved', 'hod_approved'])
    for t in qs:
        dto = thesis_to_dict(t)
        bucket = 'pending' if t.status=='dean_pending' or t.status=='hod_approved' else \
            'approved' if t.status=='dean_approved' else 'rejected'
        data[bucket].append(dto)
    return JsonResponse(data)


@api_view(['GET','POST'])
@permission_classes([IsAuthenticated])
@role_required(['Dean Academic'])
def dean_review_api(request, pk):
    thesis = get_object_or_404(ThesisTopic, pk=pk)
    if request.method == 'GET':
        data = thesis_to_dict(thesis)
        return JsonResponse(data, status=200)

    # POST
    if thesis.status not in ['dean_pending','hod_approved']:
        return JsonResponse({"error":"Forbidden or invalid stage"}, status=403)

    d = request.data
    student_discipline = (
        thesis.student.batch_id.discipline.acronym
        if thesis.student.batch_id and thesis.student.batch_id.discipline else None
    )

    if d.get('approve'):
        thesis.status       = 'dean_approved'
        thesis.dean_remarks = ''
        thesis.save()
        recipients = [thesis.student.id.user, thesis.supervisor.id.user]
        if thesis.co_supervisor:
            recipients.append(thesis.co_supervisor.id.user)
        for recipient in recipients:
            _thesis_notify(
                sender=request.user,
                recipient=recipient,
                verb='Thesis topic approved by Dean Academic',
                description=f"{thesis.student.id.user.get_full_name()}'s thesis topic has "
                            f"received final approval from Dean Academic.",
            )
        _thesis_notify(
            sender=request.user,
            recipient=_hod_users_for_discipline(student_discipline),
            verb='Thesis topic approved by Dean Academic',
            description=f"{thesis.student.id.user.get_full_name()}'s thesis topic has "
                        f"received final approval from Dean Academic.",
        )
    else:
        thesis.status       = 'dean_rejected'
        thesis.dean_remarks = d.get('remarks','')
        thesis.save()
        _thesis_notify(
            sender=request.user,
            recipient=_hod_users_for_discipline(student_discipline),
            verb='Thesis topic rejected by Dean Academic',
            description=f"Dean Academic rejected {thesis.student.id.user.get_full_name()}'s "
                        f"thesis topic. Remarks: {thesis.dean_remarks or '—'}",
        )

    return JsonResponse({"status":thesis.status}, status=200)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(['Dean Academic'])
def dean_generate_pdf_api(request, pk):
    thesis = get_object_or_404(ThesisTopic, pk=pk)
    if thesis.status != 'dean_approved':
        return HttpResponse({"error": "Not fully approved"}, status=403)

    buffer = BytesIO()
    student_roll = thesis.student.id.id.replace(' ', '_')
    filename = f"approved_thesis_{student_roll}.pdf"
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=10 * mm,
        title=filename,
    )

    styles = getSampleStyleSheet()
    normal = styles['Normal']
    normal.fontSize = 9
    normal.leading = 11
    bold = ParagraphStyle('Bold', parent=normal, fontName='Helvetica-Bold', fontSize=9)
    title_center = ParagraphStyle('TitleCenter', parent=styles['Title'], alignment=1, fontSize=12)
    heading2 = ParagraphStyle('H2', parent=styles['Heading2'], fontSize=11, spaceAfter=6)
    heading3 = ParagraphStyle('H3', parent=styles['Heading3'], fontSize=10, spaceAfter=4)

    elements = []

    # Header
    logo = Image('./media/logo2.jpg', width=22*mm, height=22*mm)
    institute = Paragraph(
        '<b>Indian Institute of Information Technology, Design and Manufacturing, Jabalpur</b>',
        title_center
    )
    header = Table([[logo, institute]], colWidths=[28*mm, 152*mm])
    header.setStyle(TableStyle([
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN',       (1, 0), (1, 0), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING',(0, 0), (-1, -1), 0),
    ]))
    elements += [header, Spacer(1, 8)]
    elements += [Paragraph('<u>Thesis Approval Summary</u>', heading2), Spacer(1, 6)]

    # Form fields
    form_data = [
        ['Roll Number', thesis.student.id.id],
        ['Student Name', thesis.student.id.user.get_full_name()],
        ['Discipline', thesis.student.specialization],
        ['Category', thesis.category],
        ['Broad Area', thesis.broad_area],
    ]
    # research theme as separate row with wrapping
    form_data.append(['Research Theme',
                      Paragraph(thesis.research_theme.replace('\n', '<br/>'), normal)])
    if thesis.co_supervisor:
        form_data.append(['Co-Supervisor', thesis.co_supervisor.id.user.get_full_name()])
    if thesis.external_name:
        form_data += [
            ['External Supervisor', thesis.external_name],
            ['Email', thesis.external_email],
            ['External Discipline', thesis.external_discipline],
            ['Institution', thesis.external_institution],
        ]

    # All tables same total width: use available width = 160mm
    total_width = 160 * mm
    col1 = 50 * mm
    col2 = total_width - col1

    # Form table with reduced row height
    form_tbl = Table(form_data, colWidths=[col1, col2], rowHeights=[8*mm]*len(form_data))
    form_tbl.setStyle(TableStyle([
        ('GRID',         (0, 0), (-1, -1), 0.4, colors.grey),
        ('VALIGN',       (0, 0), (-1, -1), 'TOP'),
        ('BACKGROUND',   (0, 0), (0, -1), colors.whitesmoke),
        ('LEFTPADDING',  (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING',   (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 3),
        ('FONTSIZE',     (0, 0), (-1, -1), 9),
    ]))
    elements += [form_tbl, Spacer(1, 8)]

    # Supervision Load
    load_data = [
        ['Category', 'Single', 'Shared'],
        ['PG', str(thesis.pg_single), str(thesis.pg_shared)],
        ['PhD', str(thesis.phd_single), str(thesis.phd_shared)],
    ]
    load_tbl = Table(load_data, colWidths=[col1, (total_width-col1)/2, (total_width-col1)/2], rowHeights=[7*mm]*3)
    load_tbl.setStyle(TableStyle([
        ('GRID',       (0, 0), (-1, -1), 0.4, colors.grey),
        ('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN',      (1, 1), (-1, -1), 'CENTER'),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING',(0,0),(-1,-1),3),('RIGHTPADDING',(0,0),(-1,-1),3),
        ('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2),
        ('FONTSIZE',   (0, 0), (-1, -1), 9),
    ]))
    elements += [Paragraph('<b>Supervision Load</b>', heading3), load_tbl, Spacer(1, 8)]

    # Committee
    comm = [['Member', 'Discipline']]
    for cm in thesis.committee.all():
        comm.append([cm.member.id.user.get_full_name(), cm.member.id.department.name or ''])
    comm_tbl = Table(comm, colWidths=[col1, col2], rowHeights=[7*mm]*len(comm))
    comm_tbl.setStyle(TableStyle([
        ('GRID',       (0, 0), (-1, -1), 0.4, colors.grey),
        ('BACKGROUND',(0, 0), (-1, 0), colors.whitesmoke),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING',(0,0),(-1,-1),3),('RIGHTPADDING',(0,0),(-1,-1),3),
        ('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2),
        ('FONTSIZE',   (0, 0), (-1, -1), 9),
    ]))
    elements += [Paragraph('<b>RPC Committee Members</b>', heading3), comm_tbl, Spacer(1, 8)]

    # Signatures: one per line (compact but readable)
    sig_labels = [
        ('Supervisor Signature:', 'Date:'),
        ('Co-Supervisor Signature:', 'Date:') if thesis.co_supervisor else None,
        ('HOD Signature:', 'Date:'),
        ('Dean Signature:', 'Date:')
    ]
    # Ensure all signature rows have equal, large vertical spacing
    # Use moderate spacing and group all signature rows to avoid page break
    sig_row_space = 18  # mm, balanced for single page
    sig_tables = []
    for label_pair in sig_labels:
        if label_pair:
            label = label_pair[0]
            sig_tbl = Table(
                [[
                    Paragraph(f'<b>{label}</b>', bold),
                    '__________________________',
                    Paragraph(f'<b>{label_pair[1]}</b>', bold),
                    '_______________'
                ]],
                colWidths=[45*mm, 55*mm, 15*mm, 45*mm],
                rowHeights=[12*mm]  # force equal height for all signature rows
            )
            sig_tbl.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'BOTTOM'),
                ('LEFTPADDING',(0,0),(-1,-1),2),
                ('RIGHTPADDING',(0,0),(-1,-1),2),
                # Remove any background for all rows
            ]))
            sig_tables += [sig_tbl, Spacer(1, sig_row_space)]
    # Use KeepTogether to prevent page break in signature block
    from reportlab.platypus import KeepTogether
    elements.append(KeepTogether(sig_tables))

    doc.build(elements)
    buffer.seek(0)
    
    # Set proper filename without spaces
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = (
        f"attachment; filename=\"{filename}\"; filename*=UTF-8''{filename}"
    )
    return response


# Seminar Views
# 1. STUDENT

def _progress_seminar_catalog_entry(semester):
    """The catalog Seminar (code/name) tied to a semester's progress seminar
    slot, if one has been configured — analogous to a thesis slot's catalog
    Thesis. Returns (code, name), either possibly None."""
    if semester is None:
        return None, None
    slot = ProgressSeminarSlot.objects.filter(semester=semester).first()
    if slot is None:
        return None, None
    catalog = slot.seminars.first()
    if catalog is None:
        return None, None
    return catalog.code, catalog.name


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_reports(request):
    thesis = get_object_or_404(ThesisTopic, student_id=request.user.username)
    data = []
    for s in thesis.seminars.order_by('version'):
        seminar_code, seminar_name = _progress_seminar_catalog_entry(s.semester)
        data.append({
            "id":            s.id,
            "version":       s.version,
            "semester_no":   s.semester.semester_no if s.semester else None,
            "seminar_code":  seminar_code,
            "seminar_name":  seminar_name,
            "status":        s.status,
            "created_at":    s.created_at.isoformat(),
        })

    return JsonResponse(data, safe=False)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_report(request, thesis_pk):
    thesis = get_object_or_404(ThesisTopic, pk=thesis_pk, student_id=request.user.username)
    if thesis.status != 'dean_approved':
        return JsonResponse({"error":"Thesis not Dean-approved."}, status=403)

    if thesis.seminars.filter(status='rpc_pending').exists():
        return JsonResponse(
            {"error": "A previous seminar report is still awaiting RPC consent."},
            status=403,
        )

    # versioning
    last = thesis.seminars.order_by('-version').first()
    version = (last.version + 1) if last else 1

    student = thesis.student
    try:
        semester = Semester.objects.get(
            curriculum=student.batch_id.curriculum,
            semester_no=student.curr_semester_no,
        )
    except Semester.DoesNotExist:
        semester = None

    if semester is None or not ProgressSeminarRegistration.objects.filter(
        student=student, semester=semester, status='verified',
    ).exists():
        return JsonResponse(
            {"error": "No verified Progress Seminar registration for the current semester."},
            status=403,
        )

    seminar = ProgressSeminarEntry.objects.create(
        thesis=thesis,
        version=version,
        semester=semester,
        status='rpc_pending',
        seminar_date  = request.data.get('date') or None,
        seminar_time  = request.data.get('time') or None,
        seminar_venue = request.data.get('venue',''),
        summary_prev  = request.data.get('prev',''),
        summary_curr  = request.data.get('curr',''),
        future_plan   = request.data.get('future',''),
        upload_doc    = request.FILES.get('doc', None),
        pub_published_or_accepted  = int(request.data.get('pub_published_or_accepted', 0) or 0),
        pub_presented_unpublished  = int(request.data.get('pub_presented_unpublished', 0) or 0),
        pub_submitted_under_review = int(request.data.get('pub_submitted_under_review', 0) or 0),
    )

    _progress_seminar_notify(
        sender=request.user,
        recipient=_rpc_committee_users(student),
        verb='Progress Seminar report pending your consent',
        description=f"{student.id.user.get_full_name()} has submitted a Progress Seminar "
                    f"report (version {version}) awaiting RPC consent.",
    )

    return JsonResponse({
        "id": seminar.id,
        "message": "Seminar submitted; awaiting RPC consent."
    }, status=201)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def detail_report(request, pk):
    s = get_object_or_404(ProgressSeminarEntry, pk=pk, thesis__student_id=request.user.username)
    return JsonResponse({
        "id":          s.id,
        "version":     s.version,
        "semester_no": s.semester.semester_no if s.semester else None,
        "status":      s.status,
        "date":    str(s.seminar_date or ""),
        "time":    str(s.seminar_time or ""),
        "venue":   s.seminar_venue,
        "prev":    s.summary_prev,
        "curr":    s.summary_curr,
        "future":  s.future_plan,
        "doc_url": s.upload_doc.url if s.upload_doc else None,
        "pub_published_or_accepted":  s.pub_published_or_accepted,
        "pub_presented_unpublished":  s.pub_presented_unpublished,
        "pub_submitted_under_review": s.pub_submitted_under_review,
    })

# 2. RPC

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def rpc_seminar_list(request):
    faculty = get_object_or_404(Faculty, id__user=request.user)
    all_entries = ProgressSeminarEntry.objects.filter(
        thesis__committee__member=faculty
    ).distinct()

    def serialize(qs):
        return [
            {
                "id":               s.id,
                "version":          s.version,
                "semester_no":      s.semester.semester_no if s.semester else None,
                "roll_number":      s.thesis.student.id.id,
                "student":          s.thesis.student.id.user.get_full_name(),
                "thesis":           s.thesis.research_theme,
                "status":           s.status,
                "my_consent_given": ProgressSeminarConsent.objects.filter(
                    seminar=s, member=faculty, consented=True
                ).exists(),
            }
            for s in qs
        ]

    return Response({
        "pending":  serialize(all_entries.filter(status='rpc_pending')),
        "approved": serialize(all_entries.filter(status='rpc_approved')),
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def rpc_detail(request, pk):
    faculty = get_object_or_404(Faculty, id__user=request.user)
    seminar = get_object_or_404(ProgressSeminarEntry, pk=pk)
    if not CommitteeMember.objects.filter(thesis=seminar.thesis, member=faculty).exists():
        return JsonResponse({"error": "Not on committee."}, status=403)
    
    student_extra = seminar.thesis.student.id
    student_name  = student_extra.user.get_full_name()
    roll_number   = student_extra.user.username
    discipline    = seminar.thesis.student.specialization
    thesis_title  = seminar.thesis.research_theme

    panel = {
        f: getattr(seminar, f) for f in [
            'quality', 'quantity', 'overall_grade', 'expected_period',
            'rec_assist', 'rec_enhance', 'rec_repeat', 'rec_open'
        ]
    }

    committee = []
    for cm in CommitteeMember.objects.filter(thesis=seminar.thesis).select_related('member__id__user', 'member__id__department'):
        fac = cm.member
        extra = fac.id
        consented = ProgressSeminarConsent.objects.filter(seminar=seminar, member=fac, consented=True).exists()
        committee.append({
            "id": extra.id,
            "name": f"{extra.user.first_name} {extra.user.last_name}",
            "discipline": extra.department.name if extra.department else "",
            "consented": consented,
        })

    comments = [
        {
            "member": c.member.id.user.get_full_name(),
            "text": c.text,
            "timestamp": c.timestamp.isoformat()
        }
        for c in seminar.comments.all()
    ]

    my_comment = ProgressSeminarComment.objects.filter(seminar=seminar, member=faculty).first()
    is_consented = ProgressSeminarConsent.objects.filter(seminar=seminar, member=faculty, consented=True).exists()

    payload = {
        "studentName":  student_name,
        "rollNumber":   roll_number,
        "discipline":   discipline,
        "thesisTitle":  thesis_title,
        "programme_category": _student_programme_category(seminar.thesis.student),
        "id": seminar.id,
        "version": seminar.version,
        "semester_no": seminar.semester.semester_no if seminar.semester else None,
        "date": seminar.seminar_date.isoformat() if seminar.seminar_date else "",
        "time": seminar.seminar_time.isoformat() if seminar.seminar_time else "",
        "venue": seminar.seminar_venue,
        "prev": seminar.summary_prev,
        "curr": seminar.summary_curr,
        "future": seminar.future_plan,
        "doc_url": seminar.upload_doc.url if seminar.upload_doc else None,
        "pub_published_or_accepted":  seminar.pub_published_or_accepted,
        "pub_presented_unpublished":  seminar.pub_presented_unpublished,
        "pub_submitted_under_review": seminar.pub_submitted_under_review,
        **panel,
        "committee": committee,
        "committeeSize": len(committee),
        "consentedCount": sum(1 for m in committee if m["consented"]),
        "comments": comments,
        "myComment": my_comment.text if my_comment else "",
        "isConsented": is_consented,
        "status": seminar.status,
    }

    return JsonResponse(payload)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def rpc_consent(request, pk):
    faculty = get_object_or_404(Faculty, id__user=request.user)
    seminar = get_object_or_404(ProgressSeminarEntry, pk=pk, status='rpc_pending')
    if not CommitteeMember.objects.filter(thesis=seminar.thesis, member=faculty).exists():
        return JsonResponse({"error": "Not on committee."}, status=403)

    data = request.data
    panel_fields = [
        'quality', 'quantity', 'overall_grade', 'expected_period',
        'rec_assist', 'rec_enhance', 'rec_repeat', 'rec_open'
    ]

    changed = any(
        field in data and getattr(seminar, field) != data[field]
        for field in panel_fields
    )
    if changed:
        ProgressSeminarConsent.objects.filter(seminar=seminar).update(consented=False)

    for field in panel_fields:
        if field in data:
            setattr(seminar, field, data[field])
    seminar.save()

    if 'comment' in data:
        ProgressSeminarComment.objects.update_or_create(
            seminar=seminar,
            member=faculty,
            defaults={'text': data['comment']}
        )

    consent_obj, _ = ProgressSeminarConsent.objects.get_or_create(seminar=seminar, member=faculty)
    consent_obj.consented = True
    consent_obj.save()

    return JsonResponse({"message": "Consent & data recorded."})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def rpc_finalize(request, pk):
    faculty = get_object_or_404(Faculty, id__user=request.user)
    seminar = get_object_or_404(ProgressSeminarEntry, pk=pk, status='rpc_pending')
    if not CommitteeMember.objects.filter(thesis=seminar.thesis, member=faculty).exists():
        return JsonResponse({"error": "Not on committee."}, status=403)

    current_committee_ids = CommitteeMember.objects.filter(thesis=seminar.thesis).values_list('member_id', flat=True)
    total = len(current_committee_ids)
    yes = ProgressSeminarConsent.objects.filter(
        seminar=seminar, consented=True, member_id__in=current_committee_ids,
    ).count()

    if total == 0 or yes < total:
        return JsonResponse({"error": "Not all consents recorded."}, status=400)

    seminar.status = 'rpc_approved'
    seminar.save()

    thesis = seminar.thesis
    for recipient in [thesis.student.id.user, thesis.supervisor.id.user]:
        _progress_seminar_notify(
            sender=request.user,
            recipient=recipient,
            verb='Progress Seminar report approved',
            description=f"The RPC has approved {thesis.student.id.user.get_full_name()}'s "
                        f"Progress Seminar report (version {seminar.version}).",
        )

    return JsonResponse({"message": "Seminar approved."})


from applications.academic_procedures.models import (
    ThesisSubmission, ReviewInvitation, ThesisReview, ExaminerBankDetails,
    ThesisRevisionRound, ThesisRevisionConsent,
)
from applications.academic_procedures.utils import (
    send_invitation_email,
    send_review_form_email,
    send_thank_you_email,
    advance_invitation,
    INVITATION_TIMEOUT_DAYS,
)


def _current_invitation(sub, examiner_type):
    """The examiner currently 'holding' this category -- the one who accepted
    (or is still pending a response) rather than one who rejected/expired and
    was passed over. There's only ever one non-finalized-out invitation per
    category once the panel has been sent out."""
    return ReviewInvitation.objects.filter(
        submission=sub, examiner_type=examiner_type
    ).exclude(status__in=['rejected', 'expired']).order_by('priority').first()

# 1. Student submits thesis
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def thesis_submit(request):
    user = request.user
    try:
        user_details = user.extrainfo
        student = Student.objects.get(id=user_details)
    except Student.DoesNotExist:
        return Response({'error': 'Student record not found.'}, 404)
    if _student_programme_category(student) != 'PHD':
        return Response(
            {'error': 'This final thesis submission workflow is for PhD students only.'},
            status=403,
        )
    thesis = ThesisTopic.objects.filter(student=student).order_by('-created_at').first()
    if thesis is None:
        return Response({'error': 'No thesis found for given submission.'}, 400)
    if not OpenSeminar.objects.filter(student=student, status='satisfactory').exists():
        return Response(
            {'error': 'Open Seminar must be completed satisfactorily before final thesis submission.'},
            status=403,
        )
    syn   = request.FILES.get('synopsis')
    rpt   = request.FILES.get('thesis_report')
    if not all([syn, rpt]):
        return Response({'error': 'Missing fields'}, 400)
    if syn.size > 5*1024*1024 or rpt.size > 25*1024*1024:
        return Response({'error': 'File too large'}, 400)
    if syn.content_type != 'application/pdf' or rpt.content_type != 'application/pdf':
        return Response({'error': 'Both files must be PDFs'}, status=400)
    if ThesisSubmission.objects.filter(thesis=thesis).exists():
        return Response({'error': 'Thesis has already been submitted and cannot be changed.'}, status=400)
    sub = ThesisSubmission.objects.create(
        thesis = thesis,
        synopsis=syn,
        thesis_report=rpt,
        status='submitted'
    )
    return Response({'submission_id': sub.id}, status=201)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def thesis_submission_status(request):
    """
    GET /thesis/submission-status/
    Returns the requesting student's own thesis submission (if any), so the
    upload screen can show existing status instead of a blank form. `thesis`
    is a OneToOneField on ThesisSubmission, so at most one can ever exist.
    """
    user = request.user
    try:
        student = Student.objects.get(id=user.extrainfo)
    except Student.DoesNotExist:
        return Response({'error': 'Student record not found.'}, status=404)

    thesis = ThesisTopic.objects.filter(student=student).order_by('-created_at').first()
    if thesis is None:
        return Response({'submission': None}, status=200)

    try:
        sub = thesis.submission
    except ThesisSubmission.DoesNotExist:
        return Response({'submission': None}, status=200)

    current_round = sub.revision_rounds.order_by('-round_number').first()
    return Response({
        'submission': {
            'id': sub.id,
            'status': sub.status,
            'status_label': sub.get_status_display(),
            'submitted_at': sub.submitted_at.isoformat(),
            'synopsis_url': sub.synopsis.url if sub.synopsis else None,
            'thesis_report_url': sub.thesis_report.url if sub.thesis_report else None,
            'dean_panel_remarks': sub.dean_panel_remarks,
            'director_remarks': sub.director_remarks,
            # Only meaningful (and only the student's turn) when status is
            # 'student_revision_pending' -- revised_thesis_url null means
            # they haven't uploaded yet.
            'pending_revision_round': (
                current_round.round_number
                if sub.status == 'student_revision_pending' and current_round else None
            ),
        },
    }, status=200)

def _serialize_invitations(sub):
    """Return (indian_examiners, foreign_examiners) lists for a submission's panel.

    Reads via the `invitations` related manager rather than a fresh filter()
    so that callers who prefetch_related('invitations', queryset=...ordered...)
    get it from the prefetch cache instead of a query per submission.
    """
    invites = sub.invitations.all()
    indian, foreign = [], []
    for inv in invites:
        data = {
            'token': str(inv.token),
            'name': inv.prof_name,
            'position': inv.prof_position,
            'address': inv.prof_address,
            'phone': inv.prof_phone,
            'fax': inv.prof_fax,
            'email': inv.prof_email,
            'priority': inv.priority,
            'status': inv.status,
        }
        if inv.examiner_type == 'foreign':
            data['time_ranking'] = inv.prof_time_ranking
            foreign.append(data)
        else:
            indian.append(data)
    return indian, foreign


# 1) Supervisor dashboard: pending vs forwarded
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def supervisor_dashboard(request):
    ex = request.user
    topics = ThesisTopic.objects.filter(
        Q(supervisor__id=ex.username) | Q(co_supervisor__id=ex.username)
    )

    def serialize(sub, action=None, action_label=None, waiting_since=None):
        return {
            'id': sub.id,
            'title': sub.thesis.research_theme,
            'student_name': sub.thesis.student.id.user.get_full_name(),
            'student_roll': sub.thesis.student.id.id,
            'status': sub.status,
            'action': action,
            'action_label': action_label,
            'waiting_since': waiting_since,
            'submitted_at': sub.submitted_at,
            'supervisor_approved_at': sub.supervisor_approved_at,
            'dean_panel_remarks': sub.dean_panel_remarks,
        }

    # status='submitted' covers two different situations: a brand new
    # submission the panel has never been assigned for, or one the Dean just
    # sent back with remarks after rejecting the proposed panel.
    action_required = []
    for sub in ThesisSubmission.objects.filter(
        status='submitted', thesis__in=topics
    ).select_related('thesis__student__id__user'):
        if sub.dean_panel_remarks:
            action, action_label = 'revise_panel', 'Revise Panel (Dean)'
        else:
            action, action_label = 'assign_examiners', 'Assign Examiners'
        action_required.append(serialize(sub, action, action_label, sub.updated_at))
    # Oldest-waiting first, so overdue items surface at the top.
    action_required.sort(key=lambda s: s['waiting_since'] or timezone.now())

    history = [
        serialize(s) for s in
        ThesisSubmission.objects.filter(thesis__in=topics)
        .exclude(status='submitted')
        .select_related('thesis__student__id__user')
    ]

    return Response({
        'action_required': action_required,
        'history': history,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def supervisor_submission_detail(request, submission_id):
    """
    Returns the already‐assigned examiners (and any Dean remarks from a
    prior rejection) so the panel can pre-fill both the read-only view and
    a resubmission after the Dean sends the panel back.
    """
    sub = get_object_or_404(ThesisSubmission, id=submission_id)
    indian, foreign = _serialize_invitations(sub)
    return Response({
        'indian_examiners': indian,
        'foreign_examiners': foreign,
        'dean_panel_remarks': sub.dean_panel_remarks,
    })


def _serialize_examiner_reports(sub):
    """Completed examiners' latest-round reports for a submission. Identity
    redaction is deliberately not applied here yet -- deferred until the
    exact rule is confirmed with officials."""
    completed = (
        ReviewInvitation.objects
        .filter(submission=sub, status='completed')
        .prefetch_related('reviews')
        .order_by('examiner_type', 'priority')
    )
    reviews = []
    for inv in completed:
        # Only the latest round's report matters here -- earlier rounds
        # are superseded once the examiner reconfirms on a revision.
        latest = inv.reviews.order_by('-round_number').first()
        if latest is None:
            continue
        reviews.append({
            'examiner_type': inv.examiner_type,
            'examiner_name': inv.prof_name,
            'examiner_email': inv.prof_email,
            'round_number': latest.round_number,
            'originality_presentation': latest.originality_presentation,
            'quality_comparable': latest.quality_comparable,
            'new_ideas_original': latest.new_ideas_original,
            'correction_severity': latest.correction_severity,
            'technical_content': latest.technical_content,
            'highlights': latest.highlights,
            'suggestions': latest.suggestions,
            'defense_questions': latest.defense_questions,
            'recommendation': latest.recommendation,
            'submitted_at': latest.submitted_at,
        })
    return reviews


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def supervisor_review_reports(request):
    """
    Submissions supervised by the caller that have at least one examiner's
    completed review, with the full report content for each.
    """
    ex = request.user
    topics = ThesisTopic.objects.filter(
        Q(supervisor__id=ex.username) | Q(co_supervisor__id=ex.username)
    )

    data = []
    for sub in ThesisSubmission.objects.filter(thesis__in=topics):
        reviews = _serialize_examiner_reports(sub)
        if reviews:
            data.append({
                'id': sub.id,
                'title': sub.thesis.research_theme,
                'student_name': sub.thesis.student.id.user.get_full_name(),
                'student_roll': sub.thesis.student.id.id,
                'reviews': reviews,
                'status': sub.status,
                'status_label': sub.get_status_display(),
                'reports_forwarded_at': sub.reports_forwarded_at,
                'current_round': _serialize_revision_round(sub, request),
            })

    return Response(data)


def _serialize_revision_round(sub, request):
    """Latest revision round + RPC consent progress, or None if the
    submission hasn't reached the revision stage."""
    round_obj = sub.revision_rounds.order_by('-round_number').first()
    if round_obj is None:
        return None

    base = request.build_absolute_uri(settings.MEDIA_URL)
    committee = CommitteeMember.objects.filter(thesis=sub.thesis).select_related('member__id__user')
    consents = {
        c.member_id: c
        for c in ThesisRevisionConsent.objects.filter(round=round_obj)
    }
    members = [
        {
            'member_name': cm.member.id.user.get_full_name(),
            'consented': consents[cm.member_id].consented if cm.member_id in consents else False,
            'remarks': consents[cm.member_id].remarks if cm.member_id in consents else '',
        }
        for cm in committee
    ]
    return {
        'round_number': round_obj.round_number,
        'revised_thesis_url': (base + round_obj.revised_thesis.name) if round_obj.revised_thesis else None,
        'revised_at': round_obj.revised_at,
        'supervisor_consented_at': round_obj.supervisor_consented_at,
        'committee': members,
        'committee_size': len(members),
        'consented_count': sum(1 for m in members if m['consented']),
    }


# 3) Supervisor assign examiners
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def supervisor_assign(request):
    data = request.data
    sub = get_object_or_404(ThesisSubmission, id=data.get('submission_id'))

    # Only the thesis's own supervisor/co-supervisor may assign its panel.
    topic = sub.thesis
    allowed_users = {topic.supervisor.id.user_id}
    if topic.co_supervisor:
        allowed_users.add(topic.co_supervisor.id.user_id)
    if request.user.id not in allowed_users:
        return Response(
            {'error': 'You are not the supervisor or co-supervisor for this thesis.'},
            status=status.HTTP_403_FORBIDDEN
        )

    # Prevent re-assignment
    if sub.status != 'submitted':
        return Response(
            {'error': 'Examiners have already been assigned.'},
            status=status.HTTP_400_BAD_REQUEST
        )

    indian = data.get('indian_examiners', [])
    foreign = data.get('foreign_examiners', [])

    if not indian or not foreign:
        return Response(
            {'error': 'At least one Indian and one foreign examiner are required.'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Save submission + invitations atomically. ReviewInvitation has a
    # unique constraint on (submission, examiner_type, priority); each
    # category gets its own independent 1..N rank.
    with transaction.atomic():
        # Wipe old invites & create new ones
        ReviewInvitation.objects.filter(submission=sub).delete()

        for idx, prof in enumerate(indian, start=1):
            ReviewInvitation.objects.create(
                submission=sub,
                examiner_type='indian',
                prof_name=prof.get('name', ''),
                prof_position=prof.get('position', ''),
                prof_address=prof.get('address', ''),
                prof_phone=prof.get('phone', ''),
                prof_fax=prof.get('fax', ''),
                prof_email=prof.get('email', ''),
                priority=idx,
            )

        for idx, prof in enumerate(foreign, start=1):
            ReviewInvitation.objects.create(
                submission=sub,
                examiner_type='foreign',
                prof_name=prof.get('name', ''),
                prof_position=prof.get('position', ''),
                prof_address=prof.get('address', ''),
                prof_phone=prof.get('phone', ''),
                prof_fax=prof.get('fax', ''),
                prof_email=prof.get('email', ''),
                prof_time_ranking=prof.get('time_ranking', 1),
                priority=idx,
            )

        # Update submission only after invitations are created successfully.
        # Clear any earlier Dean rejection remark -- this resubmission is the
        # response to it, so it shouldn't resurface as if still unaddressed.
        sub.supervisor = request.user
        sub.supervisor_approved_at = timezone.now()
        sub.status = 'dean_panel_review'
        sub.dean_panel_remarks = ''
        sub.save()

    return Response({'detail': 'Examiners assigned successfully, forwarded to Dean for approval.'}, status=status.HTTP_200_OK)


# 4) Dean panel dashboard: a single "action required" queue covering both
#    panel approval and invitation-sending, plus a read-only history.
ACTION_STATUSES = {
    'dean_panel_review':      ('approve_panel', 'Forward Panel'),
    'dean_invite_pending':    ('send_invitations', 'Send Invitations'),
    'examiner_reports_ready': ('forward_reports', 'Forward Reports to Supervisor'),
    'dean_final_review':      ('final_review', 'Final Review'),
}

# Friendly labels for the Dean's read-only History tab. 'submitted' only
# reaches history once it has been through dean_panel_review at least once
# (see the history query below), so here it always means "sent back".
STATUS_LABELS = {
    'submitted':                  'Sent Back to Supervisor',
    'director_review':            'With Director',
    'in_review':                  'In External Review',
    'supervisor_reports_review':  'With Supervisor',
    'student_revision_pending':   'Awaiting Student Revision',
    'supervisor_revision_review': 'Awaiting Supervisor & RPC Consent',
    'approved_for_defense':       'Approved for Defense',
}


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(['Dean Academic'])
def dean_panel_dashboard(request):
    def serialize(sub, action=None, action_label=None, waiting_since=None):
        indian, foreign = _serialize_invitations(sub)
        return {
            'id': sub.id,
            'title': sub.thesis.research_theme,
            'student_name': sub.thesis.student.id.user.get_full_name(),
            'student_roll': sub.thesis.student.id.id,
            'status': sub.status,
            'status_label': STATUS_LABELS.get(sub.status, sub.status),
            'action': action,
            'action_label': action_label,
            'waiting_since': waiting_since,
            'supervisor_approved_at': sub.supervisor_approved_at,
            'dean_approved_at': sub.dean_approved_at,
            'dean_panel_remarks': sub.dean_panel_remarks,
            'director_remarks': sub.director_remarks,
            'indian_examiners': indian,
            'foreign_examiners': foreign,
            'reviews': _serialize_examiner_reports(sub) if sub.status in (
                'examiner_reports_ready', 'dean_final_review', 'approved_for_defense',
            ) else [],
            'current_round': _serialize_revision_round(sub, request) if sub.status in (
                'student_revision_pending', 'supervisor_revision_review', 'dean_final_review',
                'approved_for_defense',
            ) else None,
        }

    action_required = []
    for sub in ThesisSubmission.objects.filter(
        status__in=ACTION_STATUSES.keys()
    ).select_related('thesis__student__id__user').prefetch_related('invitations'):
        # 'dean_panel_review' covers two different situations that both need
        # a Dean decision: a fresh panel from the Supervisor, or one the
        # Director just sent back with remarks. Tell them apart so the Dean
        # isn't stuck re-reading the panel to figure out which one it is.
        if sub.status == 'dean_panel_review' and sub.director_remarks:
            action, action_label = 'reconsider_panel', 'Reconsider Panel (Director)'
            waiting_since = sub.director_approved_at
        elif sub.status == 'dean_invite_pending':
            action, action_label = ACTION_STATUSES[sub.status]
            waiting_since = sub.director_approved_at
        elif sub.status == 'examiner_reports_ready':
            action, action_label = ACTION_STATUSES[sub.status]
            waiting_since = sub.examiner_reports_ready_at
        elif sub.status == 'dean_final_review':
            if _dissenting_invitations(sub):
                action, action_label = 'send_back_to_examiner', 'Send Back to Examiner(s)'
            else:
                action, action_label = 'approve_for_defense', 'Approve for Defense'
            waiting_since = sub.revision_consented_at or sub.final_review_requested_at
        else:
            action, action_label = ACTION_STATUSES[sub.status]
            waiting_since = sub.supervisor_approved_at
        action_required.append(serialize(sub, action, action_label, waiting_since))
    # Oldest-waiting first, so overdue items surface at the top.
    action_required.sort(key=lambda s: s['waiting_since'] or timezone.now())

    # 'submitted' normally means "not yet assigned by the supervisor" and
    # doesn't belong in the Dean's history — except when it got there *after*
    # a panel review (supervisor_approved_at is set), i.e. the Dean sent it
    # back. That case should still show up, with a clear status label.
    history = [
        serialize(s) for s in
        ThesisSubmission.objects.exclude(
            status__in=ACTION_STATUSES.keys()
        ).exclude(
            Q(status='submitted') & Q(supervisor_approved_at__isnull=True)
        ).select_related('thesis__student__id__user').prefetch_related('invitations')
    ]

    return Response({
        'action_required': action_required,
        'history': history,
    })


# 5) Dean approves or rejects the proposed panel
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['Dean Academic'])
def dean_panel_approve(request):
    data = request.data
    sub = get_object_or_404(ThesisSubmission, id=data.get('submission_id'))
    action = data.get('action')

    if sub.status != 'dean_panel_review':
        return Response({'error': 'This panel is not awaiting Dean approval.'}, status=status.HTTP_400_BAD_REQUEST)

    if action == 'approve':
        sub.dean = request.user
        sub.dean_approved_at = timezone.now()
        sub.status = 'director_review'
        sub.save()
        return Response({'detail': 'Panel approved, forwarded to Director for prioritization.'})

    if action == 'reject':
        remarks = (data.get('remarks') or '').strip()
        if not remarks:
            return Response(
                {'error': 'A remark is required when sending the panel back to the Supervisor.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        sub.status = 'submitted'
        # Supervisor's dashboard tells "fresh submission" apart from "sent
        # back by Dean" by checking whether dean_panel_remarks is non-empty,
        # so this must always be non-empty for a rejection to be recognized.
        sub.dean_panel_remarks = remarks
        # Starting a fresh Supervisor cycle -- any earlier Director remark no
        # longer applies and would otherwise look like a stale "sent back by
        # Director" marker on the resubmitted panel.
        sub.director_remarks = ''
        sub.save()
        return Response({'detail': 'Panel rejected, sent back to Supervisor.'})

    return Response({'error': 'Unknown action.'}, status=status.HTTP_400_BAD_REQUEST)


# 6) Director dashboard: pending prioritization vs already prioritized
@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(['Director'])
def director_dashboard(request):
    def serialize(sub, action=None, action_label=None, waiting_since=None):
        indian, foreign = _serialize_invitations(sub)
        return {
            'id': sub.id,
            'title': sub.thesis.research_theme,
            'student_name': sub.thesis.student.id.user.get_full_name(),
            'student_roll': sub.thesis.student.id.id,
            'status': sub.status,
            'action': action,
            'action_label': action_label,
            'waiting_since': waiting_since,
            'supervisor_approved_at': sub.supervisor_approved_at,
            'director_approved_at': sub.director_approved_at,
            'indian_examiners': indian,
            'foreign_examiners': foreign,
        }

    action_required = [
        serialize(sub, 'prioritize', 'Set Priorities', sub.dean_approved_at)
        for sub in ThesisSubmission.objects.filter(
            status='director_review'
        ).select_related('thesis__student__id__user').prefetch_related('invitations')
    ]
    # Oldest-waiting first, so overdue items surface at the top.
    action_required.sort(key=lambda s: s['waiting_since'] or timezone.now())

    history = [
        serialize(s) for s in
        ThesisSubmission.objects.exclude(
            status__in=['submitted', 'dean_panel_review', 'director_review']
        ).select_related('thesis__student__id__user').prefetch_related('invitations')
    ]

    return Response({
        'action_required': action_required,
        'history': history,
    })


# 7) Director sets the priority order within each examiner category, then
#    hands the submission back to the Dean to send out invitations.
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['Director'])
def director_approve(request):
    data = request.data
    sub = get_object_or_404(ThesisSubmission, id=data.get('submission_id'))

    if sub.status != 'director_review':
        return Response(
            {'error': 'This panel is not awaiting Director prioritization.'},
            status=status.HTTP_400_BAD_REQUEST
        )

    action = data.get('action', 'approve')
    if action not in ('approve', 'send_back'):
        return Response({'error': 'Unknown action.'}, status=status.HTTP_400_BAD_REQUEST)

    indian = data.get('indian_examiners', [])
    foreign = data.get('foreign_examiners', [])
    if not indian or not foreign:
        return Response(
            {'error': 'At least one Indian and one foreign examiner are required.'},
            status=status.HTTP_400_BAD_REQUEST
        )

    remarks = (data.get('remarks') or '').strip()
    if action == 'send_back' and not remarks:
        return Response(
            {'error': 'A remark is required when sending the panel back to the Dean.'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # The Director can add, remove, or edit examiners (not just re-rank the
    # ones the Supervisor originally nominated), so the panel is rebuilt from
    # the submitted lists the same way supervisor_assign builds it initially.
    # Rank is simply the row's position within its category.
    with transaction.atomic():
        ReviewInvitation.objects.filter(submission=sub).delete()

        for idx, prof in enumerate(indian, start=1):
            ReviewInvitation.objects.create(
                submission=sub,
                examiner_type='indian',
                prof_name=prof.get('name', ''),
                prof_position=prof.get('position', ''),
                prof_address=prof.get('address', ''),
                prof_phone=prof.get('phone', ''),
                prof_fax=prof.get('fax', ''),
                prof_email=prof.get('email', ''),
                priority=idx,
            )

        for idx, prof in enumerate(foreign, start=1):
            ReviewInvitation.objects.create(
                submission=sub,
                examiner_type='foreign',
                prof_name=prof.get('name', ''),
                prof_position=prof.get('position', ''),
                prof_address=prof.get('address', ''),
                prof_phone=prof.get('phone', ''),
                prof_fax=prof.get('fax', ''),
                prof_email=prof.get('email', ''),
                prof_time_ranking=prof.get('time_ranking', 1),
                priority=idx,
            )

        sub.director = request.user
        sub.director_approved_at = timezone.now()
        sub.director_remarks = remarks

        if action == 'send_back':
            sub.status = 'dean_panel_review'
            detail = 'Panel sent back to the Dean with your remarks.'
        else:
            sub.status = 'dean_invite_pending'
            detail = 'Priorities approved, sent to Dean to send invitations.'
        sub.save()

    return Response({'detail': detail})


# 7b) Dean sends the invitation to the Rank-1 Indian and Rank-1 Foreign examiners.
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['Dean Academic'])
def dean_send_invitations(request):
    data = request.data
    sub = get_object_or_404(ThesisSubmission, id=data.get('submission_id'))

    if sub.status != 'dean_invite_pending':
        return Response(
            {'error': 'This submission is not ready for invitations.'},
            status=status.HTTP_400_BAD_REQUEST
        )

    invited = []
    for examiner_type in ('indian', 'foreign'):
        inv = ReviewInvitation.objects.filter(
            submission=sub, examiner_type=examiner_type, priority=1
        ).first()
        if inv is None:
            continue
        inv.last_sent = timezone.now()
        inv.expires_at = timezone.now() + datetime.timedelta(days=INVITATION_TIMEOUT_DAYS)
        inv.save(update_fields=['last_sent', 'expires_at'])
        try:
            send_invitation_email(inv)
            invited.append(inv.prof_email)
        except Exception:
            logger.exception(f"Failed to send invitation to {inv.prof_email} for submission {sub.id}")

    sub.dean = request.user
    sub.dean_invited_at = timezone.now()
    sub.status = 'in_review'
    sub.save()

    return Response({'detail': 'Invitations sent.', 'invited': invited})


# 8. Invitation accept/reject (external examiners have no Fusion account —
#    the secret UUID token in the emailed link is the auth mechanism here)
@api_view(['POST'])
@permission_classes([AllowAny])
def invitation_action(request, token, action):
    inv = get_object_or_404(ReviewInvitation, token=token)
    if inv.is_expired() or inv.is_finalized():
        return Response({'error': 'Invalid/expired'}, 403)
    if action == 'accept':
        inv.status = 'accepted'
        inv.save()
        try:
            send_review_form_email(inv)
            inv.review_form_sent = timezone.now()
            inv.save(update_fields=['review_form_sent'])
        except Exception:
            logger.exception(f"Failed to send review-form email for token {inv.token}")
        return Response({'detail': 'Accepted'}, 200)
    if action == 'reject':
        inv.status = 'rejected'
        inv.save()
        # Fall through to the next-ranked examiner in the same category.
        advance_invitation(inv.submission, inv.examiner_type)
        return Response({'detail': 'Rejected'}, 200)
    return Response({'error': 'Unknown action'}, 400)

# 9. Review detail & submission (token-authenticated, same as invitation_action)
@api_view(['GET', 'POST'])
@permission_classes([AllowAny])
def review_detail(request, token):
    inv = get_object_or_404(ReviewInvitation, token=token)
    if inv.is_expired() or inv.is_finalized():
        return Response({'error': 'Invalid/expired'}, status=403)
    if inv.status != 'accepted':
        return Response({'error': 'This invitation has not been accepted yet.'}, status=403)

    sub = inv.submission
    topic = sub.thesis

    if request.method == 'GET':
        base = request.build_absolute_uri(settings.MEDIA_URL)
        # Reconfirmation rounds (revision_round > 1) review the student's
        # revised thesis from the matching round, not the original upload --
        # round N here pairs with ThesisRevisionRound N-1's resubmission.
        report_file = sub.thesis_report
        if inv.revision_round > 1:
            prior_round = sub.revision_rounds.filter(round_number=inv.revision_round - 1).first()
            if prior_round and prior_round.revised_thesis:
                report_file = prior_round.revised_thesis
        return Response({
            'student_name': topic.student.id.user.get_full_name(),
            'student_roll': topic.student.id.id,
            'student_discipline': topic.student.specialization,
            'thesis_title': topic.research_theme,
            'synopsis_url': base + sub.synopsis.name,
            'report_url': base + report_file.name,
            'revision_round': inv.revision_round,
            'examiner_type': inv.examiner_type,
            'examiner': {
                'name': inv.prof_name,
                'email': inv.prof_email,
                'position': inv.prof_position,
                'address': inv.prof_address,
                'phone': inv.prof_phone,
                'fax': inv.prof_fax,
            },
        }, status=200)

    # POST: record the formal evaluation and finalize this examiner's invitation
    # (for this revision round -- see ReviewInvitation.revision_round). Once
    # BOTH categories are sitting at 'completed' for their current round, the
    # submission moves to examiner_reports_ready for the Dean to forward on.
    data = request.data
    if not data.get('recommendation'):
        return Response({'error': 'A specific recommendation is required.'}, status=400)

    with transaction.atomic():
        ThesisReview.objects.update_or_create(
            invitation=inv,
            round_number=inv.revision_round,
            defaults={
                'originality_presentation': data.get('originality_presentation', ''),
                'quality_comparable': data.get('quality_comparable'),
                'new_ideas_original': data.get('new_ideas_original'),
                'correction_severity': data.get('correction_severity', ''),
                'technical_content': data.get('technical_content', ''),
                'highlights': data.get('highlights', ''),
                'suggestions': data.get('suggestions', ''),
                'defense_questions': data.get('defense_questions', ''),
                'recommendation': data['recommendation'],
            },
        )

        bank = data.get('bank_details') or {}
        if any(bank.values()):
            ExaminerBankDetails.objects.update_or_create(
                invitation=inv,
                defaults={
                    'beneficiary_name': bank.get('beneficiary_name', ''),
                    'bank_name': bank.get('bank_name', ''),
                    'bank_address': bank.get('bank_address', ''),
                    'account_no': bank.get('account_no', ''),
                    'ifsc_code': bank.get('ifsc_code', ''),
                    'pan_no': bank.get('pan_no', ''),
                    'iban_no': bank.get('iban_no', ''),
                    'swift_code': bank.get('swift_code', ''),
                },
            )

        inv.status = 'completed'
        inv.save(update_fields=['status'])

        indian_inv = _current_invitation(sub, 'indian')
        foreign_inv = _current_invitation(sub, 'foreign')
        if (indian_inv and indian_inv.status == 'completed' and
                foreign_inv and foreign_inv.status == 'completed'):
            sub.status = 'examiner_reports_ready'
            sub.examiner_reports_ready_at = timezone.now()
            sub.save(update_fields=['status', 'examiner_reports_ready_at'])

    try:
        send_thank_you_email(inv)
    except Exception:
        logger.exception(f"Failed to send thank-you email for token {inv.token}")

    return Response({'detail': 'Review submitted successfully.'}, status=200)


# 10. Acadadmin: bank details for examiners who have completed a review, so
#     the honorarium can be processed. Independent of the thesis outcome --
#     surfaces as soon as each individual examiner finishes, regardless of
#     what happens next in the review-consolidation workflow.
@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(['acadadmin'])
def examiner_honorarium_list(request):
    invs = (
        ReviewInvitation.objects
        .filter(status='completed')
        .select_related('bank_details', 'submission__thesis__student__id__user')
        .order_by('-updated_at')
    )

    data = []
    for inv in invs:
        if not hasattr(inv, 'bank_details'):
            continue
        bank = inv.bank_details
        sub = inv.submission
        data.append({
            'invitation_id': inv.id,
            'examiner_name': inv.prof_name,
            'examiner_email': inv.prof_email,
            'examiner_type': inv.examiner_type,
            'thesis_title': sub.thesis.research_theme,
            'student_name': sub.thesis.student.id.user.get_full_name(),
            'student_roll': sub.thesis.student.id.id,
            'beneficiary_name': bank.beneficiary_name,
            'bank_name': bank.bank_name,
            'bank_address': bank.bank_address,
            'account_no': bank.account_no,
            'ifsc_code': bank.ifsc_code,
            'pan_no': bank.pan_no,
            'iban_no': bank.iban_no,
            'swift_code': bank.swift_code,
        })

    return Response(data)


# ===========================================================================
# Post-evaluation workflow: both examiner reports received -> Dean forwards
# (identity redaction deliberately deferred, per instruction, until officials
# confirm the exact rule) -> Supervisor decides revision or not -> Student
# revises -> Supervisor + RPC consent -> Dean's final call, which either
# approves for defense or sends the revision back to whichever examiner(s)
# didn't give a clean accept, unlimited rounds, via ReviewInvitation.revision_round.
# ===========================================================================

# 11) Dean forwards both examiner reports to the Supervisor.
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['Dean Academic'])
def dean_forward_reports(request):
    sub = get_object_or_404(ThesisSubmission, id=request.data.get('submission_id'))
    if sub.status != 'examiner_reports_ready':
        return Response({'error': 'Examiner reports are not ready to be forwarded.'}, status=400)

    sub.status = 'supervisor_reports_review'
    sub.reports_forwarded_at = timezone.now()
    sub.save(update_fields=['status', 'reports_forwarded_at'])
    return Response({'detail': 'Reports forwarded to the Supervisor.'})


def _require_supervisor(request, sub):
    topic = sub.thesis
    allowed_users = {topic.supervisor.id.user_id}
    if topic.co_supervisor:
        allowed_users.add(topic.co_supervisor.id.user_id)
    return request.user.id in allowed_users


# 12) Supervisor decides whether the reports call for a revision, or approves
#     straight through to the Dean's final review.
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def supervisor_reports_decision(request):
    sub = get_object_or_404(ThesisSubmission, id=request.data.get('submission_id'))
    if not _require_supervisor(request, sub):
        return Response({'error': 'You are not the supervisor or co-supervisor for this thesis.'}, status=403)
    if sub.status != 'supervisor_reports_review':
        return Response({'error': 'This submission is not awaiting your decision.'}, status=400)

    action = request.data.get('action')
    if action not in ('forward_to_student', 'no_revision_needed'):
        return Response({'error': 'Unknown action.'}, status=400)

    sub.revision_requested_at = timezone.now()
    if action == 'forward_to_student':
        next_round = (sub.revision_rounds.aggregate(m=Max('round_number'))['m'] or 0) + 1
        ThesisRevisionRound.objects.create(submission=sub, round_number=next_round)
        sub.status = 'student_revision_pending'
        detail = 'Forwarded to the student for revision.'
    else:
        sub.status = 'dean_final_review'
        sub.final_review_requested_at = timezone.now()
        detail = 'No revision needed -- forwarded to the Dean for final approval.'
    sub.save()

    return Response({'detail': detail})


# 13) Student uploads the revised thesis for the current round.
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def student_submit_revision(request):
    sub = get_object_or_404(
        ThesisSubmission, id=request.data.get('submission_id'),
        thesis__student__id__user=request.user,
    )
    if sub.status != 'student_revision_pending':
        return Response({'error': 'No revision is currently pending for this submission.'}, status=400)

    round_obj = sub.revision_rounds.order_by('-round_number').first()
    if round_obj is None or round_obj.revised_thesis:
        return Response({'error': 'No pending revision round found.'}, status=400)

    file = request.FILES.get('revised_thesis')
    if not file:
        return Response({'error': 'A revised thesis file is required.'}, status=400)
    if file.content_type != 'application/pdf':
        return Response({'error': 'The revised thesis must be a PDF.'}, status=400)
    if file.size > 25 * 1024 * 1024:
        return Response({'error': 'File too large.'}, status=400)

    round_obj.revised_thesis = file
    round_obj.revised_at = timezone.now()
    round_obj.save(update_fields=['revised_thesis', 'revised_at'])

    sub.status = 'supervisor_revision_review'
    sub.revision_submitted_at = timezone.now()
    sub.save(update_fields=['status', 'revision_submitted_at'])

    return Response({'detail': 'Revised thesis submitted.'})


# 14) RPC: list revision rounds awaiting the calling faculty member's consent.
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def thesis_revision_rpc_list(request):
    faculty = get_object_or_404(Faculty, id__user=request.user)
    rounds = (
        ThesisRevisionRound.objects
        .filter(submission__thesis__committee__member=faculty, submission__status='supervisor_revision_review')
        .distinct()
        .select_related('submission__thesis__student__id__user')
    )

    data = []
    for r in rounds:
        sub = r.submission
        base = request.build_absolute_uri(settings.MEDIA_URL)
        data.append({
            'round_id': r.id,
            'submission_id': sub.id,
            'round_number': r.round_number,
            'student_name': sub.thesis.student.id.user.get_full_name(),
            'thesis_title': sub.thesis.research_theme,
            'revised_thesis_url': (base + r.revised_thesis.name) if r.revised_thesis else None,
            'revised_at': r.revised_at,
            'my_consent_given': ThesisRevisionConsent.objects.filter(
                round=r, member=faculty, consented=True
            ).exists(),
        })
    return Response(data)


# 15) RPC: record this faculty member's consent on a revision round.
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def thesis_revision_rpc_consent(request):
    faculty = get_object_or_404(Faculty, id__user=request.user)
    round_obj = get_object_or_404(ThesisRevisionRound, id=request.data.get('round_id'))
    if not CommitteeMember.objects.filter(thesis=round_obj.submission.thesis, member=faculty).exists():
        return Response({'error': 'You are not on this thesis\'s RPC.'}, status=403)
    if round_obj.submission.status != 'supervisor_revision_review':
        return Response({'error': 'This round is not awaiting RPC consent.'}, status=400)

    ThesisRevisionConsent.objects.update_or_create(
        round=round_obj, member=faculty,
        defaults={'consented': True, 'remarks': request.data.get('remarks', '')},
    )
    return Response({'detail': 'Consent recorded.'})


# 16) Supervisor forwards the RPC-consented revision to the Dean, once every
#     current committee member has consented.
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def supervisor_finalize_revision(request):
    sub = get_object_or_404(ThesisSubmission, id=request.data.get('submission_id'))
    if not _require_supervisor(request, sub):
        return Response({'error': 'You are not the supervisor or co-supervisor for this thesis.'}, status=403)
    if sub.status != 'supervisor_revision_review':
        return Response({'error': 'This submission is not awaiting revision consent.'}, status=400)

    round_obj = sub.revision_rounds.order_by('-round_number').first()
    committee_ids = list(CommitteeMember.objects.filter(thesis=sub.thesis).values_list('member_id', flat=True))
    consented = ThesisRevisionConsent.objects.filter(
        round=round_obj, consented=True, member_id__in=committee_ids
    ).count()
    if not committee_ids or consented < len(committee_ids):
        return Response({'error': 'Not all RPC members have consented yet.'}, status=400)

    round_obj.supervisor_consented_at = timezone.now()
    round_obj.save(update_fields=['supervisor_consented_at'])

    sub.status = 'dean_final_review'
    sub.revision_consented_at = timezone.now()
    sub.final_review_requested_at = timezone.now()
    sub.save(update_fields=['status', 'revision_consented_at', 'final_review_requested_at'])

    return Response({'detail': 'Forwarded to the Dean for final approval.'})


def _dissenting_invitations(sub):
    """Examiner categories whose latest review wasn't a clean accept -- these
    are the ones the Dean sends the revision back to for reconfirmation."""
    dissenting = []
    for examiner_type in ('indian', 'foreign'):
        inv = _current_invitation(sub, examiner_type)
        if inv is None:
            continue
        latest = inv.reviews.order_by('-round_number').first()
        if latest and latest.recommendation != 'accept':
            dissenting.append(inv)
    return dissenting


# 17) Dean's final call: approve for defense if every examiner is now a clean
#     accept, otherwise send the revision back to whoever still isn't (unlimited
#     rounds -- ReviewInvitation.revision_round just keeps incrementing).
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['Dean Academic'])
def dean_final_review_action(request):
    sub = get_object_or_404(ThesisSubmission, id=request.data.get('submission_id'))
    if sub.status != 'dean_final_review':
        return Response({'error': 'This submission is not awaiting your final review.'}, status=400)

    dissenting = _dissenting_invitations(sub)

    with transaction.atomic():
        if dissenting:
            for inv in dissenting:
                inv.revision_round += 1
                inv.status = 'accepted'
                inv.save(update_fields=['revision_round', 'status'])
                try:
                    send_review_form_email(inv)
                    inv.review_form_sent = timezone.now()
                    inv.save(update_fields=['review_form_sent'])
                except Exception:
                    logger.exception(f"Failed to re-send review-form email for token {inv.token}")
            sub.status = 'in_review'
            sub.save(update_fields=['status'])
            detail = 'Revised thesis sent back to the examiner(s) for reconfirmation.'
        else:
            sub.status = 'approved_for_defense'
            sub.defense_approved_at = timezone.now()
            sub.dean = request.user
            sub.save(update_fields=['status', 'defense_approved_at', 'dean'])
            detail = 'Approved for defense.'

    return Response({'detail': detail})


# ===========================================================================
# Thesis Slot Semester-Level Registration
# ===========================================================================
from applications.academic_procedures.models import (
    ThesisTopic, CommitteeMember, ProgressSeminarEntry,
    ProgressSeminarConsent, ProgressSeminarComment,
    ThesisRegistration, ProgressSeminarRegistration, TeachingCreditRegistration,
    ThesisEvaluation, ProgressSeminarEvaluation,
    ThesisExaminerPanel, ThesisExaminerCandidate, ThesisEvaluationScore, PGThesisSubmission,
    ComprehensiveExam, ComprehensiveExamAttempt,
    ComprehensiveExamConsent, ComprehensiveExamRPCComment,
    OpenSeminar, OpenSeminarAttempt,
    OpenSeminarConsent, OpenSeminarRPCComment,
    TeachingCreditAllocation, TeachingCreditEvaluationResponse,
    resolve_progress_seminar_credit, resolve_teaching_credit_credit,
)
from applications.programme_curriculum.models import (
    ThesisSlot, SeminarSlot as ProgressSeminarSlot, TeachingCreditSlot,
)
import datetime as _dt


def _resolve_discipline_matched_entry(manager, student):
    """Pick the catalog entry (thesis/seminar/teaching-credit) matching the
    student's own discipline when a slot links entries from more than one
    discipline's catalog rows, falling back to the first entry otherwise.
    Mirrors resolve_progress_seminar_catalog_entry's discipline-preference rule
    -- a slot is allowed to serve multiple disciplines with different
    code/name/credit per discipline, so callers must not just take "the first
    linked entry" as if a slot only ever served one."""
    discipline = getattr(getattr(student, 'batch_id', None), 'discipline', None)
    return (manager.filter(discipline=discipline).first() if discipline else None) or manager.first()


def _catalog_entry_to_dict(entry):
    return {'id': entry.id, 'code': entry.code, 'name': entry.name, 'credit': entry.credit} if entry else None


def _student_programme_category(student):
    """'PG', 'PHD', or None -- used to let admin screens that merge PG/PhD
    requests together (semester numbering overlaps between the two) filter
    by category."""
    try:
        return student.batch_id.curriculum.programme.category
    except AttributeError:
        return None


def _thesis_reg_to_dict(reg):
    """Serialize a ThesisRegistration instance to a plain dict."""
    if reg is None:
        return None
    slot = reg.thesis_slot
    theses_list = [
        {'id': t.id, 'code': t.code, 'name': t.name, 'credit': t.credit}
        for t in slot.theses.all()
    ]
    resolved_thesis = _catalog_entry_to_dict(_resolve_discipline_matched_entry(slot.theses, reg.student))
    return {
        'id': reg.id,
        'status': reg.status,
        'remarks': reg.remarks,
        'credits': reg.credits,
        'registered_on': reg.registered_on.isoformat(),
        'verified_on': reg.verified_on.isoformat() if reg.verified_on else None,
        'academic_session': reg.academic_session,
        'thesis_slot': {
            'id': slot.id,
            'name': slot.name,
            'info': slot.thesis_slot_info or '',
            'duration': slot.duration,
            'theses': theses_list,
            'resolved_thesis': resolved_thesis,
        },
        'student': {
            'id': reg.student.id.id,
            'name': reg.student.id.user.get_full_name(),
        },
        'semester_no': reg.semester.semester_no,
        'programme_category': _student_programme_category(reg.student),
    }


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def student_thesis_enrollment_api(request):
    """
    GET  /stu/thesis-enrollment/
         Returns the current semester's ThesisSlot, the student's
         ThesisTopic approval status, and any existing registration.

    POST /stu/thesis-enrollment/
         Creates a new ThesisRegistration for the current semester.
         Requires thesis_topic to be dean_approved.
    """
    user = request.user
    try:
        user_details = user.extrainfo
        student = Student.objects.get(id=user_details)
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student record not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'User setup error: {type(e).__name__}: {e}'}, status=400)

    if request.method == 'POST':
        _elig = get_pg_phd_registration_eligibility(
            timezone.now().date(), _student_programme_category(student),
            student.curr_semester_no, datetime.datetime.now().year)
        if isinstance(_elig, JsonResponse):
            return _elig

    try:
        # Resolve current semester
        if not student.batch_id or not student.batch_id.curriculum:
            return JsonResponse({'error': 'Student batch or curriculum is not configured'}, status=400)
        try:
            semester = Semester.objects.get(
                curriculum=student.batch_id.curriculum,
                semester_no=student.curr_semester_no,
            )
        except Semester.DoesNotExist:
            return JsonResponse({'error': 'Current semester not found in curriculum'}, status=400)

        # Thesis topic info
        topic = ThesisTopic.objects.filter(student=student).order_by('-created_at').first()
        topic_data = thesis_to_dict(topic) if topic else None

        # ThesisSlot for this semester
        thesis_slots = ThesisSlot.objects.filter(semester=semester)
        thesis_slot = thesis_slots.first()  # typically one per semester

        # Existing registration
        try:
            reg = ThesisRegistration.objects.get(student=student, semester=semester)
            reg_data = _thesis_reg_to_dict(reg)
        except ThesisRegistration.DoesNotExist:
            reg = None
            reg_data = None

        if request.method == 'GET':
            slot_data = None
            if thesis_slot:
                slot_data = {
                    'id': thesis_slot.id,
                    'name': thesis_slot.name,
                    'info': thesis_slot.thesis_slot_info or '',
                    'duration': thesis_slot.duration,
                    'evaluation_type': thesis_slot.evaluation_type,
                    'theses': [
                        {'id': t.id, 'code': t.code, 'name': t.name, 'credit': t.credit}
                        for t in thesis_slot.theses.all()
                    ],
                    'resolved_thesis': _catalog_entry_to_dict(
                        _resolve_discipline_matched_entry(thesis_slot.theses, student)
                    ),
                }
            # Include announced evaluation blocks so student can see grades
            eval_blocks = []
            if reg is not None:
                for ev in reg.evaluations.filter(announced=True).order_by('block_number'):
                    eval_blocks.append({
                        'id':           ev.id,
                        'block_number': ev.block_number,
                        'total_blocks': reg.credits // 3,
                        'grade':        ev.grade,
                        'remarks':      ev.remarks,
                        'announced_at': ev.announced_at.isoformat() if ev.announced_at else None,
                    })
            return JsonResponse({
                'thesis_topic': topic_data,
                'current_semester_no': student.curr_semester_no,
                'thesis_slot': slot_data,
                'registration': reg_data,
                'evaluations': eval_blocks,
                'programme_category': _student_programme_category(student),
            }, status=200)

    except Exception as e:
        return JsonResponse({'error': f'Internal error: {type(e).__name__}: {e}'}, status=500)

    # POST: create registration
    if reg is not None:
        return JsonResponse(
            {'error': 'Already registered for this semester', 'registration': reg_data},
            status=400,
        )

    if topic is None or topic.status != 'dean_approved':
        return JsonResponse(
            {'error': 'Thesis topic must be dean-approved before registering for a thesis slot'},
            status=403,
        )

    if thesis_slot is None:
        return JsonResponse(
            {'error': 'No thesis slot configured for your current semester'},
            status=400,
        )

    # PG students have a fixed credit value per evaluation_type -- no free
    # choice: 3 for a block-graded (S/X) semester, 12 for the decimal-graded
    # semester. The 12 is fixed regardless of any earlier block-graded
    # semesters; those are additional thesis credit, not a substitute for
    # any part of it (PG's total can range from 12 up to 18).
    if _student_programme_category(student) == 'PG':
        chosen_credits = 3 if thesis_slot.evaluation_type == 'blocks_sx' else 12
    else:
        ALLOWED_THESIS_CREDITS = [3, 6, 9, 12]
        try:
            chosen_credits = int(request.data.get('credits', 6))
        except (TypeError, ValueError):
            chosen_credits = 6
        if chosen_credits not in ALLOWED_THESIS_CREDITS:
            return JsonResponse(
                {'error': f'Invalid credit value. Choose from {ALLOWED_THESIS_CREDITS}'},
                status=400,
            )

    # Check max registration limit
    current_count = ThesisRegistration.objects.filter(
        thesis_slot=thesis_slot, status__in=['pending', 'verified']
    ).count()
    if current_count >= thesis_slot.max_registration_limit:
        return JsonResponse(
            {'error': 'Thesis slot has reached maximum capacity'},
            status=400,
        )

    now = _dt.datetime.now()
    # Build academic session string e.g. "2025-26"
    year = now.year
    month = now.month
    if month >= 7:
        session = f"{year}-{str(year + 1)[2:]}"
    else:
        session = f"{year - 1}-{str(year)[2:]}"

    reg = ThesisRegistration.objects.create(
        student=student,
        thesis_slot=thesis_slot,
        thesis_topic=topic,
        semester=semester,
        credits=chosen_credits,
        working_year=year,
        academic_session=session,
        status='pending',
    )
    return JsonResponse(_thesis_reg_to_dict(reg), status=201)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def admin_thesis_enrollment_list(request):
    """
    GET /acadadmin/thesis-enrollments/?semester=<no>&status=<status>
    Lists all ThesisRegistration entries.  Supports optional filters:
      ?semester=<semester_no>   filter by semester number
      ?status=pending|verified|rejected
    """
    qs = ThesisRegistration.objects.select_related(
        'student__id__user', 'thesis_slot', 'thesis_topic', 'semester'
    ).all().order_by('-registered_on')

    sem_no = request.GET.get('semester')
    if sem_no:
        qs = qs.filter(semester__semester_no=sem_no)

    status_filter = request.GET.get('status')
    if status_filter:
        qs = qs.filter(status__iexact=status_filter)

    data = []
    for reg in qs:
        entry = _thesis_reg_to_dict(reg)
        # Also include thesis topic approval status for admin view
        entry['topic_status'] = reg.thesis_topic.status if reg.thesis_topic else None
        data.append(entry)

    return JsonResponse({'registrations': data}, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['acadadmin'])
def admin_verify_enrollments(request):
    """
    POST /acadadmin/thesis-enrollments/verify/
    Body: { "ids": [1, 2, 3] }
    Marks the given ThesisRegistration records as 'verified'.
    """
    ids = request.data.get('ids', [])
    if not ids:
        return JsonResponse({'error': 'No registration IDs provided'}, status=400)

    now = _dt.datetime.now(_dt.timezone.utc)
    regs = ThesisRegistration.objects.filter(id__in=ids, status='pending').select_related('thesis_slot')
    count = 0
    with transaction.atomic():
        for reg in regs:
            reg.status = 'verified'
            reg.verified_on = now
            reg.save(update_fields=['status', 'verified_on'])
            if reg.thesis_slot.evaluation_type == 'decimal':
                # Single overall score (PG's final thesis semester) -- no block
                # split, average of supervisor_score/examiner_score lands in
                # numeric_grade once ThesisEvaluationScore has both.
                evaluation, _created = ThesisEvaluation.objects.get_or_create(
                    registration=reg,
                    block_number=1,
                )
                ThesisEvaluationScore.objects.get_or_create(evaluation=evaluation)
            else:
                # Block-wise S/X (PhD, or PG sem 2/3): one block per 3 credits.
                total_blocks = reg.credits // 3
                for blk in range(1, total_blocks + 1):
                    ThesisEvaluation.objects.get_or_create(
                        registration=reg,
                        block_number=blk,
                    )
            count += 1
    return JsonResponse({'verified_count': count}, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['acadadmin'])
def admin_reject_enrollments(request):
    """
    POST /acadadmin/thesis-enrollments/reject/
    Body: { "ids": [1, 2], "remarks": "Reason for rejection" }
    Marks the given ThesisRegistration records as 'rejected'.
    """
    ids = request.data.get('ids', [])
    remarks = request.data.get('remarks', '')
    if not ids:
        return JsonResponse({'error': 'No registration IDs provided'}, status=400)

    updated = ThesisRegistration.objects.filter(id__in=ids, status='pending').update(
        status='rejected',
        remarks=remarks,
    )
    return JsonResponse({'rejected_count': updated}, status=200)


# ===========================================================================
# PG Decimal Thesis Grading -- Supervisor Score + Batch-Wide Examiner Panel
#
# Applies only to ThesisRegistrations whose thesis_slot.evaluation_type is
# 'decimal' (PG's final thesis semester). Flow: supervisor scores each
# student out of 100 (this is the "forward to HOD" step) -> once every
# student in the batch has a supervisor score, HOD nominates 4 Indian
# examiner candidates for the WHOLE BATCH -> Dean ranks them and invites the
# top candidate -> whoever accepts first examines every student in the
# batch -> each student's numeric_grade = round((supervisor+examiner)/2, 1).
# ===========================================================================

from applications.academic_procedures.utils import (
    send_examiner_panel_invitation_email,
    send_examiner_panel_scoring_email,
    advance_examiner_panel_invitation,
)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def pg_thesis_submit(request):
    """
    POST /stu/pg-thesis-submit/
    PG student uploads synopsis + full thesis report. Deliberately separate
    from thesis_submit (PhD's Dean Panel/Director/foreign-examiner workflow
    doesn't apply here) -- the supervisor and the batch's accepted examiner
    reference these files directly while scoring, no approval chain of its
    own. One submission per ThesisTopic, final -- once submitted it cannot
    be changed.
    """
    user = request.user
    try:
        student = Student.objects.get(id=user.extrainfo)
    except Student.DoesNotExist:
        return Response({'error': 'Student record not found.'}, status=404)
    if _student_programme_category(student) != 'PG':
        return Response(
            {'error': 'This thesis submission workflow is for PG students only.'},
            status=403,
        )
    thesis = ThesisTopic.objects.filter(student=student).order_by('-created_at').first()
    if thesis is None:
        return Response({'error': 'No thesis found for given submission.'}, status=400)

    if PGThesisSubmission.objects.filter(thesis=thesis).exists():
        return Response({'error': 'Thesis has already been submitted and cannot be changed.'}, status=400)

    syn = request.FILES.get('synopsis')
    rpt = request.FILES.get('thesis_report')
    if not syn or not rpt:
        return Response({'error': 'Missing fields'}, status=400)
    if syn.size > 5 * 1024 * 1024:
        return Response({'error': 'File too large'}, status=400)
    if rpt.size > 25 * 1024 * 1024:
        return Response({'error': 'File too large'}, status=400)
    # Client-declared content type is untrustworthy on its own, but combined
    # with the hardcoded .pdf extension in upload_pg_synopsis/upload_pg_report
    # it keeps a renamed non-PDF file from ever being stored/served as
    # something a browser would render (stored XSS via file upload).
    if syn.content_type != 'application/pdf' or rpt.content_type != 'application/pdf':
        return Response({'error': 'Both files must be PDFs'}, status=400)

    sub = PGThesisSubmission.objects.create(thesis=thesis, synopsis=syn, thesis_report=rpt)

    return Response({'submission_id': sub.id}, status=201)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def pg_thesis_submission_status(request):
    """
    GET /stu/pg-thesis-submission-status/
    Returns the requesting student's own PG thesis submission (if any).
    """
    user = request.user
    try:
        student = Student.objects.get(id=user.extrainfo)
    except Student.DoesNotExist:
        return Response({'error': 'Student record not found.'}, status=404)

    thesis = ThesisTopic.objects.filter(student=student).order_by('-created_at').first()
    if thesis is None:
        return Response({'submission': None}, status=200)

    sub = PGThesisSubmission.objects.filter(thesis=thesis).first()
    if sub is None:
        return Response({'submission': None}, status=200)

    return Response({
        'submission': {
            'id': sub.id,
            'submitted_at': sub.submitted_at.isoformat(),
            'synopsis_url': sub.synopsis.url if sub.synopsis else None,
            'thesis_report_url': sub.thesis_report.url if sub.thesis_report else None,
        },
    }, status=200)


def _maybe_finalize_numeric_grade(evaluation):
    """If both supervisor and examiner scores are in, compute and store the
    averaged numeric_grade on the ThesisEvaluation."""
    score_inputs = getattr(evaluation, 'score_inputs', None)
    if score_inputs is None:
        return
    if score_inputs.supervisor_score is not None and score_inputs.examiner_score is not None:
        avg = (score_inputs.supervisor_score + score_inputs.examiner_score) / 2
        evaluation.numeric_grade = round(avg, 1)
        evaluation.save(update_fields=['numeric_grade'])


def _thesis_examiner_candidate_to_dict(c):
    return {
        'id': c.id,
        'name': c.name,
        'position': c.position,
        'address': c.address,
        'phone': c.phone,
        'fax': c.fax,
        'email': c.email,
        'priority': c.priority,
        'status': c.status,
        'last_sent': c.last_sent.isoformat() if c.last_sent else None,
        'expires_at': c.expires_at.isoformat() if c.expires_at else None,
    }


def _thesis_examiner_panel_to_dict(panel):
    batch = panel.batch
    return {
        'id': panel.id,
        'batch_id': batch.id,
        'batch_name': str(batch),
        'discipline_acronym': batch.discipline.acronym if batch.discipline else None,
        'year': batch.year,
        'group_name': f"{batch.discipline.acronym} {batch.year}" if batch.discipline else str(batch.year),
        'status': panel.status,
        'hod_submitted_at': panel.hod_submitted_at.isoformat() if panel.hod_submitted_at else None,
        'dean_invited_at': panel.dean_invited_at.isoformat() if panel.dean_invited_at else None,
        'candidates': [_thesis_examiner_candidate_to_dict(c) for c in panel.candidates.all().order_by('priority')],
        'accepted_candidate': (
            _thesis_examiner_candidate_to_dict(panel.accepted_candidate)
            if panel.accepted_candidate else None
        ),
    }


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def supervisor_thesis_decimal_scores(request):
    """
    GET  /academic-procedures/api/supervisor/thesis-decimal-scores/
         Lists the requesting supervisor's decimal-mode ThesisEvaluations.
    POST /academic-procedures/api/supervisor/thesis-decimal-scores/
         Body: { "evaluation_id": <id>, "score": <0-100> }
         Records the supervisor's score -- this is the "forward to HOD" step.
    """
    user = request.user
    try:
        faculty = Faculty.objects.get(id=user.extrainfo)
    except Faculty.DoesNotExist:
        return JsonResponse({'error': 'Faculty record not found'}, status=404)

    if request.method == 'GET':
        evaluations = ThesisEvaluation.objects.filter(
            registration__thesis_topic__supervisor=faculty,
            registration__thesis_slot__evaluation_type='decimal',
        ).select_related(
            'registration__student__id__user', 'registration__semester',
            'registration__thesis_topic', 'registration__thesis_topic__pg_submission', 'score_inputs',
        )
        data = []
        for ev in evaluations:
            score_inputs = getattr(ev, 'score_inputs', None)
            submission = getattr(ev.registration.thesis_topic, 'pg_submission', None)
            data.append({
                'evaluation_id': ev.id,
                'student_name': ev.registration.student.id.user.get_full_name(),
                'student_roll': ev.registration.student.id.id,
                'semester_no': ev.registration.semester.semester_no,
                'credits': ev.registration.credits,
                'supervisor_score': score_inputs.supervisor_score if score_inputs else None,
                'examiner_score': score_inputs.examiner_score if score_inputs else None,
                'numeric_grade': ev.numeric_grade,
                'synopsis_url': submission.synopsis.url if submission and submission.synopsis else None,
                'thesis_report_url': submission.thesis_report.url if submission and submission.thesis_report else None,
            })
        return JsonResponse({'evaluations': data}, status=200)

    # POST
    evaluation_id = request.data.get('evaluation_id')
    try:
        score = round(float(request.data.get('score')), 1)
    except (TypeError, ValueError):
        return JsonResponse({'error': 'Invalid score'}, status=400)
    if not (0 <= score <= 100):
        return JsonResponse({'error': 'Score must be between 0 and 100'}, status=400)

    evaluation = get_object_or_404(
        ThesisEvaluation, id=evaluation_id,
        registration__thesis_topic__supervisor=faculty,
        registration__thesis_slot__evaluation_type='decimal',
    )
    if evaluation.numeric_grade is not None:
        return JsonResponse({'error': 'This evaluation has already been finalized'}, status=403)
    if not PGThesisSubmission.objects.filter(thesis=evaluation.registration.thesis_topic).exists():
        return JsonResponse(
            {'error': 'Student must submit their thesis and synopsis before scoring'}, status=400
        )
    score_inputs, _created = ThesisEvaluationScore.objects.get_or_create(evaluation=evaluation)
    score_inputs.supervisor_score = score
    score_inputs.supervisor_scored_at = timezone.now()
    score_inputs.save(update_fields=['supervisor_score', 'supervisor_scored_at'])
    _maybe_finalize_numeric_grade(evaluation)
    return JsonResponse({'detail': 'Score recorded'}, status=200)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def hod_examiner_panel_dashboard(request):
    """
    GET /hod/thesis-examiner-panels/
    Lists (discipline, year) groups (in HOD's disciplines) with verified
    decimal-mode thesis registrations -- e.g. CSE's "AI & ML" and "Data
    Science" specialization batches admitted the same year are grouped
    together for display -- each showing its own per-batch breakdown
    (supervisor-forwarded completion, existing panel status). The grouping
    is a UI convenience only: each specialization batch still gets its own
    independent ThesisExaminerPanel (own 4 examiners, own Dean ranking, own
    accepted examiner), matching the paper form (one sheet per
    specialization). Grouping just lets HOD nominate all of them in one sitting.
    """
    hod_disciplines = get_hod_disciplines(request.user)
    if not hod_disciplines:
        return JsonResponse({'groups': []}, status=200)

    regs = ThesisRegistration.objects.filter(
        status='verified',
        thesis_slot__evaluation_type='decimal',
        student__batch_id__discipline__acronym__in=hod_disciplines,
    ).select_related(
        'student__batch_id__discipline', 'student__id__user', 'thesis_topic',
        'thesis_topic__supervisor__id__user', 'thesis_topic__co_supervisor__id__user',
    )

    evaluations_by_reg_id = {
        ev.registration_id: ev
        for ev in ThesisEvaluation.objects.filter(registration__in=regs).select_related('score_inputs')
    }

    batches = {}
    for reg in regs:
        batch = reg.student.batch_id
        entry = batches.setdefault(batch.id, {
            'batch_id': batch.id, 'batch_name': str(batch.name),
            'discipline_id': batch.discipline_id, 'discipline_acronym': batch.discipline.acronym,
            'year': batch.year, 'total': 0, 'forwarded': 0, 'students': [],
        })
        entry['total'] += 1
        evaluation = evaluations_by_reg_id.get(reg.id)
        if evaluation and getattr(evaluation, 'score_inputs', None) and evaluation.score_inputs.supervisor_score is not None:
            entry['forwarded'] += 1

        supervisors = []
        if reg.thesis_topic:
            if reg.thesis_topic.supervisor:
                supervisors.append(reg.thesis_topic.supervisor.id.user.get_full_name())
            if reg.thesis_topic.co_supervisor:
                supervisors.append(reg.thesis_topic.co_supervisor.id.user.get_full_name())
        entry['students'].append({
            'roll_no': reg.student.id.id,
            'name': reg.student.id.user.get_full_name(),
            'supervisors': ' and '.join(supervisors) or None,
            'thesis_title': reg.thesis_topic.research_theme if reg.thesis_topic else None,
        })

    panels_by_batch_id = {
        p.batch_id: p
        for p in ThesisExaminerPanel.objects.filter(batch_id__in=batches.keys())
    }
    for entry in batches.values():
        panel = panels_by_batch_id.get(entry['batch_id'])
        entry['ready_for_panel'] = entry['total'] > 0 and entry['forwarded'] == entry['total']
        entry['panel_status'] = panel.status if panel else None
        entry['panel_id'] = panel.id if panel else None

    groups = {}
    for entry in batches.values():
        key = (entry['discipline_id'], entry['year'])
        group = groups.setdefault(key, {
            'discipline_id': entry['discipline_id'],
            'discipline_acronym': entry['discipline_acronym'],
            'year': entry['year'],
            'group_name': f"{entry['discipline_acronym']} {entry['year']}",
            'batches': [],
        })
        group['batches'].append(entry)

    return JsonResponse({'groups': list(groups.values())}, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def hod_submit_examiner_panel(request):
    """
    POST /hod/thesis-examiner-panels/submit/
    Body: { "discipline_id": <id>, "year": <int>, "batches": [
        { "batch_id": <id>, "candidates": [ {name, position, address, phone, fax, email} x4 ] },
        ...
    ] }
    Every specialization batch in this discipline+year that's ready for
    nomination (fully supervisor-scored, no panel yet) must be included and
    is submitted together in one action -- but each batch still gets its own
    independent ThesisExaminerPanel (own 4 examiners, own Dean ranking, own
    accepted examiner), matching the paper form (one sheet per specialization).
    """
    user = request.user
    try:
        faculty = Faculty.objects.get(id=user.extrainfo)
    except Faculty.DoesNotExist:
        return JsonResponse({'error': 'Faculty record not found'}, status=404)

    discipline_id = request.data.get('discipline_id')
    discipline = get_object_or_404(Discipline, id=discipline_id)
    try:
        year = int(request.data.get('year'))
    except (TypeError, ValueError):
        return JsonResponse({'error': 'Invalid year'}, status=400)

    hod_disciplines = get_hod_disciplines(user)
    if discipline.acronym not in hod_disciplines:
        return JsonResponse({'error': "You are not HOD of this discipline"}, status=403)

    batches_data = request.data.get('batches', [])
    if not batches_data:
        return JsonResponse({'error': 'No batches provided'}, status=400)
    for b in batches_data:
        if len(b.get('candidates', [])) != 4:
            return JsonResponse(
                {'error': 'Exactly 4 examiner candidates are required for each specialization'}, status=400
            )

    regs = ThesisRegistration.objects.filter(
        status='verified', thesis_slot__evaluation_type='decimal',
        student__batch_id__discipline=discipline, student__batch_id__year=year,
    ).select_related('student__batch_id')

    evaluations_by_reg_id = {
        ev.registration_id: ev
        for ev in ThesisEvaluation.objects.filter(registration__in=regs).select_related('score_inputs')
    }

    batch_totals = defaultdict(lambda: {'total': 0, 'forwarded': 0})
    for reg in regs:
        bid = reg.student.batch_id_id
        batch_totals[bid]['total'] += 1
        evaluation = evaluations_by_reg_id.get(reg.id)
        if evaluation and getattr(evaluation, 'score_inputs', None) and evaluation.score_inputs.supervisor_score is not None:
            batch_totals[bid]['forwarded'] += 1

    batches_with_panels = set(
        ThesisExaminerPanel.objects.filter(batch_id__in=batch_totals.keys()).values_list('batch_id', flat=True)
    )
    ready_batch_ids = {
        bid for bid, t in batch_totals.items()
        if t['total'] > 0 and t['total'] == t['forwarded']
        and bid not in batches_with_panels
    }
    submitted_batch_ids = {b.get('batch_id') for b in batches_data}
    if submitted_batch_ids != ready_batch_ids:
        return JsonResponse(
            {'error': "Submission must include exactly the specialization batches ready for "
                      "nomination in this discipline/year -- no partial submission"},
            status=400,
        )

    created_panel_ids = []
    try:
        with transaction.atomic():
            for b in batches_data:
                batch = get_object_or_404(Batch, id=b['batch_id'])
                panel = ThesisExaminerPanel.objects.create(batch=batch)
                for idx, c in enumerate(b['candidates'], start=1):
                    ThesisExaminerCandidate.objects.create(
                        panel=panel,
                        name=c.get('name', ''),
                        position=c.get('position', ''),
                        address=c.get('address', ''),
                        phone=c.get('phone', ''),
                        fax=c.get('fax', ''),
                        email=c.get('email', ''),
                        priority=idx,
                    )
                panel.hod_submitted_by = faculty
                panel.hod_submitted_at = timezone.now()
                panel.status = 'dean_pending'
                panel.save(update_fields=['hod_submitted_by', 'hod_submitted_at', 'status'])
                created_panel_ids.append(panel.id)
    except IntegrityError:
        # One of these batches already got a panel from a concurrent
        # submission between the readiness check above and this insert --
        # the whole atomic block rolled back, so nothing was half-created.
        return JsonResponse(
            {'error': 'One of these specializations was already submitted by another request. Refresh and try again.'},
            status=409,
        )

    for dean_user in _dean_academic_users():
        academics_module_notif(
            request.user, dean_user,
            f'Thesis examiner panels pending your ranking ({discipline.acronym} {year})',
        )

    return JsonResponse(
        {'detail': 'Examiner panels submitted to Dean', 'panel_ids': created_panel_ids}, status=200
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(['Dean Academic'])
def dean_examiner_panel_dashboard(request):
    """GET /dean/thesis-examiner-panels/ -- panels awaiting Dean action.
    Each panel is one specialization batch's independent process; the
    frontend groups them by discipline+year (see each panel's group_name)
    so Dean can rank every specialization in a discipline+year on one screen.
    """
    panels = ThesisExaminerPanel.objects.exclude(status='hod_pending') \
        .select_related('batch__discipline').prefetch_related('candidates')
    return JsonResponse({'panels': [_thesis_examiner_panel_to_dict(p) for p in panels]}, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['Dean Academic'])
def dean_rank_and_invite_examiner_panel(request):
    """
    POST /dean/thesis-examiner-panels/rank-and-invite/
    Body: { "panel_id": <id>, "ranked_candidate_ids": [id1, id2, id3, id4] }
    Sets the Dean's priority order (1-4) and immediately invites the
    top-ranked candidate (no Director step for PG).
    """
    panel_id = request.data.get('panel_id')
    ranked_ids = request.data.get('ranked_candidate_ids', [])
    panel = get_object_or_404(ThesisExaminerPanel, id=panel_id)

    if panel.status != 'dean_pending':
        return JsonResponse({'error': 'This panel is not awaiting Dean ranking'}, status=400)

    candidates = list(panel.candidates.all())
    if sorted(c.id for c in candidates) != sorted(ranked_ids):
        return JsonResponse(
            {'error': "ranked_candidate_ids must include exactly this panel's candidates"}, status=400
        )

    with transaction.atomic():
        # Two-phase: reassigning priorities to a new permutation in place can
        # momentarily collide with a not-yet-updated row's current value --
        # (panel, priority) is a unique constraint enforced immediately, not
        # deferred, and priority also has a DB-level CHECK (priority >= 0) so
        # negative placeholders aren't an option -- move everything out of
        # the 1-4 range first, then set the real values.
        for offset, cid in enumerate(ranked_ids, start=1):
            ThesisExaminerCandidate.objects.filter(id=cid, panel=panel).update(priority=1000 + offset)
        for idx, cid in enumerate(ranked_ids, start=1):
            ThesisExaminerCandidate.objects.filter(id=cid, panel=panel).update(priority=idx)

        panel.dean_reviewed_by = request.user
        panel.dean_invited_at = timezone.now()
        panel.status = 'invited'
        panel.save(update_fields=['dean_reviewed_by', 'dean_invited_at', 'status'])

    top_candidate = panel.candidates.order_by('priority').first()
    top_candidate.status = 'invited'
    top_candidate.last_sent = timezone.now()
    top_candidate.expires_at = timezone.now() + datetime.timedelta(days=INVITATION_TIMEOUT_DAYS)
    top_candidate.save(update_fields=['status', 'last_sent', 'expires_at'])
    try:
        send_examiner_panel_invitation_email(top_candidate)
    except Exception:
        logger.exception(f"Failed to send examiner panel invitation to {top_candidate.email}")

    return JsonResponse({'detail': 'Ranked and invitation sent'}, status=200)


@api_view(['POST'])
@permission_classes([AllowAny])
def examiner_panel_invitation_action(request, token, action):
    """GET, token-authenticated -- external examiner has no Fusion account."""
    candidate = get_object_or_404(ThesisExaminerCandidate, token=token)
    if candidate.is_expired() or candidate.is_finalized():
        return Response({'error': 'Invalid/expired'}, status=403)
    if action == 'accept':
        if candidate.status != 'invited':
            return Response({'error': 'This invitation has not been sent yet'}, status=403)
        candidate.status = 'accepted'
        candidate.save(update_fields=['status'])
        panel = candidate.panel
        panel.accepted_candidate = candidate
        panel.status = 'accepted'
        panel.save(update_fields=['accepted_candidate', 'status'])
        try:
            send_examiner_panel_scoring_email(candidate)
        except Exception:
            logger.exception(f"Failed to send examiner scoring email for token {candidate.token}")
        return Response({'detail': 'Accepted'}, status=200)
    if action == 'reject':
        candidate.status = 'rejected'
        candidate.save(update_fields=['status'])
        advance_examiner_panel_invitation(candidate.panel)
        return Response({'detail': 'Rejected'}, status=200)
    return Response({'error': 'Unknown action'}, status=400)


@api_view(['GET'])
@permission_classes([AllowAny])
def examiner_panel_batch_detail(request, token):
    """List every student in the panel's batch needing an examiner score."""
    candidate = get_object_or_404(ThesisExaminerCandidate, token=token)
    if candidate.status != 'accepted':
        return Response({'error': 'This invitation has not been accepted'}, status=403)

    regs = ThesisRegistration.objects.filter(
        status='verified', thesis_slot__evaluation_type='decimal',
        student__batch_id=candidate.panel.batch,
    ).select_related('student__id__user', 'thesis_topic', 'thesis_topic__pg_submission')

    evaluations_by_reg_id = {
        ev.registration_id: ev
        for ev in ThesisEvaluation.objects.filter(registration__in=regs).select_related('score_inputs')
    }

    students = []
    for reg in regs:
        evaluation = evaluations_by_reg_id.get(reg.id)
        score_inputs = getattr(evaluation, 'score_inputs', None) if evaluation else None
        submission = getattr(reg.thesis_topic, 'pg_submission', None) if reg.thesis_topic else None
        students.append({
            'evaluation_id': evaluation.id if evaluation else None,
            'student_name': reg.student.id.user.get_full_name(),
            'student_roll': reg.student.id.id,
            'credits': reg.credits,
            'examiner_score': score_inputs.examiner_score if score_inputs else None,
            'synopsis_url': submission.synopsis.url if submission and submission.synopsis else None,
            'thesis_report_url': submission.thesis_report.url if submission and submission.thesis_report else None,
        })

    return Response({
        'batch_name': str(candidate.panel.batch),
        'examiner_name': candidate.name,
        'students': students,
    }, status=200)


@api_view(['POST'])
@permission_classes([AllowAny])
def examiner_panel_submit_score(request, token):
    """Body: { "evaluation_id": <id>, "score": <0-100> }"""
    candidate = get_object_or_404(ThesisExaminerCandidate, token=token)
    if candidate.status != 'accepted':
        return Response({'error': 'This invitation has not been accepted'}, status=403)

    evaluation_id = request.data.get('evaluation_id')
    try:
        score = round(float(request.data.get('score')), 1)
    except (TypeError, ValueError):
        return Response({'error': 'Invalid score'}, status=400)
    if not (0 <= score <= 100):
        return Response({'error': 'Score must be between 0 and 100'}, status=400)

    evaluation = get_object_or_404(
        ThesisEvaluation, id=evaluation_id,
        registration__student__batch_id=candidate.panel.batch,
        registration__thesis_slot__evaluation_type='decimal',
    )
    if evaluation.numeric_grade is not None:
        return Response({'error': 'This evaluation has already been finalized'}, status=403)
    score_inputs, _created = ThesisEvaluationScore.objects.get_or_create(evaluation=evaluation)
    score_inputs.examiner_candidate = candidate
    score_inputs.examiner_score = score
    score_inputs.examiner_scored_at = timezone.now()
    score_inputs.save(update_fields=['examiner_candidate', 'examiner_score', 'examiner_scored_at'])
    _maybe_finalize_numeric_grade(evaluation)

    return Response({'detail': 'Score recorded'}, status=200)


# ===========================================================================
# Progress Seminar Slot Semester-Level Registration
#
# Gated the same way as thesis enrollment: the student's ThesisTopic must be
# dean_approved. This is only the enrollment step -- the substantive report
# submission and RPC review (ProgressSeminarEntry) is separate and unaffected.
# ===========================================================================

def _progress_seminar_reg_to_dict(reg):
    """Serialize a ProgressSeminarRegistration instance to a plain dict."""
    if reg is None:
        return None
    slot = reg.progress_seminar_slot
    seminars_list = [
        {'id': s.id, 'code': s.code, 'name': s.name, 'credit': s.credit}
        for s in slot.seminars.all()
    ]
    resolved_seminar = _catalog_entry_to_dict(_resolve_discipline_matched_entry(slot.seminars, reg.student))
    return {
        'id': reg.id,
        'status': reg.status,
        'remarks': reg.remarks,
        'registered_on': reg.registered_on.isoformat(),
        'progress_seminar_slot': {
            'id': slot.id,
            'name': slot.name,
            'info': slot.seminar_slot_info or '',
            'duration': slot.duration,
            'seminars': seminars_list,
            'resolved_seminar': resolved_seminar,
        },
        'student': {
            'id': reg.student.id.id,
            'name': reg.student.id.user.get_full_name(),
        },
        'semester_no': reg.semester.semester_no,
        'programme_category': _student_programme_category(reg.student),
    }


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def student_progress_seminar_enrollment_api(request):
    """
    GET  /stu/progress-seminar-enrollment/
         Returns the current semester's SeminarSlot, the student's
         ThesisTopic approval status, and any existing registration.

    POST /stu/progress-seminar-enrollment/
         Creates a new ProgressSeminarRegistration for the current semester.
         Requires thesis_topic to be dean_approved.
    """
    user = request.user
    try:
        student = Student.objects.get(id=user.extrainfo)
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student record not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'User setup error: {type(e).__name__}: {e}'}, status=400)

    try:
        if not student.batch_id or not student.batch_id.curriculum:
            return JsonResponse({'error': 'Student batch or curriculum is not configured'}, status=400)
        try:
            semester = Semester.objects.get(
                curriculum=student.batch_id.curriculum,
                semester_no=student.curr_semester_no,
            )
        except Semester.DoesNotExist:
            return JsonResponse({'error': 'Current semester not found in curriculum'}, status=400)

        if request.method == 'POST':
            _elig = get_pg_phd_registration_eligibility(
                timezone.now().date(), _student_programme_category(student),
                student.curr_semester_no, datetime.datetime.now().year)
            if isinstance(_elig, JsonResponse):
                return _elig

        topic = ThesisTopic.objects.filter(student=student).order_by('-created_at').first()
        topic_approved = topic is not None and topic.status == 'dean_approved'

        slot = ProgressSeminarSlot.objects.filter(semester=semester).first()

        try:
            reg = ProgressSeminarRegistration.objects.get(student=student, semester=semester)
            reg_data = _progress_seminar_reg_to_dict(reg)
        except ProgressSeminarRegistration.DoesNotExist:
            reg = None
            reg_data = None

        if request.method == 'GET':
            slot_data = None
            if slot:
                slot_data = {
                    'id': slot.id,
                    'name': slot.name,
                    'info': slot.seminar_slot_info or '',
                    'duration': slot.duration,
                    'seminars': [
                        {'id': s.id, 'code': s.code, 'name': s.name, 'credit': s.credit}
                        for s in slot.seminars.all()
                    ],
                    'resolved_seminar': _catalog_entry_to_dict(
                        _resolve_discipline_matched_entry(slot.seminars, student)
                    ),
                }
            return JsonResponse({
                'thesis_topic_approved': topic_approved,
                'current_semester_no': student.curr_semester_no,
                'progress_seminar_slot': slot_data,
                'registration': reg_data,
            }, status=200)

    except Exception as e:
        return JsonResponse({'error': f'Internal error: {type(e).__name__}: {e}'}, status=500)

    # POST: create registration
    if reg is not None:
        return JsonResponse(
            {'error': 'Already registered for this semester', 'registration': reg_data},
            status=400,
        )
    if not topic_approved:
        return JsonResponse(
            {'error': 'Thesis topic must be dean-approved before registering for progress seminar'},
            status=403,
        )
    if slot is None:
        return JsonResponse(
            {'error': 'No progress seminar slot configured for your current semester'},
            status=400,
        )

    current_count = ProgressSeminarRegistration.objects.filter(
        progress_seminar_slot=slot, status__in=['pending', 'verified']
    ).count()
    if current_count >= slot.max_registration_limit:
        return JsonResponse(
            {'error': 'Progress seminar slot has reached maximum capacity'},
            status=400,
        )

    now = _dt.datetime.now()
    year, month = now.year, now.month
    session = f"{year}-{str(year + 1)[2:]}" if month >= 7 else f"{year - 1}-{str(year)[2:]}"

    reg = ProgressSeminarRegistration.objects.create(
        student=student,
        progress_seminar_slot=slot,
        semester=semester,
        working_year=year,
        status='pending',
    )
    return JsonResponse(_progress_seminar_reg_to_dict(reg), status=201)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def admin_progress_seminar_enrollment_list(request):
    """
    GET /acadadmin/progress-seminar-enrollments/?semester=<no>&status=<status>
    Lists all ProgressSeminarRegistration entries. Supports optional filters:
      ?semester=<semester_no>   filter by semester number
      ?status=pending|verified|rejected
    """
    qs = ProgressSeminarRegistration.objects.select_related(
        'student__id__user', 'progress_seminar_slot', 'semester'
    ).all().order_by('-registered_on')

    sem_no = request.GET.get('semester')
    if sem_no:
        qs = qs.filter(semester__semester_no=sem_no)

    status_filter = request.GET.get('status')
    if status_filter:
        qs = qs.filter(status__iexact=status_filter)

    data = []
    for reg in qs:
        entry = _progress_seminar_reg_to_dict(reg)
        topic = ThesisTopic.objects.filter(student=reg.student).order_by('-created_at').first()
        entry['topic_status'] = topic.status if topic else None
        data.append(entry)

    return JsonResponse({'registrations': data}, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['acadadmin'])
def admin_verify_progress_seminar_enrollments(request):
    """
    POST /acadadmin/progress-seminar-enrollments/verify/
    Body: { "ids": [1, 2, 3] }
    Marks the given ProgressSeminarRegistration records as 'verified' and
    auto-creates the single grade block (progress seminars are fixed at 3
    credits, unlike thesis's variable 3/6/9/12).
    """
    ids = request.data.get('ids', [])
    if not ids:
        return JsonResponse({'error': 'No registration IDs provided'}, status=400)

    now = _dt.datetime.now(_dt.timezone.utc)
    regs = ProgressSeminarRegistration.objects.filter(id__in=ids, status='pending')
    count = 0
    for reg in regs:
        reg.status = 'verified'
        reg.save(update_fields=['status'])
        ProgressSeminarEvaluation.objects.get_or_create(registration=reg)
        count += 1
        _progress_seminar_notify(
            sender=request.user,
            recipient=reg.student.id.user,
            verb='Progress Seminar registration verified',
            description="Your Progress Seminar registration for this semester has been verified.",
        )
    return JsonResponse({'verified_count': count}, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['acadadmin'])
def admin_reject_progress_seminar_enrollments(request):
    """
    POST /acadadmin/progress-seminar-enrollments/reject/
    Body: { "ids": [1, 2], "remarks": "Reason for rejection" }
    Marks the given ProgressSeminarRegistration records as 'rejected'.
    """
    ids = request.data.get('ids', [])
    remarks = request.data.get('remarks', '')
    if not ids:
        return JsonResponse({'error': 'No registration IDs provided'}, status=400)

    pending_qs = ProgressSeminarRegistration.objects.filter(id__in=ids, status='pending')
    to_notify = list(pending_qs.select_related('student__id__user'))
    updated = pending_qs.update(status='rejected', remarks=remarks)
    for reg in to_notify:
        _progress_seminar_notify(
            sender=request.user,
            recipient=reg.student.id.user,
            verb='Progress Seminar registration rejected',
            description=f"Your Progress Seminar registration for this semester was rejected. "
                        f"Remarks: {remarks or '—'}",
        )
    return JsonResponse({'rejected_count': updated}, status=200)


# ===========================================================================
# Teaching Credit Slot Semester-Level Registration
#
# Gated on ComprehensiveExam.status == 'passed', same precondition already
# enforced by the substantive TeachingCreditAllocation flow. This is only
# the enrollment step -- the choice-and-allocation process is separate and
# unaffected.
# ===========================================================================

def _teaching_credit_enrollment_to_dict(reg):
    """Serialize a TeachingCreditRegistration instance to a plain dict."""
    if reg is None:
        return None
    slot = reg.teaching_credit_slot
    credits_list = [
        {'id': t.id, 'code': t.code, 'name': t.name, 'credit': t.credit}
        for t in slot.teaching_credits.all()
    ]
    resolved_teaching_credit = _catalog_entry_to_dict(_resolve_discipline_matched_entry(slot.teaching_credits, reg.student))
    return {
        'id': reg.id,
        'status': reg.status,
        'remarks': reg.remarks,
        'registered_on': reg.registered_on.isoformat(),
        'academic_session': reg.academic_session,
        'teaching_credit_slot': {
            'id': slot.id,
            'name': slot.name,
            'info': slot.teaching_credit_slot_info or '',
            'duration': slot.duration,
            'teaching_credits': credits_list,
            'resolved_teaching_credit': resolved_teaching_credit,
        },
        'student': {
            'id': reg.student.id.id,
            'name': reg.student.id.user.get_full_name(),
        },
        'semester_no': reg.semester.semester_no,
        'programme_category': _student_programme_category(reg.student),
    }


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def student_teaching_credit_enrollment_api(request):
    """
    GET  /stu/teaching-credit-enrollment/
         Returns the current semester's TeachingCreditSlot, the student's
         Comprehensive Exam status, and any existing registration.

    POST /stu/teaching-credit-enrollment/
         Creates a new TeachingCreditRegistration for the current semester.
         Requires ComprehensiveExam.status == 'passed'.
    """
    user = request.user
    try:
        student = Student.objects.get(id=user.extrainfo)
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student record not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'User setup error: {type(e).__name__}: {e}'}, status=400)

    try:
        if not student.batch_id or not student.batch_id.curriculum:
            return JsonResponse({'error': 'Student batch or curriculum is not configured'}, status=400)
        try:
            semester = Semester.objects.get(
                curriculum=student.batch_id.curriculum,
                semester_no=student.curr_semester_no,
            )
        except Semester.DoesNotExist:
            return JsonResponse({'error': 'Current semester not found in curriculum'}, status=400)

        if request.method == 'POST':
            _elig = get_pg_phd_registration_eligibility(
                timezone.now().date(), _student_programme_category(student),
                student.curr_semester_no, datetime.datetime.now().year)
            if isinstance(_elig, JsonResponse):
                return _elig

        comprehensive_exam_passed = ComprehensiveExam.objects.filter(
            student=student, status='passed'
        ).exists()

        slot = TeachingCreditSlot.objects.filter(semester=semester).first()

        try:
            reg = TeachingCreditRegistration.objects.get(student=student, semester=semester)
            reg_data = _teaching_credit_enrollment_to_dict(reg)
        except TeachingCreditRegistration.DoesNotExist:
            reg = None
            reg_data = None

        if request.method == 'GET':
            slot_data = None
            if slot:
                slot_data = {
                    'id': slot.id,
                    'name': slot.name,
                    'info': slot.teaching_credit_slot_info or '',
                    'duration': slot.duration,
                    'teaching_credits': [
                        {'id': t.id, 'code': t.code, 'name': t.name, 'credit': t.credit}
                        for t in slot.teaching_credits.all()
                    ],
                    'resolved_teaching_credit': _catalog_entry_to_dict(
                        _resolve_discipline_matched_entry(slot.teaching_credits, student)
                    ),
                }
            return JsonResponse({
                'comprehensive_exam_passed': comprehensive_exam_passed,
                'current_semester_no': student.curr_semester_no,
                'teaching_credit_slot': slot_data,
                'registration': reg_data,
            }, status=200)

    except Exception as e:
        return JsonResponse({'error': f'Internal error: {type(e).__name__}: {e}'}, status=500)

    # POST: create registration
    if reg is not None:
        return JsonResponse(
            {'error': 'Already registered for this semester', 'registration': reg_data},
            status=400,
        )
    if not comprehensive_exam_passed:
        return JsonResponse(
            {'error': 'Comprehensive Examination must be passed before registering for teaching credit'},
            status=403,
        )
    if slot is None:
        return JsonResponse(
            {'error': 'No teaching credit slot configured for your current semester'},
            status=400,
        )

    current_count = TeachingCreditRegistration.objects.filter(
        teaching_credit_slot=slot, status__in=['pending', 'verified']
    ).count()
    if current_count >= slot.max_registration_limit:
        return JsonResponse(
            {'error': 'Teaching credit slot has reached maximum capacity'},
            status=400,
        )

    now = _dt.datetime.now()
    year, month = now.year, now.month
    session = f"{year}-{str(year + 1)[2:]}" if month >= 7 else f"{year - 1}-{str(year)[2:]}"

    reg = TeachingCreditRegistration.objects.create(
        student=student,
        teaching_credit_slot=slot,
        semester=semester,
        working_year=year,
        academic_session=session,
        status='pending',
    )
    return JsonResponse(_teaching_credit_enrollment_to_dict(reg), status=201)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def admin_teaching_credit_enrollment_list(request):
    """
    GET /acadadmin/teaching-credit-enrollments/?semester=<no>&status=<status>
    Lists all TeachingCreditRegistration entries. Supports optional filters:
      ?semester=<semester_no>   filter by semester number
      ?status=pending|verified|rejected
    """
    qs = TeachingCreditRegistration.objects.select_related(
        'student__id__user', 'teaching_credit_slot', 'semester'
    ).all().order_by('-registered_on')

    sem_no = request.GET.get('semester')
    if sem_no:
        qs = qs.filter(semester__semester_no=sem_no)

    status_filter = request.GET.get('status')
    if status_filter:
        qs = qs.filter(status__iexact=status_filter)

    data = [_teaching_credit_enrollment_to_dict(reg) for reg in qs]
    return JsonResponse({'registrations': data}, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['acadadmin'])
def admin_verify_teaching_credit_enrollments(request):
    """
    POST /acadadmin/teaching-credit-enrollments/verify/
    Body: { "ids": [1, 2, 3] }
    Marks the given TeachingCreditRegistration records as 'verified'. Unlike
    thesis/seminar, no grade-block is created here -- the substantive
    satisfactory/not_satisfactory result is recorded on the separate
    TeachingCreditAllocation once that process completes.
    """
    ids = request.data.get('ids', [])
    if not ids:
        return JsonResponse({'error': 'No registration IDs provided'}, status=400)

    updated = TeachingCreditRegistration.objects.filter(id__in=ids, status='pending').update(
        status='verified',
    )
    return JsonResponse({'verified_count': updated}, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['acadadmin'])
def admin_reject_teaching_credit_enrollments(request):
    """
    POST /acadadmin/teaching-credit-enrollments/reject/
    Body: { "ids": [1, 2], "remarks": "Reason for rejection" }
    Marks the given TeachingCreditRegistration records as 'rejected'.
    """
    ids = request.data.get('ids', [])
    remarks = request.data.get('remarks', '')
    if not ids:
        return JsonResponse({'error': 'No registration IDs provided'}, status=400)

    updated = TeachingCreditRegistration.objects.filter(id__in=ids, status='pending').update(
        status='rejected',
        remarks=remarks,
    )
    return JsonResponse({'rejected_count': updated}, status=200)


# ===========================================================================
# PhD Course (Coursework) Registration
#
# Standalone request-and-verify workflow, independent of the UG/PG backlog
# add-course flow (add_course / CourseAddRequest). PhD students don't go
# through pre-registration/final-registration or the backlog Add/Drop tab —
# they self-submit a request per curriculum course slot for their current
# semester, and acadadmin verifies it here.
# ===========================================================================

def _is_phd_student(student):
    """True if `student` is enrolled in a PhD or PG (M.Tech/M.Des) programme --
    the two categories that use this lightweight self-submit-and-verify
    registration flow instead of the UG-style pre-registration/allocation/
    final-registration pipeline.
    programme is stored inconsistently across seeded data ('PhD' vs 'Ph.D'),
    so normalize it; also fall back to the batch name (e.g. 'PhD (Odd)').
    PG detection falls back to the curriculum's Programme.category since PG
    students don't have an equivalent programme/batch-name shorthand."""
    programme_norm = (student.programme or '').upper().replace('.', '')
    batch_name = student.batch_id.name if student.batch_id else ''
    if programme_norm == 'PHD' or batch_name.upper().startswith('PHD'):
        return True
    try:
        return student.batch_id.curriculum.programme.category == 'PG'
    except AttributeError:
        return False


def _resolve_phd_student(request):
    """Returns (student, error_response). error_response is a JsonResponse
    if the requester isn't a valid PhD/PG student, else None."""
    try:
        student = Student.objects.select_related('batch_id__curriculum').get(
            id__user=request.user
        )
    except Student.DoesNotExist:
        return None, JsonResponse({'error': 'Student record not found'}, status=404)

    if not _is_phd_student(student):
        return None, JsonResponse(
            {'error': 'This section is for PhD/PG students only'}, status=403
        )

    if not student.batch_id or not student.batch_id.curriculum:
        return None, JsonResponse(
            {'error': 'Student batch or curriculum is not configured'}, status=400
        )

    return student, None


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(['student'])
def phd_student_status(request):
    """
    GET /stu/phd/status/
    Lightweight check used by the frontend to decide whether to show the
    "PhD Course Registration" tab at all, before fetching any curriculum data.
    """
    try:
        student = Student.objects.select_related('batch_id').get(id__user=request.user)
    except Student.DoesNotExist:
        return JsonResponse({'is_phd': False}, status=200)

    return JsonResponse({'is_phd': _is_phd_student(student)}, status=200)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(['student'])
def phd_course_slots(request):
    """
    GET /stu/phd/course-slots/
    Returns the CourseSlots in the PhD student's current-semester curriculum,
    excluding slots already registered or requested (pending/approved).
    """
    student, err = _resolve_phd_student(request)
    if err:
        return err

    try:
        semester = Semester.objects.get(
            curriculum=student.batch_id.curriculum,
            semester_no=student.curr_semester_no,
        )
    except Semester.DoesNotExist:
        return JsonResponse({'error': 'Current semester not found in curriculum'}, status=400)

    taken_slot_ids = set(
        PhDCourseRegistrationRequest.objects.filter(
            student=student, semester=semester, status__in=['Pending', 'Approved']
        ).values_list('course_slot_id', flat=True)
    )

    slots = CourseSlot.objects.filter(semester=semester).annotate(course_count=Count('courses'))
    data = [
        {'id': s.id, 'name': s.name, 'course_count': s.course_count}
        for s in slots if s.id not in taken_slot_ids
    ]
    # Registration window (same Calendar gate enforced at submit) so the form only renders when open.
    import json as _json
    _elig = get_pg_phd_registration_eligibility(
        timezone.now().date(), _student_programme_category(student),
        student.curr_semester_no, datetime.datetime.now().year)
    registration_open = _elig is None
    registration_message = '' if registration_open else _json.loads(_elig.content).get('error', 'Registration is not open.')
    return JsonResponse({
        'semester_no': semester.semester_no,
        'slots': data,
        'registration_open': registration_open,
        'registration_message': registration_message,
    }, status=200)


# Same backlog-eligibility rule as the UG add-course flow (add_course): a BL
# slot course may only be (re)registered if the student's latest grade in it is
# below C+. PG/PhD have no sections, so only this grade gate carries over.
_PHD_BL_ALLOWED_GRADES = ['F', 'X', 'CD', 'C', 'D+', 'D']


_PHD_BL_BACKLOG_GRADES = {'F', 'X', 'CD'}


def registration_type_for_grade(grade):
    if grade in _PHD_BL_BACKLOG_GRADES:
        return 'Backlog'
    if grade in _PHD_BL_ALLOWED_GRADES:
        return 'Improvement'
    return 'Regular'


def latest_grade(roll_no, course):
    """The student's most recent grade row in a course, or None."""
    return Student_grades.objects.filter(
        roll_no=roll_no, course_id=course
    ).order_by('-year', '-semester').first()


def _phd_bl_grade_status(username, course):
    """Returns (allowed, grade) for a BL course. grade is None when the student
    has no recorded grade in the course (also treated as not allowed)."""
    sg = latest_grade(username, course)
    if not sg:
        return False, None
    return sg.grade in _PHD_BL_ALLOWED_GRADES, sg.grade


def _phd_bl_source_courses(student):
    """The student's own courses eligible to be cleared through a BL slot: the
    latest grade in each is below C+. A course taken in an OE slot may be
    swapped for a different one; anything else has to be retaken as it is."""
    username = student.id.user.username
    # The latest grade decides, so the filter cannot come first: a course whose
    # latest attempt cleared it must drop out, not fall back to the older low one.
    latest = {}
    for sg in Student_grades.objects.filter(
        roll_no=username
    ).select_related('course_id').order_by('year', 'semester'):
        latest[sg.course_id_id] = sg
    latest = {course_id: sg for course_id, sg in latest.items()
              if sg.grade in _PHD_BL_ALLOWED_GRADES}

    oe_course_ids = set(course_registration.objects.filter(
        student_id=student, course_id__in=latest.keys(),
        course_slot_id__name__istartswith='OE',
    ).values_list('course_id', flat=True))

    out = []
    for course_id, sg in latest.items():
        course = sg.course_id
        out.append({
            'id': course.id,
            'code': course.code,
            'name': course.name,
            'credit': course.credit,
            'grade': sg.grade,
            'registration_type': registration_type_for_grade(sg.grade),
            'replaceable': course_id in oe_course_ids,
        })
    out.sort(key=lambda c: c['code'])
    return out


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(['student'])
def phd_course_slot_courses(request):
    """
    GET /stu/phd/course-slots/courses/?slot_id=<id>
    Returns the courses within a slot in the student's current semester.
    """
    student, err = _resolve_phd_student(request)
    if err:
        return err

    slot_id = request.query_params.get('slot_id')
    if not slot_id:
        return JsonResponse({'error': 'slot_id query parameter is required'}, status=400)

    try:
        semester = Semester.objects.get(
            curriculum=student.batch_id.curriculum,
            semester_no=student.curr_semester_no,
        )
        slot = CourseSlot.objects.get(id=slot_id, semester=semester)
    except (Semester.DoesNotExist, CourseSlot.DoesNotExist):
        return JsonResponse({'error': 'Course slot not found in current semester'}, status=404)

    # A BL slot clears a past course, so it is driven by the student's own
    # below-C+ grades rather than the slot list; the slot list is only the pool
    # a replaceable (OE) course can be swapped for.
    is_bl = slot.name.startswith('BL')
    username = student.id.user.username
    courses = []
    for c in slot.courses.all():
        entry = {'id': c.id, 'code': c.code, 'name': c.name, 'credit': c.credit,
                 'bl_eligible': True, 'bl_grade': None}
        if is_bl:
            allowed, grade = _phd_bl_grade_status(username, c)
            entry['bl_eligible'] = allowed
            entry['bl_grade'] = grade
        courses.append(entry)

    payload = {'courses': courses}
    if is_bl:
        # A source already in flight through another slot is spoken for, so the
        # dropdown can leave it out instead of failing on submit.
        claimed = dict(PhDCourseRegistrationRequest.objects.filter(
            student=student, semester=semester, status__in=['Pending', 'Approved'],
            source_course__isnull=False,
        ).exclude(course_slot=slot).values_list('source_course_id', 'course_slot__name'))
        sources = _phd_bl_source_courses(student)
        for source in sources:
            source['claimed_by'] = claimed.get(source['id'])
        payload['source_courses'] = sources
    return JsonResponse(payload, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['student'])
def phd_submit_course_request(request):
    """
    POST /stu/phd/course-request/
    Body: { "slot_id": <id>, "course_id": <id> }
    Creates a Pending PhDCourseRegistrationRequest for the student's current semester.
    One request per slot per semester.
    """
    student, err = _resolve_phd_student(request)
    if err:
        return err

    _elig = get_pg_phd_registration_eligibility(
        timezone.now().date(), _student_programme_category(student) or 'PHD',
        student.curr_semester_no, datetime.datetime.now().year)
    if isinstance(_elig, JsonResponse):
        return _elig

    slot_id = request.data.get('slot_id')
    course_id = request.data.get('course_id')
    source_course_id = request.data.get('source_course_id')
    if not slot_id or not course_id:
        return JsonResponse({'error': 'slot_id and course_id are required'}, status=400)

    try:
        semester = Semester.objects.get(
            curriculum=student.batch_id.curriculum,
            semester_no=student.curr_semester_no,
        )
        slot = CourseSlot.objects.get(id=slot_id, semester=semester)
    except Semester.DoesNotExist:
        return JsonResponse({'error': 'Current semester not found in curriculum'}, status=400)
    except CourseSlot.DoesNotExist:
        return JsonResponse({'error': 'Course slot not found in current semester'}, status=404)

    is_bl = slot.name.startswith('BL')
    source_course = None
    registration_type = ''

    if is_bl:
        if not source_course_id:
            return JsonResponse({'error': 'source_course_id is required for a BL slot'}, status=400)
        eligible = {c['id']: c for c in _phd_bl_source_courses(student)}
        source = eligible.get(int(source_course_id))
        if not source:
            return JsonResponse({'error': 'That course is not one you can clear: a BL slot needs a course you scored below C+ in.'}, status=400)
        registration_type = source['registration_type']
        # An OE course can be cleared by a different course from this slot;
        # anything else has to be the same course taken again.
        if source['replaceable']:
            # a retake is allowed even though it is not one of the stand-ins
            if int(course_id) == source['id']:
                course = Courses.objects.get(id=source['id'])
            else:
                try:
                    course = slot.courses.get(id=course_id)
                except Courses.DoesNotExist:
                    return JsonResponse({'error': 'Course not found in this slot'}, status=404)
        elif int(course_id) != source['id']:
            return JsonResponse({'error': f"{source['code']} was not taken in an open elective slot, so it has to be cleared by retaking the same course."}, status=400)
        else:
            course = Courses.objects.get(id=source['id'])
        source_course = Courses.objects.get(id=source['id'])
    else:
        try:
            course = slot.courses.get(id=course_id)
        except Courses.DoesNotExist:
            return JsonResponse({'error': 'Course not found in this slot'}, status=404)

    if PhDCourseRegistrationRequest.objects.filter(
        student=student, semester=semester, course_slot=slot, status__in=['Pending', 'Approved']
    ).exists():
        return JsonResponse({'error': 'You already have a request for this slot'}, status=400)

    # One backlog is cleared once: without this a second slot could clear the
    # same course again and earn its credits twice.
    if source_course:
        clash = PhDCourseRegistrationRequest.objects.filter(
            student=student, semester=semester, source_course=source_course,
            status__in=['Pending', 'Approved'],
        ).exclude(course_slot=slot).select_related('course_slot').first()
        if clash:
            return JsonResponse({'error': f"{source_course.code} is already being cleared "
                                          f"through {clash.course_slot.name} this semester."}, status=400)

    academic_year, semester_type = generate_current_session(
        datetime.datetime.now().year, student.curr_semester_no
    )

    # unique_together is (student, semester, course_slot) regardless of status,
    # so a prior Rejected request for this slot must be reused, not re-created.
    req, _created = PhDCourseRegistrationRequest.objects.update_or_create(
        student=student, semester=semester, course_slot=slot,
        defaults={
            'academic_year': academic_year,
            'semester_type': semester_type,
            'course': course,
            'source_course': source_course,
            'registration_type': registration_type,
            'status': 'Pending',
            'remarks': '',
            'requested_at': timezone.now(),
            'processed_at': None,
            'processed_by': None,
        },
    )

    _phd_course_registration_notify(
        sender=request.user,
        recipient=_academic_office_users(),
        verb='PhD course registration request pending approval',
        description=f"{student.id.user.get_full_name()} has requested registration in "
                    f"{course.code} - {course.name} ({slot.name}).",
    )

    return JsonResponse({
        'id': req.id,
        'slot': slot.name,
        'course': course.code,
        'course_name': course.name,
        'status': req.status,
    }, status=201)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(['student'])
def phd_my_course_requests(request):
    """
    GET /stu/phd/my-course-requests/
    Returns the PhD student's own course requests, most recent first.
    """
    student, err = _resolve_phd_student(request)
    if err:
        return err

    qs = PhDCourseRegistrationRequest.objects.filter(student=student) \
        .select_related('course', 'source_course', 'course_slot', 'semester') \
        .order_by('-requested_at')

    data = [{
        'id': r.id,
        'slot': r.course_slot.name,
        'course': r.course.code,
        'course_name': r.course.name,
        'credit': r.course.credit,
        'source_course': r.source_course.code if r.source_course_id else None,
        'registration_type': r.registration_type or None,
        'semester_no': r.semester.semester_no,
        'academic_year': r.academic_year,
        'semester_type': r.semester_type,
        'status': r.status,
        'remarks': r.remarks,
        'requested_at': r.requested_at.isoformat(),
        'processed_at': r.processed_at.isoformat() if r.processed_at else None,
        'programme_category': _student_programme_category(r.student),
    } for r in qs]

    return JsonResponse({'requests': data}, status=200)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(list(ALL_ACAD_ROLES))
def phd_admin_list_requests(request):
    """
    GET /acadadmin/phd/course-requests/?academic_year=&semester_type=&semester=&status=
    Lists all PhDCourseRegistrationRequest entries, filterable. `semester` is
    the semester number, the common filter axis with the thesis/progress-seminar/
    teaching-credit enrollment list endpoints (used by the merged admin view).
    """
    qs = PhDCourseRegistrationRequest.objects.select_related(
        'student__id__user', 'student__batch_id__discipline', 'course', 'source_course',
        'course_slot', 'semester'
    ).all().order_by('-requested_at')
    qs = scope_via_student(qs, scopes_for(request.user), 'student')

    year = request.GET.get('academic_year', '').strip()
    sem_type = request.GET.get('semester_type', '').strip()
    sem_no = request.GET.get('semester', '').strip()
    status_filter = request.GET.get('status', '').strip()

    if year:
        qs = qs.filter(academic_year=year)
    if sem_type:
        qs = qs.filter(semester_type=sem_type)
    if sem_no:
        qs = qs.filter(semester__semester_no=sem_no)
    if status_filter:
        qs = qs.filter(status__iexact=status_filter)

    # Pending first, so the rows needing action survive the cap below.
    qs = qs.order_by(
        Case(When(status__iexact='pending', then=0), default=1,
             output_field=IntegerField()),
        '-requested_at',
    )

    qs = qs[:500]

    data = [{
        'id': r.id,
        'student': r.student.id.user.username,
        'student_name': f"{r.student.id.user.first_name} {r.student.id.user.last_name}".strip(),
        'discipline': (r.student.batch_id.discipline.acronym if r.student.batch_id and r.student.batch_id.discipline_id else ''),
        'specialization': r.student.specialization or '',
        'slot': r.course_slot.name,
        'course': r.course.code,
        'course_name': r.course.name,
        'credit': r.course.credit,
        'source_course': r.source_course.code if r.source_course_id else None,
        'registration_type': r.registration_type or None,
        'semester_no': r.semester.semester_no,
        'academic_year': r.academic_year,
        'semester_type': r.semester_type,
        'status': r.status,
        'remarks': r.remarks,
        'requested_at': r.requested_at.isoformat(),
        'processed_at': r.processed_at.isoformat() if r.processed_at else None,
        'programme_category': _student_programme_category(r.student),
    } for r in qs]

    return JsonResponse({'requests': data}, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@transaction.atomic
@role_required(list(ALL_ACAD_ROLES))
def phd_admin_process_requests(request):
    """
    POST /acadadmin/phd/course-requests/process/
    Body: { "request_ids": [1, 2, 3], "action": "approve"|"reject", "remarks": "..." }
    Approving creates the real course_registration row; rejecting just marks status.
    """
    request_ids = request.data.get('request_ids', [])
    request_ids = scoped_ids(
        PhDCourseRegistrationRequest, request_ids, scopes_for(request.user))
    action = str(request.data.get('action', 'approve')).lower().strip()
    remarks = request.data.get('remarks', '')

    if not request_ids or not isinstance(request_ids, list):
        return JsonResponse({'error': 'request_ids must be a non-empty array'}, status=400)
    if action not in ['approve', 'reject']:
        return JsonResponse({'error': 'action must be either "approve" or "reject"'}, status=400)

    admin_extrainfo = getattr(request.user, 'extrainfo', None)
    results = []
    now = timezone.now()

    for req_id in request_ids:
        try:
            req_id = int(req_id)
            req = PhDCourseRegistrationRequest.objects.select_related(
                'student', 'course', 'course_slot', 'semester', 'student__batch_id'
            ).select_for_update(of=('self',)).get(id=req_id)
        except (ValueError, TypeError):
            results.append({'id': req_id, 'status': 'error', 'detail': 'Invalid ID format'})
            continue
        except PhDCourseRegistrationRequest.DoesNotExist:
            results.append({'id': req_id, 'status': 'not_found'})
            continue

        if req.status != 'Pending':
            results.append({'id': req_id, 'status': 'already_processed', 'current_status': req.status})
            continue

        if action == 'reject':
            req.status = 'Rejected'
            req.remarks = remarks
            req.processed_at = now
            req.processed_by = admin_extrainfo
            req.save(update_fields=['status', 'remarks', 'processed_at', 'processed_by'])
            _phd_course_registration_notify(
                sender=request.user,
                recipient=req.student.id.user,
                verb='PhD course registration request rejected',
                description=f"Your registration request for {req.course.code} was rejected. "
                            f"Remarks: {remarks or '—'}",
            )
            results.append({'id': req_id, 'status': 'rejected'})
            continue

        # approve
        already_registered = course_registration.objects.filter(
            student_id=req.student,
            course_id=req.course,
            session=req.academic_year,
            semester_type=req.semester_type,
        ).exists()
        if already_registered:
            req.status = 'Rejected'
            req.remarks = 'Already registered for this course'
            req.processed_at = now
            req.processed_by = admin_extrainfo
            req.save(update_fields=['status', 'remarks', 'processed_at', 'processed_by'])
            _phd_course_registration_notify(
                sender=request.user,
                recipient=req.student.id.user,
                verb='PhD course registration request rejected',
                description=f"Your registration request for {req.course.code} was rejected: "
                            f"already registered for this course.",
            )
            results.append({'id': req_id, 'status': 'error', 'detail': 'Already registered'})
            continue

        # A BL request clears its source course, which for a stand-in is not the
        # course being registered; that source carries the grade and the
        # registration this one replaces.
        cleared_course = req.source_course or req.course

        # Registration type from the prior grade (same rule as the UG add approval).
        _sg = latest_grade(req.student.id.user.username, cleared_course)
        registration_type = registration_type_for_grade(_sg.grade if _sg else None)
        if registration_type == 'Regular' and req.registration_type:
            registration_type = req.registration_type

        # The prior registration of the cleared course (the low attempt) is the one being replaced.
        old_reg = course_registration.objects.filter(
            student_id=req.student, course_id=cleared_course,
        ).order_by('-working_year', '-semester_id__semester_no').first()

        try:
            # Nested atomic block (savepoint): a unique-constraint collision
            # here (e.g. two different course_slot requests for the same
            # actual course approved concurrently) must only roll back this
            # one item, not the whole batch's already-applied results.
            with transaction.atomic():
                new_reg = course_registration.objects.create(
                    student_id=req.student,
                    course_id=req.course,
                    course_slot_id=req.course_slot,
                    semester_id=req.semester,
                    session=req.academic_year,
                    semester_type=req.semester_type,
                    working_year=datetime.datetime.now().year,
                    registration_type=registration_type,
                )
                # Mirror UG: link the replaced (old) registration to the new one.
                if old_reg:
                    course_replacement.objects.create(
                        old_course_registration=old_reg,
                        new_course_registration=new_reg,
                    )
        except IntegrityError:
            req.status = 'Rejected'
            req.remarks = 'Already registered for this course'
            req.processed_at = now
            req.processed_by = admin_extrainfo
            req.save(update_fields=['status', 'remarks', 'processed_at', 'processed_by'])
            _phd_course_registration_notify(
                sender=request.user,
                recipient=req.student.id.user,
                verb='PhD course registration request rejected',
                description=f"Your registration request for {req.course.code} was rejected: "
                            f"already registered for this course.",
            )
            results.append({'id': req_id, 'status': 'error', 'detail': 'Already registered'})
            continue

        req.status = 'Approved'
        req.remarks = remarks
        req.processed_at = now
        req.processed_by = admin_extrainfo
        req.save(update_fields=['status', 'remarks', 'processed_at', 'processed_by'])
        _phd_course_registration_notify(
            sender=request.user,
            recipient=req.student.id.user,
            verb='PhD course registration request approved',
            description=f"Your registration request for {req.course.code} - "
                        f"{req.course.name} has been approved.",
        )
        results.append({'id': req_id, 'status': 'approved'})

    return JsonResponse({'results': results}, status=200)


# ===========================================================================
# Thesis Grade Evaluation
# ===========================================================================

def _eval_to_dict(ev):
    """Serialize a ThesisEvaluation block to a plain dict."""
    reg = ev.registration
    catalog_thesis = reg.thesis_slot.theses.first()
    evaluation_type = reg.thesis_slot.evaluation_type
    is_decimal = evaluation_type == 'decimal'

    score_inputs = getattr(ev, 'score_inputs', None) if is_decimal else None
    submission = None
    if is_decimal and reg.thesis_topic:
        submission = PGThesisSubmission.objects.filter(thesis=reg.thesis_topic).first()

    return {
        'id':             ev.id,
        'block_number':   ev.block_number,
        'total_blocks':   ev.total_blocks,
        'evaluation_type': evaluation_type,
        'grade':          ev.grade,
        'numeric_grade':  ev.numeric_grade,
        'supervisor_score': score_inputs.supervisor_score if score_inputs else None,
        'examiner_score': score_inputs.examiner_score if score_inputs else None,
        'synopsis_url':   submission.synopsis.url if submission and submission.synopsis else None,
        'thesis_report_url': submission.thesis_report.url if submission and submission.thesis_report else None,
        'remarks':        ev.remarks,
        'submitted_by':   ev.submitted_by.id.user.get_full_name() if ev.submitted_by else None,
        'submitted_at':   ev.submitted_at.isoformat() if ev.submitted_at else None,
        'verified':       ev.verified,
        'verified_at':    ev.verified_at.isoformat() if ev.verified_at else None,
        'announced':      ev.announced,
        'announced_at':   ev.announced_at.isoformat() if ev.announced_at else None,
        'registration': {
            'id':           reg.id,
            'credits':      reg.credits,
            'semester_no':  reg.semester.semester_no,
            'academic_session': reg.academic_session,
            'thesis_slot':  reg.thesis_slot.name,
            'thesis_code':  catalog_thesis.code if catalog_thesis else reg.thesis_slot.name,
            'thesis_title': reg.thesis_topic.research_theme if reg.thesis_topic else None,
            'programme_category': _student_programme_category(reg.student),
            'student': {
                'id':   reg.student.id.id,
                'name': reg.student.id.user.get_full_name(),
            },
        },
    }


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def supervisor_thesis_grades(request):
    """
    GET /supervisor/thesis-grades/
    Returns all ThesisEvaluation blocks for registrations where the
    student's thesis_topic.supervisor is the requesting faculty.
    Ordered by semester, then student name.
    Optional: ?semester=<no>  ?graded=true|false
    """
    user = request.user
    try:
        faculty = Faculty.objects.get(id__user=user)
    except Faculty.DoesNotExist:
        return JsonResponse({'error': 'Faculty record not found'}, status=404)

    # Thesis registrations where this faculty is the supervisor
    qs = ThesisEvaluation.objects.select_related(
        'registration__student__id__user',
        'registration__semester',
        'registration__thesis_slot',
        'registration__thesis_topic',
        'submitted_by__id__user',
        'score_inputs',
    ).filter(
        registration__status='verified',
        registration__thesis_topic__supervisor=faculty,
    ).order_by('registration__semester__semester_no', 'registration__student__id__user__last_name')

    # Filters
    sem_no = request.GET.get('semester')
    if sem_no:
        qs = qs.filter(registration__semester__semester_no=sem_no)

    # "Graded" means different things per evaluation_type: blocks_sx uses
    # `grade`, decimal uses the supervisor's raw score on ThesisEvaluationScore
    # (numeric_grade itself only appears once the examiner also scores).
    graded_param = request.GET.get('graded')
    not_graded_q = Q(registration__thesis_slot__evaluation_type='decimal', score_inputs__supervisor_score__isnull=True) | \
        Q(registration__thesis_slot__evaluation_type='blocks_sx', grade__isnull=True)
    if graded_param == 'false':
        qs = qs.filter(not_graded_q)
    elif graded_param == 'true':
        qs = qs.exclude(not_graded_q)

    return JsonResponse({'evaluations': [_eval_to_dict(ev) for ev in qs]}, status=200)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(['acadadmin'])
def admin_thesis_grades_list(request):
    """
    GET /acadadmin/thesis-grades/?semester=<no>&status=pending|verified|announced
    Lists all ThesisEvaluation blocks with optional filters.
    status filter: pending = grade submitted but not verified
                   verified = verified but not announced
                   announced = announced
                   ungraded = no grade yet

    Decimal-mode (PG final-thesis semester) evaluations never appear here --
    they skip admin verification entirely and go straight to the HOD/examiner
    panel flow once the supervisor scores them.
    """
    qs = ThesisEvaluation.objects.select_related(
        'registration__student__id__user',
        'registration__semester',
        'registration__thesis_slot',
        'submitted_by__id__user',
        'verified_by',
    ).exclude(
        registration__thesis_slot__evaluation_type='decimal',
    ).order_by('registration__semester__semester_no', 'registration__student__id__user__last_name', 'block_number')

    sem_no = request.GET.get('semester')
    if sem_no:
        qs = qs.filter(registration__semester__semester_no=sem_no)

    status_param = request.GET.get('status')
    if status_param == 'ungraded':
        qs = qs.filter(grade__isnull=True)
    elif status_param == 'pending':
        qs = qs.exclude(grade__isnull=True).filter(verified=False)
    elif status_param == 'verified':
        qs = qs.filter(verified=True, announced=False)
    elif status_param == 'announced':
        qs = qs.filter(announced=True)

    return JsonResponse({'evaluations': [_eval_to_dict(ev) for ev in qs]}, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['acadadmin'])
def admin_verify_thesis_grades(request):
    """
    POST /acadadmin/thesis-grades/verify/
    Body: { "ids": [1, 2, 3] }
    Verifies submitted grades (grade must already be set by supervisor).
    """
    ids = request.data.get('ids', [])
    if not ids:
        return JsonResponse({'error': 'No evaluation IDs provided'}, status=400)

    now = _dt.datetime.now(_dt.timezone.utc)
    count = 0
    for ev in ThesisEvaluation.objects.filter(id__in=ids, verified=False).exclude(grade=None):
        ev.verified    = True
        ev.verified_by = request.user
        ev.verified_at = now
        ev.save(update_fields=['verified', 'verified_by', 'verified_at'])
        count += 1
    return JsonResponse({'verified_count': count}, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['acadadmin'])
def admin_announce_thesis_grades(request):
    """
    POST /acadadmin/thesis-grades/announce/
    Body: { "ids": [1, 2, 3] }
    Announces grades — makes them visible to students.
    Only verified grades can be announced.
    """
    ids = request.data.get('ids', [])
    if not ids:
        return JsonResponse({'error': 'No evaluation IDs provided'}, status=400)

    now = _dt.datetime.now(_dt.timezone.utc)
    count = 0
    for ev in ThesisEvaluation.objects.filter(id__in=ids, verified=True, announced=False):
        ev.announced    = True
        ev.announced_at = now
        ev.save(update_fields=['announced', 'announced_at'])
        count += 1
    return JsonResponse({'announced_count': count}, status=200)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def supervisor_download_all_thesis_grades_template(request):
    """
    GET /supervisor/thesis-grades-all-template/
    Downloads Excel template with student name, roll number, and grade columns for ALL blocks.
    Pre-fills with all students who have ungraded evaluations across any block.
    """
    user = request.user

    try:
        # Get faculty record
        faculty = Faculty.objects.get(id__user=user)
    except Faculty.DoesNotExist:
        return JsonResponse({'error': 'Faculty record not found'}, status=404)

    # Fetch all ungraded evaluations for this supervisor across all blocks
    try:
        evals = ThesisEvaluation.objects.select_related(
            'registration__student__id'
        ).filter(
            registration__thesis_topic__supervisor=faculty,
            registration__status='verified',
            grade__isnull=True
        ).exclude(
            registration__thesis_slot__evaluation_type='decimal',
        ).order_by('registration__student__id__id', 'block_number')

        if not evals.exists():
            return JsonResponse({'error': 'No ungraded evaluations found'}, status=400)

        # Group by student to get unique students and their blocks
        from collections import defaultdict
        student_blocks = defaultdict(lambda: {'name': '', 'blocks': {}})

        for eval in evals:
            student = eval.registration.student
            roll_no = student.id.id

            if roll_no not in student_blocks:
                student_blocks[roll_no]['name'] = student.id.user.get_full_name()

            student_blocks[roll_no]['blocks'][eval.block_number] = eval.id

        # Determine all blocks present
        all_blocks = set()
        for student_data in student_blocks.values():
            all_blocks.update(student_data['blocks'].keys())
        all_blocks = sorted(list(all_blocks))

        # Generate Excel template
        import openpyxl
        output = BytesIO()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'All Grades'

        # Headers: Name, Roll Number, Grade 1, Grade 2, ..., Remarks
        headers = ['Student Name', 'Roll Number']
        headers.extend([f'Grade {b}' for b in all_blocks])
        headers.append('Remarks')

        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

        # Add student data
        for row, (roll_no, student_data) in enumerate(sorted(student_blocks.items()), 2):
            try:
                worksheet.cell(row=row, column=1, value=student_data['name'])
                worksheet.cell(row=row, column=2, value=roll_no)
                # Columns 3+ are grades for each block (leave empty for supervisor to fill)
                # Last column is remarks (leave empty)
            except Exception as e:
                logger.error(f"Error writing row for {roll_no}: {str(e)}", exc_info=True)

        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Thesis_Grades_All_{_dt.datetime.now().strftime("%Y%m%d")}.xlsx"'
        return response

    except Exception as e:
        return JsonResponse({'error': f'Failed to generate template: {str(e)}'}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def supervisor_upload_all_thesis_grades(request):
    """
    POST /supervisor/thesis-grades-all/upload/
    Uploads and validates Excel file with grades for multiple blocks.
    Expected columns: Name, Roll Number, Grade 1, Grade 2, ..., Remarks
    Returns valid and invalid rows.
    """
    user = request.user
    uploaded_file = request.FILES.get('file')

    if not uploaded_file:
        return JsonResponse({'error': 'file is required'}, status=400)

    try:
        faculty = Faculty.objects.get(id__user=user)
    except Faculty.DoesNotExist:
        return JsonResponse({'error': 'Faculty record not found'}, status=404)

    # Parse Excel file
    try:
        df = pd.read_excel(uploaded_file, engine='openpyxl')
    except Exception:
        try:
            df = pd.read_excel(uploaded_file, engine='xlrd')
        except Exception as e:
            return JsonResponse({'error': f'Failed to read Excel file: {str(e)}'}, status=400)

    # Normalize column names
    df.columns = [col.strip().lower() for col in df.columns]

    # Find roll number and remarks columns
    roll_col = None
    remarks_col = None
    grade_cols = {}  # {block_number: column_name}

    for col in df.columns:
        if 'roll' in col and not roll_col:
            roll_col = col
        elif 'remark' in col and not remarks_col:
            remarks_col = col
        elif 'grade' in col:
            # Extract the grade number from "grade N" or similar
            match = re.search(r'grade\s*(\d+)', col)
            if match:
                block_num = int(match.group(1))
                grade_cols[block_num] = col

    if not roll_col:
        return JsonResponse({'error': 'Excel must contain "Roll Number" column'}, status=400)
    if not grade_cols:
        return JsonResponse({'error': 'Excel must contain at least one "Grade N" column'}, status=400)

    # Fetch all evaluations for this supervisor grouped by student and block
    evals = ThesisEvaluation.objects.select_related(
        'registration__student__id'
    ).filter(
        registration__thesis_topic__supervisor=faculty,
        registration__status='verified',
        grade__isnull=True
    ).exclude(
        registration__thesis_slot__evaluation_type='decimal',
    )

    # Create lookup: {roll_no: {block_num: eval_id}}
    eval_lookup = defaultdict(dict)
    for eval in evals:
        roll_no = eval.registration.student.id.id
        eval_lookup[roll_no][eval.block_number] = eval.id

    valid_rows = []
    invalid_rows = []

    # Validate each row
    for idx, row in df.iterrows():
        roll_no = str(row[roll_col]).strip() if pd.notna(row[roll_col]) else None
        remarks = str(row[remarks_col]).strip() if remarks_col and pd.notna(row[remarks_col]) else ''
        row_errors = []

        if not roll_no:
            row_errors.append('Roll number is required')
            invalid_rows.append({
                'row_num': idx + 2,
                'roll_no': 'N/A',
                'errors': row_errors
            })
            continue

        if roll_no not in eval_lookup:
            invalid_rows.append({
                'row_num': idx + 2,
                'roll_no': roll_no,
                'errors': ['No student found with this roll number']
            })
            continue

        # Validate grades for each block
        row_submissions = []
        for block_num, grade_col in grade_cols.items():
            grade = str(row[grade_col]).strip().upper() if pd.notna(row[grade_col]) else ''

            # Grade is optional if student doesn't have evaluation for that block
            if not grade:
                if block_num in eval_lookup[roll_no]:
                    row_errors.append(f'Grade {block_num} is required for this student')
                continue

            # If grade provided, validate it
            if grade not in ('S', 'X'):
                row_errors.append(f'Grade {block_num} must be S or X, got {grade}')
                continue

            # Check if evaluation exists for this student and block
            if block_num not in eval_lookup[roll_no]:
                row_errors.append(f'No evaluation found for Grade {block_num}')
                continue

            row_submissions.append({
                'evaluation_id': eval_lookup[roll_no][block_num],
                'block_number': block_num,
                'grade': grade,
                'remarks': remarks
            })

        if row_errors:
            invalid_rows.append({
                'row_num': idx + 2,
                'roll_no': roll_no,
                'errors': row_errors
            })
        elif row_submissions:
            valid_rows.extend(row_submissions)

    return JsonResponse({
        'valid_rows': valid_rows,
        'invalid_rows': invalid_rows
    }, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def supervisor_bulk_submit_all_thesis_grades(request):
    """
    POST /supervisor/thesis-grades-all/bulk-submit/
    Submits multiple grades across multiple blocks in one request.
    Body: { "submissions": [{"evaluation_id": 123, "grade": "S", "remarks": "..."}, ...] }
    """
    user = request.user
    submissions = request.data.get('submissions', [])

    if not submissions:
        return JsonResponse({'error': 'submissions list is required'}, status=400)

    try:
        faculty = Faculty.objects.get(id__user=user)
    except Faculty.DoesNotExist:
        return JsonResponse({'error': 'Faculty record not found'}, status=404)

    now = _dt.datetime.now(_dt.timezone.utc)

    # Batch fetch all evaluations
    eval_ids = [sub.get('evaluation_id') for sub in submissions if sub.get('evaluation_id')]
    evaluations_dict = ThesisEvaluation.objects.select_related(
        'registration__thesis_topic', 'registration__thesis_slot',
    ).filter(id__in=eval_ids).in_bulk(field_name='id')

    success_count = 0
    errors = []
    evaluations_to_update = []

    # Process each submission
    for idx, submission in enumerate(submissions):
        eval_id = submission.get('evaluation_id')
        grade = submission.get('grade', '').upper()
        remarks = submission.get('remarks', '')

        try:
            if not eval_id:
                errors.append({'index': idx, 'error': 'evaluation_id is required'})
                continue
            if grade not in ('S', 'X'):
                errors.append({'index': idx, 'evaluation_id': eval_id, 'error': 'grade must be S or X'})
                continue

            if eval_id not in evaluations_dict:
                errors.append({'index': idx, 'evaluation_id': eval_id, 'error': 'Evaluation not found'})
                continue

            evaluation = evaluations_dict[eval_id]

            # Verify ownership and permissions
            if evaluation.registration.thesis_topic.supervisor != faculty:
                errors.append({'index': idx, 'evaluation_id': eval_id, 'error': 'Not authorized for this evaluation'})
                continue

            if evaluation.registration.thesis_slot.evaluation_type == 'decimal':
                errors.append({'index': idx, 'evaluation_id': eval_id, 'error': 'This is a decimal-mode thesis and cannot take an S/X grade'})
                continue

            if evaluation.verified or evaluation.announced:
                errors.append({'index': idx, 'evaluation_id': eval_id, 'error': 'Grade already verified/announced; cannot be changed'})
                continue

            # Update evaluation
            evaluation.grade = grade
            evaluation.remarks = remarks
            evaluation.submitted_by = faculty
            evaluation.submitted_at = now

            evaluations_to_update.append(evaluation)
            success_count += 1

        except Exception as e:
            errors.append({'index': idx, 'evaluation_id': eval_id, 'error': str(e)})

    # Batch update all at once
    if evaluations_to_update:
        ThesisEvaluation.objects.bulk_update(
            evaluations_to_update,
            fields=['grade', 'remarks', 'submitted_by', 'submitted_at'],
            batch_size=500
        )

    return JsonResponse({
        'success_count': success_count,
        'error_count': len(errors),
        'errors': errors if errors else None
    }, status=200)


# ===========================================================================
# Comprehensive Examination
# ===========================================================================
# Workflow: Supervisor proposes eligibility -> Academic Office verifies ->
# Convener DPGC (HOD of the student's department) approves -> attempt 1 is
# auto-created (no committee to propose -- the student's existing RPC,
# fetched live via their ThesisTopic, doubles as the examination committee)
# -> RPC collectively records the result + qualitative comments, each member
# consenting like Progress Seminar (any panel edit resets everyone else's
# consent) -> Convener PGCS (also HOD) reviews the finalized result: reject
# sends it back to the RPC for fresh consensus, approve forwards to Dean
# Academic -> Dean Academic gives a forward-only final approval, closing the
# attempt as passed/failed. On failure with attempts remaining (max
# ComprehensiveExam.MAX_ATTEMPTS), the next attempt auto-creates starting
# directly at RPC review -- the Academic Office/DPGC eligibility gate is
# one-time on the exam as a whole, not per-attempt.

def _exam_rpc_committee(student):
    """The student's RPC (Progress Seminar committee, via their most recent
    ThesisTopic) -- reused as the Comprehensive Exam examination committee.
    Read-only here; RPC membership itself is managed via the Progress
    Seminar flow (supervisor_review_api)."""
    thesis_topic = ThesisTopic.objects.filter(student=student).order_by('-created_at').first()
    if not thesis_topic:
        return CommitteeMember.objects.none()
    return CommitteeMember.objects.filter(thesis=thesis_topic).select_related(
        'member__id__user', 'member__id__department'
    )


def _rpc_committee_users(student):
    """Users for the student's live RPC committee -- for notifying committee members."""
    return User.objects.filter(
        pk__in=_exam_rpc_committee(student).values_list('member__id__user', flat=True)
    ).distinct()


def _is_thesis_supervisor_or_co(faculty, student):
    """Whether `faculty` is the supervisor/co-supervisor on the student's
    most recent ThesisTopic -- used to gate Comprehensive Exam / Open
    Seminar proposal so an unrelated faculty member can't self-assign as
    supervisor for someone else's student."""
    thesis_topic = ThesisTopic.objects.filter(student=student).order_by('-created_at').first()
    if not thesis_topic:
        return False
    return faculty.pk in (thesis_topic.supervisor_id, thesis_topic.co_supervisor_id)


def _comprehensive_exam_attempt_to_dict(a):
    return {
        'id': a.id,
        'attempt_number': a.attempt_number,
        'status': a.status,
        'exam_date': a.exam_date.isoformat() if a.exam_date else None,
        'result': a.result,
        'fundamentals_comment': a.fundamentals_comment,
        'problem_identification_comment': a.problem_identification_comment,
        'plan_of_work_comment': a.plan_of_work_comment,
        'suggestions_comment': a.suggestions_comment,
        'additional_literature_comment': a.additional_literature_comment,
        'milestone_plan_url': a.milestone_plan_upload.url if a.milestone_plan_upload else None,
        'reported_at': a.reported_at.isoformat() if a.reported_at else None,
        'pgcs_remarks': a.pgcs_remarks,
        'pgcs_reviewed_at': a.pgcs_reviewed_at.isoformat() if a.pgcs_reviewed_at else None,
        'dean_approved_at': a.dean_approved_at.isoformat() if a.dean_approved_at else None,
        'consented_count': a.consents.filter(
            consented=True,
            member_id__in=_exam_rpc_committee(a.exam.student).values_list('member_id', flat=True),
        ).count(),
        'committee_size': _exam_rpc_committee(a.exam.student).count(),
    }


def _is_exam_supervisor_or_co(request, exam):
    """True if request.user is the exam's supervisor or co-supervisor.

    Mirrors the ownership check in supervisor_assign (ThesisSubmission flow):
    compare Django auth User pks directly instead of going through Faculty,
    which avoids Faculty.id resolving to the related ExtraInfo object rather
    than its raw pk.
    """
    allowed_users = {exam.supervisor.id.user_id}
    if exam.co_supervisor:
        allowed_users.add(exam.co_supervisor.id.user_id)
    return request.user.id in allowed_users


def _can_set_exam_date(request, attempt):
    """Supervisor/co-supervisor or any RPC member may set/update the exam date."""
    return _is_exam_supervisor_co_or_rpc_member(request, attempt.exam)


def _is_exam_supervisor_co_or_rpc_member(request, exam):
    """Supervisor/co-supervisor or any RPC committee member -- the Manage
    Comprehensive Examination modal (supervisor_comprehensive_exam_detail) is
    shared by both audiences (RPC-only members reach it via the "RPC Member"
    tab's Consent button, per SupervisorComprehensiveExamDashboard)."""
    if _is_exam_supervisor_or_co(request, exam):
        return True
    try:
        faculty = Faculty.objects.get(id__user=request.user)
    except Faculty.DoesNotExist:
        return False
    return _exam_rpc_committee(exam.student).filter(member=faculty).exists()


def _student_completed_credits(student):
    """Sum of credits for regular courses the student has a passing grade for.
    Coursework only -- callers needing PhD Thesis/Progress Seminar/Teaching
    Credit too should add those separately (see _student_total_credits_completed)
    or, like _compute_open_seminar_eligibility, report them as their own fields.

    Reads Student_grades (the table the current React grade-submission flow
    and transcript both actually use), not SemesterMarks -- that table is
    only ever written by an old, disconnected legacy Django-template grading
    view (academic_procedures/views.py::course_marks_data) and is empty for
    students graded through the current flow.
    """
    total = Student_grades.objects.filter(roll_no=student.id_id).exclude(
        grade__isnull=True
    ).exclude(grade__in=['F', 'X']).aggregate(total=Sum('course_id__credit'))['total']
    return total or 0


def _progress_seminar_credits_completed(student):
    """Sum of each of the student's rpc_approved ProgressSeminarEntry records at its
    own catalog credit value (see resolve_progress_seminar_credit -- do not hardcode
    this number, it varies by the Seminar catalog row)."""
    thesis_topic = ThesisTopic.objects.filter(student=student).order_by('-created_at').first()
    if not thesis_topic:
        return 0
    approved_seminars = thesis_topic.seminars.filter(status='rpc_approved')
    return sum(
        resolve_progress_seminar_credit(student, s.semester) for s in approved_seminars
    )


def _thesis_research_credits_completed(student):
    """Registered credits (ThesisRegistration.credits) are just what the student
    signed up for -- actual earned credit only counts once a block is graded
    Satisfactory and the result announced (each block = 3 credits)."""
    return ThesisEvaluation.objects.filter(
        registration__student=student, grade='S', announced=True,
    ).count() * 3


def _teaching_credits_completed(student):
    """Registering for teaching credit (TeachingCreditRegistration) or being
    allocated a course (TeachingCreditAllocation) isn't earning the credit --
    only a semester marked 'completed'/'satisfactory' counts, at that
    semester's own catalog credit value."""
    return sum(
        resolve_teaching_credit_credit(student, alloc.semester)
        for alloc in TeachingCreditAllocation.objects.filter(
            student=student, status='completed', result='satisfactory',
        )
    )


def _student_total_credits_completed(student):
    """Full credits-completed figure: coursework plus PhD Thesis/Progress
    Seminar/Teaching Credit -- the same four numbers _compute_open_seminar_eligibility
    reports separately (to avoid double-counting there), summed into one total for
    eligibility checks (e.g. Comprehensive Exam) that just need a single figure."""
    return (
        _student_completed_credits(student)
        + _progress_seminar_credits_completed(student)
        + _thesis_research_credits_completed(student)
        + _teaching_credits_completed(student)
    )


def _student_current_cpi(student):
    """Live CPI as of now, via the same canonical calculate_cpi_for_student() the
    transcript uses. Student.cpi itself is a dead field -- nothing in the codebase
    ever writes to it after grades are submitted -- so it can't be trusted here."""
    cpi, _, _ = calculate_cpi_for_student(student, student.curr_semester_no, None)
    return cpi


def comprehensive_exam_to_dict(exam):
    """Serialize a ComprehensiveExam (with RPC committee & attempts) for JSON responses."""
    return {
        'id': exam.id,
        'student_roll': exam.student.id.id,
        'student_name': exam.student.id.user.get_full_name(),
        'student_discipline': exam.student.specialization,
        'semester_no': exam.student.curr_semester_no,
        'supervisor': {
            'id': exam.supervisor.id.id,
            'name': str(exam.supervisor),
            'discipline': exam.supervisor.id.department.name if exam.supervisor.id.department else '',
        },
        'co_supervisor': (
            {
                'id': exam.co_supervisor.id.id,
                'name': str(exam.co_supervisor),
                'discipline': exam.co_supervisor.id.department.name if exam.co_supervisor.id.department else '',
            }
            if exam.co_supervisor else None
        ),
        'possible_thesis_title': exam.possible_thesis_title,
        # A freshly-`.create()`d instance holds whatever raw value was passed
        # in (e.g. a plain date string) until reloaded from the DB, so this
        # can't assume `.isoformat()` is always safe to call.
        'proposed_exam_date': (
            exam.proposed_exam_date.isoformat()
            if hasattr(exam.proposed_exam_date, 'isoformat')
            else exam.proposed_exam_date
        ),
        'entry_qualification': exam.entry_qualification,
        'required_credits': exam.required_credits,
        'credits_completed': exam.credits_completed,
        'current_cpi': str(exam.current_cpi) if exam.current_cpi is not None else None,
        'research_methodology_completed': exam.research_methodology_completed,
        'credits_verified': exam.credits_verified,
        'cpi_verified': exam.cpi_verified,
        'research_methodology_verified': exam.research_methodology_verified,
        'academic_office_remarks': exam.academic_office_remarks,
        'dpgc_remarks': exam.dpgc_remarks,
        'status': exam.status,
        'current_attempt_number': exam.current_attempt_number,
        'max_attempts': ComprehensiveExam.MAX_ATTEMPTS,
        'committee': [
            {
                'id': cm.member.id.id,
                'name': str(cm.member),
                'discipline': cm.member.id.department.name if cm.member.id.department else '',
            }
            for cm in _exam_rpc_committee(exam.student)
        ],
        'attempts': [_comprehensive_exam_attempt_to_dict(a) for a in exam.attempts.order_by('attempt_number')],
    }


# 1. Student

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_comprehensive_exam_api(request):
    """GET /stu/comprehensive-exam/ -> fetch the requesting student's exam ({} if none)."""
    try:
        student = Student.objects.get(id=request.user.extrainfo)
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student record not found'}, status=404)

    exam = ComprehensiveExam.objects.filter(student=student).first()
    return JsonResponse(comprehensive_exam_to_dict(exam) if exam else {}, status=200)


# 2. Supervisor

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def supervisor_comprehensive_exam_dashboard(request):
    """GET /supervisor/comprehensive-exam/dashboard/ -> exams supervised or co-supervised by the requester."""
    try:
        faculty = Faculty.objects.get(id__user=request.user)
    except Faculty.DoesNotExist:
        return JsonResponse({'error': 'Faculty record not found'}, status=404)

    qs = ComprehensiveExam.objects.filter(
        Q(supervisor=faculty) | Q(co_supervisor=faculty)
    ).select_related('student__id__user', 'supervisor__id__user').prefetch_related('attempts')

    return JsonResponse({'exams': [comprehensive_exam_to_dict(e) for e in qs]}, status=200)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def supervisor_student_academic_info(request, roll_no):
    """
    GET /supervisor/comprehensive-exam/student-info/<roll_no>/
    Read-only credits-completed & CPI, computed from the student's own
    academic records -- never manually entered.
    """
    if getattr(getattr(request.user, 'extrainfo', None), 'user_type', None) != 'faculty':
        return JsonResponse({'error': 'Only faculty can view student academic info.'}, status=403)
    try:
        student = Student.objects.get(id=roll_no)
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student not found'}, status=404)

    return JsonResponse({
        'credits_completed': _student_total_credits_completed(student),
        'current_cpi': str(_student_current_cpi(student)),
    }, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def supervisor_propose_comprehensive_exam(request):
    """
    POST /supervisor/comprehensive-exam/propose/
    Body: { roll_no, co_supervisor_id, possible_thesis_title, entry_qualification, proposed_exam_date }
    credits_completed / current_cpi are computed server-side from the
    student's own records, not accepted from the client. Research Methodology
    completion is Academic Office's call (set via the verify endpoint), not
    the supervisor's -- not accepted here either. No committee is proposed --
    the student's existing RPC (see _exam_rpc_committee) doubles as the
    examination committee.
    """
    try:
        faculty = Faculty.objects.get(id__user=request.user)
    except Faculty.DoesNotExist:
        return JsonResponse({'error': 'Faculty record not found'}, status=404)

    data = request.data
    roll_no = data.get('roll_no')
    if not roll_no:
        return JsonResponse({'error': 'roll_no is required'}, status=400)

    try:
        student = Student.objects.get(id=roll_no)
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student not found'}, status=404)

    if not _is_thesis_supervisor_or_co(faculty, student):
        return JsonResponse({'error': 'You are not this student\'s supervisor or co-supervisor'}, status=403)

    if ComprehensiveExam.objects.filter(student=student).exists():
        return JsonResponse({'error': 'Comprehensive exam already exists for this student'}, status=400)

    entry_qualification = data.get('entry_qualification')
    if entry_qualification not in dict(ComprehensiveExam.ENTRY_QUALIFICATION_CHOICES):
        return JsonResponse({'error': 'Invalid entry_qualification'}, status=400)

    exam = ComprehensiveExam.objects.create(
        student=student,
        supervisor=faculty,
        co_supervisor_id=data.get('co_supervisor_id') or None,
        possible_thesis_title=data.get('possible_thesis_title', ''),
        proposed_exam_date=data.get('proposed_exam_date') or None,
        entry_qualification=entry_qualification,
        credits_completed=_student_total_credits_completed(student),
        current_cpi=_student_current_cpi(student),
    )

    _comprehensive_exam_notify(
        sender=request.user,
        recipient=_academic_office_users(),
        verb='Comprehensive Exam proposal pending verification',
        description=f"{student.id.user.get_full_name()}'s Comprehensive Exam proposal is "
                    f"awaiting Academic Office eligibility verification.",
    )

    return JsonResponse(comprehensive_exam_to_dict(exam), status=201)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def supervisor_comprehensive_exam_detail(request, pk):
    """GET /supervisor/comprehensive-exam/<pk>/ -> full detail (also used to prefill a
    resubmission, and shared with the RPC Member "Consent" view -- see
    _is_exam_supervisor_co_or_rpc_member)."""
    exam = get_object_or_404(ComprehensiveExam, pk=pk)
    if not _is_exam_supervisor_co_or_rpc_member(request, exam):
        return JsonResponse({'error': 'Not authorized'}, status=403)
    return JsonResponse(comprehensive_exam_to_dict(exam), status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def supervisor_resubmit_proposal(request, pk):
    """
    POST /supervisor/comprehensive-exam/<pk>/resubmit/
    Edits eligibility fields after an Academic Office or Convener (DPGC)
    rejection, and resends for Academic Office verification.
    """
    exam = get_object_or_404(ComprehensiveExam, pk=pk)
    if not _is_exam_supervisor_or_co(request, exam):
        return JsonResponse({'error': 'Not authorized'}, status=403)
    if exam.status not in ('academic_office_rejected', 'dpgc_rejected'):
        return JsonResponse({'error': 'Cannot edit at this stage'}, status=403)

    data = request.data
    if 'possible_thesis_title' in data:
        exam.possible_thesis_title = data['possible_thesis_title']
    if 'proposed_exam_date' in data:
        exam.proposed_exam_date = data['proposed_exam_date'] or None
    if 'entry_qualification' in data:
        exam.entry_qualification = data['entry_qualification']
    if 'co_supervisor_id' in data:
        exam.co_supervisor_id = data['co_supervisor_id'] or None

    # Re-derive from the student's own records rather than trusting client input.
    exam.credits_completed = _student_total_credits_completed(exam.student)
    exam.current_cpi = _student_current_cpi(exam.student)

    exam.status = 'academic_office_pending'
    exam.credits_verified = False
    exam.cpi_verified = False
    exam.research_methodology_verified = False
    exam.academic_office_remarks = ''
    exam.dpgc_remarks = ''
    exam.save()

    return JsonResponse(comprehensive_exam_to_dict(exam), status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def supervisor_set_exam_date(request, attempt_pk):
    """
    POST /supervisor/comprehensive-exam/attempt/<attempt_pk>/set-exam-date/
    Body: { exam_date }
    Settable by the supervisor/co-supervisor or any RPC member, any time
    before Dean Academic's final approval -- including while the RPC is
    still finalizing their report.
    """
    attempt = get_object_or_404(ComprehensiveExamAttempt, pk=attempt_pk)
    if attempt.status in ('passed', 'failed'):
        return JsonResponse({'error': 'Attempt is already closed'}, status=403)
    if not _can_set_exam_date(request, attempt):
        return JsonResponse({'error': 'Not authorized'}, status=403)

    exam_date = request.data.get('exam_date')
    if not exam_date:
        return JsonResponse({'error': 'exam_date is required'}, status=400)
    attempt.exam_date = exam_date
    attempt.save()

    return JsonResponse(comprehensive_exam_to_dict(attempt.exam), status=200)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_courses_for_dropdown(request):
    """
    GET /courses/dropdown/?search=<text>
    Lightweight {id, code, name} course list for populating dropdowns (used
    by Teaching Credit's course-choice pickers). Deliberately not
    acadadmin-gated -- faculty need this too.
    """
    qs = Courses.objects.filter(working_course=True, latest_version=True)
    search = request.GET.get('search', '').strip()
    if search:
        qs = qs.filter(Q(code__icontains=search) | Q(name__icontains=search))
    qs = qs.order_by('code')[:100]
    return JsonResponse({
        'courses': [{'id': c.id, 'code': c.code, 'name': c.name} for c in qs],
    }, status=200)


# 3. Academic Office (acadadmin)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(['acadadmin'])
def academic_office_comprehensive_exam_list(request):
    """GET /acadadmin/comprehensive-exam/?status=<status>"""
    qs = ComprehensiveExam.objects.select_related('student__id__user', 'supervisor__id__user').prefetch_related(
        'attempts'
    ).all()
    status_param = request.GET.get('status')
    if status_param:
        qs = qs.filter(status=status_param)
    exams = list(qs)
    # Still-pending exams may have been proposed before the student earned more
    # credits / their CPI changed -- refresh from current records (and persist,
    # so what's shown here matches exactly what the verify action will check)
    # rather than displaying whatever was snapshotted at proposal time.
    for exam in exams:
        if exam.status == 'academic_office_pending':
            exam.credits_completed = _student_total_credits_completed(exam.student)
            exam.current_cpi = _student_current_cpi(exam.student)
            exam.save(update_fields=['credits_completed', 'current_cpi'])
    return JsonResponse({'exams': [comprehensive_exam_to_dict(e) for e in exams]}, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['acadadmin'])
def academic_office_verify_comprehensive_exam(request, pk):
    """
    POST /acadadmin/comprehensive-exam/<pk>/verify/
    Body: { approve: true|false, credits_verified, cpi_verified,
            research_methodology_verified, remarks }
    """
    exam = get_object_or_404(ComprehensiveExam, pk=pk, status='academic_office_pending')
    data = request.data

    # Refresh from the student's current records rather than trusting whatever was
    # snapshotted at proposal time -- the student may have earned more credits
    # (or their CPI may have changed) since then.
    exam.credits_completed = _student_total_credits_completed(exam.student)
    exam.current_cpi = _student_current_cpi(exam.student)

    exam.credits_verified = bool(data.get('credits_verified', False))
    exam.cpi_verified = bool(data.get('cpi_verified', False))
    exam.research_methodology_verified = bool(data.get('research_methodology_verified', False))
    exam.academic_office_remarks = data.get('remarks', '')
    exam.academic_office_verified_by = request.user
    exam.academic_office_verified_at = timezone.now()

    if data.get('approve'):
        if not (exam.credits_verified and exam.cpi_verified and exam.research_methodology_verified):
            return JsonResponse({
                'error': 'All three eligibility checks (credits, CPI, Research Methodology) '
                         'must be confirmed before approving.',
            }, status=400)
        if exam.credits_completed < exam.required_credits:
            return JsonResponse({
                'error': f'Credits completed ({exam.credits_completed}) is below the '
                         f'{exam.required_credits} required for this entry qualification.',
            }, status=400)
        if exam.current_cpi is None or exam.current_cpi < ComprehensiveExam.MIN_CPI:
            return JsonResponse({
                'error': f'Current CPI ({exam.current_cpi}) is below the required minimum '
                         f'of {ComprehensiveExam.MIN_CPI}.',
            }, status=400)
        exam.status = 'dpgc_pending'
        exam.save()
        student = exam.student
        acronym = student.batch_id.discipline.acronym if student.batch_id and student.batch_id.discipline else None
        _comprehensive_exam_notify(
            sender=request.user,
            recipient=_hod_users_for_discipline(acronym),
            verb='Comprehensive Exam pending your (DPGC) approval',
            description=f"{student.id.user.get_full_name()}'s Comprehensive Exam has been "
                        f"verified by Academic Office and is awaiting Convener (DPGC) approval.",
        )
        _comprehensive_exam_notify(
            sender=request.user,
            recipient=exam.supervisor.id.user,
            verb='Comprehensive Exam verified by Academic Office',
            description=f"{student.id.user.get_full_name()}'s Comprehensive Exam has been "
                        f"verified by Academic Office and forwarded to Convener (DPGC).",
        )
    else:
        exam.status = 'academic_office_rejected'
        exam.save()
        _comprehensive_exam_notify(
            sender=request.user,
            recipient=exam.supervisor.id.user,
            verb='Comprehensive Exam rejected by Academic Office',
            description=f"Academic Office rejected {exam.student.id.user.get_full_name()}'s "
                        f"Comprehensive Exam proposal. Remarks: {exam.academic_office_remarks or '—'}",
        )

    return JsonResponse(comprehensive_exam_to_dict(exam), status=200)


# 4. Convener DPGC (HOD of the student's department stands in)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def hod_dpgc_comprehensive_exam_dashboard(request):
    """GET /hod/comprehensive-exam/dpgc-dashboard/ -> exams pending DPGC approval,
    plus a history of already-decided ones, scoped to the HOD's own discipline."""
    hod_disciplines = get_hod_disciplines(request.user)

    def _scoped(qs):
        result = []
        for exam in qs:
            student = exam.student
            acronym = student.batch_id.discipline.acronym if student.batch_id and student.batch_id.discipline else None
            if acronym and acronym in hod_disciplines:
                result.append(exam)
        return result

    pending_qs = ComprehensiveExam.objects.filter(status='dpgc_pending').select_related(
        'student__id__user', 'student__batch_id__discipline', 'supervisor__id__user'
    ).prefetch_related('attempts')
    pending = [comprehensive_exam_to_dict(e) for e in _scoped(pending_qs)]

    history_qs = ComprehensiveExam.objects.filter(dpgc_by__isnull=False).select_related(
        'student__id__user', 'student__batch_id__discipline', 'supervisor__id__user', 'dpgc_by',
    ).prefetch_related('attempts').order_by('-dpgc_at')
    history = [
        {
            **comprehensive_exam_to_dict(e),
            'decision': 'Rejected' if e.status == 'dpgc_rejected' else 'Approved',
            'decided_by': e.dpgc_by.get_full_name() if e.dpgc_by else None,
            'decided_at': e.dpgc_at.isoformat() if e.dpgc_at else None,
            'remarks': e.dpgc_remarks,
        }
        for e in _scoped(history_qs)
    ]

    return JsonResponse({'pending': pending, 'history': history}, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def hod_dpgc_approve_comprehensive_exam(request, pk):
    """
    POST /hod/comprehensive-exam/<pk>/dpgc-approve/
    Body: { approve: true|false, remarks }
    Approving auto-creates attempt 1, starting directly at RPC review --
    there is no committee to propose, the student's RPC is fetched live.
    """
    exam = get_object_or_404(ComprehensiveExam, pk=pk, status='dpgc_pending')
    student = exam.student
    acronym = student.batch_id.discipline.acronym if student.batch_id and student.batch_id.discipline else None
    if not is_hod_of_discipline(request.user, acronym):
        return JsonResponse({'error': 'Not authorized'}, status=403)

    data = request.data
    exam.dpgc_remarks = data.get('remarks', '')
    exam.dpgc_by = request.user
    exam.dpgc_at = timezone.now()

    if data.get('approve'):
        exam.status = 'in_progress'
        exam.save()
        ComprehensiveExamAttempt.objects.get_or_create(
            exam=exam, attempt_number=exam.current_attempt_number,
            defaults={'exam_date': exam.proposed_exam_date},
        )
        _comprehensive_exam_notify(
            sender=request.user,
            recipient=_rpc_committee_users(student),
            verb='Comprehensive Exam pending your review',
            description=f"{student.id.user.get_full_name()}'s Comprehensive Exam has been "
                        f"approved by Convener (DPGC) and is awaiting RPC review.",
        )
        _comprehensive_exam_notify(
            sender=request.user,
            recipient=exam.supervisor.id.user,
            verb='Comprehensive Exam approved by Convener (DPGC)',
            description=f"{student.id.user.get_full_name()}'s Comprehensive Exam has been "
                        f"approved by Convener (DPGC) and forwarded to the RPC.",
        )
    else:
        exam.status = 'dpgc_rejected'
        exam.save()
        _comprehensive_exam_notify(
            sender=request.user,
            recipient=exam.supervisor.id.user,
            verb='Comprehensive Exam rejected by Convener (DPGC)',
            description=f"Convener (DPGC) rejected {student.id.user.get_full_name()}'s "
                        f"Comprehensive Exam. Remarks: {exam.dpgc_remarks or '—'}",
        )

    return JsonResponse(comprehensive_exam_to_dict(exam), status=200)


# 5. RPC (the student's existing committee, fetched live)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def rpc_comprehensive_exam_list(request):
    """GET /faculty/comprehensive-exam/rpc/ -> attempts where the requester is an RPC member."""
    faculty = get_object_or_404(Faculty, id__user=request.user)

    thesis_ids = ThesisTopic.objects.filter(committee__member=faculty).values_list('id', flat=True)
    student_ids = ThesisTopic.objects.filter(id__in=thesis_ids).values_list('student_id', flat=True)

    qs = ComprehensiveExamAttempt.objects.filter(exam__student_id__in=student_ids).select_related(
        'exam__student__id__user', 'exam__supervisor__id__user'
    ).distinct()

    def serialize(attempts):
        return [
            {
                **_comprehensive_exam_attempt_to_dict(a),
                'exam_id': a.exam.id,
                'student_roll': a.exam.student.id.id,
                'student_name': a.exam.student.id.user.get_full_name(),
                'my_consent_given': ComprehensiveExamConsent.objects.filter(
                    attempt=a, member=faculty, consented=True
                ).exists(),
            }
            for a in attempts
        ]

    return JsonResponse({
        'pending': serialize(qs.filter(status='rpc_pending')),
        'history': serialize(qs.exclude(status='rpc_pending')),
    }, status=200)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def rpc_comprehensive_exam_detail(request, attempt_pk):
    """GET /faculty/comprehensive-exam/rpc/<attempt_pk>/"""
    faculty = get_object_or_404(Faculty, id__user=request.user)
    attempt = get_object_or_404(ComprehensiveExamAttempt, pk=attempt_pk)
    if not _exam_rpc_committee(attempt.exam.student).filter(member=faculty).exists():
        return JsonResponse({'error': 'Not on committee'}, status=403)

    committee = []
    for cm in _exam_rpc_committee(attempt.exam.student):
        fac = cm.member
        extra = fac.id
        consented = ComprehensiveExamConsent.objects.filter(attempt=attempt, member=fac, consented=True).exists()
        committee.append({
            'id': extra.id,
            'name': f"{extra.user.first_name} {extra.user.last_name}",
            'discipline': extra.department.name if extra.department else '',
            'consented': consented,
        })

    comments = [
        {
            'member': c.member.id.user.get_full_name(),
            'text': c.text,
            'timestamp': c.timestamp.isoformat(),
        }
        for c in attempt.rpc_comments.all()
    ]

    my_comment = ComprehensiveExamRPCComment.objects.filter(attempt=attempt, member=faculty).first()
    is_consented = ComprehensiveExamConsent.objects.filter(attempt=attempt, member=faculty, consented=True).exists()

    exam = attempt.exam
    payload = {
        **_comprehensive_exam_attempt_to_dict(attempt),
        'exam_id': exam.id,
        'student_name': exam.student.id.user.get_full_name(),
        'student_roll': exam.student.id.id,
        'student_discipline': exam.student.specialization,
        'possible_thesis_title': exam.possible_thesis_title,
        'supervisor': {
            'id': exam.supervisor.id.id,
            'name': str(exam.supervisor),
        },
        'co_supervisor': (
            {'id': exam.co_supervisor.id.id, 'name': str(exam.co_supervisor)}
            if exam.co_supervisor else None
        ),
        'committee': committee,
        'committee_size': len(committee),
        'consented_count': sum(1 for m in committee if m['consented']),
        'comments': comments,
        'my_comment': my_comment.text if my_comment else '',
        'is_consented': is_consented,
    }
    return JsonResponse(payload, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def rpc_comprehensive_exam_consent(request, attempt_pk):
    """
    POST /faculty/comprehensive-exam/rpc/<attempt_pk>/consent/
    Body: { result, fundamentals_comment, problem_identification_comment,
            plan_of_work_comment, suggestions_comment,
            additional_literature_comment, exam_date, comment, milestone_plan (file) }
    Accepts either JSON (no file) or multipart (to attach milestone_plan) --
    DRF's default parsers handle both, unlike the file-only endpoints
    elsewhere in this app that pin MultiPartParser/FormParser explicitly.
    Any edit to the shared panel resets everyone else's consent -- mirrors
    Progress Seminar's rpc_consent.
    """
    faculty = get_object_or_404(Faculty, id__user=request.user)
    attempt = get_object_or_404(ComprehensiveExamAttempt, pk=attempt_pk, status='rpc_pending')
    if not _exam_rpc_committee(attempt.exam.student).filter(member=faculty).exists():
        return JsonResponse({'error': 'Not on committee'}, status=403)

    data = request.data
    if data.get('result') and data['result'] not in dict(ComprehensiveExamAttempt.RESULT_CHOICES):
        return JsonResponse({'error': 'Invalid result value'}, status=400)

    panel_fields = [
        'result', 'fundamentals_comment', 'problem_identification_comment',
        'plan_of_work_comment', 'suggestions_comment', 'additional_literature_comment',
    ]

    changed = any(
        field in data and getattr(attempt, field) != data[field]
        for field in panel_fields
    )
    if data.get('exam_date'):
        old_exam_date = attempt.exam_date.isoformat() if attempt.exam_date else None
        if data['exam_date'] != old_exam_date:
            changed = True
    if request.FILES.get('milestone_plan'):
        changed = True
    if changed:
        ComprehensiveExamConsent.objects.filter(attempt=attempt).update(consented=False)

    for field in panel_fields:
        if field in data:
            setattr(attempt, field, data[field])
    if data.get('exam_date'):
        attempt.exam_date = data['exam_date']
    if request.FILES.get('milestone_plan'):
        attempt.milestone_plan_upload = request.FILES['milestone_plan']
    attempt.save()

    if 'comment' in data:
        ComprehensiveExamRPCComment.objects.update_or_create(
            attempt=attempt, member=faculty, defaults={'text': data['comment']},
        )

    consent_obj, _created = ComprehensiveExamConsent.objects.get_or_create(attempt=attempt, member=faculty)
    consent_obj.consented = True
    consent_obj.save()

    return JsonResponse({'message': 'Consent & data recorded.'}, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def rpc_comprehensive_exam_finalize(request, attempt_pk):
    """
    POST /faculty/comprehensive-exam/rpc/<attempt_pk>/finalize/
    Requires every RPC member to have consented and a result to have been
    recorded; forwards to Convener PGCS.
    """
    faculty = get_object_or_404(Faculty, id__user=request.user)
    attempt = get_object_or_404(ComprehensiveExamAttempt, pk=attempt_pk, status='rpc_pending')
    if not _exam_rpc_committee(attempt.exam.student).filter(member=faculty).exists():
        return JsonResponse({'error': 'Not on committee'}, status=403)

    if not attempt.result:
        return JsonResponse({'error': 'Record a result before finalizing'}, status=400)

    current_committee_ids = _exam_rpc_committee(attempt.exam.student).values_list('member_id', flat=True)
    total = len(current_committee_ids)
    yes = ComprehensiveExamConsent.objects.filter(
        attempt=attempt, consented=True, member_id__in=current_committee_ids,
    ).count()
    if total == 0 or yes < total:
        return JsonResponse({'error': 'Not all RPC members have consented'}, status=400)

    attempt.status = 'pgcs_pending'
    attempt.reported_by = request.user
    attempt.reported_at = timezone.now()
    # Starting a fresh PGCS review cycle -- an earlier rejection's
    # reviewed_by/at/remarks no longer apply and would otherwise make this
    # attempt look like already-decided history while it's still pending.
    attempt.pgcs_reviewed_by = None
    attempt.pgcs_reviewed_at = None
    attempt.pgcs_remarks = ''
    attempt.save()

    exam = attempt.exam
    student = exam.student
    acronym = student.batch_id.discipline.acronym if student.batch_id and student.batch_id.discipline else None
    _comprehensive_exam_notify(
        sender=request.user,
        recipient=_hod_users_for_discipline(acronym),
        verb='Comprehensive Exam pending your (PGCS) review',
        description=f"The RPC has finalized {student.id.user.get_full_name()}'s Comprehensive "
                    f"Exam (attempt {attempt.attempt_number}); awaiting Convener (PGCS) review.",
    )
    _comprehensive_exam_notify(
        sender=request.user,
        recipient=exam.supervisor.id.user,
        verb='Comprehensive Exam RPC review finalized',
        description=f"The RPC has finalized {student.id.user.get_full_name()}'s Comprehensive "
                    f"Exam (attempt {attempt.attempt_number}); forwarded to Convener (PGCS).",
    )

    return JsonResponse(comprehensive_exam_to_dict(attempt.exam), status=200)


# 6. Convener PGCS (HOD of the student's department stands in)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def hod_pgcs_comprehensive_exam_dashboard(request):
    """GET /hod/comprehensive-exam/pgcs-dashboard/ -> attempts pending PGCS review,
    plus a history of already-decided ones, scoped to the HOD's own discipline."""
    hod_disciplines = get_hod_disciplines(request.user)

    def _scoped(qs):
        result = []
        for attempt in qs:
            student = attempt.exam.student
            acronym = student.batch_id.discipline.acronym if student.batch_id and student.batch_id.discipline else None
            if acronym and acronym in hod_disciplines:
                result.append(attempt)
        return result

    pending_qs = ComprehensiveExamAttempt.objects.filter(status='pgcs_pending').select_related(
        'exam__student__id__user', 'exam__student__batch_id__discipline', 'exam__supervisor__id__user'
    ).prefetch_related('exam__attempts')
    pending = [comprehensive_exam_to_dict(a.exam) for a in _scoped(pending_qs)]

    history_qs = ComprehensiveExamAttempt.objects.filter(pgcs_reviewed_by__isnull=False).select_related(
        'exam__student__id__user', 'exam__student__batch_id__discipline', 'exam__supervisor__id__user',
        'pgcs_reviewed_by',
    ).prefetch_related('exam__attempts').order_by('-pgcs_reviewed_at')
    history = [
        {
            **comprehensive_exam_to_dict(a.exam),
            'decision': 'Rejected' if a.status == 'rpc_pending' else 'Approved',
            'decided_by': a.pgcs_reviewed_by.get_full_name() if a.pgcs_reviewed_by else None,
            'decided_at': a.pgcs_reviewed_at.isoformat() if a.pgcs_reviewed_at else None,
            'attempt_number': a.attempt_number,
            'remarks': a.pgcs_remarks,
        }
        for a in _scoped(history_qs)
    ]

    return JsonResponse({'pending': pending, 'history': history}, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def hod_pgcs_review_comprehensive_exam(request, attempt_pk):
    """
    POST /hod/comprehensive-exam/attempt/<attempt_pk>/pgcs-review/
    Body: { approve: true|false, remarks }
    Rejecting sends it back to the RPC for fresh consensus (all consents reset).
    """
    attempt = get_object_or_404(ComprehensiveExamAttempt, pk=attempt_pk, status='pgcs_pending')
    student = attempt.exam.student
    acronym = student.batch_id.discipline.acronym if student.batch_id and student.batch_id.discipline else None
    if not is_hod_of_discipline(request.user, acronym):
        return JsonResponse({'error': 'Not authorized'}, status=403)

    data = request.data
    attempt.pgcs_reviewed_by = request.user
    attempt.pgcs_reviewed_at = timezone.now()
    if data.get('approve'):
        attempt.pgcs_remarks = ''
        attempt.status = 'dean_pending'
        attempt.save()
        _comprehensive_exam_notify(
            sender=request.user,
            recipient=_dean_academic_users(),
            verb='Comprehensive Exam pending your final approval',
            description=f"{student.id.user.get_full_name()}'s Comprehensive Exam has been "
                        f"approved by Convener (PGCS) and is awaiting your final approval.",
        )
        _comprehensive_exam_notify(
            sender=request.user,
            recipient=attempt.exam.supervisor.id.user,
            verb='Comprehensive Exam approved by Convener (PGCS)',
            description=f"{student.id.user.get_full_name()}'s Comprehensive Exam has been "
                        f"approved by Convener (PGCS) and forwarded to Dean Academic.",
        )
    else:
        attempt.pgcs_remarks = data.get('remarks', '')
        attempt.status = 'rpc_pending'
        attempt.save()
        ComprehensiveExamConsent.objects.filter(attempt=attempt).update(consented=False)
        _comprehensive_exam_notify(
            sender=request.user,
            recipient=_rpc_committee_users(student),
            verb='Comprehensive Exam sent back by Convener (PGCS)',
            description=f"Convener (PGCS) sent {student.id.user.get_full_name()}'s Comprehensive "
                        f"Exam back to the RPC for fresh consensus. Remarks: {attempt.pgcs_remarks or '—'}",
        )

    return JsonResponse(comprehensive_exam_to_dict(attempt.exam), status=200)


# 7. Dean Academic (forward-only final approval)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(['Dean Academic'])
def dean_comprehensive_exam_dashboard(request):
    """GET /dean/comprehensive-exam/dashboard/ -> attempts approved by PGCS, pending final approval."""
    qs = ComprehensiveExamAttempt.objects.filter(status='dean_pending').select_related(
        'exam__student__id__user', 'exam__supervisor__id__user'
    ).prefetch_related('exam__attempts')

    return JsonResponse({
        'pending': [comprehensive_exam_to_dict(a.exam) for a in qs],
    }, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['Dean Academic'])
def dean_approve_comprehensive_exam(request, attempt_pk):
    """
    POST /dean/comprehensive-exam/attempt/<attempt_pk>/approve/
    Forward-only -- closes the attempt as passed/failed (whichever the RPC
    already decided). On failure with attempts remaining, auto-creates the
    next attempt starting directly at RPC review.
    """
    attempt = get_object_or_404(ComprehensiveExamAttempt, pk=attempt_pk, status='dean_pending')
    exam = attempt.exam

    attempt.dean_approved_by = request.user
    attempt.dean_approved_at = timezone.now()
    attempt.status = attempt.result
    attempt.save()

    if attempt.result == 'passed':
        exam.status = 'passed'
        exam.save()
        result_desc = "passed the Comprehensive Exam."
    elif exam.current_attempt_number < ComprehensiveExam.MAX_ATTEMPTS:
        exam.current_attempt_number += 1
        exam.save()
        ComprehensiveExamAttempt.objects.get_or_create(exam=exam, attempt_number=exam.current_attempt_number)
        result_desc = (f"not cleared attempt {attempt.attempt_number} of the Comprehensive Exam. "
                       f"A new attempt has been created, starting at RPC review.")
    else:
        exam.status = 'failed_final'
        exam.save()
        result_desc = "failed the Comprehensive Exam with all attempts exhausted."

    student = exam.student
    acronym = student.batch_id.discipline.acronym if student.batch_id and student.batch_id.discipline else None
    for recipient in [student.id.user, exam.supervisor.id.user]:
        _comprehensive_exam_notify(
            sender=request.user,
            recipient=recipient,
            verb='Comprehensive Exam result declared',
            description=f"{student.id.user.get_full_name()} has {result_desc}",
        )
    _comprehensive_exam_notify(
        sender=request.user,
        recipient=_hod_users_for_discipline(acronym),
        verb='Comprehensive Exam result declared',
        description=f"{student.id.user.get_full_name()} has {result_desc}",
    )

    return JsonResponse(comprehensive_exam_to_dict(exam), status=200)


# ===========================================================================
# Open Seminar
# ===========================================================================
# Workflow: Supervisor proposes eligibility -> Convener DPGC (HOD of the
# student's department) reviews -> Dean Academic appoints the Dean Nominee
# and approves -> attempt 1 auto-creates, starting directly at RPC review
# (no committee to propose -- the student's existing RPC, fetched live via
# their ThesisTopic, doubles as the examination committee) -> RPC
# collectively records the result + comments, each member consenting like
# Comprehensive Exam/Progress Seminar -> Convener DPGC reviews the finalized
# result a second time: reject sends it back to the RPC for fresh
# consensus, approve forwards to Dean Academic -> Dean Academic's dashboard
# shows the committee's verdict together with the Dean Nominee's
# confidential report, and gives a forward-only final approval, closing the
# attempt as satisfactory/not_satisfactory. On not_satisfactory, the next
# attempt auto-creates starting directly at RPC review -- the Convener/Dean
# early gate (and Dean Nominee appointment) is one-time on the OpenSeminar
# as a whole, not per-attempt.

def _is_open_seminar_supervisor_or_co(request, seminar):
    """Mirrors _is_exam_supervisor_or_co (Comprehensive Exam) for OpenSeminar."""
    allowed_users = {seminar.supervisor.id.user_id}
    if seminar.co_supervisor:
        allowed_users.add(seminar.co_supervisor.id.user_id)
    return request.user.id in allowed_users


def _can_set_seminar_date(request, attempt):
    """Supervisor/co-supervisor or any RPC member may set/update the seminar date."""
    return _is_seminar_supervisor_co_or_rpc_member(request, attempt.open_seminar)


def _is_seminar_supervisor_co_or_rpc_member(request, seminar):
    """Supervisor/co-supervisor or any RPC committee member -- the Manage Open
    Seminar modal (supervisor_open_seminar_detail) is shared by both audiences
    (RPC-only members reach it via the "RPC Member" tab's Consent button,
    mirroring the Comprehensive Exam flow's _is_exam_supervisor_co_or_rpc_member)."""
    if _is_open_seminar_supervisor_or_co(request, seminar):
        return True
    try:
        faculty = Faculty.objects.get(id__user=request.user)
    except Faculty.DoesNotExist:
        return False
    return _exam_rpc_committee(seminar.student).filter(member=faculty).exists()


def _compute_open_seminar_eligibility(student):
    """Auto-derive the Constitution form's credit breakdown + RPC recommendation.

    The four credit numbers are reported separately here (not summed into one
    total) -- see _student_completed_credits / _progress_seminar_credits_completed /
    _thesis_research_credits_completed / _teaching_credits_completed for what each
    one means. rpc_recommended_open_seminar reads the latest approved seminar's
    rec_open field.
    """
    course_work_credits = _student_completed_credits(student)
    progress_seminar_credits = _progress_seminar_credits_completed(student)
    thesis_research_credits = _thesis_research_credits_completed(student)
    teaching_credits = _teaching_credits_completed(student)

    thesis_topic = ThesisTopic.objects.filter(student=student).order_by('-created_at').first()
    rpc_recommended_open_seminar = False
    if thesis_topic:
        latest_approved = thesis_topic.seminars.filter(
            status='rpc_approved'
        ).order_by('-version').first()
        if latest_approved:
            rpc_recommended_open_seminar = (latest_approved.rec_open == 'Yes')

    return {
        'course_work_credits': course_work_credits,
        'progress_seminar_credits': progress_seminar_credits,
        'thesis_research_credits': thesis_research_credits,
        'teaching_credits': teaching_credits,
        'semesters_completed': student.curr_semester_no or 0,
        'rpc_recommended_open_seminar': rpc_recommended_open_seminar,
    }


def _open_seminar_attempt_to_dict(a, include_confidential=False):
    d = {
        'id': a.id,
        'attempt_number': a.attempt_number,
        'status': a.status,
        'seminar_date': a.seminar_date.isoformat() if a.seminar_date else None,
        'result': a.result,
        'committee_comments': a.committee_comments,
        'reported_at': a.reported_at.isoformat() if a.reported_at else None,
        'hod_review_remarks': a.hod_review_remarks,
        'hod_reviewed_at': a.hod_reviewed_at.isoformat() if a.hod_reviewed_at else None,
        'dean_approved_at': a.dean_approved_at.isoformat() if a.dean_approved_at else None,
        'dean_nominee': (
            {'id': a.dean_nominee.id.id, 'name': str(a.dean_nominee)}
            if a.dean_nominee else None
        ),
        'dn_submitted_at': a.dn_submitted_at.isoformat() if a.dn_submitted_at else None,
        'consented_count': a.consents.filter(
            consented=True,
            member_id__in=_exam_rpc_committee(a.open_seminar.student).values_list('member_id', flat=True),
        ).count(),
        'committee_size': _exam_rpc_committee(a.open_seminar.student).count(),
        'rpc_comments': [
            {
                'member': c.member.id.user.get_full_name(),
                'text': c.text,
                'timestamp': c.timestamp.isoformat(),
            }
            for c in a.rpc_comments.all()
        ],
    }
    if include_confidential:
        d.update({
            'dn_quality': a.dn_quality,
            'dn_quantity': a.dn_quantity,
            'dn_publications': a.dn_publications,
            'dn_overall': a.dn_overall,
            'dn_comments': a.dn_comments,
        })
    return d


def open_seminar_to_dict(seminar, include_confidential=False):
    """Serialize an OpenSeminar (with RPC committee & attempts). Confidential
    Dean-Nominee fields are only included for Dean/Dean-Nominee-facing
    endpoints."""
    return {
        'id': seminar.id,
        'student_roll': seminar.student.id.id,
        'student_name': seminar.student.id.user.get_full_name(),
        'student_discipline': seminar.student.specialization,
        'semester_no': seminar.student.curr_semester_no,
        'supervisor': {
            'id': seminar.supervisor.id.id,
            'name': str(seminar.supervisor),
            'discipline': seminar.supervisor.id.department.name if seminar.supervisor.id.department else '',
        },
        'co_supervisor': (
            {'id': seminar.co_supervisor.id.id, 'name': str(seminar.co_supervisor)}
            if seminar.co_supervisor else None
        ),
        'possible_thesis_title': seminar.possible_thesis_title,
        # A freshly-`.create()`d instance holds whatever raw value was passed
        # in (e.g. a plain date string) until reloaded from the DB, so this
        # can't assume `.isoformat()` is always safe to call.
        'proposed_date': (
            seminar.proposed_date.isoformat()
            if hasattr(seminar.proposed_date, 'isoformat')
            else seminar.proposed_date
        ),
        'course_work_credits': seminar.course_work_credits,
        'progress_seminar_credits': seminar.progress_seminar_credits,
        'thesis_research_credits': seminar.thesis_research_credits,
        'teaching_credits': seminar.teaching_credits,
        'total_credits': seminar.total_credits,
        'semesters_completed': seminar.semesters_completed,
        'rpc_recommended_open_seminar': seminar.rpc_recommended_open_seminar,
        'first_draft_document_url': seminar.first_draft_document.url if seminar.first_draft_document else None,
        'hod_remarks': seminar.hod_remarks,
        'dean_remarks': seminar.dean_remarks,
        'status': seminar.status,
        'current_attempt_number': seminar.current_attempt_number,
        'committee': [
            {
                'id': cm.member.id.id,
                'name': str(cm.member),
                'discipline': cm.member.id.department.name if cm.member.id.department else '',
            }
            for cm in _exam_rpc_committee(seminar.student)
        ],
        'attempts': [
            _open_seminar_attempt_to_dict(a, include_confidential)
            for a in seminar.attempts.order_by('attempt_number')
        ],
    }


# 0. Shared

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def open_seminar_eligibility_preview(request, roll_no):
    """
    GET /supervisor/open-seminar/eligibility/<roll_no>/
    Read-only preview of the auto-computed credit breakdown + RPC
    recommendation, so the supervisor can see them before proposing --
    never manually entered.
    """
    try:
        student = Student.objects.get(id=roll_no)
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student not found'}, status=404)

    return JsonResponse(_compute_open_seminar_eligibility(student), status=200)


# 1. Student

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_open_seminar_api(request):
    """GET /stu/open-seminar/ -> fetch the requesting student's Open Seminar ({} if none)."""
    try:
        student = Student.objects.get(id=request.user.extrainfo)
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student record not found'}, status=404)

    seminar = OpenSeminar.objects.filter(student=student).first()
    return JsonResponse(open_seminar_to_dict(seminar) if seminar else {}, status=200)


# 2. Supervisor

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def supervisor_open_seminar_dashboard(request):
    """GET /supervisor/open-seminar/dashboard/ -> Open Seminars supervised or co-supervised by the requester."""
    try:
        faculty = Faculty.objects.get(id__user=request.user)
    except Faculty.DoesNotExist:
        return JsonResponse({'error': 'Faculty record not found'}, status=404)

    qs = OpenSeminar.objects.filter(
        Q(supervisor=faculty) | Q(co_supervisor=faculty)
    ).select_related('student__id__user', 'supervisor__id__user').prefetch_related('attempts')

    return JsonResponse({'seminars': [open_seminar_to_dict(s) for s in qs]}, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def supervisor_propose_open_seminar(request):
    """
    POST /supervisor/open-seminar/propose/
    Body: { roll_no, possible_thesis_title, co_supervisor_id, proposed_date,
            first_draft_document (file) }
    course_work/progress_seminar/thesis_research/teaching credits, semesters_completed,
    and rpc_recommended_open_seminar are computed server-side. No committee
    is proposed -- the student's existing RPC (see _exam_rpc_committee)
    doubles as the examination committee.
    """
    try:
        faculty = Faculty.objects.get(id__user=request.user)
    except Faculty.DoesNotExist:
        return JsonResponse({'error': 'Faculty record not found'}, status=404)

    data = request.data
    roll_no = data.get('roll_no')
    if not roll_no:
        return JsonResponse({'error': 'roll_no is required'}, status=400)

    try:
        student = Student.objects.get(id=roll_no)
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student not found'}, status=404)

    if not _is_thesis_supervisor_or_co(faculty, student):
        return JsonResponse({'error': 'You are not this student\'s supervisor or co-supervisor'}, status=403)

    if OpenSeminar.objects.filter(student=student).exists():
        return JsonResponse({'error': 'Open Seminar already exists for this student'}, status=400)

    if not ComprehensiveExam.objects.filter(student=student, status='passed').exists():
        return JsonResponse(
            {'error': 'Comprehensive Examination must be passed before proposing Open Seminar.'},
            status=403,
        )

    eligibility = _compute_open_seminar_eligibility(student)
    if not eligibility['rpc_recommended_open_seminar']:
        return JsonResponse(
            {'error': 'The RPC has not recommended this student for Open Seminar yet.'},
            status=403,
        )

    seminar = OpenSeminar.objects.create(
        student=student,
        supervisor=faculty,
        co_supervisor_id=data.get('co_supervisor_id') or None,
        possible_thesis_title=data.get('possible_thesis_title', ''),
        proposed_date=data.get('proposed_date') or None,
        first_draft_document=request.FILES.get('first_draft_document'),
        **eligibility,
    )

    acronym = student.batch_id.discipline.acronym if student.batch_id and student.batch_id.discipline else None
    _open_seminar_notify(
        sender=request.user,
        recipient=_hod_users_for_discipline(acronym),
        verb='Open Seminar constitution pending your review',
        description=f"{student.id.user.get_full_name()}'s Open Seminar has been proposed and "
                    f"is awaiting Convener (DPGC) review.",
    )

    return JsonResponse(open_seminar_to_dict(seminar), status=201)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def supervisor_open_seminar_detail(request, pk):
    """GET /supervisor/open-seminar/<pk>/ -> full detail (also used to prefill a
    resubmission, and shared with the RPC Member "Consent" view -- see
    _is_seminar_supervisor_co_or_rpc_member)."""
    seminar = get_object_or_404(OpenSeminar, pk=pk)
    if not _is_seminar_supervisor_co_or_rpc_member(request, seminar):
        return JsonResponse({'error': 'Not authorized'}, status=403)
    return JsonResponse(open_seminar_to_dict(seminar), status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def supervisor_resubmit_open_seminar(request, pk):
    """
    POST /supervisor/open-seminar/<pk>/resubmit/
    Edits eligibility fields after a Convener (DPGC) or Dean Academic
    rejection, and resends for Convener (DPGC) review.
    """
    seminar = get_object_or_404(OpenSeminar, pk=pk)
    if not _is_open_seminar_supervisor_or_co(request, seminar):
        return JsonResponse({'error': 'Not authorized'}, status=403)
    if seminar.status not in ('hod_rejected', 'dean_rejected'):
        return JsonResponse({'error': 'Cannot edit at this stage'}, status=403)

    data = request.data
    if 'possible_thesis_title' in data:
        seminar.possible_thesis_title = data['possible_thesis_title']
    if 'co_supervisor_id' in data:
        seminar.co_supervisor_id = data['co_supervisor_id'] or None
    if 'proposed_date' in data:
        seminar.proposed_date = data['proposed_date'] or None
    if request.FILES.get('first_draft_document'):
        seminar.first_draft_document = request.FILES['first_draft_document']

    # Re-derive from the student's own records rather than trusting client input.
    eligibility = _compute_open_seminar_eligibility(seminar.student)
    if not eligibility['rpc_recommended_open_seminar']:
        return JsonResponse(
            {'error': 'The RPC has not recommended this student for Open Seminar yet.'},
            status=403,
        )
    for field, value in eligibility.items():
        setattr(seminar, field, value)

    seminar.status = 'hod_pending'
    seminar.hod_remarks = ''
    seminar.dean_remarks = ''
    seminar.save()

    return JsonResponse(open_seminar_to_dict(seminar), status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def supervisor_set_seminar_date(request, attempt_pk):
    """
    POST /supervisor/open-seminar/attempt/<attempt_pk>/set-seminar-date/
    Body: { seminar_date }
    Settable by the supervisor/co-supervisor or any RPC member, any time
    before Dean Academic's final approval -- including while the RPC is
    still finalizing their report.
    """
    attempt = get_object_or_404(OpenSeminarAttempt, pk=attempt_pk)
    if attempt.status in ('satisfactory', 'not_satisfactory'):
        return JsonResponse({'error': 'Attempt is already closed'}, status=403)
    if not _can_set_seminar_date(request, attempt):
        return JsonResponse({'error': 'Not authorized'}, status=403)

    seminar_date = request.data.get('seminar_date')
    if not seminar_date:
        return JsonResponse({'error': 'seminar_date is required'}, status=400)
    attempt.seminar_date = seminar_date
    attempt.save()

    return JsonResponse(open_seminar_to_dict(attempt.open_seminar), status=200)


# 3. Convener DPGC, early review (HOD of the student's department stands in)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def hod_dpgc_open_seminar_dashboard(request):
    """GET /hod/open-seminar/dpgc-dashboard/ -> seminars pending DPGC review,
    plus a history of already-decided ones, scoped to the HOD's own discipline."""
    hod_disciplines = get_hod_disciplines(request.user)

    def _scoped(qs):
        result = []
        for seminar in qs:
            student = seminar.student
            acronym = student.batch_id.discipline.acronym if student.batch_id and student.batch_id.discipline else None
            if acronym and acronym in hod_disciplines:
                result.append(seminar)
        return result

    pending_qs = OpenSeminar.objects.filter(status='hod_pending').select_related(
        'student__id__user', 'student__batch_id__discipline', 'supervisor__id__user'
    )
    pending = [open_seminar_to_dict(s) for s in _scoped(pending_qs)]

    history_qs = OpenSeminar.objects.filter(hod_by__isnull=False).select_related(
        'student__id__user', 'student__batch_id__discipline', 'supervisor__id__user', 'hod_by',
    ).order_by('-hod_at')
    history = [
        {
            **open_seminar_to_dict(s),
            'decision': 'Rejected' if s.status == 'hod_rejected' else 'Approved',
            'decided_by': s.hod_by.get_full_name() if s.hod_by else None,
            'decided_at': s.hod_at.isoformat() if s.hod_at else None,
            'remarks': s.hod_remarks,
        }
        for s in _scoped(history_qs)
    ]

    return JsonResponse({'pending': pending, 'history': history}, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def hod_dpgc_review_open_seminar(request, pk):
    """
    POST /hod/open-seminar/<pk>/dpgc-review/
    Body: { approve: true|false, remarks }
    """
    seminar = get_object_or_404(OpenSeminar, pk=pk, status='hod_pending')
    student = seminar.student
    acronym = student.batch_id.discipline.acronym if student.batch_id and student.batch_id.discipline else None
    if not is_hod_of_discipline(request.user, acronym):
        return JsonResponse({'error': 'Not authorized'}, status=403)

    data = request.data
    seminar.hod_remarks = data.get('remarks', '')
    seminar.hod_by = request.user
    seminar.hod_at = timezone.now()
    seminar.status = 'dean_pending' if data.get('approve') else 'hod_rejected'
    seminar.save()

    if data.get('approve'):
        _open_seminar_notify(
            sender=request.user,
            recipient=_dean_academic_users(),
            verb='Open Seminar pending nominee appointment',
            description=f"{student.id.user.get_full_name()}'s Open Seminar has been approved by "
                        f"Convener (DPGC) and needs a Dean Nominee appointed.",
        )
        _open_seminar_notify(
            sender=request.user,
            recipient=seminar.supervisor.id.user,
            verb='Open Seminar approved by Convener (DPGC)',
            description=f"{student.id.user.get_full_name()}'s Open Seminar has been approved by "
                        f"Convener (DPGC) and forwarded to Dean Academic.",
        )
    else:
        _open_seminar_notify(
            sender=request.user,
            recipient=seminar.supervisor.id.user,
            verb='Open Seminar rejected by Convener (DPGC)',
            description=f"Convener (DPGC) rejected {student.id.user.get_full_name()}'s Open "
                        f"Seminar. Remarks: {seminar.hod_remarks or '—'}",
        )

    return JsonResponse(open_seminar_to_dict(seminar), status=200)


# 4. Dean Academic (appoints the Dean Nominee early; forward-only final approval)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(['Dean Academic'])
def dean_open_seminar_dashboard(request):
    """
    GET /dean/open-seminar/dashboard/
    -> pending nominee appointments + pending final approvals (the latter
    include the committee's verdict together with the Dean Nominee's
    confidential report, shown side by side).
    """
    pending_appointment = OpenSeminar.objects.filter(status='dean_pending').select_related(
        'student__id__user', 'supervisor__id__user'
    ).prefetch_related('attempts')
    pending_final = OpenSeminarAttempt.objects.filter(status='dean_pending').select_related(
        'open_seminar__student__id__user', 'open_seminar__supervisor__id__user'
    ).prefetch_related('open_seminar__attempts')

    return JsonResponse({
        'pending_appointment': [open_seminar_to_dict(s) for s in pending_appointment],
        'pending_final': [open_seminar_to_dict(a.open_seminar, include_confidential=True) for a in pending_final],
    }, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['Dean Academic'])
def dean_appoint_nominee_open_seminar(request, pk):
    """
    POST /dean/open-seminar/<pk>/appoint-nominee/
    Body: { approve: true|false, dean_nominee_id, remarks }
    Approving requires appointing a Dean Nominee and auto-creates attempt 1,
    starting directly at RPC review -- there is no committee to propose,
    the student's RPC is fetched live.
    """
    seminar = get_object_or_404(OpenSeminar, pk=pk, status='dean_pending')
    data = request.data

    seminar.dean_remarks = data.get('remarks', '')
    seminar.dean_by = request.user
    seminar.dean_at = timezone.now()

    if data.get('approve'):
        dean_nominee_id = data.get('dean_nominee_id')
        if not dean_nominee_id:
            return JsonResponse({'error': 'A Dean Nominee must be appointed to approve.'}, status=400)

        conflicted_ids = {seminar.supervisor_id}
        if seminar.co_supervisor_id:
            conflicted_ids.add(seminar.co_supervisor_id)
        conflicted_ids.update(_exam_rpc_committee(seminar.student).values_list('member_id', flat=True))
        if dean_nominee_id in conflicted_ids:
            return JsonResponse({
                'error': 'The Dean Nominee must be independent of the student\'s supervisor, '
                         'co-supervisor, and RPC committee.',
            }, status=400)

        seminar.status = 'in_progress'
        seminar.save()
        OpenSeminarAttempt.objects.get_or_create(
            open_seminar=seminar, attempt_number=seminar.current_attempt_number,
            defaults={'seminar_date': seminar.proposed_date, 'dean_nominee_id': dean_nominee_id},
        )
        student = seminar.student
        nominee = Faculty.objects.filter(pk=dean_nominee_id).select_related('id__user').first()
        if nominee:
            _open_seminar_notify(
                sender=request.user,
                recipient=nominee.id.user,
                verb='You have been appointed Dean Nominee',
                description=f"You have been appointed Dean Nominee for "
                            f"{student.id.user.get_full_name()}'s Open Seminar.",
            )
        _open_seminar_notify(
            sender=request.user,
            recipient=_rpc_committee_users(student),
            verb='Open Seminar pending your review',
            description=f"{student.id.user.get_full_name()}'s Open Seminar has been approved by "
                        f"Dean Academic and is awaiting RPC review.",
        )
        _open_seminar_notify(
            sender=request.user,
            recipient=seminar.supervisor.id.user,
            verb='Open Seminar approved by Dean Academic',
            description=f"{student.id.user.get_full_name()}'s Open Seminar has been approved by "
                        f"Dean Academic and forwarded to the RPC.",
        )
    else:
        seminar.status = 'dean_rejected'
        seminar.save()
        _open_seminar_notify(
            sender=request.user,
            recipient=seminar.supervisor.id.user,
            verb='Open Seminar rejected by Dean Academic',
            description=f"Dean Academic rejected {seminar.student.id.user.get_full_name()}'s "
                        f"Open Seminar. Remarks: {seminar.dean_remarks or '—'}",
        )

    return JsonResponse(open_seminar_to_dict(seminar), status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@role_required(['Dean Academic'])
def dean_approve_open_seminar(request, attempt_pk):
    """
    POST /dean/open-seminar/attempt/<attempt_pk>/approve/
    Forward-only -- closes the attempt as satisfactory/not_satisfactory
    (whichever the RPC already decided). On not_satisfactory, auto-creates
    the next attempt starting directly at RPC review (no new Dean Nominee).
    """
    attempt = get_object_or_404(OpenSeminarAttempt, pk=attempt_pk, status='dean_pending')
    seminar = attempt.open_seminar

    attempt.dean_approved_by = request.user
    attempt.dean_approved_at = timezone.now()
    attempt.status = attempt.result
    attempt.save()

    if attempt.result == 'satisfactory':
        seminar.status = 'satisfactory'
        seminar.save()
        result_desc = "completed the Open Seminar satisfactorily."
    else:
        next_number = seminar.current_attempt_number + 1
        seminar.current_attempt_number = next_number
        seminar.save()
        OpenSeminarAttempt.objects.get_or_create(open_seminar=seminar, attempt_number=next_number)
        result_desc = (f"not cleared attempt {attempt.attempt_number} of the Open Seminar. "
                       f"A new attempt has been created, starting at RPC review.")

    student = seminar.student
    acronym = student.batch_id.discipline.acronym if student.batch_id and student.batch_id.discipline else None
    for recipient in [student.id.user, seminar.supervisor.id.user]:
        _open_seminar_notify(
            sender=request.user,
            recipient=recipient,
            verb='Open Seminar result declared',
            description=f"{student.id.user.get_full_name()} has {result_desc}",
        )
    _open_seminar_notify(
        sender=request.user,
        recipient=_hod_users_for_discipline(acronym),
        verb='Open Seminar result declared',
        description=f"{student.id.user.get_full_name()} has {result_desc}",
    )

    return JsonResponse(open_seminar_to_dict(seminar), status=200)


# 5. RPC (the student's existing committee, fetched live)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def rpc_open_seminar_list(request):
    """GET /faculty/open-seminar/rpc/ -> attempts where the requester is an RPC member."""
    faculty = get_object_or_404(Faculty, id__user=request.user)

    thesis_ids = ThesisTopic.objects.filter(committee__member=faculty).values_list('id', flat=True)
    student_ids = ThesisTopic.objects.filter(id__in=thesis_ids).values_list('student_id', flat=True)

    qs = OpenSeminarAttempt.objects.filter(open_seminar__student_id__in=student_ids).select_related(
        'open_seminar__student__id__user', 'open_seminar__supervisor__id__user'
    ).distinct()

    def serialize(attempts):
        return [
            {
                **_open_seminar_attempt_to_dict(a),
                'seminar_id': a.open_seminar.id,
                'student_roll': a.open_seminar.student.id.id,
                'student_name': a.open_seminar.student.id.user.get_full_name(),
                'my_consent_given': OpenSeminarConsent.objects.filter(
                    attempt=a, member=faculty, consented=True
                ).exists(),
            }
            for a in attempts
        ]

    return JsonResponse({
        'pending': serialize(qs.filter(status='rpc_pending')),
        'history': serialize(qs.exclude(status='rpc_pending')),
    }, status=200)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def rpc_open_seminar_detail(request, attempt_pk):
    """GET /faculty/open-seminar/rpc/<attempt_pk>/"""
    faculty = get_object_or_404(Faculty, id__user=request.user)
    attempt = get_object_or_404(OpenSeminarAttempt, pk=attempt_pk)
    if not _exam_rpc_committee(attempt.open_seminar.student).filter(member=faculty).exists():
        return JsonResponse({'error': 'Not on committee'}, status=403)

    committee = []
    for cm in _exam_rpc_committee(attempt.open_seminar.student):
        fac = cm.member
        extra = fac.id
        consented = OpenSeminarConsent.objects.filter(attempt=attempt, member=fac, consented=True).exists()
        committee.append({
            'id': extra.id,
            'name': f"{extra.user.first_name} {extra.user.last_name}",
            'discipline': extra.department.name if extra.department else '',
            'consented': consented,
        })

    comments = [
        {
            'member': c.member.id.user.get_full_name(),
            'text': c.text,
            'timestamp': c.timestamp.isoformat(),
        }
        for c in attempt.rpc_comments.all()
    ]

    my_comment = OpenSeminarRPCComment.objects.filter(attempt=attempt, member=faculty).first()
    is_consented = OpenSeminarConsent.objects.filter(attempt=attempt, member=faculty, consented=True).exists()

    seminar = attempt.open_seminar
    payload = {
        **_open_seminar_attempt_to_dict(attempt),
        'seminar_id': seminar.id,
        'student_name': seminar.student.id.user.get_full_name(),
        'student_roll': seminar.student.id.id,
        'student_discipline': seminar.student.specialization,
        'possible_thesis_title': seminar.possible_thesis_title,
        'supervisor': {'id': seminar.supervisor.id.id, 'name': str(seminar.supervisor)},
        'co_supervisor': (
            {'id': seminar.co_supervisor.id.id, 'name': str(seminar.co_supervisor)}
            if seminar.co_supervisor else None
        ),
        'committee': committee,
        'committee_size': len(committee),
        'consented_count': sum(1 for m in committee if m['consented']),
        'comments': comments,
        'my_comment': my_comment.text if my_comment else '',
        'is_consented': is_consented,
    }
    return JsonResponse(payload, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def rpc_open_seminar_consent(request, attempt_pk):
    """
    POST /faculty/open-seminar/rpc/<attempt_pk>/consent/
    Body: { result, committee_comments, seminar_date, comment }
    Any edit to the shared panel resets everyone else's consent -- mirrors
    Progress Seminar's rpc_consent / Comprehensive Exam's RPC consent.
    """
    faculty = get_object_or_404(Faculty, id__user=request.user)
    attempt = get_object_or_404(OpenSeminarAttempt, pk=attempt_pk, status='rpc_pending')
    if not _exam_rpc_committee(attempt.open_seminar.student).filter(member=faculty).exists():
        return JsonResponse({'error': 'Not on committee'}, status=403)

    data = request.data
    if data.get('result') and data['result'] not in dict(OpenSeminarAttempt.RESULT_CHOICES):
        return JsonResponse({'error': 'Invalid result value'}, status=400)

    panel_fields = ['result', 'committee_comments']

    changed = any(
        field in data and getattr(attempt, field) != data[field]
        for field in panel_fields
    )
    if changed:
        OpenSeminarConsent.objects.filter(attempt=attempt).update(consented=False)

    for field in panel_fields:
        if field in data:
            setattr(attempt, field, data[field])
    if data.get('seminar_date'):
        attempt.seminar_date = data['seminar_date']
    attempt.save()

    if 'comment' in data:
        OpenSeminarRPCComment.objects.update_or_create(
            attempt=attempt, member=faculty, defaults={'text': data['comment']},
        )

    consent_obj, _created = OpenSeminarConsent.objects.get_or_create(attempt=attempt, member=faculty)
    consent_obj.consented = True
    consent_obj.save()

    return JsonResponse({'message': 'Consent & data recorded.'}, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def rpc_open_seminar_finalize(request, attempt_pk):
    """
    POST /faculty/open-seminar/rpc/<attempt_pk>/finalize/
    Requires every RPC member to have consented and a result to have been
    recorded; forwards to Convener (DPGC).
    """
    faculty = get_object_or_404(Faculty, id__user=request.user)
    attempt = get_object_or_404(OpenSeminarAttempt, pk=attempt_pk, status='rpc_pending')
    if not _exam_rpc_committee(attempt.open_seminar.student).filter(member=faculty).exists():
        return JsonResponse({'error': 'Not on committee'}, status=403)

    if not attempt.result:
        return JsonResponse({'error': 'Record a result before finalizing'}, status=400)

    current_committee_ids = _exam_rpc_committee(attempt.open_seminar.student).values_list('member_id', flat=True)
    total = len(current_committee_ids)
    yes = OpenSeminarConsent.objects.filter(
        attempt=attempt, consented=True, member_id__in=current_committee_ids,
    ).count()
    if total == 0 or yes < total:
        return JsonResponse({'error': 'Not all RPC members have consented'}, status=400)

    attempt.status = 'hod_review_pending'
    attempt.reported_by = request.user
    attempt.reported_at = timezone.now()
    # Starting a fresh Convener (DPGC) review cycle -- an earlier rejection's
    # reviewed_by/at/remarks no longer apply and would otherwise make this
    # attempt look like already-decided history while it's still pending.
    attempt.hod_reviewed_by = None
    attempt.hod_reviewed_at = None
    attempt.hod_review_remarks = ''
    attempt.save()

    seminar = attempt.open_seminar
    student = seminar.student
    acronym = student.batch_id.discipline.acronym if student.batch_id and student.batch_id.discipline else None
    _open_seminar_notify(
        sender=request.user,
        recipient=_hod_users_for_discipline(acronym),
        verb='Open Seminar pending your (post-RPC) review',
        description=f"The RPC has finalized {student.id.user.get_full_name()}'s Open Seminar "
                    f"(attempt {attempt.attempt_number}); awaiting Convener (DPGC) review.",
    )
    _open_seminar_notify(
        sender=request.user,
        recipient=seminar.supervisor.id.user,
        verb='Open Seminar RPC review finalized',
        description=f"The RPC has finalized {student.id.user.get_full_name()}'s Open Seminar "
                    f"(attempt {attempt.attempt_number}); forwarded to Convener (DPGC).",
    )

    return JsonResponse(open_seminar_to_dict(attempt.open_seminar), status=200)


# 6. Convener DPGC, second review (HOD of the student's department stands in)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def hod_review_open_seminar_dashboard(request):
    """GET /hod/open-seminar/review-dashboard/ -> attempts pending post-RPC review,
    plus a history of already-decided ones, scoped to the HOD's own discipline."""
    hod_disciplines = get_hod_disciplines(request.user)

    def _scoped(qs):
        result = []
        for attempt in qs:
            student = attempt.open_seminar.student
            acronym = student.batch_id.discipline.acronym if student.batch_id and student.batch_id.discipline else None
            if acronym and acronym in hod_disciplines:
                result.append(attempt)
        return result

    pending_qs = OpenSeminarAttempt.objects.filter(status='hod_review_pending').select_related(
        'open_seminar__student__id__user', 'open_seminar__student__batch_id__discipline', 'open_seminar__supervisor__id__user'
    )
    pending = [open_seminar_to_dict(a.open_seminar) for a in _scoped(pending_qs)]

    history_qs = OpenSeminarAttempt.objects.filter(hod_reviewed_by__isnull=False).select_related(
        'open_seminar__student__id__user', 'open_seminar__student__batch_id__discipline',
        'open_seminar__supervisor__id__user', 'hod_reviewed_by',
    ).order_by('-hod_reviewed_at')
    history = [
        {
            **open_seminar_to_dict(a.open_seminar),
            'decision': 'Rejected' if a.status == 'rpc_pending' else 'Approved',
            'decided_by': a.hod_reviewed_by.get_full_name() if a.hod_reviewed_by else None,
            'decided_at': a.hod_reviewed_at.isoformat() if a.hod_reviewed_at else None,
            'attempt_number': a.attempt_number,
            'remarks': a.hod_review_remarks,
        }
        for a in _scoped(history_qs)
    ]

    return JsonResponse({'pending': pending, 'history': history}, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def hod_review_open_seminar(request, attempt_pk):
    """
    POST /hod/open-seminar/attempt/<attempt_pk>/review/
    Body: { approve: true|false, remarks }
    Rejecting sends it back to the RPC for fresh consensus (all consents reset).
    """
    attempt = get_object_or_404(OpenSeminarAttempt, pk=attempt_pk, status='hod_review_pending')
    student = attempt.open_seminar.student
    acronym = student.batch_id.discipline.acronym if student.batch_id and student.batch_id.discipline else None
    if not is_hod_of_discipline(request.user, acronym):
        return JsonResponse({'error': 'Not authorized'}, status=403)

    data = request.data
    attempt.hod_reviewed_by = request.user
    attempt.hod_reviewed_at = timezone.now()
    if data.get('approve'):
        attempt.hod_review_remarks = ''
        attempt.status = 'dean_pending'
        attempt.save()
        _open_seminar_notify(
            sender=request.user,
            recipient=_dean_academic_users(),
            verb='Open Seminar pending your final approval',
            description=f"{student.id.user.get_full_name()}'s Open Seminar has been approved by "
                        f"Convener (DPGC) and is awaiting your final approval.",
        )
        _open_seminar_notify(
            sender=request.user,
            recipient=attempt.open_seminar.supervisor.id.user,
            verb='Open Seminar approved by Convener (DPGC)',
            description=f"{student.id.user.get_full_name()}'s Open Seminar has been approved by "
                        f"Convener (DPGC) and forwarded to Dean Academic.",
        )
    else:
        attempt.hod_review_remarks = data.get('remarks', '')
        attempt.status = 'rpc_pending'
        attempt.save()
        OpenSeminarConsent.objects.filter(attempt=attempt).update(consented=False)
        _open_seminar_notify(
            sender=request.user,
            recipient=_rpc_committee_users(student),
            verb='Open Seminar sent back by Convener (DPGC)',
            description=f"Convener (DPGC) sent {student.id.user.get_full_name()}'s Open Seminar "
                        f"back to the RPC for fresh consensus. Remarks: {attempt.hod_review_remarks or '—'}",
        )

    return JsonResponse(open_seminar_to_dict(attempt.open_seminar), status=200)


# 7. Dean Nominee (ad-hoc faculty appointment)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dean_nominee_open_seminar_dashboard(request):
    """
    GET /faculty/open-seminar-nominee/dashboard/
    Attempts where the requester is the appointed Dean Nominee, pending their
    own confidential report.
    """
    try:
        faculty = Faculty.objects.get(id__user=request.user)
    except Faculty.DoesNotExist:
        return JsonResponse({'error': 'Faculty record not found'}, status=404)

    qs = OpenSeminarAttempt.objects.filter(
        dean_nominee=faculty, dn_submitted_at__isnull=True,
    ).select_related('open_seminar')

    # Include the specific attempt id the nominee was appointed to and still
    # owes a report for -- the seminar's *current* attempt may have moved on
    # (e.g. a retry) since this nominee was appointed, so the report must
    # target this attempt, not whichever one is current now.
    return JsonResponse({
        'pending': [
            {**open_seminar_to_dict(a.open_seminar, include_confidential=True), 'nominee_attempt_id': a.id}
            for a in qs
        ],
    }, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def dean_nominee_submit_open_seminar_report(request, attempt_pk):
    """
    POST /faculty/open-seminar-nominee/attempt/<attempt_pk>/report/
    Body: { quality, quantity, publications, overall, comments }
    Only the appointed Dean Nominee can submit this -- confidential, kept
    separate from (and not gating) the committee's own verdict.
    """
    attempt = get_object_or_404(OpenSeminarAttempt, pk=attempt_pk)
    try:
        faculty = Faculty.objects.get(id__user=request.user)
    except Faculty.DoesNotExist:
        return JsonResponse({'error': 'Faculty record not found'}, status=404)

    if attempt.dean_nominee_id != faculty.pk:
        return JsonResponse({'error': 'Not authorized'}, status=403)

    if attempt.dn_submitted_at:
        return JsonResponse({'error': 'Report already submitted'}, status=403)

    data = request.data
    attempt.dn_quality = data.get('quality', '')
    attempt.dn_quantity = data.get('quantity', '')
    attempt.dn_publications = data.get('publications', '')
    attempt.dn_overall = data.get('overall', '')
    attempt.dn_comments = data.get('comments', '')
    attempt.dn_submitted_at = timezone.now()
    attempt.save()

    return JsonResponse({'message': 'Report submitted.'}, status=200)


# ===========================================================================
# Teaching Credit
# ===========================================================================
# Workflow: [Precondition: ComprehensiveExam.status == 'passed'] -> Student
# submits 4 course choices for a semester -> HOD allocates one of the 4 (or
# sends it back with remarks, student edits and resubmits) -> [offline
# teaching] -> any student registered for the allocated course that semester
# submits one anonymous evaluation -> HOD reviews the aggregated (anonymized)
# evaluations and marks the registration completed with a satisfactory/
# not_satisfactory result. Not_satisfactory is terminal -- no retry, a fresh
# attempt would just be a new semester's registration.

def _teaching_credit_choice_dict(course):
    if not course:
        return None
    return {'id': course.id, 'code': course.code, 'name': course.name}


def teaching_credit_to_dict(reg, include_evaluations=False):
    """Serialize a TeachingCreditAllocation. Evaluation respondents are
    never included -- only aggregated/anonymized responses, and only when
    include_evaluations is explicitly requested (HOD-facing endpoints)."""
    d = {
        'id': reg.id,
        'student_roll': reg.student.id.id,
        'student_name': reg.student.id.user.get_full_name(),
        'student_discipline': reg.student.specialization,
        'semester_no': reg.semester.semester_no,
        'choices': [
            _teaching_credit_choice_dict(reg.choice_1),
            _teaching_credit_choice_dict(reg.choice_2),
            _teaching_credit_choice_dict(reg.choice_3),
            _teaching_credit_choice_dict(reg.choice_4),
        ],
        'status': reg.status,
        'allocated_course': _teaching_credit_choice_dict(reg.allocated_course),
        'hod_remarks': reg.hod_remarks,
        'result': reg.result,
        'evaluation_count': reg.evaluations.count(),
    }
    if include_evaluations:
        d['evaluations'] = [
            {
                'punctuality_band': e.punctuality_band,
                'schedule_adherence_band': e.schedule_adherence_band,
                'topics_sequence': e.topics_sequence,
                'teaching_aids': e.teaching_aids,
                'questions_answered': e.questions_answered,
                'overall_effectiveness': e.overall_effectiveness,
                'strengths_weaknesses': e.strengths_weaknesses,
            }
            for e in reg.evaluations.all()
        ]
    return d


def _hod_discipline_acronyms(user):
    """Mirrors get_hod_disciplines used elsewhere -- discipline acronyms this user is HOD of."""
    hod_designations = HoldsDesignation.objects.filter(
        working=user, designation__name__icontains='HOD'
    ).values_list('designation__name', flat=True)
    acronyms = []
    for des_name in hod_designations:
        if '(' in des_name and ')' in des_name:
            acronyms.append(des_name[des_name.index('(') + 1:des_name.index(')')].strip())
    return acronyms


def _is_hod_of_student(user, student):
    acronym = student.batch_id.discipline.acronym if student.batch_id and student.batch_id.discipline else None
    if not acronym:
        return False
    return HoldsDesignation.objects.filter(working=user, designation__name=f"HOD ({acronym})").exists()


# 1. Student

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_teaching_credit_api(request):
    """GET /stu/teaching-credit/ -> this student's own registrations (all semesters)."""
    try:
        student = Student.objects.get(id=request.user.extrainfo)
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student record not found'}, status=404)

    regs = TeachingCreditAllocation.objects.filter(student=student).order_by('-semester__semester_no')
    return JsonResponse({
        'registrations': [teaching_credit_to_dict(r) for r in regs],
        'comprehensive_exam_passed': ComprehensiveExam.objects.filter(student=student, status='passed').exists(),
    }, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def student_propose_teaching_credit(request):
    """
    POST /stu/teaching-credit/propose/
    Body: { choice_1, choice_2, choice_3, choice_4 }
    Precondition: ComprehensiveExam.status == 'passed'. The semester is
    resolved server-side from the student's current curriculum position
    (same pattern as student_thesis_enrollment_api), not taken from the client.
    """
    try:
        student = Student.objects.get(id=request.user.extrainfo)
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student record not found'}, status=404)

    if not ComprehensiveExam.objects.filter(student=student, status='passed').exists():
        return JsonResponse(
            {'error': 'Comprehensive Examination must be passed before registering for teaching credit.'},
            status=403,
        )

    if not student.batch_id or not student.batch_id.curriculum:
        return JsonResponse({'error': 'Student batch or curriculum is not configured'}, status=400)
    try:
        semester = Semester.objects.get(
            curriculum=student.batch_id.curriculum,
            semester_no=student.curr_semester_no,
        )
    except Semester.DoesNotExist:
        return JsonResponse({'error': 'Current semester not found in curriculum'}, status=400)

    data = request.data
    choice_1 = data.get('choice_1')
    if not choice_1:
        return JsonResponse({'error': 'choice_1 is required'}, status=400)

    if TeachingCreditAllocation.objects.filter(student=student, semester=semester).exists():
        return JsonResponse({'error': 'Already registered for teaching credit this semester'}, status=400)

    reg = TeachingCreditAllocation.objects.create(
        student=student,
        semester=semester,
        choice_1_id=choice_1,
        choice_2_id=data.get('choice_2') or None,
        choice_3_id=data.get('choice_3') or None,
        choice_4_id=data.get('choice_4') or None,
    )

    acronym = student.batch_id.discipline.acronym if student.batch_id and student.batch_id.discipline else None
    _teaching_credit_notify(
        sender=request.user,
        recipient=_hod_users_for_discipline(acronym),
        verb='Teaching Credit registration pending allocation',
        description=f"{student.id.user.get_full_name()} has submitted teaching credit course "
                    f"choices for your allocation.",
    )

    return JsonResponse(teaching_credit_to_dict(reg), status=201)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_teaching_credit_detail(request, pk):
    """GET /stu/teaching-credit/<pk>/ -> full detail (also used to prefill a resubmission)."""
    reg = get_object_or_404(TeachingCreditAllocation, pk=pk)
    try:
        student = Student.objects.get(id=request.user.extrainfo)
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student record not found'}, status=404)

    if reg.student_id != student.pk:
        return JsonResponse({'error': 'Not authorized'}, status=403)

    return JsonResponse(teaching_credit_to_dict(reg), status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def student_resubmit_teaching_credit(request, pk):
    """
    POST /stu/teaching-credit/<pk>/resubmit/
    Edits choices after HOD sends it back, resends for HOD decision.
    """
    reg = get_object_or_404(TeachingCreditAllocation, pk=pk)
    try:
        student = Student.objects.get(id=request.user.extrainfo)
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student record not found'}, status=404)

    if reg.student_id != student.pk:
        return JsonResponse({'error': 'Not authorized'}, status=403)
    if reg.status != 'sent_back':
        return JsonResponse({'error': 'Cannot edit at this stage'}, status=403)

    data = request.data
    if 'choice_1' in data:
        reg.choice_1_id = data['choice_1']
    if 'choice_2' in data:
        reg.choice_2_id = data['choice_2'] or None
    if 'choice_3' in data:
        reg.choice_3_id = data['choice_3'] or None
    if 'choice_4' in data:
        reg.choice_4_id = data['choice_4'] or None
    reg.status = 'pending'
    reg.hod_remarks = ''
    reg.save()

    acronym = student.batch_id.discipline.acronym if student.batch_id and student.batch_id.discipline else None
    _teaching_credit_notify(
        sender=request.user,
        recipient=_hod_users_for_discipline(acronym),
        verb='Teaching Credit registration resubmitted',
        description=f"{student.id.user.get_full_name()} has resubmitted teaching credit course "
                    f"choices for your allocation.",
    )

    return JsonResponse(teaching_credit_to_dict(reg), status=200)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_teaching_credit_evaluation_targets(request):
    """
    GET /stu/teaching-credit/evaluation-targets/
    Allocated (or completed) registrations for courses the requesting
    student is registered for this semester -- i.e. whose Research Scholar
    they're eligible to evaluate -- excluding ones already submitted.
    """
    try:
        student = Student.objects.get(id=request.user.extrainfo)
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student record not found'}, status=404)

    registered_course_ids = course_registration.objects.filter(student_id=student).values_list('course_id', flat=True)
    already_evaluated = TeachingCreditEvaluationResponse.objects.filter(
        respondent=student
    ).values_list('registration_id', flat=True)

    qs = TeachingCreditAllocation.objects.filter(
        status__in=['allocated', 'completed'],
        allocated_course_id__in=registered_course_ids,
    ).exclude(id__in=already_evaluated)

    return JsonResponse({'targets': [teaching_credit_to_dict(r) for r in qs]}, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def student_submit_teaching_credit_evaluation(request, pk):
    """
    POST /stu/teaching-credit/<pk>/evaluate/
    Body: { punctuality_band, schedule_adherence_band, topics_sequence,
            teaching_aids, questions_answered, overall_effectiveness,
            strengths_weaknesses }
    Only a student registered for the allocated course may submit, once.
    Anonymous -- respondent identity is never exposed via API.

    Note: eligibility is checked by course only, not by matching
    `reg.semester` -- that field is the PhD registrant's own semester
    (resolved from their curriculum), which is a different `Semester` row
    than the respondent's `course_registration.semester_id` whenever the
    two students are in different curricula (e.g. a UG respondent taking a
    course a PhD scholar is teaching) -- `Semester` is scoped per-curriculum,
    so exact FK matching across curricula can never succeed.
    """
    reg = get_object_or_404(TeachingCreditAllocation, pk=pk, status__in=['allocated', 'completed'])
    try:
        student = Student.objects.get(id=request.user.extrainfo)
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student record not found'}, status=404)

    is_registered = course_registration.objects.filter(
        student_id=student, course_id=reg.allocated_course,
    ).exists()
    if not is_registered:
        return JsonResponse({'error': 'You are not registered for this course'}, status=403)

    if TeachingCreditEvaluationResponse.objects.filter(registration=reg, respondent=student).exists():
        return JsonResponse({'error': 'You have already submitted an evaluation for this course'}, status=400)

    data = request.data
    TeachingCreditEvaluationResponse.objects.create(
        registration=reg,
        respondent=student,
        punctuality_band=data.get('punctuality_band', ''),
        schedule_adherence_band=data.get('schedule_adherence_band', ''),
        topics_sequence=data.get('topics_sequence', ''),
        teaching_aids=data.get('teaching_aids', ''),
        questions_answered=data.get('questions_answered', ''),
        overall_effectiveness=data.get('overall_effectiveness', ''),
        strengths_weaknesses=data.get('strengths_weaknesses', ''),
    )
    return JsonResponse({'message': 'Evaluation submitted.'}, status=201)


# 2. HOD

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def hod_teaching_credit_dashboard(request):
    """
    GET /hod/teaching-credit/dashboard/
    Pending decisions + allocated-awaiting-completion, scoped to the HOD's
    own discipline.
    """
    user = request.user
    hod_disciplines = _hod_discipline_acronyms(user)

    qs = TeachingCreditAllocation.objects.filter(
        status__in=['pending', 'allocated']
    ).select_related('student__id__user', 'student__batch_id__discipline')

    pending, awaiting_completion = [], []
    for reg in qs:
        student = reg.student
        acronym = student.batch_id.discipline.acronym if student.batch_id and student.batch_id.discipline else None
        if not acronym or acronym not in hod_disciplines:
            continue
        if reg.status == 'pending':
            pending.append(teaching_credit_to_dict(reg))
        else:
            awaiting_completion.append(teaching_credit_to_dict(reg, include_evaluations=True))

    return JsonResponse({'pending': pending, 'awaiting_completion': awaiting_completion}, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def hod_decide_teaching_credit(request, pk):
    """
    POST /hod/teaching-credit/<pk>/decide/
    Body: { allocate: true|false, allocated_course (required if allocate), remarks }
    allocated_course must be one of the student's 4 submitted choices.
    """
    reg = get_object_or_404(TeachingCreditAllocation, pk=pk, status='pending')
    user = request.user

    if not _is_hod_of_student(user, reg.student):
        return JsonResponse({'error': 'Not authorized'}, status=403)

    data = request.data
    reg.decided_by = user
    reg.decided_at = timezone.now()

    if data.get('allocate'):
        allocated_course_id = data.get('allocated_course')
        valid_choices = {
            str(c) for c in (reg.choice_1_id, reg.choice_2_id, reg.choice_3_id, reg.choice_4_id) if c
        }
        if str(allocated_course_id) not in valid_choices:
            return JsonResponse(
                {'error': "Allocated course must be one of the student's 4 choices"}, status=400,
            )
        reg.allocated_course_id = allocated_course_id
        reg.hod_remarks = ''
        reg.status = 'allocated'
        reg.save()
        _teaching_credit_notify(
            sender=user,
            recipient=reg.student.id.user,
            verb='Teaching Credit course allocated',
            description=f"You have been allocated {reg.allocated_course.code} - "
                        f"{reg.allocated_course.name} for teaching credit.",
        )
    else:
        reg.hod_remarks = data.get('remarks', '')
        reg.status = 'sent_back'
        reg.save()
        _teaching_credit_notify(
            sender=user,
            recipient=reg.student.id.user,
            verb='Teaching Credit registration sent back',
            description=f"Your teaching credit choices were sent back by the HOD. "
                        f"Remarks: {reg.hod_remarks or '—'}",
        )

    return JsonResponse(teaching_credit_to_dict(reg), status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def hod_complete_teaching_credit(request, pk):
    """
    POST /hod/teaching-credit/<pk>/complete/
    Body: { result: satisfactory|not_satisfactory }
    Satisfactory awards the credit; not_satisfactory is terminal.
    """
    reg = get_object_or_404(TeachingCreditAllocation, pk=pk, status='allocated')
    user = request.user

    if not _is_hod_of_student(user, reg.student):
        return JsonResponse({'error': 'Not authorized'}, status=403)

    result = request.data.get('result')
    if result not in ('satisfactory', 'not_satisfactory'):
        return JsonResponse({'error': 'result must be satisfactory or not_satisfactory'}, status=400)

    if not reg.evaluations.exists():
        return JsonResponse(
            {'error': 'At least one student evaluation is required before completing this registration.'},
            status=400,
        )

    reg.result = result
    reg.status = 'completed'
    reg.completed_by = user
    reg.completed_at = timezone.now()
    reg.save()

    _teaching_credit_notify(
        sender=user,
        recipient=reg.student.id.user,
        verb='Teaching Credit result declared',
        description=f"Your teaching credit registration has been marked "
                    f"{'satisfactory' if result == 'satisfactory' else 'not satisfactory'} by the HOD.",
    )

    return JsonResponse(teaching_credit_to_dict(reg, include_evaluations=True), status=200)


# 3. Supervisor (read-only)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def supervisor_teaching_credit_list(request):
    """GET /supervisor/teaching-credit/ -> read-only list for the requester's thesis students."""
    try:
        faculty = Faculty.objects.get(id__user=request.user)
    except Faculty.DoesNotExist:
        return JsonResponse({'error': 'Faculty record not found'}, status=404)

    student_ids = ThesisTopic.objects.filter(
        Q(supervisor=faculty) | Q(co_supervisor=faculty)
    ).values_list('student_id', flat=True)

    qs = TeachingCreditAllocation.objects.filter(student_id__in=student_ids)
    return JsonResponse({'registrations': [teaching_credit_to_dict(r) for r in qs]}, status=200)


# 4. Academic Office (read-only)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(['acadadmin'])
def academic_office_teaching_credit_list(request):
    """GET /acadadmin/teaching-credit/ -> read-only list of all teaching-credit registrations."""
    qs = TeachingCreditAllocation.objects.select_related('student__id__user').all().order_by('-created_at')
    return JsonResponse({'registrations': [teaching_credit_to_dict(r) for r in qs]}, status=200)
