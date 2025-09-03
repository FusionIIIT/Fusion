import datetime
import json
import io
import logging
import time
import pandas as pd
from io import BytesIO
from xlsxwriter.workbook import Workbook
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
import xlsxwriter
from applications.academic_procedures.models import course_registration
from applications.academic_information.utils import allocate, check_for_registration_complete
from applications.globals.models import (HoldsDesignation,Designation)
from django.shortcuts import get_object_or_404
from django.forms.models import model_to_dict
from rest_framework.permissions import IsAuthenticated
from rest_framework.authentication import TokenAuthentication
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes,authentication_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from applications.globals.models import User,ExtraInfo
from applications.academic_information.models import Student, Course, Curriculum, Curriculum_Instructor, Student_attendance, Meeting, Calendar, Holiday, Grades, Spi, Timetable, Exam_timetable
from applications.programme_curriculum.models import Course as Courses, CourseSlot, Batch, Semester, CourseInstructor
from applications.academic_procedures.models import InitialRegistration, Assignment, StipendRequest
from . import serializers
from rest_framework.generics import ListCreateAPIView
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from applications.academic_procedures.api.views import role_required
from django.core.cache import cache
from django.db import connection

logger = logging.getLogger(__name__)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def student_api(request):

    if request.method == 'GET':
        student = Student.objects.all()
        student_serialized = serializers.StudentSerializers(student,many=True).data
        resp = {
            'students' : student_serialized,
        }
        return Response(data=resp,status=status.HTTP_200_OK)


@api_view(['GET'])
# @permission_classes([IsAuthenticated])
# @authentication_classes([TokenAuthentication])
def course_api(request):

    if request.method == 'GET':
        course = Course.objects.all()
        course_serialized = serializers.CourseSerializer(course,many=True).data
        resp = {
            'courses' : course_serialized,
        }
        return Response(data=resp,status=status.HTTP_200_OK)
 


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def curriculum_api(request):

    if request.method == 'GET':
        curriculum = Curriculum.objects.all()
        curriculum_serialized = serializers.CurriculumSerializer(curriculum,many=True).data
        resp = {
            'curriculum' : curriculum_serialized,
        }
        return Response(data=resp,status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def meeting_api(request):

    if request.method == 'GET':
        meeting = Meeting.objects.all()
        meeting_serialized = serializers.MeetingSerializers(meeting,many=True).data
        resp = {
            'meeting' : meeting_serialized,
        }
        return Response(data=resp,status=status.HTTP_200_OK)


# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# @authentication_classes([TokenAuthentication])
# def calendar_api(request):

#     if request.method == 'GET':
#         calendar = Calendar.objects.all()
#         calendar_serialized = serializers.CalendarSerializers(calendar,many=True).data
#         resp = {
#             'calendar' :calendar_serialized,
#         }
#         return Response(data=resp,status=status.HTTP_200_OK)
    
# class ListCalendarView(ListCreateAPIView):
#     permission_classes = [IsAuthenticated]
#     authentication_classes=[TokenAuthentication]
#     serializer_class = serializers.CalendarSerializers
#     queryset = Calendar.objects.all()

# @api_view(['PUT'])
# @permission_classes([IsAuthenticated])
# @authentication_classes([TokenAuthentication])
# def update_calendar(request):
#     if request.method == "PUT":
#         id = request.data.get("id")
#         instance = Calendar.objects.get(pk = id)
#         instance.from_date = request.data.get("from_date")
#         instance.to_date = request.data.get("to_date")
#         instance.description = request.data.get("description")
#         instance.save()
        
#         return Response({"message": "Updated successfully!"})

# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# @authentication_classes([TokenAuthentication])
# def add_calendar(request):
#     if request.method == "POST":
#         from_date = request.data.get("from_date")
#         to_date = request.data.get("to_date")
#         description = request.data.get("description")
#         Calendar.objects.create(from_date=from_date, to_date=to_date, description=description)
        
#         return Response({"message": "Created successfully!"})
    
# @api_view(['DELETE'])
# @permission_classes([IsAuthenticated])
# @authentication_classes([TokenAuthentication])
# def delete_calendar(request):
#     id = request.data.get("id")  # Get the ID from request body

#     try:
#         instance = Calendar.objects.get(pk=id)
#         instance.delete()
#         return Response({"message": "Deleted successfully!"}, status=200)
#     except Calendar.DoesNotExist:
#         return Response({"error": "Calendar entry not found"}, status=404)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def holiday_api(request):

    if request.method == 'GET':
        holiday = Holiday.objects.all()
        holiday_serialized = serializers.HolidaySerializers(holiday,many=True).data
        resp = {
            'holiday' : holiday_serialized,
        }
        return Response(data=resp,status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def timetable_api(request):

    if request.method == 'GET':
        timetable = Timetable.objects.all()
        timetable_serialized = serializers.TimetableSerializers(timetable,many=True).data
        resp = {
            'timetable' : timetable_serialized,
        }
        return Response(data=resp,status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def exam_timetable_api(request):

    if request.method == 'GET':
        exam_timetable = Exam_timetable.objects.all()
        exam_timetable_serialized = serializers.Exam_timetableSerializers(exam_timetable,many=True).data
        resp = {
            'exam_timetable' : exam_timetable_serialized,
        }
        return Response(data=resp,status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def curriculum_instructor_api(request):

    if request.method == 'GET':
        curriculum_instructor = Curriculum_Instructor.objects.all()
        curriculum_instructor_serialized = serializers.CurriculumInstructorSerializer(curriculum_instructor,many=True).data
        resp = {
            'curriculum_instructor' : curriculum_instructor_serialized,
        }
        return Response(data=resp,status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def student_attendance_api(request):

    if request.method == 'GET':
        student_attendance = Student_attendance.objects.all()
        student_attendance_serialized = serializers.Student_attendanceSerializers(student_attendance,many=True).data
        resp = {
            'student_attendance' : student_attendance_serialized,
        }
        return Response(data=resp,status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def grades_api(request):

    if request.method == 'GET':
        grades = Grades.objects.all()
        grades_serialized = serializers.GradesSerializers(grades,many=True).data
        resp = {
            'grades' : grades_serialized,
        }
        return Response(data=resp,status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
def spi_api(request):

    if request.method == 'GET':
        spi = Spi.objects.all()
        spi_serialized = serializers.SpiSerializers(spi,many=True).data
        resp = {
            'spi' : spi_serialized,
        }
        return Response(data=resp,status=status.HTTP_200_OK)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
@role_required(['acadadmin'])
def check_allocation_api(request):
    """
    API to check the allocation status for a given batch, semester, and year.
    Uses the utility function to avoid code repetition.
    """
    try:
        batch = request.data.get('batch')
        sem = request.data.get('sem')
        year = request.data.get('year')

        if not batch or not sem or not year:
            return Response(
                {"status": -3, "message": "Batch, semester, and year are required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            batch = int(batch)
            sem = int(sem)
        except ValueError:
            return Response(
                {"status": -3, "message": "Invalid batch or semester value"},
                status=status.HTTP_400_BAD_REQUEST
            )

        result = check_for_registration_complete(batch, sem, year)

        # Map status values to appropriate HTTP codes
        status_code_map = {
            -3: status.HTTP_404_NOT_FOUND,
            -2: status.HTTP_403_FORBIDDEN,
            -1: status.HTTP_200_OK,
             1: status.HTTP_200_OK,
             2: status.HTTP_200_OK,
        }

        return Response(result, status=status_code_map.get(result["status"], status.HTTP_500_INTERNAL_SERVER_ERROR))

    except Exception as e:
        return Response(
            {"status": -3, "message": f"Internal Server Error: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
@role_required(['acadadmin'])
def start_allocation_api(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method"}, status=400)

    try:
        data = json.loads(request.body.decode('utf-8'))
        batch = data.get("batch")
        semester = data.get("semester")
        year = data.get("year")

        if not batch or not semester or not year:
            return JsonResponse({"error": "Missing required fields"}, status=400)

        batch = int(batch)
        semester = int(semester)

        mock_request = type('MockRequest', (), {})()
        mock_request.POST = {'batch': batch, 'sem': semester, 'year': year}

        return allocate(mock_request)

    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON format"}, status=400)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def parse_academic_year(academic_year, semester_type):
    """
    Parse academic_year string (e.g., "2024-25") and determine the working_year based on semester type.
    For Odd Semester, working_year = first part (e.g., 2024).
    For Even Semester, working_year = second part prefixed by '20' (e.g., 2025 if academic_year is "2024-25").
    The session is set to the academic_year string.
    """
    parts = academic_year.split("-")
    if len(parts) != 2:
        raise ValueError("Invalid academic year format. Expected format like '2024-25'.")
    first_year = parts[0].strip()
    second_year = parts[1].strip()
    if semester_type == "Odd Semester":
        working_year = int(first_year)
    elif semester_type == "Even Semester":
        working_year = int("20" + second_year)
    else:
        # For any other semester type (e.g., Summer Semester) use the first year by default.
        working_year = int("20"+second_year)
    return working_year


@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
@role_required(['acadadmin', "Associate Professor", "Professor", "Assistant Professor", "Dean Academic"])
def generate_xlsheet_api(request):
    try:
        start_time = time.time()
        
        # Debug: Check request data
        print(f"DEBUG: Request data: {request.data}")
        
        # Get parameters
        course_id = int(request.data.get('course'))
        academic_year = request.data.get('academic_year')
        semester_type = request.data.get('semester_type')
        list_type = request.data.get('list_type', '').strip()
        preview_only = request.data.get('preview_only', False)

        if not list_type:
            list_type = None
        
        if not all([course_id, academic_year, semester_type]):
            return Response({
                'error': 'Missing required parameters: course, academic_year, semester_type'
            }, status=status.HTTP_400_BAD_REQUEST)
        cache_key = f"student_list_{course_id}_{academic_year.replace('-', '_')}_{semester_type.replace(' ', '_')}_{(list_type or 'all').replace(' ', '_')}"
        
        cached_data = cache.get(cache_key)
        if cached_data and not preview_only:
            students, course_info = cached_data
        else:
            sql = """
            SELECT DISTINCT
                u.username as roll_no,
                u.first_name,
                u.last_name,
                CONCAT(u.first_name, ' ', u.last_name) as full_name,
                COALESCE(s.specialization, 'General') as discipline,
                u.email,
                cr.registration_type,
                c.code as course_code,
                c.name as course_name
            FROM course_registration cr
            INNER JOIN globals_extrainfo ei ON cr.student_id_id = ei.id
            INNER JOIN auth_user u ON ei.user_id = u.id
            LEFT JOIN academic_information_student s ON ei.id = s.id_id
            INNER JOIN programme_curriculum_course c ON cr.course_id_id = c.id
            WHERE cr.session = %s 
                AND cr.semester_type = %s 
                AND cr.course_id_id = %s
            """
            
            params = [academic_year, semester_type, course_id]
            
            # Add list_type filter if specified
            if list_type:
                if list_type.lower() == 'backlog_improvement':
                    sql += " AND cr.registration_type IN ('Backlog', 'Improvement')"
                else:
                    sql += " AND cr.registration_type = %s"
                    params.append(list_type)
            
            sql += " ORDER BY u.username"
            try:
                with connection.cursor() as cursor:
                    start_sql = time.time()

                    cursor.execute(sql, params)
                    columns = [col[0] for col in cursor.description]
                    raw_results = cursor.fetchall()
                    students = []
                    for row in raw_results:
                        data = dict(zip(columns, row))
                        students.append({
                            'roll_no': data['roll_no'],
                            'first_name': data['first_name'],
                            'last_name': data['last_name'],
                            'full_name': data['full_name'],
                            'discipline': data['discipline'],
                            'email': data['email'],
                            'registration_type': data['registration_type']
                        })
                    
                    sql_time = time.time() - start_sql
                    print(f"DEBUG: SQL Execution Time: {sql_time:.3f}s for {len(students)} students")

                    reg_types = set(student['registration_type'] for student in students)

                    if raw_results:
                        course_info = {
                            'code': raw_results[0][7],  # course_code
                            'name': raw_results[0][8]   # course_name
                        }
                    else:
                        cursor.execute("SELECT code, name FROM programme_curriculum_course WHERE id = %s", [course_id])
                        result = cursor.fetchone()
                        course_info = {'code': result[0], 'name': result[1]} if result else {'code': 'N/A', 'name': 'Unknown'}
            except Exception as e:
                # Debug: logger.error(f"SQL Error in generate_xlsheet_api: {str(e)}")
                return Response({
                    'error': f'Database query failed: {str(e)}',
                    'sql_debug': sql,
                    'params': params
                }, status=500)
            
            cache.set(cache_key, (students, course_info), 300)

        if not list_type:
            list_type_display = "All Enrolled Students"
        elif list_type.lower() == 'backlog_improvement':
            list_type_display = "Backlog & Improvement Students"
        else:
            list_type_display = f"{list_type} Students"
        
        processing_time = time.time() - start_time
        if preview_only:
            preview_students = students[:350] if len(students) > 350 else students
            
            return Response({
                'students': preview_students,
                'course_info': {
                    'code': course_info['code'],
                    'name': course_info['name'],
                    'academic_year': academic_year,
                    'semester_type': semester_type,
                    'list_type': list_type_display,
                    'total_count': len(students),
                    'preview_count': len(preview_students),
                    'processing_time_ms': round(processing_time * 1000, 2)
                }
            }, status=status.HTTP_200_OK)
        
        # OPTIMIZATION 7: Fast Excel generation with minimal formatting
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Student List"
        
        # Set column widths once
        column_widths = [8, 15, 30, 15, 35, 18, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # Minimal header formatting (single operation)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Add title rows efficiently
        ws.merge_cells('A1:G1')
        ws['A1'] = "PDPM INDIAN INSTITUTE OF INFORMATION TECHNOLOGY, DESIGN AND MANUFACTURING JABALPUR"
        ws['A1'].font = Font(bold=True, size=9)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A2:G2')
        ws['A2'] = f"{semester_type.upper()}, {academic_year}"
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # Course details
        ws['A3'] = "Course No:"
        ws['B3'] = course_info['code']
        ws['A4'] = "Course Title:"
        ws.merge_cells('B4:G4')
        ws['B4'] = course_info['name']
        ws['A5'] = "List Type:"
        ws.merge_cells('B5:G5')
        ws['B5'] = list_type_display

        headers = ['Sl. No', 'Roll No', 'Name', 'Discipline', 'Email', 'Reg. Type', 'Signature']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for idx, student in enumerate(students, 1):
            row_data = [
                idx,
                student['roll_no'],
                student['full_name'],
                student['discipline'],
                student['email'],
                student['registration_type'],
                ''  # Signature
            ]
            ws.append(row_data)

        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        if not list_type:
            filename_suffix = "All_Enrolled_Students"
        elif list_type.lower() == 'backlog_improvement':
            filename_suffix = "Backlog_Improvement_Students"
        else:
            filename_suffix = f"{list_type.replace(' ', '_')}_Students"
        
        filename = f"{course_info['code']}_{filename_suffix}_CourseList.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        total_time = time.time() - start_time
        response['X-Processing-Time'] = f"{round(total_time * 1000, 2)}ms"
        response['X-Student-Count'] = str(len(students))
        
        return response
        
    except Exception as e:
        try:
            error_time = time.time() - start_time
        except:
            error_time = 0
            
        print(f"DETAILED ERROR: {str(e)}")
        print(f"ERROR TYPE: {type(e)}")
        import traceback
        print(f"TRACEBACK: {traceback.format_exc()}")
        
        return Response({
            'error': f'Error: {str(e)}',
            'error_type': str(type(e)),
            'processing_time_ms': round(error_time * 1000, 2),
            'debug_info': 'Check server console for detailed traceback'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
@role_required(['acadadmin'])
def generate_preregistration_report(request):
    """
    to generate preresgistration report after pre-registration

    @param:
        request - contains metadata about the requested page

    @variables:
        sem - get current semester from current time
        now - get current time
        year - getcurrent year
        batch - gets the batch from form
        sem - stores the next semester
        obj - All the registration details appended into one
        data - Formated data for context
        m - counter for Sl. No (in formated data)
        z - temporary array to add data to variable data
        k -temporary array to add data to formatted array/variable
        output - io Bytes object to write to xlsx file
        book - workbook of xlsx file
        title - formatting variable of title the workbook
        subtitle - formatting variable of subtitle the workbook
        normaltext - formatting variable for normal text
        sheet - xlsx sheet to be rendered
        titletext - formatting variable of title text
        dep - temporary variables
        z - temporary variables for final output
        b - temporary variables for final output
        c - temporary variables for final output
        st - temporary variables for final output

    """
        
    if request.method == "POST":
        sem = request.data.get('semester_no')
        batch_id=request.data.get('batch_branch')
        batch = Batch.objects.filter(id = batch_id).first()
        obj = InitialRegistration.objects.filter(student_id__batch_id=batch_id, semester_id__semester_no=sem)



        registered_students = set()
        unregistered_students = set()

        # registered students contains objects of type InitialRegistration
        for stu in obj:
            registered_students.add(stu.student_id)

        students = Student.objects.filter(batch_id = batch_id)

        for stu in students:
            if stu not in registered_students:
                unregistered_students.add(stu)
        

        # for stu in obj:
        #     registered_students.add(stu.student_id)
        # students = Student.objects.filter(batch_id = batch_id)
        # for stu in students:
        #     if stu not in registered_students:
        #         unregistered_students.add(stu)



        data = []
        m = 1
        for i in unregistered_students:
            # z is a row in excel
            z = []
            z.append(m)
            m += 1
            z.append(i.id.user.username)
            z.append(str(i.id.user.first_name)+" "+str(i.id.user.last_name))
            z.append(i.id.department.name)
            z.append('Not Registered')
            data.append(z)

        sem_id = Semester.objects.get(curriculum = batch.curriculum, semester_no = sem)
        course_slots = CourseSlot.objects.all().filter(semester = sem_id)
        max_width = 1
        for student in registered_students:
            #z = []
            # z.append(m)
            # m += 1
            # z.append(i.id.user.username)
            # z.append(str(i.id.user.first_name)+" "+str(i.id.user.last_name))
            # z.append(i.id.department.name)
            # z.append('Registered')
            # data.append(z)
            current_student_registered_courses = InitialRegistration.objects.filter(student_id=student, semester_id__semester_no=sem).all()
            timestamp = current_student_registered_courses.first().timestamp
            #print("current student is ",student.id.user.username)
            #print("timstamp value ",timestamp)
            for slot in course_slots:
                #print("current slot belongs to ",slot)
                z = []
                z.append(m)
                z.append(student.id.user.username)
                z.append(str(student.id.user.first_name)+" "+str(student.id.user.last_name))
                z.append(student.id.department.name)
                z.append('Registered')
                z.append(str(timestamp))
                z.append(str(slot.name))
                
                choices_of_current_student = InitialRegistration.objects.filter(student_id=student, semester_id__semester_no=sem,course_slot_id = slot).all()
                max_width = max(max_width,len(choices_of_current_student))

                for choice in range(1,len(choices_of_current_student)+1):
                    try:
                        current_choice = InitialRegistration.objects.get(student_id=student, semester_id__semester_no=sem, course_slot_id=slot, priority=choice)
                        z.append(str(current_choice.course_id.code) + "-" + str(current_choice.course_id.name))
                    except :
                        z.append("No registration found")
                    # current_choice = InitialRegistration.objects.get(student_id=student, semester_id__semester_no=sem,course_slot_id = slot,priority = choice)
                    # # #print("current choice is ",current_choice)
                    # z.append(str(current_choice.course_id.code)+"-"+str(current_choice.course_id.name))
                
                data.append(z)
                m+=1
        output = BytesIO()

        book = xlsxwriter.Workbook(output,{'in_memory':True})
        title = book.add_format({'bold': True,
                                    'font_size': 22,
                                    'align': 'center',
                                    'valign': 'vcenter'})
        subtitle = book.add_format({'bold': True,
                                    'font_size': 15,
                                    'align': 'center',
                                    'valign': 'vcenter'})
        normaltext = book.add_format({'bold': False,
                                    'font_size': 15,
                                    'align': 'center',
                                    'valign': 'vcenter'})
        sheet = book.add_worksheet()

        # add semester too in title text
        title_text = ("Pre-registeration : "+ batch.name + str(" ") + batch.discipline.acronym + str(" ") + str(batch.year) + " Semester : "+str(sem))
        # ??
        sheet.set_default_row(25)
        characters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        # text, formatting
        sheet.merge_range('A2:E2', title_text, title)
        sheet.write_string('A3',"Sl. No",subtitle)
        sheet.write_string('B3',"Roll No",subtitle)
        sheet.write_string('C3',"Name",subtitle)
        sheet.write_string('D3',"Discipline",subtitle)
        sheet.write_string('E3','Status',subtitle)
        sheet.write_string('F3','TimeStamp',subtitle)
        sheet.write_string('G3','Course Slot ID',subtitle)
        for choice_num  in range(7,7+max_width):
            sheet.write_string(characters[choice_num]+'3','Choice '+str(choice_num-6),subtitle)

        
        # Width of column
        sheet.set_column('A:A',20)
        sheet.set_column('B:B',20)
        sheet.set_column('C:C',50)
        sheet.set_column('D:D',15)
        sheet.set_column('E:E',20)
        sheet.set_column('F:F',40)
        sheet.set_column('G:G',30)
        sheet.set_column('H:H',70)
        sheet.set_column('I:I',70)
        sheet.set_column('J:J',70)
        sheet.set_column('K:K',70)
        sheet.set_column('L:L',70)
        sheet.set_column('M:M',70)
        #rows numbers
        k = 4
        # SERIAL numbers S.no 1,2,3...
        num = 1
        for i in data:
            sheet.write_number('A'+str(k),num,normaltext)
            num+=1
            z,b,c = str(i[0]),i[1],i[2]
            if(len(i) > 5):
                a,b,c,d,e,f,g = str(i[0]),str(i[1]),str(i[2]),str(i[3]),str(i[4]),str(i[5]),str(i[6])
                temp = str(i[3]).split()
                sheet.write_string('B'+str(k),b,normaltext)
                sheet.write_string('C'+str(k),c,normaltext)
                sheet.write_string('D'+str(k),d,normaltext)
                sheet.write_string('E'+str(k),e,normaltext)
                sheet.write_string('F'+str(k),f,normaltext)
                sheet.write_string('G'+str(k),g,normaltext)
                characters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
                # for character in characters
                for temp_num in range(7,len(i)):
                    sheet.write_string(characters[temp_num]+str(k),str(i[temp_num]),normaltext)
            else:
                a,b,c,d,e= str(i[0]),str(i[1]),str(i[2]),str(i[3]),str(i[4])
                temp = str(i[3]).split()
                sheet.write_string('B'+str(k),b,normaltext)
                sheet.write_string('C'+str(k),c,normaltext)
                sheet.write_string('D'+str(k),d,normaltext)
                sheet.write_string('E'+str(k),e,normaltext)

            k+=1
        book.close()
        # ?? 
        output.seek(0)
        response = HttpResponse(output.read(),content_type = 'application/vnd.ms-excel')
        st = 'attachment; filename = ' + batch.name + batch.discipline.acronym + str(batch.year) + '-preresgistration.xlsx'
        response['Content-Disposition'] = st
        return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])

def list_calendar(request):
    events = Calendar.objects.all().values('id', 'description', 'from_date', 'to_date')
    return Response(list(events))

@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
@role_required(['acadadmin'])
def add_calendar(request):
    Calendar.objects.create(
        description=request.data.get('description'),
        from_date=request.data.get('from_date'),
        to_date=request.data.get('to_date'),
    )
    return Response({'message': 'Created successfully!'})

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
@role_required(['acadadmin'])
def update_calendar(request):
    try:
        cal = Calendar.objects.get(pk=request.data.get('id'))
    except Calendar.DoesNotExist:
        return Response({'error': 'Not found'}, status=404)
    cal.description = request.data.get('description')
    cal.from_date = request.data.get('from_date')
    cal.to_date = request.data.get('to_date')
    cal.save()
    return Response({'message': 'Updated successfully!'})

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
@role_required(['acadadmin'])
def delete_calendar(request):
    try:
        cal = Calendar.objects.get(pk=request.data.get('id'))
        cal.delete()
        return Response({'message': 'Deleted successfully!'})
    except Calendar.DoesNotExist:
        return Response({'error': 'Not found'}, status=404)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
@role_required(['acadadmin'])
def clear_calendar(request):
    Calendar.objects.all().delete()
    return Response({'message': 'All events deleted!'})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
@role_required(['acadadmin'])
def export_calendar(request):
    qs = Calendar.objects.all().values('description', 'from_date', 'to_date')
    df = pd.DataFrame(list(qs))
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename="calendar.xlsx"'
    return resp

@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([TokenAuthentication])
@role_required(['acadadmin'])
def import_calendar(request):
    file = request.FILES.get('file')
    if not file:
        return Response({'error': 'Excel file required'}, status=400)
    df = pd.read_excel(file)
    errors = []
    for _, row in df.iterrows():
        try:
            Calendar.objects.create(
                description=row['description'],
                from_date=row['from_date'],
                to_date=row['to_date'],
            )
        except Exception as e:
            errors.append(str(e))
    if errors:
        return Response({'errors': errors}, status=400)
    return Response({'message': 'Imported successfully!'})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@role_required(['acadadmin'])
def available_courses(request):
    """
    GET /api/available-courses/?academic_year=2024-25&semester_type=Odd+Semester
    Returns unique courses for which the student has registrations.
    """
    year = request.query_params.get('academic_year')
    sem  = request.query_params.get('semester_type')
    if not year or not sem:
        return Response({"detail": "academic_year and semester_type required"}, status=400)

    regs = course_registration.objects.filter(session=year, semester_type=sem)
    course_ids = regs.values_list('course_id', flat=True).distinct()
    courses = Courses.objects.filter(id__in=course_ids)
    data = [{"id": c.id, "code": c.code, "name": c.name} for c in courses]
    return Response(data)