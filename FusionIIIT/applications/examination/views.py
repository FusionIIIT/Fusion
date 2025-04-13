from notifications.signals import notify
from django.views import View
from django.views.generic import View
import traceback
from django.http import HttpResponse
from django.conf import settings
from django.contrib.auth import get_user_model
import csv
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from io import BytesIO,StringIO
from django.db.models import IntegerField
from django.db.models.functions import Cast
from django.db.models.query_utils import Q
from django.http import request, HttpResponse
from django.shortcuts import get_object_or_404, render, HttpResponse, redirect
from django.http import HttpResponse, HttpResponseRedirect
import itertools
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from datetime import date
import requests
from django.db.models import Q
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth.models import User
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from applications.academic_information.models import Spi, Student, Curriculum
from applications.globals.models import (
    Designation,
    ExtraInfo,
    HoldsDesignation,
    Faculty,
)
from applications.eis.models import faculty_about, emp_research_projects
from applications.academic_information.models import Course
from applications.academic_procedures.models import course_registration, Register,Semester
from applications.programme_curriculum.filters import CourseFilter
from notification.views import examination_notif
from applications.department.models import SpecialRequest, Announcements
from applications.globals.models import (
    DepartmentInfo,
    Designation,
    ExtraInfo,
    Faculty,
    HoldsDesignation,
)
from jsonschema import validate
from jsonschema.exceptions import ValidationError
from django.shortcuts import render, redirect, HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import hidden_grades, grade
from .forms import StudentGradeForm
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import hidden_grades, authentication
from rest_framework.permissions import AllowAny
from applications.online_cms.models import Student_grades
from django.http import JsonResponse
import csv
from applications.programme_curriculum.models import Course as Courses, CourseInstructor,Discipline,Batch, CourseSlot
from django.urls import reverse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.colors import HexColor
from reportlab.lib.units import inch 
@login_required(login_url="/accounts/login")
def exam(request):
    """
    This function is used to Differenciate acadadmin and all other user.

    @param:
        request - contains metadata about the requested page

    @variables:
        user_details - Gets the information about the logged in user.
        des - Gets the designation about the looged in user.
    #"""
    user_details = ExtraInfo.objects.get(user=request.user)
    des = request.session.get("currentDesignationSelected")
    if (
        str(des) == "Associate Professor"
        or str(des) == "Professor"
        or str(des) == "Assistant Professor"
    ):
        return HttpResponseRedirect("/examination/submitGradesProf/")
    elif request.session.get("currentDesignationSelected") == "acadadmin":
        return HttpResponseRedirect("/examination/updateGrades/")
    elif request.session.get("currentDesignationSelected") == "Dean Academic":
        return HttpResponseRedirect("/examination/verifyGradesDean/")
    # elif request.session.get("currentDesignationSelected") == "student":
    #     return HttpResponseRedirect("/examination/checkresult/")
    return HttpResponseRedirect("/dashboard/")


@login_required(login_url='/accounts/login')
def submit(request):
    des = request.session.get("currentDesignationSelected")
    if des == "acadadmin" or des=="Dean Academic" :
        pass
    else:
        return HttpResponseRedirect('/dashboard/')
    unique_course_ids = course_registration.objects.values(
        'course_id').distinct()

    # Cast the course IDs to integers
    unique_course_ids = unique_course_ids.annotate(
        course_id_int=Cast('course_id', IntegerField()))

    # Retrieve course names and course codes based on unique course IDs
    courses_info = Course.objects.filter(
        id__in=unique_course_ids.values_list('course_id_int', flat=True))

    return render(request, '../templates/examination/submit.html', {'courses_info': courses_info})


@login_required(login_url='/accounts/login')
def verify(request):
    unique_course_ids = hidden_grades.objects.values('course_id').distinct()

    unique_course_ids = unique_course_ids.annotate(
        course_id_int=Cast('course_id', IntegerField()))

    courses_info = Course.objects.filter(
        id__in=unique_course_ids.values_list('course_id_int', flat=True))

    return render(request, '../templates/examination/verify.html', {'courses_info': courses_info})


@login_required(login_url='/accounts/login')
def publish(request):
    return render(request, '../templates/examination/publish.html', {})


@login_required(login_url='/accounts/login')
def notReady_publish(request):
    return render(request, '../templates/examination/notReady_publish.html', {})


@login_required(login_url='/accounts/login')
def timetable(request):
    return render(request, '../templates/examination/timetable.html', {})


def browse_announcements():
    """
    This function is used to browse Announcements Department-Wise
    made by different faculties and admin.

    @variables:
        cse_ann - Stores CSE Department Announcements
        ece_ann - Stores ECE Department Announcements
        me_ann - Stores ME Department Announcements
        sm_ann - Stores SM Department Announcements
        all_ann - Stores Announcements intended for all Departments
        context - Dictionary for storing all above data

    """
    cse_ann = Announcements.objects.filter(department="CSE")
    ece_ann = Announcements.objects.filter(department="ECE")
    me_ann = Announcements.objects.filter(department="ME")
    sm_ann = Announcements.objects.filter(department="SM")
    all_ann = Announcements.objects.filter(department="ALL")
    print(cse_ann)
    context = {
        "cse": cse_ann,
        "ece": ece_ann,
        "me": me_ann,
        "sm": sm_ann,
        "all": all_ann
    }

    return context


def get_to_request(username):
    """
    This function is used to get requests for the receiver

    @variables:
        req - Contains request queryset

    """
    req = SpecialRequest.objects.filter(request_receiver=username)
    return req


def entergrades(request):
    course_id = request.GET.get('course')
    semester_id = request.GET.get('semester')

    course_present = hidden_grades.objects.filter(
        course_id=course_id, semester_id=semester_id)

    if (course_present):
        return render(request, 'examination/all_course_grade_filled.html', {})

    registrations = course_registration.objects.filter(
        course_id__id=course_id, semester_id=semester_id)

    context = {
        'registrations': registrations
    }

    return render(request, 'examination/entergrades.html', context)


def verifygrades(request):
    course_id = request.GET.get('course')
    semester_id = request.GET.get('semester')

    registrations = hidden_grades.objects.filter(
        course_id=course_id, semester_id=semester_id)

    context = {
        'registrations': registrations
    }

    return render(request, 'examination/verifygrades.html', context)


def authenticate(request):  # new
    # Retrieve unique course IDs from hidden_grades
    unique_course_ids = Student_grades.objects.values('course_id').distinct()

    # Cast the course IDs to integers
    unique_course_ids = unique_course_ids.annotate(
        course_id_int=Cast('course_id', IntegerField()))

    # Retrieve course names and course codes based on unique course IDs
    courses_info = Courses.objects.filter(
        id__in=unique_course_ids.values_list('course_id_int', flat=True))
    working_years = Student_grades.objects.values(
        'year').distinct()
    context = {
        'courses_info': courses_info,
        'working_years':working_years

    }
    print(working_years)
    return render(request, '../templates/examination/authenticate.html', context)


@login_required(login_url='/accounts/login')
def authenticategrades(request):  # new
    course_id = request.GET.get('course')
    year = request.GET.get('year')

    print(course_id)
    print(year)

    course_instance = Courses.objects.get(id=course_id)
    registrations = authentication.objects.filter(
        course_id=course_instance, course_year=year)

    if registrations:
        # Registrations exist, pass them to the template context
        context = {
            'registrations': registrations,
            'year': year
        }
    else:
        course_instance = Courses.objects.get(id=course_id)
        course_present = Student_grades.objects.filter(
            course_id=course_id, year=year)
        if (course_present):
            authentication_object = authentication.objects.create(
                course_id=course_instance, course_year=year)
            registrations = authentication.objects.filter(
                course_id=course_instance, course_year=year)

            context = {
                'registrations': registrations,
                'course_year': year,
            }

    context = {
        'registrations': registrations,
        'year': year
    }

    return render(request, 'examination/authenticategrades.html', context)


# def examination_notif(sender, recipient, type):
#     try:
#         url = 'examination:examination'
#         module = 'examination'
#         verb = type
#         flag = "examination"

#         notify.send(sender=sender,
#                     recipient=recipient,
#                     url=url,
#                     module=module,
#                     verb=verb,
#                     flag=flag)

#     except Exception as e:
#         print("Error sending notification:", e)


@login_required(login_url='/accounts/login')
def announcement(request):
    des = request.session.get("currentDesignationSelected")
    if des == "acadadmin" :
        pass
    else:
        return HttpResponseRedirect('/dashboard/')
    """
    This function is contains data for Requests and Announcement Related methods.
    Data is added to Announcement Table using this function.

    @param:
        request - contains metadata about the requested page

    @variables:
        usrnm, user_info, ann_maker_id - Stores data needed for maker
        batch, programme, message, upload_announcement,
        department, ann_date, user_info - Gets and store data from FORM used for Announcements for Students.

    """
    try:
        usrnm = get_object_or_404(User, username=request.user.username)
        user_info = usrnm.extrainfo
        ann_maker_id = user_info.id
        requests_received = get_to_request(usrnm)

        if request.method == 'POST':
            batch = request.POST.get('batch', '')
            programme = request.POST.get('programme', '')
            message = request.POST.get('announcement', '')
            upload_announcement = request.FILES.get('upload_announcement')
            department = request.POST.get('department')
            ann_date = date.today()

            obj1 = Announcements.objects.get_or_create(
                maker_id=user_info,
                batch=batch,
                programme=programme,
                message=message,
                upload_announcement=upload_announcement,
                department=department,
                ann_date=ann_date
            )

            recipients = User.objects.all()  # Modify this query as per your requirements
            examination_notif(sender=usrnm, recipient=recipients, type=message)
            return render(request,'department/browse_announcements_staff.html')
        print(user_info.user_type)
        context = browse_announcements()
        return render(request, 'examination/announcement_req.html', {
            "user_designation": user_info.user_type,
            "announcements": context,
            "request_to": requests_received
        })
    except Exception as e:

        return render(request, 'examination/announcement_req.html', {"error_message": "An error occurred. Please try again later."})


class Updatehidden_gradesMultipleView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        student_ids = request.POST.getlist('student_ids[]')
        semester_ids = request.POST.getlist('semester_ids[]')
        course_ids = request.POST.getlist('course_ids[]')
        grades = request.POST.getlist('grades[]')

        if len(student_ids) != len(semester_ids) != len(course_ids) != len(grades):
            return Response({'error': 'Invalid grade data provided'}, status=status.HTTP_400_BAD_REQUEST)

        for student_id, semester_id, course_id, grade in zip(student_ids, semester_ids, course_ids, grades):
            # Create an instance of hidden_grades model and save the data

            try:
                hidden_grade = hidden_grades.objects.get(
                    course_id=course_id, student_id=student_id, semester_id=semester_id)
                hidden_grade.grade = grade
                hidden_grade.save()
            except hidden_grades.DoesNotExist:
                # If the grade doesn't exist, create a new one
                hidden_grade = hidden_grades.objects.create(
                    course_id=course_id, student_id=student_id, semester_id=semester_id, grade=grade)
                hidden_grade.save()

            hidden_grade.save()

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="grades.csv"'

        # Write data to CSV
        writer = csv.writer(response)
        writer.writerow(['Student ID', 'Semester ID', 'Course ID', 'Grade'])
        for student_id, semester_id, course_id, grade in zip(student_ids, semester_ids, course_ids, grades):
            writer.writerow([student_id, semester_id, course_id, grade])

        return response
        return render(request, '../templates/examination/grades_updated.html', {})


class Submithidden_gradesMultipleView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        student_ids = request.POST.getlist('student_ids[]')
        semester_ids = request.POST.getlist('semester_ids[]')
        course_ids = request.POST.getlist('course_ids[]')
        grades = request.POST.getlist('grades[]')

        if len(student_ids) != len(semester_ids) != len(course_ids) != len(grades):
            return Response({'error': 'Invalid grade data provided'}, status=status.HTTP_400_BAD_REQUEST)

        for student_id, semester_id, course_id, grade in zip(student_ids, semester_ids, course_ids, grades):
            # Create an instance of hidden_grades model and save the data

            try:
                hidden_grade = hidden_grades.objects.get(
                    course_id=course_id, student_id=student_id, semester_id=semester_id)
                hidden_grade.grade = grade
                hidden_grade.save()
            except hidden_grades.DoesNotExist:
                # If the grade doesn't exist, create a new one
                hidden_grade = hidden_grades.objects.create(
                    course_id=course_id, student_id=student_id, semester_id=semester_id, grade=grade)
                hidden_grade.save()

            hidden_grade.save()

        return render(request, '../templates/examination/grades_updated.html', {})


class update_authentication(View):
    def post(self, request, *args, **kwargs):
        # Extract data from the POST request
        course = request.POST.get('course')
        course_year = request.POST.get('course_year')
        authenticator1 = request.POST.get('authenticator1')
        authenticator2 = request.POST.get('authenticator2')
        authenticator3 = request.POST.get('authenticator3')

        # Retrieve the registration object
        try:

            course_instance = Courses.objects.get(id=course)
            registration = authentication.objects.get(
                course_id=course_instance, course_year=course_year)
        except authentication.DoesNotExist:
            # Redirect if registration does not exist
            return redirect('examination:submitGrades')

        # Update authenticators if the values have changed
        if authenticator1 is not None:
            registration.authenticator_1 = (authenticator1 == '1')
        else:
            registration.authenticator_1 = 0
        if authenticator2 is not None:
            registration.authenticator_2 = (authenticator2 == '1')
        else:
            registration.authenticator_2 = 0
        if authenticator3 is not None:
            registration.authenticator_3 = (authenticator3 == '1')
        else:
            registration.authenticator_3 = 0

        # Save the changes
        registration.save()

        # Redirect to the appropriate page
        return redirect('examination:authenticate')


class DownloadExcelView(View):
    def post(self, request, *args, **kwargs):
        # Retrieve form data
        student_ids = request.POST.getlist('student_ids[]')
        semester_ids = request.POST.getlist('semester_ids[]')
        course_ids = request.POST.getlist('course_ids[]')
        grades = request.POST.getlist('grades[]')

        # Create a CSV response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="grades.csv"'

        # Write data to CSV
        writer = csv.writer(response)
        writer.writerow(['Student ID', 'Semester ID', 'Course ID', 'Grade'])
        for student_id, semester_id, course_id, grade in zip(student_ids, semester_ids, course_ids, grades):
            writer.writerow([student_id, semester_id, course_id, grade])

        return response

@login_required(login_url="/accounts/login")
def generate_transcript(request):
    des = request.session.get("currentDesignationSelected")
    if des == "acadadmin" :
        pass
    else:
        return HttpResponseRedirect('/dashboard/')

     
    student_id = request.GET.get('student')
    semester = request.GET.get('semester')
    courses_registered = Student_grades.objects.filter(
        roll_no=student_id, semester=semester)

    # Initialize a dictionary to store course grades
    course_grades = {}

    total_course_registered = Student_grades.objects.filter(
        roll_no=student_id, semester__lte=semester)
    # for each_course in total_course_registered:
    #     course_name = Curriculum.objects.filter(
    #         curriculum_id=each_course.curr_id_id)
    
    

    for course in courses_registered:
        try:
            # Attempt to fetch the grade for the course from Student_grades
            grade = Student_grades.objects.get(
                roll_no=student_id, course_id=course.course_id)

            # course_detail = Curriculum.objects.get(
            #     course_id=course.course_id, batch=grade.batch)
            course_instance = Courses.objects.get(id=course.course_id_id)
            # check_authentication_object = authentication.objects.filter(
            #     course_id=course_instance, course_year=grade.year)
            # all_authenticators_true = True

            # if check_authentication_object:
            #     # Iterate over each authentication object
            #     for auth_obj in check_authentication_object:
            #         # Check if all authenticators are true
            #         if not (auth_obj.authenticator_1 and auth_obj.authenticator_2 and auth_obj.authenticator_3):
            #             all_authenticators_true = False
            #             break  # No need to check further if any authenticator is False
            # else:
                # Create authentication object if it doesn't exist
                # authentication_object = authentication.objects.create(
                #     course_id=course_instance, course_year=grade.year)
                # Get all registrations for the course and year
                # registrations = authentication.objects.filter(
                #     course_id=course_instance, course_year=grade.year)
                # all_authenticators_true = False

            course_grades[course_instance] = {
                'grade': grade,
                # 'all_authenticators_true': all_authenticators_true
            }  # Store the grade
        except Student_grades.DoesNotExist:
            # Grade not available
            course_grades[course] = "Grading not done yet"

    context = {
        'courses_grades': course_grades,
        'total_course_registered':total_course_registered
    }
    #   print(context)

    return render(request, 'examination/generate_transcript.html', context)

# new

@login_required(login_url="/accounts/login")
def generate_transcript_form(request):
    des = request.session.get("currentDesignationSelected")
    if des == "acadadmin" :
        pass
    else:
        return HttpResponseRedirect('/dashboard/')
    if request.method == 'POST':
        programme = request.POST.get('programme')
        batch = request.POST.get('batch')
        specialization = request.POST.get('specialization')
        semester = request.POST.get('semester')

        if specialization == None:
            students = Student.objects.filter(
                programme=programme, batch=batch)
        else:
            students = Student.objects.filter(
                programme=programme, batch=batch, specialization=specialization)

        # Pass the filtered students to the template
        context = {
            'students': students,
            'semester': semester
        }
        return render(request, 'examination/generate_transcript_students.html', context)
    else:
        programmes = Student.objects.values_list(
            'programme', flat=True).distinct()
        specializations = Student.objects.exclude(
            specialization__isnull=True).values_list('specialization', flat=True).distinct()
        batches = Student.objects.values_list('batch', flat=True).distinct()
        context = {
            'programmes': programmes,
            'batches': batches,
            'specializations': specializations,
        }

        return render(request, 'examination/generate_transcript_form.html', context)


@login_required(login_url="/accounts/login")
def updateGrades(request):
    des = request.session.get("currentDesignationSelected")
    if des == "acadadmin":
        pass
    else:
        if request.is_ajax():
            return JsonResponse({"success": False, "error": "Access denied."}, status=403)
        else:  # For non-AJAX requests
            return HttpResponseRedirect('/dashboard/')
    unique_course_ids = Student_grades.objects.filter(verified=False).values("course_id").distinct()

    # Cast the course IDs to integers
    unique_course_ids = unique_course_ids.annotate(
        course_id_int=Cast("course_id", IntegerField())
    )

    # Retrieve course names and course codes based on unique course IDs

    # print(unique_course_ids)
    courses_info = Courses.objects.filter(
        id__in=unique_course_ids.values_list("course_id_int", flat=True)
    )

    unique_year_ids = Student_grades.objects.values('year').distinct()
    # print(unique_year_ids)
    context = {
        "courses_info": courses_info,
        "unique_year_ids": unique_year_ids,
    }

    return render(request, "../templates/examination/submitGrade.html", context)

@login_required(login_url="/accounts/login")
def updateEntergrades(request):
    des = request.session.get("currentDesignationSelected")
    if des == "acadadmin":
        pass
    else:
        if request.is_ajax():  # For AJAX or JSON requests
            return JsonResponse({"success": False, "error": "Access denied."}, status=403)
        else:  # For non-AJAX requests
            return HttpResponseRedirect('/dashboard/')
    course_id = request.GET.get("course")
    
    year = request.GET.get("year")
    # print(course_id,semester_id ,year)
    course_present = Student_grades.objects.filter(
        course_id=course_id, year=year
    )

    if not course_present:
        context = {"message": "THIS COURSE IS NOT SUBMITTED BY THE INSTRUCTOR"}
        return render(request, "../templates/examination/message.html", context)

    verification = course_present.first().verified
    print(verification)
    if verification:
        context = {"message": "THIS COURSE IS VERIFIED"}
        return render(request, "../templates/examination/message.html", context)

    context = {"registrations": course_present}

    return render(request, "../templates/examination/updateEntergrades.html", context)

class moderate_student_grades(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        des = request.session.get("currentDesignationSelected")
        if des == "acadadmin" or des=="Dean Academic":
         pass
        else:
         if request.is_ajax():  # For AJAX or JSON requests
            return JsonResponse({"success": False, "error": "Access denied."}, status=403)
         else:  # For non-AJAX requests
            return HttpResponseRedirect('/dashboard/')
        student_ids = request.POST.getlist('student_ids[]')
        semester_ids = request.POST.getlist('semester_ids[]')
        course_ids = request.POST.getlist('course_ids[]')
        grades = request.POST.getlist('grades[]')
        allow_resubmission = request.POST.get('allow_resubmission', 'NO')
        
        if len(student_ids) != len(semester_ids) != len(course_ids) != len(grades):
            return Response({'error': 'Invalid grade data provided'}, status=status.HTTP_400_BAD_REQUEST)

        for student_id, semester_id, course_id, grade in zip(student_ids, semester_ids, course_ids, grades):

            try:
                grade_of_student = Student_grades.objects.get(
                    course_id=course_id, roll_no=student_id, semester=semester_id)
                grade_of_student.grade = grade
                grade_of_student.verified = True
                if allow_resubmission == 'YES':
                    grade_of_student.reSubmit = True
                grade_of_student.save()
            except Student_grades.DoesNotExist:
                # If the grade doesn't exist, create a new one
                hidden_grade = hidden_grades.objects.create(
                    course_id=course_id, student_id=student_id, semester_id=semester_id, grade=grade)
                hidden_grade.save()

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="grades.csv"'

        # Write data to CSV
        writer = csv.writer(response)
        writer.writerow(['Student ID', 'Semester ID', 'Course ID', 'Grade'])
        for student_id, semester_id, course_id, grade in zip(student_ids, semester_ids, course_ids, grades):
            writer.writerow([student_id, semester_id, course_id, grade])
        print("HELLO")
        return response
        return render(request, '../templates/examination/grades_updated.html', {})


class submitGrades(APIView):
    login_url = "/accounts/login"
    
    def get(self, request):
        des = request.session.get("currentDesignationSelected")
        if des == "acadadmin":
         pass
        else:
         if request.is_ajax():  # For AJAX or JSON requests
            return JsonResponse({"success": False, "error": "Access denied."}, status=403)
         else:  # For non-AJAX requests
            return HttpResponseRedirect('/dashboard/')
        academic_year = request.GET.get('academic_year')
        
        if academic_year:
            if academic_year is None or not academic_year.isdigit():
                return JsonResponse({})
            # Filter course registration based on the academic year and get unique course IDs
            unique_course_ids = course_registration.objects.filter(
                working_year=academic_year
            ).values("course_id").distinct()
            
            # Cast the course IDs to integers
            unique_course_ids = unique_course_ids.annotate(
                course_id_int=Cast("course_id", IntegerField())
            )
            
            # Retrieve course information based on the unique course IDs
            courses_info = Courses.objects.filter(
              id__in=unique_course_ids.values_list("course_id_int", flat=True)
            ).order_by('code')
            
            # Return the course information as JSON response
            return JsonResponse({"courses": list(courses_info.values())})
        
        # If no academic year is provided, return the working years for the dropdown
        working_years = course_registration.objects.values("working_year").distinct()
        
        context = {"working_years": working_years}
        
        return render(request, "../templates/examination/gradeSubmission.html", context)

@login_required(login_url="/accounts/login")
def submitEntergrades(request):

    course_id = request.GET.get("course")
    year = request.GET.get("year")
    if year is None or not year.isdigit():
        message = "YEAR SHOULD NOT BE NONE"
        context = {"message": message}

        return render(request, "../templates/examination/message.html", context)
        return HttpResponse("Invalid year parameter")
        # Handle invalid year parameter
        # You can return an error response or redirect the user to an error page
    courses_info = Courses.objects.get(id=course_id)

    courses = Student_grades.objects.filter(course_id=courses_info.id, year=year)

    if courses:
        message = "THIS Course was Already Submitted"
        context = {"message": message}

        return render(request, "../templates/examination/message.html", context)

    students = course_registration.objects.filter(
        course_id_id=course_id, working_year=year
    )

    # print(students)

    context = {"registrations": students, "curr_id": course_id, "year": year}

    return render(request, "../templates/examination/gradeSubmissionForm.html", context)


class submitEntergradesStoring(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        student_ids = request.POST.getlist("student_ids[]")
        batch_ids = request.POST.getlist("batch_ids[]")
        course_ids = request.POST.getlist("course_ids[]")
        semester_ids = request.POST.getlist("semester_ids[]")
        year_ids = request.POST.getlist("year_ids[]")
        marks = request.POST.getlist("marks[]")
        Student_grades = request.POST.getlist("Student_grades[]")

        if (
            len(student_ids)
            != len(batch_ids)
            != len(course_ids)
            != len(semester_ids)
            != len(year_ids)
            != len(marks)
            != len(Student_grades)
        ):
            return Response(
                {"error": "Invalid Student_grades data provided"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        for (
            student_id,
            batch_id,
            course_id,
            semester_id,
            year_id,
            mark,
            Student_grades,
        ) in zip(
            student_ids,
            batch_ids,
            course_ids,
            semester_ids,
            year_ids,
            marks,
            Student_grades,
        ):
            # Create an instance of hidden_grades model and save the data

            try:
                grade_of_student = Student_grades.objects.get(
                    course_id=course_id, roll_no=student_id, semester=semester_id
                )
            except Student_grades.DoesNotExist:
                # If the Student_grades doesn't exist, create a new one
                course_instance = Courses.objects.get(id=course_id)
                student_grade = Student_grades.objects.create(
                    course_id=course_instance,
                    roll_no=student_id,
                    semester=semester_id,
                    Student_grades=Student_grades,
                    batch=batch_id,
                    year=year_id,
                    total_marks=mark,
                )
                student_grade.save()

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="Student_grades.csv"'

        # Write data to CSV
        writer = csv.writer(response)
        writer.writerow(
            [
                "student_id",
                "batch_ids",
                "course_id",
                "semester_id",
                "year_ids",
                "marks",
                "Student_grades",
            ]
        )
        for (
            student_id,
            batch_id,
            course_id,
            semester_id,
            year_id,
            mark,
            Student_grades,
        ) in zip(
            student_ids,
            batch_ids,
            course_ids,
            semester_ids,
            year_ids,
            marks,
            Student_grades,
        ):
            writer.writerow(
                [
                    student_id,
                    batch_id,
                    course_id,
                    semester_id,
                    year_id,
                    mark,
                    Student_grades,
                ]
            )

        return response
        return render(request, "../templates/examination/grades_updated.html", {})

@login_required(login_url="/accounts/login")
def upload_grades(request):
    if request.method == "POST" and request.FILES.get("csv_file"):
        des = request.session.get("currentDesignationSelected")
        if des == "acadadmin":
         pass
        else:
         if request.is_ajax():  # For AJAX or JSON requests
            return JsonResponse({"success": False, "error": "Access denied."}, status=403)
         else:  # For non-AJAX requests
            return HttpResponseRedirect('/dashboard/')
        csv_file = request.FILES["csv_file"]

        if not csv_file.name.endswith(".csv"):
            return JsonResponse(
                {"error": "Invalid file format. Please upload a CSV file."}, status=400
            )

        course_id = request.POST.get("course_id")
        academic_year = request.POST.get("academic_year")
        # semester = request.POST.get('semester')

        if academic_year == "None" or not academic_year.isdigit():
            return JsonResponse(
                {"error": "Academic year must be a valid number."}, status=400
            )

        if not course_id or not academic_year:
            return JsonResponse(
                {"error": "Course ID and Academic Year are required."}, status=400
            )

        courses_info = Courses.objects.get(id=course_id)

        courses = Student_grades.objects.filter(
            course_id=courses_info.id, year=academic_year
        )
        students = course_registration.objects.filter(
            course_id_id=course_id, working_year=academic_year
        )

        if not students:
            message = "NO STUDENTS REGISTERED IN THIS COURSE THIS SEMESTER"
            redirect_url = reverse("examination:message") + f"?message={message}"
            return JsonResponse(
                {"error": message, "redirect_url": redirect_url}, status=400
            )

        if courses and not courses.first().reSubmit:
            
            message = "THIS Course was Already Submitted"
            redirect_url = reverse("examination:message") + f"?message={message}"
            return JsonResponse(
                {"error": message, "redirect_url": redirect_url}, status=400
            )

        

        try:
            # Parse the CSV file
            decoded_file = csv_file.read().decode("utf-8").splitlines()
            reader = csv.DictReader(decoded_file)

            required_columns = ["roll_no", "grade", "remarks"]
            if not all(column in reader.fieldnames for column in required_columns):
                return JsonResponse(
                    {
                        "error": "CSV file must contain the following columns: roll_no, grade, remarks."
                    },
                    status=400,
                )

            for row in reader:
                roll_no = row["roll_no"]
                grade = row["grade"]
                remarks = row["remarks"]
                semester = row["semester"] if "semester" in row and row["semester"] else None
                stud = Student.objects.get(id_id=roll_no)
                semester = semester or stud.curr_semester_no
                batch=stud.batch
                reSubmit=False

                Student_grades.objects.update_or_create(
                 roll_no=roll_no,
                 course_id_id=course_id,
                 year=academic_year,
                 semester=semester,
                 batch=batch,
        # Fields that will be updated if a match is found
                 defaults={
                    'grade': grade,
                    'remarks': remarks,
                    'reSubmit': reSubmit,
                }
                )
            des = request.session.get("currentDesignationSelected")
            if (
             str(des) == "Associate Professor"
             or str(des) == "Professor"
             or str(des) == "Assistant Professor"
              ):
             return JsonResponse(
                {
                    "message": "Grades uploaded successfully.",
                    "redirect_url": "/examination/submitGradesProf",
                }
             )
            return JsonResponse(
                {
                    "message": "Grades uploaded successfully.",
                    "redirect_url": "/examination/submitGrades",
                }
            )

        except Courses.DoesNotExist:
            return JsonResponse({"error": "Invalid course ID."}, status=400)

        except Exception as e:
            return JsonResponse({"error": f"An error occurred: {e}"}, status=500)

    return JsonResponse(
        {"error": "Invalid request. Please upload a CSV file."}, status=400
    )

@login_required(login_url="/accounts/login")
def show_message(request):
    des = request.session.get("currentDesignationSelected")
    if des == "acadadmin" or str(des) == "Associate Professor" or str(des) == "Professor" or str(des) == "Assistant Professor" :
         pass
    else:
         if request.is_ajax():  # For AJAX or JSON requests
            return JsonResponse({"success": False, "error": "Access denied."}, status=403)
         else:  # For non-AJAX requests
            return HttpResponseRedirect('/dashboard/')
    message = request.GET.get("message", "Default message if none provided.")
    des = request.session.get("currentDesignationSelected")
    if (
        str(des) == "Associate Professor"
        or str(des) == "Professor"
        or str(des) == "Assistant Professor"
    ):
        return render(request, "examination/messageProf.html", {"message": message})
    return render(request, "examination/message.html", {"message": message})


@login_required(login_url="/accounts/login")
def submitGradesProf(request):
    des = request.session.get("currentDesignationSelected")
    if  str(des) == "Associate Professor" or str(des) == "Professor" or str(des) == "Assistant Professor" :
         pass
    else:
         if request.is_ajax():  # For AJAX or JSON requests
            return JsonResponse({"success": False, "error": "Access denied."}, status=403)
         else:  # For non-AJAX requests
            return HttpResponseRedirect('/dashboard/')
    # print(request.user,1)
    unique_course_ids = (
        CourseInstructor.objects.filter(instructor_id_id=request.user.username)
        .values("course_id_id")
        .distinct()
    )
    # unique_course_ids = course_registration.objects.values(
    #     'course_id').distinct()
    working_years = course_registration.objects.values("working_year").distinct()
    course_ids_final=course_registration.objects.filter()
    # Cast the course IDs to integers
    unique_course_ids = unique_course_ids.annotate(
        course_id_int=Cast("course_id", IntegerField())
    )

    # Retrieve course names and course codes based on unique course IDs

    # print(unique_course_ids)
    courses_info = Courses.objects.filter(
        id__in=unique_course_ids.values_list("course_id_int", flat=True)
    )

    context = {"courses_info": courses_info, "working_years": working_years}

    print(working_years)

    return render(request, "../templates/examination/submitGradesProf.html", context)

@login_required(login_url="/accounts/login")
def download_template(request):
    des = request.session.get("currentDesignationSelected")
    if des not in ["acadadmin", "Associate Professor", "Professor", "Assistant Professor", "Dean Academic"]:
        if request.is_ajax():
            return JsonResponse({"success": False, "error": "Access denied."}, status=403)
        else:
            return HttpResponseRedirect('/dashboard/')
    
    course = request.GET.get('course')
    year = request.GET.get('year')

    if not course or not year:
        return JsonResponse({'error': 'Course and year are required'}, status=400)
    
    try:
        # Fetching the custom user model
        User = get_user_model()
        
        course_info = course_registration.objects.filter(
            course_id_id=course, 
            working_year=year
        )
        
        if not course_info.exists():
            return JsonResponse({'error': 'No registration data found for the provided course and year'}, status=404)

        course_obj = course_info.first().course_id
        response = HttpResponse(content_type="text/csv")
        filename = f"{course_obj.code}_template_{year}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Create CSV writer
        writer = csv.writer(response)
        
        # Write header
        writer.writerow(["roll_no", "name", "grade", "remarks"])
        
        # Write student roll numbers and names
        for entry in course_info:
            student_entry = entry.student_id
            # Fetching the user instance dynamically
            student_user = User.objects.get(username=student_entry.id_id)
            writer.writerow([student_entry.id_id, student_user.first_name+" "+student_user.last_name, "", ""])
        
        return response
    
    except Exception as e:
        print(f"Error in download_template: {str(e)}")
        return JsonResponse({'error': 'An unexpected error occurred'}, status=500)


@login_required(login_url="/accounts/login")
def verifyGradesDean(request):
    des = request.session.get("currentDesignationSelected")
    if des=="Dean Academic" :
         pass
    else:
         if request.is_ajax():  # For AJAX or JSON requests
            return JsonResponse({"success": False, "error": "Access denied."}, status=403)
         else:  # For non-AJAX requests
            return HttpResponseRedirect('/dashboard/')
    unique_course_ids = Student_grades.objects.filter(verified=True).values("course_id").distinct()

    # Cast the course IDs to integers
    unique_course_ids = unique_course_ids.annotate(
        course_id_int=Cast("course_id", IntegerField())
    )

    # Retrieve course names and course codes based on unique course IDs

    # print(unique_course_ids)
    courses_info = Courses.objects.filter(
        id__in=unique_course_ids.values_list("course_id_int", flat=True)
    ).order_by("code")

    unique_year_ids = Student_grades.objects.values("year").distinct()

    context = {
        "courses_info": courses_info,
        "unique_year_ids": unique_year_ids,
    }

    return render(request, "../templates/examination/submitGradeDean.html", context)

@login_required(login_url="/accounts/login")
def updateEntergradesDean(request):
    des = request.session.get("currentDesignationSelected")
    if des=="Dean Academic" :
         pass
    else:
         if request.is_ajax():  # For AJAX or JSON requests
            return JsonResponse({"success": False, "error": "Access denied."}, status=403)
         else:  # For non-AJAX requests
            return HttpResponseRedirect('/dashboard/')
    course_id = request.GET.get("course")
    year = request.GET.get("year")
    course_present = Student_grades.objects.filter(
        course_id=course_id, year=year
    )

    if not course_present:
        context = {"message": "THIS COURSE IS NOT SUBMITTED BY THE INSTRUCTOR"}
        return render(request, "../templates/examination/messageDean.html", context)

    context = {"registrations": course_present}

    return render(request, "../templates/examination/updateEntergradesDean.html", context)


@login_required(login_url="/accounts/login")
def upload_grades_prof(request):
    des = request.session.get("currentDesignationSelected")
    if str(des) == "Associate Professor" or str(des) == "Professor" or str(des) == "Assistant Professor" :
         pass
    else:
         if request.is_ajax():  # For AJAX or JSON requests
            return JsonResponse({"success": False, "error": "Access denied."}, status=403)
         else:  # For non-AJAX requests
            return HttpResponseRedirect('/dashboard/')
    if request.method == "POST" and request.FILES.get("csv_file"):
        csv_file = request.FILES["csv_file"]

        if not csv_file.name.endswith(".csv"):
            return JsonResponse(
                {"error": "Invalid file format. Please upload a CSV file."}, status=400
            )

        course_id = request.POST.get("course_id")
        academic_year = request.POST.get("academic_year")
        # semester = request.POST.get('semester')

        if academic_year == "None" or not academic_year.isdigit():
            return JsonResponse(
                {"error": "Academic year must be a valid number."}, status=400
            )

        if not course_id or not academic_year:
            return JsonResponse(
                {"error": "Course ID and Academic Year are required."}, status=400
            )

        courses_info = Courses.objects.get(id=course_id)

        courses = Student_grades.objects.filter(
            course_id=courses_info.id, year=academic_year
        )
        students = course_registration.objects.filter(
            course_id_id=course_id, working_year=academic_year
        )

        if not students:
            message = "NO STUDENTS REGISTERED IN THIS COURSE THIS SEMESTER"
            redirect_url = reverse("examination:message") + f"?message={message}"
            return JsonResponse(
                {"error": message, "redirect_url": redirect_url}, status=400
            )
        
        if courses and not courses.first().reSubmit:
            
            message = "THIS Course was Already Submitted"
            redirect_url = reverse("examination:message") + f"?message={message}"
            return JsonResponse(
                {"error": message, "redirect_url": redirect_url}, status=400
            )

        

        try:
            # Parse the CSV file
            decoded_file = csv_file.read().decode("utf-8").splitlines()
            reader = csv.DictReader(decoded_file)

            required_columns = ["roll_no", "grade", "remarks"]
            if not all(column in reader.fieldnames for column in required_columns):
                return JsonResponse(
                    {
                        "error": "CSV file must contain the following columns: roll_no, grade, remarks."
                    },
                    status=400,
                )

            for row in reader:
                roll_no = row["roll_no"]
                grade = row["grade"]
                remarks = row["remarks"]
                semester = row["semester"] if "semester" in row and row["semester"] else None
                stud = Student.objects.get(id_id=roll_no)
                semester = semester or stud.curr_semester_no
                batch=stud.batch
                reSubmit=False

                Student_grades.objects.update_or_create(
                 roll_no=roll_no,
                 course_id_id=course_id,
                 year=academic_year,
                 semester=semester,
                 batch=batch,
        # Fields that will be updated if a match is found
                 defaults={
                    'grade': grade,
                    'remarks': remarks,
                    'reSubmit': reSubmit,
                }
                )
            des = request.session.get("currentDesignationSelected")
            if (
             str(des) == "Associate Professor"
             or str(des) == "Professor"
             or str(des) == "Assistant Professor"
              ):
             return JsonResponse(
                {
                    "message": "Grades uploaded successfully.",
                    "redirect_url": "/examination/submitGradesProf",
                }
             )
            return JsonResponse(
                {
                    "message": "Grades uploaded successfully.",
                    "redirect_url": "/examination/submitGradesProf",
                }
            )

        except Courses.DoesNotExist:
            return JsonResponse({"error": "Invalid course ID."}, status=400)

        except Exception as e:
            return JsonResponse({"error": f"An error occurred: {e}"}, status=500)

    return JsonResponse(
        {"error": "Invalid request. Please upload a CSV file."}, status=400
    )


@login_required(login_url="/accounts/login")
def validateDean(request):
    des = request.session.get("currentDesignationSelected")
    if des=="Dean Academic" :
         pass
    else:
         if request.is_ajax():  # For AJAX or JSON requests
            return JsonResponse({"success": False, "error": "Access denied."}, status=403)
         else:  # For non-AJAX requests
            return HttpResponseRedirect('/dashboard/')
    unique_course_ids = Student_grades.objects.filter(verified=True).values("course_id").distinct()

    # Cast the course IDs to integers
    unique_course_ids = unique_course_ids.annotate(
        course_id_int=Cast("course_id", IntegerField())
    )

    # Retrieve course names and course codes based on unique course IDs

    # print(unique_course_ids)
    courses_info = Courses.objects.filter(
        id__in=unique_course_ids.values_list("course_id_int", flat=True)
    )
    working_years = course_registration.objects.values("working_year").distinct()

    unique_batch_ids = Student_grades.objects.values("batch").distinct()

    context = {"courses_info": courses_info, "working_years": working_years}

    return render(request, "../templates/examination/validation.html", context)

@login_required(login_url="/accounts/login")
def validateDeanSubmit(request):
    des = request.session.get("currentDesignationSelected")
    if des=="Dean Academic" :
         pass
    else:
         if request.is_ajax():  # For AJAX or JSON requests
            return JsonResponse({"success": False, "error": "Access denied."}, status=403)
         else:  # For non-AJAX requests
            return HttpResponseRedirect('/dashboard/')
    if request.method == "POST" and request.FILES.get("csv_file"):
        csv_file = request.FILES["csv_file"]

        if not csv_file.name.endswith(".csv"):
            message = "Please Submit a csv file "
            context = {
                 "message":message,
                }
            return render(request, "../templates/examination/messageDean.html", context)

        course_id = request.POST.get("course")
        academic_year = request.POST.get("year")
        # semester = request.POST.get('semester')
        print(academic_year)
        if academic_year is None or not academic_year.isdigit():
            message = "Academic Year must be a number"
            context = {
                 "message":message,
                }
            return render(request, "../templates/examination/messageDean.html", context)

        if not course_id or not academic_year:
            message = "Course and Academic year are required"
            context = {
                 "message":message,
                }
            return render(request, "../templates/examination/messageDean.html", context)

        # courses_info = Courses.objects.get(id=course_id)

        # courses = Student_grades.objects.filter(
        #     course_id=courses_info.id, year=academic_year
        # )
        students = course_registration.objects.filter(
            course_id_id=course_id, working_year=academic_year
        )
        
        
        
        try:
            # Parse the CSV file
            decoded_file = csv_file.read().decode("utf-8").splitlines()
            reader = csv.DictReader(decoded_file)

            required_columns = ["roll_no", "grade", "remarks"]
            if not all(column in reader.fieldnames for column in required_columns):
                message = "CSV file must contain the following columns: roll_no, grade, remarks."
                context = {
                 "message":message,
                }
                return render(request, "../templates/examination/messageDean.html", context)
                   
            
            mismatch=[]
            for row in reader:
                roll_no = row["roll_no"]
                grade = row["grade"]
                remarks = row["remarks"]
                stud=Student.objects.get(id_id=roll_no)
                semester=stud.curr_semester_no
                batch=stud.batch
                Student_grades.objects.filter(
                 roll_no=roll_no,
                 course_id_id=course_id,
                 year=academic_year,
                 batch=batch,
                )
                student_grade = Student_grades.objects.get(
                roll_no=roll_no,
                course_id_id=course_id,
                year=academic_year,
                batch=batch
                )
                if student_grade.grade != grade:
                 mismatch.append({
                    "roll_no": roll_no,
                    "csv_grade": grade,
                    "db_grade": student_grade.grade,
                    "remarks": remarks,
                    "batch": batch,
                    "semester": semester,
                    "course_id": course_id, 
                 })
            if not mismatch:
                message = "There Are no Mismatches"
                context = {
                 "message":message,
                }
                return render(request, "../templates/examination/messageDean.html", context)
            context = {
                 "mismatch": mismatch,
                }
            return render(request, "../templates/examination/validationSubmit.html", context)
            
        except Exception as e:

            error_message = f"An error occurred while processing the file: {str(e)}"
            context = {
                "message": error_message,
            }
            return render(request, "../templates/examination/messageDean.html", context)

@login_required(login_url="/accounts/login")
def downloadGrades(request):
  des = request.session.get("currentDesignationSelected")
  if str(des) == "Associate Professor" or str(des) == "Professor" or str(des) == "Assistant Professor" :
         pass
  else:
         if request.is_ajax():  # For AJAX or JSON requests
            return JsonResponse({"success": False, "error": "Access denied."}, status=403)
         else:  # For non-AJAX requests
            return HttpResponseRedirect('/dashboard/')
  academic_year = request.GET.get('academic_year')
        
  if academic_year:
    if academic_year is None or not academic_year.isdigit():
     return JsonResponse({})
    # print(request.user,1)
    unique_course_ids = (
        CourseInstructor.objects.filter(instructor_id_id=request.user.username)
        .values("course_id_id")
        .distinct()
    )
    # unique_course_ids = course_registration.objects.values(
    #     'course_id').distinct()
    
    # Cast the course IDs to integers
    unique_course_ids = unique_course_ids.annotate(
        course_id_int=Cast("course_id", IntegerField())
    )
    # Retrieve course names and course codes based on unique course IDs

    courses_info = Student_grades.objects.filter(
        year=academic_year,
        course_id_id__in=unique_course_ids.values_list("course_id_int", flat=True)
    )
    courses_details=Courses.objects.filter(
        id__in=courses_info.values_list("course_id_id", flat=True)
    )
    # print(courses_info.values(),'abcd')
    return JsonResponse({"courses": list(courses_details.values())})
    
    
  working_years = course_registration.objects.values("working_year").distinct()
        
  context = {"working_years": working_years}
        
  return render(request, "../templates/examination/download_resultProf.html", context)


@login_required(login_url="/accounts/login")
def generate_pdf(request):
    des = request.session.get("currentDesignationSelected")
    if str(des) == "Associate Professor" or str(des) == "Professor" or str(des) == "Assistant Professor" :
         pass
    else:
         if request.is_ajax():  # For AJAX or JSON requests
            return JsonResponse({"success": False, "error": "Access denied."}, status=403)
         else:  # For non-AJAX requests
            return HttpResponseRedirect('/dashboard/')
    course_id = request.POST.get('course_id')
    academic_year = request.POST.get('academic_year')
    course_info = get_object_or_404(Courses, id=course_id)
    grades = Student_grades.objects.filter(course_id_id=course_id, year=academic_year).order_by("roll_no")
    course=CourseInstructor.objects.filter(course_id_id=course_id,year=academic_year,instructor_id_id=request.user.username)
    if not course:
         return JsonResponse({"success": False, "error": "course not found."}, status=404)
    semester=course.first().semester_no
    
    all_grades = ["O", "A+", "A", "B+", "B", "C+", "C", "D+", "D", "F", "I", "S", "X"]
    grade_counts = {grade: grades.filter(grade=grade).count() for grade in all_grades}

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{course_info.code}_grades.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # Custom Header Style
    header_style = ParagraphStyle(
    "HeaderStyle",
    parent=styles["Heading1"],
    fontName="Helvetica-Bold",
    fontSize=16,
    textColor=HexColor("#333333"),
    spaceAfter=20,
    alignment=1,  # Center alignment
    )
    subheader_style = ParagraphStyle(
        "SubheaderStyle",
        parent=styles["Normal"],
        fontSize=12,
        textColor=HexColor("#666666"),
        spaceAfter=10,
    )
    instructor = request.user.first_name + " " + request.user.last_name

    # Add Header
    elements.append(Paragraph(f"Grade Sheet", header_style))
    field_label_style = ParagraphStyle(
    "FieldLabelStyle",
    parent=styles["Normal"],
    fontSize=12,
    textColor=colors.black,  # Black text color for labels
    spaceAfter=5,
)
    field_value_style = ParagraphStyle(
    "FieldValueStyle",
    parent=styles["Normal"],
    fontSize=12,
    textColor=HexColor("#666666"),  # Gray text color for values
    spaceAfter=10,
)

# Add fields with labels in black and values in gray
    elements.append(Paragraph(f"<b>Session:</b> {academic_year}", field_label_style))
    elements.append(Paragraph(f"<b>Semester:</b> {semester}", field_label_style))
    elements.append(Paragraph(f"<b>Course Code:</b> {course_info.code}", field_label_style))
    elements.append(Paragraph(f"<b>Course Name:</b> {course_info.name}", field_label_style))
    elements.append(Paragraph(f"<b>Instructor:</b> {instructor}", field_label_style))

    # Table Data with Wider Column Widths
    data = [["S.No.", "Roll Number", "Grade"]]
    for i, grade in enumerate(grades, 1):
        data.append([i, grade.roll_no, grade.grade])
    table = Table(data, colWidths=[80, 300, 100])

    # Improved Table Style
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), HexColor("#E0E0E0")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 14),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                ("BACKGROUND", (0, 1), (-1, -1), HexColor("#F9F9F9")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [HexColor("#F9F9F9"), colors.white]),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 12),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elements.append(table)
    elements.append(Spacer(1, 20))

    # Add Grade Distribution with Row Splitting
    elements.append(Paragraph(f"Grade Distribution:", header_style))

    # First Grade Table
    grade_data1 = [["O", "A+", "A", "B+", "B", "C+", "C", "D+"]]
    grade_data1.append([grade_counts[grade] for grade in grade_data1[0]])
    grade_table1 = Table(grade_data1, colWidths=[60] * 8)
    grade_table1.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), HexColor("#E0E0E0")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 12),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
            ]
        )
    )
    elements.append(grade_table1)
    elements.append(Spacer(1, 10))

    # Second Grade Table
    grade_data2 = [["D", "F", "I", "S", "X"]]
    grade_data2.append([grade_counts[grade] for grade in grade_data2[0]])
    grade_table2 = Table(grade_data2, colWidths=[60] * 5)
    grade_table2.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), HexColor("#E0E0E0")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 12),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
            ]
        )
    )
    elements.append(grade_table2)
    elements.append(Spacer(1, 40))

    verified_style = ParagraphStyle(
    "VerifiedStyle",
    parent=styles["Normal"],
    fontSize=13,
    textColor=HexColor("#333333"),
    alignment=0,  # Center alignment
    spaceAfter=20,
      )
    elements.append(Paragraph("I have carefully checked and verified the submitted grade. The grade distribution and submitted grades are correct. [Please mention any exception below.]", verified_style))

    # Footer Signatures
    def draw_signatures(canvas, doc):
        canvas.saveState()
        width, height = letter
        canvas.drawString(inch, 0.75 * inch, "")
        canvas.drawString(inch, 0.5 * inch, "Date")
        canvas.drawString(width - 4 * inch, 0.75 * inch, "")
        canvas.drawString(width - 4 * inch, 0.5 * inch, "Course Instructor's Signature")
        canvas.restoreState()

    doc.build(elements, onLaterPages=draw_signatures, onFirstPage=draw_signatures)
    return response


@login_required(login_url="/accounts/login")
def generate_result(request):
    if request.method == 'POST':
        des = request.session.get("currentDesignationSelected")
        if des == "acadadmin":
         pass
        else:
         if request.is_ajax():  # For AJAX or JSON requests
            return JsonResponse({"success": False, "error": "Access denied."}, status=403)
         else:  # For non-AJAX requests
            return HttpResponseRedirect('/dashboard/')
        try:
            data = json.loads(request.body)
            semester = data.get('semester')
            branch = data.get('specialization')
            batch = data.get('batch')

            branch_info = Discipline.objects.filter(acronym=branch).first()
            if not branch_info:
                return JsonResponse({'error': 'Branch not found'}, status=404)

            curriculum_id = Batch.objects.filter(
                year=batch, discipline_id=branch_info.id
            ).values_list('curriculum_id', flat=True).first()
            if not curriculum_id:
                return JsonResponse({'error': 'Curriculum not found'}, status=404)

            semester_info = Semester.objects.filter(
                curriculum_id=curriculum_id, semester_no=semester
            ).first()
            if not semester_info:
                return JsonResponse({'error': 'Semester not found'}, status=404)
            # print(batch, branch)
            course_slots = CourseSlot.objects.filter(semester_id=semester_info)
            course_ids_from_slots = course_slots.values_list('courses', flat=True)
            course_ids_from_grades = Student_grades.objects.filter(
            batch=batch,
            semester=semester
             ).values_list('course_id_id', flat=True)
            course_ids = set(course_ids_from_slots).union(set(course_ids_from_grades))
            courses = Courses.objects.filter(id__in=course_ids)
            courses_map={}
            for course in courses:
                courses_map[course.id]=(course.credit)
            students = Student.objects.filter(batch=batch, specialization=branch).order_by('id')
            # print(students.first().id_id,"studejt id")
      
            wb = Workbook()
            ws = wb.active
            ws.title = "Student Grades"

        
            # ws.merge_cells(start_row=1, start_column=1, end_row=4, end_column=1) 
            # ws.merge_cells(start_row=1, start_column=2, end_row=4, end_column=2) 
            ws["A1"] = "S. No"
            ws["B1"] = "Roll No"
            for cell in ("A1", "B1"):
                ws[cell].alignment = Alignment(horizontal="center", vertical="center")
                ws[cell].font = Font(bold=True)

      
            ws.column_dimensions[get_column_letter(1)].width = 12  
            ws.column_dimensions[get_column_letter(2)].width = 18 
            col_idx = 3
            for course in courses:
           
                ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 1)
                ws.merge_cells(start_row=2, start_column=col_idx, end_row=2, end_column=col_idx + 1)
                ws.merge_cells(start_row=3, start_column=col_idx, end_row=3, end_column=col_idx + 1)

                ws.cell(row=1, column=col_idx).value = course.code
                ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=1, column=col_idx).font = Font(bold=True)
                ws.cell(row=2, column=col_idx).value = course.name
                ws.cell(row=2, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=2, column=col_idx).font = Font(bold=True)
                
                ws.cell(row=3, column=col_idx).value=course.credit
                ws.cell(row=3, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=3, column=col_idx).font = Font(bold=True)
                ws.cell(row=4, column=col_idx).value = "Grade"
                ws.cell(row=4, column=col_idx + 1).value = "Remarks"
                ws.cell(row=4, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=4, column=col_idx+1).alignment = Alignment(horizontal="center", vertical="center")
                ws.column_dimensions[get_column_letter(col_idx)].width = 25
                ws.column_dimensions[get_column_letter(col_idx+1)].width = 25 
                col_idx += 2

            # ws.merge_cells(start_row=1, start_column=col_idx, end_row=4, end_column=col_idx)  # SPI
            ws.cell(row=1, column=col_idx).value = "SPI"
            ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=1, column=col_idx).font = Font(bold=True)

            # ws.merge_cells(start_row=1, start_column=col_idx + 1, end_row=4, end_column=col_idx + 1)  # CPI
            ws.cell(row=1, column=col_idx + 1).value = "CPI"
            ws.cell(row=1, column=col_idx + 1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=1, column=col_idx + 1).font = Font(bold=True)

         
            row_idx = 5
            for idx, student in enumerate(students, start=1):
                ws.cell(row=row_idx, column=1).value = idx
                ws.cell(row=row_idx, column=2).value = student.id_id

                ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center", vertical="center")
                student_grades = Student_grades.objects.filter(
                    roll_no=student.id_id, course_id_id__in=course_ids, semester=semester
                )
               
                grades_map = {}
                for grade in student_grades:
                    grades_map[grade.course_id_id] = (grade.grade, grade.remarks,courses_map.get(grade.course_id_id) )

                col_idx = 3
                gained_credit=0
                total_credit=0
                for course in courses:
                    grade, remark, credits = grades_map.get(course.id, ("N/A", "N/A",0))
                    ws.cell(row=row_idx, column=col_idx).value = grade
                    ws.cell(row=row_idx, column=col_idx + 1).value = remark
                    ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=row_idx, column=col_idx+1).alignment = Alignment(horizontal="center", vertical="center")
                    if grade=="O" or grade=="A+":
                        gained_credit+=1*credits
                        total_credit+=credits
                    elif grade=="A":
                        gained_credit+=0.9*credits
                        total_credit+=credits
                    elif grade=="B+":
                        gained_credit+=0.8*credits
                        total_credit+=credits
                    elif grade=="B":
                        gained_credit+=0.7*credits
                        total_credit+=credits
                    elif grade=="C+":
                        gained_credit+=0.6*credits
                        total_credit+=credits
                    elif grade=="C":
                        gained_credit+=0.5*credits
                        total_credit+=credits
                    elif grade=="D+":
                        gained_credit+=0.4*credits
                        total_credit+=credits
                    elif grade=="D":
                        gained_credit+=0.3*credits
                        total_credit+=credits
                    elif grade=="F":
                        gained_credit+=0.2*credits
                        total_credit+=credits
                    
                    
                    col_idx += 2
                if total_credit==0 :
                    ws.cell(row=row_idx, column=col_idx).value =0
                else:
                 ws.cell(row=row_idx, column=col_idx).value = 10*(gained_credit/total_credit)
                ws.cell(row=row_idx, column=col_idx + 1).value = 0
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=row_idx, column=col_idx+1).alignment = Alignment(horizontal="center", vertical="center")

                row_idx += 1

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="student_grades.xlsx"'
            wb.save(response)
            return response

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)
            

    return JsonResponse({'error': 'Invalid request method'}, status=405)

def checkresult(request):
    des = request.session.get("currentDesignationSelected")
    if des == "student":
        pass
    else:
        if request.is_ajax():  # For AJAX or JSON requests
            return JsonResponse({"success": False, "error": "Access denied."}, status=403)
        else:  # For non-AJAX requests
            return HttpResponseRedirect('/dashboard/')
    return render(request, "../templates/examination/check_result.html")


def grades_report(request):
    if request.method == 'POST':
        des = request.session.get("currentDesignationSelected")
        if des == "student":
         pass
        else:
         if request.is_ajax():  # For AJAX or JSON requests
            return JsonResponse({"success": False, "error": "Access denied."}, status=403)
         else:  # For non-AJAX requests
            return HttpResponseRedirect('/dashboard/')
        roll_number=request.user.username
        semester=request.POST.get('semester')
        grades_info=Student_grades.objects.filter(roll_no=roll_number,semester=semester).select_related('course_id')
        gained_credit=0
        total_credit=0
        all_credits=0
        for grades in grades_info:
            credits=grades.course_id.credit
            grade=grades.grade
            if grade=="O" or grade=="A+":
                        gained_credit+=1*credits
                        total_credit+=credits
            elif grade=="A":
                        gained_credit+=0.9*credits
                        total_credit+=credits
            elif grade=="B+":
                        gained_credit+=0.8*credits
                        total_credit+=credits
            elif grade=="B":
                        gained_credit+=0.7*credits
                        total_credit+=credits
            elif grade=="C+":
                        gained_credit+=0.6*credits
                        total_credit+=credits
            elif grade=="C":
                        gained_credit+=0.5*credits
                        total_credit+=credits
            elif grade=="D+":
                        gained_credit+=0.4*credits
                        total_credit+=credits
            elif grade=="D":
                        gained_credit+=0.3*credits
                        total_credit+=credits
            elif grade=="F":
                        gained_credit+=0.2*credits
                        total_credit+=credits 
            all_credits+=credits
        spi = 10*(gained_credit/total_credit) if total_credit > 0 else 0
        all_grades = Student_grades.objects.filter(roll_no=roll_number)
        total_units = sum(grade.course_id.credit for grade in all_grades)
        student_user = request.user.first_name+' '+request.user.last_name 
        student = Student.objects.filter(id=roll_number).first()
        # count=Student.objects.filter(id=roll.id).count()
        # student=students.count()
        
        context = {
                'student': student,
                'student_user': student_user,
                'grades': grades_info,
                'spi': round(spi, 2),
                'semester_units': all_credits,
                'total_units': total_units,
            }
        return render(request, '../templates/examination/grades_report.html', context)


