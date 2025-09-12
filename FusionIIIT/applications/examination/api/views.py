from django.db.models.query_utils import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, HttpResponse
from django.http import HttpResponse
from django.http import JsonResponse
from decimal import Decimal, ROUND_HALF_UP
from applications.academic_procedures.models import(course_registration, course_replacement)
from applications.programme_curriculum.models import Course as Courses ,  Batch, CourseInstructor
from applications.examination.models import(hidden_grades , ResultAnnouncement)
from applications.academic_information.models import(Student)
from applications.online_cms.models import(Student_grades)
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view,permission_classes
from rest_framework.response import Response
import csv
from io import StringIO, BytesIO
from django.contrib.auth import get_user_model
from rest_framework.views import APIView
from django.db.models import IntegerField
from django.db.models.functions import Cast
from rest_framework.parsers import MultiPartParser, FormParser
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import traceback
from applications.academic_information.models import Course
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.colors import HexColor
from reportlab.lib.units import inch
from django.core.exceptions import ObjectDoesNotExist
from collections import defaultdict
from django.db.models import Case, When, IntegerField

grade_conversion = {
    "O": 1.0, "A+": 1.0, "A": 0.9, "B+": 0.8, "B": 0.7,
    "C+": 0.6, "C": 0.5, "D+": 0.4, "D": 0.3, "F": 0.2, "S": 0.0,
    **{f"A{i}": Decimal(str(0.9 + i * 0.01)) for i in range(1, 11)},
    **{f"B{i}": Decimal(str(0.8 + i * 0.01)) for i in range(1, 11)},
    **{
        f"{x/10:.1f}": Decimal(f"{x/100:.2f}")
        for x in range(20, 101)
    }
}

ALLOWED_GRADES = {
    "O", "A+", "A",
    "B+", "B",
    "C+", "C",
    "D+", "D", "F",
    "CD", "S", "X"
}

PBI_AND_BTP_ALLOWED_GRADES = {
    f"{x:.1f}" for x in [i / 10 for i in range(20, 101)]
}

# Helper function to format semester display for PDFs
def format_semester_display(semester_no, semester_type=None, semester_label=None):
    if semester_label and 'summer' in semester_label.lower():
        return semester_label
    if semester_type and 'summer' in semester_type.lower():
        if semester_no == 2:
            return "Summer 1"
        elif semester_no == 4:
            return "Summer 2" 
        elif semester_no == 6:
            return "Summer 3"
        elif semester_no == 8:
            return "Summer 4"
        else:
            return f"Summer {semester_no // 2}"
    else:
        return str(semester_no)

def round_from_last_decimal(number, decimal_places=1):
    d = Decimal(str(number))
    return Decimal(d).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)
    # d = Decimal(str(number))
    # current_places = abs(d.as_tuple().exponent)

    # # Keep rounding from the last decimal place until we reach the desired one
    # while current_places > decimal_places:
    #     quantize_str = '0.' + '0' * (current_places - 1) + '1'
    #     d = d.quantize(Decimal(quantize_str), rounding=ROUND_HALF_UP)
    #     current_places -= 1

    # # Final rounding to target place
    # final_quantize = '0.' + '0' * (decimal_places - 1) + '1'
    # return float(d.quantize(Decimal(final_quantize), rounding=ROUND_HALF_UP))
    

def calculate_spi_for_student(student, selected_semester, semester_type):
    semester_unit = Decimal('0')
    grades = (
        Student_grades.objects
            .filter(
                roll_no=student.id_id,
                semester=selected_semester,
                semester_type=semester_type
            )
            .annotate(
                semester_type_order=Case(
                    When(semester_type="Odd Semester",    then=0),
                    When(semester_type="Even Semester",   then=1),
                    When(semester_type="Summer Semester", then=2),
                    default=3,
                    output_field=IntegerField(),
                )
            )
            .order_by('semester', 'semester_type_order')
    )
    total_points = Decimal('0')
    total_credits = Decimal('0')
    for g in grades:
        credit = Decimal(str(g.course_id.credit))
        factor = grade_conversion.get(g.grade.strip(), -1)
        if factor >= 0:
            if factor != 0:
                factor = Decimal(str(factor))
                total_points += factor * credit
                total_credits += credit
            semester_unit += credit
    return round_from_last_decimal(Decimal('10') * (total_points / total_credits)) if total_credits else 0, semester_unit, (total_points*10)

def trace_registration(reg_id, mapping):
    seen = set()
    while reg_id in mapping and reg_id not in seen:
        seen.add(reg_id)
        reg_id = mapping[reg_id]
    return reg_id

def calculate_cpi_for_student(student, selected_semester, semester_type):
    total_unit = Decimal('0')
    if selected_semester % 2 == 0 and semester_type == 'Summer Semester':
        grades = (
            Student_grades.objects
                .filter(roll_no=student.id_id, semester__lte=selected_semester)
                .annotate(
                    semester_type_order=Case(
                        When(semester_type="Odd Semester",  then=0),
                        When(semester_type="Even Semester", then=1),
                        When(semester_type="Summer Semester", then=2),
                        default=3,
                        output_field=IntegerField(),
                    )
                )
                .order_by('semester', 'semester_type_order')
        )
        registrations = (
            course_registration.objects
                .select_related('course_id', 'semester_id')
                .filter(
                    student_id=student,
                    semester_id__semester_no__lte=selected_semester,
                )
                .annotate(
                    semester_type_order=Case(
                        When(semester_type="Odd Semester",    then=0),
                        When(semester_type="Even Semester",   then=1),
                        When(semester_type="Summer Semester", then=2),
                        default=3,
                        output_field=IntegerField(),
                    )
                )
                .order_by('semester_id__semester_no', 'semester_type_order')
        )
    else :
        grades = Student_grades.objects.filter(
            roll_no=student.id_id, semester__lte=selected_semester,
        ).exclude(semester_type = 'Summer Semester', semester = selected_semester)

        registrations = course_registration.objects.select_related('course_id', 'semester_id').filter(
            student_id=student,
            semester_id__semester_no__lte=selected_semester
        ).exclude(semester_type = 'Summer Semester', semester_id__semester_no = selected_semester)
    reg_mapping = {}
    for reg in registrations:
        key = (reg.course_id.code.strip(), reg.semester_id.semester_no, reg.semester_type)
        reg_mapping[key] = reg.id
    replacements = course_replacement.objects.filter(
        Q(old_course_registration__student_id=student) |
        Q(new_course_registration__student_id=student)
    ).select_related('old_course_registration', 'new_course_registration')
    reg_replacement_map = {}
    for rep in replacements:
        old_reg_id = rep.old_course_registration.id
        new_reg_id = rep.new_course_registration.id
        if new_reg_id != old_reg_id:
            reg_replacement_map[new_reg_id] = old_reg_id
    grade_groups = defaultdict(list)
    for g in grades:
        key = (g.course_id.code.strip(), g.semester, g.semester_type)
        reg_id = reg_mapping.get(key)
        if reg_id is None:
            continue
        original_reg_id = trace_registration(reg_id, reg_replacement_map)
        grade_groups[original_reg_id].append(g)
    total_points = Decimal('0')
    total_credits = Decimal('0')
    for orig_reg, g_list in grade_groups.items():
        best_record = max(g_list, key=lambda r: grade_conversion.get(r.grade.strip(), -1))
        grade_factor = grade_conversion.get(best_record.grade.strip(), -1)
        credit = Decimal(str(getattr(best_record.course_id, 'credit', 3)))
        if grade_factor >=  0:
            if grade_factor != 0:
                grade_factor =  Decimal(str(grade_factor))
                total_points += grade_factor * credit
                total_credits += credit
            total_unit += credit
    return round_from_last_decimal(Decimal('10') * (total_points / total_credits)) if total_credits else 0, total_unit, (total_points*10)

def parse_academic_year(academic_year, semester_type):
    """
    Parse academic_year string (e.g., "2024-25") and determine the working_year based on semester type.
    For Odd Semester, working_year = first part (e.g., 2024).
    For Even Semester, working_year = second part prefixed by '20' (e.g., 2025 if academic_year is "2024-25").
    The session is set to the academic_year string.
    """
    parts = academic_year.split("-")
    if len(parts) != 2:
        raise ValueError("Invalid academic year format. Expected format like '2024-25'.")
    first_year = parts[0].strip()
    second_year = parts[1].strip()
    if semester_type == "Odd Semester":
        working_year = int(first_year)
    elif semester_type == "Even Semester":
        working_year = int("20" + second_year)
    else:
        # For any other semester type (e.g., Summer Semester) use the first year by default.
        working_year = int("20"+second_year)
    session = academic_year  # Use the complete academic year string as session.
    return working_year, session

def is_valid_grade(grade: str, course_code: str) -> bool:
    """
    Returns True if the grade is valid for the given course code.
    Special grades apply to PR4001, PR4002 and BTP4001.
    """
    if not grade or not course_code:
        return False

    code = course_code.strip().upper()
    grade = grade.strip().upper()

    if code in {"PR4001","PR4002", "BTP4001"}:
        return grade in PBI_AND_BTP_ALLOWED_GRADES
    return grade in ALLOWED_GRADES


def gather_related_registrations(initial_reg, max_semester):
    """
    Using BFS, collect all course_registration objects related by replacements
    up to the given semester, ignoring semester_type.
    """
    related = set()
    queue = [initial_reg]
    while queue:
        reg = queue.pop(0)
        if reg.id in related:
            continue
        related.add(reg.id)
        olds = course_replacement.objects.filter(old_course_registration=reg)
        news = course_replacement.objects.filter(new_course_registration=reg)
        for rep in list(olds) + list(news):
            for neighbor in (rep.old_course_registration, rep.new_course_registration):
                if (neighbor.student_id == initial_reg.student_id and
                    neighbor.semester_id.semester_no <= max_semester):
                    queue.append(neighbor)
    return course_registration.objects.filter(id__in=related).exclude(id=initial_reg.id)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def exam_view(request):
    """
    API to differentiate roles and provide appropriate redirection links.
    """
    role = request.data.get('Role')

    if not role:
        return Response({"error": "Role parameter is required."}, status=status.HTTP_400_BAD_REQUEST)

    if role in ["Associate Professor", "Professor", "Assistant Professor"]:
        return Response({"redirect_url": "/examination/submitGradesProf/"})
    elif role == "acadadmin":
        return Response({"redirect_url": "/examination/updateGrades/"})
    elif role == "Dean Academic":
        return Response({"redirect_url": "/examination/verifyGradesDean/"})
    else:
        return Response({"redirect_url": "/dashboard/"})


class UniqueStudentGradeYearsView(APIView):
    """
    GET: Return all distinct academic_year values from Student_grades.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        years = (
            Student_grades.objects
            .values_list('academic_year', flat=True)
            .distinct()
            .order_by('academic_year')
        )
        return Response({'academic_years': list(years)}, status=200)


class UniqueRegistrationYearsView(APIView):
    """
    GET: Return all distinct working_year values from course_registration.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        years = (
            course_registration.objects
            .values_list('session', flat=True)
            .distinct()
            .order_by('session').exclude(session__isnull = True)
        )
        return Response({'academic_years': list(years)}, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def download_template(request):
    """
    API to download a CSV template for a course based on the provided role, course, academic year, and semester type.
    
    Expected request data:
      - Role: User role (only allowed roles can access, e.g., "acadadmin")
      - course: Course ID
      - year: Academic year (session) in the format "YYYY-YY" (e.g., "2023-24")
      - semester_type: Semester type (e.g., "Odd Semester", "Even Semester", "Summer Semester")
    """
    role = request.data.get('Role')
    course = request.data.get('course')
    session_year = request.data.get('year')
    semester_type = request.data.get('semester_type')

    if not role:
        return Response({"error": "Role parameter is required."}, status=status.HTTP_400_BAD_REQUEST)
    if not course or not session_year or not semester_type:
        return Response(
            {"error": "Course, academic year, and semester type are required."},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Check access for allowed roles.
    allowed_roles = [
        "acadadmin", "Associate Professor", "Professor",
        "Assistant Professor", "Dean Academic"
    ]
    if role not in allowed_roles:
        return Response({"error": "Access denied."}, status=status.HTTP_403_FORBIDDEN)

    try:
        User = get_user_model()

        # Filter course_registration records using course, session (academic year), and semester_type.
        course_info = course_registration.objects.filter(
            course_id_id=course,
            session=session_year,
            semester_type=semester_type
        ).order_by("student_id_id")

        if not course_info.exists():
            return Response(
                {"error": "No registration data found for the provided course, academic year, and semester type."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Get course information from the first matched registration.
        course_obj = course_info.first().course_id
        response = HttpResponse(content_type="text/csv")
        filename = f"{course_obj.code}_template_{session_year}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        writer.writerow(["roll_no", "name", "grade", "remarks", "semester"])

        # Write a CSV row for each student registration.
        for entry in course_info:
            student_entry = entry.student_id
            # Assuming student_entry.id_id is the student's roll number.
            student_user = User.objects.get(username=student_entry.id_id)
            writer.writerow([
                student_entry.id_id,
                f"{student_user.first_name} {student_user.last_name}",
                "",
                "",
                ""
            ])

        return response

    except Exception as e:
        return Response({'error': 'An unexpected error occurred'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

class SubmitGradesView(APIView):
    """
    API to retrieve course information for a given academic year session and semester type.

    If both academic_year (formatted as "YYYY-YY") and semester_type are provided in the request,
    the API filters courses from course_registration records based on:
      - session: must equal the provided academic_year, and
      - semester_type: must match the provided semester type.

    Otherwise, if academic_year is not provided, it returns available sessions.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        designation = request.data.get("Role")
        academic_year = request.data.get("academic_year")
        semester_type = request.data.get("semester_type")

        # Only allow access to 'acadadmin'
        if designation != "acadadmin":
            return Response(
                {"success": False, "error": "Access denied."},
                status=status.HTTP_403_FORBIDDEN
            )

        # If both academic_year and semester_type are provided, filter courses.
        if academic_year and semester_type:
            # Use academic_year to match the session field.
            unique_course_ids = course_registration.objects.filter(
                session=academic_year,
                semester_type=semester_type
            ).values("course_id").distinct()

            courses_info = Courses.objects.filter(
                id__in=unique_course_ids.values_list("course_id", flat=True)
            ).order_by("code")

            return Response(
                {"courses": list(courses_info.values())},
                status=status.HTTP_200_OK
            )

        # If academic_year is not provided, return available sessions.
        sessions = course_registration.objects.values("session").distinct()
        return Response(
            {"sessions": list(sessions)},
            status=status.HTTP_200_OK
        )


from django.db import transaction
class UploadGradesAPI(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        # Validate the role (only allow "acadadmin" in this example).
        des = request.data.get("Role")
        if des != "acadadmin":
            return Response(
                {"success": False, "error": "Access denied."},
                status=status.HTTP_403_FORBIDDEN,
            )

        csv_file = request.FILES.get("csv_file")
        if not csv_file:
            return Response(
                {"error": "No file provided. Please upload a CSV file."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not csv_file.name.endswith(".csv"):
            return Response(
                {"error": "Invalid file format. Please upload a CSV file."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Extract course_id, academic_year, and semester_type from the request.
        course_id = request.data.get("course_id")
        academic_year = request.data.get("academic_year")
        semester_type = request.data.get("semester_type")
        if not course_id or not academic_year or not semester_type:
            return Response(
                {"error": "Course ID, Academic Year, and Semester Type are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # Parse academic_year to determine working_year and session.
            working_year, session = parse_academic_year(academic_year, semester_type)

            # Fetch the course.
            courses_info = Courses.objects.get(id=course_id)

            # Check if any student is registered for this course, working_year, and semester_type.
            registrations = course_registration.objects.filter(
                course_id=courses_info,
                session = academic_year,
                semester_type=semester_type
            )
            if not registrations.exists():
                message = "NO STUDENTS REGISTERED IN THIS COURSE FOR THE SELECTED SEMESTER."
                return Response(
                    {"error": message,},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Check if grades already exist and cannot be resubmitted.
            existing_grades = Student_grades.objects.filter(
                course_id=courses_info.id, academic_year = academic_year, semester_type = semester_type
            )
            if existing_grades.exists() and not existing_grades.first().reSubmit:
                message = "THIS COURSE HAS ALREADY BEEN SUBMITTED."
                return Response(
                    {"error": message},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Parse the CSV file."20"
            decoded_file = csv_file.read().decode("utf-8").splitlines()
            reader = csv.DictReader(decoded_file)
            required_columns = ["roll_no", "grade", "remarks"]
            if not all(column in reader.fieldnames for column in required_columns):
                return Response(
                    {
                        "error": "CSV file must contain the following columns: roll_no, grade, remarks."
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            errors = []  # To track errors for each CSV row.
            allowed_list = ", ".join(sorted(ALLOWED_GRADES))
            # Wrap the upload process in an atomic transaction.
            with transaction.atomic():
                for index, row in enumerate(reader, start=1):
                    roll_no = row.get("roll_no")
                    grade = row.get("grade")
                    remarks = row.get("remarks", "")
                    semester = row.get("semester", None)

                    # Validate student existence.
                    try:
                        stud = Student.objects.get(id_id=roll_no)
                    except Student.DoesNotExist:
                        errors.append(f"Row {index}: Student with roll_no {roll_no} does not exist.")
                        continue

                    # Check that the student is registered for the course.
                    registration_exists = course_registration.objects.filter(
                        student_id=stud,
                        course_id=courses_info,
                        semester_type=semester_type,
                        session = academic_year
                    ).exists()
                    if not registration_exists:
                        errors.append(
                            f"Row {index}: Student with roll_no {roll_no} is not registered for this course in the selected semester."
                        )
                        continue


                    if not is_valid_grade(grade, courses_info.code):
                        errors.append(
                            f"Row {index}: Invalid grade '{grade}' for roll_no {roll_no}. "
                            f"Allowed grades are: {allowed_list}."
                        )
                        continue

                    # Determine the semester for the grade (use the provided value or fall back to student's current semester).
                    semester = semester or stud.curr_semester_no
                    batch = stud.batch
                    reSubmit = False

                    # Create or update the grade record.
                    try:
                        Student_grades.objects.update_or_create(
                            roll_no=roll_no,
                            course_id_id=course_id,
                            year=working_year,  # stored as academic year string
                            semester=semester,
                            batch=batch,
                            academic_year = academic_year,
                            semester_type = semester_type,
                            defaults={
                                'grade': grade,
                                'remarks': remarks,
                                'reSubmit': reSubmit,
                                'academic_year': session,        
                                'semester_type': semester_type,
                            }
                        )
                    except Exception as create_err:
                        errors.append(
                            f"Row {index}: Error creating/updating grade for student with roll_no {roll_no} - {str(create_err)}"
                        )
                        continue

                # If errors were encountered in any row, rollback and return error summary.
                if errors:
                    error_summary = "\n".join(f"- {msg}" for msg in errors)
                    raise Exception(error_summary)

            return Response(
                {"message": "Grades uploaded successfully."},
                status=status.HTTP_200_OK,
            )

        except Courses.DoesNotExist:
            return Response(
                {"error": "Invalid course ID."}, status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {"error": f"An error occurred: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
"""
API to fetch courses with unverified grades along with unique academic years.

- Only users with the role of 'acadadmin' can access this endpoint.
- Retrieves courses where at least one student's grades are unverified.
- Fetches the academic years associated with unverified grades.

Expected Request:
Headers:
    Authorization: Token <your_auth_token>

Body (JSON):
    {
        "Role": "acadadmin"
    }

Response:
    200 OK - {
        "courses_info": [{"id": 1, "course_name": "Data Structures", ...}],
        "unique_year_ids": [{"year": "2024"}, {"year": "2025"}]
    }
    403 Forbidden - {"success": false, "error": "Access denied."}
"""

class UpdateGradesAPI(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        role = request.data.get("Role")
        academic_year = request.data.get("academic_year")
        semester_type = request.data.get("semester_type")

        if role != "acadadmin":
            return Response(
                {"success": False, "error": "Access denied."},
                status=403,
            )

        if not academic_year or not semester_type:
            return Response(
                {"success": False, "error": "Academic year and semester type are required."},
                status=400,
            )

        # Filter unverified grades based on academic_year and semester_type, then get unique course IDs.
        unique_course_ids = (
            Student_grades.objects.filter(
                verified=False,
                academic_year=academic_year,
                semester_type=semester_type,
            )
            .values("course_id")
            .distinct()
            .annotate(course_id_int=Cast("course_id", IntegerField()))
        )

        # Retrieve courses that are linked to these student grades.
        courses_info = Courses.objects.filter(
            id__in=unique_course_ids.values_list("course_id_int", flat=True)
        )

        # Optionally, get unique year values for the selected filters.
        unique_year_ids = Student_grades.objects.filter(
            academic_year=academic_year,
            semester_type=semester_type
        ).values("year").distinct()

        return Response(
            {
                "courses_info": list(courses_info.values()),
                "unique_year_ids": list(unique_year_ids),
            },
            status=200,
        )


"""
API to check if grades have been submitted and are verified for a given course and academic year.

- Only users with the role of 'acadadmin' can access this endpoint.
- Verifies if grades exist for the requested course and year.
- Returns student grades if unverified, else indicates if already verified.

Expected Request:
Headers:
    Authorization: Token <your_auth_token>

Body (JSON):
    {
        "Role": "acadadmin",
        "course": <course_id>,
        "year": <academic_year>
    }

Response:
    200 OK - If grades exist but are unverified:
        {
            "registrations": [
                {"id": 1, "roll_no": "CS101001", "grade": "A", ...}
            ]
        }
    200 OK - If already verified:
        {"message": "This course is already verified."}
    400 Bad Request - If required fields are missing:
        {"error": "Both 'course' and 'year' are required."}
    404 Not Found - If no grades exist:
        {"message": "This course is not submitted by the instructor."}
    403 Forbidden - If access is denied:
        {"success": false, "error": "Access denied."}
"""

class UpdateEnterGradesAPI(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        des = request.data.get("Role")

        
        if des != "acadadmin":
            return Response(
                {"success": False, "error": "Access denied."},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Get course_id and year from request body
        course_id = request.data.get("course")
        year = request.data.get("year")
        semester_type = request.data.get("semester_type")

        # Validate course_id and year
        if not course_id or not year:
            return Response(
                {"error": "Both 'course' and 'year' are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check if the course exists with grades for the given year
        course_present = Student_grades.objects.filter(course_id=course_id, academic_year=year, semester_type=semester_type)

        if not course_present.exists():
            return Response(
                {"message": "This course is not submitted by the instructor."},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Check if the course is already verified
        verification = course_present.first().verified
        if verification:
            return Response(
                {"message": "This course is already verified."},
                status=status.HTTP_200_OK,
            )

        
        registrations = course_present.values()

        return Response(
            {"registrations": list(registrations)},
            status=status.HTTP_200_OK,
        )
    


"""
API for moderating student grades (updating, verifying, or creating hidden grades).

- Only users with the roles 'acadadmin' or 'Dean Academic' can access this endpoint.
- Updates grades for students based on course and semester.
- Marks grades as verified and supports resubmission if enabled.
- If a student grade doesn't exist, it creates a hidden grade record.
- Returns the updated grades in a downloadable CSV file.

Expected Request:
Headers:
    Authorization: Token <your_auth_token>

Body (JSON):
    {
        "Role": "acadadmin",
        "student_ids": ["20231001", "20231002"],
        "semester_ids": [5, 5],
        "course_ids": [101, 101],
        "grades": ["A", "B"],
        "allow_resubmission": "YES"
    }

Response:
    200 OK - CSV file with updated grades
    403 Forbidden - If access is denied:
        {"success": false, "error": "Access denied."}
    400 Bad Request - If required fields are missing or mismatched:
        {"error": "Invalid or incomplete grade data provided."}
    500 Internal Server Error - If any unexpected error occurs:
        {"error": "An error occurred: <error_message>"}
"""

class ModerateStudentGradesAPI(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        
        des = request.data.get("Role")
        if des not in ["acadadmin", "Dean Academic"]:
            return Response(
                {"success": False, "error": "Access denied."},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Extract data from the request
        student_ids = request.data.get("student_ids", [])
        semester_ids = request.data.get("semester_ids", [])
        course_ids = request.data.get("course_ids", [])
        grades = request.data.get("grades", [])
        remarks=request.data.get("remarks",[])
        allow_resubmission = request.data.get("allow_resubmission", "NO")

       
        if (
            not student_ids
            or not semester_ids
            or not course_ids
            or not grades
            or len(student_ids) != len(semester_ids)
            or len(semester_ids) != len(course_ids)
            or len(course_ids) != len(grades)
        ):
            return Response(
                {"error": "Invalid or incomplete grade data provided."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Update or create grades
        for student_id, semester_id, course_id, grade,remark in zip(
            student_ids, semester_ids, course_ids, grades,remarks
        ):
            try:
                grade_of_student = Student_grades.objects.get(
                    course_id=course_id, roll_no=student_id, semester=semester_id
                )
                grade_of_student.remarks=remark
                grade_of_student.grade = grade
                grade_of_student.verified = True
                if allow_resubmission.upper() == "YES":
                    grade_of_student.reSubmit = True
                grade_of_student.save()
            except Student_grades.DoesNotExist:
                # Create a new hidden grade if the student grade doesn't exist
                hidden_grades.objects.create(
                    course_id=course_id,
                    student_id=student_id,
                    semester_id=semester_id,
                    grade=grade,
                )

        # Generate CSV file as the response
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="grades.csv"'

        writer = csv.writer(response)
        writer.writerow(["Student ID", "Semester ID", "Course ID", "Grade"])
        for student_id, semester_id, course_id, grade in zip(
            student_ids, semester_ids, course_ids, grades
        ):
            writer.writerow([student_id, semester_id, course_id, grade])

        # Return the CSV response
        return response
    


"""
API to generate a student's academic transcript for a specific semester.

- Only users with the role 'acadadmin' can access this endpoint.
- Fetches the courses and grades of the student for the requested semester.
- Also retrieves all courses registered by the student up to that semester.
- If a grade is unavailable for a course, it returns "Grading not done yet".

Expected Request:
Headers:
    Authorization: Token <your_auth_token>

Body (JSON):
    {
        "Role": "acadadmin",
        "student": "20231001",
        "semester": 3
    }

Response:
    200 OK - Transcript data
    403 Forbidden - If access is denied:
        {"error": "Access denied."}
    400 Bad Request - If required fields are missing:
        {"error": "Student ID and Semester are required."}
    500 Internal Server Error - If any unexpected error occurs:
        {"error": "An error occurred: <error_message>"}
"""

import json
class GenerateTranscript(APIView):
    permission_classes = [IsAuthenticated] 

    def post(self, request):
        des = request.data.get("Role")
        student_id = request.data.get("student")
        semester = request.data.get("semester")
        semester = json.loads(semester)
        semester_number = semester.get('no')
        semester_type = semester.get('type')

        if des != "acadadmin":
            return Response({"error": "Access denied."}, status=status.HTTP_403_FORBIDDEN)

        if not student_id or not semester:
            return Response({"error": "Student ID and Semester are required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            student = Student.objects.get(id_id=student_id)
            cpi, tu, _ = calculate_cpi_for_student(student, semester_number, semester_type)
            spi, su, _ = calculate_spi_for_student(student, semester_number, semester_type)
        except:
            return Response({"error": "Student ID does not exist."}, status=status.HTTP_400_BAD_REQUEST)

        course_grades = {}
        courses_registered = Student_grades.objects.filter(roll_no=student_id, semester=semester_number, semester_type = semester_type)

        # Get academic year from the first grade record
        academic_year = None
        if courses_registered.exists():
            academic_year = courses_registered.first().academic_year

        for reg in courses_registered:
            course = reg.course_id
            course_grades[course.id] = {
                "course_name": course.name,
                "course_code": course.code,
                "credit": course.credit,
                "grade": reg.grade,
                "points": Decimal(str(grade_conversion.get(reg.grade, 0) * 10)).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP),
            }

        # Add complete student information like CheckResultView
        student_info = {
            "name": f"{student.id.user.first_name} {student.id.user.last_name}".strip(),
            "student_name": f"{student.id.user.first_name} {student.id.user.last_name}".strip(),
            "rollNumber": student.id.user.username,
            "roll_number": student.id.user.username,
            "programme": student.programme,
            "batch": str(student.batch_id) if student.batch_id else str(student.batch),
            "branch": student.id.department.name if student.id.department else "",
            "department": student.id.department.name if student.id.department else "",
            "semester": student.curr_semester_no,
            "academicYear": academic_year or "",
            "academic_year": academic_year or ""
        }

        response_data = {
            "name": f"{student.id.user.first_name} {student.id.user.last_name}".strip(),
            "student_name": f"{student.id.user.first_name} {student.id.user.last_name}".strip(),
            "roll_number": student.id.user.username,
            "programme": student.programme,
            "department": student.id.department.name if student.id.department else "",
            "branch": student.id.department.name if student.id.department else "",
            "academic_year": academic_year or "",
            "student_info": student_info,
            "courses_grades": course_grades,
            "spi": spi,
            "cpi": cpi,
            "tu": tu,
            "su": su,
        }

        return Response(response_data, status=status.HTTP_200_OK)
    

"""
API to fetch available academic details and retrieve students for generating transcripts.

- Only users with the role 'acadadmin' can access this endpoint.
- GET: Retrieves the list of available programmes, batches, and specializations.
- POST: Fetches students based on programme, batch, specialization (optional), and semester.

Expected Requests:
1. GET /api/generate-transcript-form/
   Headers:
       Authorization: Token <your_auth_token>
       X-User-Role: acadadmin
   Response:
       200 OK - List of programmes, batches, and specializations
       403 Forbidden - If access is denied:
           {"error": "Access denied. Invalid or missing role."}

2. POST /api/generate-transcript-form/
   Headers:
       Authorization: Token <your_auth_token>
       X-User-Role: acadadmin
   Body (JSON):
       {
           "programme": "B.Tech",
           "batch": 2021,
           "specialization": "AI & ML",
           "semester": 5
       }
   Response:
       200 OK - List of students in the given filters
       403 Forbidden - If access is denied:
           {"error": "Access denied. Invalid or missing role."}
       400 Bad Request - If required fields are missing:
           {"error": "Programme, batch, and semester are required fields."}
"""

class GenerateTranscriptForm(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        role = request.GET.get("role")
        if not role or role != "acadadmin":
            return Response(
                {"error": "Access denied. Invalid or missing role."},
                status=status.HTTP_403_FORBIDDEN
            )

        # Query only running batches from the Batch table.
        batches_queryset = Batch.objects.filter(running_batch=True)
        # Create a display label that combines batch name, discipline, and year.
        batch_list = [
            {
                "id": batch.id,
                "label": f"{batch.name} - {batch.discipline} {batch.year}"
            }
            for batch in batches_queryset
        ]

        # Get programmes from Student table.
        programmes = Student.objects.values_list('programme', flat=True).distinct()
        # Get unique, non-null (and non-empty) specializations.
        specializations = Student.objects.exclude(specialization__isnull=True)\
                                         .exclude(specialization__exact="")\
                                         .values_list('specialization', flat=True)\
                                         .distinct()
        return Response({
            "programmes": list(programmes),
            "batches": batch_list,
            "specializations": list(specializations),
        }, status=status.HTTP_200_OK)

    def post(self, request):
        role = request.data.get("Role")
        if not role or role != "acadadmin":
            return Response(
                {"error": "Access denied. Invalid or missing role."},
                status=status.HTTP_403_FORBIDDEN
            )

        batch = request.data.get('batch')
        specialization = request.data.get('specialization')
        semester = request.data.get('semester')

        if not batch or not semester:
            return Response(
                {"error": "batch, and semester are required fields."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Use the batch from the Batch table (passed as ID).
        if specialization:
            students = Student.objects.filter(
                batch_id=batch, specialization=specialization
            ).order_by('id')
        else:
            students = Student.objects.filter(
                batch_id=batch
            ).order_by('id')

        return Response({
            "students": list(students.values()),
            "semester": semester
        }, status=status.HTTP_200_OK)


class GenerateResultAPI(APIView):
    """
    API endpoint to generate an Excel file containing student grades with SPI and CPI.
    
    Request Body (JSON):
      - Role: must be "acadadmin"
      - semester: integer (selected semester)
      - specialization: string (branch acronym, e.g., "CSE") [optional]
      - batch: integer (Batch primary key)
    
    Response:
      - An Excel file (student_grades.xlsx) in the required format.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            role = request.data.get("Role")
            if role != "acadadmin":
                return Response({"error": "Access denied."}, status=403)

            semester = request.data.get("semester")
            branch = request.data.get("specialization")  # optional now
            batch_id = request.data.get("batch")
            semester_type = request.data.get("semester_type")

            if not semester or not batch_id:
                return Response({"error": "Semester and Batch are required."}, status=400)

            # Get the Batch record using its primary key.
            batch_obj = Batch.objects.filter(id=batch_id).first()
            if not batch_obj:
                return Response({"error": "Batch not found."}, status=404)

            # Fetch all students for this Batch.
            if branch:
                students = Student.objects.filter(batch_id=batch_id, specialization=branch).order_by('id')
            else:
                students = Student.objects.filter(batch_id=batch_id).order_by('id')

            # Fetch course_ids for which the grade is not empty.
            course_ids = Student_grades.objects.filter(
                batch=batch_obj.year, 
                semester=semester,
                roll_no__in = students,
                semester_type = semester_type
            ).exclude(grade__isnull=True).exclude(grade="").values_list('course_id_id', flat=True).distinct()
            
            courses = Courses.objects.filter(id__in=course_ids)
            courses_map = {course.id: course.credit for course in courses}

            wb = Workbook()
            ws = wb.active
            ws.title = "Student Grades"

            # Define a fill style for header cells: light grey background.
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )


            # Setup header rows: S. No and Roll No in columns A and B.
            ws["A1"] = "S. No"
            ws["B1"] = "Roll No"
            for col in ("A", "B"):
                cell = ws[col + "1"]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True)
                cell.fill = header_fill
            ws.column_dimensions[get_column_letter(1)].width = 12
            ws.column_dimensions[get_column_letter(2)].width = 18

            # Starting from column 3, add headers for each course (each course uses 2 columns for Grade and Remarks).
            col_idx = 3
            for course in courses:
                # Merge cells for the course code header.
                ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+1)
                cell = ws.cell(row=1, column=col_idx)
                cell.value = course.code
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True)
                cell.fill = header_fill

                ws.merge_cells(start_row=2, start_column=col_idx, end_row=2, end_column=col_idx+1)
                cell = ws.cell(row=2, column=col_idx)
                cell.value = course.name
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True)
                cell.fill = header_fill

                ws.merge_cells(start_row=3, start_column=col_idx, end_row=3, end_column=col_idx+1)
                cell = ws.cell(row=3, column=col_idx)
                cell.value = course.credit
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True)
                cell.fill = header_fill

                cell = ws.cell(row=4, column=col_idx)
                cell.value = "Grade"
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True)
                cell.fill = header_fill

                cell = ws.cell(row=4, column=col_idx+1)
                cell.value = "Remarks"
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True)
                cell.fill = header_fill

                ws.column_dimensions[get_column_letter(col_idx)].width = 25
                ws.column_dimensions[get_column_letter(col_idx+1)].width = 25
                col_idx += 2

            # Append headers for SPI and CPI.
            cell = ws.cell(row=1, column=col_idx)
            cell.value = "SPI"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            cell.fill = header_fill

            cell = ws.cell(row=1, column=col_idx+1)
            cell.value = "CPI"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            cell.fill = header_fill

            cell = ws.cell(row=1, column=col_idx+2)
            cell.value = "SU"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            cell.fill = header_fill

            cell = ws.cell(row=1, column=col_idx+3)
            cell.value = "TU"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            cell.fill = header_fill

            cell = ws.cell(row=1, column=col_idx+4)
            cell.value = "SP"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            cell.fill = header_fill

            cell = ws.cell(row=1, column=col_idx+5)
            cell.value = "TP"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            cell.fill = header_fill

            # Ensure full header rows (1 to 4) are highlighted.
            max_col = ws.max_column
            for row in range(1, 5):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = header_fill
                    cell.border = thin_border

            # Fill in student rows, starting from row 5.
            row_idx = 5
            for idx, student in enumerate(students, start=1):
                ws.cell(row=row_idx, column=1).value = idx
                ws.cell(row=row_idx, column=2).value = student.id_id
                for c in [1, 2]:
                    ws.cell(row=row_idx, column=c).alignment = Alignment(horizontal="center", vertical="center")
                
                # Get the student’s grade records for the current semester.
                student_grades = Student_grades.objects.filter(
                    roll_no=student.id_id,
                    course_id_id__in=course_ids,
                    semester_type=semester_type,
                    semester=semester
                )
                grades_map = {g.course_id_id: g for g in student_grades}
                col_ptr = 3
                for course in courses:
                    grade_entry = grades_map.get(course.id)
                    grade_val = grade_entry.grade if grade_entry else '-'

                    remark = '-'
                    if grade_entry:
                        reg = course_registration.objects.filter(
                            student_id=student,
                            course_id=course,
                            semester_id__semester_no=semester,
                            semester_type=semester_type,
                            session=grade_entry.academic_year,
                        ).first()
                        if reg:
                            related_regs = gather_related_registrations(reg, semester)
                            attempts = []
                            for r in related_regs:
                                g = Student_grades.objects.filter(
                                    roll_no=student.id_id,
                                    course_id__code=r.course_id.code,
                                    semester=r.semester_id.semester_no,
                                    semester_type = r.semester_type,
                                    academic_year = r.session
                                ).order_by('-semester').first()
                                if g:
                                    attempts.append((r.course_id.code, g.grade))

                            if len(attempts) >= 1:
                                scored = sorted(
                                    attempts,
                                    key=lambda x: grade_conversion.get(x[1], -1),
                                    reverse=True
                                )
                                first_code, first_grade = scored[0]
                                if first_grade == 'F' or first_grade == 'X':
                                    remark = 'R(BL)' if first_code == course.code else 'S(BL)'
                                else:
                                    remark = 'R(IM)' if first_code == course.code else 'S(IM)'
                    ws.cell(row=row_idx, column=col_ptr).value = grade_val
                    ws.cell(row=row_idx, column=col_ptr+1).value = remark
                    for c in [col_ptr, col_ptr+1]:
                        ws.cell(row=row_idx, column=c).alignment = Alignment(horizontal="center", vertical="center")
                    col_ptr += 2

                # Calculate SPI and CPI.
                spi_val, SU, SP = calculate_spi_for_student(student, semester, semester_type)
                cpi_val, TU, TP = calculate_cpi_for_student(student, semester, semester_type)
                ws.cell(row=row_idx, column=col_ptr).value = spi_val
                ws.cell(row=row_idx, column=col_ptr+1).value = cpi_val
                ws.cell(row=row_idx, column=col_ptr+2).value = SU
                ws.cell(row=row_idx, column=col_ptr+3).value = TU
                ws.cell(row=row_idx, column=col_ptr+4).value = SP
                ws.cell(row=row_idx, column=col_ptr+5).value = TP
                for c in [col_ptr, col_ptr+1]:
                    ws.cell(row=row_idx, column=c).alignment = Alignment(horizontal="center", vertical="center")
                row_idx += 1

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="student_grades.xlsx"'
            wb.save(response)
            return response

        except Exception as e:
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)


class SubmitAPI(APIView):

    """
    API endpoint to fetch the list of unique courses available for submission.
    Accessible only by users with the 'acadadmin' or 'Dean Academic' role.

    ### Headers:
    - **Authorization**: `Token <your_token>` (Required, for authentication)
    
    ### Request Body (JSON):
    - **Role** (string, required) → User role (must be `"acadadmin"` or `"Dean Academic"`)

    ### Response:
    - **Success (200 OK)**: Returns a JSON object with `courses_info`, which contains course details.
    - **Errors**:
      - `403 Forbidden`: If the user does not have `"acadadmin"` or `"Dean Academic"` role.
      - `500 Internal Server Error`: For any unexpected errors.
    """

    permission_classes = [IsAuthenticated]

    def post(self, request):
        role = request.data.get("Role")

        if role not in ["acadadmin", "Dean Academic"]:
            return Response(
                {"error": "Access denied."},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Get unique course IDs
        unique_course_ids = (
            course_registration.objects.values("course_id")
            .distinct()
            .annotate(course_id_int=Cast("course_id", IntegerField()))
        )

        courses_info = Course.objects.filter(
            id__in=unique_course_ids.values_list("course_id_int", flat=True)
        )

        return Response(
            {"courses_info": list(courses_info.values())},
            status=status.HTTP_200_OK,
        )


class DownloadExcelAPI(APIView):

    """
    API endpoint to generate and download a CSV file containing student grades.

    ### Headers:
    - **Authorization**: `Token <your_token>` (Required, for authentication)

    ### Request Body (JSON):
    - **student_ids** (list, required) → List of student IDs.
    - **semester_ids** (list, required) → Corresponding semester IDs.
    - **course_ids** (list, required) → Corresponding course IDs.
    - **grades** (list, required) → Corresponding grades for the students.

    ### Response:
    - **Success (200 OK)**: Returns a downloadable CSV file with student grades.
    - **Errors**:
      - `400 Bad Request`: If input data is missing or inconsistent in length.
      - `500 Internal Server Error`: For any unexpected errors.
    """

    permission_classes = [IsAuthenticated]

    def post(self, request):
        student_ids = request.data.get("student_ids", [])
        semester_ids = request.data.get("semester_ids", [])
        course_ids = request.data.get("course_ids", [])
        grades = request.data.get("grades", [])

        if (
            not student_ids
            or not semester_ids
            or not course_ids
            or not grades
            or len(student_ids) != len(semester_ids)
            or len(semester_ids) != len(course_ids)
            or len(course_ids) != len(grades)
        ):
            return Response(
                {"error": "Invalid or incomplete grade data provided."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="grades.csv"'

        writer = csv.writer(response)
        writer.writerow(["Student ID", "Semester ID", "Course ID", "Grade"])
        for student_id, semester_id, course_id, grade in zip(
            student_ids, semester_ids, course_ids, grades
        ):
            writer.writerow([student_id, semester_id, course_id, grade])

        return response


class SubmitGradesProfAPI(APIView):

    """
    API endpoint to fetch courses assigned to a professor and available academic years.

    ### Headers:
    - **Authorization**: `Token <your_token>` (Required for authentication)

    ### Request Body (JSON):
    - **Role** (string, required) → User's role (Must be one of `Associate Professor`, `Professor`, or `Assistant Professor`).

   
    """

    permission_classes = [IsAuthenticated]

    def post(self, request):
        
        role = request.data.get("Role")
        academic_year = request.data.get("academic_year")
        semester_type = request.data.get("semester_type")
        
        if role not in ["Associate Professor", "Professor", "Assistant Professor"]:
            return Response(
                {"success": False, "error": "Access denied."},
                status=status.HTTP_403_FORBIDDEN,
            )

        if not academic_year or not semester_type:
            return Response(
                {"success": False, "error": "Academic year and semester type are required."},
                status=400,
            )
        
        instructor_id = request.user.username

        working_year, _ = parse_academic_year(academic_year=academic_year, semester_type=semester_type)

        unique_course_ids = (
            CourseInstructor.objects.filter(instructor_id_id=instructor_id, year = working_year, semester_type=semester_type)
            .values("course_id_id")
            .distinct()
            .annotate(course_id_int=Cast("course_id_id", IntegerField()))
        )

        # Retrieve course details
        courses_info = Courses.objects.filter(
            id__in=unique_course_ids.values_list("course_id_int", flat=True)
        )

        return Response(
            {
                "courses_info": list(courses_info.values()),
            },
            status=status.HTTP_200_OK,
        )


class UploadGradesProfAPI(APIView):
    """
    Upload grades CSV by the assigned instructor.
    - Role check
    - File & header validation
    - Registration & duplicate submit checks
    - Instructor ownership check
    - All or nothing: reset reSubmit and update/create rows in one atomic block
    """
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        try:
            # 1) ROLE CHECK
            role = request.data.get("Role")
            if role not in ["Associate Professor", "Professor", "Assistant Professor"]:
                return Response({"error": "Access denied."},
                                status=status.HTTP_403_FORBIDDEN)

            # 2) FILE CHECK
            csv_file = request.FILES.get("csv_file")
            if not csv_file:
                return Response(
                    {"error": "No file provided. Please upload a CSV file."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            if not csv_file.name.lower().endswith(".csv"):
                return Response(
                    {"error": "Invalid file format. Please upload a CSV file."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 3) REQUIRED PARAMS
            course_id     = request.data.get("course_id")
            academic_year = request.data.get("academic_year")
            semester_type = request.data.get("semester_type")
            if not (course_id and academic_year and semester_type):
                return Response(
                    {"error": "Course ID, Academic Year, and Semester Type are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 4) PARSE academic_year → working_year & session
            try:
                working_year, session = parse_academic_year(academic_year, semester_type)
            except Exception:
                return Response(
                    {"error": "Invalid academic year or semester type."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 5) FETCH COURSE
            try:
                course = Courses.objects.get(id=course_id)
            except Courses.DoesNotExist:
                return Response({"error": "Invalid course ID."},
                                status=status.HTTP_400_BAD_REQUEST)

            # 6) CHECK STUDENT REGISTRATIONS
            regs = course_registration.objects.filter(
                course_id=course,
                working_year=working_year,
                semester_type=semester_type
            )
            if not regs.exists():
                return Response(
                    {"error": "NO STUDENTS REGISTERED IN THIS COURSE FOR THE SELECTED SEMESTER."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 7) DUPLICATE‐SUBMIT CHECK
            existing = Student_grades.objects.filter(
                course_id=course_id,
                academic_year=academic_year,
                semester_type=semester_type
            )
            if existing.exists() and not existing.first().reSubmit:
                return Response(
                    {"error": "THIS COURSE HAS ALREADY BEEN SUBMITTED."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 8) INSTRUCTOR‐OWNERSHIP CHECK
            if not CourseInstructor.objects.filter(
                course_id_id=course_id,
                instructor_id_id=request.user.username,
                year=working_year
            ).exists():
                return Response(
                    {"error": "Access denied: you are not assigned as instructor for this course."},
                    status=status.HTTP_403_FORBIDDEN
                )

            # 9) PARSE CSV HEADER
            decoded = csv_file.read().decode("utf-8").splitlines()
            reader  = csv.DictReader(decoded)
            required_cols = {"roll_no", "grade", "remarks"}
            if not required_cols.issubset(reader.fieldnames or []):
                return Response(
                    {"error": "CSV file must contain columns: roll_no, grade, remarks."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 10) ATOMIC PROCESSING
            errors = []
            with transaction.atomic():
                # ─── Reset ALL reSubmit flags for this course/year/semester ───
                Student_grades.objects.filter(
                    course_id_id=course_id,
                    academic_year=academic_year,
                    semester_type=semester_type
                ).update(reSubmit=False)

                # ─── Process each CSV row ───
                for idx, row in enumerate(reader, start=1):
                    roll_no = row.get("roll_no", "").strip()
                    grade   = row.get("grade", "").strip()
                    remarks = row.get("remarks", "").strip()
                    sem_csv = row.get("semester", "").strip() or None

                    # a) STUDENT EXISTS?
                    try:
                        stud = Student.objects.get(id_id=roll_no)
                    except Student.DoesNotExist:
                        errors.append(f"Row {idx}: Student with roll_no {roll_no} does not exist.")
                        continue

                    # b) STUDENT REGISTERED FOR COURSE?
                    if not course_registration.objects.filter(
                        student_id=stud,
                        course_id=course,
                        semester_type=semester_type,
                        session=academic_year
                    ).exists():
                        errors.append(
                            f"Row {idx}: Student {roll_no} not registered for this course/semester."
                        )
                        continue

                    # c) VALID GRADE?
                    if grade not in ALLOWED_GRADES:
                        allowed = ", ".join(sorted(ALLOWED_GRADES))
                        errors.append(
                            f"Row {idx}: Invalid grade '{grade}'. Allowed: {allowed}."
                        )
                        continue

                    # d) DETERMINE SEMESTER & BATCH
                    semester = sem_csv or stud.curr_semester_no
                    batch    = stud.batch
                    reSubmit = False  # enforce false

                    # e) UPSERT GRADE
                    try:
                        Student_grades.objects.update_or_create(
                            roll_no=roll_no,
                            course_id_id=course_id,
                            year=working_year,
                            batch = batch,
                            academic_year=academic_year,
                            semester_type=semester_type,
                            semester=semester,
                            defaults={
                                "grade": grade,
                                "remarks": remarks,
                                "reSubmit": reSubmit,
                            }
                        )
                    except Exception as exc:
                        errors.append(
                            f"Row {idx}: Error saving grade for {roll_no}: {str(exc)}"
                        )
                        continue

                # If any row errors occurred, rollback & return them
                if errors:
                    summary = "\n".join(f"- {e}" for e in errors)
                    raise Exception(f"Upload failed with the following errors:\n{summary}")

            # 11) SUCCESS RESPONSE
            return Response(
                {"message": "Grades uploaded successfully."},
                status=status.HTTP_200_OK
            )

        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )


class DownloadGradesAPI(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            role = request.data.get("Role")
            academic_year = request.data.get("academic_year")
            semester_type = request.data.get("semester_type")

            if role not in ["Associate Professor", "Professor", "Assistant Professor"]:
                return Response({"error": "Access denied."}, status=status.HTTP_403_FORBIDDEN)

            if not academic_year or not semester_type:
                return Response(
                    {"success": False, "error": "Academic year and semester type are required."},
                    status=400,
                )

            working_year, _ = parse_academic_year(academic_year=academic_year, semester_type=semester_type)
            instructor_id = request.user.username

            unique_course_ids = (
                CourseInstructor.objects
                    .filter(instructor_id_id=instructor_id, year=working_year, semester_type=semester_type)
                    .values("course_id_id")
                    .distinct()
                    .annotate(course_id_int=Cast("course_id_id", IntegerField()))
            )

            grades_qs = Student_grades.objects.filter(
                academic_year = academic_year,
                semester_type = semester_type,
                course_id_id__in=unique_course_ids.values_list("course_id_int", flat=True)
            )

            course_ids = grades_qs.values_list("course_id_id", flat=True).distinct()
            courses_details = Courses.objects.filter(id__in=course_ids)

            return Response({"courses": list(courses_details.values())}, status=status.HTTP_200_OK)

        except Exception as e:
            # Optionally log the exception here
            return Response({"error": "Internal server error."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class GeneratePDFAPI(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            # Check if this is a student result request (no Role field or Role is student)
            role = request.data.get("Role")
            
            # If no role specified or role is student, handle as student result PDF
            if not role or role.lower() == "student" or "student_info" in request.data:
                # Delegate to student result PDF generation
                return self.generate_student_result_pdf(request)
            
            # Faculty role check for course grade sheets
            if role not in ["Associate Professor", "Professor", "Assistant Professor"]:
                return Response({"error": "Access denied."}, status=status.HTTP_403_FORBIDDEN)

            # Existing faculty course grade sheet logic
            course_id     = request.data.get("course_id")
            academic_year = request.data.get("academic_year")
            semester_type = request.data.get("semester_type")

            course_info     = get_object_or_404(Courses, id=course_id)
            working_year, _ = parse_academic_year(academic_year, semester_type)

            grades = Student_grades.objects.filter(
                course_id_id=course_id,
                academic_year=academic_year,
                semester_type=semester_type
            ).order_by("roll_no")

            ci = CourseInstructor.objects.filter(
                course_id_id=course_id,
                year=working_year,
                semester_type=semester_type,
                instructor_id_id=request.user.username
            )
            if not ci.exists():
                return Response({"success": False, "error": "Course not found."}, status=404)

            # semester   = ci.first().semester_no
            instructor = f"{request.user.first_name} {request.user.last_name}"

            # count grades
            all_grades   = ["O","A+","A","B+","B","C+","C","D+","D","F","S","X","CD"]
            grade_counts = {g: grades.filter(grade=g).count() for g in all_grades}

            # prepare PDF response
            response = HttpResponse(content_type="application/pdf")
            response["Content-Disposition"] = f'attachment; filename="{course_info.code}_grades.pdf"'

            # Create PDF metadata to fix "(anonymous)" title issue
            pdf_title = f"Course Grades - {course_info.code} - {course_info.name}"
            
            # ↑ Increase topMargin to 2" for header + spaceBetween
            doc = SimpleDocTemplate(
                response,
                pagesize=letter,
                leftMargin=inch,
                rightMargin=inch,
                topMargin=2 * inch,
                bottomMargin=inch,
                title=pdf_title,
                author="PDPM IIITDM Jabalpur",
                subject=f"Course Grade Report - {course_info.code}",
                creator="Fusion Academic System"
            )

            elements = []
            styles   = getSampleStyleSheet()

            # keep your original field_label_style
            field_label_style = ParagraphStyle(
                "FieldLabelStyle",
                parent=styles["Normal"],
                fontSize=12,
                textColor=colors.black,
                spaceAfter=5,
            )

            # Title style for body (you can leave this or remove, it's not used in header now)
            header_style = ParagraphStyle(
                "HeaderStyle",
                parent=styles["Heading1"],
                fontName="Helvetica-Bold",
                fontSize=16,
                textColor=HexColor("#333333"),
                spaceAfter=20,
                alignment=1,
            )
            # You can optionally remove the next line if you don't want "Grade Sheet"
            # in the body—it's now drawn in the page header.

            # Main grades table (unchanged)
            data = [["S.No.","Roll Number","Grade"]]
            for i, g in enumerate(grades, 1):
                data.append([i, g.roll_no, g.grade])
            tbl = Table(data, colWidths=[80, 300, 100])
            tbl.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), HexColor("#E0E0E0")),
                ("TEXTCOLOR", (0,0), (-1,0), colors.black),
                ("ALIGN", (0,0), (-1,0), "CENTER"),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE", (0,0), (-1,0), 14),
                ("BOTTOMPADDING", (0,0), (-1,0), 8),
                ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0,1), (-1,-1), [HexColor("#F9F9F9"), colors.white]),
                ("TEXTCOLOR", (0,1), (-1,-1), colors.black),
                ("FONTNAME", (0,1), (-1,-1), "Helvetica"),
                ("FONTSIZE", (0,1), (-1,-1), 12),
                ("ALIGN", (0,1), (-1,-1), "CENTER"),
            ]))
            elements.append(tbl)
            elements.append(Spacer(1, 20))

            # Grade distribution 1 & 2 (unchanged)…
            grade_data1 = [["O","A+","A","B+","B","C+","C","D+"]]
            grade_data1.append([grade_counts[g] for g in grade_data1[0]])
            grade_table1 = Table(grade_data1, colWidths=[60]*8)
            grade_table1.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), HexColor("#E0E0E0")),
                ("TEXTCOLOR", (0,0), (-1,0), colors.black),
                ("ALIGN", (0,0), (-1,-1), "CENTER"),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE", (0,0), (-1,0), 12),
                ("BOTTOMPADDING", (0,0), (-1,0), 8),
                ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
            ]))
            elements.append(grade_table1)
            elements.append(Spacer(1, 10))

            grade_data2 = [["D","F","S","X","CD"]]
            grade_data2.append([grade_counts[g] for g in grade_data2[0]])
            grade_table2 = Table(grade_data2, colWidths=[80]*5)
            grade_table2.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), HexColor("#E0E0E0")),
                ("TEXTCOLOR", (0,0), (-1,0), colors.black),
                ("ALIGN", (0,0), (-1,-1), "CENTER"),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE", (0,0), (-1,0), 12),
                ("BOTTOMPADDING", (0,0), (-1,0), 8),
                ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
            ]))
            elements.append(grade_table2)
            elements.append(Spacer(1, 40))

            # Verified statement (unchanged)
            verified_style = ParagraphStyle(
                "VerifiedStyle",
                parent=styles["Normal"],
                fontSize=13,
                textColor=HexColor("#333333"),
                spaceAfter=20,
            )
            elements.append(Paragraph(
                "I have carefully checked and verified the submitted grades. "
                "The grade distribution and submitted grades are correct. "
                "[Please mention any exception below.]",
                verified_style
            ))

            # ——— Page Header & Footer ———
            def draw_page(canvas, doc):
                canvas.setTitle(f"Grade Sheet - {course_info.code}")
                canvas.saveState()
                width, height = letter

                # 1) Draw "Grade Sheet" at top center
                p_title = Paragraph("Grade Sheet", header_style)
                w, h = p_title.wrap(doc.width, doc.topMargin)
                p_title.drawOn(canvas, doc.leftMargin, height - h)
                course_details = f"L:{course_info.lecture_hours}, T:{course_info.tutorial_hours}, P:{course_info.project_hours}, C:{course_info.credit}"
                # 2) Draw your five field_label-style lines below it
                hdr_texts = [
                    f"<b>Session:</b> {academic_year}",
                    f"<b>Course Code:</b> {course_info.code}",
                    f"<b>Course Name:</b> {course_info.name} ({course_details})",
                    f"<b>Instructor:</b> {instructor}",
                ]
                y = height - h - header_style.spaceAfter
                for txt in hdr_texts:
                    p = Paragraph(txt, field_label_style)
                    w2, h2 = p.wrap(doc.width, doc.topMargin)
                    p.drawOn(canvas, doc.leftMargin, y - h2)
                    y -= (h2 + field_label_style.spaceAfter)

                # Footer: page number
                canvas.setFont("Helvetica", 9)
                canvas.drawRightString(width - inch, 0.3 * inch, f"Page {doc.page}")

                # Footer: date & signature placeholders
                canvas.setFont("Helvetica", 12)
                canvas.drawString(inch, 0.75 * inch, "")
                canvas.drawString(inch, 0.5 * inch, "Date")
                canvas.drawString(width - 4 * inch, 0.75 * inch, "")
                canvas.drawString(width - 4 * inch, 0.5 * inch, "Course Instructor's Signature")

                canvas.restoreState()

            # Build PDF with header/footer on every page
            doc.build(elements, onFirstPage=draw_page, onLaterPages=draw_page)
            return response

        except Exception as e:
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)
    
    def generate_student_result_pdf(self, request):
        """Generate PDF for student examination results"""
        try:
            # Get data from request
            data = request.data
            student_info = data.get('student_info', {})
            courses = data.get('courses', [])
            spi = float(data.get('spi', 0))
            cpi = float(data.get('cpi', 0))
            su = int(data.get('su', 0))
            tu = int(data.get('tu', 0))
            semester_no = data.get('semester_no', 1)
            semester_type = data.get('semester_type', '')
            semester_label = data.get('semester_label', '')
            formatted_semester = format_semester_display(semester_no, semester_type, semester_label)
            
            # Create PDF buffer
            buffer = BytesIO()
            
            # Create PDF metadata to fix "(anonymous)" title issue
            pdf_title = f"Student Result - {student_info.get('name', student_info.get('rollNumber', 'Student'))} - {formatted_semester}"
            
            doc = SimpleDocTemplate(
                buffer,
                pagesize=letter,
                rightMargin=0.5*inch,
                leftMargin=0.5*inch,
                topMargin=0.5*inch,
                bottomMargin=0.5*inch,
                title=pdf_title,
                author="PDPM IIITDM Jabalpur",
                subject=f"Student Result Report - {formatted_semester}",
                creator="Fusion Academic System"
            )
            
            story = []
            styles = getSampleStyleSheet()
            
            # Header styles
            title_style = ParagraphStyle(
                'Title',
                parent=styles['Heading1'],
                fontSize=14,
                spaceAfter=8,
                alignment=1,  # Center
                fontName='Times-Bold'
            )
            
            subtitle_style = ParagraphStyle(
                'Subtitle',
                parent=styles['Normal'],
                fontSize=11,
                spaceAfter=6,
                alignment=1,  # Center
                fontName='Times-Bold'
            )
            
            # Institution Header
            story.append(Paragraph("PDPM Indian Institute of Information Technology, Design &", title_style))
            story.append(Paragraph("Manufacturing, Jabalpur", title_style))
            story.append(Paragraph("(An Institute of National Importance under MoE, Govt. of India)", subtitle_style))
            story.append(Paragraph("Semester Grade Report / Marksheet", subtitle_style))
            story.append(Spacer(1, 20))
            
            # Student Information Table
            student_data = [
                ['Name of Student:', student_info.get('name', 'N/A'), 'Roll No.:', student_info.get('roll_number', 'N/A')],
                ['Programme:', student_info.get('programme', 'N/A'), 'Branch:', student_info.get('department', 'N/A')],
                ['Semester:', formatted_semester, 'Academic Year:', student_info.get('academic_year', 'N/A')]
            ]
            
            student_table = Table(student_data, colWidths=[1.75*inch, 1.75*inch, 1.75*inch, 1.75*inch])
            student_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Times-Roman'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 5),
                ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            
            story.append(student_table)
            story.append(Spacer(1, 15))
            
            # Courses Table
            headers = ['S. No.', 'Course Code', 'Course Title', 'Credits', 'Grade', 'Grade Points']
            course_data = [headers]
            
            for i, course in enumerate(courses, 1):
                course_data.append([
                    str(i),
                    course.get('coursecode', ''),
                    course.get('coursename', ''),
                    str(course.get('credits', '')),
                    course.get('grade', ''),
                    str(course.get('points', ''))
                ])
            
            course_table = Table(course_data, colWidths=[0.5*inch, 1*inch, 3.2*inch, 0.7*inch, 0.6*inch, 1*inch])
            course_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),  # Header
                ('FONTNAME', (0, 1), (-1, -1), 'Times-Roman'),  # Data
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (2, 1), (2, -1), 'LEFT'),  # Course title left aligned
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            
            story.append(course_table)
            story.append(Spacer(1, 15))
            
            # Summary Table
            summary_data = [
                ['Total Credits Registered:', str(tu), 'Semester Credits Earned:', str(su)],
                ['SPI:', f"{spi:.1f}", 'CPI:', f"{cpi:.1f}"]
            ]
            
            summary_table = Table(summary_data, colWidths=[2.2*inch, 0.8*inch, 2.2*inch, 0.8*inch])
            summary_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Times-Roman'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 5),
                ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 25))
            
            # Footer
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                alignment=1,  # Center
                fontName='Times-Italic'
            )
            
            from datetime import datetime
            current_date = datetime.now().strftime("%d/%m/%Y")
            story.append(Paragraph(f"This is a computer-generated document. Generated on {current_date}", footer_style))
            
            # Build PDF
            doc.build(story)
            
            # Return PDF response
            pdf_data = buffer.getvalue()
            buffer.close()
            
            response = HttpResponse(pdf_data, content_type='application/pdf')
            # Create filename with formatted semester for clarity
            semester_suffix = formatted_semester.replace(' ', '_').replace(':', '').lower() 
            filename = f"result_{student_info.get('rollNumber', student_info.get('roll_number', 'student'))}_{semester_suffix}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Content-Length'] = len(pdf_data)
            
            return response
            
        except Exception as e:
            return JsonResponse({'error': f'PDF generation failed: {str(e)}'}, status=500)


"""
API for the Dean Academic to verify if a course has been submitted by the instructor.

Only users with the role 'Dean Academic' can access this endpoint.

POST: Retrieves courses that have been verified by the Dean Academic.
Expected Requests:
1. POST /api/verify-grades-dean/
    Headers:
        Authorization: Token <your_auth_token>  
        X-User-Role: Dean Academic  
        Body (JSON):
        {
        "Role": "Dean Academic"
        }
    Response:
        200 OK - List of verified courses and unique years
            {
            "courses_info": [
                {
                "id": 1,
                "name": "Mathematics",
                "code": "MATH101"
                }
            ],
            "unique_year_ids": [
                {"year": 2024},
                {"year": 2023}
            ]
            }
        403 Forbidden - Access Denied
            {"error": "Access denied. Invalid or missing role."}
        400 Bad Request - Missing Required Fields
            {"error": "Role is required."}
        500 Internal Server Error - Unexpected Error
            {"error": "An error occurred: <error_message>"}
"""


class VerifyGradesDeanView(APIView):
    """
    API for Dean Academic to verify student grades.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        role = request.data.get("Role")
        academic_year = request.data.get("academic_year")
        semester_type = request.data.get("semester_type")

        if role != "Dean Academic":
            return Response({"error": "Access denied."}, status=status.HTTP_403_FORBIDDEN)
        if not academic_year or not semester_type:
            return Response({"error": "Both academic_year and semester_type are required."},
                            status=status.HTTP_400_BAD_REQUEST)

        # Filter only verified grades matching year & semester_type
        qs = Student_grades.objects.filter(
            academic_year=academic_year,
            semester_type=semester_type,
        )
        course_ids = qs.values_list("course_id_id", flat=True).distinct()
        courses = Courses.objects.filter(id__in=course_ids).order_by("code")
        courses_info = list(courses.values("id", "code", "name"))

        return Response({"courses_info": courses_info}, status=status.HTTP_200_OK)


"""
API for the Dean Academic to verify if a course has been submitted by the instructor.

Only users with the role 'Dean Academic' can access this endpoint.

POST: Checks if a course has been submitted by the instructor.
    If found, returns student registrations for that course.
    If not found, returns a message that the grading is not submitted.
Expected Requests:
1. POST /api/update-enter-grades-dean/
    Headers:
        Authorization: Token <your_auth_token>  
        X-User-Role: Dean Academic  
        Body (JSON):
        {
        "Role": "Dean Academic",
        "course": "CS101",
        "year": 2024
        }
    Response:
        200 OK - Course Found & Student Registrations
        {
        "registrations": [
            {
            "id": 1,
            "course_id": "CS101",
            "year": 2024,
            "grade": "A"
            }
        ]
        }
        404 Not Found - Course Not Submitted
            {"message": "THIS COURSE IS NOT SUBMITTED BY THE INSTRUCTOR"}
        403 Forbidden - Access Denied
            {"error": "Access denied. Invalid or missing role."}
        400 Bad Request - Missing Required Fields
            {"error": "Role, course, and year are required fields."}
        500 Internal Server Error - Unexpected Error
            {"error": "An error occurred: <error_message>"}
"""


class UpdateEnterGradesDeanView(APIView):
    """
    API for Dean Academic to fetch all student registrations for a course & year.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        role = request.data.get("Role")
        course_id = request.data.get("course")
        year = request.data.get("year")
        semester_type = request.data.get("semester_type")

        if role != "Dean Academic":
            return Response({"success": False, "error": "Access denied."}, status=status.HTTP_403_FORBIDDEN)

        qs = Student_grades.objects.filter(course_id=course_id, academic_year=year, semester_type = semester_type)
        if not qs.exists():
            return Response({"message": "THIS COURSE IS NOT SUBMITTED BY THE INSTRUCTOR"})

        return Response({"registrations": list(qs.values())}, status=status.HTTP_200_OK)

"""
Only users with the role 'Dean Academic' can access this endpoint.

POST: Returns a list of verified courses and available working years.
Expected Request:
1. POST /api/validate-dean/
    Headers:
        Authorization: Token <your_auth_token>  
        X-User-Role: Dean Academic  
        Body (JSON):
            {
            "Role": "Dean Academic"
            }
    Response:
        200 OK - List of Verified Courses & Working Years
            {
            "courses_info": [
                {
                "id": 1,
                "name": "Computer Networks",
                "code": "CN101"
                }
            ],
            "working_years": [
                {"working_year": 2024},
                {"working_year": 2023}
            ]
            }
        403 Forbidden - Access Denied
            {"error": "Access denied."}
        400 Bad Request - Missing Required Fields
            {"error": "Role is required."}
        500 Internal Server Error - Unexpected Error
            {"error": "An error occurred: <error_message>"}

"""
class ValidateDeanView(APIView):
    """
    API for Dean Academic to validate courses and fetch working years & batches.
    
    - Only users with the role 'Dean Academic' can access this endpoint.
    - Returns a list of verified courses, available working years, and batch details.
    """

    permission_classes = [IsAuthenticated]

    def post(self, request):
        role = request.data.get("Role")

        if role != "Dean Academic":
            return Response(
                {"error": "Access denied."}, 
                status=status.HTTP_403_FORBIDDEN
            )

        # Fetch unique verified course IDs
        unique_course_ids = Student_grades.objects.filter(verified=True).values("course_id").distinct()
        unique_course_ids = unique_course_ids.annotate(course_id_int=Cast("course_id", IntegerField()))

        # Retrieve course names and codes
        courses_info = Courses.objects.filter(
            id__in=unique_course_ids.values_list("course_id_int", flat=True)
        )

        # Fetch available working years and batch IDs
        working_years = course_registration.objects.values("working_year").distinct()
        unique_batch_ids = Student_grades.objects.values("batch").distinct()

        return Response(
            {
                "courses_info": list(courses_info.values()),
                "working_years": list(working_years),
                # "unique_batch_ids": list(unique_batch_ids),
            },
            status=status.HTTP_200_OK
        )


class ValidateDeanSubmitView(APIView):
    """
    API for Dean Academic to submit and validate student grades from a CSV file.
    
    - Only users with the role 'Dean Academic' can access this endpoint.
    - Accepts a CSV file containing roll numbers, grades, and remarks.
    - Validates grades against the existing database records and finds mismatches.
    """

    permission_classes = [IsAuthenticated]

    def post(self, request):
        role = request.data.get("Role")

        if role != "Dean Academic":
            return Response(
                {"error": "Access denied."},
                status=status.HTTP_403_FORBIDDEN
            )

        # Validate file existence
        if "csv_file" not in request.FILES:
            return Response(
                {"error": "CSV file is required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        csv_file = request.FILES["csv_file"]

        # Validate file type
        if not csv_file.name.endswith(".csv"):
            return Response(
                {"error": "Please submit a valid CSV file."},
                status=status.HTTP_400_BAD_REQUEST
            )

        course_id = request.data.get("course")
        academic_year = request.data.get("year")

        # Validate required fields
        if not course_id or not academic_year:
            return Response(
                {"error": "Course and Academic Year are required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not academic_year.isdigit():
            return Response(
                {"error": "Academic Year must be a number."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Fetch students registered for the course
        students = course_registration.objects.filter(
            course_id_id=course_id, working_year=academic_year
        )

        try:
            # Parse the CSV file
            decoded_file = csv_file.read().decode("utf-8").splitlines()
            reader = csv.DictReader(decoded_file)

            required_columns = ["roll_no", "grade", "remarks"]
            if not all(column in reader.fieldnames for column in required_columns):
                return Response(
                    {"error": "CSV file must contain the following columns: roll_no, grade, remarks."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            mismatches = []

            for row in reader:
                roll_no = row["roll_no"]
                grade = row["grade"]
                remarks = row["remarks"]

                try:
                    student = Student.objects.get(id_id=roll_no)
                    semester = student.curr_semester_no
                    batch = student.batch

                    student_grade = Student_grades.objects.get(
                        roll_no=roll_no,
                        course_id_id=course_id,
                        year=academic_year,
                        batch=batch
                    )

                    if student_grade.grade != grade:
                        mismatches.append({
                            "roll_no": roll_no,
                            "csv_grade": grade,
                            "db_grade": student_grade.grade,
                            "remarks": remarks,
                            "batch": batch,
                            "semester": semester,
                            "course_id": course_id
                        })
                except ObjectDoesNotExist:
                    return Response(
                        {"error": f"Student or grade record not found for roll number: {roll_no}"},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            if not mismatches:
                return Response(
                    {"message": "There are no mismatches."},
                    status=status.HTTP_200_OK
                )

            return Response(
                {"mismatches": mismatches},
                status=status.HTTP_200_OK
            )

        except Exception as e:
            return Response(
                {"error": f"An error occurred while processing the file: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class CheckResultView(APIView):
    """
    API endpoint to retrieve student result information including grades and personal details.
    
    Returns:
        JsonResponse: Contains student personal information, course grades, SPI, CPI, and credit units
        
    Response Structure:
        {
            "success": True,
            "student_info": {
                "name": "Student Full Name",
                "roll_number": "Student Roll Number",
                "programme": "Programme Name",
                "batch": "Batch Information",
                "department": "Department Name",
                "semester": "Current Semester Number",
                "academic_year": "Academic Year (e.g., 2023-24)"
            },
            "courses": [
                {
                    "coursecode": "Course Code",
                    "courseid": "Course ID", 
                    "coursename": "Course Name",
                    "credits": "Course Credits",
                    "grade": "Letter Grade",
                    "points": "Grade Points"
                }
            ],
            "spi": "Semester Performance Index",
            "cpi": "Cumulative Performance Index", 
            "su": "Semester Credits",
            "tu": "Total Credits"
        }
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        roll_number = request.user.username
        semester_no = request.data.get('semester_no')
        semester_type = request.data.get('semester_type')

        if semester_no is None or semester_type is None:
            return JsonResponse(
                {"success": False, "message": "semester_no and semester_type are required."},
                status=400,
            )

        try:
            student = Student.objects.get(id_id=roll_number)
        except Student.DoesNotExist:
            return JsonResponse(
                {"success": False, "message": "Student record not found."},
                status=404,
            )

        # Find the announcement for this batch, number and type
        ann = ResultAnnouncement.objects.filter(
            batch=student.batch_id,
            semester=semester_no,
        ).first()

        if not ann or not ann.announced:
            return JsonResponse(
                {"success": False, "message": "Results not announced yet."},
                status=200,
            )

        # Filter grades on both fields
        grades_info = Student_grades.objects.filter(
            roll_no=roll_number,
            semester=semester_no,
            semester_type=semester_type
        ).select_related('course_id')

        # Get academic year from the first grade record
        academic_year = None
        if grades_info.exists():
            academic_year = grades_info.first().academic_year
        else:
            pass

        spi, su, _ = calculate_spi_for_student(student, semester_no, semester_type)
        cpi, tu, _ = calculate_cpi_for_student(student, semester_no, semester_type)

        # Add student personal information to the response
        student_info = {
            "name": f"{student.id.user.first_name} {student.id.user.last_name}".strip(),
            "rollNumber": student.id.user.username,  # Frontend uses camelCase
            "roll_number": student.id.user.username,  # Backend uses snake_case
            "programme": student.programme,
            "batch": str(student.batch_id) if student.batch_id else str(student.batch),
            "branch": student.id.department.name if student.id.department else "",  # Frontend uses branch
            "department": student.id.department.name if student.id.department else "",  # Backend uses department
            "semester": student.curr_semester_no,
            "academicYear": academic_year or "",  # Frontend uses camelCase
            "academic_year": academic_year or ""   # Backend uses snake_case
        }

        response_data = {
            "success": True,
            "student_info": student_info,
            "courses": [
                {
                    "coursecode": grade.course_id.code,
                    "courseid": grade.course_id.id,
                    "coursename": grade.course_id.name,
                    "credits": grade.course_id.credit,
                    "grade":grade.grade,
                    "points": Decimal(str(grade_conversion.get(grade.grade, 0) * 10)).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP),
                }
                for grade in grades_info
            ],
            "spi": spi,
            "cpi": cpi,
            "su": su,
            "tu": tu,
        }
        return JsonResponse(response_data)


class PreviewGradesAPI(APIView):
    permission_classes = [IsAuthenticated] 
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        # Validate user role
        user_role = request.data.get("Role")
        if user_role != "acadadmin" and user_role!='Assistant Professor' and user_role != 'Professor' and user_role!='Associate Professor':
            return Response(
                {"error": "Access denied."},
                status=status.HTTP_403_FORBIDDEN,
            )

        csv_file = request.FILES.get("csv_file")
        if not csv_file:
            return Response(
                {"error": "No file provided. Please upload a CSV file."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not csv_file.name.endswith(".csv"):
            return Response(
                {"error": "Invalid file format. Please upload a CSV file."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Fetch the parameters from the request
        course_id = request.data.get("course_id")
        academic_year = request.data.get("academic_year")  # e.g., "2023-24"
        semester_type = request.data.get("semester_type")
        if not course_id or not academic_year or not semester_type:
            return Response(
                {"error": "course_id, academic_year and semester_type are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Extract the starting working year from academic_year.
        try:
            working_year = int(academic_year.split("-")[0])
        except Exception as e:
            return Response(
                {"error": "Invalid academic_year format. Expected format like 2023-24."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Get course registration records for the given course and working_year and semester_type.
        registrations = course_registration.objects.filter(
            course_id=course_id,
            session=academic_year,
            semester_type=semester_type,
        )
        # Build a set of registered roll numbers for fast lookup.
        registered_rollnos = set()
        for reg in registrations.select_related("student_id"):
            # Assume the Student model has a roll_no attribute.
            if hasattr(reg.student_id, 'id_id'):
                registered_rollnos.add(reg.student_id.id_id)
            else:
                # If the Student model uses a different field, adjust here.
                registered_rollnos.add(str(reg.student_id_id))

        # Parse CSV file
        try:
            decoded_file = csv_file.read().decode("utf-8")
            io_string = StringIO(decoded_file)
            reader = csv.DictReader(io_string)
        except Exception as e:
            return Response(
                {"error": f"Error reading CSV file: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validate required columns
        required_columns = ["roll_no", "name", "grade", "remarks", "semester"]
        if not all(column in reader.fieldnames for column in required_columns):
            return Response(
                {"error": f"CSV file must contain the following columns: {', '.join(required_columns)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        preview_rows = []
        for row in reader:
            roll_no = row["roll_no"]
            # Check if the current roll number is registered.
            is_registered = roll_no in registered_rollnos

            # Add additional data (e.g. name, grades, remarks, semester) as in CSV
            preview_rows.append({
                "roll_no": roll_no,
                "name": row["name"],
                "grades": row["grade"],
                "remarks": row["remarks"],
                "semester": row["semester"],
                "is_registered": is_registered
            })

        return Response({"preview": preview_rows}, status=status.HTTP_200_OK)
    


class ResultAnnouncementListAPI(APIView):
    """
    GET /api/result-announcements/?role=acadadmin

    Returns an object with:
      - announcements: a list of announcement records with batch info.
      - batches: a list of available batch options (computed label).
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        role = request.query_params.get("role")
        if role != "acadadmin" and role != "Dean Academic":
            return Response({"error": "Access denied."}, status=status.HTTP_403_FORBIDDEN)
        
        # Get announcements sorted by creation date (most recent first)
        announcements = ResultAnnouncement.objects.all().order_by("-created_at")
        ann_data = []
        for ann in announcements:
            # Compute the batch label.
            batch = ann.batch
            batch_label = f"{batch.name} - {batch.discipline.acronym} {batch.year}"
            ann_data.append({
                "id": ann.id,
                "batch": {
                    "id": batch.id,
                    "name": batch.name,
                    "discipline": batch.discipline.acronym,
                    "year": batch.year,
                    "label": batch_label
                },
                "semester": ann.semester,
                "announced": ann.announced,
                "created_at": ann.created_at,
            })
        
        # Fetch available batches (running batches)
        batch_objs = Batch.objects.filter(running_batch=True)
        batch_options = []
        for b in batch_objs:
            # Compute a label exactly as above.
            label = f"{b.name} - {b.discipline.acronym} {b.year}"
            batch_options.append({"id": b.id, "label": label})
        
        return Response({"announcements": ann_data, "batches": batch_options}, status=status.HTTP_200_OK)


class UpdateAnnouncementAPI(APIView):
    """
    POST /api/update-announcement/
    Request Body:
      - id: announcement record ID
      - announced: boolean value
      - Role: must be "acadadmin"
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        role = request.data.get("Role")
        if role != "acadadmin" and role != "Dean Academic":
            return Response({"error": "Access denied."}, status=status.HTTP_403_FORBIDDEN)
        announcement_id = request.data.get("id")
        announced = request.data.get("announced")
        try:
            ann = ResultAnnouncement.objects.get(id=announcement_id)
            ann.announced = announced
            ann.save()
            return Response({"success": True}, status=status.HTTP_200_OK)
        except ResultAnnouncement.DoesNotExist:
            return Response({"error": "Announcement not found."}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            traceback.print_exc()
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CreateAnnouncementAPI(APIView):
    """
    POST /api/create-announcement/
    Body (JSON):
      - Role: must be "acadadmin"
      - batch: integer (Batch PK)
      - semester: integer
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            role = request.data.get("Role")
            if role != "acadadmin" and role != "Dean Academic":
                return Response({"error": "Access denied."}, status=status.HTTP_403_FORBIDDEN)

            batch_id = request.data.get("batch")
            semester = request.data.get("semester")
            if not batch_id or not semester:
                return Response({"error": "Batch and Semester are required."}, status=status.HTTP_400_BAD_REQUEST)

            batch_obj = Batch.objects.filter(id=batch_id).first()
            if not batch_obj:
                return Response({"error": "Batch not found."}, status=status.HTTP_404_NOT_FOUND)

            # Try to get existing announcement, or create if missing
            ann, created = ResultAnnouncement.objects.get_or_create(
                batch=batch_obj,
                semester=semester,
                defaults={"announced": False}
            )

            batch_label = f"{batch_obj.name} - {batch_obj.discipline.acronym} {batch_obj.year}"
            data = {
                "id": ann.id,
                "batch": {"id": batch_obj.id, "label": batch_label},
                "semester": ann.semester,
                "announced": ann.announced,
                "created_at": ann.created_at,
            }

            if created:
                return Response(data, status=status.HTTP_201_CREATED)
            else:
                # Already existed
                return Response(data, status=status.HTTP_200_OK)

        except Exception as e:
            traceback.print_exc()
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

from collections import OrderedDict


def make_label(no: int, sem_type: str) -> str:
    """
    - odd → "Semester <no>"
    - even & Even Semester → "Semester <no>"
    - even & Summer Semester → "Summer <no//2>"
    """
    if no % 2 == 1:
        return f"Semester {no}"
    if sem_type == "Summer Semester":
        return f"Summer {no // 2}"
    return f"Semester {no}"

class StudentSemesterListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        roll_number = request.user.username
        qs = (Student_grades.objects
              .filter(roll_no=roll_number)
              .values_list('semester', 'semester_type')
              .distinct()
              .order_by('semester'))
        unique = OrderedDict()
        for sem_no, sem_type in qs:
            label = make_label(sem_no, sem_type or "")
            unique[(sem_no, sem_type)] = label

        semesters = [
            {"semester_no": no, "semester_type": typ, "label": lbl}
            for (no, typ), lbl in unique.items()
        ]

        return JsonResponse({"success": True, "semesters": semesters})


class GenerateStudentResultPDFAPI(APIView):
    """
    API endpoint to generate PDF report for student examination results
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            # Get data from request
            data = request.data
            
            # Check if student_info is provided, if not fetch it like CheckResultView
            student_info = data.get('student_info', {})
            courses = data.get('courses', [])
            
            # If student_info or courses are not provided, fetch them from database
            if not student_info or not courses:
                # Fetch student data like CheckResultView does
                roll_number = request.user.username
                semester_no = data.get('semester_no')
                semester_type = data.get('semester_type')

                if semester_no is None or semester_type is None:
                    return JsonResponse(
                        {"success": False, "message": "semester_no and semester_type are required."},
                        status=400,
                    )

                try:
                    student = Student.objects.get(id_id=roll_number)
                except Student.DoesNotExist:
                    return JsonResponse(
                        {"success": False, "message": "Student record not found."},
                        status=404,
                    )

                ann = ResultAnnouncement.objects.filter(
                    batch=student.batch_id,
                    semester=semester_no,
                ).first()

                if not ann or not ann.announced:
                    return JsonResponse(
                        {"success": False, "message": "Results not announced yet."},
                        status=200,
                    )

                grades_info = Student_grades.objects.filter(
                    roll_no=roll_number,
                    semester=semester_no,
                    semester_type=semester_type
                ).select_related('course_id')

                academic_year = None
                if grades_info.exists():
                    academic_year = grades_info.first().academic_year

                spi, su, _ = calculate_spi_for_student(student, semester_no, semester_type)
                cpi, tu, _ = calculate_cpi_for_student(student, semester_no, semester_type)

                student_info = {
                    "name": f"{student.id.user.first_name} {student.id.user.last_name}".strip(),
                    "rollNumber": student.id.user.username,
                    "roll_number": student.id.user.username,
                    "programme": student.programme,
                    "batch": str(student.batch_id) if student.batch_id else str(student.batch),
                    "branch": student.id.department.name if student.id.department else "",
                    "department": student.id.department.name if student.id.department else "",
                    "semester": student.curr_semester_no,
                    "academicYear": academic_year or "",
                    "academic_year": academic_year or ""
                }

                # Build courses list like CheckResultView
                from applications.academic_information.models import grade_conversion
                from decimal import Decimal, ROUND_HALF_UP
                
                courses = [
                    {
                        "coursecode": grade.course_id.code,
                        "courseid": grade.course_id.id,
                        "coursename": grade.course_id.name,
                        "credits": grade.course_id.credit,
                        "grade": grade.grade,
                        "points": Decimal(str(grade_conversion.get(grade.grade, 0) * 10)).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP),
                    }
                    for grade in grades_info
                ]
            else:
                # Use provided data
                spi = float(data.get('spi', 0))
                cpi = float(data.get('cpi', 0))
                su = int(data.get('su', 0))
                tu = int(data.get('tu', 0))

            semester_no = data.get('semester_no', 1)
            semester_type = data.get('semester_type', '')
            semester_label = data.get('semester_label', '')

            formatted_semester = format_semester_display(semester_no, semester_type, semester_label)
            
            # Create PDF buffer
            buffer = BytesIO()
            
            # Create PDF metadata to fix "(anonymous)" title issue
            is_transcript = data.get('is_transcript', False) or request.path.find('transcript') != -1 or data.get('document_type') == 'transcript'
            doc_type = "Transcript" if is_transcript else "Student Result"
            pdf_title = f"{doc_type} - {student_info.get('name', student_info.get('rollNumber', 'Student'))} - {formatted_semester}"
            
            doc = SimpleDocTemplate(
                buffer,
                pagesize=letter,
                rightMargin=0.5*inch,
                leftMargin=0.5*inch,
                topMargin=0.5*inch,
                bottomMargin=0.5*inch,
                title=pdf_title,
                author="PDPM IIITDM Jabalpur",
                subject=f"{doc_type} Report - {formatted_semester}",
                creator="Fusion Academic System"
            )
            
            story = []
            styles = getSampleStyleSheet()
            
            # Header styles
            title_style = ParagraphStyle(
                'Title',
                parent=styles['Heading1'],
                fontSize=14,
                spaceAfter=8,
                alignment=1,  # Center
                fontName='Times-Bold'
            )
            
            subtitle_style = ParagraphStyle(
                'Subtitle',
                parent=styles['Normal'],
                fontSize=11,
                spaceAfter=6,
                alignment=1,  # Center
                fontName='Times-Roman'
            )
            
            # Header with logo on the left side
            header_data = []
            
            # Try to add logo
            try:
                from django.conf import settings
                import os
                logo_path = os.path.join(settings.MEDIA_ROOT, 'logo2.jpg')
                if os.path.exists(logo_path):
                    # Make logo smaller to match website appearance
                    logo = Image(logo_path, width=0.8*inch, height=0.8*inch)
                    
                    # Create separate paragraphs for better text formatting
                    title_para = Paragraph("PDPM Indian Institute of Information Technology, Design &", title_style)
                    subtitle1_para = Paragraph("Manufacturing, Jabalpur", title_style)
                    subtitle2_para = Paragraph("(An Institute of National Importance under MoE, Govt. of India)", subtitle_style)
                    subtitle3_para = Paragraph("Semester Grade Report / Marksheet", subtitle_style)
                    
                    # Create header table with logo and text
                    header_table_data = [
                        [logo, [title_para, subtitle1_para, subtitle2_para, subtitle3_para]]
                    ]
                    header_table = Table(header_table_data, colWidths=[1*inch, 6*inch])
                    header_table.setStyle(TableStyle([
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('ALIGN', (0, 0), (0, 0), 'CENTER'),  # Logo center aligned
                        ('ALIGN', (1, 0), (1, 0), 'CENTER'),  # Text center aligned
                        ('LEFTPADDING', (0, 0), (-1, -1), 5),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                        ('TOPPADDING', (0, 0), (-1, -1), 5),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                    ]))
                    story.append(header_table)
                else:
                    # No logo, just text
                    story.append(Paragraph("PDPM Indian Institute of Information Technology, Design &", title_style))
                    story.append(Paragraph("Manufacturing, Jabalpur", title_style))
                    story.append(Paragraph("(An Institute of National Importance under MoE, Govt. of India)", subtitle_style))
                    story.append(Paragraph("Semester Grade Report / Marksheet", subtitle_style))
            except Exception as e:
                # If logo fails, continue with text only
                story.append(Paragraph("PDPM Indian Institute of Information Technology, Design &", title_style))
                story.append(Paragraph("Manufacturing, Jabalpur", title_style))
                story.append(Paragraph("(An Institute of National Importance under MoE, Govt. of India)", subtitle_style))
                story.append(Paragraph("Semester Grade Report / Marksheet", subtitle_style))
            
            story.append(Spacer(1, 12))
            
            # Student Information Table
            student_data = [
                ['Name of Student:', student_info.get('name', 'N/A'), 'Roll No.:', student_info.get('rollNumber', student_info.get('roll_number', 'N/A'))],
                ['Programme:', student_info.get('programme', 'N/A'), 'Branch:', student_info.get('branch', student_info.get('department', 'N/A'))],
                ['Semester:', formatted_semester, 'Academic Year:', student_info.get('academicYear', student_info.get('academic_year', 'N/A'))]
            ]
            
            student_table = Table(student_data, colWidths=[1.75*inch, 1.75*inch, 1.75*inch, 1.75*inch])
            student_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Times-Roman'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            
            story.append(student_table)
            story.append(Spacer(1, 12))
            
            # Courses Table
            headers = ['S. No.', 'Course Code', 'Course Title', 'Credits', 'Grade', 'Grade Points']
            course_data = [headers]
            
            for i, course in enumerate(courses, 1):
                course_data.append([
                    str(i),
                    course.get('coursecode', ''),
                    course.get('coursename', ''),
                    str(course.get('credits', '')),
                    course.get('grade', ''),
                    str(course.get('points', ''))
                ])
            
            course_table = Table(course_data, colWidths=[0.5*inch, 1*inch, 3.2*inch, 0.7*inch, 0.6*inch, 1*inch])
            course_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),  # Header
                ('FONTNAME', (0, 1), (-1, -1), 'Times-Roman'),  # Data
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (2, 1), (2, -1), 'LEFT'),  # Course title left aligned
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            
            story.append(course_table)
            story.append(Spacer(1, 15))
            
            # Summary Table
            summary_data = [
                ['Total Credits Registered:', str(tu), 'Semester Credits Earned:', str(su)],
                ['SPI:', f"{spi:.1f}", 'CPI:', f"{cpi:.1f}"]
            ]
            
            summary_table = Table(summary_data, colWidths=[2.2*inch, 0.8*inch, 2.2*inch, 0.8*inch])
            summary_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Times-Roman'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 5),
                ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 25))
            
            # Footer
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                alignment=1,  # Center
                fontName='Times-Italic'
            )
            
            from datetime import datetime
            current_date = datetime.now().strftime("%d/%m/%Y")
            story.append(Paragraph(f"This is a computer-generated document. Generated on {current_date}", footer_style))
            
            # Build PDF
            doc.build(story)
            
            # Return PDF response
            pdf_data = buffer.getvalue()
            buffer.close()
            
            response = HttpResponse(pdf_data, content_type='application/pdf')
            
            # Check if this is a transcript request (could be determined by various factors)
            # If called from transcript context, use 'transcript_' prefix, otherwise 'result_'
            is_transcript = data.get('is_transcript', False) or request.path.find('transcript') != -1 or data.get('document_type') == 'transcript'
            
            # Create filename with formatted semester for clarity
            semester_suffix = formatted_semester.replace(' ', '_').replace(':', '').lower()
            
            # Choose prefix based on document type
            prefix = "transcript_" if is_transcript else "result_"
            filename = f"{prefix}{student_info.get('rollNumber', student_info.get('roll_number', 'student'))}_{semester_suffix}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Content-Length'] = len(pdf_data)
            
            return response
            
        except Exception as e:
            return JsonResponse({'error': f'PDF generation failed: {str(e)}'}, status=500)